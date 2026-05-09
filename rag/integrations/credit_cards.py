"""Credit-card statements ETL — extracted from rag/cross_source_etls.py 2026-05-09.

Parses the bank's monthly ``Último resumen - <Marca> <Last4>.xlsx`` exports
that land in ``TARJETAS_BACKUP_DIR`` (``CloudDocs/Finances`` by default,
override via ``OBSIDIAN_RAG_FINANCE_DIR``) and renders one markdown note per
``(card, cycle)`` pair under
``99-obsidian/99-AI/external-ingest/Finanzas/Tarjetas/Tarjeta-<brand>-<last4>-<YYYY-MM>.md``
so the regular ``_run_index`` rglob absorbs them.

Silent-fail contract: helpers return ``None`` /
``{ok: False, reason: "..."}`` instead of raising. ``_etl_log_swallow`` is
lazy-imported from ``rag.cross_source_etls`` to avoid circular import.
``openpyxl`` is imported lazily inside ``_parse_credit_card_xlsx`` so the
module loads without it (the credit-card sync is a no-op when the dep is
missing — same contract as ``spotipy`` for Spotify, ``youtube-transcript-api``
for YouTube transcripts).

The xlsx parser is heuristic (not schema-driven): the bank changes columns
between issuers, so we walk rows looking for anchor strings ("Total a pagar",
"Mínimo a pagar", "Próximo resumen", etc.) and read the rows immediately
below. The parser tolerates moved columns; cells that don't match are
silently skipped.

Override paths (env vars, read at module-import time):
  - ``OBSIDIAN_RAG_FINANCE_DIR`` → ``TARJETAS_BACKUP_DIR``.
  - ``OBSIDIAN_RAG_TARJETAS_FOLDER`` → ``TARJETAS_VAULT_SUBPATH``.
"""
from __future__ import annotations

import os
import re
from datetime import datetime
from pathlib import Path

from rag._constants import _EXTERNAL_INGEST_BASE

__all__ = [
    "TARJETAS_BACKUP_DIR",
    "TARJETAS_VAULT_SUBPATH",
    "_CARD_BRAND_RE",
    "_CARD_LAST4_RE",
    "_parse_ars_or_usd",
    "_parse_card_date",
    "_parse_credit_card_xlsx",
    "_card_note_filename",
    "_card_render_note",
    "_sync_credit_cards_notes",
]

# Tarjetas de crédito (xlsx del banco) + PDFs de transferencias siguen en
# `CloudDocs/Finances`. Override: `OBSIDIAN_RAG_FINANCE_DIR` (nombre legacy
# preservado para no romper setups existentes que ya lo tenían apuntando ahí).
TARJETAS_BACKUP_DIR = Path(
    os.environ.get("OBSIDIAN_RAG_FINANCE_DIR", "")
    or (Path.home() / "Library/Mobile Documents/com~apple~CloudDocs/Finances")
)

TARJETAS_VAULT_SUBPATH = os.environ.get(
    "OBSIDIAN_RAG_TARJETAS_FOLDER",
    f"{_EXTERNAL_INGEST_BASE}/Finanzas/Tarjetas",
)

# Regex para extraer marca + últimos 4 del nombre de archivo o sheet name.
_CARD_BRAND_RE = re.compile(
    r"\b(Visa|Master(?:card)?|Amex|American\s*Express|Cabal|Maestro|Naranja)\b",
    re.IGNORECASE,
)
_CARD_LAST4_RE = re.compile(r"(\d{4})(?!\d)")


def _parse_ars_or_usd(raw: object) -> tuple[float | None, str | None]:
    """Parsea celdas tipo `$549.438,75`, `U$S98,93`, `-$926,15`, o numérico
    crudo (openpyxl puede devolver float si la celda tiene formato número).
    Retorna `(amount, currency)` donde currency ∈ {"ARS", "USD", None}.

    Heurística decimal: si el string tiene `,` lo asumimos formato ES
    (decimal coma, miles punto) y normalizamos. Si solo tiene `.` puede ser
    formato US (`24.99`) — solo strippeamos puntos como miles si hay 3
    dígitos exactos después, sino el punto es decimal.
    """
    if raw is None:
        return (None, None)
    if isinstance(raw, (int, float)):
        return (float(raw), None)
    s = str(raw).strip()
    if not s:
        return (None, None)
    cur: str | None = None
    if s.upper().startswith("U$S") or "U$S" in s.upper() or s.upper().startswith("USD"):
        cur = "USD"
    elif s.startswith("$") or "ARS" in s.upper():
        cur = "ARS"
    # Strippeamos símbolos no-numéricos.
    cleaned = re.sub(r"[^\d,.\-]", "", s)
    if "," in cleaned:
        # Formato ES: punto = miles, coma = decimal.
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif cleaned.count(".") > 1:
        # Múltiples puntos → todos son miles ES sin decimal explícito.
        cleaned = cleaned.replace(".", "")
    # Else: un solo punto → decimal US (`24.99`), dejarlo como está.
    try:
        return (float(cleaned), cur)
    except ValueError:
        return (None, cur)


def _parse_card_date(raw: object) -> str | None:
    """DD/MM/YYYY → ISO `YYYY-MM-DD`. None si no parsea. Acepta
    `datetime.date|datetime` (openpyxl normaliza fechas si la celda tiene
    formato fecha) y strings con prefijo (`"Cierre: 26/03/2026"`).
    """
    if raw is None:
        return None
    if hasattr(raw, "strftime"):
        try:
            return raw.strftime("%Y-%m-%d")
        except Exception:
            return None
    s = str(raw).strip()
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if not m:
        return None
    try:
        d = datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        return d.strftime("%Y-%m-%d")
    except ValueError:
        return None


def _parse_credit_card_xlsx(path: Path) -> dict | None:
    """Parsea un `Último resumen - <Marca> <Últimos4>.xlsx` del banco a un
    dict normalizado. None si openpyxl no está disponible o el xlsx no
    tiene la estructura esperada (sin Total a pagar reconocible).
    """
    from rag.cross_source_etls import _etl_log_swallow

    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        _etl_log_swallow("credit_card_xlsx_load", exc)
        return None

    rows: list[tuple] = []
    sheet_name = ""
    try:
        ws = wb.active
        sheet_name = ws.title or ""
        # Lectura completa: el archivo es <100 filas — read_only modo
        # streaming, sin riesgo de memoria.
        for row in ws.iter_rows(values_only=True):
            rows.append(row)
    except Exception as exc:
        _etl_log_swallow("credit_card_xlsx_iter_rows", exc)
        return None
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if not rows:
        return None

    # Identificar marca + últimos 4 — primero del sheet name, sino del
    # nombre de archivo, sino del primer row con "terminada en".
    brand = None
    last4 = None
    for source in (sheet_name, path.stem):
        m_brand = _CARD_BRAND_RE.search(source)
        m_last4 = _CARD_LAST4_RE.search(source)
        if m_brand and not brand:
            brand = m_brand.group(1).title()
            if brand.lower() in ("master", "mastercard"):
                brand = "Mastercard"
            elif brand.lower() in ("amex", "american express"):
                brand = "Amex"
        if m_last4 and not last4:
            last4 = m_last4.group(1)
        if brand and last4:
            break

    # Pasada por filas: extraer secciones por anchor text.
    holder = None
    closing_date = None
    due_date = None
    next_closing_date = None
    next_due_date = None
    total_ars = None
    total_usd = None
    minimum_ars = None
    minimum_usd = None
    top_purchases: list[dict] = []

    def _row_text(r: tuple) -> str:
        return " ".join(str(c) for c in r if c is not None).strip()

    n = len(rows)
    i = 0
    in_purchases_block = False
    while i < n:
        row = rows[i]
        text = _row_text(row).lower()

        # Header: "Tarjeta <Marca> ... terminada en NNNN"
        if not last4 and "terminada en" in text:
            m = _CARD_LAST4_RE.search(text)
            if m:
                last4 = m.group(1)
            m_brand = _CARD_BRAND_RE.search(text)
            if m_brand and not brand:
                brand = m_brand.group(1).title()

        # "Fecha de cierre" / "Fecha de vencimiento" — fila siguiente las trae
        if "fecha de cierre" in text and "fecha de vencimiento" in text and i + 1 < n:
            nxt = rows[i + 1]
            closing_date = _parse_card_date(nxt[0] if len(nxt) > 0 else None)
            due_date = _parse_card_date(nxt[1] if len(nxt) > 1 else None)
            i += 2
            continue

        # "Total a pagar" — fila siguiente: ARS | USD
        if "total a pagar" in text and i + 1 < n:
            nxt = rows[i + 1]
            v0, c0 = _parse_ars_or_usd(nxt[0] if len(nxt) > 0 else None)
            v1, c1 = _parse_ars_or_usd(nxt[1] if len(nxt) > 1 else None)
            for v, c in ((v0, c0), (v1, c1)):
                if v is None:
                    continue
                if c == "USD":
                    total_usd = v
                else:
                    total_ars = v if total_ars is None else total_ars
            i += 2
            continue

        # "Mínimo a pagar"
        if "mínimo a pagar" in text or "minimo a pagar" in text:
            if i + 1 < n:
                nxt = rows[i + 1]
                v0, c0 = _parse_ars_or_usd(nxt[0] if len(nxt) > 0 else None)
                v1, c1 = _parse_ars_or_usd(nxt[1] if len(nxt) > 1 else None)
                for v, c in ((v0, c0), (v1, c1)):
                    if v is None:
                        continue
                    if c == "USD":
                        minimum_usd = v
                    else:
                        minimum_ars = v if minimum_ars is None else minimum_ars
                i += 2
                continue

        # "Próximo resumen" — la fila tiene "Cierre: DD/MM/YYYY" en col 2
        if "próximo resumen" in text or "proximo resumen" in text:
            # Las dos siguientes filas tienen "Cierre: ..." y "Vencimiento: ..."
            for j in range(i + 1, min(i + 4, n)):
                trow = _row_text(rows[j]).lower()
                if "cierre" in trow:
                    cell = rows[j][1] if len(rows[j]) > 1 else None
                    next_closing_date = _parse_card_date(cell)
                if "vencimiento" in trow:
                    cell = rows[j][1] if len(rows[j]) > 1 else None
                    next_due_date = _parse_card_date(cell)
            i += 1
            continue

        # Holder: "<Marca> Crédito terminada en NNNN" + col 2 con "(Titular)"
        if holder is None and len(row) > 1 and row[1] and "titular" in str(row[1]).lower():
            holder = re.sub(r"\s*\(Titular\)\s*$", "", str(row[1])).strip() or None

        # Bloques de movimientos: "Pago de tarjeta y devoluciones" (skipear) /
        # "Tarjeta de <holder>" (capturar movimientos hasta "Total de ...")
        if "pago de tarjeta y devoluciones" in text:
            in_purchases_block = False
        elif text.startswith("tarjeta de ") and "terminada en" in text:
            in_purchases_block = True
            if holder is None:
                m = re.search(r"tarjeta de (.+?)\s*-", text, re.IGNORECASE)
                if m:
                    holder = m.group(1).strip().title()
        elif text.startswith("total de ") and ("terminada en" in text or "tarjeta" in text):
            in_purchases_block = False
        elif text.startswith("otros conceptos"):
            in_purchases_block = False
        elif in_purchases_block and len(row) >= 5:
            # Fila de movimiento: (fecha, descripción, cuotas, comprobante, ARS, USD)
            desc = (str(row[1]).strip() if row[1] else "").strip()
            if desc and desc.lower() not in ("descripción", "descripcion"):
                amt_ars, _ = _parse_ars_or_usd(row[4] if len(row) > 4 else None)
                amt_usd, _ = _parse_ars_or_usd(row[5] if len(row) > 5 else None)
                amount = None
                currency = None
                if amt_ars is not None:
                    amount = abs(amt_ars)
                    currency = "ARS"
                elif amt_usd is not None:
                    amount = abs(amt_usd)
                    currency = "USD"
                if amount is not None and amount > 0:
                    top_purchases.append({
                        "date": _parse_card_date(row[0] if len(row) > 0 else None),
                        "description": desc,
                        "amount": amount,
                        "currency": currency,
                    })

        i += 1

    # Sin total ni fechas reconocibles → xlsx no es un resumen válido.
    if total_ars is None and total_usd is None and not closing_date and not due_date:
        return None

    # Top 5 movimientos por monto absoluto, separando ARS/USD.
    ars_purchases = sorted(
        (p for p in top_purchases if p["currency"] == "ARS"),
        key=lambda p: p["amount"],
        reverse=True,
    )[:5]
    usd_purchases = sorted(
        (p for p in top_purchases if p["currency"] == "USD"),
        key=lambda p: p["amount"],
        reverse=True,
    )[:3]

    try:
        mtime = path.stat().st_mtime
    except Exception:
        mtime = 0.0

    return {
        "brand": brand,
        "last4": last4,
        "holder": holder,
        "closing_date": closing_date,
        "due_date": due_date,
        "next_closing_date": next_closing_date,
        "next_due_date": next_due_date,
        "total_ars": total_ars,
        "total_usd": total_usd,
        "minimum_ars": minimum_ars,
        "minimum_usd": minimum_usd,
        "top_purchases_ars": ars_purchases,
        "top_purchases_usd": usd_purchases,
        "source_file": path.name,
        "source_mtime": datetime.fromtimestamp(mtime).isoformat(timespec="seconds") if mtime else None,
        "all_purchases": top_purchases,  # extra para el render de nota
    }


def _card_note_filename(card: dict) -> str | None:
    """Construye el filename de la nota .md para una tarjeta parseada."""
    brand = (card.get("brand") or "").strip()
    last4 = (card.get("last4") or "").strip()
    if not brand or not last4:
        return None
    cycle = card.get("closing_date") or card.get("due_date")
    if not cycle:
        return None
    yyyy_mm = cycle[:7]  # "2026-03-26" → "2026-03"
    return f"Tarjeta-{brand}-{last4}-{yyyy_mm}.md"


def _card_render_note(card: dict) -> str:
    """Renderiza una nota markdown para un resumen de tarjeta parseado."""
    brand = card.get("brand") or "?"
    last4 = card.get("last4") or "????"
    holder = card.get("holder") or ""
    closing = card.get("closing_date") or ""
    due = card.get("due_date") or ""
    next_closing = card.get("next_closing_date") or ""
    next_due = card.get("next_due_date") or ""
    total_ars = card.get("total_ars")
    total_usd = card.get("total_usd")
    minimum_ars = card.get("minimum_ars")
    minimum_usd = card.get("minimum_usd")
    src_file = card.get("source_file") or ""
    cycle_yyyy_mm = (closing or due or "")[:7]

    def _fmt_ars(v: float | None) -> str:
        if v is None:
            return "—"
        s = f"{v:,.2f}"  # `1,234,567.89` → en-US
        return s.replace(",", "_TMP_").replace(".", ",").replace("_TMP_", ".")

    def _fmt_usd(v: float | None) -> str:
        if v is None:
            return "—"
        return f"{v:,.2f}".replace(",", "_TMP_").replace(".", ",").replace("_TMP_", ".")

    fm: list[str] = ["---"]
    fm.append("type: finanzas")
    fm.append("source: tarjeta")
    fm.append(f"brand: {brand}")
    fm.append(f"last4: \"{last4}\"")
    if holder:
        fm.append(f"holder: {holder}")
    if cycle_yyyy_mm:
        fm.append(f"cycle: {cycle_yyyy_mm}")
    if closing:
        fm.append(f"closing_date: {closing}")
    if due:
        fm.append(f"due_date: {due}")
    if next_closing:
        fm.append(f"next_closing_date: {next_closing}")
    if next_due:
        fm.append(f"next_due_date: {next_due}")
    if total_ars is not None:
        fm.append(f"total_ars: {total_ars:.2f}")
    if total_usd is not None:
        fm.append(f"total_usd: {total_usd:.2f}")
    if minimum_ars is not None:
        fm.append(f"minimum_ars: {minimum_ars:.2f}")
    if minimum_usd is not None:
        fm.append(f"minimum_usd: {minimum_usd:.2f}")
    tags = ["finanzas", "tarjeta", brand.lower()]
    fm.append(f"tags: [{', '.join(tags)}]")
    fm.append("ambient: skip")
    if src_file:
        fm.append(f"source_file: {src_file}")
    fm.append("---")
    fm.append("")

    title = f"# Tarjeta {brand} ·{last4} — ciclo {cycle_yyyy_mm or 'sin fecha'}"
    body: list[str] = [title, ""]
    if holder:
        body.append(f"Titular: **{holder}**.")
        body.append("")

    parts: list[str] = []
    if total_ars is not None:
        parts.append(f"total ARS **${_fmt_ars(total_ars)}**")
    if total_usd is not None:
        parts.append(f"total USD **U$S{_fmt_usd(total_usd)}**")
    if parts:
        body.append("Total a pagar este ciclo: " + " · ".join(parts) + ".")
    if closing or due:
        body.append(f"Fecha de cierre: **{closing or '—'}** · Fecha de vencimiento: **{due or '—'}**.")
    if minimum_ars is not None or minimum_usd is not None:
        mins: list[str] = []
        if minimum_ars is not None:
            mins.append(f"ARS **${_fmt_ars(minimum_ars)}**")
        if minimum_usd is not None:
            mins.append(f"USD **U$S{_fmt_usd(minimum_usd)}**")
        body.append("Mínimo a pagar: " + " · ".join(mins) + ".")
    if next_closing or next_due:
        body.append(f"Próximo ciclo — cierre: {next_closing or '—'} · vencimiento: {next_due or '—'}.")
    body.append("")

    ars = card.get("top_purchases_ars") or []
    if ars:
        body.append("## Top movimientos ARS")
        body.append("")
        body.append("| Fecha | Descripción | Monto |")
        body.append("|---|---|---:|")
        for p in ars:
            d = p.get("date") or ""
            desc = (p.get("description") or "").replace("|", "\\|")
            amt = p.get("amount")
            body.append(f"| {d} | {desc} | ${_fmt_ars(amt)} |")
        body.append("")

    usd = card.get("top_purchases_usd") or []
    if usd:
        body.append("## Top movimientos USD")
        body.append("")
        body.append("| Fecha | Descripción | Monto |")
        body.append("|---|---|---:|")
        for p in usd:
            d = p.get("date") or ""
            desc = (p.get("description") or "").replace("|", "\\|")
            amt = p.get("amount")
            body.append(f"| {d} | {desc} | U$S{_fmt_usd(amt)} |")
        body.append("")

    all_p = card.get("all_purchases") or []
    if all_p:
        body.append("## Todos los movimientos")
        body.append("")
        for p in all_p:
            d = p.get("date") or ""
            desc = (p.get("description") or "").replace("|", "\\|")
            cur = p.get("currency") or ""
            amt = p.get("amount")
            if cur == "USD":
                body.append(f"- {d} — {desc} — U$S{_fmt_usd(amt)}")
            else:
                body.append(f"- {d} — {desc} — ${_fmt_ars(amt)}")
        body.append("")

    body.append(f"_Fuente: `{src_file}` (export del banco)._")
    body.append("")

    return "\n".join(fm + body)


def _sync_credit_cards_notes(vault_root: Path) -> dict:
    """Regenerate per-cycle credit-card notes from the xlsx exports in
    `TARJETAS_BACKUP_DIR` (`CloudDocs/Finances` por default). Mirrors the
    MOZE pattern: one note per (card, cycle), hash-skip if content is
    unchanged, prune notes whose source xlsx no longer exists.
    """
    import sys as _sys

    from rag.cross_source_etls import _etl_log_swallow

    # Re-resolve constantes en call-time via `sys.modules.get("rag")` para
    # que los `monkeypatch.setattr(rag, "TARJETAS_BACKUP_DIR", ...)` de los
    # tests propaguen al call site real (el binding local quedaría cacheado
    # con el path de iCloud sin esto).
    _rag = _sys.modules.get("rag")
    _backup_dir = getattr(_rag, "TARJETAS_BACKUP_DIR", TARJETAS_BACKUP_DIR)
    _vault_subpath = getattr(_rag, "TARJETAS_VAULT_SUBPATH", TARJETAS_VAULT_SUBPATH)

    try:
        seen: set[Path] = set()
        for pattern in ("Último resumen*.xlsx", "Ultimo resumen*.xlsx"):
            for p in _backup_dir.glob(pattern):
                seen.add(p)
        files = sorted(seen, key=lambda p: p.name)
    except Exception as exc:
        return {"ok": False, "reason": f"glob: {exc}"}
    if not files:
        return {"ok": False, "reason": "no_xlsx"}

    target_dir = vault_root / _vault_subpath
    try:
        _vr = vault_root.resolve()
        _td = target_dir.resolve()
        _td.relative_to(_vr)
    except Exception:
        return {"ok": False, "reason": f"target escapa vault: {target_dir}"}
    try:
        target_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        return {"ok": False, "reason": f"mkdir: {exc}"}

    written = 0
    skipped = 0
    parse_failed = 0
    current_set: set[str] = set()

    for xlsx in files:
        card = _parse_credit_card_xlsx(xlsx)
        if not card:
            parse_failed += 1
            continue
        fname = _card_note_filename(card)
        if not fname:
            parse_failed += 1
            continue
        current_set.add(fname)
        body = _card_render_note(card)
        path = target_dir / fname
        existing = path.read_text(encoding="utf-8") if path.is_file() else ""
        if existing == body:
            skipped += 1
            continue
        path.write_text(body, encoding="utf-8")
        written += 1

    for p in target_dir.glob("Tarjeta-*.md"):
        if p.name not in current_set:
            try:
                p.unlink()
            except Exception as exc:
                _etl_log_swallow("credit_card_prune_stale", exc)

    if not current_set and parse_failed:
        return {"ok": False, "reason": "no_parsed"}

    return {
        "ok": True,
        "files_total": len(files),
        "files_written": written,
        "files_skipped": skipped,
        "files_parse_failed": parse_failed,
        "target": str(target_dir.relative_to(vault_root)),
    }
