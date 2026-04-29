"""Web UI mínima para `rag chat`.

Espeja el pipeline del CLI (multi_retrieve → command-r streaming → sources)
sobre HTTP + SSE. Sin build step en el frontend; vanilla JS contra este
endpoint. Sesiones persistidas en el mismo store que el CLI — el session_id
de la web es `web:<uuid>` así no colisiona con `tg:<chat_id>` ni con los
ids del chat interactivo.
"""
from __future__ import annotations

import asyncio
import json
import os
import re
import subprocess
import sys
import threading
import time
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Callable

# === LOKY SEMAPHORE LEAK FIX (2026-04-25) ===================================
# `tqdm` (transitively pulled by sentence-transformers / transformers) crea
# un `multiprocessing.RLock()` lazy en su primer `tqdm.get_lock()` para
# coordinar progress bars en multi-process. Ese RLock es un POSIX named
# semaphore que NO se libera al shutdown del web daemon — Python no tiene
# atexit hook que lo unlinkee.
#
# Adicionalmente, `joblib.externals.loky.backend.__init__` monkey-patcha
# `multiprocessing.synchronize.SemLock._make_name` para que TODOS los
# SemLocks creados después usen el prefix `/loky-PID-XXX`. Resultado: la
# warning `leaked semaphore objects: {/loky-PID-XXX}` aparece en cada
# clean shutdown del web daemon — 247 leaks acumulados en `web.error.log`
# pre-fix.
#
# Trace empírico (2026-04-25 con monkey-patch de SemLock.__init__) confirma:
# el SemLock se crea en `tqdm/std.py:121 create_mp_lock` durante el primer
# `cls.get_lock()` que dispara sentence-transformers al cargar el reranker.
#
# Fix: pre-set el lock de tqdm a un `threading.RLock` (in-process, no
# semaphore POSIX) ANTES de que cualquier dep pesado (sentence-transformers,
# transformers, etc.) toque tqdm. La condición `not hasattr(cls, '_lock')`
# en `tqdm.tqdm.get_lock()` deja nuestro lock alone y nunca llama
# `TqdmDefaultWriteLock()` que es donde se crea el SemLock.
#
# Side effect: las progress bars de tqdm en este proceso ya no son
# inter-process safe — pero el web daemon es single-process (uvicorn con
# workers=1), así que no hay loss real. Si alguna vez se introduce
# multi-process workers, este lock necesita re-evaluación.
try:
    import tqdm as _tqdm
    _tqdm.tqdm.set_lock(threading.RLock())
except Exception:  # pragma: no cover - tqdm not installed
    pass
# === END LOKY SEMAPHORE LEAK FIX ============================================

# pillow-heif registra el HEIC/HEIF reader/writer en PIL. Audit 2026-04-25
# R2-OCR #4 followup: sin esto, las fotos del iPhone (HEIC default) eran
# passthrough en `_sanitize_image_exif` y conservaban GPS coords al
# copiarse al vault iCloud. Lo registramos UNA sola vez al import del
# módulo — `register_heif_opener()` es idempotente pero igual lo
# guardeamos detrás del flag `_HEIC_AVAILABLE` para que el sanitizer
# (y sus tests) puedan detectar cuándo el plugin está disponible.
try:
    import pillow_heif as _pillow_heif  # noqa: PLC0415
    _pillow_heif.register_heif_opener()
    _HEIC_AVAILABLE = True
except ImportError:
    _HEIC_AVAILABLE = False

from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field, field_validator

# rag.py vive en el root del proyecto; lo importamos como módulo.
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import ollama  # noqa: E402

from rag import (  # noqa: E402
    CHAT_OPTIONS,
    CONFIDENCE_RERANK_MIN,
    CONTRADICTION_LOG_PATH,
    LOG_PATH,
    SILENT_ERRORS_LOG_PATH,
    _LOG_QUEUE,
    _LOOKUP_MODEL,
    _LOOKUP_NUM_CTX,
    RAG_STATE_SQL,
    _SQL_STATE_ERROR_LOG,
    _enqueue_background_sql,
    _log_sql_state_error,
    _map_cpu_row,
    _map_memory_row,
    _ragvec_state_conn,
    _sql_append_event,
    MORNING_FOLDER,
    OLLAMA_KEEP_ALIVE,
    chat_keep_alive,
    SESSION_HISTORY_WINDOW,
    _collect_screentime,
    _fmt_hm,
    VAULT_PATH,
    WHATSAPP_BOT_JID,
    WHATSAPP_DB_PATH,
    _apple_enabled,
    _build_tasks_system_rules as _rag_build_tasks_system_rules,
    _collect_scoped_tasks_evidence_multi as _rag_collect_scoped_tasks_evidence_multi,
    _collect_today_evidence,
    _fetch_calendar_ahead,
    _fetch_calendar_today,
    _fetch_chrome_bookmarks_used,
    _fetch_drive_evidence,
    _fetch_gmail_today,
    _fetch_reminders_due,
    _fetch_vault_activity,
    _fetch_weather_forecast,
    _fetch_whatsapp_today,
    _fetch_whatsapp_unread,
    _fetch_youtube_today,
    _format_scoped_tasks_context as _rag_format_scoped_tasks_context,
    _generate_today_narrative,
    _icalbuddy_path,
    _load_corpus,
    _load_vaults_config,
    _path_to_title,
    _pendientes_collect,
    _pendientes_urgent,
    _render_today_prompt,
    _round_timing_ms,
    _tasks_services_consulted as _rag_tasks_services_consulted,
    append_turn,
    ensure_session,
    _extract_followup_loops,
    _note_created_ts,
    find_contradictions_for_note,
    find_followup_loops,
    find_related,
    find_wikilink_suggestions,
    get_db,
    get_pagerank,
    log_behavior_event,
    log_query_event,
    multi_retrieve,
    new_turn_id,
    record_feedback,
    reformulate_query,
    resolve_chat_model,
    resolve_vault_paths,
    save_session,
    session_history,
    TOPIC_SHIFT_COSINE,
)

# Cosine band for the borderline reform-LLM gate. The lower bound matches
# `TOPIC_SHIFT_COSINE` (0.40) — anything below already gets `history = []`
# from `detect_topic_shift`, so the borderline case is "history kept but
# cosine ≤ this upper bound". 0.7 chosen as the empirical threshold above
# which user follow-ups are paraphrases / self-contained restatements
# rather than elliptical references — measured 2026-04-26 on the chains
# golden set: 0.585 ("listame los gastos en pesos" after "Cuanto devo a
# la visa?") needed reform; 0.887 ("listame los gastos en pesos de la
# visa") was already self-contained. 0.7 splits the two.
REFORM_COSINE_HIGH = 0.70

from web.conversation_writer import write_turn, TurnData  # noqa: E402
from web.tools import (  # noqa: E402
    CHAT_TOOLS,
    CHAT_TOOL_OPTIONS,
    PARALLEL_SAFE,
    PROPOSAL_TOOL_NAMES,
    TOOL_FNS,
    _WEB_TOOL_ADDENDUM,
)


# Pre-router: keyword → tool forced execution. Defeats the system prompt's
# "engancháte SIEMPRE con el CONTEXTO" bias which made the LLM decline to
# call tools when retrieval already returned anything (even low-conf vault
# notes). Regex-based, O(len(q)), runs before the LLM sees the query.
#
# Matches are independent — "gastos de este mes en agenda" fires both
# finance_summary and calendar_ahead.
# Patterns covering "planning" queries — "cómo viene mi semana", "qué
# tengo hoy", "qué hay mañana". These are ambiguous enough that we fire
# BOTH reminders_due AND calendar_ahead (via the shared pattern below)
# so the LLM has tasks + events to build a picture of the timeframe.
# Plurals: `\bX\b` no matchea "Xs" porque la `s` es word-char y el
# segundo `\b` requiere transición word→non-word. User report 2026-04-24:
# "cuales son mis ultimos mails?" devolvía info de WhatsApp porque el
# pre-router regex fallaba en plurales españoles ("mails", "correos",
# "eventos", "días", "semanas"…). Fix: sufijo `s?` antes del word-boundary
# en cada token que aparece flexionado en español rioplatense. "clima" y
# "tiempo" no tienen plural natural en este contexto pero los normalizamos
# por consistencia (no introduce falsos positivos relevantes).
_PLANNING_PAT = (
    r"\bsemanas?\b|\bhoy\b|\bma[ñn]ana\b|pasado\s+ma[ñn]ana|\bd[ií]as?\b|c[oó]mo\s+viene"
    # "qué tengo / hay / tenés" only counts as planning when followed by a
    # temporal token (hoy, mañana, semana, día, agenda, pendiente) so that
    # "qué tengo sobre coaching" doesn't fire calendar + reminders.
    r"|qu[eé]\s+(tengo|hay|ten[eé]s)\b(?=.{0,40}\b(hoy|ma[ñn]ana|semanas?|d[ií]as?|agendas?|pendient|tarea|recordator)\b)"
)

_TOOL_INTENT_RULES: tuple[tuple[str, dict, str], ...] = (
    ("finance_summary", {}, r"gast[oéó]s?|gast[aá][mn]os|gastar|presupuesto|plata|finanz|moze"),
    # Tarjetas: keywords inequívocos del dominio "resumen de tarjeta de
    # crédito" → fuerza credit_cards_summary.
    #
    # Triggers (en orden de especificidad):
    #   1. tarjeta(s) | visa | master(card) | amex | crédito           — marca
    #   2. saldo a pagar | fecha de cierre | fecha de vencimiento      — ciclo
    #   3. resumen de tarjeta | cuánto debo                            — directo
    #   4. último/reciente + gasto/consumo/movimiento/compra/cargo     — recientes
    #   5. consumo/movimiento/cargo + de/del/en/con/mi                 — fraseo
    #
    # 2026-04-26: ampliado tras user report — el LLM alucinaba "Helado en
    # Manalu $15K" cuando la query era "cuál fue mi último gasto" (sin
    # "tarjeta" explícito) porque solo finance_summary disparaba (MOZE).
    # Triggers 4+5 disparan AMBOS tools (finance_summary + credit_cards_
    # summary): el LLM ve los dos contextos y prioriza el más específico
    # via REGLA 1.b del system prompt (datos transaccionales: cita literal,
    # no inventes).
    #
    # No matchean por diseño: "saldo" solo (puede ser cuenta/billetera),
    # "tiempo" (clima), "gasto" solo sin marca (MOZE genérico).
    ("credit_cards_summary", {}, r"\btarjet|\bvisa\b|\bmaster(?:card)?\b|\bamex\b|\bcr[eé]dito\b|saldo.*paga|fecha.*cierre|fecha.*vencim|resumen.*tarjeta|cu[aá]nto.*deb[oe]|\b(?:[uú]ltim[oa]s?|recientes?)\s+(?:gastos?|consumos?|movimientos?|compras?|cargos?|transac)\b|\b(?:consumos?|movimientos?|cargos?)\s+(?:de|del|en|con|mi)"),
    # NOTE: "recordame" / "agendáme" used to be here as query triggers, but
    # they're CREATE intents now (propose_reminder / propose_calendar_event).
    # Moved to `_PROPOSE_INTENT_RE` below. `recordator` still matches
    # "qué recordatorios tengo" → list.
    ("reminders_due",   {}, r"pendient|tarea|to.?do|recordator|" + _PLANNING_PAT),
    # "inbox" / "bandeja" alone are too generic — Obsidian PARA uses
    # "00-Inbox" heavily. Require an explicit mail signal.
    ("gmail_recent",    {}, r"\b(mail|correo|e.?mail|gmail)s?\b|bandeja\s+de\s+entrada"),
    ("calendar_ahead",  {}, r"calendari|\beventos?\b|\bcitas?\b|reuni[oó]n|\bagendas?\b|pr[oó]xim[ao]s?\s+d[ií]as|" + _PLANNING_PAT),
    ("weather",         {}, r"\bclimas?\b|\btiempos?\b|llov|lluvia|temperatur|pron[oó]stico"),
    # Drive: "drive" alone es ambiguo (también es una palabra EN genérica,
    # "drive-thru", "hard drive"), pero "drive" precedido por
    # google/en/mi/tu/del/al/a/sobre es señal clara de la intent
    # "buscar en Google Drive". También disparan keywords
    # Drive-específicas (planilla, spreadsheet, sheet, presentación).
    # "drive_search" es el único tool del pre-router que recibe args
    # dinámicos — la query cruda del user se inyecta en
    # `_detect_tool_intent` con key `query`, y el helper de rag filtra
    # stopwords + resuelve aliases internamente. Ver
    # `_agent_tool_drive_search` para la lógica completa.
    ("drive_search",    {}, r"(?:google\s*|(?:en|mi|tu|del|al|a|sobre)\s+)drive\b|\bplanillas?\b|\bspreadsheets?\b|\bsheets?\b|\bpresentaci[oó]n\b|\bpresentaciones\b"),
    # WhatsApp: "qué tengo pendiente" sin contexto WA explícito también
    # dispara el tool — los chats sin respuesta son parte del "pending
    # bucket" semántico del user (reportado 2026-04-24 Fer F.: "me
    # faltan las conversaciones de wzp relativa a la pregunta" cuando
    # preguntó por pendientes de la semana). Keywords explícitos
    # (whatsapp/wzp/wsp/chats) también gatillan. `pendient` plain
    # (sin `\b`) matcha "pendiente(s)/pendient(e) de contestar/etc" —
    # mismo patrón que usa `reminders_due`, así el scope de "pending
    # bucket" está consistente entre tasks/events y chats. `chats?` por
    # sí solo es aceptable porque las otras apps de chat del user
    # (Slack, Messages) no están integradas al RAG todavía — en la
    # práctica "chat" == WhatsApp en este setup.
    ("whatsapp_pending", {}, r"whats.?app|\bwzp\b|\bwsp\b|\bchats?\b|mensajes?\s+sin|pendient|" + _PLANNING_PAT),
    # `whatsapp_search` — buscar DENTRO del contenido de los mensajes
    # (corpus indexado, ~4500 chunks), distinto de `whatsapp_pending`
    # (que sólo lista chats sin respuesta). Triggers son frases de
    # "comunicación pasada" que no tocan los keywords de pending:
    #   - "qué me/te/le dijo Juan" / "qué me dijeron"
    #   - "qué me mandó/escribió María" / "qué me comentó"
    #   - "cuándo quedamos / hablamos / charlamos / acordamos"
    #   - "el chat donde X mencionó Y" / "dónde mencionó / habló de"
    # Importante: NO solapamos con `whatsapp_pending` porque ahí los
    # keywords son `whatsapp/wzp/wsp/chats?/mensajes sin/pendient`. Las
    # frases de acá hablan de pasado (dijo/mandó/quedamos), no de
    # estado (pendiente/sin responder). La cobertura adicional la pone
    # el LLM via el addendum cuando el pre-router no engancha. Pasamos
    # la query cruda al tool — el helper resuelve `contact` desde el
    # texto natural a futuro; por ahora deja `contact=None` y el
    # retrieve unfiltered ya devuelve los matches relevantes.
    ("whatsapp_search", {}, (
        r"qu[eé]\s+(me|te|le|nos)?\s*"
        r"(dij[oe]|dij[oe]ron|mand[oó]|mandaron|escribi[oó]|escribieron|coment[oó]|comentaron|cont[oó]|contaron)"
        r"|cu[aá]ndo\s+(quedam|hablam|charlam|acordam)"
        r"|d[oó]nde\s+(hablam|charlam|menci[oó]n|qued[oó])"
        r"|el\s+chat\s+(donde|en\s+el\s+que)"
    )),
)
_TOOL_INTENT_COMPILED = tuple(
    (name, args, re.compile(pat, re.IGNORECASE)) for name, args, pat in _TOOL_INTENT_RULES
)


# Pre-router heuristic for read_note (eval 2026-04-28 fix). El LLM
# tiende a ignorar read_note y prefiere search_vault aunque el user
# pidió leer un archivo específico. Disparamos el tool a mano cuando
# el patrón es clarísimo.
#
# Conservative match: SOLO dispara si hay (a) qualifier explícito
# "la nota / el archivo / el md / el markdown" + nombre, O (b) un
# nombre que termina en `.md`. Sin esto el regex levantaba "abrí
# gmail" → read_note('gmail.md') (rompía test_web_tool_intent_plurals
# y test_whatsapp_search_tool, eval 2026-04-28).
_READ_NOTE_TRIGGER_RE = re.compile(
    r"\b(?:le[eé]|abr[ií]|mostrame|muestra|mu[eé]strame|"
    r"showme|show\s+me)\s+"
    r"(?:"
    # (a) qualifier-prefixed: "la nota CLAUDE", "el archivo notas",
    # "el md plan", "la nota 02-Areas/Foo".
    r"(?:la\s+nota|el\s+archivo|el\s+md|el\s+markdown)\s+"
    r"([A-Za-z0-9_\-./áéíóúñÁÉÍÓÚÑ][A-Za-z0-9_\-./áéíóúñÁÉÍÓÚÑ]{1,79})"
    r"|"
    # (b) bare path con extensión `.md` explícita.
    r"([A-Za-z0-9_\-./áéíóúñÁÉÍÓÚÑ][A-Za-z0-9_\-./áéíóúñÁÉÍÓÚÑ]{1,79}\.md)"
    r")",
    re.IGNORECASE,
)


def _detect_read_note_intent(q: str) -> tuple[str, dict] | None:
    """Detect 'leé la nota X / abrí X.md / mostrame el archivo X' patterns.

    Returns ('read_note', {'path': resolved_path}) si matchea, None si no.

    Conservative: solo dispara si el pattern es clarísimo (qualifier
    explícito tipo "la nota / el archivo" o sufijo `.md`). False
    positives son peor que false negatives — si dudoso, el LLM puede
    igual elegir read_note vía el tool addendum.
    """
    if not q or not q.strip():
        return None
    m = _READ_NOTE_TRIGGER_RE.search(q)
    if not m:
        return None
    raw = (m.group(1) or m.group(2) or "").strip()
    if not raw:
        return None
    # Si no termina en .md, agregarlo. Y si el user pasó solo el nombre,
    # asumimos que es relativo al root del vault.
    path = raw if raw.lower().endswith(".md") else f"{raw}.md"
    return ("read_note", {"path": path})


# Eval 2026-04-28 BUG-2b: queries "clima en CIUDAD" disparaban 4 tools
# en paralelo (weather + reminders + calendar + wa_pending) porque
# "hoy" matchea _PLANNING_PAT. Cuando el user es explícito sobre la
# ciudad, queremos SOLO weather con esa location, sin morning-brief
# noise.
#
# 2026-04-28 wave-3: relajado el regex porque la versión inicial era
# demasiado estricta:
#   1. "qué clima hace en BA hoy" no matcheaba — "hace" intercalado
#      entre keyword y prep cortaba el match. Ahora aceptamos hasta 3
#      palabras intermedias.
#   2. Location lowercase ("buenos aires") fallaba porque pedía mayúscula
#      al principio. Ahora aceptamos cualquier letra inicial.
#   3. Lookahead final era demasiado restrictivo. Ahora dejamos que la
#      location sea cualquier cosa hasta `?` / `,` / fin de string /
#      keyword temporal.
_WEATHER_LOCATION_RE = re.compile(
    r"\b(?:clima|tiempos?|pron[oó]stico|temperatur[ao])\b"
    r"(?:\s+\w+){0,3}?"  # palabras intermedias opcionales (hace, hay, está, etc.)
    r"\s+(?:en|de|para|para\s+el?)\s+"
    r"(?P<loc>[A-Za-zÁÉÍÓÚÑáéíóúñ][\wáéíóúñÁÉÍÓÚÑ\s]{1,40}?)"
    r"(?=\s*(?:[?.,!]|$|\bhoy\b|\bmañana\b|\bahora\b|\bpasado\s+mañana\b))",
    re.IGNORECASE,
)


# 2026-04-28 wave-8 (eval Conv 4 T2): patrones de follow-up anafórico
# temporal que indican "lo mismo del turno previo, pero shifteado en
# tiempo o en el slot de location". Repro: "y mañana?" tras "qué hago
# hoy?" — el LLM alucinaba propose_reminder. Usado por el carry-over
# pre-router para re-fire los mismos read-intent tools del turno previo.
_ANAPHORIC_TEMPORAL_RE = re.compile(
    r"^\s*"
    r"(?:y\s+)?"  # opcional "y ..."
    r"(?:"
    # Adverbios temporales puros (con o sin trailing words cortos)
    r"(?:hoy|ma[ñn]ana|ayer|anoche|antier|pasado\s+ma[ñn]ana)"
    # Frases temporales "la/esta/proxima semana/tarde/noche/mañana"
    r"|(?:la|esta|esa|el|este|pr[oó]xim[oa]|la\s+pr[oó]xima|la\s+que\s+viene)"
    r"\s+(?:semana|tarde|noche|ma[ñn]ana|finde|fin\s+de\s+semana|mes|a[ñn]o|d[ií]a)"
    r"(?:\s+que\s+viene)?"
    # Location follow-up: "en X"
    r"|en\s+\w+(?:\s+\w+)?"
    # Weather attribute follow-up: "cuánta lluvia/nieve" (con cualquier
    # trailing — "cuánta lluvia se espera?")
    r"|cu[aá]nta?\s+(?:lluvia|nieve|tormenta|llovizna|fr[ií]o|calor)"
    r")"
    # Trailing chars hasta cap razonable — permite "se espera?", "habrá?",
    # etc. sin saltar a frases largas que ya no son anafóricas.
    r"\b.{0,30}\s*[?!.]?\s*$",
    re.IGNORECASE,
)


# 2026-04-28 wave-8 (P3 #10): patrones meta/referenciales de follow-up.
# Distintos del temporal — acá el user pregunta SOBRE la lista del turno
# previo en vez de cambiar la ventana temporal. Ejemplos del eval Conv 2:
#
#   T1: "qué eventos esta semana?" → calendar_ahead
#   T2: "y de eso qué puedo postergar?" → ❌ off-topic (sin fix)
#
# Match patterns:
#   - "y de eso/esos/esas/esto/estas"
#   - "cuál podría/puedo/podés/sería/es el más X"
#   - "los demás", "el primero", "el último", "el de X"
#   - "el más urgente/importante/...", "qué prioridad...", "qué es lo más X"
#
# Cuando matchea: re-fire el read tool del turno previo Y propagamos su
# output renderizado al CONTEXTO con un header "DATOS DEL TURNO ANTERIOR".
_ANAPHORIC_REFERENCE_RE = re.compile(
    r"^\s*"
    r"(?:y\s+)?"
    r"(?:"
    # "de eso/eso(s)/esto/estas/aquellos" — referencia directa al turno previo
    r"(?:de|sobre|en)\s+(?:eso|esos?|esto|estas?|aquellos?|aquellas?)"
    # "los/las anteriores", "los demás", "el primero", "el último"
    r"|(?:los|las)\s+(?:anteriores?|dem[aá]s|primeros?|[uú]ltimos?)"
    r"|el\s+(?:primero|[uú]ltimo|m[aá]s\s+\w+|de\s+\w+)"
    # "cuál es/podría/sería/podés/puedo el más X" / "cuál podría posponer"
    r"|cu[aá]l\s+(?:es|podr[ií]a|ser[ií]a|pod[eé]s|puedo|deber[ií]a|tendr[ií]a)"
    # "qué tan X", "qué X es lo más Y", "qué es lo más X"
    r"|qu[eé]\s+(?:tan|es\s+lo\s+m[aá]s|es\s+m[aá]s)"
    # "cuáles son los más X", "cuáles puedo X"
    r"|cu[aá]les\s+(?:son|puedo|podr[ií]a|deber[ií]a)"
    r")"
    r"\b.{0,80}\s*[?!.]?\s*$",
    re.IGNORECASE,
)


def _resolve_anaphoric_args(tool_name: str, question: str, last_location: str | None) -> dict:
    """Compute args para el tool re-fire en follow-up anafórico.

    Para `weather` específicamente: si la query menciona una nueva location
    (ej. "y en Barcelona?"), usa esa; sino usa la `last_location` del turno
    previo. Esto preserva el contexto de la ciudad sin re-prompt.

    Para otros tools (calendar_ahead, reminders_due, etc.) los args default
    son fine — el tool ya respeta horizonte de "hoy/esta semana" interno.
    """
    args: dict = {}
    if tool_name == "weather":
        # Detectar si query menciona nueva location.
        m = re.search(r"\ben\s+([A-ZÁÉÍÓÚÑa-záéíóúñ][\wáéíóúñÁÉÍÓÚÑ\s]{1,40}?)(?=\s*[?!.]?\s*$)", question, re.IGNORECASE)
        if m:
            args["location"] = m.group(1).strip().rstrip(",.;:")
        elif last_location:
            args["location"] = last_location
    return args


def _detect_weather_explicit_location_intent(q: str) -> tuple[str, dict] | None:
    """Detect 'clima/tiempo en CIUDAD' patterns. Returns ('weather', {'location': X})
    si matchea, None si no. Cuando matchea, debe forzar SOLO weather sin
    disparar morning-brief tools.

    Conservative pero no demasiado: keyword ('clima'/'tiempo'/'pronóstico')
    + (opcional palabras intermedias) + preposición ('en'/'de'/'para')
    + ciudad. Acepta lowercase ("buenos aires") y mayúscula ("BA").
    """
    if not q or not q.strip():
        return None
    m = _WEATHER_LOCATION_RE.search(q)
    if not m:
        return None
    loc = (m.group("loc") or "").strip().rstrip(",.;:")
    # Strip trailing common time words si quedaron pegados al match
    for tail in (" hoy", " mañana", " ahora", " ahorita"):
        if loc.lower().endswith(tail):
            loc = loc[: -len(tail)].strip()
    # 2026-04-28 wave-4: strip trailing preposiciones que quedaron pegadas
    # cuando el regex non-greedy capturó "Mendoza para" en "clima en Mendoza
    # para mañana". Una prep al final NUNCA es parte del nombre de ciudad.
    for prep in (" para el", " para", " de", " en", " hasta"):
        if loc.lower().endswith(prep):
            loc = loc[: -len(prep)].strip()
    if not loc or len(loc) < 2:
        return None
    return ("weather", {"location": loc})


# Eval 2026-04-28 BUG-3: queries "qué WA programados" se ruteaban a
# whatsapp_pending (incoming) en vez de whatsapp_list_scheduled
# (outgoing). Causa: la palabra "mañana" matcheaba _PLANNING_PAT del
# whatsapp_pending. Fix: helper específico que detecta keywords claros
# de "mensajes que YO programé" y fuerza whatsapp_list_scheduled.
_WA_LIST_SCHEDULED_TRIGGER_RE = re.compile(
    r"\b(?:programad\w*|scheduled|"
    r"(?:quedan|que\s+quedan)\s+por\s+mandar|"
    r"pendientes?\s+de\s+mandar|"
    r"qu[eé]\s+(?:program[eé]|tengo\s+programado\w*))\b",
    re.IGNORECASE,
)


def _detect_whatsapp_list_scheduled_intent(q: str) -> tuple[str, dict] | None:
    """Detect 'qué mensajes WA programados / quedan por mandar' patterns.

    OUTGOING que YO programé (símil a whatsapp_list_scheduled), NO confundir
    con whatsapp_pending (INCOMING sin contestar). Conservative: solo dispara
    con keywords explícitos.

    Returns ('whatsapp_list_scheduled', {}) si matchea, None si no.
    """
    if not q or not q.strip():
        return None
    if not _WA_LIST_SCHEDULED_TRIGGER_RE.search(q):
        return None
    # Sanity: la query debe tocar el dominio WhatsApp. Aceptamos plurales
    # ("wsps", "wzps") y variantes como "whatsapp"/"whatsapps".
    if not re.search(r"\b(?:whats\w*|wzps?|wsps?|mensajes?)\b", q, re.IGNORECASE):
        return None
    return ("whatsapp_list_scheduled", {})


def _detect_tool_intent(q: str) -> list[tuple[str, dict]]:
    """Deterministic keyword → tool routing. Returns (name, args) tuples
    to execute BEFORE the LLM tool-deciding call. Empty list = no forced
    tools (LLM decides freely).

    `drive_search` y `whatsapp_search` reciben la query cruda como
    `query` arg — los tool helpers tokenizan / fan-out internamente.
    All other tools have static args from their rule entry.
    """
    if not q:
        return []
    # read_note tiene prioridad: si el user pidió explícitamente leer
    # una nota específica, devolvemos eso solo y skippeamos las regex
    # genéricas (que podrían disparar reminders_due / calendar_ahead /
    # etc. spuriamente sobre la misma frase).
    _read = _detect_read_note_intent(q)
    if _read is not None:
        return [_read]
    # weather con location explícita tiene prioridad: skipea morning-brief
    # spurioso disparado por "hoy"/"mañana" en _PLANNING_PAT.
    _wx = _detect_weather_explicit_location_intent(q)
    if _wx is not None:
        return [_wx]
    # whatsapp_list_scheduled tiene prioridad: keywords claros de "outgoing
    # que YO programé" no deben disparar whatsapp_pending por "mañana".
    _wa_sched = _detect_whatsapp_list_scheduled_intent(q)
    if _wa_sched is not None:
        return [_wa_sched]
    out: list[tuple[str, dict]] = []
    for name, args, rx in _TOOL_INTENT_COMPILED:
        if not rx.search(q):
            continue
        if name == "drive_search":
            out.append((name, {"query": q}))
        elif name == "whatsapp_search":
            out.append((name, {"query": q}))
        else:
            out.append((name, dict(args)))
    return out


# Metadata por tool source-specific, usada para componer el hint de
# "intención explícita" que se le pasa al LLM cuando el pre-router
# disparó un tool. Campos:
#   label         — cómo nombramos la fuente al user ("tus mails/correos").
#   live_section  — header de la sección fresca en CONTEXTO que renderea
#                   `_format_forced_tool_output` (p.ej. "### Mails").
#   digest_hint   — dónde más buscar items indexados en el CONTEXTO si la
#                   live section está vacía. Para mails: las notas de
#                   03-Resources/Gmail/YYYY-MM-DD.md usan un `## <asunto>`
#                   por mail con From/Date/Snippet debajo — extraer esos
#                   H2 da un listado crudo de "mis últimos mails" que el
#                   user espera. Otros sources tienen su propio formato.
#   item_shape    — ejemplo del formato que debe usar cada bullet en la
#                   respuesta final, para que el LLM no invente prosa
#                   cuando el user pidió un listado.
#   empty_phrase  — frase explícita cuando NO hay nada ni live ni en
#                   digest. Reemplaza el vago "te dejo otras fuentes".
#
# Weather / finance_summary NO están acá porque no son "fuentes" que el
# user busca — son resúmenes autogenerados sin concepto de ausencia.
_SOURCE_INTENT_META: dict[str, dict[str, str]] = {
    "gmail_recent": {
        "label": "tus mails/correos",
        "live_section": "### Mails",
        "digest_hint": (
            "Si en el CONTEXTO hay notas del vault del tipo "
            "`03-Resources/Gmail/YYYY-MM-DD.md` (cada `## <asunto>` "
            "dentro es UN mail, con su **From:**, **Date:** y **Snippet:**), "
            "extraé esos asuntos y listálos uno por línea — son LITERALMENTE "
            "los últimos mails del usuario. NO digas 'en tu nota' ni "
            "menciones la ruta de la nota: los asuntos SON los mails."
        ),
        "item_shape": "- <asunto> (de <remitente>)",
        "empty_phrase": "No encontré mails recientes en tu corpus",
    },
    "calendar_ahead": {
        "label": "tu calendario/agenda/eventos",
        "live_section": "### Calendario",
        "digest_hint": (
            "Si en el CONTEXTO hay notas con eventos (morning brief, "
            "agenda del día), extraé los títulos de los eventos y listálos."
        ),
        "item_shape": "- <título> (<fecha/hora>)",
        "empty_phrase": "No tenés eventos en el horizonte",
    },
    "reminders_due": {
        "label": "tus recordatorios/pendientes",
        "live_section": "### Recordatorios",
        "digest_hint": (
            "Si en el CONTEXTO hay notas que mencionan tareas pendientes "
            "(morning/evening brief, PARA projects), extraelas y listálas."
        ),
        "item_shape": "- <tarea> (<fecha si tiene>)",
        "empty_phrase": "No tenés recordatorios pendientes",
    },
    "drive_search": {
        "label": "tu Google Drive",
        "live_section": "### Google Drive",
        "digest_hint": (
            "La sección live trae los archivos encontrados con su body "
            "exportado. Si el user pidió un dato concreto (precio, deuda, "
            "cantidad), citálo TEXTUAL del body si está; si no aparece, "
            "decí explícitamente que buscaste y no encontraste ese dato."
        ),
        "item_shape": "- <nombre del archivo> (<tipo>) · <dato relevante o 'sin match'>",
        "empty_phrase": "No encontré nada en tu Google Drive que matchee",
    },
    "whatsapp_pending": {
        "label": "tus chats de WhatsApp esperando respuesta",
        "live_section": "### WhatsApp",
        "digest_hint": (
            "La sección live trae los chats donde el user debe el próximo "
            "mensaje (último inbound sin reply). Si en el CONTEXTO hay "
            "notas de `03-Resources/WhatsApp/<contacto>/YYYY-MM.md` con "
            "más contexto de esos chats, podés complementar. NUNCA "
            "inventes conversaciones de WhatsApp — si la sección live "
            "está vacía decilo explícitamente en vez de citar otras "
            "fuentes como si fueran WA."
        ),
        "item_shape": "- <contacto> (hace <Xh/d>): <último mensaje>",
        "empty_phrase": "No hay chats de WhatsApp esperando tu respuesta",
    },
    "whatsapp_search": {
        "label": "tus mensajes de WhatsApp (búsqueda por contenido)",
        "live_section": "### WhatsApp",
        "digest_hint": (
            "La sección live trae los mensajes WhatsApp matcheantes a la "
            "query, ordenados por relevancia. Cada bullet tiene "
            "`[<contacto> · <fecha>] <snippet>`; si arranca con `yo →` el "
            "mensaje lo mandó el user, no el contacto. NUNCA inventes "
            "conversaciones — citá TEXTUAL de los snippets que aparecen, "
            "y si la sección live está vacía decilo explícitamente."
        ),
        "item_shape": "- <contacto> (<fecha>): <cita textual del snippet>",
        "empty_phrase": "No encontré mensajes de WhatsApp que matcheen tu búsqueda",
    },
}


# Retrocompatibilidad: el helper previo `_SOURCE_INTENT_LABEL` es un
# mapping más chico (label + section) que usan los tests directos.
# Lo mantenemos derivado de `_SOURCE_INTENT_META` para no romper imports
# viejos ni tener dos sources of truth que puedan divergir.
_SOURCE_INTENT_LABEL: dict[str, tuple[str, str]] = {
    name: (meta["label"], meta["live_section"])
    for name, meta in _SOURCE_INTENT_META.items()
}


def _build_source_intent_hint(forced_tool_names: list[str]) -> str | None:
    """Compone un system message turn-scoped que le dice al LLM cómo
    responder cuando el user preguntó explícitamente por una fuente
    concreta (mails / calendario / recordatorios).

    El hint combina:

    1. Dónde buscar primero (la sección live del tool: "### Mails").
    2. Dónde buscar si la live está vacía (notas indexadas del vault con
       formato conocido — p.ej. 03-Resources/Gmail/YYYY-MM-DD.md tiene
       un H2 por mail, listar esos H2 == listar los últimos mails).
    3. Formato de respuesta esperado (viñetas, shape por item).
    4. Qué PROHIBIR explícitamente (decir "tus notas" / "otras fuentes" /
       "te dejo esto por si ayuda" — vocabulario abstracto que no le
       sirve al user cuando pidió una lista concreta).
    5. Frase canned cuando NO hay nada (reemplaza el vago "te dejo
       otras fuentes").

    Motivación histórica (2026-04-24, user report iter 1-3):

    - Iter 1: regex plurales faltaba → `gmail_recent` no disparaba → el
      sistema respondía con WhatsApp sin reconocer intent.
    - Iter 2: con el tool disparando pero vacío, el CONTEXTO se reemplazaba
      → LLM sin material para fallback → "te dejo otras fuentes" abstracto.
    - Iter 3 (este): con el CONTEXTO preservado, el LLM tiene notas de
      `03-Resources/Gmail/*.md` disponibles, PERO hablaba de "tu nota
      del 22 de abril" y "fuentes" en lugar de extraer los asuntos de
      los mails. User feedback textual: "en vez de fuentes (que no tiene
      sentido porque son notas de obsidian) trae los titulos de los
      mails". Este hint ahora explicita el formato deseado.

    Devuelve None si ninguna tool es source-specific (solo weather o
    finance_summary). En ese caso no hay hint que agregar y el system
    prompt default alcanza.
    """
    metas = [_SOURCE_INTENT_META[n] for n in forced_tool_names
             if n in _SOURCE_INTENT_META]
    if not metas:
        return None

    def _join(parts: list[str]) -> str:
        if len(parts) == 1:
            return parts[0]
        if len(parts) == 2:
            return f"{parts[0]} y {parts[1]}"
        return ", ".join(parts[:-1]) + f" y {parts[-1]}"

    labels = [m["label"] for m in metas]
    sections = [m["live_section"] for m in metas]
    joined_labels = _join(labels)
    joined_sections = _join(sections)
    digest_block = "\n".join(f"  • {m['digest_hint']}" for m in metas)
    shape_block = "\n".join(f"  • {m['item_shape']}" for m in metas)
    empty_phrase = _join([m["empty_phrase"] for m in metas])

    return (
        f"INTENCIÓN EXPLÍCITA DEL USUARIO: pidió {joined_labels}. "
        f"Tu tarea es devolver un LISTADO CONCRETO, no un resumen abstracto "
        f"ni una referencia a 'fuentes' del sistema.\n\n"
        f"ORDEN DE BÚSQUEDA en el CONTEXTO:\n"
        f"  1. Sección live {joined_sections}: si tiene items, listálos.\n"
        f"  2. Si la sección live está vacía, buscá en el resto del "
        f"CONTEXTO data indexada de esta fuente:\n{digest_block}\n\n"
        f"FORMATO DE RESPUESTA — lista con viñetas, un item por línea:\n"
        f"{shape_block}\n\n"
        f"PROHIBIDO:\n"
        f"  • Decir 'en tu nota X', 'en tus fuentes', 'te dejo otras "
        f"fuentes que podrían ayudarte', 'revisá tus notas' — el user "
        f"pidió los items directamente, no una meta-referencia.\n"
        f"  • Mencionar rutas del vault (`03-Resources/...`, "
        f"`04-Archive/...`) ni el sistema PARA.\n"
        f"  • Resumir en prosa cuando la pregunta exige un listado.\n"
        f"  • Responder sobre WhatsApp u otras fuentes como si fueran la "
        f"respuesta principal cuando el usuario pidió {joined_labels}.\n\n"
        f"SI NO HAY DATA NI LIVE NI INDEXADA: respondé exactamente "
        f"'{empty_phrase}' — sin agregar sugerencias de fallback."
    )


# Empty-output detection por tool — shape-aware. Usado para decidir si el
# pre-router debe REEMPLAZAR el CONTEXTO (comportamiento original cuando
# el tool trae data fresca) o PRESERVARLO como fallback (cuando el tool
# vino vacío y las notas retrieveadas pueden servir).
#
# Motivación (2026-04-24, iteración 2 del mismo user report): tras el fix
# de plurales + hint, el user preguntó "cuales son mis ultimos mails?"
# y el sistema respondió "Busqué en tus mails y no encontré nada, te
# dejo otras fuentes que podrían ayudarte" — correcto en el
# reconocimiento de intent, pero GENÉRICO en el fallback. Causa: el pre-
# router reemplazó el CONTEXTO entero con el output de gmail_recent
# (que vino vacío → sólo "_Sin mails pendientes._"), descartando la
# retrieve del vault que había devuelto `03-Resources/Gmail/2026-04-22.md`
# (el digest de mails indexado). Sin material en CONTEXTO, el LLM
# resolvió el fallback con una frase abstracta. Este helper permite
# detectar empty-state y preservar el vault retrieve en ese caso.
def _is_empty_tool_output(name: str, raw: str) -> bool:
    """True si el tool devolvió un shape que semánticamente significa
    'no hay nada'. Empty-test por tool:

    - ``gmail_recent`` → ``threads == []`` AND ``unread_count == 0``.
    - ``calendar_ahead`` → lista vacía.
    - ``reminders_due`` → ``dated == []`` AND ``undated == []``.
    - ``credit_cards_summary`` → lista vacía (sin xlsx en /Finances).
    - ``finance_summary`` / ``weather`` → siempre ``False`` (tienen output
      útil aun cuando los números son todos cero — un mes sin gastos es
      data válida, no ausencia de data).

    Malformed JSON (no debería pasar — los tools escriben shapes bien
    definidos) devuelve ``False`` conservativamente: si no puedo parsear,
    no asumo que es empty; dejo que el LLM decida con el string crudo.
    """
    try:
        data = json.loads(raw)
    except (json.JSONDecodeError, TypeError, ValueError):
        return False
    if name == "gmail_recent":
        if not isinstance(data, dict):
            return False
        threads = data.get("threads") or []
        try:
            unread = int(data.get("unread_count") or 0)
        except (TypeError, ValueError):
            unread = 0
        return not threads and unread == 0
    if name == "calendar_ahead":
        return isinstance(data, list) and not data
    if name == "credit_cards_summary":
        return isinstance(data, list) and not data
    if name == "reminders_due":
        if not isinstance(data, dict):
            return False
        return (
            not (data.get("dated") or [])
            and not (data.get("undated") or [])
        )
    if name == "drive_search":
        # Empty = no files AND an error OR no files without an error (both
        # mean "nothing to show"). Auth errors also count as empty so the
        # fallback CONTEXTO-preserve branch kicks in.
        if not isinstance(data, dict):
            return False
        return not (data.get("files") or [])
    if name == "whatsapp_pending":
        # Shape: list of chat dicts. Empty list = no chats waiting for reply.
        return isinstance(data, list) and not data
    if name == "whatsapp_search":
        # Shape: dict con `messages` list. Empty list = no matches encontrados
        # en el corpus WA. Igual que drive_search: errores también cuentan
        # como "empty" para que el fallback CONTEXTO-preserve kick-in.
        if not isinstance(data, dict):
            return False
        return not (data.get("messages") or [])
    return False


# ── Forced-tool output renderer (2026-04-22, Fer F. report) ──────────
# Before: the pre-router dumped raw JSON under a `## {tool_name}` header
# into the CONTEXTO block. qwen2.5:7b reacted badly — dropped `undated`
# items, invented reminders that weren't in the feed, and occasionally
# seeded citation artifacts like `[[calendar_ahead]]` because the tool
# name leaked as a wikilink-ish token.
#
# The helper below renders each forced-tool result as tidy markdown the
# LLM can cite without inventing: Spanish bucket labels, explicit empty
# states, dedup by (name, due), no tool-name leak, graceful fallback on
# malformed JSON (so a tool exception never crashes the request). Plays
# nicely with REGLA 1 ("engancháte SIEMPRE con el CONTEXTO") because the
# block lives INSIDE the CONTEXTO slot the LLM is pinned on.
#
# Shapes assumed (see web/tools.py):
#   reminders_due  → {"dated": [{name,due,bucket,list}], "undated":[...]}
#   calendar_ahead → [{title, date_label, time_range}]
#   gmail_recent   → {"unread_count": int, "threads": [{kind,from,...}]}
#   finance_summary→ dict with variable fields (passthrough JSON, pretty)
#   weather        → plain string (already friendly)
#   unknown tool   → header + raw content (name allowed as disambiguator)
_BUCKET_ES: dict[str, str] = {
    "overdue": "vencido",
    "today": "hoy",
    "upcoming": "próximo",
}


def _format_forced_tool_output(name: str, raw: str) -> str:
    """Render one forced-tool result as markdown for the CONTEXTO block.

    Never raises — malformed / non-JSON input falls through to a raw
    passthrough under the tool's Spanish section header. Tool name is
    NEVER leaked for known tools (prevents `[[calendar_ahead]]` citation
    artifacts observed in production). Unknown tools include the name
    as a disambiguator since we can't invent a friendly label.
    """
    if name == "reminders_due":
        return _format_reminders_block(raw)
    if name == "calendar_ahead":
        return _format_calendar_block(raw)
    if name == "gmail_recent":
        return _format_gmail_block(raw)
    if name == "finance_summary":
        return _format_finance_block(raw)
    if name == "credit_cards_summary":
        return _format_cards_block(raw)
    if name == "weather":
        return _format_weather_block(raw)
    if name == "drive_search":
        return _format_drive_block(raw)
    if name == "whatsapp_pending":
        return _format_whatsapp_block(raw)
    if name == "whatsapp_search":
        return _format_whatsapp_search_block(raw)
    # Unknown tool → keep raw JSON available but wrap in a labeled
    # section so it doesn't merge visually with the next block.
    return f"### Datos ({name})\n{raw}\n"


def _format_reminders_block(raw: str) -> str:
    """Render Apple Reminders JSON as two sub-sections: 'Con fecha' and
    'Sin fecha'. Dedupes by (name, due) — AppleScript occasionally returns
    the same reminder twice across recurring-instance boundaries.
    """
    try:
        data = json.loads(raw)
    except (json.JSONDecodeError, TypeError, ValueError):
        return f"### Recordatorios\n{raw}\n"
    if not isinstance(data, dict):
        return f"### Recordatorios\n{raw}\n"

    def _dedup(items):
        seen: set[tuple[str, str]] = set()
        out: list[dict] = []
        for it in items or []:
            if not isinstance(it, dict):
                continue
            nm = str(it.get("name", "")).strip()
            du = str(it.get("due", "")).strip()
            if not nm:
                continue
            key = (nm, du)
            if key in seen:
                continue
            seen.add(key)
            out.append(it)
        return out

    dated = _dedup(data.get("dated"))
    undated = _dedup(data.get("undated"))

    if not dated and not undated:
        return "### Recordatorios\n_Sin recordatorios pendientes._\n"

    # Header con counts explícitos. Pre-fix qwen2.5:7b a veces contaba
    # mal ("tres tareas para llamar al dentista" cuando había una sola);
    # emitir N/M literal en el header le da al modelo un ancla numérica
    # para no alucinar totales. Medido 2026-04-23 en scratch_eval.
    header_bits: list[str] = []
    if dated:
        header_bits.append(f"{len(dated)} con fecha")
    if undated:
        header_bits.append(f"{len(undated)} sin fecha")
    header = f"### Recordatorios ({', '.join(header_bits)})"
    lines: list[str] = [header]
    if dated:
        lines.append("**Con fecha:**")
        for it in dated:
            nm = str(it.get("name", "")).strip()
            due = str(it.get("due", "")).strip()
            bucket = str(it.get("bucket", "")).strip().lower()
            date_part, _, time_part = due.partition("T")
            time_str = time_part[:5] if time_part else ""
            stamp = f"{date_part} {time_str}".strip() if date_part else ""
            tag_es = _BUCKET_ES.get(bucket, "")
            # Only prefix a tag for overdue/today (urgency signal).
            # Upcoming is the default forward-looking state; leaving it
            # unmarked reduces noise without losing information.
            prefix = f"[{tag_es}] " if tag_es in ("vencido", "hoy") else ""
            body = f"{stamp} · {nm}" if stamp else nm
            lines.append(f"- {prefix}{body}")
    if undated:
        lines.append("**Sin fecha:**")
        for it in undated:
            nm = str(it.get("name", "")).strip()
            lines.append(f"- {nm}")
    return "\n".join(lines) + "\n"


# Mapeo de date_labels en inglés (formato icalBuddy) → español. icalBuddy
# emite "today" / "tomorrow" / "day after tomorrow" para fechas relativas
# corto plazo, y YYYY-MM-DD o "lunes, abril 28, 2026" para fechas lejanas.
# El LLM (qwen2.5:7b) traducía "tomorrow" a "mañana" pero NO se daba
# cuenta de la inconsistencia con la pregunta del user ("hoy"). Pre-fix
# 2026-04-28 (Playwright autónomo): pregunta "qué tengo agendado para
# hoy específicamente" → respuesta "- Psiquiatra (mañana)" sin
# disclaimer de que NO había nada hoy. Forzar el label en español + tag
# relativo explícito le da al LLM el anclaje temporal que le faltaba.
_CALENDAR_LABEL_ES = {
    "today": "HOY",
    "tomorrow": "MAÑANA",
    "day after tomorrow": "PASADO MAÑANA",
}


def _translate_calendar_label(date_label: str, time_range: str = "") -> str:
    """Normaliza un date_label de icalBuddy a español y, cuando es un
    label relativo, ANCLA el tag absoluto entre corchetes para que el LLM
    no confunda "tomorrow" con "hoy". Retorna string vacío si no hay
    label útil.
    """
    if not date_label:
        return ""
    low = date_label.strip().lower()
    if low in _CALENDAR_LABEL_ES:
        # Tag explícito en MAYÚSCULAS — llama atención del LLM y se
        # mantiene incluso si el modelo lo cita literal.
        return _CALENDAR_LABEL_ES[low]
    # icalBuddy también puede devolver "Wednesday, April 30, 2026" o un
    # YYYY-MM-DD plano para fechas más lejanas. No reformateamos —
    # devolvemos como está; el LLM puede leerlo igual.
    return date_label.strip()


def _format_calendar_block(raw: str, now: datetime | None = None) -> str:
    """Render calendar events as a bulleted list. Does NOT dedup — the
    same recurring event across multiple days must show up N times so the
    user sees 'cumpleaños de Astor' twice if it falls on two rendered
    dates.

    Header incluye la fecha de "hoy" (DD/MM/YYYY + día de la semana en
    español) para que el LLM tenga anclaje temporal explícito al
    interpretar items con date_label relativo. Pre-2026-04-28 sólo se
    emitía "### Calendario (N eventos)" sin fecha → el LLM no podía
    distinguir "hoy" en la pregunta del user vs "today/tomorrow" en los
    items, y caía en respuestas crudas tipo "- Psiquiatra (mañana)" para
    "qué tengo hoy".
    """
    try:
        data = json.loads(raw)
    except (json.JSONDecodeError, TypeError, ValueError):
        return f"### Calendario\n{raw}\n"
    if not isinstance(data, list):
        return f"### Calendario\n{raw}\n"

    # Día actual para el header — usamos el now inyectado (testeable) o
    # datetime.now() del huso local. _DAYS_ES traduce el nombre del
    # weekday inglés (datetime.strftime("%A")) a español rioplatense.
    if now is None:
        now = datetime.now()
    _DAYS_ES = {
        "Monday": "lunes", "Tuesday": "martes", "Wednesday": "miércoles",
        "Thursday": "jueves", "Friday": "viernes", "Saturday": "sábado",
        "Sunday": "domingo",
    }
    today_str = f"{_DAYS_ES.get(now.strftime('%A'), now.strftime('%A').lower())} {now.strftime('%d/%m/%Y')}"

    if not data:
        return f"### Calendario (hoy: {today_str})\n_Sin eventos en el horizonte._\n"

    # Count explícito en header (ver comentario en _format_reminders_block).
    valid_events = [ev for ev in data if isinstance(ev, dict) and str(ev.get("title", "")).strip()]
    header = (
        f"### Calendario (hoy: {today_str} · "
        f"{len(valid_events)} evento{'s' if len(valid_events) != 1 else ''})"
    )
    lines: list[str] = [header]
    has_today = False
    for ev in valid_events:
        title = str(ev.get("title", "")).strip()
        date_label = _translate_calendar_label(
            str(ev.get("date_label", "")).strip(),
            str(ev.get("time_range", "")).strip(),
        )
        time_range = str(ev.get("time_range", "")).strip()
        if date_label == "HOY":
            has_today = True
        parts: list[str] = []
        if date_label:
            parts.append(date_label)
        if time_range:
            parts.append(time_range)
        prefix = " ".join(parts)
        if prefix:
            lines.append(f"- {prefix} · {title}")
        else:
            lines.append(f"- {title}")
    # Hint explícito al LLM cuando la lista NO tiene eventos para HOY.
    # Sin esto, el LLM ve "- MAÑANA · Psiquiatra" y lo cita literal sin
    # aclarar que la respuesta a "qué tengo hoy" es "nada hoy".
    if valid_events and not has_today:
        lines.append("_(no hay eventos hoy; los items de arriba son posteriores)_")
    return "\n".join(lines) + "\n"


# 2026-04-28 wave-8: noise detectors para gmail_recent. Agrupar CI/security/
# marketing reduce ruido en el CONTEXTO y pistas el LLM hacia los mails que
# importan ("tenés 4 mails de CI fallando + 1 de OSDE + 1 personal" en vez
# de listar cada notificación con su SHA).
_GMAIL_GITHUB_BRACKET_RE = re.compile(r"^\s*\[[\w.-]+/[\w.-]+\]")  # [user/repo]
_GMAIL_CI_RE = re.compile(
    r"\b(?:run\s+(?:failed|cancelled|succeeded|skipped)|ci\s+(?:failed|passed)|"
    r"build\s+(?:failed|passed|succeeded)|workflow\s+run|pipeline)\b",
    re.IGNORECASE,
)
_GMAIL_SECURITY_RE = re.compile(
    r"\b(?:security\s+alert|vulnerability|cve-?\d{4}|dependabot|advisory|"
    r"alerta\s+de\s+seguridad)\b",
    re.IGNORECASE,
)
_GMAIL_MARKETING_RE = re.compile(
    r"\b(?:newsletter|unsubscribe|promo(?:ci[oó]n)?|oferta|"
    r"discount|sale|black\s+friday|cyber\s+monday)\b",
    re.IGNORECASE,
)
_GMAIL_NOREPLY_RE = re.compile(
    # Conservative: solo no-reply explícito + alias claramente automatizados.
    # `info@` se quitó porque OSDE/Galicia/etc usan `info@org` para
    # notificaciones legítimas que al user le importan (turnos médicos,
    # avisos bancarios). Mejor un false negative en automated que
    # ocultar un mail real.
    r"(?:no.?reply|noreply|donot.?reply|news(?:letter)?@|marketing@|"
    r"updates?@|notifications?@)",
    re.IGNORECASE,
)


def _classify_gmail_thread(frm: str, subj: str) -> str:
    """Clasifica un thread de gmail en una bucket de ruido.

    Returns "ci", "security", "marketing", "automated", o "personal".
    Personal es el bucket default — los mails que NO matchean ninguno de
    los patrones de ruido son los que importan al user.
    """
    full = f"{frm} {subj}".strip()
    # Order matters: CI antes que automated porque CI es más específico.
    if _GMAIL_GITHUB_BRACKET_RE.search(subj) or _GMAIL_CI_RE.search(full):
        return "ci"
    if _GMAIL_SECURITY_RE.search(full):
        return "security"
    if _GMAIL_MARKETING_RE.search(full):
        return "marketing"
    if _GMAIL_NOREPLY_RE.search(frm):
        return "automated"
    return "personal"


def _format_gmail_block(raw: str) -> str:
    """Render Gmail evidence as 'Mails' section with awaiting-reply and
    starred threads as bullets. Keeps the tool name out of the output.

    2026-04-28 wave-8: agrupar mails de noise (CI, security alerts, marketing,
    automated notifications) en summary lines. El user pidió "qué mails
    pendientes tengo" y antes recibía 4 líneas separadas con commit SHAs +
    `[jagoff/rag-obsidian]` brackets. Ahora resume "4 mails de CI: 2 fail,
    1 ok, 1 cancelled" + lista solo los mails personales.
    """
    try:
        data = json.loads(raw)
    except (json.JSONDecodeError, TypeError, ValueError):
        return f"### Mails\n{raw}\n"
    if not isinstance(data, dict):
        return f"### Mails\n{raw}\n"
    threads = data.get("threads") or []
    unread = int(data.get("unread_count") or 0)
    if not threads and unread == 0:
        return "### Mails\n_Sin mails pendientes._\n"
    # Count explícito en header (ver comentario en _format_reminders_block).
    valid_threads = [t for t in threads if isinstance(t, dict)]
    header_bits: list[str] = []
    if valid_threads:
        header_bits.append(f"{len(valid_threads)} hilo{'s' if len(valid_threads) != 1 else ''}")
    if unread:
        header_bits.append(f"{unread} no leído{'s' if unread != 1 else ''}")
    header = f"### Mails ({', '.join(header_bits)})" if header_bits else "### Mails"
    lines = [header]

    # Agrupar threads por bucket. Personal va listado uno por uno; los
    # demás se resumen en una sola línea con count.
    buckets: dict[str, list[dict]] = {
        "ci": [], "security": [], "marketing": [],
        "automated": [], "personal": [],
    }
    for t in valid_threads:
        frm = str(t.get("from", "")).strip()
        subj = str(t.get("subject", "")).strip()
        bucket = _classify_gmail_thread(frm, subj)
        buckets[bucket].append(t)

    # Resumen de noise (siempre primero — los grupos compactos arriba).
    noise_labels = {
        "ci": "CI/builds",
        "security": "alertas de seguridad",
        "marketing": "marketing/newsletters",
        "automated": "notificaciones automáticas",
    }
    for bucket_name in ("ci", "security", "marketing", "automated"):
        items = buckets[bucket_name]
        if not items:
            continue
        n = len(items)
        label = noise_labels[bucket_name]
        # Para CI agregar el desglose por outcome (failed/passed/cancelled).
        if bucket_name == "ci":
            outcomes = {"failed": 0, "passed": 0, "cancelled": 0, "other": 0}
            for it in items:
                s = str(it.get("subject", "")).lower()
                if "failed" in s or "fail:" in s or " fail" in s:
                    outcomes["failed"] += 1
                elif "passed" in s or "succeeded" in s or "success" in s:
                    outcomes["passed"] += 1
                elif "cancelled" in s or "skipped" in s:
                    outcomes["cancelled"] += 1
                else:
                    outcomes["other"] += 1
            outcome_bits = [f"{v} {k}" for k, v in outcomes.items() if v > 0]
            lines.append(f"- {n} {label} ({', '.join(outcome_bits)})")
        else:
            lines.append(f"- {n} {label}")

    # Mails personales: uno por línea con tag de bucket si corresponde.
    for t in buckets["personal"]:
        frm = str(t.get("from", "")).strip()
        subj = str(t.get("subject", "")).strip()
        kind = str(t.get("kind", "")).strip()
        tag = {
            "awaiting_reply": "esperando respuesta",
            "starred": "starred",
            "recent": "",
        }.get(kind, kind)
        tag_str = f" [{tag}]" if tag else ""
        lines.append(f"- {frm} · {subj}{tag_str}")
    return "\n".join(lines) + "\n"


def _format_finance_block(raw: str) -> str:
    """Finance summary is a dict with month, totals, top categories. Render
    under 'Gastos' section. Passthrough pretty JSON if shape is unexpected
    — the LLM handles a raw dict fine when there's a section header.
    """
    try:
        data = json.loads(raw)
    except (json.JSONDecodeError, TypeError, ValueError):
        return f"### Gastos\n{raw}\n"
    if not isinstance(data, dict) or not data:
        return "### Gastos\n_Sin datos financieros._\n"
    try:
        pretty = json.dumps(data, ensure_ascii=False, indent=2)
    except (TypeError, ValueError):
        pretty = raw
    return f"### Gastos\n```json\n{pretty}\n```\n"


def _format_weather_block(raw: str) -> str:
    """Render weather JSON output as human-readable markdown for the LLM.

    Eval 2026-04-28 BUG-2a: el LLM no mencionaba la ciudad porque la
    location quedaba enterrada en el JSON. Fix: parsear y exponer un
    text_summary "Ciudad: descripción (temp°C)" prominente.
    """
    body = (raw or "").strip()
    if not body:
        return "### Clima\n_Sin datos del clima._\n"
    try:
        data = json.loads(body)
    except (json.JSONDecodeError, ValueError):
        # No es JSON parseable, dejar el raw como antes.
        return f"### Clima\n{body}\n"

    if not isinstance(data, dict):
        return f"### Clima\n{body}\n"

    # Si el output ya tiene text_summary (post-fix), usarlo como header.
    summary = data.get("text_summary")
    if not summary and data.get("current"):
        loc = data.get("location", "Ubicación no especificada")
        cur = data["current"]
        desc = cur.get("description", "Sin datos")
        temp = cur.get("temp_C") or cur.get("temp")
        summary = f"{loc}: {desc} ({temp}°C)" if temp else f"{loc}: {desc}"

    # Render: summary (prominente) + JSON full por si el LLM quiere los días.
    out = "### Clima\n"
    if summary:
        out += f"{summary}\n"
    out += body  # mantiene el JSON crudo para el LLM si necesita días
    out += "\n"
    return out


def _format_cards_block(raw: str) -> str:
    """Credit cards summary es una lista de dicts (uno por tarjeta). Renderea
    una sub-sección por tarjeta con saldo a pagar (ARS/USD), vencimiento,
    cierre y top 3 consumos. Es passthrough en formato humano para que el
    LLM cite los números directamente sin hacer su propia matemática.
    """
    try:
        cards = json.loads(raw)
    except (json.JSONDecodeError, TypeError, ValueError):
        return f"### Tarjetas\n{raw}\n"
    if not isinstance(cards, list) or not cards:
        return "### Tarjetas\n_Sin resúmenes de tarjeta disponibles._\n"

    def _fmt_ars(n: float | None) -> str:
        if n is None:
            return "—"
        v = int(round(abs(n)))
        s = f"{v:,}".replace(",", ".")
        return f"${s}"

    def _fmt_usd(n: float | None) -> str:
        if n is None:
            return "—"
        return f"U$S{abs(n):,.2f}"

    lines = ["### Tarjetas"]
    for c in cards:
        if not isinstance(c, dict):
            continue
        brand = c.get("brand") or "Tarjeta"
        last4 = c.get("last4") or "----"
        title = f"{brand} ····{last4}"
        lines.append(f"- **{title}**")
        # Saldo a pagar
        total_bits = []
        if c.get("total_ars"):
            total_bits.append(_fmt_ars(c["total_ars"]))
        if c.get("total_usd"):
            total_bits.append(_fmt_usd(c["total_usd"]))
        if total_bits:
            lines.append(f"  - Total a pagar: {' + '.join(total_bits)}")
        # Mínimo (solo si difiere del total — banks suelen igualarlos)
        min_bits = []
        if c.get("minimum_ars") and c.get("minimum_ars") != c.get("total_ars"):
            min_bits.append(_fmt_ars(c["minimum_ars"]))
        if c.get("minimum_usd") and c.get("minimum_usd") != c.get("total_usd"):
            min_bits.append(_fmt_usd(c["minimum_usd"]))
        if min_bits:
            lines.append(f"  - Mínimo: {' + '.join(min_bits)}")
        # Fechas: vencimiento es la accionable
        if c.get("due_date"):
            lines.append(f"  - Vence: {c['due_date']}")
        if c.get("closing_date"):
            lines.append(f"  - Cierre: {c['closing_date']}")
        if c.get("next_due_date"):
            lines.append(f"  - Próximo venc.: {c['next_due_date']}")
        # Top consumos: ARS + USD (3 + 2 max para no inundar)
        ars_top = c.get("top_purchases_ars") or []
        usd_top = c.get("top_purchases_usd") or []
        if ars_top or usd_top:
            lines.append("  - Top consumos:")
            for p in ars_top[:3]:
                desc = p.get("description") or "—"
                lines.append(f"    - {desc} · {_fmt_ars(p.get('amount'))}")
            for p in usd_top[:2]:
                desc = p.get("description") or "—"
                lines.append(f"    - {desc} · {_fmt_usd(p.get('amount'))}")
    return "\n".join(lines) + "\n"


def _format_drive_block(raw: str) -> str:
    """Render `drive_search` output as a `### Google Drive` section.

    Shape esperado del tool (web/tools.drive_search → rag._agent_tool_drive_search):
        {tokens: [...], query_used: "a b c", files: [{name, mime_label,
         modified, link, body}], error?: "..."}

    Por archivo: header con nombre + tipo + fecha + link (un solo bullet),
    seguido de un bloque de body entre líneas en blanco. Cap defensivo del
    body a 2500 chars acá (aunque el helper ya lo capea a 3500) por si
    algún día elevamos el cap en la tool pero no queremos explotar el
    CONTEXTO acá. Auth / API errors se renderizan como "_Sin resultados
    (motivo: ...)._" para que el LLM lo comunique al user honestamente
    en vez de inventar contenido.
    """
    try:
        data = json.loads(raw)
    except (json.JSONDecodeError, TypeError, ValueError):
        return f"### Google Drive\n{raw}\n"
    if not isinstance(data, dict):
        return f"### Google Drive\n{raw}\n"

    files = data.get("files") or []
    tokens = data.get("tokens") or []
    err = data.get("error") or ""
    query_used = data.get("query_used") or ""

    header = "### Google Drive"
    if query_used:
        header = f"{header} (búsqueda: {query_used})"

    if not files:
        reason_bits: list[str] = []
        if err == "no_google_credentials":
            reason_bits.append("auth de Drive no configurada")
        elif err.startswith("search_failed"):
            reason_bits.append("falló la API de Drive")
        elif err == "query vacía después de filtrar stopwords":
            reason_bits.append("la pregunta no tenía keywords buscables")
        elif err:
            reason_bits.append(err)
        else:
            reason_bits.append(
                f"ningún archivo matchea {tokens!r}" if tokens
                else "ningún archivo matchea la búsqueda"
            )
        return f"{header}\n_Sin resultados ({'; '.join(reason_bits)})._\n"

    # Defensive body cap (ver docstring).
    BODY_CAP = 2500

    lines: list[str] = [header]
    for f in files:
        if not isinstance(f, dict):
            continue
        name = str(f.get("name", "")).strip() or "(sin nombre)"
        mime = str(f.get("mime_label", "")).strip() or "archivo"
        modified = str(f.get("modified", "")).strip()
        link = str(f.get("link", "")).strip()
        date_part = modified.split("T")[0] if modified else ""
        bits: list[str] = [f"**{name}**", f"({mime})"]
        if date_part:
            bits.append(f"· modificado {date_part}")
        if link:
            bits.append(f"· [abrir]({link})")
        lines.append("- " + " ".join(bits))
        body = str(f.get("body", "")).strip()
        if body:
            lines.append("")
            lines.append(body[:BODY_CAP])
            lines.append("")
    return "\n".join(lines) + "\n"


def _format_whatsapp_block(raw: str) -> str:
    """Render `whatsapp_pending` output as a `### WhatsApp` section.

    Shape esperado: list of `{jid, name, last_snippet, hours_waiting}`
    (ver `_fetch_whatsapp_unreplied`). Formatea cada chat como bullet
    con contacto bold, tiempo esperando humanizado ("hace 3h" / "hace
    2d"), y el snippet del último mensaje. Dedupe por (jid, snippet)
    porque el shape del fetcher ya dedupea por jid — redundante pero
    defensivo ante shape drift.
    """
    try:
        data = json.loads(raw)
    except (json.JSONDecodeError, TypeError, ValueError):
        return f"### WhatsApp\n{raw}\n"
    if not isinstance(data, list):
        return f"### WhatsApp\n{raw}\n"
    if not data:
        return "### WhatsApp\n_Sin chats esperando tu respuesta._\n"

    chats = [c for c in data if isinstance(c, dict) and str(c.get("name", "")).strip()]
    header = f"### WhatsApp ({len(chats)} chat{'s' if len(chats) != 1 else ''} esperando respuesta)"
    lines: list[str] = [header]
    for chat in chats:
        name = str(chat.get("name", "")).strip()
        snippet = str(chat.get("last_snippet", "")).strip()
        try:
            hours = float(chat.get("hours_waiting") or 0)
        except (TypeError, ValueError):
            hours = 0.0
        # Humanize wait time: <24h → "Xh", ≥24h → "Xd".
        if hours < 1:
            age = "recién"
        elif hours < 24:
            age = f"hace {int(hours)}h"
        else:
            age = f"hace {int(hours / 24)}d"
        line = f"- **{name}** ({age})"
        if snippet:
            line += f": {snippet}"
        lines.append(line)
    return "\n".join(lines) + "\n"


def _format_whatsapp_search_block(raw: str) -> str:
    """Render `whatsapp_search` output as a `### WhatsApp` section with
    one bullet per matched message.

    Shape esperado: ``{query, contact_filter, messages: [{jid, contact,
    ts, who, text, score}], warning?, error?}`` (ver
    `_agent_tool_whatsapp_search`).

    Cada bullet sigue el shape ``- [<contacto> · <fecha>] <snippet>``;
    si el chunk es outbound (sender == "yo") prefija ``yo →`` para que
    el LLM diferencie quién dijo qué. Snippet se renderea capado
    defensivamente a 300 chars (el tool ya lo capa a 400, pero el LLM
    digestivo prefiere bullets más cortos cuando hay 5–8 hits).

    Empty / errored results muestran un mensaje explícito en vez de
    raw JSON — same pattern que `_format_drive_block` para que el LLM
    no invente conversaciones cuando el corpus no devolvió match.
    """
    try:
        data = json.loads(raw)
    except (json.JSONDecodeError, TypeError, ValueError):
        return f"### WhatsApp\n{raw}\n"
    if not isinstance(data, dict):
        return f"### WhatsApp\n{raw}\n"

    messages = data.get("messages") or []
    contact_filter = data.get("contact_filter")
    err = data.get("error") or ""
    warning = data.get("warning") or ""

    # Empty path — surface why so the LLM can tell the user.
    if not messages:
        reason_bits: list[str] = []
        if err == "query vacía":
            reason_bits.append("la pregunta vino vacía")
        elif err.startswith("retrieve_failed"):
            reason_bits.append("falló el retrieval")
        elif err:
            reason_bits.append(err)
        else:
            if contact_filter:
                reason_bits.append(f"sin matches para {contact_filter!r}")
            else:
                reason_bits.append("sin matches en el corpus de WhatsApp")
        if warning and warning not in reason_bits:
            reason_bits.append(warning)
        return f"### WhatsApp\n_Sin resultados ({'; '.join(reason_bits)})._\n"

    valid = [m for m in messages if isinstance(m, dict) and str(m.get("text", "")).strip()]
    header_bits: list[str] = [f"{len(valid)} mensaje{'s' if len(valid) != 1 else ''}"]
    if contact_filter:
        header_bits.append(f"contacto: {contact_filter}")
    header = f"### WhatsApp ({', '.join(header_bits)})"

    lines: list[str] = [header]
    if warning:
        lines.append(f"_Aviso: {warning}_")

    SNIPPET_CAP = 300  # Defensive — tool capa a 400, pero bullets cortos = LLM más limpio.
    for m in valid:
        contact = str(m.get("contact", "")).strip() or "(sin contacto)"
        ts = str(m.get("ts", "")).strip()
        who = str(m.get("who", "")).strip().lower()
        text = str(m.get("text", "")).strip()
        # Date-only — el LLM se confunde con timestamps largos en
        # bullets; la hora exacta vive en el JSON crudo si lo necesita.
        date_part = ts.split("T")[0] if ts else ""
        prefix = f"[{contact}"
        if date_part:
            prefix += f" · {date_part}"
        prefix += "]"
        # Outbound = lo dijo el user; importante distinguirlo para que
        # el LLM no diga "Juan te dijo X" cuando el "X" lo dijiste vos.
        speaker_marker = "yo → " if who == "outbound" else ""
        snippet = text[:SNIPPET_CAP].replace("\n", " ⏎ ")
        lines.append(f"- {prefix} {speaker_marker}{snippet}")
    return "\n".join(lines) + "\n"


# Create-intent detection moved to rag.py so both the web chat endpoint
# and the CLI `rag chat` loop can share it without inverting the
# web → rag import direction. See rag.py `_detect_propose_intent` for
# the full regex list + three-branch logic.
from rag import _detect_propose_intent, _detect_metachat_intent  # noqa: E402


# Canned replies for the meta-chat short-circuit. Buckets keyed by the
# class of input; within a bucket we pick one variant by hashing the
# message + the current minute so the same phrase in a tight window is
# stable (no user surprise if they re-send) but repeat visits pick
# different variants (feels alive, not scripted).
_METACHAT_GREETING = (
    "¡Hola! Preguntame lo que quieras sobre tus notas, o decime *recordame …* / *agendá …* si querés crear algo.",
    "Hola 👋 ¿en qué te ayudo? Probá una pregunta sobre tus notas o pedime *recordame …* / *agendá …*.",
    "¡Buenas! Tirame una pregunta, o decime *recordame X* / *el viernes 20hs X* para crear un recordatorio o evento.",
)
_METACHAT_THANKS = (
    "¡De nada!",
    "¡Cuando quieras!",
    "👌",
)
_METACHAT_META = (
    "Puedo responder sobre tus notas, crear recordatorios (*recordame X*) y eventos de calendar (*el viernes 20hs …*). Probá algo.",
    "Consulto tu vault de Obsidian y creo recordatorios / eventos desde texto libre. ¿Qué necesitás?",
    "Leo tus notas y armo recordatorios o eventos cuando se los pedís en lenguaje natural. Tirame algo concreto.",
)
_METACHAT_BYE = (
    "¡Hasta luego!",
    "¡Nos vemos!",
    "👋",
)


def _metachat_bucket(q: str) -> tuple[str, ...]:
    """Classify the meta-chat message into a response bucket."""
    s = q.strip().lower().lstrip("¿ ").lstrip()
    if s.startswith(("gracias", "muchas gracias", "mil gracias",
                     "dale gracias", "thanks", "thx")):
        return _METACHAT_THANKS
    if s.startswith(("chau", "bye", "adiós", "adios", "nos vemos")):
        return _METACHAT_BYE
    if s.startswith(("qué podés", "que podes", "qué sabés", "que sabes",
                     "cómo funcion", "como funcion", "cómo te us",
                     "como te us", "qué comandos", "que comandos",
                     "ayuda", "help", "quién sos", "quien sos",
                     "quién es este", "quien es este")):
        return _METACHAT_META
    return _METACHAT_GREETING


def _pick_metachat_reply(q: str, *, now: float | None = None) -> str:
    """Pick a canned reply for a meta-chat turn.

    Variation seed = hash(q) XOR minute-bucket. Same input within the
    same minute returns the same variant (stable on retry); different
    inputs or different minutes rotate. Tests monkey-patch with fixed
    `now` for determinism.
    """
    import hashlib
    import time as _time
    bucket = _metachat_bucket(q)
    ts = now if now is not None else _time.time()
    minute = int(ts // 60)
    seed = int(hashlib.sha256(f"{q}|{minute}".encode()).hexdigest()[:8], 16)
    return bucket[seed % len(bucket)]


# ── Degenerate query short-circuit (2026-04-23) ──────────────────────
# Queries con <2 caracteres alfanuméricos (ej. "x", "?", "?¡@#") caían
# al retrieve + rerank y devolvían chunks random de WhatsApp porque el
# matching semántico sobre un input casi vacío es puro ruido. Medido en
# scratch_eval: `?¡@#` → 395 chars de contenido WA sin relación. Ahora
# devolvemos una respuesta canned antes de tocar retrieve/LLM,
# invitando al usuario a reformular.
_DEGENERATE_REPLIES: tuple[str, ...] = (
    "No entendí tu pregunta. Podés reformularla con más detalle?",
    "Necesito un poco más de contexto. Qué querés consultar de tus notas?",
    "Tu mensaje parece muy corto o sin contenido. Preguntame algo concreto sobre el vault.",
)


def _is_degenerate_query(q: str) -> bool:
    """True si la query tiene <2 caracteres alfanuméricos totales.

    Evita que `"x"`, `"?"`, `"?¡@#"`, strings de puro símbolo, o cadenas
    vacías disparen el pipeline full — no hay suficiente señal para que
    el retrieve devuelva algo útil ni para que el LLM produzca una
    respuesta honesta. Metachat tiene su propio short-circuit; esta
    función se encarga de lo que no alcanza ni siquiera a ser saludo.
    """
    if not q or not q.strip():
        return True
    alphanum = sum(1 for c in q if c.isalnum())
    return alphanum < 2


def _pick_degenerate_reply(q: str, *, now: float | None = None) -> str:
    """Pick a canned reply for a degenerate-query turn. Same seeding as
    metachat (hash(q) XOR minute-bucket) so retries stay stable.
    """
    import hashlib
    import time as _time
    ts = now if now is not None else _time.time()
    minute = int(ts // 60)
    seed = int(hashlib.sha256(f"{q}|{minute}".encode()).hexdigest()[:8], 16)
    return _DEGENERATE_REPLIES[seed % len(_DEGENERATE_REPLIES)]


# Post-stream filter enforcing REGLA 0 at the byte level. qwen2.5:7b and
# command-r leak CJK / Cyrillic / Arabic tokens under context pressure; the
# chat path strips them from every streamed delta so the client never sees
# them regardless of what the LLM emits. Whitelists ASCII, Latin (including
# Spanish diacritics and the Latin-Extended blocks), box-drawing, general
# punctuation used in markdown, and everything ≥ U+1F000 (emoji + symbols).
# Kills CJK ideographs + kana + hangul, Cyrillic, Hebrew, Arabic, Syriac,
# and the fullwidth/halfwidth CJK punctuation block.
_FOREIGN_SCRIPT_RE = re.compile(
    "["
    "\u0400-\u052f"     # Cyrillic + Cyrillic Supplement
    "\u0530-\u058f"     # Armenian
    "\u0590-\u05ff"     # Hebrew
    "\u0600-\u06ff"     # Arabic
    "\u0700-\u074f"     # Syriac
    "\u0750-\u077f"     # Arabic Supplement
    "\u0780-\u07bf"     # Thaana
    "\u3000-\u303f"     # CJK Symbols and Punctuation (includes 。、)
    "\u3040-\u309f"     # Hiragana
    "\u30a0-\u30ff"     # Katakana
    "\u3400-\u4dbf"     # CJK Unified Ideographs Extension A
    "\u4e00-\u9fff"     # CJK Unified Ideographs
    "\uac00-\ud7af"     # Hangul Syllables
    "\uff00-\uffef"     # Halfwidth and Fullwidth Forms
    "]"
)


def _strip_foreign_scripts(text: str) -> str:
    """Remove characters from non-allowed scripts (CJK, Cyrillic, Hebrew,
    Arabic, …). Preserves Spanish diacritics, ASCII, markdown punctuation,
    and emoji (≥U+1F000). Safe to call per-token on a streaming delta —
    characters are dropped individually (no stateful lookahead)."""
    if not text:
        return text
    return _FOREIGN_SCRIPT_RE.sub("", text)


def _own_conversation_path(session_id: str) -> str | None:
    """Look up the episodic conversation .md for `session_id` in
    rag_conversations_index. Used for the self-citation exclusion guard in
    /api/chat. SQL-only since T10 (the old conversations_index.json path is
    gone)."""
    try:
        from web.conversation_writer import get_conversation_path
        return get_conversation_path(session_id)
    except Exception:
        return None


STATIC_DIR = Path(__file__).resolve().parent / "static"


# Lifespan handler para reemplazar `@app.on_event("startup"/"shutdown")`
# (deprecated en FastAPI 0.93+, será eliminado en una mayor futura).
# 2026-04-24 audit hardening: las 4 callbacks que estaban dispersas
# por el archivo (warmup home cache, drain conversation writers,
# memory sampler, cpu sampler) se registran ahora vía decorators
# `@_on_startup` / `@_on_shutdown` que populan estas listas. La
# lifespan async-iterates para preservar el contract original (orden
# de registro = orden de ejecución, idéntico al behavior de los 4
# `on_event` antes).
#
# Beneficio operacional: cero deprecation warnings en cada test run
# (eran 4 + 16 propagated = 20 warnings/run que ensuciaban el output).
# Beneficio estructural: lifespan maneja excepciones de startup
# uniformemente — antes una callback que raisea bloqueaba el server
# silently; ahora la traza queda visible.
_startup_callbacks: "list[Callable[[], None]]" = []
_shutdown_callbacks: "list[Callable[[], None]]" = []


def _on_startup(fn: "Callable[[], None]") -> "Callable[[], None]":
    """Decorator: registra `fn` para correr al startup. Reemplazo
    drop-in de `@app.on_event("startup")`."""
    _startup_callbacks.append(fn)
    return fn


def _on_shutdown(fn: "Callable[[], None]") -> "Callable[[], None]":
    """Decorator: registra `fn` para correr al shutdown. Reemplazo
    drop-in de `@app.on_event("shutdown")`."""
    _shutdown_callbacks.append(fn)
    return fn


from contextlib import asynccontextmanager
from typing import Callable, AsyncIterator


@asynccontextmanager
async def _lifespan(_app) -> "AsyncIterator[None]":
    """FastAPI lifespan: corre todos los callbacks registrados via
    `_on_startup` antes del primer request, y los de `_on_shutdown` al
    apagar el server. Errores en startup callbacks NO bloquean el
    server (catch + log); errores en shutdown se loguean y siguen.
    """
    for fn in _startup_callbacks:
        try:
            fn()
        except Exception as exc:
            print(f"[lifespan-startup] {fn.__name__} falló: {exc}", flush=True)
    yield
    for fn in _shutdown_callbacks:
        try:
            fn()
        except Exception as exc:
            print(f"[lifespan-shutdown] {fn.__name__} falló: {exc}", flush=True)


app = FastAPI(
    title="obsidian-rag web", docs_url=None, redoc_url=None,
    lifespan=_lifespan,
)


# Cache-Control para /static/* — assets inmutables en prod (versionados por
# fichero, reload del server pushea bytes nuevos al disk). Pre-2026-04-22 no
# había header → cada reload del browser refetcheaba el bundle entero
# (~50-100ms por asset, peor en mobile/WiFi lento). Con `max-age=3600` el
# browser cachea 1h y la navegación entre páginas (home/dashboard/chat) es
# instant. `public` permite que proxies intermediate cacheen (irrelevante en
# localhost pero es higiene). ETag lo maneja StaticFiles automáticamente
# vía If-None-Match → con cache válido el response es 304 sin body.
#
# Trade-off: si el user pushea un hotfix a `web/static/app.js` y tiene el
# browser abierto, no lo ve hasta que expire el cache o haga hard-reload
# (⌘⇧R). Para dev usar `OBSIDIAN_RAG_STATIC_NO_CACHE=1` que setea max-age=0.
class _CachedStaticFiles(StaticFiles):
    """StaticFiles subclass que agrega Cache-Control a cada response."""

    def __init__(self, *args, max_age: int = 3600, **kwargs):
        super().__init__(*args, **kwargs)
        self._max_age = max_age

    async def get_response(self, path, scope):
        response = await super().get_response(path, scope)
        if hasattr(response, "headers"):
            response.headers["Cache-Control"] = (
                f"public, max-age={self._max_age}"
                if self._max_age > 0
                else "no-cache, no-store, must-revalidate"
            )
        return response


_STATIC_MAX_AGE = 0 if os.environ.get("OBSIDIAN_RAG_STATIC_NO_CACHE") == "1" else 3600
app.mount(
    "/static",
    _CachedStaticFiles(directory=STATIC_DIR, max_age=_STATIC_MAX_AGE),
    name="static",
)

# ── /memory: mem-vault UI mounted as a sub-application ───────────────────
# Boot mem-vault's FastAPI app once (Qdrant + mem0 are expensive) and let
# it serve everything under /memory/*. Same origin as the rest of the web
# UI so links + cookies + CSP work without surprises. Best-effort: if the
# mem-vault package isn't installed in this venv (it lives in a separate
# repo, optional dep) we just skip the mount and the link in the navbar
# will 404 — easy to spot but doesn't crash the rest of the server.
try:
    from mem_vault.ui.server import create_app as _mem_vault_create_app  # type: ignore[import-not-found]

    _mem_vault_app = _mem_vault_create_app()
    app.mount("/memory", _mem_vault_app, name="mem-vault-ui")
    print("[mem-vault] UI mounted at /memory", file=sys.stderr, flush=True)
except Exception as _exc:  # pragma: no cover — best-effort wiring
    print(
        f"[mem-vault] UI not mounted ({_exc}). "
        "Install with `uv pip install --editable /path/to/mem-vault[ui,hybrid]` "
        "if you want /memory to work.",
        file=sys.stderr,
        flush=True,
    )

# CORS: same-origin only. The server is bound to 127.0.0.1 by the
# launchd plist, so cross-origin requests would come from a browser
# running an untrusted origin (e.g. a malicious local webpage trying
# to hit our API). FastAPI without CORSMiddleware rejects cross-origin
# by default, but adding the middleware explicitly:
#   1. documents the intent for future maintainers who might otherwise
#      assume CORS wasn't thought about,
#   2. whitelists the exact origins we DO serve (127.0.0.1 / localhost
#      on any port — the plist uses 8765 but dev sometimes uses 8766),
#   3. prevents a future contributor from dropping in
#      `CORSMiddleware(allow_origins=["*"])` thinking "we need CORS",
#      which would open the API to every page in the browser.
from fastapi.middleware.cors import CORSMiddleware  # noqa: E402 — must go after `app = FastAPI(...)`

# LAN IP privada (192.168.x.x / 10.x.x.x / 172.16-31.x.x) se permite sólo
# cuando OBSIDIAN_RAG_ALLOW_LAN=1 — empareja con OBSIDIAN_RAG_BIND_HOST=
# 0.0.0.0 para exponer la PWA al iPhone vía `http://192.168.x.x:8765` o
# `https://192.168.x.x:8765` (Caddy tls internal). Sin el flag el regex
# se mantiene igual que antes (localhost only).
#
# Bug fix 2026-04-27: http → https? en ambas ramas para no bloquear Caddy
# tls internal (LAN) ni HTTPS local (localhost:8765 vía Caddy).
#
# OBSIDIAN_RAG_ALLOW_TUNNEL=1 añade las URLs de Cloudflare Quick Tunnel
# (`https://<slug>.trycloudflare.com`) al regex para que el iPhone pueda
# usar el SW + full PWA sobre HTTPS público. Se empareja con la env var
# que propaga el watcher (`cloudflared_watcher.sh`). Default OFF — la URL
# random de trycloudflare.com es unguessable pero expone el server al
# internet público (sin auth). Solo activar junto con el túnel.
_allow_lan = os.environ.get("OBSIDIAN_RAG_ALLOW_LAN", "").strip().lower() in ("1", "true", "yes")
_allow_tunnel = os.environ.get("OBSIDIAN_RAG_ALLOW_TUNNEL", "").strip().lower() in ("1", "true", "yes")
if _allow_lan:
    # RFC1918 ranges: 10.0.0.0/8, 172.16.0.0/12, 192.168.0.0/16. Nada más.
    # https? cubre http (LAN plain) y https (Caddy tls internal).
    # Refuses file:// y cualquier IP pública.
    _cors_regex = (
        r"^https?://("
        r"127\.0\.0\.1|localhost|"
        r"10\.\d{1,3}\.\d{1,3}\.\d{1,3}|"
        r"172\.(1[6-9]|2\d|3[01])\.\d{1,3}\.\d{1,3}|"
        r"192\.168\.\d{1,3}\.\d{1,3}"
        r")(:[0-9]+)?$"
    )
else:
    _cors_regex = r"^https?://(127\.0\.0\.1|localhost)(:[0-9]+)?$"
# Cloudflare Quick Tunnel: cualquier subdominio *.trycloudflare.com sobre
# HTTPS solamente. La URL cambia en cada restart de cloudflared pero el
# patrón es estable — solo HTTPS, nunca HTTP (Cloudflare fuerza TLS).
if _allow_tunnel:
    _cors_regex = (
        r"(?:" + _cors_regex + r")"
        r"|^https://[a-z0-9-]+\.trycloudflare\.com$"
    )

app.add_middleware(
    CORSMiddleware,
    allow_origin_regex=_cors_regex,
    allow_credentials=True,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["*"],
)

# Chat model for /chat. Delegated to `resolve_chat_model()` so it tracks
# whatever rag.py decides is best on this host (command-r > qwen2.5:14b >
# phi4). command-r:35b prefill is slower (~5-10s on a 5k-char context)
# but is the only model in the local set that reliably:
#   - stays in Spanish instead of drifting to Portuguese / Italian
#   - resolves pronouns from conversation history ("ella", "eso") against
#     prior turns instead of treating follow-ups as standalone
#   - follows the "no inline citations / no refusal" rules of this prompt
# qwen2.5:3b was used here for a while as a speed optimization (~600ms
# prefill, ~10× faster) but failed those three constraints consistently.
# Set OBSIDIAN_RAG_WEB_CHAT_MODEL to override if you need speed over quality.
WEB_CHAT_MODEL = os.environ.get("OBSIDIAN_RAG_WEB_CHAT_MODEL") or None

# Runtime model override — persisted to disk so it survives server restarts.
# Takes priority over env / resolve_chat_model() so the user can A/B different
# models from the UI without editing the launchd plist or redeploying.
# Format: {"model": "qwen3.6", "set_at": "2026-04-20T15:00:00"}. Absent or
# malformed → no override active.
_CHAT_MODEL_OVERRIDE_PATH = Path.home() / ".local/share/obsidian-rag" / "chat-model.json"


def _read_chat_model_override() -> str | None:
    """Load the persistent runtime override if any. Silent on any error —
    a corrupt or missing file means "no override, use defaults"."""
    try:
        raw = _CHAT_MODEL_OVERRIDE_PATH.read_text(encoding="utf-8")
        data = json.loads(raw)
        model = (data or {}).get("model")
        if isinstance(model, str) and model.strip():
            return model.strip()
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        pass
    return None


def _write_chat_model_override(model: str | None) -> None:
    """Persist or clear the runtime override. None → delete the file."""
    _CHAT_MODEL_OVERRIDE_PATH.parent.mkdir(parents=True, exist_ok=True)
    if model is None:
        _CHAT_MODEL_OVERRIDE_PATH.unlink(missing_ok=True)
        return
    payload = {"model": model,
               "set_at": datetime.now(timezone.utc).isoformat(timespec="seconds")}
    tmp = _CHAT_MODEL_OVERRIDE_PATH.with_suffix(_CHAT_MODEL_OVERRIDE_PATH.suffix + ".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    os.replace(tmp, _CHAT_MODEL_OVERRIDE_PATH)


def _resolve_web_chat_model() -> str:
    """Model resolution priority (first match wins):
      1. Runtime override file (_CHAT_MODEL_OVERRIDE_PATH) — set via
         POST /api/chat/model from the UI; persists across restarts.
      2. OBSIDIAN_RAG_WEB_CHAT_MODEL env var — for launchd plist overrides.
      3. rag.resolve_chat_model() — the CHAT_MODEL_PREFERENCE fallback chain.
    Not cached: the override file is cheap to re-read (JSON, <200 bytes)
    and we want runtime flips to take effect on the next /api/chat call
    with no server restart.
    """
    override = _read_chat_model_override()
    if override:
        return override
    return WEB_CHAT_MODEL or resolve_chat_model()


# Module-level system prompt. Kept BYTE-IDENTICAL across all /api/chat
# requests so ollama's prefix cache hits — when the system message tokens
# match a prior request, llama.cpp skips KV recomputation for those ~900
# tokens. Cached prefill is ~50× faster (0.23ms/tok vs 12.7ms/tok measured
# on command-r:35b), saving up to 10s on the first user-facing token.
# Any per-request variation (context, signals, history) goes in the user
# message AFTER this static block.
# v1: versión original 5231 chars (~1307 tok). Preservada como fallback
# documentado vía env var RAG_WEB_PROMPT_VERSION=v1 — útil si al medir
# v2 en producción aparece regresión en algún caso no cubierto por el
# bench. Retirar cuando v2 sume suficiente historia sin incidentes.
_WEB_SYSTEM_PROMPT_V1 = (
    "Eres un asistente de consulta sobre las notas personales de "
    "Obsidian del usuario. NO sos un modelo de conocimiento general.\n\n"
    "REGLA 0 — IDIOMA: respondé SIEMPRE en español rioplatense. "
    "TOTALMENTE PROHIBIDO emitir tokens en chino, japonés, coreano, "
    "árabe, ruso, alemán, portugués, italiano, francés o cualquier "
    "idioma que no sea español — aunque el CONTEXTO contenga chats "
    "de WhatsApp con contactos brasileros, argentinos de otros "
    "países o notas con fragmentos no-español, la respuesta TIENE QUE "
    "estar íntegramente en español rioplatense. Si el contexto "
    "recuperado contiene fragmentos en otros idiomas (ej. citas en "
    "inglés, nombres propios, código), citalos textualmente entre "
    "comillas pero el resto de tu respuesta TIENE QUE estar en "
    "español. Si la pregunta del usuario está en otro idioma, "
    "traducila a español mentalmente y respondé en español. Esta "
    "regla es ABSOLUTA — ni siquiera caracteres sueltos en otro "
    "alfabeto (汉字, русский, etc.) están permitidos en tu output.\n\n"
    "REGLA 1 — ENGÁNCHATE CON EL CONTEXTO QUE RECIBÍS:\n"
    "  • El CONTEXTO abajo es lo que el retriever consideró más "
    "relacionado con la pregunta. Tu trabajo es leerlo y resumir lo "
    "que aporta, siempre. Incluso si los chunks son breves, parciales "
    "o tangenciales — engancháte con ellos.\n"
    "  • Preguntas tipo '¿tengo algo sobre X?' / '¿qué hay de X?' / "
    "'¿tenés notas para X?' se responden afirmativamente apenas el "
    "CONTEXTO mencione X en título o cuerpo. Listá brevemente lo que "
    "hay.\n"
    "  • Si el CONTEXTO es genuinamente pobre o tangencial para lo "
    "que se pregunta, describí lo que sí aparece y aclaralo en prosa "
    "('las notas que encontré mencionan X pero no detallan Y'). "
    "NUNCA uses fórmulas de refusal del tipo 'no tengo información' "
    "— esa vía de escape está PROHIBIDA en este asistente, siempre "
    "hay que devolver el mejor resumen posible del CONTEXTO.\n"
    "  • Fuera del CONTEXTO no inventes (eso queda para REGLA 3 con "
    "su marcador `<<ext>>`).\n\n"
    "REGLA 2 — NO CITAR NOTAS INLINE: la UI ya muestra la lista de "
    "fuentes recuperadas debajo de cada respuesta (nota, score, "
    "ruta). Por lo tanto TOTALMENTE PROHIBIDO escribir:\n"
    "  • markdown links tipo `[Título](ruta.md)`\n"
    "  • nombres de archivo con extensión (`algo.md`)\n"
    "  • rutas con carpetas PARA (`03-Resources/…`, `02-Areas/…`, etc.)\n"
    "  • el nombre completo de la nota como título dentro de la prosa\n"
    "Podés referirte implícitamente: 'según tus notas', 'en tu nota "
    "sobre X', 'tenés una nota al respecto'. Todo lo demás se ve en "
    "la lista de fuentes — no la dupliques en el cuerpo.\n\n"
    "REGLA 3 — MARCAR EXTERNO: el marcador `<<ext>>...<</ext>>` es "
    "EXCEPCIONAL, no rutinario. Usalo SOLO cuando agregues una de "
    "estas 3 cosas: (a) conocimiento general externo al CONTEXTO "
    "(ej: 'React es una librería de UI de Meta' cuando el CONTEXTO "
    "no tiene React), (b) opinión tuya / inferencia que NO se "
    "deriva del CONTEXTO, (c) link a docs oficiales permitido por "
    "REGLA 4.6. Parafraseo rutinario del CONTEXTO, reordenamientos, "
    "síntesis, conectores ('también', 'además', 'en resumen'), y "
    "resúmenes — TODO eso NO lleva marcador. Si podés defender una "
    "oración como 'esto es lo que dice el CONTEXTO con otras "
    "palabras' o 'es una reorganización de datos del CONTEXTO', NO "
    "la envuelvas. Usar `<<ext>>` en cada oración es un BUG — marca "
    "sólo lo genuinamente externo. Ante duda, NO marques.\n\n"
    "REGLA 4 — FORMATO: 2-4 oraciones o lista corta de viñetas. Dato "
    "clave en la primera oración; contexto mínimo (qué hace, cómo se "
    "invoca) después. Si la pregunta apunta a un comando, "
    "herramienta, parámetro o valor puntual Y el CONTEXTO incluye "
    "su uso (firma, parámetros, ejemplo, en qué MCP/server vive), "
    "ese uso es OBLIGATORIO en la respuesta — no te quedes en un "
    "token mínimo cuando el chunk lo explica.\n\n"
    "REGLA 4.5 — PRESERVAR LINKS DEL CONTENIDO: si el CONTEXTO "
    "contiene URLs (http://, https://) o wikilinks ([[Nota]]) que "
    "viven DENTRO del cuerpo de una nota (no son la ruta del chunk), "
    "copialos LITERAL en la respuesta — son clickeables y útiles para "
    "el usuario. Aclaración respecto a REGLA 2: REGLA 2 prohíbe "
    "citar las notas-fuente (la lista de abajo ya las muestra). Los "
    "links/URLs que aparecen como contenido de una nota son data, "
    "no citas; tienen que aparecer en la respuesta. Ejemplo: si un "
    "chunk dice 'tutorial: https://x.com/y', escribí 'tutorial: "
    "https://x.com/y', no 'tutorial disponible'.\n\n"
    "REGLA 4.6 — LINK A DOCS OFICIALES (raro, MUY acotado): "
    "TOTALMENTE PROHIBIDO para queries sobre personas ('qué sabés "
    "de X', 'hablame de Y'), eventos, recordatorios, mails, gastos, "
    "WhatsApp, calendario o cualquier dato del vault. SOLO aplica "
    "cuando (a) la pregunta nombra EXPLÍCITAMENTE un software / "
    "herramienta / producto externo (ej: 'cómo configuro OmniFocus', "
    "'qué features tiene Obsidian'), (b) el CONTEXTO del vault se "
    "queda corto para lo que se pide, y (c) tenés certeza del dominio "
    "raíz oficial. Formato: una sola línea al final, "
    "`<<ext>>Más info: <dominio-raíz></ext>>`. NUNCA inventes paths "
    "profundos. En TODOS los demás casos NO agregues link externo, "
    "aunque la respuesta quede breve. Ante duda, NO lo incluyas.\n\n"
    "REGLA 5 — SEGUÍ EL HILO: esta es una conversación, no preguntas "
    "sueltas. Los mensajes previos (user/assistant de arriba) son "
    "contexto vivo. Si la pregunta nueva usa pronombres ('ella', "
    "'eso', 'ahí'), referencias elípticas ('y de X?', 'profundizá'), "
    "o asume un tema del turn anterior, resolvé la referencia usando "
    "el historial y respondé sobre ESE tema. No trates la pregunta "
    "como si empezara de cero.\n\n"
    "REGLA 6 — TRATAMIENTO: le hablás DIRECTAMENTE al usuario en 2da "
    "persona singular, tuteo rioplatense ('vos', 'tu', 'tenés', 'te'). "
    "El usuario ES la persona que hace la pregunta — no es un tercero "
    "del que hablás. TOTALMENTE PROHIBIDO escribir 'el usuario', 'del "
    "usuario', 'le' refiriéndose al usuario, ni describirlo en 3ra "
    "persona. Traducí automáticamente: 'la hija del usuario' → 'tu "
    "hija'; 'las notas del usuario' → 'tus notas'; 'el proyecto X "
    "del usuario' → 'tu proyecto X'. Si una nota dice 'Grecia es la "
    "hija de Fernando', y Fernando es el usuario, escribí 'Grecia es "
    "tu hija'.\n\n"
    "REGLA 7 — NO FUSIONAR PERSONAS: si el CONTEXTO menciona varias "
    "personas distintas (ej: una 'María' contacto + otra 'María' de "
    "otro chat + un 'Mario'), NUNCA mezcles sus atributos. Si no "
    "podés distinguir a quién pertenece cada dato, decí explícitamente "
    "'hay varias personas con ese nombre en tus notas' y listá lo más "
    "seguro. PROHIBIDO inventar parentesco ('María es tu hermana/o', "
    "'es tu prima') si el CONTEXTO no lo afirma LITERALMENTE con esa "
    "palabra — si una nota dice 'mi prima María' y otra 'María "
    "Fernández, colega', NO unifiques. Respetá el género y pronombres "
    "como aparecen en cada cita — no los 'corrijas' al género de la "
    "persona preguntada.\n"
)

# v2: comprimido 2898 chars (~724 tok), −44% tokens. Mismas REGLAs
# preservando los anchors visuales que qwen reconoce como patrón
# (`<<ext>>...<</ext>>`, `[Título](ruta.md)`, `03-Resources/`, `[[Nota]]`).
# Medido 2026-04-20: prefill cae 1737ms → ~1100ms en el bench A/B
# gracias al ahorro de ~600 tok del system prompt.
#
# 2026-04-23: endurecemos REGLA 4.6 + agregamos REGLA 7 (anti-fusión de
# personas). El scratch_eval automático mostró que el LLM pegaba
# `<<ext>>Más info: https://omnifocus.com</ext>>` en queries sobre
# personas ("hablame de María", "qué eventos tengo") — leyó la URL del
# prompt como ejemplo copiable. También fusionaba info de 2+ personas
# del CONTEXTO ("María es tu hermano"). Endurecemos REGLA 0 para
# incluir portugués e italiano explícitamente (contagion bajo fast-path
# WA con contactos brasileros).
_WEB_SYSTEM_PROMPT_V2 = 'Eres un asistente de consulta sobre las notas personales de Obsidian del usuario. NO sos un modelo de conocimiento general.\n\nREGLA 0 — IDIOMA: respondé SIEMPRE en español rioplatense. PROHIBIDO emitir tokens en portugués, inglés, italiano, ni otros idiomas/alfabetos (汉字, русский, etc.); caracteres fuera del alfabeto latino sólo se permiten dentro de una cita literal entre comillas. Si el CONTEXTO contiene mensajes en otros idiomas (ej. WhatsApp con contactos brasileros), traducilos al responder. Si la pregunta viene en otro idioma, traducila y respondé en español.\n\nREGLA 1 — ENGANCHÁTE CON EL CONTEXTO: el CONTEXTO de abajo es lo que el retriever consideró más cercano. Resumí SIEMPRE lo que aporta, aun si es breve o tangencial. Preguntas tipo "¿tengo algo sobre X?" se responden afirmativo apenas X aparezca en título o cuerpo — listá brevemente. Si el CONTEXTO es pobre, describí lo que sí aparece ("las notas mencionan X pero no detallan Y"). PROHIBIDO refusal tipo "no tengo información" — siempre devolvé el mejor resumen posible del CONTEXTO. Fuera del CONTEXTO no inventes (ver REGLA 3).\n\nREGLA 1.b — DATOS TRANSACCIONALES FINANCIEROS (excepción ESPECÍFICA y ACOTADA a REGLA 1): SOLO aplica cuando la pregunta nombra EXPLÍCITAMENTE banco/tarjeta/visa/mastercard/amex/MOZE/dolares/pesos/USD/ARS/montos/consumos/movimientos. En esos casos: NUNCA inventes ni copies de notas tangenciales. SOLO cita números/fechas/comercios que aparecen literalmente bajo ### Gastos o ### Tarjetas. Si esas secciones NO están, o están pero vacías, respondé "No tengo data fresca de [X] — el último export del banco puede no estar al día." y CORTÁ ahí. Si hay AMBAS secciones (MOZE + Tarjetas) y la pregunta menciona "tarjeta/visa/master/amex/crédito", priorizá ### Tarjetas. Para CALENDARIO/REMINDERS/MAILS/WHATSAPP/CLIMA/DRIVE — citá literal de la sección correspondiente cuando esté presente, pero NUNCA uses el template de "data fresca/export" (esos no tienen "exports" — los devolvés ya frescos via tool calls).\n\nREGLA 1.b.1 — ALCANCE de REGLA 1.b (PROHIBICIÓN EXPLÍCITA): REGLA 1.b NO aplica a preguntas sobre notas, conceptos, temas, ideas, técnicas, proyectos, conocimiento, conceptos abstractos, métodos, frameworks, libros, autores, citas. Para esas: REGLA 1 (engancháte con el CONTEXTO). El template "No tengo data fresca de [X] — el último export del [...] puede no estar al día" está PROHIBIDO para queries que NO sean financieras explícitas. Si el vault no tiene matches sobre un concepto/técnica/tema, respondé en LENGUAJE NATURAL: "No encontré nada en tus notas sobre [X]. Probá buscar como [variante1] o [variante2], o agregá una nota si querés trackearlo." NO uses la palabra "export", NO digas "no tengo data fresca", NO menciones "el último [algo]".\n\nREGLA 1.c — NO FALSE CONFIRMATIONS (CRÍTICA, security-related): NUNCA digas "se ha programado/agregado/creado/cancelado/eliminado/modificado/actualizado [X]" si NO ejecutaste literalmente la tool correspondiente en este turno. Si el user pide editar/sumar/cambiar/cancelar algo PREVIAMENTE creado y NO tenés tool para hacer ese edit (no existe `propose_reminder_edit`, `propose_calendar_cancel`, etc.), DECILE textualmente: "No puedo editar/cancelar lo anterior desde acá — abrí Apple Reminders/Calendar y modificá manualmente. Si querés, puedo crear uno nuevo con [X] (el viejo queda como está)." PROHIBIDO confirmar acciones que no se ejecutaron.\n\nREGLA 2 — NO CITAR NOTAS INLINE: la UI ya muestra la lista de fuentes (nota, score, ruta) debajo. PROHIBIDO markdown links `[Título](ruta.md)`, nombres con extensión (`algo.md`), rutas PARA (`03-Resources/…`, `02-Areas/…`) ni el título completo como header. Referencias implícitas OK: "según tus notas", "en tu nota sobre X".\n\nREGLA 3 — MARCAR EXTERNO (excepcional, no rutinario): usá `<<ext>>...<</ext>>` SOLO para (a) conocimiento general externo al CONTEXTO (ej: \'React es una librería de UI de Meta\' si el CONTEXTO no tiene React), (b) opinión/inferencia tuya que NO se deriva del CONTEXTO, (c) link a docs oficiales permitido por REGLA 4.6. Parafraseo rutinario, reordenamientos, conectores (\'también\', \'además\', \'en resumen\'), síntesis — TODO eso NO lleva marcador. Marcar cada oración con `<<ext>>` es un BUG. Ante duda, NO marques.\n\nREGLA 4 — FORMATO: 2-4 oraciones o lista corta. Dato clave primero, contexto mínimo (qué hace, cómo se invoca) después. Si piden un comando, herramienta o parámetro Y el CONTEXTO tiene su uso (firma, ejemplo, en qué MCP vive), ese uso es OBLIGATORIO en la respuesta.\n\nREGLA 4.5 — PRESERVAR LINKS DEL CONTENIDO: URLs (http://, https://) y wikilinks ([[Nota]]) que vivan DENTRO del cuerpo de una nota son data, no citas-fuente — copialos LITERAL. REGLA 2 sólo prohíbe citar la ruta del chunk; los links internos son clickeables.\n\nREGLA 4.6 — LINK A DOCS OFICIALES (raro, MUY acotado): TOTALMENTE PROHIBIDO en queries sobre personas ("qué sabés de X", "hablame de Y"), eventos, recordatorios, mails, gastos, WhatsApp, calendar, o cualquier dato del vault. SOLO aplica cuando (a) la pregunta nombra EXPLÍCITAMENTE un software/herramienta/producto externo (ej. "cómo configuro OmniFocus", "qué features tiene Obsidian"), (b) el CONTEXTO del vault se queda corto, y (c) tenés certeza del dominio raíz oficial. Formato: `<<ext>>Más info: <dominio-raíz></ext>>`. En TODOS los demás casos NO agregues link externo, aunque la respuesta sea breve. Ante duda, NO lo incluyas.\n\nREGLA 5 — SEGUÍ EL HILO: es una conversación. Pronombres ("ella", "eso"), referencias elípticas ("y de X?", "profundizá") o temas asumidos se resuelven con los turns previos. No trates la pregunta como si empezara de cero.\n\nREGLA 6 — TRATAMIENTO: hablale DIRECTAMENTE al usuario en 2da persona, tuteo rioplatense ("vos", "tenés", "te"). El usuario ES quien pregunta. PROHIBIDO 3ra persona ("el usuario", "la hija del usuario", "le"). Traducí: "la hija del usuario" → "tu hija"; "las notas del usuario" → "tus notas".\n\nREGLA 7 — NO FUSIONAR PERSONAS: si el CONTEXTO menciona varias personas (ej. una "María" contacto + otra "María" de otro chat + un "Mario"), NUNCA mezcles sus atributos. Si no podés distinguir a quién pertenece cada dato, decí "hay varias personas con ese nombre en tus notas" y listá lo más seguro. PROHIBIDO inventar parentesco ("María es tu hermana/o") si el CONTEXTO no lo afirma LITERALMENTE con esa palabra — si una nota dice "mi prima María" y otra "María Fernández, colega", NO unifiques. Respetá el género/pronombre tal como aparece en cada cita — no los "corrijas" al género preguntado.'

# Selector con fallback seguro a v1 si el env var toma un valor raro.
_WEB_SYSTEM_PROMPT = (
    _WEB_SYSTEM_PROMPT_V1
    if os.environ.get("RAG_WEB_PROMPT_VERSION", "v2").strip() == "v1"
    else _WEB_SYSTEM_PROMPT_V2
)

# Turn-scoped override appended when `_detect_propose_intent` fires.
# Neutralises REGLA 1 (engancháte con CONTEXTO) + REGLA 2/3/4.5 (citar
# notas, preservar links) which don't apply to create-intent turns —
# there's no CONTEXTO attached, there are no notes to cite, there are
# no sources. Without this override qwen2.5:7b reverts to its "resumí
# lo que el retriever trajo" default and hallucinates a path like
# "04-Archive/Calendar.md" to open instead of calling propose_*.
#
# Critical detail learned the hard way: do NOT give the model a literal
# response example in the override. qwen2.5:7b will copy the example
# verbatim and skip the tool call entirely ("Listo, quedó agendado."
# was the literal example → model responded with exactly that text,
# tool_rounds=0). The override now makes the tool call the PRIMARY
# action and describes the response shape abstractly.
_WD_ES = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]


def _build_propose_create_override(today: datetime | None = None) -> str:
    """Build the create-intent system override string with today's date
    injected. The LLM (qwen2.5:7b) is bad at knowing what "today" is
    out-of-the-box — sin esto, pidiéndole "a las 12:55" emite ISO con
    año/mes random ("2028-11-29T12:55:00-03:00" observado 2026-04-28).
    Pinearle la fecha actual al inicio del prompt evita el
    hallucinated-date failure mode.

    `today` se inyecta por request (el caller pasa `datetime.now()`).
    Esto rompe el prefix-cache de ollama por turn (la fecha cambia día
    a día), pero es trade-off aceptable: create-intent turns son raros
    comparado con read-intent, y el cache cold de la primera frase no
    impacta la latencia perceived (el LLM igual se toma 1-2s en
    arrancar a generar tool_calls).

    También enseña al modelo a pasar `scheduled_for` como NL ("a las
    12:55") en lugar de ISO — _validate_scheduled_for ahora hace
    fallback a _parse_natural_datetime, así no le exigimos al LLM que
    haga date math (que históricamente le sale mal).
    """
    if today is None:
        today = datetime.now()
    today_str = f"{_WD_ES[today.weekday()]} {today.strftime('%Y-%m-%d')}"
    now_str = today.strftime("%H:%M")
    iso_today = today.strftime("%Y-%m-%d")

    return (
        f"Hoy es {today_str}, son las {now_str} (Argentina, UTC-3). "
        "Usá esto como anchor para resolver fechas relativas como "
        "'mañana', 'el viernes', 'a las HH:MM'.\n\n"
        "El usuario te pide REGISTRAR algo (recordatorio, evento de "
        "calendario, mensaje de WhatsApp, mail). Las tools disponibles "
        "están en el schema que te pasa Ollama — invocalas como tool_calls "
        "del protocolo, no como texto ni como prosa.\n\n"
        "Criterio de selección:\n"
        "  - propose_calendar_event → visitas, cumpleaños, reuniones, "
        "    turnos médicos, viajes, vacaciones, feriados (cosas con "
        "    fecha/hora que vivirán en Apple Calendar).\n"
        "  - propose_reminder → tareas, pagos, llamadas, cosas para "
        "    acordarse (Apple Reminders).\n"
        "  - propose_whatsapp_send → mandar un mensaje a un tercero. "
        "    Verbos rioplatenses: 'mandale/enviale/decile/escribile a "
        "    <Contacto>'. Ver reglas detalladas abajo.\n"
        "  - propose_whatsapp_reply → responder a UN mensaje específico "
        "    que el user recibió ('respondele/contestale a <X> al "
        "    del...'). Distinto al send.\n"
        "  - propose_whatsapp_send_note → mandar el contenido de una "
        "    nota del vault ('pasale a <X> la receta de Y').\n"
        "  - propose_whatsapp_send_contact_card → mandar el tel/email/"
        "    dir de un contacto ('pasale a <X> el teléfono de <Y>').\n"
        "  - propose_mail_send → mandar un email via Gmail.\n\n"
        "Para el título de eventos/reminders: extraé el sustantivo + "
        "contexto del mensaje (sin fechas, sin horas). Ej: 'cumpleaños "
        "de Astor el viernes' → título 'cumpleaños de Astor'.\n\n"
        "Para la fecha/hora de eventos/reminders: pasala EXACTAMENTE "
        "como la dijo el usuario ('el viernes', 'mañana a las 10', 'el "
        "miercoles'). La tool la parsea con anchor de hoy. Si el "
        "usuario no dio hora, no pasés ningún campo de hora — la tool "
        "detecta sola que es all-day.\n\n"
        # ── Reminder/Calendar slim rules — historia 2026-04-28: intenté
        # agregar reglas detalladas + ejemplos para que el modelo no
        # omita `when` (reminder) ni `recurrence_text="yearly"` (cumples).
        # Resultado: el override se infló a 5300+ chars y el modelo
        # se confundía rutiando reminders a propose_whatsapp_send. La
        # versión minimal abajo es suficiente para que el modelo
        # entienda los basics — las reglas detalladas viven en los
        # docstrings de los tools (que el modelo SÍ ve via el JSON
        # schema). Si en el futuro hace falta endurecer, mejor un
        # override DEDICADO por tipo de propose-intent que uno gigante
        # único — ver `_detect_propose_intent` para discriminar.
        "Para reminders: si el user mencionó CUALQUIER hint temporal "
        "(día concreto como lunes/.../domingo, 'mañana'/'hoy'/'pasado mañana', "
        "hora), pasalo EXACTAMENTE en `when` tal como lo dijo. NO truques al "
        "`title` partes del hint temporal. Si el user dice 'recordame regar "
        "el sábado a las 11am' → title='regar', when='el sábado a las 11am' "
        "(NO when='a las 11am' sin el día). Si no mencionó ningún hint, "
        "dejá `when` vacío.\n\n"
        "Para cumples y aniversarios: usá `propose_calendar_event` "
        "(no reminder) con `all_day` true y `recurrence_text` 'yearly' "
        "para que el evento se repita cada año.\n\n"
        # ── WhatsApp-specific rules — el bug del 2026-04-28 mostró que
        # el docstring de propose_whatsapp_send + reglas genéricas no
        # alcanza con qwen2.5:7b. Hay que ser MUY explícito acá porque
        # el modelo tiende a (a) truncar el body cuando ve palabras
        # tipo "programado"/"urgente" pensando que son meta-descrip-
        # ciones, (b) ignorar "a las HH:MM" como horario de envío, y
        # (c) hallucinar fechas (emitir "2028-11-29" para "a las 12:55"
        # de hoy). Las reglas abajo + el anchor de fecha al inicio
        # del prompt mitigan los tres modos de falla.
        "REGLAS DE WHATSAPP (críticas, el modelo se equivoca seguido):\n\n"
        "  1) `message_text` es LITERAL hasta el FINAL del prompt. "
        "Cuando el user dice 'diciendo: X' / 'que diga: X' / ': X', "
        "TODO lo que viene después de los dos puntos es el cuerpo, "
        "hasta el final, incluyendo signos de exclamación, palabras "
        "tipo 'programado'/'urgente'/'de aviso', emojis, todo. NO "
        "interpretes 'mensaje programado' como meta-descripción del "
        "envío y la saques del cuerpo: es texto literal que el user "
        "quiere mandar. Si el user dice 'mandale: Hola! mensaje "
        "programado' → message_text=\"Hola! mensaje programado\" (no "
        "\"Hola!\").\n\n"
        "  2) `scheduled_for` (campo OPCIONAL): si el user mencionó "
        "hora/fecha futura, pasala. **La tool acepta DOS formatos** y "
        "tu mejor opción es la primera:\n"
        "     a) **Lenguaje natural EXACTO** como lo dijo el user — "
        "        'a las 12:55', 'mañana 14:30', 'el viernes 18hs', "
        "        'en 2 horas'. La tool corre _parse_natural_datetime "
        "        con anchor=hoy, así que NO necesitás computar la "
        "        fecha. Esta es la forma RECOMENDADA — el modelo "
        "        consistentemente alucina años random cuando intenta "
        "        emitir ISO ('2028-11-29' para algo de hoy, observado).\n"
        f"     b) ISO8601 completo con offset -03:00 ('{iso_today}"
        "T12:55:00-03:00'). Sólo si tenés ALTA confianza en la fecha. "
        "Si tenés duda, usá (a).\n"
        "     Patterns que SIEMPRE disparan scheduled_for (cualquier "
        "formato):\n"
        "       - 'a las HH:MM' / 'a las HHhs'\n"
        "       - 'mañana <hora>' / 'el <día> <hora>' / 'el <fecha> a "
        "         las <hora>'\n"
        "       - 'en N horas/minutos/días'\n"
        "       - 'esta tarde' / 'esta noche' / 'mañana temprano'\n"
        "     Si el user NO mencionó fecha/hora, OMITÍ el arg. NO "
        "pongas null. NO inventes fecha.\n\n"
        "  3) EJEMPLO CANÓNICO con NL (preferido):\n"
        "       'mandale a mi mamá a las 11:55 diciendo: Hola! mensaje "
        "programado' →\n"
        "       propose_whatsapp_send(\n"
        "           contact_name=\"mama\",\n"
        "           message_text=\"Hola! mensaje programado\",\n"
        "           scheduled_for=\"a las 11:55\"\n"
        "       )\n"
        "     NUNCA truncar a \"Hola!\". NUNCA omitir scheduled_for. "
        "NUNCA inventar año/mes random.\n\n"
        "Después del tool call, el siguiente turn del assistant (sin "
        "tool calls) es UNA oración breve de confirmación. No repitas "
        "los campos (el usuario los ve en el chip inline del chat).\n\n"
        # 2026-04-28 wave-7 (eval Conv 3 — wrong cancel tool): rule corta
        # para evitar inflar el override (5300+ chars regresiona routing).
        "Cancelar/editar: SOLO `propose_whatsapp_cancel_scheduled` para "
        "WhatsApp scheduled. Reminders/calendar NO tienen edit/cancel tools "
        "— para esos NO llames tool, el siguiente turn dice: \"No puedo "
        "editar/cancelar desde acá, abrí Apple Reminders/Calendar.\"\n\n"
        "No cites notas del vault, no inventes paths, no menciones "
        "REGLA 1. El único output válido acá es un tool_call en el "
        "protocolo; imprimir el call como texto es un bug."
    )


@_on_startup
def _warmup() -> None:
    """Hydrate the home cache and kick the bg prewarmer; DO NOT pin the
    chat model or reranker on boot.

    The previous version called `ollama.chat(model=command-r, ...)` on
    startup which pinned ~19 GB of unified memory even when the user was
    only browsing /dashboard or /. On a 36 GB Mac that's enough wired
    memory to starve the kernel and cause host freezes. Warmup of the
    expensive bits (reranker MPS init, chat model load) now happens
    on-demand from the first /api/chat request — 2-3s extra cost on one
    query in exchange for not holding the machine hostage when idle.

    Corpus/BM25/PageRank are cheap (RAM only, no VRAM), so we still
    preload them so the first retrieve is fast.
    """
    import threading

    _load_home_cache()
    _ensure_home_prewarmer()
    _ensure_chat_model_prewarmer()
    _ensure_reranker_prewarmer()
    _ensure_corpus_prewarmer()

    # Record this daemon startup in rag_ambient so restart count is queryable
    # via SQL instead of grepping web.log.
    try:
        import importlib.metadata as _imeta
        _rag_ver: str | None = _imeta.version("obsidian-rag")
    except Exception:
        _rag_ver = None
    try:
        _startup_payload: dict = {
            "pid": os.getpid(),
            "ts": datetime.now().isoformat(timespec="seconds"),
        }
        if _rag_ver:
            _startup_payload["version"] = _rag_ver
        with _ragvec_state_conn() as _sc:
            _sql_append_event(_sc, "rag_ambient", {
                "ts": datetime.now().isoformat(timespec="seconds"),
                "cmd": "serve.startup",
                "payload_json": _startup_payload,
            })
    except Exception as _exc:
        print(f"[warmup] startup event skipped: {_exc}", flush=True)

    # Memory-pressure watchdog — evita beachballs si el server + otras apps
    # saturan los 36 GB unified memory. Fires keep_alive=0 sobre el chat
    # model a los >85%, force-unload del reranker si sigue alto. Ver doc
    # extensa en rag.py alrededor de `_system_memory_used_pct()`.
    try:
        from rag import start_memory_pressure_watchdog
        start_memory_pressure_watchdog()
    except Exception as _exc:
        print(f"[warmup] memory-pressure watchdog skipped: {_exc}", flush=True)

    # Latency degradation watchdog — detecta wedge del daemon ollama
    # tras queries tool-heavy consecutivas y reinicia automáticamente.
    # Eval 2026-04-28: tras 5-7 queries con tools, p95 sube de ~30s a
    # ~70s. El watchdog detecta y bouncea el daemon. Lazy import para
    # no romper si el módulo no existe (deploy parcial OK).
    try:
        from rag._ollama_health import start_latency_degradation_watchdog
        start_latency_degradation_watchdog()
    except Exception as _exc:
        print(f"[warmup] ollama-health-watchdog skipped: {_exc}", flush=True)

    # WAL checkpointer — libera páginas del WAL cada 30s para que los
    # writers concurrentes (queries, behavior, cache) no peguen contra el
    # busy_timeout bajo carga sostenida. Audit 2026-04-24.
    try:
        from rag import start_wal_checkpointer
        start_wal_checkpointer()
    except Exception as _exc:
        print(f"[warmup] wal-checkpointer skipped: {_exc}", flush=True)

    def _do_warmup() -> None:
        try:
            from rag import get_db_for
            for _name, path in resolve_vault_paths(None):
                try:
                    col = get_db_for(path)
                    if col.count():
                        _load_corpus(col)
                        get_pagerank(col)
                except Exception:
                    pass
                break
            # End-to-end warmup: load the expensive singletons that the
            # first /api/chat would otherwise pay for (reranker on MPS +
            # bge-m3 SentenceTransformer + one dummy embed pass). Only
            # fires when the operator has opted into pinning them via the
            # same flags that keep them resident.
            if os.environ.get("RAG_RERANKER_NEVER_UNLOAD", "").strip() not in ("", "0", "false", "no"):
                try:
                    from rag import get_reranker as _get_rr
                    _get_rr()
                    print("[warmup] reranker loaded on MPS", flush=True)
                except Exception as _exc:
                    print(f"[warmup] reranker skipped: {_exc}", flush=True)
            if os.environ.get("RAG_LOCAL_EMBED", "").strip() not in ("", "0", "false", "no"):
                try:
                    # `_warmup_local_embedder()` es el único helper que, además
                    # de cargar el modelo y hacer un dummy encode, setea el
                    # `_local_embedder_ready.Event` que `query_embed_local()`
                    # checkea como gate non-blocking. Antes del 2026-04-22
                    # este startup llamaba `_get_local_embedder()` + `.encode`
                    # directo, lo que cargaba el modelo pero NO seteaba el
                    # Event → cada `/api/chat` caía al fallback ollama
                    # (~140ms vs ~10-30ms local). Ver
                    # test_web_local_embed_warmup.py para el contrato.
                    from rag import _warmup_local_embedder
                    if _warmup_local_embedder():
                        print("[warmup] bge-m3 local embedder ready (event set)", flush=True)
                    else:
                        print("[warmup] bge-m3 local embedder skipped (load/encode failed)", flush=True)
                except Exception as _exc:
                    print(f"[warmup] local embed skipped: {_exc}", flush=True)
            # Drain any conversation turns that failed to persist on previous
            # runs (transient SQL busy / disk full / etc). Best-effort;
            # survivors stay in the file for the next startup.
            try:
                _retried = _retry_pending_conversation_turns()
                if _retried:
                    print(f"[warmup] retried {_retried} pending conversation turn(s)",
                          flush=True)
            except Exception as _exc:
                print(f"[warmup] conversation-turn retry failed: {_exc}", flush=True)
        except Exception:
            pass

    threading.Thread(target=_do_warmup, daemon=True).start()

    def _idle_sweeper() -> None:
        """Evict the reranker from MPS after `_RERANKER_IDLE_TTL` of no
        activity. Keeps the Mac responsive when the user walks away from
        chat — 2-3 GB of unified memory freed on idle.

        After eviction, schedules a deferred pre-warm 60s later so the
        reranker is hot again before the likely next request (user returns
        to chat after a short break). The pre-warm runs in its own daemon
        thread and MUST NOT block the sweeper loop.

        When `RAG_RERANKER_NEVER_UNLOAD=1` (env var) the sweeper loop
        still runs but skips `maybe_unload_reranker()` entirely — the
        reranker stays pinned in MPS VRAM at all times. Use this on a
        36 GB unified-memory Mac where the ~2-3 GB cost is acceptable and
        eliminating the 9s cold-reload hit after idle eviction is worth it.
        """
        from rag import maybe_unload_reranker, get_reranker
        _never_unload = os.environ.get("RAG_RERANKER_NEVER_UNLOAD", "").strip() not in ("", "0", "false", "no")
        if _never_unload:
            print("[idle-sweep] RAG_RERANKER_NEVER_UNLOAD=1 — reranker pinned, sweeper disabled", flush=True)
        while True:
            try:
                time.sleep(120)
                if _never_unload:
                    continue
                if maybe_unload_reranker():
                    print("[idle-sweep] reranker unloaded from MPS", flush=True)

                    def _deferred_rewarm() -> None:
                        time.sleep(60)
                        try:
                            get_reranker()
                            print("[idle-sweep] reranker pre-warmed (deferred)", flush=True)
                        except Exception:
                            pass

                    threading.Thread(
                        target=_deferred_rewarm, name="reranker-rewarm", daemon=True
                    ).start()
            except Exception:
                pass

    threading.Thread(target=_idle_sweeper, name="idle-sweeper", daemon=True).start()


@app.get("/")
def home_page() -> FileResponse:
    """Home dashboard (mission-control terminal aesthetic). Servido por
    `home.v2.html` desde 2026-04-27 después del refactor + 4 commits de
    iteración (cross-source correlator, prompt enriquecido, voice
    normalizer, tier visual hierarchy). Usa `/api/home` como backend.
    """
    return FileResponse(STATIC_DIR / "home.v2.html")


# Legacy ruta del home v1 — accesible en `/v1` durante la transición.
# Si rompe algo en producción podemos volver al viejo apuntando `/` a
# `home.html` con un revert del swap. Borrar esta ruta en el commit
# siguiente confirmando que `/` v2 funciona en producción ~1 semana.
@app.get("/v1")
def home_v1_page() -> FileResponse:
    return FileResponse(STATIC_DIR / "home.html")


# `/v2` mantenido como alias de `/` para no romper bookmarks o links
# externos que apuntan a la URL preview durante la fase de iteración.
@app.get("/v2")
def home_v2_page() -> FileResponse:
    return FileResponse(STATIC_DIR / "home.v2.html")


@app.get("/chat")
def chat_page() -> FileResponse:
    return FileResponse(STATIC_DIR / "index.html")


# ── PWA endpoints ─────────────────────────────────────────────────────
# El manifest y el service worker DEBEN servirse desde la raíz para que
# el SW pueda controlar el scope "/" (home + /chat + /dashboard). Un SW
# servido en /static/sw.js sólo cubre /static/** — inútil para nosotros.
# Podríamos mandar el header `Service-Worker-Allowed: /` para extender
# el scope sin mover el archivo, pero tener los archivos físicos en
# /static/ + un proxy controller acá es más simple que toquetear
# middlewares de StaticFiles.
#
# Cache del manifest: 1 día (cambia raro, pero no queremos que un bump
# de icons quede invisible por semanas). El SW mismo: `no-cache` para
# que updates al archivo JS se detecten al toque (browser lo chequea
# a cada navegación por spec). Sin este `no-cache` los updates del SW
# pueden tardar hasta 24h en llegar al device del usuario — incluso con
# `updateViaCache: "none"` en el register().
@app.get("/manifest.webmanifest")
def manifest() -> FileResponse:
    return FileResponse(
        STATIC_DIR / "manifest.webmanifest",
        media_type="application/manifest+json",
        headers={"Cache-Control": "public, max-age=86400"},
    )


@app.get("/sw.js")
def service_worker() -> FileResponse:
    return FileResponse(
        STATIC_DIR / "sw.js",
        media_type="application/javascript",
        headers={
            # no-cache fuerza el browser a revalidar con ETag/Last-Modified
            # en cada request. No es no-store — si el server responde 304,
            # el browser usa el JS cacheado. Este es el setting que
            # recomienda la spec SW para archivos SW.
            "Cache-Control": "no-cache",
            # Permite cambiar el scope del SW en el futuro sin mover
            # el archivo (no lo necesitamos hoy porque ya está en root,
            # pero lo dejamos como hint declarativo).
            "Service-Worker-Allowed": "/",
        },
    )


_CHAT_SESSION_RE = re.compile(r"^[A-Za-z0-9_.:@\-]{1,80}$")
_TURN_ID_RE = re.compile(r"^[A-Za-z0-9_\-]{1,64}$")
_CHAT_QUESTION_MAX = 16000  # ~4k tokens — bumped from 8000 (2026-04-20) because users sometimes paste long doc excerpts into chat; 16000 still well under CHAT_OPTIONS num_ctx=4096 × 4 chars/token


class ChatRequest(BaseModel):
    question: str
    session_id: str | None = None
    # None → vault activo; "all" → todos los registrados; "name" → ese puntual.
    vault_scope: str | None = None
    # Regeneration: pass a previous turn_id to re-ask that same question.
    # The handler resolves `q` from rag_queries (by turn_id in extra_json)
    # and uses it as the effective question, so the client can call /redo
    # without having kept the question text locally. `hint` is an optional
    # soft-steer that gets concatenated ("la pregunta — enfocá: <hint>")
    # to redirect the answer without reformulating from scratch.
    # When redo_turn_id is set, `question` becomes a placeholder — the
    # client should still send "(redo)" or similar non-empty string to
    # satisfy the non-empty validator.
    redo_turn_id: str | None = Field(None, max_length=80)
    hint: str | None = Field(None, max_length=500)
    # User-facing mode selector (2026-04-24): "auto" | "fast" | "deep".
    # Gates `_fast_path` in the streaming block. None → treated as "auto"
    # (adaptive routing from `retrieve()`). Invalid strings → silent
    # fallback to "auto" (validated in the endpoint body, NOT in a
    # field_validator — we explicitly tolerate garbage here so that old
    # curl scripts, MCP clients, and PWA builds that pre-date the UI
    # control keep working without 400s). The legacy /api/chat/model
    # endpoints remain as a developer escape hatch for forcing a specific
    # chat model tag.
    mode: str | None = Field(None, max_length=16)

    @field_validator("question")
    @classmethod
    def _check_question(cls, v: str) -> str:
        if not v or not v.strip():
            raise ValueError("question must be non-empty")
        if len(v) > _CHAT_QUESTION_MAX:
            raise ValueError(f"question too long (>{_CHAT_QUESTION_MAX} chars)")
        return v

    @field_validator("redo_turn_id")
    @classmethod
    def _check_redo_turn_id(cls, v: str | None) -> str | None:
        if v is None or v == "":
            return None
        if not _TURN_ID_RE.match(v):
            raise ValueError("invalid redo_turn_id format")
        return v

    @field_validator("session_id")
    @classmethod
    def _check_session(cls, v: str | None) -> str | None:
        if v is None or v == "":
            return None
        if not _CHAT_SESSION_RE.match(v):
            raise ValueError("invalid session_id format")
        return v

    @field_validator("vault_scope")
    @classmethod
    def _check_vault_scope(cls, v: str | None) -> str | None:
        if v is None or v == "":
            return None
        # Same character class as session_id but shorter — vault names are
        # paths / identifiers registered via `rag vault add`.
        if len(v) > 200 or not re.match(r"^[A-Za-z0-9_./\-]{1,200}$", v):
            raise ValueError("invalid vault_scope format")
        return v


class FeedbackRequest(BaseModel):
    turn_id: str
    rating: int                   # +1 thumbs up / -1 thumbs down
    # Post-hoc parse-time size caps so a multi-MB payload rejects with
    # 422 before we allocate it into memory. The downstream handler
    # trims/caps further as a belt-and-suspenders.
    q: str | None = Field(None, max_length=2000)
    paths: list[str] | None = Field(None, max_length=50)
    session_id: str | None = None
    reason: str | None = Field(None, max_length=500)     # short free-text, optional (≤200 chars after trim)
    # Vault-relative path the user marked as the correct one when the
    # retrieve failed. Mirrors the CLI corrective_path flow (rag.py:~18997,
    # commit 23f2899 / 2026-04-22). Before this, rag_feedback had 0 web-sourced
    # rows with corrective_path — the fine-tune of the reranker was starved of
    # clean (query, positive, negative) triplets from the 80% of traffic that
    # comes through the web. Cap at 512 chars (longest realistic vault path).
    corrective_path: str | None = Field(None, max_length=512)

    @field_validator("turn_id")
    @classmethod
    def _check_turn_id(cls, v: str) -> str:
        if not _TURN_ID_RE.match(v):
            raise ValueError("invalid turn_id format")
        return v

    @field_validator("session_id")
    @classmethod
    def _check_session(cls, v: str | None) -> str | None:
        if v is None or v == "":
            return None
        if not _CHAT_SESSION_RE.match(v):
            raise ValueError("invalid session_id format")
        return v


@app.post("/api/feedback")
def submit_feedback(req: FeedbackRequest) -> dict:
    """Record a +1/-1 rating for a chat turn into feedback.jsonl.

    The rating signal feeds ranker tuning (`rag tune`) and the golden
    cache that biases future retrievals. Client sends the turn_id from
    the `done` SSE event plus the question and the source paths it saw,
    so we don't need to re-scan queries.jsonl on every click.

    Optional `reason` lets the user jot a short note when clicking 👎
    ("genérico", "faltó la nota X", "mezcla temas") — grouped later by
    `rag insights` to surface common failure modes.

    Optional `corrective_path` is a vault-relative path the user picked
    as the correct one among the top-5 shown. When present it overrides
    `reason` to "corrective" (the sentinel `_feedback_augmented_cases()`
    keys on to mine augmentation triplets for `rag tune`). This is the
    web-side counterpart of the CLI chat corrective prompt (2026-04-22)
    — without it the fine-tune was starved of signal from 80% of the
    user traffic.
    """
    if not req.turn_id or req.rating not in (1, -1):
        raise HTTPException(status_code=400, detail="turn_id + rating ±1 requeridos")
    corrective_path = (req.corrective_path or "").strip() or None
    if corrective_path and "://" in corrective_path:
        # Reject cross-source native ids (calendar://, whatsapp://, gmail://).
        # Not a vault-relative path — `_feedback_augmented_cases()` can't
        # consume it as a positive. Silent drop + keep the rating anyway.
        corrective_path = None
    # When the user picked a corrective_path we override the reason to
    # "corrective" (this is the sentinel _feedback_augmented_cases() keys
    # on to mine augmentation pairs for `rag tune`). If they only wrote a
    # free-text reason without picking a path, keep their reason as-is.
    if corrective_path:
        reason: str | None = "corrective"
    else:
        reason = (req.reason or "").strip()[:200] or None
    record_feedback(
        turn_id=req.turn_id,
        rating=req.rating,
        q=(req.q or "").strip(),
        paths=req.paths or [],
        reason=reason,
        corrective_path=corrective_path,
        session_id=req.session_id,
    )
    return {"ok": True}


# ── /api/behavior ─────────────────────────────────────────────────────────────
# In-memory token bucket for rate limiting: 120 events/min per IP.
# Stored as {ip: [timestamp, ...]} with a 60s sliding window. No new deps.
import collections as _collections
import threading as _threading

# Audit 2026-04-25 finding R2-Performance #1: bound de IPs en buckets
# de rate-limit. Antes usábamos `defaultdict(deque)` directo y crecía
# sin límite — cada IP nueva creaba una entrada y nunca se removía.
# Bajo ataque (rotación de proxies, scanners) o uso prolongado, el dict
# acumulaba decenas de miles de keys (cada deque vacío ≈ 64 bytes +
# overhead de la key). El wrapper `_LRURateBucket` mantiene un
# `OrderedDict` ordenado por LRU; cuando supera `max_size`, popea el
# menos recientemente accedido. API compatible con el código viejo que
# hace `bucket[ip].append(...)` / `bucket[ip].popleft()`. El default de
# 5000 IPs es generoso para tráfico humano normal (~10s de IPs por día)
# y suficientemente estricto para que un atacante no agote memoria.
_LRU_RATE_BUCKET_MAX = 5000


class _LRURateBucket:
    """Bucket per-IP con bound LRU. Audit 2026-04-25 finding R2-Performance #1.

    Wrapper sobre `OrderedDict` que mantiene como mucho `max_size`
    entradas. Cada acceso a `bucket[ip]` cuenta como "recently used":
    la key se mueve al final del orden con `move_to_end`. Cuando se
    inserta una key nueva y el dict ya está full, se popea la primera
    (la menos recientemente accedida).

    La API es compatible con `defaultdict(deque)` para los call sites
    existentes (`_check_rate_limit`, tests con `.clear()`, etc.):
    ``bucket[ip]`` siempre devuelve un ``collections.deque`` (creándolo
    si no existía). Los tests que monkeypatchen el atributo entero con
    un ``defaultdict(list)`` siguen funcionando porque reemplazan la
    referencia, no el wrapper.
    """

    def __init__(self, max_size: int = _LRU_RATE_BUCKET_MAX) -> None:
        self._data: _collections.OrderedDict[str, _collections.deque] = (
            _collections.OrderedDict()
        )
        self._max_size = max_size

    def __getitem__(self, key: str) -> _collections.deque:
        # Hit: marcamos como recently-used y devolvemos el deque existente.
        if key in self._data:
            self._data.move_to_end(key)
            return self._data[key]
        # Miss: creamos un deque vacío y, si llenamos, eviccionamos LRU.
        bucket: _collections.deque = _collections.deque()
        self._data[key] = bucket
        if len(self._data) > self._max_size:
            self._data.popitem(last=False)
        return bucket

    def __setitem__(self, key: str, value) -> None:
        # Compat con tests/legacy: aceptamos cualquier iterable y lo
        # envolvemos en deque para preservar la API esperada por
        # `_check_rate_limit`.
        if not isinstance(value, _collections.deque):
            value = _collections.deque(value)
        self._data[key] = value
        self._data.move_to_end(key)
        if len(self._data) > self._max_size:
            self._data.popitem(last=False)

    def __contains__(self, key: object) -> bool:
        return key in self._data

    def __len__(self) -> int:
        return len(self._data)

    def __iter__(self):
        return iter(self._data)

    def clear(self) -> None:
        self._data.clear()

    def keys(self):
        return self._data.keys()

    def values(self):
        return self._data.values()

    def items(self):
        return self._data.items()


_BEHAVIOR_BUCKETS: _LRURateBucket = _LRURateBucket()
_BEHAVIOR_RATE_LIMIT = 120
_BEHAVIOR_RATE_WINDOW = 60.0  # seconds

# Chat endpoint bucket — stricter because each hit pins the reranker, chat
# model, and embedder on MPS; 30 chats/min per IP is plenty for human use
# and stops an adversarial loop (e.g. browser extension gone rogue) from
# starving the daemon. Separate from behavior bucket so clicks don't
# consume chat budget.
_CHAT_BUCKETS: _LRURateBucket = _LRURateBucket()
_CHAT_RATE_LIMIT = 30
_CHAT_RATE_WINDOW = 60.0

# Single lock protects both buckets — contention is effectively nil
# (deque.append under GIL is fast enough that a lock hold measures in µs).
# 2026-04-24 audit: cambio `list` → `collections.deque` para que la
# expiración del sliding window sea O(1) con `popleft()` en vez de
# O(n) con `list.pop(0)`. Bajo carga moderada (1000+ req/min) el
# loop `while events and events[0] < cutoff: events.pop(0)` podía
# converger a O(n²) y dominar CPU del handler. Deque hace el shift
# constant-time. Los tests que hacen `_CHAT_BUCKETS.clear()` siguen
# funcionando igual (deque soporta `.clear()`).
_RATE_LIMIT_LOCK = _threading.Lock()


def _check_rate_limit(bucket: dict[str, "_collections.deque"], ip: str,
                      limit: int, window: float) -> None:
    """Sliding-window rate limit per-IP. Raises HTTPException 429 on breach."""
    now = time.time()
    cutoff = now - window
    with _RATE_LIMIT_LOCK:
        events = bucket[ip]
        # `popleft()` O(1) en deque. Antes con `list.pop(0)` cada remoción
        # era O(len(events)) → bajo carga con 30 events en ventana el
        # loop completo caía a O(n²). Cambio a deque baja el cost a O(n).
        while events and events[0] < cutoff:
            events.popleft()
        if len(events) >= limit:
            raise HTTPException(status_code=429, detail="rate limit exceeded")
        events.append(now)

_BEHAVIOR_KNOWN_EVENTS = frozenset({
    "open", "open_external", "positive_implicit",
    "negative_implicit", "kept", "deleted", "save",
    # "copy" (2026-04-22): user copied text from a RAG response (web Cmd+C
    # selection inside a .turn, or CLI `/copy`). Implicit positive signal
    # — typically stronger than a 👍 because the user did something with
    # the content, not just reacted to it. `path` carries the top source
    # at the time of the copy (best-effort inference: we can't exactly
    # attribute the copied substring to a chunk, so we pin rank=1). The
    # copied text length is gated client-side (≥20 chars) to skip
    # accidental copies of short labels / path fragments.
    "copy",
})
_BEHAVIOR_SESSION_RE = re.compile(r"^[A-Za-z0-9_.@:-]{1,80}$")


class BehaviorRequest(BaseModel):
    source: str
    event: str
    query: str | None = None
    path: str | None = None
    rank: int | None = None
    dwell_ms: int | None = None
    session: str | None = None


@app.post("/api/behavior")
def submit_behavior(req: BehaviorRequest, request: Request) -> dict:
    """Record a user-behavior event from the web dashboard into behavior.jsonl."""
    # Validate source
    if req.source != "web":
        raise HTTPException(status_code=400, detail="source must be 'web'")

    # Validate event
    if req.event not in _BEHAVIOR_KNOWN_EVENTS:
        raise HTTPException(
            status_code=400,
            detail=f"unknown event '{req.event}'; valid: {sorted(_BEHAVIOR_KNOWN_EVENTS)}",
        )

    # Validate session regex if present
    if req.session is not None and not _BEHAVIOR_SESSION_RE.match(req.session):
        raise HTTPException(status_code=400, detail="session id format invalid")

    # Validate path stays inside vault (no traversal)
    if req.path is not None:
        p = req.path
        # Cross-source native ids (calendar://, whatsapp://, gmail://) are
        # not vault-relative paths — they can't be consumed by the ranker-
        # vivo CTR aggregator (keyed on vault-relative paths). Reject with
        # 400 so clients get immediate feedback instead of silent drop.
        # Also catches any pathological attempt to sneak a URI via the path
        # field. Belt + suspenders: the client-side listener in app.js
        # already filters these out before POSTing.
        if "://" in p:
            raise HTTPException(
                status_code=400,
                detail="path must be vault-relative (no URI schemes)",
            )
        if p.startswith("/") or ".." in p.split("/"):
            raise HTTPException(status_code=400, detail="path must be vault-relative")
        try:
            resolved = (VAULT_PATH / p).resolve()
            VAULT_PATH.resolve()  # ensure VAULT_PATH itself resolves
            resolved.relative_to(VAULT_PATH.resolve())
        except ValueError:
            raise HTTPException(status_code=400, detail="path escapes vault root")

    # Rate limit: 120 events/60s per client IP
    client_ip = (request.client.host if request.client else "unknown")
    _check_rate_limit(_BEHAVIOR_BUCKETS, client_ip,
                      _BEHAVIOR_RATE_LIMIT, _BEHAVIOR_RATE_WINDOW)

    # The BehaviorRequest surface uses `dwell_ms` (higher-resolution,
    # matches the JS IntersectionObserver timer output), but the SQL
    # schema column is `dwell_s` (per _map_behavior_row at rag.py:4962).
    # The aggregator reader `_compute_behavior_priors_from_rows` only
    # SELECTs `dwell_s` from the table (rag.py:2902), so any dwell_ms
    # that falls through to `extra_json` is effectively lost to the
    # ranker-vivo priors. Convert here so the dwell signal lands in
    # the column the aggregator actually reads.
    dwell_s = (req.dwell_ms / 1000.0) if req.dwell_ms is not None else None
    # Hidrate `query` from the latest rag_queries row in this session when the
    # client didn't pass one (race con `appendSources` antes del set de
    # turn.dataset.q, panels de home.js sin context, etc.). Sin query el
    # ranker-vivo ignora el evento entero (`_behavior_augmented_cases`
    # require query NOT NULL); 122 events del web en 7d quedaron inertes
    # por este bug. Best-effort lookup: si falla, deja query=None y se
    # comporta como antes.
    query_val = req.query
    if (query_val is None or query_val == "") and req.session:
        try:
            with _ragvec_state_conn() as _conn:
                _row = _conn.execute(
                    "SELECT q FROM rag_queries WHERE session = ? AND q IS NOT NULL "
                    "AND q <> '' ORDER BY ts DESC LIMIT 1",
                    (req.session,),
                ).fetchone()
                if _row and _row[0]:
                    query_val = _row[0]
        except Exception:
            pass  # silent — peor caso queda como antes (query=None)
    try:
        log_behavior_event({
            "source": req.source,
            "event": req.event,
            "query": query_val,
            "path": req.path,
            "rank": req.rank,
            "dwell_s": dwell_s,
            "session": req.session,
        })
    except Exception as exc:
        # Never 500 on I/O failure — degrade gracefully
        print(f"[behavior] write error: {exc}", flush=True)
        raise HTTPException(status_code=503, detail="event log unavailable")

    return {"ok": True}


def _ollama_alive(timeout: float = 2.0) -> bool:
    """Fast probe: does the ollama daemon answer `/api/tags` within `timeout`s?
    When the daemon hits its stuck-load state the HTTP listener accepts but
    never replies — /api/chat then hangs forever. A short /api/tags probe
    catches that state without waiting for a model load.
    """
    import urllib.request
    import urllib.error
    try:
        with urllib.request.urlopen("http://127.0.0.1:11434/api/tags", timeout=timeout) as r:
            return r.status == 200
    except Exception:
        return False


# Dedicated streaming client with a per-chunk read timeout. httpx applies
# `read` to each chunk of a streaming response: if no bytes arrive within
# the budget, ReadTimeout fires and our `except Exception` surfaces an
# error SSE to the frontend. Without this, a stuck-load ollama daemon
# silently wedges the /api/chat stream forever (spinner never clears).
#
# Budget rationale (2026-04-28 update — was 45.0):
# La streaming call es la fase MÁS pesada del turno: prefill sobre el
# contexto post-tools (típico 25-30k chars cuando hay whatsapp_pending /
# reminders_due / calendar_ahead) + decode de la respuesta. Empíricamente
# `[chat-timing]` registra ttft hasta 38.5s en queries con tool_rounds=1,
# lo cual está peligrosamente cerca de 45s — un cold-load del modelo
# (~5-10s extra de KV reinit cuando num_ctx adaptive cambia respecto del
# value loaded) y/o memory pressure en MPS lo empuja >45s y dispara
# "LLM falló: timed out" en el frontend. Repro autónomo Playwright
# 2026-04-28: 3 de 5 queries (whatsapp/RAG/pendientes) cayeron en
# 59-62s wall time → el stream client cortaba a los 45s del primer
# read, antes de que llegara el primer token de la synthesis call.
#
# Subimos a 90s para alinear con `_OLLAMA_TOOL_TIMEOUT` (que ya estaba
# en 90s tras commit b0d140e). Argumento: el tool decision call mide el
# 1er-y-único chunk no-streaming, mientras que el stream final mide
# per-chunk; ambos pueden saturar el mismo budget cuando el modelo cold-
# loads o el contexto explota. Si realmente hay un wedge del daemon,
# `_ollama_chat_probe` lo detecta antes con 6s — el budget de 90s no
# expone al user a colgues largos por daemon stuck.
_OLLAMA_STREAM_TIMEOUT = 90.0
_OLLAMA_STREAM_CLIENT = ollama.Client(timeout=_OLLAMA_STREAM_TIMEOUT)

# Tool-decision call (non-streaming, ~once per turn, with `tools=` schema
# of all 13 chat tools). Observed 2026-04-24 (Fer F. "LLM falló: timed
# out"): long queries (~50 palabras + 2KB de _WEB_TOOL_ADDENDUM + JSON
# schemas de 13 tools) hacen que qwen2.5:7b en MPS tarde >45s en
# samplear la decisión — se disparaba el timeout del cliente streaming
# compartido y el turno fallaba antes de llegar a generar output.
# Cliente separado con budget más amplio: tool-decision es una call
# NO-streaming (no hay UX de tokens flowing) y low-frequency (1 vez por
# turn, máx 3 rounds), así que podemos permitirnos más budget sin
# degradar la percepción de "chat congelado". Si realmente hay un hang
# del daemon, `_ollama_chat_probe` lo detecta antes con 6s.
#
# Audit 2026-04-25 finding R1 #9: bajamos el budget de 120s → 45s. El
# valor previo era demasiado amplio: si qwen2.5:7b se cuelga (OOM,
# memory pressure, daemon wedge), el chat queda freeze 2 minutos enteros
# desde la perspectiva del user antes de fallar. qwen2.5 con
# num_ctx=4096 tarda 1-3s warm + 8-10s cold-load; 45s cubre cold-load +
# 1-2 retries internos del cliente sin colgar el chat 2 min entero.
#
# Eval autónomo 2026-04-28 (12 queries vía Playwright): 45s no
# alcanzaba para queries post-tool con outputs grandes. Concretamente
# gmail_recent + whatsapp_search + drive_search consistentemente
# timeouteaban en la 2da ronda (donde el LLM sintetiza el output del
# tool en prosa) — observado 60-80s de wall time. Sube a 90s para
# cubrir el caso. Si en el futuro queremos cap más agresivo, mejor
# truncar el tool output ANTES de mandarlo al LLM (no tiene sentido
# pasarle 30+ items con full body — pasarle un summary y los top-N).
_OLLAMA_TOOL_TIMEOUT = 90.0
_OLLAMA_TOOL_CLIENT = ollama.Client(timeout=_OLLAMA_TOOL_TIMEOUT)

# Shared num_ctx for every call to the chat model from this server — the
# real /api/chat, the preflight probe, and the background prewarmer must
# all pass the same value. ollama re-initialises the KV cache when a
# model call arrives with a different num_ctx than the currently-loaded
# one, and qwen2.5:7b's default (2048) differs from the value the real
# chat path needs (4096 — see _WEB_CHAT_OPTIONS below). A single drift
# between preflight and /api/chat forces a full KV reinit and pushes
# prefill from ~1.5s to ~4.9s (measured 2026-04-20). Defined at module
# scope so there's one source of truth.
_WEB_CHAT_NUM_CTX = 4096

# Critique loop env-gate (rollout 2026-04-25, default OFF).
#
# El critique pass vive en `rag.run_parallel_post_process` (ver
# rag/__init__.py:17671) y corre post-stream sobre la respuesta del LLM
# para detectar/regenerar respuestas mal-citadas o débiles. El CLI lo
# expone via `--critique` flag; el endpoint web NUNCA lo invocó (audit
# telemetría 7d / 919 web queries: `critique_fired=0` en TODOS los rows).
#
# Esta constante habilita el wiring del critique en `/api/chat` cuando
# el operador exporta `RAG_CRITIQUE_ENABLED=1`. Default OFF para que el
# rollout sea reversible sin redeploy. Hacelo via env var (no per-request
# field en ChatRequest) para evitar surface-area en la API pública
# durante la fase de validación; si el critique demuestra mejora medible
# en grounding, lo promocionamos a un knob user-facing en una segunda
# pasada.
#
# NOTA (developer-1, 2026-04-25): la constante está definida pero el
# wiring del critique pass al stream NO está aplicado todavía. El CLI
# usa `run_parallel_post_process` post-stream con re-emisión condicional
# del answer corregido — replicar eso en el endpoint web requiere
# decidir UX (¿re-emitimos un `update` SSE event? ¿pre-buffereamos la
# respuesta?) + propagar args (docs/metas/scores) al post-process. Es
# >5 líneas de cambio + decisión de producto sobre el shape del SSE
# stream, así que queda gateado por esta env var pero sin consumer
# hasta que rag-llm valide el approach.
_CRITIQUE_ENABLED = (
    os.environ.get("RAG_CRITIQUE_ENABLED", "0").strip().lower()
    in ("1", "true", "yes")
)


def _ollama_chat_probe(timeout_s: float = 6.0) -> bool:
    """Deep probe: does `/api/chat` actually stream a token within `timeout_s`?

    The shallow `/api/tags` probe is insufficient — we've seen the daemon
    accept /api/tags while /api/chat hangs indefinitely (stuck-load mid-run).
    This probe sends a 1-token chat against the pinned chat model; if the
    daemon is wedged, httpx raises within the budget. Reuses the streaming
    client so the budget applies uniformly.

    CRITICAL: MUST pass num_ctx=_WEB_CHAT_NUM_CTX. Without it, ollama uses
    its default (2048 for qwen2.5:7b) which differs from the real /api/chat
    call (4096), forcing ollama to reinit the KV cache on the next user
    request. Measured impact: prefill 1.5s → 4.9s (3x slower). See
    `num_ctx mismatch` comment in the main chat path for the same trap.

    Also MUST include _WEB_SYSTEM_PROMPT as the system message. The probe
    runs before every /api/chat via _ollama_restart_if_stuck, and if it
    sends a different prompt (even just [user:"."]) ollama overwrites the
    slot's KV cache with that short prompt. The next real request then
    cold-prefills the 1300-token system prompt from scratch. Measured
    2026-04-20: probe w/o system → prefill 4.9s; probe WITH system →
    prefill 3.4s (−1.5s per request). The probe itself pays the cold
    prefill once at startup (~2.5s) and then ~80-100ms on every
    subsequent call since the system cache already exists.
    """
    try:
        _probe_model = _resolve_web_chat_model()
        _OLLAMA_STREAM_CLIENT.chat(
            model=_probe_model,
            messages=[
                {"role": "system", "content": _WEB_SYSTEM_PROMPT},
                {"role": "user", "content": "."},
            ],
            options={"num_predict": 1, "num_ctx": _WEB_CHAT_NUM_CTX,
                     "temperature": 0, "seed": 42},
            stream=False,
            think=False,   # thinking-capable models would otherwise emit
                           # <think> blocks as "tokens" with empty content
            keep_alive=chat_keep_alive(_probe_model),
        )
        return True
    except Exception:
        return False


def _ollama_restart_if_stuck() -> bool:
    """Heal the ollama daemon. Returns True on successful restart. Blocks
    3-10s for the bounce.

    Detect which deployment is running:
      1. If `homebrew.mxcl.ollama` is loaded in launchd → use `brew services restart`
      2. Else if Ollama.app is running → quit + reopen the app
      3. Else → try kickstart of homebrew first (in case it just bootstrapped),
         fallback to `open -a Ollama`.

    Pre-2026-04-28 hardcoded `brew services restart`, falló silenciosamente
    cuando el user usaba la .app version (sin homebrew daemon). Repro
    Playwright detectó hangs recurrentes que no se auto-curaban porque el
    restart no aplicaba al daemon real.
    """
    # Step 1: kill all runners forcefully (the .app's serve will respawn them
    # on next request). This handles the case where the runner is wedged but
    # the serve itself is healthy.
    try:
        subprocess.run(
            ["pkill", "-9", "-f", "ollama runner"],
            capture_output=True, timeout=5, check=False,
        )
    except Exception:
        pass

    # Step 2: detect deployment and restart serve.
    restarted = False
    # 2a) Homebrew daemon path.
    try:
        result = subprocess.run(
            ["launchctl", "list"],
            capture_output=True, text=True, timeout=5, check=False,
        )
        if result.returncode == 0 and "homebrew.mxcl.ollama" in result.stdout:
            try:
                subprocess.run(
                    ["/opt/homebrew/bin/brew", "services", "restart", "ollama"],
                    check=True, capture_output=True, timeout=30,
                )
                restarted = True
            except Exception:
                pass
    except Exception:
        pass

    # 2b) Ollama.app path — quit + reopen.
    if not restarted:
        try:
            # Kill all ollama processes (.app + serve + runner).
            subprocess.run(
                ["pkill", "-9", "-f", "ollama"],
                capture_output=True, timeout=5, check=False,
            )
            # Brief pause for clean exit.
            time.sleep(2)
            # Reopen the .app (background launches the serve daemon).
            subprocess.run(
                ["open", "-a", "Ollama"],
                capture_output=True, timeout=10, check=False,
            )
            restarted = True
        except Exception:
            pass

    if not restarted:
        return False

    # Wait up to 12s for the daemon to accept traffic again (Ollama.app
    # cold start takes 5-8s on Apple Silicon).
    for _ in range(24):
        if _ollama_alive(timeout=1.0):
            return True
        time.sleep(0.5)
    return False


@app.post("/api/ollama/restart")
def ollama_restart() -> dict:
    """Panic button #2: brew-services-restart the ollama daemon. Use when
    /api/chat is hanging forever (stuck-load state: daemon accepts HTTP
    but never streams a reply). Blocks ~5-10s.
    """
    ok = _ollama_restart_if_stuck()
    return {"ok": ok, "alive": _ollama_alive()}


@app.post("/api/ollama/unload")
def ollama_unload() -> dict:
    """Panic button: evict every loaded model from ollama + drop the
    reranker from MPS. Frees 15-25 GB of wired unified memory in one
    call when the Mac starts beachballing.

    Drops each loaded model with keep_alive=0 (ollama's immediate-evict
    sentinel) and clears the process-local reranker. Next chat query
    pays 3-8s cold reload but the host stays responsive until then.
    """
    freed = []
    try:
        ps = ollama.ps()
        for m in getattr(ps, "models", []) or []:
            name = getattr(m, "model", None) or getattr(m, "name", None)
            if not name:
                continue
            try:
                # A bare generate with keep_alive=0 tells ollama to unload
                # immediately after the (empty) prompt returns.
                ollama.generate(model=name, prompt="", keep_alive=0)
                freed.append(name)
            except Exception as exc:
                freed.append(f"{name} (fail: {exc})")
    except Exception as exc:
        return {"ok": False, "error": str(exc), "freed": freed}
    try:
        from rag import maybe_unload_reranker, _reranker_last_use  # noqa: F401
        import rag as _rag
        _rag._reranker_last_use = 0.0   # force eviction regardless of idle_ttl
        reranker_dropped = maybe_unload_reranker()
    except Exception:
        reranker_dropped = False
    return {"ok": True, "freed_models": freed, "reranker_dropped": reranker_dropped}


class ReminderCreateRequest(BaseModel):
    text: str
    due: str | None = None         # legacy token: "tomorrow" or None
    due_iso: str | None = None     # ISO-8601 datetime; wins over `due`
    list: str | None = None        # Reminders list name; None → default
    priority: int | None = None    # 1 high / 5 medium / 9 low / None
    notes: str | None = None       # body text
    recurrence: dict | None = None  # {freq, interval, byday?}


@app.post("/api/reminders/create")
def create_reminder(req: ReminderCreateRequest) -> dict:
    """Create an Apple Reminder.

    Dual-use:
      - Daily brief's "Para mañana" items POST `{text, due:"tomorrow"}` —
        the legacy path, preserved byte-for-byte.
      - Chat proposal cards POST `{text, due_iso, list, priority, notes,
        recurrence}` after the user clicks ✓ on a `proposal` SSE card.

    If both `due` and `due_iso` are present, `due_iso` wins. `recurrence`
    is best-effort: Reminders.app's AppleScript dictionary inconsistently
    accepts the property across macOS versions — the reminder is created
    regardless (silent-fail on recurrence), caller should advise the user
    to verify in Reminders.app.
    """
    from datetime import datetime as _dt

    from rag import _create_reminder  # noqa: PLC0415

    due_dt: _dt | None = None
    if req.due_iso:
        try:
            due_dt = _dt.fromisoformat(req.due_iso)
        except ValueError as exc:
            raise HTTPException(
                status_code=400, detail=f"due_iso inválido: {exc}",
            ) from exc

    ok, res = _create_reminder(
        req.text,
        due_token=req.due if not due_dt else None,
        list_name=req.list,
        due_dt=due_dt,
        priority=req.priority,
        notes=req.notes,
        recurrence=req.recurrence,
    )
    if not ok:
        raise HTTPException(status_code=400, detail=res)
    # Bust the home cache so the new reminder shows up in the reminders
    # panel on next load without waiting for SWR.
    try:
        _HOME_STATE["ts"] = 0.0
    except Exception:
        pass
    return {"ok": True, "id": res}


class CalendarCreateRequest(BaseModel):
    title: str
    start_iso: str                 # required
    end_iso: str | None = None     # None → start + 1h
    calendar: str | None = None    # None → first writable
    location: str | None = None
    notes: str | None = None
    all_day: bool = False
    recurrence: dict | None = None  # {freq, interval, byday?}


@app.post("/api/calendar/create")
def create_calendar_event(req: CalendarCreateRequest) -> dict:
    """Create a Calendar.app event. Called from chat proposal cards after
    the user confirms. Returns the event UID so the UI can surface a
    deep-link or deletion path.

    Writes through Calendar.app → iCloud CalDAV calendars are reachable
    (unlike the JXA EventKit read path which can't see them without
    entitlement). If `calendar` is omitted we write to the first writable
    calendar — usually iCloud's default.
    """
    from datetime import datetime as _dt

    from rag import _create_calendar_event  # noqa: PLC0415

    try:
        start_dt = _dt.fromisoformat(req.start_iso)
    except ValueError as exc:
        raise HTTPException(
            status_code=400, detail=f"start_iso inválido: {exc}",
        ) from exc
    end_dt: _dt | None = None
    if req.end_iso:
        try:
            end_dt = _dt.fromisoformat(req.end_iso)
        except ValueError as exc:
            raise HTTPException(
                status_code=400, detail=f"end_iso inválido: {exc}",
            ) from exc

    ok, res = _create_calendar_event(
        req.title,
        start_dt,
        end_dt,
        calendar=req.calendar,
        location=req.location,
        notes=req.notes,
        all_day=req.all_day,
        recurrence=req.recurrence,
    )
    if not ok:
        raise HTTPException(status_code=400, detail=res)
    return {"ok": True, "uid": res}


class ReminderCompleteRequest(BaseModel):
    reminder_id: str


@app.post("/api/reminders/complete")
def complete_reminder(req: ReminderCompleteRequest) -> dict:
    """Mark an Apple Reminder as completed by its stable id.

    Called from the home page when the user ticks a reminder. Invalidates
    the home cache so the next /api/home doesn't return the now-completed
    item from SWR.
    """
    from rag import _complete_reminder  # noqa: PLC0415
    ok, msg = _complete_reminder(req.reminder_id)
    if not ok:
        raise HTTPException(status_code=400, detail=msg)
    # Bust the home cache so the reminder disappears on next load without
    # waiting for the 60s SWR cycle.
    try:
        _HOME_STATE["ts"] = 0.0
    except Exception:
        pass
    return {"ok": True, "message": msg}


class ReminderDeleteRequest(BaseModel):
    reminder_id: str


@app.post("/api/reminders/delete")
def delete_reminder(req: ReminderDeleteRequest) -> dict:
    """Permanently delete an Apple Reminder. Called from the chat's
    auto-create toast Deshacer button — the user just saw a reminder
    land in Reminders.app and wants to undo it within the 10s window.

    POST (not DELETE verb) because the reminder_id is an
    `x-apple-reminderkit://` URI which is painful to URL-encode.
    """
    from rag import _delete_reminder  # noqa: PLC0415
    ok, msg = _delete_reminder(req.reminder_id)
    if not ok:
        raise HTTPException(status_code=400, detail=msg)
    try:
        _HOME_STATE["ts"] = 0.0
    except Exception:
        pass
    return {"ok": True, "message": msg}


class CalendarDeleteRequest(BaseModel):
    event_uid: str


@app.post("/api/calendar/delete")
def delete_calendar_event(req: CalendarDeleteRequest) -> dict:
    """Permanently delete a Calendar.app event by its UID. Companion
    to the chat's auto-create toast Deshacer button for events.
    """
    from rag import _delete_calendar_event  # noqa: PLC0415
    ok, msg = _delete_calendar_event(req.event_uid)
    if not ok:
        raise HTTPException(status_code=400, detail=msg)
    return {"ok": True, "message": msg}


# Cap defensivo de upload — fotos típicas de iPhone son 4-8MB, dejamos
# margen de manija. Above this returns 413.
_CHAT_UPLOAD_MAX_BYTES = 12 * 1024 * 1024  # 12 MB
_CHAT_UPLOAD_DIR = Path.home() / ".local" / "share" / "obsidian-rag" / "chat-uploads"
# Confidence threshold para auto-crear vs mostrar card de confirmación.
# Decidido con el user 2026-04-25: ≥0.85 auto-crea con undo, <0.85
# va a card. Mantiene el balance entre velocidad y seguridad — el
# detector de citas es bastante conservador con su confidence ya, así
# que confidence ≥ 0.85 quiere decir "esto es claramente un evento/
# reminder, ahorrale el click al user".
_CHAT_UPLOAD_AUTOCREATE_CONFIDENCE = 0.85

# Audit 2026-04-25 finding R2-OCR #2: si la fecha detectada está más
# vieja que este umbral (días en el pasado), bajamos el confidence
# para forzar `needs_confirmation` aunque el detector estuviera muy
# seguro. Caso real del audit: foto de un ticket de cine de hace 2
# años → "Cine 15/06 20:00" → auto-creaba un evento histórico en el
# calendario sin que el user pueda intervenir. 30 días es un trade-off
# razonable: cubre tickets/recibos viejos pero no penaliza fotos de
# eventos de "el mes pasado" que el user quiere agendar como referencia.
_CHAT_UPLOAD_HISTORICAL_DAYS_DEFAULT = 30
try:
    _CHAT_UPLOAD_HISTORICAL_DAYS = int(
        os.environ.get("RAG_OCR_HISTORICAL_MAX_DAYS", _CHAT_UPLOAD_HISTORICAL_DAYS_DEFAULT)
    )
    if _CHAT_UPLOAD_HISTORICAL_DAYS < 0:
        _CHAT_UPLOAD_HISTORICAL_DAYS = _CHAT_UPLOAD_HISTORICAL_DAYS_DEFAULT
except (TypeError, ValueError):
    _CHAT_UPLOAD_HISTORICAL_DAYS = _CHAT_UPLOAD_HISTORICAL_DAYS_DEFAULT

# Mapeo suffix → format param de PIL.save. HEIC/HEIF se agregan
# condicionalmente si `pillow-heif` está instalado (audit 2026-04-25
# R2-OCR #4 followup) y se mappean a "JPEG" porque PIL puede *leer* HEIC
# con pillow-heif pero el writer HEIF requiere libheif con el encoder
# habilitado, que no siempre está. Re-encodear a JPEG es lossy pero
# ganamos privacidad (EXIF/GPS strippeado) + compatibilidad universal
# (todo browser y Obsidian renderean JPEG sin plugins).
#
# Trade-off deliberado: el archivo en el vault termina con extensión
# `.heic` PERO el contenido es JPEG. Es confuso pero el alternativo
# (cambiar la extensión) requiere romper el contrato `(bytes, suffix)
# → bytes` del sanitizer. Obsidian/Safari renderean igual porque
# detectan por magic bytes, no por extensión. Si en el futuro queremos
# extensión consistente, el caller (`upload_chat_image`) debería
# decidir el suffix final basado en `_HEIC_AVAILABLE`.
_SANITIZABLE_FORMATS: dict[str, str] = {
    ".jpg": "JPEG",
    ".jpeg": "JPEG",
    ".png": "PNG",
    ".webp": "WEBP",
    ".gif": "GIF",
}
if _HEIC_AVAILABLE:
    _SANITIZABLE_FORMATS[".heic"] = "JPEG"
    _SANITIZABLE_FORMATS[".heif"] = "JPEG"


def _sanitize_image_exif(raw_bytes: bytes, suffix: str) -> bytes:
    """Re-encode la imagen sin EXIF/GPS/metadata para evitar fugas de
    privacidad cuando se copia al vault iCloud.

    PIL al cargar con ``Image.open()`` lee el EXIF en ``img.info``;
    cuando re-encodeamos creando una nueva imagen con ``putdata()`` y
    saving sin pasar ``exif=...`` explícito, los metadatos no
    sobreviven al roundtrip.

    Args:
      raw_bytes: bytes de la imagen original (del upload).
      suffix: extensión con punto (ej. ``.jpg``). Determina el formato
        de re-encoding.

    Returns:
      Bytes sanitizados si el formato es soportado, o ``raw_bytes`` sin
      cambios si:

      - El suffix no está en ``_SANITIZABLE_FORMATS`` (HEIC/HEIF solo si
        pillow-heif está instalado; ver ``_HEIC_AVAILABLE``)
      - PIL falla al cargar la imagen (corrupta, formato exótico)
      - Hay cualquier otra excepción durante el re-encoding

      Falla suave porque la sanitización es nice-to-have — preferimos
      copiar la foto con EXIF a NO copiarla. Si el user pierde
      privacidad, al menos no pierde la nota.
    """
    fmt = _SANITIZABLE_FORMATS.get(suffix.lower())
    if fmt is None:
        return raw_bytes
    try:
        from PIL import Image  # noqa: PLC0415
        import io  # noqa: PLC0415

        with Image.open(io.BytesIO(raw_bytes)) as img:
            img.load()  # force decode (PIL is lazy)
            # Crear imagen NUEVA con los mismos píxeles raw pero sin
            # heredar el `info` dict del source (que contiene EXIF,
            # ICC, XMP, etc.). `tobytes()` + `frombytes()` es la API
            # estable que no usa el `getdata()` deprecated en Pillow 14.
            stripped = Image.frombytes(img.mode, img.size, img.tobytes())
            buf = io.BytesIO()
            stripped.save(buf, format=fmt)
            return buf.getvalue()
    except Exception:
        # Cualquier error → fallback al raw original. Privacidad
        # degradada pero la nota se guarda.
        return raw_bytes


@app.post("/api/chat/upload-image")
async def upload_chat_image(file: UploadFile = File(...)) -> dict:
    """Subir una imagen al chat para que la procesemos con OCR + VLM
    fallback + detector de cita, y o bien (a) creemos directamente el
    evento/recordatorio si el detector está muy seguro (confidence
    ≥0.85, mismo flujo que el chip auto-creado), o bien (b) emitamos
    un proposal card para que el user confirme manualmente cuando el
    detector está dudoso.

    Acción según el contenido detectado:

    - ``kind == "note"`` o no se detectó nada agendable → ``action="noop"``,
      el frontend muestra un mini-card minimal "imagen procesada, sin
      fecha detectada" (o simplemente nada).
    - ``kind == "event"`` o ``kind == "reminder"`` con confidence
      ≥0.85 y fecha parseable → llamamos ``propose_calendar_event`` /
      ``propose_reminder`` que CREAN el evento real → ``action="created"``,
      response incluye ``event_uid`` o ``reminder_id`` para que el
      frontend muestre el chip auto-creado con [↩ deshacer].
    - Cualquier otro caso (confidence baja, NL no parseable) →
      ``action="needs_confirmation"`` con los fields normalizados
      (mismo shape que la respuesta de ``propose_*`` cuando emite
      ``needs_clarification: true``). El frontend renderiza la card
      estándar de proposal y el user clickea [✓ Crear] para confirmar.

    Persistimos la imagen en
    ``~/.local/share/obsidian-rag/chat-uploads/<sha256>.<ext>``
    (key por hash → dedup natural). NO se copia al vault iCloud
    automáticamente — eso lo hace ``rag capture --image`` cuando el
    user pide capturar explícitamente; acá la imagen es solo input
    para el detector.

    Validaciones: tipo MIME ``image/*``, tamaño ≤12MB. Errores de OCR
    o detector se manejan silenciosamente (devuelve ``action="noop"``
    con un ``reason``) — la idea es nunca bloquear al user que
    simplemente quiso compartir una imagen.
    """
    content_type = (file.content_type or "").lower().strip()
    if not content_type.startswith("image/"):
        raise HTTPException(
            status_code=400,
            detail=f"solo imágenes (recibí content_type={content_type!r})",
        )
    # Audit security 2026-04-26 (HIGH-DoS): pre-fix `await file.read()` sin
    # cap → un cliente con multipart de 500MB ahoga la RAM del proceso ANTES
    # de que el gate de tamaño dispare. Uvicorn no tiene default body-size
    # limit. Lectura cap'd-to-MAX+1: si el body es > MAX, leemos sólo lo
    # suficiente para detectar el overflow y rechazamos.
    raw = await file.read(_CHAT_UPLOAD_MAX_BYTES + 1)
    if not raw:
        raise HTTPException(status_code=400, detail="imagen vacía")
    if len(raw) > _CHAT_UPLOAD_MAX_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"imagen muy grande (>{_CHAT_UPLOAD_MAX_BYTES} bytes)",
        )
    # Magic-byte verification — pre-fix confiabamos en `Content-Type` header
    # del cliente (trivial spoofear). Atacante mandaba `image/jpeg` con
    # body de ZIP/pickle/PDF → llegaba a disco + ocrmac (Apple Vision).
    # Whitelist de signatures comunes:
    _IMAGE_MAGIC_PREFIXES = (
        b"\xff\xd8\xff",          # JPEG
        b"\x89PNG\r\n\x1a\n",     # PNG
        b"GIF87a", b"GIF89a",     # GIF
        b"RIFF",                   # WebP (RIFF...WEBP) — verifica WEBP en bytes 8-12
        b"\x00\x00\x00",           # HEIC/HEIF (ftypheic, ftypheix, ftypmif1, etc — bytes 4-12)
    )
    _has_magic = any(raw.startswith(p) for p in _IMAGE_MAGIC_PREFIXES)
    if _has_magic and raw.startswith(b"RIFF"):
        # WebP requiere WEBP en bytes 8-12
        _has_magic = len(raw) >= 12 and raw[8:12] == b"WEBP"
    if _has_magic and raw.startswith(b"\x00\x00\x00"):
        # HEIC/HEIF requiere "ftyp" en bytes 4-8 + heic/heix/mif1/msf1/heim/etc
        _has_magic = (
            len(raw) >= 12 and raw[4:8] == b"ftyp"
            and raw[8:12] in (b"heic", b"heix", b"hevc", b"hevx",
                              b"mif1", b"msf1", b"heim", b"heis",
                              b"hevm", b"hevs", b"avif")
        )
    if not _has_magic:
        raise HTTPException(
            status_code=400,
            detail="contenido no es una imagen válida (magic bytes no matchean)",
        )

    import hashlib  # noqa: PLC0415
    img_hash = hashlib.sha256(raw).hexdigest()[:32]
    suffix = Path(file.filename or "img.jpg").suffix.lower() or ".jpg"
    if suffix not in {".jpg", ".jpeg", ".png", ".heic", ".heif", ".webp", ".gif"}:
        suffix = ".jpg"  # forzar extensión conocida para que ocrmac no se confunda
    _CHAT_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    img_path = _CHAT_UPLOAD_DIR / f"{img_hash}{suffix}"
    if not img_path.exists():
        img_path.write_bytes(raw)

    # Copia al vault 00-Inbox para consistencia con CLI `rag capture
    # --image`. El CLI siempre copia al vault porque "una foto que el
    # user subió" es info que vive en el vault de Obsidian, no en un
    # cache de runtime. Antes el endpoint web NO lo hacía → las fotos
    # subidas via /chat eran invisibles desde Obsidian. Ahora ambos
    # paths terminan en el mismo lugar (decidido 2026-04-25, C.2).
    #
    # Naming: `<timestamp>-<hash8>.<ext>` para que sea ordenable
    # cronológicamente en el inbox. Idempotente — si ya existe (hash
    # igual = mismo contenido), no lo re-escribimos.
    #
    # **EXIF sanitization (audit 2026-04-25 #5 + #4 followup)**: las fotos
    # del iPhone incluyen EXIF con GPS coords + fecha + cámara. El vault
    # sincroniza con iCloud, así que sin sanitización los metadatos
    # viajan a la nube. PIL re-codifica la imagen sin EXIF (JPEG/PNG/
    # WebP/GIF nativos; HEIC/HEIF si `pillow-heif` está instalado, en
    # cuyo caso re-encodeamos a JPEG — ver comentario en
    # `_SANITIZABLE_FORMATS`). Si pillow-heif no está disponible, HEIC
    # cae al passthrough original con EXIF (mejor que rechazar el upload
    # entero).
    try:
        import rag as _rag  # noqa: PLC0415
        from datetime import datetime as _dt  # noqa: PLC0415
        vault_inbox = _rag.VAULT_PATH / "00-Inbox"
        if vault_inbox.is_dir() or _rag.VAULT_PATH.is_dir():
            vault_inbox.mkdir(parents=True, exist_ok=True)
            ts = _dt.now().strftime("%Y%m%d-%H%M%S")
            vault_img_name = f"{ts}-{img_hash[:8]}{suffix}"
            vault_img_path = vault_inbox / vault_img_name
            if not vault_img_path.exists():
                # Buscar si ya hay un archivo con el mismo hash8 (subida
                # repetida del mismo binary) — si está, dedup natural.
                existing = list(vault_inbox.glob(f"*-{img_hash[:8]}{suffix}"))
                if not existing:
                    sanitized = _sanitize_image_exif(raw, suffix)
                    vault_img_path.write_bytes(sanitized)
    except Exception:
        # Silent-fail: copia al vault es nice-to-have, no debe romper
        # el flujo principal de OCR + detección.
        pass

    # OCR + VLM fallback. Reset del budget VLM por sesión para que el
    # cap "max captions per run" no nos bloquee si el daemon llevaba
    # rato corriendo y ya consumió su budget.
    try:
        from rag.ocr import (  # noqa: PLC0415
            _image_text_or_caption,
            _detect_cita_from_ocr,
            _vlm_caption_budget_reset,
        )
    except ImportError as e:
        raise HTTPException(status_code=500, detail=f"OCR backend no disponible: {e}") from e

    try:
        _vlm_caption_budget_reset()
    except Exception:
        pass

    ocr_text = ""
    ocr_source = ""
    # OCR + VLM caption es síncrono (ocrmac llama a Apple Vision via
    # PyObjC; VLM hace HTTP a Ollama). Si lo llamamos directo desde el
    # handler async, BLOQUEA el event loop entero hasta 60s (timeout
    # de _vlm_client). Lo movemos a un thread y le ponemos timeout
    # duro de 20s — VLM normal tarda 2-4s, cold-load del modelo ~10s,
    # 20s es margen para casos lentos sin colgar el server. Si excede,
    # devolvemos noop con razón explícita y la imagen igual quedó
    # cacheada en disco para retry manual.
    try:
        ocr_text, ocr_source = await asyncio.wait_for(
            asyncio.to_thread(_image_text_or_caption, img_path),
            timeout=20.0,
        )
    except asyncio.TimeoutError:
        return {
            "action": "noop",
            "reason": "ocr_timeout",
            "image_path": str(img_path),
        }
    except Exception as e:
        return {
            "action": "noop",
            "reason": f"ocr_error: {e}",
            "image_path": str(img_path),
        }

    if not ocr_text or len(ocr_text.strip()) < 20:
        return {
            "action": "noop",
            "reason": "ocr_empty_or_short",
            "ocr_text": ocr_text,
            "ocr_source": ocr_source,
            "image_path": str(img_path),
        }

    # ── Dedup check: si ya procesamos este texto OCR antes y ya
    #    creamos un evento/reminder, NO crear duplicado. La tabla
    #    `rag_cita_detections` (compartida con CLI `rag capture --image`)
    #    indexa por sha256 del OCR normalizado. Usuario sube misma foto
    #    2 veces → 2da vez devuelve `action="noop"` con el id original.
    try:
        from rag.ocr import _ocr_hash_key, _persist_cita_detection  # noqa: PLC0415
        from rag import _ragvec_state_conn  # noqa: PLC0415
        ocr_key = _ocr_hash_key(ocr_text)
        with _ragvec_state_conn() as _conn:
            prior = _conn.execute(
                "SELECT decision, kind, title, start_text, location, "
                "confidence, event_uid, reminder_id "
                "FROM rag_cita_detections WHERE ocr_hash = ?",
                (ocr_key,),
            ).fetchone()
    except Exception:
        prior = None
        ocr_key = ""

    if prior is not None:
        prior_decision = prior[0]
        # Solo cortocircuitamos si ya hubo creación real (cita/reminder).
        # Para `low_confidence`/`ambiguous`/`note`, re-procesamos: el
        # detector puede haber mejorado o el user puede querer
        # re-ofrecer la card de confirmación.
        if prior_decision in ("cita", "reminder"):
            return {
                "action": "noop",
                "reason": f"already_processed_{prior_decision}",
                "kind": prior[1] or "",
                "fields": {
                    "title": prior[2],
                    "start_text": prior[3],
                    "location": prior[4],
                },
                "event_uid": prior[6],
                "reminder_id": prior[7],
                "confidence": prior[5],
                "ocr_text": ocr_text[:500],
                "ocr_source": ocr_source,
                "image_path": str(img_path),
            }

    # Detector LLM (qwen2.5:3b por default) también es blocking.
    # Mismo tratamiento: thread + timeout 15s. El detector tarda 1-3s
    # en warm + ~10s en cold-load del modelo. 15s cubre casos lentos
    # sin freezar el server.
    try:
        detection = await asyncio.wait_for(
            asyncio.to_thread(_detect_cita_from_ocr, ocr_text),
            timeout=15.0,
        )
    except asyncio.TimeoutError:
        return {
            "action": "noop",
            "reason": "detector_timeout",
            "ocr_text": ocr_text[:500],
            "ocr_source": ocr_source,
            "image_path": str(img_path),
        }
    except Exception as e:
        return {
            "action": "noop",
            "reason": f"detector_error: {e}",
            "ocr_text": ocr_text[:500],
            "ocr_source": ocr_source,
            "image_path": str(img_path),
        }

    if not detection:
        return {
            "action": "noop",
            "reason": "no_cita_detected",
            "ocr_text": ocr_text[:500],
            "ocr_source": ocr_source,
            "image_path": str(img_path),
        }

    kind = detection.get("kind") or ""
    title = (detection.get("title") or "").strip() or "(sin título)"
    when = (detection.get("when") or "").strip()
    location = (detection.get("location") or "").strip()
    confidence = float(detection.get("confidence") or 0.0)
    notes = (
        f"Detectado de imagen subida al chat ({ocr_source.upper() or 'OCR'})\n\n"
        f"{ocr_text[:1000]}"
    )

    # Helper local para persistir en rag_cita_detections, idempotente.
    # El dedup check al inicio del endpoint solo cortocircuita si ya
    # creamos algo. Persistimos AHORA (después de procesar) para que
    # uploads futuros del mismo OCR encuentren el row.
    def _persist(decision: str, *, event_uid=None, reminder_id=None):
        if not ocr_key:
            return
        try:
            import time as _time  # noqa: PLC0415
            _persist_cita_detection(
                ocr_hash=ocr_key,
                image_path=str(img_path),
                source="web-chat",
                decision=decision,
                kind=kind,
                title=title,
                start_text=when,
                location=location,
                confidence=confidence,
                event_uid=event_uid,
                reminder_id=reminder_id,
                created_at=_time.time(),
            )
        except Exception:
            pass  # silent-fail: dedup miss en próximo upload no es crítico

    if kind not in {"event", "reminder"}:
        # kind == "note" o algo raro → no proponemos nada agendable.
        # Persistimos para evitar re-detectar el mismo texto.
        _persist("note")
        return {
            "action": "noop",
            "reason": "kind_not_actionable",
            "kind": kind,
            "ocr_text": ocr_text[:500],
            "ocr_source": ocr_source,
            "detection": detection,
            "image_path": str(img_path),
        }

    # Audit 2026-04-25 finding R2-OCR #2: si el detector está muy
    # seguro PERO la fecha detectada es histórica (>30 días en el
    # pasado), forzamos needs_confirmation. Foto de un ticket viejo no
    # debería crear un evento sin confirmación explícita. Parsing del
    # `when` reusa `_parse_natural_datetime` (la misma función que el
    # fallback de needs_confirmation usa más abajo). Si no parsea, lo
    # dejamos pasar — el routing normal por confidence decide.
    if (
        confidence >= _CHAT_UPLOAD_AUTOCREATE_CONFIDENCE
        and when
        and _CHAT_UPLOAD_HISTORICAL_DAYS > 0
    ):
        try:
            from rag import _parse_natural_datetime  # noqa: PLC0415
            from datetime import datetime as _dt  # noqa: PLC0415
            _now = _dt.now()
            _when_dt = _parse_natural_datetime(when, now=_now)
            if _when_dt is not None:
                # Normalizar a naive si vino con tz, para comparar
                # contra `_dt.now()` que también es naive (ambos
                # comparten la misma referencia de "ahora local AR").
                if _when_dt.tzinfo is not None:
                    _when_dt = _when_dt.replace(tzinfo=None)
                age_days = (_now - _when_dt).total_seconds() / 86400.0
                if age_days > _CHAT_UPLOAD_HISTORICAL_DAYS:
                    # Downgrade — forzamos needs_confirmation. 0.5 es
                    # un valor neutro: por debajo del umbral de
                    # auto-create pero no tan bajo que el frontend lo
                    # marque como "ruido".
                    confidence = 0.5
                    try:
                        import rag as _rag  # noqa: PLC0415
                        _rag._ambient_log_event({
                            "cmd": "ocr_historical_downgrade",
                            "kind": kind,
                            "title": title[:200],
                            "when": when[:100],
                            "age_days": round(age_days, 1),
                            "threshold_days": _CHAT_UPLOAD_HISTORICAL_DAYS,
                            "ocr_source": ocr_source,
                        })
                    except Exception:
                        pass
        except Exception:
            # Cualquier error de parsing → no bloqueamos, dejamos
            # que el routing normal decida (best-effort guardrail).
            pass

    # Routing por confidence.
    if confidence >= _CHAT_UPLOAD_AUTOCREATE_CONFIDENCE:
        # Auto-crear via propose_*: estos parsean el NL y crean al toque.
        # Si el NL no parsea, propose_* devuelve needs_clarification:true
        # — caemos al fallback de needs_confirmation abajo.
        try:
            if kind == "event":
                from rag import propose_calendar_event  # noqa: PLC0415
                resp_json = propose_calendar_event(
                    title=title,
                    start=when or "",
                    location=location or None,
                    notes=notes,
                )
            else:
                from rag import propose_reminder  # noqa: PLC0415
                resp_json = propose_reminder(
                    title=title,
                    when=when or "",
                    notes=notes,
                )
            resp = json.loads(resp_json)
        except Exception as e:
            resp = {"created": False, "error": f"propose_*_failed: {e}"}

        if resp.get("created"):
            id_field = "event_uid" if kind == "event" else "reminder_id"
            new_id = resp.get(id_field)
            # Persistir como cita/reminder con el id real → próximas
            # subidas de la misma foto detectan el dedup y devuelven
            # noop sin re-crear.
            if kind == "event":
                _persist("cita", event_uid=new_id)
            else:
                _persist("reminder", reminder_id=new_id)
            return {
                "action": "created",
                "kind": kind,
                "fields": resp.get("fields", {}),
                id_field: new_id,
                "ocr_text": ocr_text[:500],
                "ocr_source": ocr_source,
                "confidence": confidence,
                "image_path": str(img_path),
            }
        # Si auto-crear falló (NL no parseable), caemos a confirmación.

    # Confidence baja o auto-crear no pudo: emitir proposal sin crear.
    # Hacemos un best-effort de parsear el `when` para popular
    # ``start_iso`` / ``due_iso`` y que el frontend pueda mandar el
    # POST a /api/calendar/create directamente cuando el user confirme.
    parsed_iso: str | None = None
    if when:
        try:
            from rag import _parse_natural_datetime  # noqa: PLC0415
            from datetime import datetime as _dt  # noqa: PLC0415
            parsed = _parse_natural_datetime(when, now=_dt.now())
            if parsed:
                parsed_iso = parsed.isoformat()
        except Exception:
            parsed_iso = None

    if kind == "event":
        fields = {
            "title": title,
            "start_iso": parsed_iso,
            "start_text": when,
            "end_iso": None,
            "calendar": None,
            "location": location or None,
            "notes": notes,
            "all_day": False,
            "recurrence": None,
            "recurrence_text": None,
        }
    else:  # reminder
        fields = {
            "title": title,
            "due_iso": parsed_iso,
            "due_text": when,
            "list": None,
            "priority": None,
            "notes": notes,
            "recurrence": None,
            "recurrence_text": None,
        }

    # Persistir como 'ambiguous' (low_confidence o NL no parseable):
    # NO se creó nada, así que si re-suben la foto queremos darle la
    # chance al detector de re-procesar. El dedup check al inicio
    # SOLO short-circuita decisions in {cita, reminder}, no estos.
    _persist("ambiguous")

    return {
        "action": "needs_confirmation",
        "kind": kind,
        "fields": fields,
        "needs_clarification": parsed_iso is None and not when,
        "ocr_text": ocr_text[:500],
        "ocr_source": ocr_source,
        "confidence": confidence,
        "image_path": str(img_path),
    }


class WhatsAppReplyTarget(BaseModel):
    """Optional reply context for `/api/whatsapp/send` — when the user
    clicked [Enviar] on a `propose_whatsapp_reply` card, the UI ships
    the resolved original message so the bridge can wire a native quote
    when it gains support (currently the bridge ignores these fields,
    see ``rag._whatsapp_send_to_jid`` docstring; the audit log keeps
    `message_id` regardless so we can correlate later)."""
    message_id: str
    original_text: str | None = ""
    sender_jid: str | None = ""


class WhatsAppSendRequest(BaseModel):
    jid: str                      # e.g. "5491234567890@s.whatsapp.net"
    message_text: str             # literal message body (no anti-loop prefix)
    proposal_id: str | None = None  # for audit — id of the draft the UI confirmed
    reply_to: WhatsAppReplyTarget | None = None  # populated when this is a reply
    # Si viene populado, en vez de mandar al bridge, programamos el
    # envío via `rag.wa_scheduled.schedule()`. Acepta ISO8601 con
    # offset (LLM emite "2026-04-26T09:00:00-03:00") o sin offset
    # (datetime-local del HTML, asumimos TZ Argentina). Vacío / null →
    # legacy path (envío inmediato).
    scheduled_for: str | None = None
    contact_name: str | None = None  # opcional, informativo (lo guarda la tabla)


_AR_OFFSET = "-03:00"
"""Default TZ Argentina (no DST). Aplicado a inputs sin offset."""


def _parse_scheduled_for_to_utc(raw: str) -> str:
    """Acepta el ``scheduled_for`` que viene del frontend y devuelve un
    ISO8601 UTC ``"YYYY-MM-DDTHH:MM:SS+00:00"`` listo para guardar en
    ``rag_whatsapp_scheduled``.

    Formatos aceptados:
      - ISO8601 con offset: ``"2026-04-26T09:00:00-03:00"`` (lo que
        emite el LLM en la proposal).
      - ISO8601 sin offset: ``"2026-04-26T09:00"`` o
        ``"2026-04-26T09:00:00"`` (lo que emite ``<input
        type="datetime-local">``). Asumimos TZ Argentina (-03:00).

    Raisea ``ValueError`` con mensaje legible si no parsea — el caller
    convierte a HTTPException 400. NO valida que esté en el futuro
    (eso lo hace ``wa_scheduled.schedule()`` con margen para clock
    skew y cap de horizonte).
    """
    if not isinstance(raw, str):
        raise ValueError("scheduled_for debe ser string ISO8601")
    s = raw.strip()
    if not s:
        raise ValueError("scheduled_for vacío")
    candidate = s
    # ``datetime.fromisoformat`` 3.11+ acepta "Z" como sufijo, pero
    # forzamos consistencia: si no hay offset, asumimos AR.
    has_offset = (
        candidate.endswith("Z")
        or "+" in candidate[10:]
        or candidate[10:].count("-") > 0
    )
    if not has_offset:
        candidate = candidate + _AR_OFFSET
    try:
        from datetime import datetime, timezone  # noqa: PLC0415
        dt = datetime.fromisoformat(candidate.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            raise ValueError("falta timezone tras normalización")
        return dt.astimezone(timezone.utc).isoformat(timespec="seconds")
    except ValueError as e:
        raise ValueError(f"scheduled_for inválido: {e}") from e


@app.post("/api/whatsapp/send")
def whatsapp_send(req: WhatsAppSendRequest) -> dict:
    """Execute a WhatsApp send after the user clicked [Enviar] (envío
    inmediato) o [Programar] (envío diferido) on a ``propose_whatsapp_send``
    / ``propose_whatsapp_reply`` proposal card.

    Flow:
      1. `propose_whatsapp_send` (chat tool) resolves contact → JID and
         emits an SSE `proposal` event with ``kind="whatsapp_message"``.
         For replies the tool is `propose_whatsapp_reply` and the kind is
         ``"whatsapp_reply"`` — the card carries an extra ``reply_to``
         field with the resolved original message.
      2. The UI renders a card with the drafted text and [Enviar] /
         [Editar] / [Cancelar] buttons. When it's a reply the original
         message is shown above the textarea as a styled blockquote.
      3. On [Enviar] the UI POSTs to this endpoint with `{jid,
         message_text, proposal_id, reply_to?, scheduled_for?}`.
         - Sin ``scheduled_for``: llamamos `_whatsapp_send_to_jid` con
           ``anti_loop=False`` y se entrega ya.
         - Con ``scheduled_for``: NO mandamos al bridge — guardamos el
           row en ``rag_whatsapp_scheduled`` (status='pending') y el
           plist ``com.fer.obsidian-rag-wa-scheduled-send`` lo entrega
           cuando vence. Devolvemos ``{ok, scheduled: true, id,
           scheduled_for_utc}``.

    Replies (``reply_to`` present):
      - The bridge today (`whatsapp-mcp/whatsapp-bridge/main.go:707`)
        does NOT support `ContextInfo`/`QuotedMessage` — the message
        ships as plain text without WhatsApp's native boxed-quote UI.
        We still pass the field forward so when the bridge adds quote
        support it works without client changes.
      - The audit log records ``reply_to_id`` so we can correlate the
        outbound message with the original incoming message later.

    Returns ``{ok: true, jid}`` (immediate) or ``{ok: true, scheduled:
    true, id, scheduled_for_utc}`` (deferred). Raises 400 on shape
    problems, 502 on bridge unreachable. Rate-limited by the same IP
    bucket as ``/api/chat``.
    """
    jid = (req.jid or "").strip()
    body = (req.message_text or "").strip()
    if not jid or "@" not in jid:
        raise HTTPException(status_code=400, detail="jid inválido (debe ser '<digits>@s.whatsapp.net' o '<id>@g.us')")
    if not body:
        raise HTTPException(status_code=400, detail="message_text vacío")
    reply_to_payload: dict | None = None
    reply_to_id = ""
    if req.reply_to:
        reply_to_id = (req.reply_to.message_id or "").strip()
        if not reply_to_id:
            raise HTTPException(status_code=400, detail="reply_to.message_id vacío")
        reply_to_payload = {
            "message_id": reply_to_id,
            "original_text": req.reply_to.original_text or "",
            "sender_jid": req.reply_to.sender_jid or "",
        }

    # ── Deferred path: si viene scheduled_for lo guardamos y salimos. ─
    if req.scheduled_for:
        try:
            scheduled_for_utc = _parse_scheduled_for_to_utc(req.scheduled_for)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e
        try:
            from rag import wa_scheduled  # noqa: PLC0415
            row = wa_scheduled.schedule(
                jid=jid,
                message_text=body,
                scheduled_for_utc=scheduled_for_utc,
                contact_name=(req.contact_name or "").strip() or None,
                reply_to=reply_to_payload,
                proposal_id=req.proposal_id or None,
                source="chat",
            )
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e
        return {"ok": True, "scheduled": True, **row}

    # ── Immediate path: legacy (sin cambios de comportamiento). ───────
    from rag import _whatsapp_send_to_jid  # noqa: PLC0415
    ok = _whatsapp_send_to_jid(
        jid, body,
        anti_loop=False,
        reply_to=reply_to_payload,
    )
    if not ok:
        # Can't distinguish "bridge down" from "bridge rejected" cheaply
        # — the helper swallows the HTTP code. Return 502 (bad gateway)
        # to signal an upstream issue vs 400 (client-side) for shape
        # problems caught above.
        raise HTTPException(status_code=502, detail="bridge WhatsApp no respondió (localhost:8080 no disponible o rechazó el mensaje)")
    # Log minimal audit trail — we don't persist the message body for
    # privacy, just the event + jid + proposal_id so we can correlate
    # with the chat transcript later if needed.
    try:
        import rag as _rag  # noqa: PLC0415
        _rag._ambient_log_event({
            "cmd": "whatsapp_user_reply" if reply_to_payload else "whatsapp_user_send",
            "jid": jid,
            "len": len(body),
            "proposal_id": req.proposal_id or "",
            "reply_to_id": reply_to_id,
            "sent": True,
        })
    except Exception:
        pass
    return {"ok": True, "jid": jid}


class WhatsAppMatchRequest(BaseModel):
    text: str = Field(..., description="Texto del input del usuario tal cual lo escribe.")


# Verbos de envío de WhatsApp + capturas del candidato (rioplatense + neutral).
# Diseñado para matchear en MEDIO del texto (no anchored al inicio) — el user
# puede escribir "Recordame que ... y mandale a Mama que diga ..." y queremos
# detectar "Mama" igual.
_WA_SEND_VERB_RE = re.compile(
    r"\b(mand[aá]le|manda|escribi?le|dec[iíĩ]le|avis[aá]le|record[aá]le)\s+"
    r"(?:un\s+)?(?:(?:mensaje|msj|msg|wzp|wa|whatsapp)(?:\s+(?:de|por|via|v[ií]a)\s+(?:wa|whatsapp))?\s+)?"
    r"a\s+",
    re.IGNORECASE,
)


@app.post("/api/whatsapp/contacts/match")
def whatsapp_contacts_match(req: WhatsAppMatchRequest) -> dict:
    """Detección incremental de destinatario mientras el user escribe.

    Llamado con debounce desde el frontend cada vez que cambia el textarea
    del chat. Si el texto contiene un verbo de envío + un nombre que el
    sistema puede resolver (vía Related Names del My Card o Apple Contacts),
    devuelve el match para que el frontend highlightee el nombre y muestre
    una confirmación visual ANTES del envío.

    Diseñado para ser fast (osascript cached) — el frontend lo invoca cada
    ~300ms. Silent-fail para todo error (devuelve `match: None`).

    Returns::

        {
          "match": {
            "name": "Mamá",                      # nombre canónico del contacto
            "phone": "+54 9 342 547 6623",
            "jid": "5493425476623@s.whatsapp.net",
            "source": "relations" | "contacts",  # de dónde salió
            "matched_token": "Mama",             # lo que el user escribió
            "match_offset": 22,                  # posición en `text` (chars)
            "match_length": 4
          },
          "trigger": {                           # opcional
            "verb": "mandale",
            "offset": 0
          }
        } | { "match": None, "trigger": None }
    """
    text = (req.text or "").strip()
    if not text:
        return {"match": None, "trigger": None}

    m_verb = _WA_SEND_VERB_RE.search(text)
    if not m_verb:
        return {"match": None, "trigger": None}

    # Después del "a " del verbo, viene el candidate string.
    candidate_start = m_verb.end()
    candidate_text = text[candidate_start:].strip()
    if not candidate_text:
        return {"match": None, "trigger": {
            "verb": m_verb.group(1), "offset": m_verb.start(),
        }}

    # Strip posesivos al inicio del candidate ("mi Mama", "a mi Mama", "la
    # Mama") — lookup downstream también lo hace, pero acá adelantamos para
    # que el offset/length que devolvemos al frontend matchee la palabra
    # real que el highlight debe pintar (si dejamos "mi Mama", el highlight
    # incluiría "mi" que NO es parte del nombre).
    from rag.integrations.whatsapp import (
        _whatsapp_jid_from_contact,
        _strip_possessive_prefix,
    )
    candidate_stripped = _strip_possessive_prefix(candidate_text).strip()
    if not candidate_stripped:
        return {"match": None, "trigger": {
            "verb": m_verb.group(1), "offset": m_verb.start(),
        }}

    tokens = candidate_stripped.split()
    if not tokens:
        return {"match": None, "trigger": {
            "verb": m_verb.group(1), "offset": m_verb.start(),
        }}

    # Probar 1 palabra primero — la mayoría de los casos ("Mama", "Maria",
    # "Sebas"). Si no resuelve, probar 2 ("Maria José", "Tía Carmen") y
    # finalmente 3 ("Maria José Pérez"). Sin esto el endpoint matcheaba 3
    # palabras espurias ("Maria que ya") y resolvía via fuzzy a otro
    # contacto distinto — error observado 2026-04-26.
    _RELATION_LOWER = {
        "mama", "mami", "papa", "papi", "madre", "padre",
        "hermana", "hermano", "esposa", "esposo", "marido", "mujer",
        "hijo", "hija", "abuela", "abuelo", "tia", "tio",
        "prima", "primo", "suegra", "suegro",
    }
    best_match = None
    matched_token = ""
    for n in range(1, min(4, len(tokens) + 1)):
        candidate = " ".join(tokens[:n])
        # Filter: si el primer token no arranca con mayúscula Y no es alias
        # de parentesco → skip (es prosa, no nombre).
        first_lower = tokens[0].lower()
        if not tokens[0][:1].isupper() and first_lower not in _RELATION_LOWER:
            break  # sin sentido seguir agregando palabras
        try:
            lookup = _whatsapp_jid_from_contact(candidate)
        except Exception:
            continue
        if lookup.get("jid") and not lookup.get("error") and not lookup.get("is_group"):
            best_match = lookup
            matched_token = candidate
            break

    if not best_match:
        return {"match": None, "trigger": {
            "verb": m_verb.group(1), "offset": m_verb.start(),
        }}

    # Calcular offset/length del matched_token en el texto original.
    # `candidate_text` empieza en `candidate_start`. El primer token ocupa
    # `len(matched_token)` chars (después de strip).
    # Necesitamos la posición real en `text` — buscamos `matched_token`
    # case-insensitive desde `candidate_start`.
    needle_lower = matched_token.lower()
    haystack_lower = text.lower()
    match_offset = haystack_lower.find(needle_lower, candidate_start)
    match_length = len(matched_token)

    # Decidir source: si el resolver vino vía Related Names (full_name distinto
    # del input limpio), es "relations"; si match exact/fuzzy de Contacts directo,
    # es "contacts".
    source = "relations" if (
        best_match.get("full_name") and matched_token.lower() not in best_match["full_name"].lower()
    ) else "contacts"

    return {
        "match": {
            "name": best_match.get("full_name") or matched_token,
            "phone": (best_match.get("phones") or [""])[0],
            "jid": best_match["jid"],
            "source": source,
            "matched_token": matched_token,
            "match_offset": match_offset,
            "match_length": match_length,
        },
        "trigger": {
            "verb": m_verb.group(1),
            "offset": m_verb.start(),
        },
    }


@app.get("/api/whatsapp/context")
def whatsapp_context(jid: str, limit: int = 5) -> dict:
    """Últimos ``limit`` mensajes intercambiados con ``jid`` para mostrar
    contexto al lado del card del chat (cuando el LLM propone un mensaje,
    el user ve la conversación reciente antes de mandar/programar).

    Reemplaza el ruido de los chips ``"seguir con ›"`` que aparecían
    debajo de los proposals — esos preguntan sobre el chat RAG, pero
    cuando ya hay un proposal estructurado lo útil es la conversación
    pasada con el destinatario, no más preguntas al RAG.

    Devuelve estructura siempre válida (``messages_count: 0`` si no hay
    bridge / contacto / mensajes) — el frontend renderiza la sección
    aunque no haya datos.
    """
    j = (jid or "").strip()
    if not j or "@" not in j:
        raise HTTPException(status_code=400, detail="jid inválido")
    try:
        from rag.integrations.whatsapp import _fetch_whatsapp_recent_with_jid  # noqa: PLC0415
        return _fetch_whatsapp_recent_with_jid(j, limit=int(limit))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"error leyendo bridge: {e}") from e


@app.get("/api/whatsapp/scheduled")
def whatsapp_scheduled_list(
    status: str | None = None,
    limit: int = 200,
) -> dict:
    """Lista los mensajes de WhatsApp programados para mostrar en el
    dashboard.

    Query params:
      - ``status``: filtra por estado (``pending`` | ``sent`` |
        ``sent_late`` | ``failed`` | ``cancelled``). Sin filtro = todos.
      - ``limit``: cap de filas devueltas (default 200, sirve hasta
        que la cola supere ese tamaño y queramos paginación).

    Devuelve ``{items: [...]}`` ordenado por ``scheduled_for_utc ASC``
    cuando ``status=pending`` (lo más urgente arriba) o por
    ``created_at DESC`` en otros casos (timeline). El cuerpo de cada
    mensaje viene truncado a 500 chars en este endpoint para no inflar
    payloads del dashboard.
    """
    try:
        from rag import wa_scheduled  # noqa: PLC0415
        items = wa_scheduled.list_scheduled(
            status=status if (status or "").strip() else None,
            limit=max(1, min(int(limit), 1000)),
        )
        return {"items": items, "count": len(items)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"error listando scheduled: {e}") from e


@app.post("/api/whatsapp/scheduled/{scheduled_id}/cancel")
def whatsapp_scheduled_cancel(scheduled_id: int) -> dict:
    """Cancela un mensaje pending. Idempotente: si ya está en otro
    estado, devuelve ``{ok: false, reason}`` sin error 4xx/5xx (el
    dashboard puede simplemente refrescar la fila).
    """
    try:
        from rag import wa_scheduled  # noqa: PLC0415
        ok = wa_scheduled.cancel(int(scheduled_id), reason="user_cancel_dashboard")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"error cancelando: {e}") from e
    if not ok:
        return {"ok": False, "reason": "not_pending_or_not_found"}
    return {"ok": True, "id": int(scheduled_id), "status": "cancelled"}


class WhatsAppRescheduleRequest(BaseModel):
    scheduled_for: str  # mismo formato que WhatsAppSendRequest


@app.post("/api/whatsapp/scheduled/{scheduled_id}/reschedule")
def whatsapp_scheduled_reschedule(
    scheduled_id: int,
    req: WhatsAppRescheduleRequest,
) -> dict:
    """Reprograma un mensaje pending para una nueva fecha/hora. Solo
    funciona si el row está en ``status='pending'`` (mismo guard que
    cancel — no rescheduleamos `sent`/`failed`/`cancelled`).
    """
    try:
        scheduled_for_utc = _parse_scheduled_for_to_utc(req.scheduled_for)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    try:
        from rag import wa_scheduled  # noqa: PLC0415
        ok = wa_scheduled.reschedule(int(scheduled_id), scheduled_for_utc)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"error rescheduleando: {e}") from e
    if not ok:
        return {"ok": False, "reason": "not_pending_or_not_found"}
    return {"ok": True, "id": int(scheduled_id), "scheduled_for_utc": scheduled_for_utc}


class MailSendRequest(BaseModel):
    to: str
    subject: str
    body: str
    cc: str | None = None
    bcc: str | None = None
    proposal_id: str | None = None  # for audit — id of the draft the UI confirmed


@app.post("/api/mail/send")
def mail_send(req: MailSendRequest) -> dict:
    """Execute a Gmail send after the user clicked [Enviar] on a
    ``propose_mail_send`` proposal card.

    Flow idéntico al de WhatsApp:
      1. `propose_mail_send` (chat tool) genera la propuesta + emite SSE
         `proposal` con ``kind="mail"``.
      2. La UI renderiza la card con [Enviar] / [Editar] / [Descartar]
         + textareas editables para subject / body.
      3. En [Enviar] la UI postea acá con `{to, subject, body, cc?,
         bcc?, proposal_id?}`. Llamamos a ``rag._send_gmail`` que usa
         los creds gmail.modify de ``~/.gmail-mcp/credentials.json``.

    Errores:
      - 400 si `to` vacío / inválido, o `body` vacío (subject vacío se
        rellena con "(sin asunto)" silenciosamente).
      - 502 si Gmail API rechaza (creds revocadas, rate limit, etc.) —
        devolvemos el ``error`` del helper como detail.

    Rate-limit implícito: mismo bucket que ``/api/chat`` porque reusa la
    auth del mismo OAuth client. No se loggea el body (privacy) —
    solo el event + to + proposal_id para correlacionar con el turno.
    """
    to_clean = (req.to or "").strip()
    body_clean = (req.body or "").strip()
    if not to_clean or "@" not in to_clean:
        raise HTTPException(status_code=400, detail="to inválido (debe ser un email con @)")
    if not body_clean:
        raise HTTPException(status_code=400, detail="body vacío")
    from rag import _send_gmail  # noqa: PLC0415
    result = _send_gmail(
        to_clean,
        (req.subject or "").strip() or "(sin asunto)",
        body_clean,
        cc=req.cc,
        bcc=req.bcc,
    )
    if not result.get("ok"):
        detail = result.get("error") or "Gmail API no respondió"
        raise HTTPException(status_code=502, detail=f"mail no enviado: {detail}")
    try:
        import rag as _rag  # noqa: PLC0415
        _rag._ambient_log_event({
            "cmd": "mail_user_send",
            "to": to_clean,
            "len": len(body_clean),
            "proposal_id": req.proposal_id or "",
            "message_id": result.get("message_id", ""),
            "sent": True,
        })
    except Exception:
        pass
    return {
        "ok": True,
        "to": to_clean,
        "message_id": result.get("message_id", ""),
        "thread_id": result.get("thread_id", ""),
    }


@app.get("/api/contacts")
def list_contacts(q: str = "", kind: str = "any", limit: int = 20) -> dict:
    """Contact picker para el popover de ``/wzp`` y ``/mail`` del web chat.

    Filtra el cache local de Apple Contacts (``_load_contacts_phone_index``)
    por substring normalizado del nombre — sin acentos, case-insensitive.
    No usa embeddings a propósito: los nombres son strings cortos (≤20
    chars) donde los embeddings no agregan relevancia real y sí latencia.

    Query params:
        q: Substring a matchear (plegado con ``_fold``). Vacío → primeros
            ``limit`` contactos ordenados alfabéticamente (útil para
            mostrar algo apenas el user typea ``/wzp ``).
        kind: ``phone`` | ``email`` | ``any``. Filtra contactos que
            tengan el canal requerido — para ``/wzp`` el popover pide
            ``phone`` (sin teléfono no podemos mandar WA) y para
            ``/mail`` pide ``email``.
        limit: Máx. resultados (1–100, default 20).

    Returns:
        ``{contacts: [{name, phones, emails, score}], query, kind}``.
        ``score`` es 3 (exacto) / 2 (prefix) / 1 (substring) / 0 (sin
        query — orden alfabético). Si el cache está frío / osascript
        falla silently, devuelve lista vacía con los mismos headers.

    **Seguridad**: este endpoint expone nombre + teléfono + email de
    todos los contactos del user. No hay auth — el server está bound
    a ``127.0.0.1`` por default; si se expone al LAN (ver
    ``OBSIDIAN_RAG_ALLOW_LAN`` en ``web/server.py`` y ``CLAUDE.md``)
    el riesgo queda en el mismo perímetro que el resto de la data
    del vault. No loggear los resultados.
    """
    from rag import _fuzzy_filter_contacts  # noqa: PLC0415
    kind_norm = (kind or "any").strip().lower()
    if kind_norm not in {"any", "phone", "email"}:
        kind_norm = "any"
    # `int(limit or 20)` no anda con limit=0 (cae al default de 20). FastAPI
    # ya nos da 20 cuando el query param falta — acá sólo clampamos el rango.
    limit_norm = max(1, min(100, int(limit) if isinstance(limit, int) else 20))
    try:
        contacts = _fuzzy_filter_contacts(q or "", kind=kind_norm, limit=limit_norm)
    except Exception as exc:
        # Silent-fail: el popover no puede bloquear el chat. Devolvemos
        # vacío y loggeamos para debugging.
        try:
            import rag as _rag  # noqa: PLC0415
            _rag._silent_log("api_contacts", exc)
        except Exception:
            pass
        contacts = []
    return {
        "contacts": contacts,
        "query": q or "",
        "kind": kind_norm,
    }


@app.get("/api/model")
def get_chat_model() -> dict:
    """Return the chat model that /api/chat would use right now."""
    return {"model": _resolve_web_chat_model()}


@app.get("/api/session/{sid}")
def get_session_info(sid: str) -> dict:
    """Return session summary for the chat UI `/session` slash command."""
    from rag import load_session  # noqa: PLC0415
    sess = load_session(sid)
    if not sess:
        raise HTTPException(status_code=404, detail="session no encontrada")
    turns = sess.get("turns", [])
    return {
        "id": sess.get("id", sid),
        "mode": sess.get("mode", ""),
        "created_at": sess.get("created_at", ""),
        "updated_at": sess.get("updated_at", ""),
        "turns": len(turns),
        "first_q": (turns[0].get("q", "") if turns else "")[:120],
        "last_q": (turns[-1].get("q", "") if turns else "")[:120],
    }


@app.get("/api/sessions")
def list_web_sessions(limit: int = 40) -> dict:
    """List recent chat sessions for the sidebar (claude.ai-style history).

    Filters to sessions that originated from the web chat (id starts with
    `web:` — see `/api/chat` where `sid = req.session_id or f"web:{uuid…}"`).
    CLI sessions (mode="ask", "do") are excluded so the sidebar doesn't show
    noise from `rag ask` / `rag do` experiments.

    Returns newest-first with: id, title (first non-empty question, trimmed),
    updated_at, turns count. The UI uses `title` as the display label and
    renders "sin título" as fallback for empty sessions.
    """
    from rag import list_sessions  # noqa: PLC0415
    limit = max(1, min(int(limit or 40), 200))
    # Over-fetch 2x so filtering (web-only, non-empty) still yields `limit`.
    raw = list_sessions(limit=limit * 2)
    out: list[dict] = []
    for s in raw:
        sid = (s.get("id") or "").strip()
        if not sid.startswith("web:"):
            continue
        first_q = (s.get("first_q") or "").strip()
        turns = int(s.get("turns") or 0)
        if turns == 0 and not first_q:
            continue  # empty session — skip
        title = first_q[:80] if first_q else "sin título"
        out.append({
            "id": sid,
            "title": title,
            "turns": turns,
            "updated_at": s.get("updated_at", ""),
            "created_at": s.get("created_at", ""),
        })
        if len(out) >= limit:
            break
    return {"sessions": out}


@app.get("/api/session/{sid}/turns")
def get_session_turns(sid: str) -> dict:
    """Return full turn history for rehydrating the chat UI from sidebar click.

    Unlike `/api/session/{sid}` (metadata only), this returns the
    `{q, a, paths, ts}` tuples needed to re-render the conversation bubbles.
    `paths` is kept for source-row rendering; other fields (citations,
    scoring metadata) are left out — the user is viewing a historical
    snapshot, not reopening the retrieval context.
    """
    from rag import load_session  # noqa: PLC0415
    sess = load_session(sid)
    if not sess:
        raise HTTPException(status_code=404, detail="session no encontrada")
    turns = sess.get("turns", []) or []
    out_turns: list[dict] = []
    for t in turns:
        out_turns.append({
            "q": (t.get("q") or ""),
            "a": (t.get("a") or ""),
            "paths": list(t.get("paths") or []),
            "ts": t.get("ts", ""),
        })
    return {
        "id": sess.get("id", sid),
        "mode": sess.get("mode", ""),
        "created_at": sess.get("created_at", ""),
        "updated_at": sess.get("updated_at", ""),
        "turns": out_turns,
    }


class SaveRequest(BaseModel):
    session_id: str
    title: str | None = None


@app.post("/api/save")
def save_conversation(req: SaveRequest) -> dict:
    """Persist a chat session into the vault as a single Markdown note.

    The CLI's `/save` only saves the last turn; here we flatten the whole
    session so the user can export a back-and-forth thread. Lands in
    00-Inbox with source_metas aggregated across all turns.
    """
    from rag import get_db_for, load_session, save_note  # noqa: PLC0415
    sess = load_session(req.session_id)
    if not sess or not sess.get("turns"):
        raise HTTPException(status_code=404, detail="session vacía o inexistente")
    turns = sess["turns"]
    body_parts: list[str] = []
    for t in turns:
        q = (t.get("q") or "").strip()
        a = (t.get("a") or "").strip()
        if not q and not a:
            continue
        body_parts.append(f"**Pregunta:** {q}\n\n{a}")
    body = "\n\n---\n\n".join(body_parts)
    question = turns[0].get("q", "") or "chat"
    vaults = resolve_vault_paths(None)
    if not vaults:
        raise HTTPException(status_code=400, detail="no hay vault activo")
    col = get_db_for(vaults[0][1])
    seen: set[str] = set()
    source_metas: list[dict] = []
    for t in turns:
        for p in (t.get("paths") or []):
            if p and p not in seen:
                seen.add(p)
                source_metas.append({
                    "file": p,
                    "note": Path(p).stem,
                    "tags": "",
                })
    path = save_note(col, req.title, body, question, source_metas)
    try:
        rel = str(path.relative_to(vaults[0][1]))
    except ValueError:
        rel = str(path)
    return {"ok": True, "path": rel}


@app.post("/api/reindex")
def trigger_reindex() -> dict:
    """Fire-and-forget incremental reindex. `--reset` not exposed via web."""
    try:
        subprocess.Popen(
            ["rag", "index"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            start_new_session=True,
        )
    except FileNotFoundError:
        raise HTTPException(status_code=500, detail="rag CLI no encontrado en PATH")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    return {"ok": True, "message": "index lanzado en background"}


class FollowupsRequest(BaseModel):
    session_id: str


@app.post("/api/followups")
def followups(req: FollowupsRequest) -> dict:
    """Generate 3 follow-up question chips from the last turn.

    Uses the helper model (qwen2.5:3b) with deterministic temp to keep
    suggestions tight and cheap (~400-800ms on M-series). Returns empty
    list on any error so the UI fails silently.
    """
    from rag import load_session, get_db  # noqa: PLC0415
    sess = load_session(req.session_id)
    if not sess or not sess.get("turns"):
        return {"followups": []}
    last = sess["turns"][-1]
    q = (last.get("q") or "").strip()
    a = (last.get("a") or "").strip()[:800]
    paths = (last.get("paths") or [])[:3]
    snippets: list[str] = []
    if paths:
        try:
            col = get_db()
            for p in paths:
                res = col.get(where={"file": p}, include=["documents", "metadatas"])
                docs = res.get("documents") or []
                metas = res.get("metadatas") or []
                if not docs:
                    continue
                title = (metas[0].get("note") if metas else None) or Path(p).stem
                body = (docs[0] or "").strip().replace("\n", " ")
                snippets.append(f"- {title}: {body[:280]}")
        except Exception:
            snippets = []
    ctx_bits = [f"Pregunta previa: {q}", f"Respuesta: {a}"]
    if snippets:
        ctx_bits.append("Fragmentos de las notas que aparecieron:\n" + "\n".join(snippets))
    ctx = "\n\n".join(ctx_bits)
    prompt = (
        "Sugerí 3 preguntas de seguimiento concretas que el usuario podría "
        "hacer para profundizar usando su vault de Obsidian. Las preguntas "
        "DEBEN anclarse en hechos, nombres, herramientas o conceptos que "
        "aparezcan literalmente en los fragmentos de arriba — no inventes "
        "ángulos no presentes. Cada pregunta ≤70 caracteres, en español "
        "rioplatense (tuteo). Devolvé SOLO un JSON con la forma "
        '{"followups": ["...", "...", "..."]}. Sin texto extra.\n\n'
        f"{ctx}\n\nJSON:"
    )
    try:
        resp = ollama.chat(
            model=resolve_chat_model(),
            messages=[{"role": "user", "content": prompt}],
            options={"temperature": 0.3, "seed": 42, "num_predict": 220, "num_ctx": 2048},
            format="json",
            keep_alive=chat_keep_alive(),
        )
        raw = (resp.message.content or "").strip() or "{}"
        data = json.loads(raw)
        if isinstance(data, dict):
            arr = data.get("followups") or data.get("questions") or \
                  next((v for v in data.values() if isinstance(v, list)), [])
        elif isinstance(data, list):
            arr = data
        else:
            arr = []
        out: list[str] = []
        for s in arr:
            if not isinstance(s, str):
                continue
            s = s.strip().strip('"').strip("'")
            if s and len(s) <= 90:
                out.append(s)
            if len(out) >= 3:
                break
        return {"followups": out}
    except Exception:
        return {"followups": []}


# Related-context (Deezer + YouTube) -----------------------------------------
# Deezer's public search API needs no auth (CORS-restricted but server-side
# is fine). Returned every relevant track has a `link` to deezer.com.


def _deezer_search(query: str, limit: int = 2) -> list[dict]:
    import urllib.request
    import urllib.parse
    qs = urllib.parse.urlencode({"q": query, "limit": limit})
    try:
        with urllib.request.urlopen(f"https://api.deezer.com/search?{qs}", timeout=4) as r:
            data = json.loads(r.read())
    except Exception:
        return []
    items: list[dict] = []
    for tr in (data.get("data") or [])[:limit]:
        artist = (tr.get("artist") or {}).get("name") or ""
        album = (tr.get("album") or {}).get("title") or ""
        items.append({
            "source": "deezer",
            "kind": "track",
            "title": tr.get("title") or "",
            "subtitle": f"{artist} · {album}".strip(" ·"),
            "url": tr.get("link") or "",
        })
    return [i for i in items if i.get("url") and i.get("title")]


def _youtube_search(query: str, limit: int = 2) -> list[dict]:
    key = os.environ.get("YOUTUBE_API_KEY", "").strip()
    if not key:
        return []
    import urllib.request
    import urllib.parse
    qs = urllib.parse.urlencode({
        "part": "snippet", "q": query, "maxResults": limit,
        "type": "video", "key": key, "relevanceLanguage": "es", "safeSearch": "none",
    })
    try:
        with urllib.request.urlopen(f"https://www.googleapis.com/youtube/v3/search?{qs}", timeout=4) as r:
            data = json.loads(r.read())
    except Exception:
        return []
    items: list[dict] = []
    for it in (data.get("items") or [])[:limit]:
        vid = (it.get("id") or {}).get("videoId")
        sn = it.get("snippet") or {}
        if not vid:
            continue
        items.append({
            "source": "youtube",
            "kind": "video",
            "title": sn.get("title") or "",
            "subtitle": sn.get("channelTitle") or "",
            "url": f"https://www.youtube.com/watch?v={vid}",
        })
    return items


class RelatedRequest(BaseModel):
    query: str
    sources: list[str] | None = None  # subset of ["deezer","youtube"]; None = all


@app.post("/api/related")
def related(req: RelatedRequest) -> dict:
    """External enrichment: Deezer + YouTube. Returns merged items list.
    Empty if no API keys present or query is empty. The frontend decides
    when to call this (low-confidence answers, empty retrieval, etc.) so
    the endpoint stays dumb and side-effect-free.
    """
    q = (req.query or "").strip()
    if not q or len(q) < 3:
        return {"items": []}
    wanted = set(req.sources or ["youtube"])
    items: list[dict] = []
    if "deezer" in wanted:
        items.extend(_deezer_search(q, limit=2))
    if "youtube" in wanted:
        items.extend(_youtube_search(q, limit=2))
    return {"items": items}


# ── Helper compartido: excluded folders ──────────────────────────────────────
#
# Los endpoints `/api/notes/{related,contradictions,wikilink-suggestions}`
# aceptan un query param `exclude_folders` (CSV) que filtra ítems cuyo
# `path` arranca con cualquiera de los prefijos. Pensado para el plugin
# Obsidian: el user configura "no me sugieras nada de 04-Archive ni
# 00-Inbox" en settings y cada call pasa esos folders como filtro.
#
# Decisión: filter SERVER-SIDE (no client) porque:
#   - Para contradictions cada call cuesta 5-10s del LLM. Si los items
#     después se descartan client-side, el budget del LLM se desperdicia.
#   - El backend ya conoce VAULT_PATH y los paths normalizados; aplicar
#     aquí es más confiable que reimplementar match en el plugin.
#
# El match es por **prefijo + separador** — "04-Archive" matchea
# "04-Archive/X.md" y "04-Archive/Sub/Y.md" pero NO "04-Archive2/Z.md"
# (que sería un folder vecino sin relación). Por eso normalizamos
# agregando "/" cuando no termina en uno.

def _parse_exclude_folders(s: str | None) -> list[str]:
    """Parsea el query param `exclude_folders` (CSV) en una lista de
    prefijos normalizados con "/" al final."""
    if not s:
        return []
    out: list[str] = []
    for raw in s.split(","):
        folder = raw.strip().strip("/")
        if not folder:
            continue
        out.append(folder + "/")
    return out


def _is_in_excluded_folder(path: str, excluded: list[str]) -> bool:
    """True si `path` arranca con cualquiera de los prefijos `excluded`.

    `excluded` debe venir ya normalizado con `/` al final
    (vía `_parse_exclude_folders`).
    """
    if not excluded:
        return False
    p = path.lstrip("/")
    return any(p.startswith(prefix) for prefix in excluded)


# ── /api/notes/related ────────────────────────────────────────────────────────
# Distinto del `/api/related` de arriba (que es enrichment externo Deezer/
# YouTube). Este endpoint devuelve **notas del vault** relacionadas a una
# nota dada, vía shared-tags + graph hops (find_related). Pensado para el
# panel "Notas relacionadas" del plugin Obsidian — el sidebar lo querya
# cada vez que cambia la nota activa.
#
# No agregamos cache adicional encima de _load_corpus porque ese ya está
# cached con la invalidation correcta (watchdog → reindex bumpa col.id,
# que dispara rebuild). Cold call ~1-2s en vaults grandes; cached <100ms.
@app.get("/api/notes/related")
def notes_related(
    path: str,
    limit: int = 10,
    exclude_folders: str | None = None,
    request: Request = None,  # type: ignore[assignment]
) -> dict:
    """Notas relacionadas a `path` por shared tags + graph hops.

    Wrap delgado de `find_related()`. El ranking real vive en rag/__init__.py
    — este endpoint sólo valida la path, agarra la lista de chunks de la
    nota source, llama a find_related, y formatea el resultado para JSON.

    Args:
        path: Vault-relative (ej. "02-Areas/Coaching/Autoridad.md").
              Debe terminar en .md y no escapar el vault root.
        limit: Cantidad máxima de items (1-50, default 10).

    Returns:
        {"items": [{path, note, folder, tags, shared_tags, score, reason},...],
         "source_path": str}
        Si la nota no está indexada o el vault no existe, items=[] con un
        campo "reason" describiendo el motivo (sin status 4xx — el plugin
        debe poder distinguir "no hay relacionadas" de "no indexada" sin
        try/catch).
    """
    # Audit 2026-04-26 (BUG #33): rate limit per-IP. Pre-fix: plugin
    # buggy o atacante podía floodear (find_related con _load_corpus
    # tarda 1-2s en vault grande).
    if request is not None:
        client_ip = (request.client.host if request.client else "unknown")
        _check_rate_limit(_BEHAVIOR_BUCKETS, client_ip,
                          _BEHAVIOR_RATE_LIMIT, _BEHAVIOR_RATE_WINDOW)
    if not path or not path.endswith(".md"):
        raise HTTPException(status_code=400, detail="path debe terminar en .md")
    if not VAULT_PATH.exists():
        raise HTTPException(
            status_code=503,
            detail=f"vault no encontrado en {VAULT_PATH}",
        )
    # Path-traversal guard: resolver y verificar que cae dentro del vault.
    # Sigue el mismo shape que `rag_read_note` en mcp_server.py.
    try:
        full = (VAULT_PATH / path).resolve()
        full.relative_to(VAULT_PATH.resolve())
    except (ValueError, OSError) as exc:
        raise HTTPException(status_code=400, detail=f"path inválido: {exc}")

    limit = max(1, min(int(limit), 50))

    col = get_db()
    if col.count() == 0:
        return {"items": [], "source_path": path, "reason": "empty_index"}

    # Pasamos TODOS los chunks de la nota source — find_related agrega
    # tags + títulos por toda la lista para construir el set source. Si
    # pasáramos sólo el primer chunk, una nota con N chunks de tags
    # heterogéneos perdería signal.
    c = _load_corpus(col)
    source_metas = [m for m in c["metas"] if m.get("file") == path]
    if not source_metas:
        return {"items": [], "source_path": path, "reason": "not_indexed"}

    # Union de tags de la source — para computar `shared_tags` por
    # vecino y devolverlo al UI (chips clickeables).
    src_tags: set[str] = set()
    for m in source_metas:
        for t in (m.get("tags") or "").split(","):
            t = t.strip()
            if t:
                src_tags.add(t)

    excluded = _parse_exclude_folders(exclude_folders)
    # Pedimos `limit + len(excluded)*N` para cubrir el caso pessimista
    # donde el filtro descarte la mayoría. En la práctica el filter
    # es ligero (prefix match) y limit alcanza casi siempre, así que
    # un cap de 2× del limit es más que suficiente sin desperdiciar.
    fetch_limit = limit if not excluded else min(limit * 2, 100)
    results = find_related(col, source_metas, limit=fetch_limit)
    items: list[dict] = []
    for meta, score, reason in results:
        item_path = meta.get("file", "")
        if _is_in_excluded_folder(item_path, excluded):
            continue
        nei_tags = [
            t.strip() for t in (meta.get("tags") or "").split(",") if t.strip()
        ]
        shared = sorted(src_tags.intersection(nei_tags))
        items.append({
            "path": item_path,
            "note": meta.get("note", ""),
            "folder": meta.get("folder", ""),
            "tags": nei_tags,
            "shared_tags": shared,
            "score": int(score),
            "reason": reason,
        })
        if len(items) >= limit:
            break
    return {"items": items, "source_path": path}


# ── /api/notes/contradictions ────────────────────────────────────────────────
#
# Endpoint pensado para el panel "Posibles contradicciones" del plugin
# Obsidian (Track A del roadmap). Detecta fragmentos de OTRAS notas
# del vault que contradicen afirmaciones de la nota dada.
#
# Shape de la respuesta:
#   {items: [{path, note, folder, snippet, why}], source_path, reason?}
#
# Por qué NO es reactive (active-leaf-change) como `/api/notes/related`:
#   `find_contradictions_for_note` usa el chat LLM (command-r/qwen2.5)
#   para el paso de clasificación "qué es una contradicción genuina
#   vs complementaria". Cold-load del modelo + inferencia = 5-10s
#   por call en M3 Max. Dispararlo en cada cambio de nota quema
#   batería + rompe la UX. El panel es MANUAL (refresh button) con
#   cache agresivo por path.
@app.get("/api/notes/contradictions")
def notes_contradictions(
    path: str, limit: int = 5, exclude_folders: str | None = None,
) -> dict:
    """Posibles contradicciones entre `path` y otras notas del vault.

    Usa `find_contradictions_for_note` del rag.py — wrap delgado para que
    el plugin consuma el mismo shape via HTTP o CLI.

    Args:
        path: Vault-relative (ej. "02-Areas/Coaching/Autoridad.md").
        limit: Máximo de items a devolver (1-10, default 5). El LLM
            devuelve conservador, raramente supera 3.

    Returns:
        items: [{path, note, folder, snippet, why}, ...]
          - `why`: razón del LLM (<20 palabras) de por qué contradice.
          - `snippet`: fragmento del texto que contradice (primeros ~280 chars).
        source_path: echo del input.
        reason?: "empty_index" | "not_indexed" | "too_short" si items=[].

    Performance: 5-10s por call (LLM-bound). El plugin cachea 30min por
    path para que re-entrar a la misma nota sea instant.
    """
    if not path or not path.endswith(".md"):
        raise HTTPException(status_code=400, detail="path debe terminar en .md")
    if not VAULT_PATH.exists():
        raise HTTPException(
            status_code=503,
            detail=f"vault no encontrado en {VAULT_PATH}",
        )
    # Path-traversal guard (mismo shape que /api/notes/related).
    try:
        full = (VAULT_PATH / path).resolve()
        full.relative_to(VAULT_PATH.resolve())
    except (ValueError, OSError) as exc:
        raise HTTPException(status_code=400, detail=f"path inválido: {exc}")
    if not full.is_file():
        return {"items": [], "source_path": path, "reason": "not_indexed"}

    limit = max(1, min(int(limit), 10))
    col = get_db()
    if col.count() == 0:
        return {"items": [], "source_path": path, "reason": "empty_index"}

    # Leer el body. find_contradictions_for_note ya valida longitud
    # mínima internamente pero adelantamos el chequeo para devolver un
    # `reason: "too_short"` explícito en vez de una lista vacía sin
    # contexto — el plugin renderea empty states distintos según el
    # reason, y "muy corta" es accionable para el user ("escribí más
    # o amplíá") vs "no hay contradicciones" (ya evaluado).
    try:
        body = full.read_text(encoding="utf-8", errors="ignore")
    except OSError as exc:
        raise HTTPException(status_code=500, detail=f"read failed: {exc}")
    if len(body.strip()) < 200:
        return {"items": [], "source_path": path, "reason": "too_short"}

    # Excluir la misma nota del set de candidatos. find_contradictions_for_note
    # hace match por chunk embed, así que si no excluimos, los chunks de
    # la nota source van a aparecer como "contradicciones de sí misma".
    excluded = _parse_exclude_folders(exclude_folders)
    # Para contradictions cada call al LLM cuesta ~5-10s. Si los excluded
    # folders descartan muchas, NO pedimos más al LLM (sería re-ejecutar
    # el clasificador entero); aceptamos que el panel devuelva menos
    # items que el `limit` cuando hay filter agresivo.
    results = find_contradictions_for_note(
        col, body, exclude_paths={path}, k=limit,
    )
    items: list[dict] = []
    for r in results:
        p = r.get("path", "")
        if _is_in_excluded_folder(p, excluded):
            continue
        folder = "/".join(p.split("/")[:-1]) if "/" in p else ""
        items.append({
            "path": p,
            "note": r.get("note", ""),
            "folder": folder,
            "snippet": r.get("snippet", ""),
            "why": r.get("why", ""),
        })
    return {"items": items, "source_path": path}


# ── /api/notes/loops ─────────────────────────────────────────────────────────
#
# Endpoint para el panel "Loops abiertos" del plugin Obsidian (Track A
# del roadmap). Devuelve los loops sin cerrar de UNA nota: TODOs en
# frontmatter, checkboxes `- [ ]` sin marcar, y clausulas imperativas
# en el body ("tengo que X", "preguntar Y", etc.).
#
# Diferencia con `find_followup_loops` (rag/__init__.py):
#   - Esa función walks TODO el vault + classifica cada loop con LLM
#     judge (resolved/stale/activo). Pesado (~30s+ en vault grande).
#   - Acá usamos `_extract_followup_loops` sobre 1 archivo solo. No LLM.
#     Cheap (<5ms por nota típica). Apto para reactive trigger del panel.
#   - Trade-off: no clasificamos resolved/stale, solo extraemos. El user
#     ve la lista cruda con age_days y juzga visualmente. Para clasificar
#     existe el endpoint /api/pendientes que hace el walk completo.
@app.get("/api/notes/loops")
def notes_loops(path: str, limit: int = 50) -> dict:
    """Loops abiertos en `path` — frontmatter todos + checkboxes + imperativas.

    Reactive-friendly: O(N) sobre el body de la nota, sin LLM ni embed.

    Args:
        path: Vault-relative (ej. "02-Areas/Coaching/Plan-2026.md").
        limit: Máximo de loops a devolver (1-100, default 50). El plugin
            puede mandar 50 sin riesgo — la mayoría de notas tienen <10.

    Returns:
        items: [{loop_text, kind, age_days, extracted_at}, ...]
          - `kind`: "todo" (frontmatter) | "checkbox" | "inline" (imperative).
          - `age_days`: días desde extracted_at (0 = hoy).
          - El plugin ordena visualmente: stale (age >14d) primero, etc.
        source_path: echo del input.
        reason?: "not_found" si la nota no existe en el vault.

    Performance: <5ms para notas típicas. Sin embed, sin LLM, sin
    sqlite-vec — solo fs + regex + frontmatter parse.
    """
    if not path or not path.endswith(".md"):
        raise HTTPException(status_code=400, detail="path debe terminar en .md")
    if not VAULT_PATH.exists():
        raise HTTPException(
            status_code=503,
            detail=f"vault no encontrado en {VAULT_PATH}",
        )
    try:
        full = (VAULT_PATH / path).resolve()
        full.relative_to(VAULT_PATH.resolve())
    except (ValueError, OSError) as exc:
        raise HTTPException(status_code=400, detail=f"path inválido: {exc}")
    if not full.is_file():
        return {"items": [], "source_path": path, "reason": "not_found"}

    limit = max(1, min(int(limit), 100))
    try:
        raw = full.read_text(encoding="utf-8", errors="ignore")
        st = full.stat()
    except OSError as exc:
        raise HTTPException(status_code=500, detail=f"read failed: {exc}")

    extracted_ts = _note_created_ts(raw, st.st_mtime)
    loops = _extract_followup_loops(raw, path, extracted_ts)
    if not loops:
        return {"items": [], "source_path": path}

    # Anotar age_days y devolver. El now lo capturamos UNA vez para que
    # todos los loops del response sean coherentes entre sí.
    now = datetime.now()
    items: list[dict] = []
    for loop in loops[:limit]:
        # extracted_at es ISO; parseamos defensivo (algunos formatos
        # podrían venir con/sin tz).
        try:
            ex_dt = datetime.fromisoformat(loop.get("extracted_at", ""))
            if ex_dt.tzinfo is not None:
                ex_dt = ex_dt.astimezone().replace(tzinfo=None)
            age_days = max(0, (now - ex_dt).days)
        except Exception:
            age_days = 0
        items.append({
            "loop_text": loop.get("loop_text", ""),
            "kind": loop.get("kind", "inline"),
            "age_days": age_days,
            "extracted_at": loop.get("extracted_at", ""),
        })
    return {"items": items, "source_path": path}


# ── /api/notes/wikilink-suggestions ──────────────────────────────────────────
#
# Endpoint para el panel "Wikilinks sugeridos" del plugin Obsidian
# (Track A #4). Detecta strings en el body de la nota que matchean
# títulos de OTRAS notas del corpus pero NO están linkeadas con
# `[[...]]`. Útil para combatir el patrón "escribí 'autoridad' pero
# olvidé linkear `[[Autoridad]]`" que pasa cuando uno escribe rápido.
#
# Cheap (sin LLM, sin embed). El backend usa un regex multi-pattern
# pre-compilado por título; típicamente <50ms para body de 5KB +
# corpus de 2K títulos. Reactive-friendly.
@app.get("/api/notes/wikilink-suggestions")
def notes_wikilink_suggestions(
    path: str, limit: int = 30, exclude_folders: str | None = None,
) -> dict:
    """Wikilinks sugeridos para `path` — strings en el body que matchean
    títulos de otras notas pero no están linkeadas.

    Wrap delgado de `find_wikilink_suggestions` del rag.py. El ranking
    + filtering (longest-first, ambiguous skip, code-fence skip,
    self-link skip) viven allá; este endpoint solo expone el shape.

    Args:
        path: Vault-relative.
        limit: Max sugerencias a devolver (1-50, default 30). El
            algoritmo internamente respeta el cap, así que pedir más
            no agrega trabajo del lado del server.

    Returns:
        items: [{title, target, line, char_offset, context}, ...]
          - `title`: el texto detectado en la nota (== basename de la
            target sin .md).
          - `target`: path destino (vault-relative).
          - `line`: número de línea 1-indexed.
          - `char_offset`: offset absoluto en bytes desde el inicio
            del archivo. El plugin lo convierte a editor position
            con `editor.offsetToPos()`.
          - `context`: ±60 chars alrededor del match (sin newlines,
            para preview en el card).
        source_path: echo del input.
        reason?: "empty_index" | "not_found" si items=[].
    """
    if not path or not path.endswith(".md"):
        raise HTTPException(status_code=400, detail="path debe terminar en .md")
    if not VAULT_PATH.exists():
        raise HTTPException(
            status_code=503,
            detail=f"vault no encontrado en {VAULT_PATH}",
        )
    try:
        full = (VAULT_PATH / path).resolve()
        full.relative_to(VAULT_PATH.resolve())
    except (ValueError, OSError) as exc:
        raise HTTPException(status_code=400, detail=f"path inválido: {exc}")
    if not full.is_file():
        return {"items": [], "source_path": path, "reason": "not_found"}

    limit = max(1, min(int(limit), 50))
    col = get_db()
    if col.count() == 0:
        return {"items": [], "source_path": path, "reason": "empty_index"}

    excluded = _parse_exclude_folders(exclude_folders)
    # Cuando hay folders excluidos, pedimos un poco más para no quedarnos
    # cortos al filtrar. find_wikilink_suggestions ya tiene un cap interno
    # contra abuso, así que pedir 1.5× es seguro.
    fetch_limit = limit if not excluded else min(int(limit * 1.5), 50)
    suggestions = find_wikilink_suggestions(col, path, max_per_note=fetch_limit)
    if excluded:
        # El item del wikilink suggestion tiene `target` (el path de la
        # nota destino), no `path`. Filtramos sobre ese.
        suggestions = [
            s for s in suggestions
            if not _is_in_excluded_folder(s.get("target", ""), excluded)
        ]
        suggestions = suggestions[:limit]
    return {"items": suggestions, "source_path": path}


class TTSRequest(BaseModel):
    text: str
    voice: str = "Monica"


@app.post("/api/tts")
def tts(req: TTSRequest):
    """Render text with macOS `say` (default voice Mónica) to WAV bytes.

    Local-only: no cloud. WAV chosen over AIFF because browser support is
    universal. Text capped and markdown-stripped so the voice doesn't
    read out backticks and brackets.
    """
    import tempfile  # noqa: PLC0415
    from fastapi.responses import Response  # noqa: PLC0415
    text = (req.text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="texto vacío")
    text = text[:1500]
    clean = re.sub(r"```[\s\S]*?```", " ", text)
    clean = re.sub(r"`[^`]*`", " ", clean)
    clean = re.sub(r"[#*_\[\]<>]", "", clean)
    clean = re.sub(r"\(https?://[^)]+\)", "", clean)
    clean = re.sub(r"\s+", " ", clean).strip()
    if not clean:
        raise HTTPException(status_code=400, detail="texto vacío post-limpieza")
    voice = req.voice if re.match(r"^[A-Za-zÀ-ÿ]{1,32}$", req.voice or "") else "Monica"
    fd = None
    out_path = None
    try:
        fh = tempfile.NamedTemporaryFile(suffix=".wav", delete=False)
        out_path = fh.name
        fh.close()
        subprocess.run(
            ["say", "-v", voice, "--file-format=WAVE",
             "--data-format=LEI16@22050", "-o", out_path, clean],
            check=True, capture_output=True, timeout=45,
        )
        data = Path(out_path).read_bytes()
        return Response(content=data, media_type="audio/wav")
    except subprocess.CalledProcessError as exc:
        detail = (exc.stderr or b"").decode("utf-8", errors="replace")[:200]
        raise HTTPException(status_code=500, detail=f"say falló: {detail}")
    except subprocess.TimeoutExpired:
        raise HTTPException(status_code=504, detail="say timeout")
    finally:
        if out_path:
            try:
                Path(out_path).unlink(missing_ok=True)
            except Exception:
                pass


@app.get("/api/vaults")
def list_vaults() -> dict:
    """Expone el registry + vault activo para el picker de la UI."""
    cfg = _load_vaults_config()
    active = resolve_vault_paths(None)
    active_name = active[0][0] if active else None
    registered = sorted(cfg.get("vaults", {}).keys())
    # Si el activo no está en el registry (p.ej. OBSIDIAN_RAG_VAULT apuntando
    # a uno no registrado), lo incluimos igual para que el picker lo muestre.
    if active_name and active_name not in registered:
        registered = [active_name] + registered
    return {
        "active": active_name,
        "registered": registered,
        "current": cfg.get("current"),
    }


# ── Chat model runtime switch ────────────────────────────────────────────────
# Qwen / Llama families ship a new tag every few weeks; hard-coding the
# default in a launchd plist and restarting the daemon for each A/B is
# enough friction to stop users from experimenting. These two endpoints
# let the UI list available local models and pick one at runtime. No
# server restart, no plist edit — the override is persisted to
# ~/.local/share/obsidian-rag/chat-model.json and survives daemon restarts.

# Families we want to surface as chat-capable. Everything else (embeddings,
# rerankers, vision-only, helpers) is filtered out for the picker. This is
# a denylist by prefix/pattern — easier to maintain as new model families
# appear than whitelisting exact tags.
_CHAT_MODEL_FAMILY_DENYLIST = (
    "bge-",          # embedding models
    "nomic-embed",
    "all-minilm",
    "snowflake-arctic-embed",
    "mxbai-embed",
    "jina-embeddings",
)


class ChatModelRequest(BaseModel):
    model: str | None = None  # None / empty → clear override, revert to defaults


@app.get("/api/chat/model")
def get_chat_model() -> dict:
    """Return the currently-active chat model + the local catalog.

    Shape:
      {
        "current": "qwen2.5:7b",      # what _resolve_web_chat_model() returns
        "override": "qwen2.5:7b",     # only set if the runtime file has one
        "env_override": null,         # OBSIDIAN_RAG_WEB_CHAT_MODEL if set
        "default": "qwen2.5:7b",      # resolve_chat_model() baseline
        "available": ["qwen2.5:7b", "qwen3.6", ...]   # local, chat-capable
      }

    The UI uses this to render the selector with the active option marked.
    """
    override = _read_chat_model_override()
    try:
        available = sorted(
            m.model for m in ollama.list().models
            if not any(m.model.startswith(p) for p in _CHAT_MODEL_FAMILY_DENYLIST)
        )
    except Exception:
        available = []
    try:
        default = resolve_chat_model()
    except Exception:
        default = None
    return {
        "current": _resolve_web_chat_model(),
        "override": override,
        "env_override": WEB_CHAT_MODEL,
        "default": default,
        "available": available,
    }


@app.post("/api/chat/model")
def set_chat_model(req: ChatModelRequest) -> dict:
    """Switch the chat model at runtime. Persisted to disk.

    Empty / null `model` clears the override and reverts to the default
    resolution chain (env var → resolve_chat_model). Validates that the
    requested model is actually installed locally; rejects unknown tags
    with 400 so the UI can surface the error instead of the next /api/chat
    crashing on a missing model.

    Side effect: invalidates the response LRU cache so follow-up queries
    don't replay a cached response produced by a different model.
    """
    requested = (req.model or "").strip() or None
    if requested is not None:
        try:
            available = {m.model for m in ollama.list().models}
        except Exception:
            available = set()
        if requested not in available:
            raise HTTPException(
                status_code=400,
                detail=f"model '{requested}' not installed locally "
                       f"(try: ollama pull {requested.split(':')[0]})",
            )
    try:
        _write_chat_model_override(requested)
    except OSError as exc:
        raise HTTPException(status_code=500, detail=f"could not persist override: {exc}")
    # Invalidate response cache so the next identical query doesn't replay
    # the previous model's answer.
    try:
        _CHAT_CACHE.clear()
    except Exception:
        pass
    return {
        "ok": True,
        "current": _resolve_web_chat_model(),
        "override": _read_chat_model_override(),
    }



@app.get("/api/history")
def query_history(limit: int = 200) -> dict:
    """Recent chat questions for terminal-style up-arrow nav in the web UI.

    Returns oldest→newest after deduping consecutive identical questions.
    Filters to chat-bound commands so /save, /reindex, internal eval runs,
    etc. don't pollute the history list.

    Source of truth post-cutover 2026-04-19 is the SQL `rag_queries` table
    (JSONL writes are gated off by RAG_STATE_SQL=1). Falls back to
    `queries.jsonl` if SQL path is off or fails, so pre-cutover installs
    keep working.
    """
    limit = max(1, min(int(limit or 200), 1000))
    # `web` = web chat endpoint (default cmd in /api/chat log_query_event);
    # `query`/`chat`/`ask` = CLI paths; older rows may have empty cmd.
    keep_cmds = {"query", "chat", "ask", "web"}
    out: list[str] = []

    if RAG_STATE_SQL:
        try:
            # Pull ≥limit rows and dedup, newest-first, then reverse to
            # oldest→newest for the UI. Over-fetch by 4x so consecutive
            # duplicates don't starve the returned window.
            with _ragvec_state_conn() as conn:
                rows = conn.execute(
                    "SELECT q, cmd FROM rag_queries "
                    "ORDER BY id DESC LIMIT ?",
                    (limit * 4,),
                ).fetchall()
            seen_prev: str | None = None
            picked: list[str] = []
            for q, cmd in rows:  # newest → oldest
                cmd = (cmd or "").strip()
                if cmd and cmd not in keep_cmds:
                    continue
                q = (q or "").strip()
                if not q or q == seen_prev:
                    continue
                picked.append(q)
                seen_prev = q
                if len(picked) >= limit:
                    break
            out = list(reversed(picked))
            if out:
                return {"history": out}
            # SQL returned zero rows — fall through to JSONL for older data.
        except Exception as exc:
            _log_sql_state_error("history_sql_read_failed", err=repr(exc))

    if not LOG_PATH.is_file():
        return {"history": []}
    try:
        with LOG_PATH.open("r", encoding="utf-8", errors="replace") as fh:
            for raw in fh:
                raw = raw.strip()
                if not raw:
                    continue
                try:
                    rec = json.loads(raw)
                except json.JSONDecodeError:
                    continue
                cmd = rec.get("cmd") or ""
                if cmd and cmd not in keep_cmds:
                    continue
                q = (rec.get("q") or "").strip()
                if not q:
                    continue
                if out and out[-1] == q:
                    continue
                out.append(q)
    except OSError:
        return {"history": []}
    return {"history": out[-limit:]}


def _resolve_scope(scope: str | None) -> list[tuple[str, "Path"]]:
    if scope is None or scope == "":
        return resolve_vault_paths(None)
    if scope == "all":
        return resolve_vault_paths(["all"])
    return resolve_vault_paths([scope])


# Palabras portuguesas/gallegas que qwen2.5:7b ocasionalmente leakea
# bajo contextos WhatsApp con contactos brasileros (o notas scrapeadas
# de fuentes en portugués). REGLA 0 del prompt ya las prohíbe
# textualmente pero el modelo igual se contagia del lenguaje del
# CONTEXTO. Este filter es la última barrera: reemplaza palabra-por-
# palabra a su equivalente español. Medido 2026-04-23 en scratch_eval:
# 1/20 respuestas tenía "do´mañá" literal (galego) pese al prompt
# endurecido. Conservador: sólo pares alta-confianza donde la palabra
# portuguesa/gallega NO existe (o es muy rara) en español.
_IBERIAN_LEAK_REPLACEMENTS: tuple[tuple[str, str], ...] = (
    # Orden crítico: frases multi-palabra PRIMERO. Si aplicáramos las
    # reglas atomicas antes, "em março" → "em marzo" (palabra "em"
    # quedaría como galego en la respuesta).
    (r"\buma\s+conversa\b", "una conversación"),
    (r"\buma\s+conversação\b", "una conversación"),
    (r"\bem\s+março\b", "en marzo"),
    (r"\bem\s+maio\b", "en mayo"),
    (r"\bem\s+junho\b", "en junio"),
    (r"\bem\s+julho\b", "en julio"),
    (r"\bem\s+setembro\b", "en septiembre"),
    (r"\bem\s+outubro\b", "en octubre"),
    (r"\bem\s+novembro\b", "en noviembre"),
    (r"\bem\s+dezembro\b", "en diciembre"),
    (r"\bem\s+fevereiro\b", "en febrero"),
    (r"\bcontigo\s+em\b", "contigo en"),
    # Meses (único sentido en portugués — en español todos tienen
    # otra grafía).
    (r"\bmarço\b", "marzo"),
    (r"\bmaio\b", "mayo"),
    (r"\bjunho\b", "junio"),
    (r"\bjulho\b", "julio"),
    (r"\bsetembro\b", "septiembre"),
    (r"\boutubro\b", "octubre"),
    (r"\bnovembro\b", "noviembre"),
    (r"\bdezembro\b", "diciembre"),
    (r"\bfevereiro\b", "febrero"),
    # Tiempo (palabras que NO existen en español).
    (r"\bhoje\b", "hoy"),
    (r"\bontem\b", "ayer"),
    (r"\bamanhã\b", "mañana"),
    # Galego: "mañá" + variantes con apóstrofe ascii / unicode prime /
    # backtick. Incluimos formas truncadas que el LLM emite cuando
    # "trata" de españolizar el galego a medias ("do´man", "do´mañ",
    # "do´mana") — captura cualquier "do[apóstrofe]ma[nñ][a|á]?".
    (r"\bdo['´`]ma[nñ][áa]?\b", "mañana"),
    (r"\bmañá\b", "mañana"),
    # Pronombres claros (pt).
    (r"\bnão\b", "no"),
    # "sim" podría ser "sim" de simulación en español técnico — prefix
    # la palabra con word-boundary y usamos case-insensitive para no
    # cazar SIM en siglas.
    (r"\bsim\b", "sí"),
    # Cantidad (pt).
    (r"\bmuito\b", "mucho"),
    (r"\bmuita\b", "mucha"),
    (r"\bmuitos\b", "muchos"),
    (r"\bmuitas\b", "muchas"),
    # Cortesía.
    (r"\bobrigado\b", "gracias"),
    (r"\bobrigada\b", "gracias"),
    # Verbos comunes (pt — conjugaciones que no existen en español).
    (r"\besqueças\b", "olvides"),       # "no te esqueças" → "no te olvides"
    (r"\besqueça\b", "olvide"),
    (r"\bdessas\b", "de esas"),         # "no te esqueças dessas"
    (r"\bdesses\b", "de esos"),
    # 2026-04-28 wave-5: leaks observados en Playwright batch sobre nota
    # Ikigai. El LLM mezcló pt en respuestas que sí debían ser español.
    # "tem" (3ra persona singular del verbo ter en pt) → "tiene". CUIDADO:
    # "tem" puede aparecer en frases como "sistema TEM-1" — usamos word
    # boundary + lookahead que pida espacio + minúscula después para no
    # tocar siglas técnicas.
    (r"\btem\s+(?=[a-záéíóúñ])", "tiene "),
    (r"\bté\b", "té"),                  # idempotent (té ya está en español)
    # Demostrativos/adjetivos (pt) — diferenciable porque las grafías
    # con doble s NO existen en español.
    (r"\besse\b", "ese"),
    (r"\bessa\b", "esa"),
    (r"\besses\b", "esos"),
    (r"\bessas\b", "esas"),
    (r"\bisso\b", "eso"),
    (r"\bisto\b", "esto"),
    (r"\baquilo\b", "aquello"),
    # Adverbios/locuciones pt comunes en respuestas.
    (r"\bAqui\s+estão\b", "Acá están"),
    (r"\baqui\s+está\b", "acá está"),
    (r"\baqui\b", "acá"),               # SOLO al inicio de frase + minúscula
    (r"\bestão\b", "están"),
    (r"\bmelhor\b", "mejor"),
    (r"\bpior\b", "peor"),
    # Verbos de movimiento/uso (pt → es).
    (r"\bajudar\b", "ayudar"),
    (r"\bajuda\b", "ayuda"),
    (r"\bvocê\b", "vos"),
    (r"\bvocês\b", "ustedes"),
    # Conectores/preposiciones (pt — claramente no español).
    (r"\bcom\s+(?=[a-záéíóúñ])", "con "),
    (r"\bde\s+(?=[a-záéíóúñ]\w+ção\b)", "de "),  # idempotente — solo prep
    # Sufijos -ção / -ções son siempre pt; convertir a -ción/-ciones en
    # contexto. Caso específico observado.
    (r"\bação\b", "acción"),
    (r"\bações\b", "acciones"),
    (r"\bsolução\b", "solución"),
    (r"\bsoluções\b", "soluciones"),
    (r"\bquestão\b", "cuestión"),
    (r"\bquestões\b", "cuestiones"),
    # Francés/italiano stray words (hallucination genuina, no contagio).
    # "Voulu" observado al inicio de respuesta — claramente falso.
    (r"\bVoulu,?\s+", ""),
    (r"\bvoilà\b", "acá"),
)
_IBERIAN_LEAK_COMPILED: tuple[tuple[re.Pattern, str], ...] = tuple(
    (re.compile(pat, re.IGNORECASE), repl)
    for pat, repl in _IBERIAN_LEAK_REPLACEMENTS
)

# Palabras que INICIAN una frase multi-palabra del dict anterior. Cuando
# el streaming filter ve un candidate que TERMINA con una de estas más
# whitespace opcional, retiene la palabra en el buffer porque la
# próxima llegada podría completar el compound (`em ` + `março` →
# `em março` → `en marzo`). Mantener en sync con los compounds de
# `_IBERIAN_LEAK_REPLACEMENTS` que son multi-palabra.
_COMPOUND_STARTER_TAIL_RE = re.compile(
    r"\b(uma|em|contigo)(\s+\S*)?\s*$",
    re.IGNORECASE,
)


def _replace_iberian_leaks(text: str) -> str:
    """Apply the _IBERIAN_LEAK_REPLACEMENTS regexes in order. Safe on
    non-string / empty input. Preserves case for common cases via
    `IGNORECASE` on the regex side — but the replacement is lowercase,
    so mixed-case originals ("Março" → "marzo") normalise to lowercase.
    That's acceptable: the leak itself is a model quirk, not a stylistic
    choice we want to preserve.
    """
    if not text:
        return text
    out = text
    for pat, repl in _IBERIAN_LEAK_COMPILED:
        out = pat.sub(repl, out)
    return out


# 2026-04-28 wave-4: el LLM (qwen2.5:7b) tipea mal los nombres de archivos
# canónicos del proyecto. Repro Playwright: pregunta "leé CLAUDE.md y
# resumime", response empieza con "Basado en el contenido de `CLAIDE.md`..."
# (transposición I↔U en el tokenizer). Normalizamos el filename a la forma
# canónica via regex case-insensitive — corrige el typo sin afectar prosa.
# Todos los archivos canónicos viven en project root o docs/ (ver
# `_agent_tool_read_note` whitelist).
_CANONICAL_FILENAME_TYPOS: tuple[tuple[re.Pattern, str], ...] = (
    # CLAUDE typos: variantes de 5 letras con C+L+ (A|O) + (U|I) + (D|E)
    # incluyendo CLAIDE, CLODE, CLAUE, CLAUD, CLODA, CLOUDE, etc.
    (re.compile(r"\bCL[AO][UI]?D?E?\.md\b", re.IGNORECASE), "CLAUDE.md"),
    # AGENTS typos: AGUNTS, AGNTS, AGENT, AGNETS, AGENS, etc.
    (re.compile(r"\bAG[EU]?N?[ET]?[ST]?S?\.md\b", re.IGNORECASE), "AGENTS.md"),
    # README typos: READEM, REAME, READMI, REDME, etc.
    (re.compile(r"\bREA?D?[EM]?[EM]?[EI]?\.md\b", re.IGNORECASE), "README.md"),
)


def _normalize_canonical_filenames(text: str) -> str:
    """Corrige typos del LLM sobre nombres canónicos de archivos del proyecto.
    Idempotente sobre texto que ya tiene el filename correcto.
    """
    if not text or ".md" not in text.lower():
        return text
    out = text
    for pat, canonical in _CANONICAL_FILENAME_TYPOS:
        out = pat.sub(canonical, out)
    return out


# 2026-04-28 wave-7 (eval Conv 5 — CRITICAL privacy leak): el vault del
# user puede tener notas de contacto con credenciales en plain text
# (passwords, DNI, números de teléfono). Si el user pregunta "quién es
# mi mamá", el LLM lee la nota y emite TODO incluido el password. Esto
# es un security/privacy issue grave.
#
# Fix: post-process filter en el streaming pipeline que detecta patrones
# claros de credenciales y los redacta. Conservador — sólo matchea
# labels explícitos seguidos del valor (ej. "Contraseña: foo123"). NO
# redacta números/strings sueltos sin contexto. False positives son
# baratos (un número que no era PII queda redactado, el user pregunta
# de nuevo). False negatives serían el bug.
#
# Greppable: cuando se redacta algo, el filter incrementa
# `_PII_REDACT_COUNT` (atómico-ish, log periódico).
_PII_REDACT_PATTERNS: tuple[tuple[re.Pattern, str], ...] = (
    # 2026-04-28 wave-7 iter2: el LLM no siempre emite "Label: valor" — a
    # veces dice "Su contraseña es valor", "contraseña guardada como valor",
    # "el password de X es valor". Patterns label-PROXIMITY (label + hasta
    # ~30 chars de prosa + valor) cubren los modos de falla observados en
    # eval Conv 5.

    # Passwords / contraseñas — label + connector ("es"/"como"/"="/":") +
    # valor mixto alphanum 4+ chars. Greedy hasta whitespace/punctuation.
    # NOTA: catch "Contraseña: X", "su contraseña es X", "contraseña
    # guardada como X", "password = X", "su contraseña en tus notas como X".
    # Allow 0-6 palabras intermedias entre label y connector (cubre frases
    # tipo "guardada en tus notas como").
    (
        re.compile(
            r"(?i)\b(?:contrase[ñn]a|password|pwd|passphrase)"
            r"(?:\s+\w+){0,6}?"  # 0-6 palabras intermedias
            r"\s*(?:[:=]|\bes\b|\bcomo\b)\s*"  # connector
            r"\S{4,}",  # valor (4+ chars no-whitespace)
        ),
        "[contraseña REDACTADA]",
    ),
    # DNI (Argentina) — formato 7-8 dígitos con/sin puntos.
    # Pattern 1: label + connector + valor.
    (
        re.compile(
            r"(?i)\bDNI(?:\s+\w+){0,3}?\s*(?:[:=]|\bes\b)\s*[\d.]{7,12}",
        ),
        "DNI: [REDACTADO]",
    ),
    # Pattern 2: 7-8 dígitos seguidos con puntos típicos de DNI (xx.xxx.xxx).
    # Conservador: requiere el formato exacto AR.
    (
        re.compile(
            r"\b\d{2}\.\d{3}\.\d{3}\b",
        ),
        "[DNI REDACTADO]",
    ),
    # Teléfono — label + valor. Acepta connector "es"/"como"/":".
    (
        re.compile(
            r"(?i)\b(?:tel(?:[eé]fono)?|cel(?:ular)?|m[oó]vil|whatsapp|n[uú]mero(?:\s+de)?\s+(?:tel(?:[eé]fono)?|cel(?:ular)?))"
            r"(?:\s+\w+){0,3}?"
            r"\s*(?:[:=]|\bes\b|\bcomo\b)\s*"
            r"\+?\d[\d\s\-()]{7,20}",
        ),
        "[teléfono REDACTADO]",
    ),
    # Pattern 3: número AR completo (+54 9 ... ) sin label. Conservador:
    # requiere prefijo +54 + 9 dígitos+.
    (
        re.compile(
            r"\+54\s*9\s*\d{2,4}[\s-]?\d{3,4}[\s-]?\d{3,4}\b",
        ),
        "[teléfono REDACTADO]",
    ),
    # CUIT/CUIL — 11 dígitos, formato XX-XXXXXXXX-X o sin guiones.
    (
        re.compile(
            r"(?i)\b(?:cuit|cuil)(?:\s+\w+){0,3}?\s*(?:[:=]|\bes\b)?\s*\d{2}[\s-]?\d{8}[\s-]?\d{1}",
        ),
        "[CUIT REDACTADO]",
    ),
    (
        re.compile(
            r"\b\d{2}-\d{8}-\d\b",  # standalone CUIT format
        ),
        "[CUIT REDACTADO]",
    ),
    # Tokens / API keys — label + connector + valor.
    (
        re.compile(
            r"(?i)\b(?:token|api[_\s-]?key|secret|access[_\s-]?key)"
            r"(?:\s+\w+){0,3}?\s*(?:[:=]|\bes\b)\s*\S{4,}",
        ),
        "[token REDACTADO]",
    ),
    # CBU / CVU / IBAN.
    (
        re.compile(
            r"(?i)\b(?:cbu|cvu|iban)(?:\s+\w+){0,3}?\s*(?:[:=]|\bes\b)?\s*[\d-]{16,32}",
        ),
        "[CBU REDACTADO]",
    ),
    # Tarjeta de crédito 16 dígitos.
    (
        re.compile(
            r"\b(?:\d{4}[-\s]?){3}\d{4}\b",
        ),
        "[tarjeta REDACTADA]",
    ),
    # CVV/CVC.
    (
        re.compile(
            r"(?i)\b(?:cvv|cvc|c[oó]digo\s+seguridad)(?:\s+\w+){0,3}?\s*(?:[:=]|\bes\b)?\s*\d{3,4}",
        ),
        "[CVV REDACTADO]",
    ),
)


# 2026-04-28 wave-7 (eval Conv 6 — CRITICAL internal error leak): assertion
# messages internas (con references a CLAUDE.md, ThreadPoolExecutor, M3 Max,
# stack frames) llegaban directo al user via SSE error event. Repro:
# "cómo funciona el sistema RAG" → "retrieve falló: bm25_search llamado en
# paralelo — es GIL-serialised por diseño (CLAUDE.md línea 126...)".
#
# Fix: detectar señales claras de "esto es un dev error, no un user error" y
# reemplazar con mensaje user-friendly. El error ORIGINAL queda loggeado en
# web.log con tag `[chat-error-sanitized]` para debug.
_DEV_ERROR_SIGNALS = (
    "claude.md", "agents.md", "thread", "executor", "gil-serialised",
    "gil-serialized", "asserterror", "assertionerror", "traceback",
    "  file ", "stack", "m3 max", "callable", "noneType",
    "in __init__", "lambda", "wrapped function",
)


def _sanitize_error_for_user(exc: Exception, *, phase: str = "unknown") -> str:
    """Convertir un Exception en un mensaje user-friendly. Si el error parece
    interno/dev-facing (matchea _DEV_ERROR_SIGNALS), reemplazar por un
    mensaje genérico. Sino, devolver el mensaje del exception (asumido
    user-safe — ej. "ollama no responde", "vault no encontrado").
    """
    raw = str(exc) if exc else ""
    raw_lower = raw.lower()
    for signal in _DEV_ERROR_SIGNALS:
        if signal in raw_lower:
            # Mensaje user-friendly por phase. Si phase desconocido, fallback.
            user_msgs = {
                "retrieve": "No pude buscar en el vault — probá reformular o intentar de nuevo en un momento.",
                "synthesis": "Tuve un problema generando la respuesta — probá de nuevo.",
                "tool_decision": "Hubo un error decidiendo qué hacer — reformulá la pregunta o probá de nuevo.",
                "tasks_brief": "Falló el resumen de pendientes — probá de nuevo.",
            }
            return user_msgs.get(phase, "Hubo un error procesando tu consulta — probá de nuevo o reformulá.")
    # No matchea — el mensaje del exception es probablemente safe (ej.
    # "vault X no encontrado", "ollama timeout"). Devolverlo prefixado
    # con el phase para context.
    if not raw:
        return f"Error desconocido en {phase}."
    # Sanitize: nunca exponemos tracebacks o paths absolutos. Cortar al
    # primer newline + cap a 200 chars.
    safe = raw.split("\n")[0][:200]
    return f"{phase} falló: {safe}"


def _redact_pii(text: str) -> tuple[str, int]:
    """Redact PII labels en texto. Retorna (texto_redactado, count_redacciones).

    Patrones soportados (ver `_PII_REDACT_PATTERNS`):
      - Contraseña/password/clave: cualquier valor tras label
      - DNI: dígitos 7-12 con/sin puntos
      - Teléfono/celular/móvil/whatsapp con label
      - Token/API key/secret/access key con label
      - CBU/CVU/IBAN con label
      - Números de tarjeta (16 dígitos)
      - CVV/CVC con label

    Idempotente: re-aplicar sobre texto ya redactado no rompe nada.
    """
    if not text:
        return text, 0
    out = text
    n = 0
    for pat, repl in _PII_REDACT_PATTERNS:
        if callable(repl):
            new_out, k = pat.subn(repl, out)
        else:
            new_out, k = pat.subn(repl, out)
        if k:
            n += k
            out = new_out
    return out, n


# Streaming pipeline: PII labels may arrive split across chunks
# (ej. chunk1="Contraseña: " + chunk2="secret123"). Si emitimos chunk1
# antes de ver chunk2, el regex no matchea (label sin valor adyacente).
# Este regex detecta cuando el candidate de emit TERMINA con un PII
# label parcial que necesita el próximo chunk para redactar bien.
_PII_TAIL_HOLDBACK_RE = re.compile(
    r"(?i)\b(?:contrase[ñn]a|clave|password|pwd|pass|"
    r"dni|tel(?:[eé]fono)?|cel(?:ular)?|m[oó]vil|whatsapp|"
    r"token|api[_\s-]?key|secret|access[_\s-]?key|"
    r"cbu|cvu|iban|cvv|cvc|c[oó]digo\s+seguridad)"
    r"\s*[:=]?\s*\S{0,40}\s*$",  # label opcionalmente seguida de inicio del valor
    re.IGNORECASE,
)


class _PiiRedactFilter:
    """Streaming filter al final del pipeline que redacta credenciales/PII
    cruzando chunk boundaries. Mantiene un buffer si el candidate termina
    con un label de PII conocido — el valor llegará en el próximo chunk.

    Usado DESPUÉS del IberianLeakFilter (que ya hace redact inline para
    mismo-chunk). Este filter cubre los casos de chunks pequeños donde
    label y valor llegan separados.
    """

    _MAX_HOLD = 200  # cap de emergencia

    def __init__(self) -> None:
        self._buf = ""
        self._redact_count = 0

    def feed(self, chunk: str) -> str:
        if not chunk:
            return ""
        self._buf += chunk
        # Si el buffer entero termina con un PII label (con o sin valor
        # parcial), retenemos hasta el próximo chunk.
        m = _PII_TAIL_HOLDBACK_RE.search(self._buf)
        if m and m.end() == len(self._buf):
            # Tail es PII label en construcción — emergency flush si demasiado largo
            if len(self._buf) > self._MAX_HOLD:
                out, n = _redact_pii(self._buf)
                self._redact_count += n
                self._buf = ""
                return out
            return ""
        # No hay label pendiente — aplicar redact y emitir todo.
        out, n = _redact_pii(self._buf)
        self._redact_count += n
        self._buf = ""
        return out

    def flush(self) -> str:
        if not self._buf:
            return ""
        out, n = _redact_pii(self._buf)
        self._redact_count += n
        self._buf = ""
        return out


class _IberianLeakFilter:
    """Streaming filter chained después de `_InlineCitationStripper` que
    reemplaza leaks portugueses/gallegos con su equivalente español.

    Problema de diseño: las frases multi-palabra ("em março", "contigo
    em", "uma conversa") llegan partidas entre chunks del stream (peor
    caso: ollama chunk_size=1). Si emitimos cada palabra al llegar a un
    boundary, la regex compuesta nunca matchea porque los fragmentos
    ya fueron emitidos por separado.

    Solución: **retener compound starters**. Cuando el candidate de
    emit TERMINA con "em ", "uma " o "contigo " (las 3 palabras que
    inician compounds en `_IBERIAN_LEAK_REPLACEMENTS`), retenemos ese
    starter en el buffer esperando la siguiente palabra. Cuando llega,
    el buffer tiene "em março" completa y la regla dispara. Ver
    `_COMPOUND_STARTER_TAIL_RE` — mantenerlo sincronizado con los
    compounds multi-palabra al agregar frases nuevas.

    API:
      - `.feed(chunk)` acumula; emite hasta el último boundary menos
        cualquier compound starter pendiente al final.
      - `.flush()` drena el buffer aplicando replace.
      - Idempotente: texto ya en español pasa sin modificar.
    """

    _MAX_HOLD = 200  # cap de emergencia contra tokens gigantes sin espacios.
    _BOUNDARY_CHARS = " \t\n.,!?;:()[]{}\"'·"

    def __init__(self) -> None:
        self._buf = ""

    def feed(self, chunk: str) -> str:
        if not chunk:
            return ""
        self._buf += chunk
        # Último boundary (espacio / puntuación) en el buffer.
        last_boundary = -1
        for ch in self._BOUNDARY_CHARS:
            idx = self._buf.rfind(ch)
            if idx > last_boundary:
                last_boundary = idx
        if last_boundary == -1:
            # Nada que emitir todavía. Flush de emergencia si el buffer
            # explotó (un token gigante sin espacios — caso raro).
            if len(self._buf) > self._MAX_HOLD:
                out = _replace_iberian_leaks(self._buf)
                self._buf = ""
                return out
            return ""
        candidate = self._buf[:last_boundary + 1]
        # Clave para el streaming de compounds: si el candidate TERMINA
        # con un "starter" de frase compuesta ("em ", "uma ", "contigo
        # "), lo retenemos en el buffer porque podría completar una
        # frase cuando lleguen más chars. Sin este safeguard, con
        # chunk_size=1 el candidate "hola em " emitiría "em" antes de
        # ver "março" y la regla `em\s+março` nunca dispararía.
        m = _COMPOUND_STARTER_TAIL_RE.search(candidate)
        if m:
            starter_start = m.start()
            if starter_start == 0:
                # El candidate es sólo el starter + espacios — no hay
                # nada que emitir por ahora. Esperamos más chunks.
                if len(self._buf) > self._MAX_HOLD:
                    out = _normalize_canonical_filenames(_replace_iberian_leaks(self._buf))
                    self._buf = ""
                    return out
                return ""
            # Retener desde el starter; emitir todo lo anterior.
            to_emit = self._buf[:starter_start]
            self._buf = self._buf[starter_start:]
            return _redact_pii(_normalize_canonical_filenames(_replace_iberian_leaks(to_emit)))[0]
        to_emit = candidate
        self._buf = self._buf[last_boundary + 1:]
        return _redact_pii(_normalize_canonical_filenames(_replace_iberian_leaks(to_emit)))[0]

    def flush(self) -> str:
        tail = self._buf
        self._buf = ""
        return _redact_pii(_normalize_canonical_filenames(_replace_iberian_leaks(tail)))[0]


# 2026-04-28 wave-5: el LLM (qwen2.5:7b) ocasionalmente emite la SINTAXIS
# de un tool call como TEXTO en vez de invocar el tool real via el
# protocol. Repro Playwright:
#   - "agendá mi cumpleaños el 12 de febrero" → response:
#     `propose_calendar_event(title="cumpleaños", date="el 12 de febrero")`
#   - "recordame X" → response: `propose_reminder(title='X', when='')`
#   - "clima en Mendoza para mañana" → "工具调用：propose_weather_forecast(...)"
# El tool nunca se ejecuta, el user ve el call como prosa. Es un model bug
# que ningún prompt arregla 100%. Fix: stripper que detecta el patrón
# `<tool_name>(args...)` en los primeros chars del stream y lo reemplaza
# por una clarificación.
_RAW_TOOL_CALL_RE = re.compile(
    r"^\s*"
    r"(?:工具调用[：:]\s*)?"  # opcional prefijo en chino
    r"(?P<name>(?:propose_|search_|read_|gmail_|whatsapp_|calendar_|reminders_|weather|drive_|finance_|credit_|record_)\w*)"
    r"\s*\([^)]{0,300}\)?",
    re.IGNORECASE,
)


def _detect_raw_tool_call(text: str) -> str | None:
    """Returns the matched tool name if `text` starts with a raw tool-call
    syntax. None otherwise. Used to short-circuit synthesis output that
    leaked tool-call markup instead of executing it.
    """
    if not text:
        return None
    m = _RAW_TOOL_CALL_RE.match(text)
    if not m:
        return None
    return (m.group("name") or "").lower()


class _RawToolCallStripper:
    """Streaming filter que detecta raw tool-call syntax al inicio del
    stream y la reemplaza con una clarificación amigable. Buffer first
    ~120 chars antes de decidir; si match → swallow + emit clarification;
    si no → flush + pass-through del resto.
    """

    _BUFFER_CAP = 120

    def __init__(self) -> None:
        self._buf = ""
        self._decided = False
        self._was_raw_tool_call = False

    def feed(self, chunk: str) -> str:
        if not chunk:
            return ""
        if self._decided:
            # Si detectamos raw tool call al inicio, swallow TODO el resto del
            # stream — el LLM va a seguir emitiendo el resto del call (params,
            # cierre de paréntesis, etc.) que NO queremos mostrar al user. Ya
            # emitimos la clarificación, no agregamos nada más.
            if self._was_raw_tool_call:
                return ""
            return chunk
        self._buf += chunk
        # Decisión rápida si ya tenemos suficientes chars.
        if len(self._buf) >= self._BUFFER_CAP:
            return self._decide_and_flush()
        return ""

    def _decide_and_flush(self) -> str:
        self._decided = True
        m = _RAW_TOOL_CALL_RE.match(self._buf)
        if m:
            self._was_raw_tool_call = True
            tool = (m.group("name") or "tool").lower()
            self._buf = ""  # discard buffered raw tool call syntax
            # Clarification messages tailored al tool detectado.
            clarif = {
                "propose_reminder": "Necesito un poco más para crearte el recordatorio. Decime qué querés que te recuerde y cuándo (ej: 'recordame llamar al banco el lunes a las 9am').",
                "propose_calendar_event": "Para agendar el evento necesito el detalle. Decime qué evento y cuándo (ej: 'agendá reunión con Pedro el miércoles a las 4pm').",
                "propose_whatsapp_send": "Para mandar el WhatsApp decime a quién y qué (ej: 'mensajeale a María: nos vemos a las 8').",
                "weather": "Para el clima decime qué ciudad (ej: 'clima en Buenos Aires' o 'pronóstico Madrid mañana').",
            }.get(tool, f"Necesito un poco más de detalle para usar la tool `{tool}`. Probá reformular con más contexto.")
            return clarif
        # No es raw tool call — flush normal.
        out = self._buf
        self._buf = ""
        return out

    def flush(self) -> str:
        if self._decided:
            if self._was_raw_tool_call:
                return ""
            tail = self._buf
            self._buf = ""
            return tail
        return self._decide_and_flush()


class _InlineCitationStripper:
    """Streaming filter that elides `[Title](path.md)` from LLM output.

    Command-r ignores "don't cite inline" instructions about ~15-20% of
    the time even when the prompt is unambiguous. Since the UI renders a
    dedicated sources block below every answer, inline citations show
    the same note twice. This filter is a safety net: it buffers any
    text starting at an open bracket until we know whether the bracket
    opens a `[…](…md)` link or something else. If it does, we drop the
    whole pattern; if it doesn't, we flush the buffered text unchanged.

    Kept mid-module (not inside the chat handler) so tests can import it
    in isolation.
    """

    _MAX_TITLE = 200   # guard against runaway buffers on malformed output
    _MAX_TARGET = 400

    def __init__(self) -> None:
        self._buf = ""

    def feed(self, chunk: str) -> str:
        self._buf += chunk
        out: list[str] = []
        i = 0
        n = len(self._buf)
        while i < n:
            c = self._buf[i]
            if c != "[":
                out.append(c)
                i += 1
                continue
            # Try to resolve a `[title](target)` pattern starting here.
            close_br = self._buf.find("]", i + 1)
            if close_br == -1:
                if n - i > self._MAX_TITLE:
                    # Not a link — too long to be a title. Flush.
                    out.append(self._buf[i])
                    i += 1
                    continue
                break  # incomplete; wait for more.
            if close_br + 1 >= n:
                break  # don't know yet what follows `]`.
            if self._buf[close_br + 1] != "(":
                # Plain `[…]` text, not a link. Emit and continue.
                out.append(self._buf[i:close_br + 1])
                i = close_br + 1
                continue
            close_pa = self._buf.find(")", close_br + 2)
            if close_pa == -1:
                if n - (close_br + 2) > self._MAX_TARGET:
                    out.append(self._buf[i:close_br + 2])
                    i = close_br + 2
                    continue
                break  # incomplete link tail; wait for more.
            target = self._buf[close_br + 2:close_pa]
            if target.endswith(".md"):
                # Drop the entire `[title](path.md)`.
                i = close_pa + 1
                # Optional: also elide a trailing ", " or " (" that was
                # a separator between consecutive inline citations.
                while i < n and self._buf[i] in (" ", ","):
                    i += 1
            else:
                # External URL or non-note target — keep as-is.
                out.append(self._buf[i:close_pa + 1])
                i = close_pa + 1
        self._buf = self._buf[i:]
        return "".join(out)

    def flush(self) -> str:
        tail = self._buf
        self._buf = ""
        return tail


def _sse(event: str, data: dict) -> str:
    return f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"


def _maybe_emit_proposal(name: str, out: str) -> str | None:
    """If `name` is a proposal tool and `out` parses as the expected JSON
    payload, return a pre-encoded SSE frame; else None.

    Routes based on payload kind:
      - `needs_clarification: true` → `proposal` event (UI shows card,
        user confirms/edits).
      - `created: true` → `created` event (UI shows toast with Deshacer).
      - `created: false` → `proposal` event too, but with `error` field
        set so the UI can surface the failure in the card.

    Keeps the gen() body clean — callers just
    `if frame := _maybe_emit_proposal(...): yield frame`. A malformed
    payload is treated as "not a proposal" (the LLM still sees the tool
    output normally via the `role:"tool"` message; the UI just won't
    render anything). No logging here — caller decides what to do.
    """
    if name not in PROPOSAL_TOOL_NAMES:
        return None
    try:
        payload = json.loads(out)
    except Exception:
        return None
    if not isinstance(payload, dict):
        return None
    # Auto-created happy path → toast + undo.
    if payload.get("created") is True:
        return _sse("created", payload)
    # Ambiguous / needs-clarification path → confirmation card.
    if payload.get("needs_clarification") or payload.get("proposal_id"):
        return _sse("proposal", payload)
    # created=false with error set → show proposal card with error state
    # so the user can retry. Emit as proposal event (UI renders card).
    if payload.get("created") is False and payload.get("error"):
        # Inject a synthetic proposal_id so the UI handler treats it as
        # a normal proposal. Keep the error field for visibility.
        payload.setdefault("proposal_id", f"prop-err-{payload.get('kind', 'x')}")
        payload["needs_clarification"] = True
        return _sse("proposal", payload)
    return None


def _source_payload(meta: dict, score: float) -> dict:
    return {
        "file": meta.get("file", ""),
        "note": meta.get("note", ""),
        "folder": meta.get("folder", ""),
        "score": round(float(score), 3),
    }


# Thin wrappers al par canónico en rag.py — mantienen la API estable acá
# sin duplicar la calibración (2026-04-21: rag.py expone thresholds
# calibrados contra la distribución real de `rag_queries.top_score`).
from rag import confidence_badge as _rag_confidence_badge  # noqa: E402
from rag import score_bar as _rag_score_bar  # noqa: E402


def _confidence_badge(score: float) -> tuple[str, str]:
    return _rag_confidence_badge(score)


def _score_bar(score: float, width: int = 5) -> str:
    return _rag_score_bar(score, width=width)


# ── Intent detection: tasks / agenda / pendientes ────────────────────────────
# Short-circuits the normal RAG retrieve when the question is about what the
# user has to do. RAG over vault notes produces hallucinations ("pedido de
# speeds") because the vault doesn't store live calendar/reminders; the
# services layer does.
#
# Tasks-mode is scope-aware: "mañana" queries fetch tomorrow's reminders +
# calendar, "esta semana" fetches 7 days, default fetches 2 weeks. Without
# this, a "qué tengo para mañana" query got today's reminders and the LLM
# correctly concluded "nada para mañana" — the fix is to bring the right
# window, not to patch the LLM.

_TASKS_INTENT_TOKENS = re.compile(
    r"\b(pendientes?|agenda|recordatorios?|to-?dos?|tasks?|reminders?|"
    r"tareas?|reuniones?|meetings?|prioridad(?:es)?|prioriz|"
    r"compromisos?|obligaciones?|deberes?)",
    re.IGNORECASE,
)
_TASKS_INTENT_PHRASES = re.compile(
    # "qué tengo/debo/hay/etc" REQUIERE qualifier de tarea/tiempo después
    # — la versión laxa anterior matcheaba "qué hay sobre X" / "qué tengo
    # en notas sobre Y" y mis-routeaba topic queries al fetcher de agenda.
    r"(qu[eé]\s+(tengo|debo|me\s+falta|me\s+queda|hay)\s+"
    r"(para|que\s+(hacer|recordar|completar|priorizar)|"
    r"hacer|recordar|completar|"
    r"hoy|ma[nñ]ana|esta\s+semana|pendiente|priorizar|gestionar|organizar)|"
    r"what\s+do\s+i\s+(have|need)\s+(to\s+)?(do|complete|finish)|"
    r"(tengo|hay)\s+(algo|cosas?|eventos?|citas?)\s+(pendiente|hoy|ma[nñ]ana|esta\s+semana)|"
    r"organiz(ar|ame)\s+(el\s+d[ií]a|la\s+semana))",
    re.IGNORECASE,
)


def _is_tasks_query(q: str) -> bool:
    if not q:
        return False
    return bool(_TASKS_INTENT_TOKENS.search(q) or _TASKS_INTENT_PHRASES.search(q))


# Anaphora cues that mean "this question depends on the prior turn" — pronouns,
# demonstratives, ellipsis markers. If a follow-up has any of these, route it
# through reformulate_query so retrieval gets the resolved entity. Otherwise
# the original wording is already a self-contained search target.
#
# Deliberately excluded: `el`/`la`/`lo` — they're common articles and would
# fire on virtually every Spanish sentence, causing the reformat call on fully
# self-contained questions (empirically regressed quality by ~0.15 rerank).
_FOLLOWUP_CUES = re.compile(
    r"\b(eso|esa|ese|esto|esta|este|aquello|ella|[eé]l\b(?!\s+[A-ZÁÉÍÓÚ])|"
    r"ahí|ahi|allí|alli|allá|alla|"
    r"it\b|this\b|that\b|he\b|she\b|"
    r"y\s+(de|sobre|con|para|en)|"
    r"profundizá|profundiza|ampliá|amplia|seguí|segui|continuá|continua|"
    r"más\s+(sobre|de|al\s+respecto)|mas\s+(sobre|de|al\s+respecto)|"
    # 2026-04-28 wave-6: sync con _TOPIC_SHIFT_FOLLOWUP_RE en rag/__init__.py.
    # Antes: detect_topic_shift veía "qué otros materiales..." como anaphoric
    # (mantenía history) PERO _looks_like_followup retornaba False, así que el
    # search_question quedaba raw → retrieve buscaba "materiales" sueltos sin
    # anchor del turno anterior. Sincronizar acá hace que CONCAT(prev, current)
    # sea el search-question, recuperando el contexto.
    r"qu[eé]\s+otros?|qu[eé]\s+otras?|"
    r"cu[aá]l\s+(?:era|fue|es|ser[ií]a|ser[ií]an)\s+(?:el|la|los|las)?|"
    r"dame\s+(?:un|el|los|otro)|un\s+ejemplo|otro\s+ejemplo|"
    r"explic[aá]me\s+(?:mejor|bien|m[aá]s|otra\s+vez)|"
    r"resum[ií]me|tradu[cz]i?[ií]?(?:me|lo|la)?|"
    r"y\s+despu[eé]s|y\s+entonces|y\s+ahora|"
    r"recomend[aá]ri[aá]s|sugerir[ií]as)\b",
    re.IGNORECASE,
)

# Extended follow-up patterns added 2026-04-26 to close the recall gap on
# elliptical follow-ups that don't have explicit pronoun cues. Three buckets:
#
# (a) Leading "y + article/wh-word": "y el icbc", "y la retrospectiva del
#     barco", "y qué pendientes tengo del sistema RAG", "y cómo las indexás",
#     "y las versiones de la herrumbre". The base regex already catches
#     "y + de/sobre/con/para/en" but missed these continuations because
#     "el/la/los/las/qué/cómo" are way too common as leading articles in
#     standalone questions ("la nota que dice X"). We anchor with `^` so
#     ONLY question-initial "y + token" fires — confirmed safe on the
#     queries.yaml golden set (no false positives in standalone questions).
#
# (b) Leading verb + definite article + noun: imperative shorthand like
#     "listame los gastos", "mostrame las cuentas", "detallame los pagos".
#     The user pattern is "I just asked about X, now elaborate without
#     re-naming X". Reproduced 2026-04-26 from session web:cee69e81829c:
#     after "Cuanto devo a la visa?" the user asked "listame los gastos
#     en pesos" expecting Visa context — got MOZE chunks because
#     `_looks_like_followup` returned False and the retriever searched
#     the raw question.
#
# (c) "tengo algún" / "qué <noun> tengo|usé|tenía": cross-source follow-ups
#     from queries.yaml `rag-pendientes-cross` / `rag-workshops-cross` chains
#     ("tengo algún workshop agendado sobre RAG" after "obsidian-rag
#     pipelines"). The continuation references the prior turn's entity
#     implicitly via "tengo".
_FOLLOWUP_CUES_EXT_LEADING_Y = re.compile(
    r"^y\s+(el|la|los|las|qué|que|cómo|como|cu[áa]l|otra|otro|otras|otros)\b",
    re.IGNORECASE,
)
_FOLLOWUP_CUES_EXT_LEADING_VERB_DEFART = re.compile(
    r"^(listame|dame|mostrame|contame|explicame|detallame|decime|enumerame|"
    r"resume|resumi|resumí)\s+(el|la|los|las)\s+\w+",
    re.IGNORECASE,
)
_FOLLOWUP_CUES_EXT_TENGO_ALGUN = re.compile(
    r"^tengo\s+alg[uú]n\b|^qu[ée]\s+\w+\s+(tengo|us[oé]|tenía|tuve)\b",
    re.IGNORECASE,
)

# Leading-pronoun pattern: question starts with a pronoun/demonstrative that
# requires prior context to resolve. Only checked at position 0 (after strip).
_LEADING_PRONOUN_RE = re.compile(
    r"^(eso|esa|ese|esto|esta|este|ella|[eé]l|it|this|that|he|she)\b",
    re.IGNORECASE,
)

# Short interrogative without an explicit subject — "qué hay de X?", "cuál
# es?", "cómo funciona?" where X ≤ 1 token. Only fires for ≤5-word questions
# so long complete questions pass through.
_SHORT_INTERROGATIVE_RE = re.compile(
    r"^(qué|que|cuál|cual|cómo|como|por\s+qué|por\s+que|"
    r"what|how|which|why)\b",
    re.IGNORECASE,
)


# List-intent: query pide explícitamente una LISTA / ENUMERACIÓN amplia.
# Cuando matchea, el endpoint sube k y rerank_pool en `multi_retrieve`
# para devolver más fuentes — "mostrame todas las notas sobre X" con
# k=4 starveaba al LLM, que terminaba alucinando o devolviendo respuestas
# triviales. Pre-fix Playwright autónomo 2026-04-28: "mostrame todas las
# notas relacionadas con RAG y embeddings" trajo solo 4 fuentes con
# confidence 0.334 y el LLM alucinó "RAG (Rapid Application Generation)"
# por falta de anclaje. Con k=12 + multi_query=True el retrieval cubre
# decenas de notas reales del corpus y el LLM tiene contexto suficiente
# para no inventar. NO matchea queries de "lista" cortas tipo "qué
# tengo" (esos van por el pre-router de tools); este regex solo dispara
# cuando hay un cuantificador explícito (todas/todos/varias/cuántas).
_LIST_INTENT_RE = re.compile(
    r"\b(todas|todos|cu[áa]ntas|cu[áa]ntos|enumera|listame|mostrame|"
    r"dame|deja(?:me)?\s+ver)\s+(las|los|mis|tus|cu[áa]nt[oa]s)?\s*"
    r"(notas|archivos|menciones|documentos|referencias|hilos|entradas|"
    r"tareas|projectos|proyectos|recordatorios|reuniones)",
    re.IGNORECASE,
)


def _looks_like_followup(q: str) -> bool:
    """Heuristic: is this question elliptical / context-dependent?

    When True, the chat handler concatenates the prior user turn into the
    retrieve query (`{last_user_q} {q}`) — cheap (0ms), empirically
    anchors retrieval to the right entity (~+0.20 rerank).
    Concretely fires when the question matches any of:

      (a) ≤2 words — genuine ellipsis: "y eso?", "más?", "cuál?"
      (b) Starts with a pronoun/demonstrative: "eso que dijiste", "ella cómo?"
      (c) ≤5-word question starting with an interrogative word (incomplete
          subject implied by context): "qué hay de eso?"
      (d) Anaphora cue anywhere: pronoun, "y de/sobre/...", "profundizá",
          "más sobre", "amplía", etc. (see `_FOLLOWUP_CUES`).
      (e) Question-initial "y + article/wh-word": "y el icbc", "y la
          retrospectiva", "y qué pendientes tengo", "y cómo las indexás".
          Anchored to ^ so only follow-up continuations match — standalone
          questions starting with mid-sentence "y" don't fire.
      (f) Imperative + definite article + noun: "listame los gastos",
          "mostrame las cuentas", "detallame los pagos". The user pattern
          is "I asked about X, now elaborate without re-naming X".
      (g) "tengo algún X" / "qué <noun> tengo|usé|tenía": implicit cross-
          turn entity reference via possessive "tengo".

    Long standalone questions (≥4 words without leading cues) skip the
    concat — the original query is already a good search target. Borderline
    turns (history exists, no regex match, cosine 0.4-0.7) get a separate
    `reformulate_query` LLM rewrite — the hybrid path is in `chat()`.
    """
    if not q:
        return False
    words = q.split()
    # (a) very short — almost certainly anaphoric
    if len(words) <= 2:
        return True
    # (b) starts with a pronoun/demonstrative that needs antecedent resolution
    stripped = q.strip().lstrip("¿").strip()
    if _LEADING_PRONOUN_RE.match(stripped):
        return True
    # (c) short interrogative (≤5 words) — subject likely implicit from context
    if len(words) <= 5 and _SHORT_INTERROGATIVE_RE.match(stripped):
        return True
    # (d) explicit anaphora cue anywhere in the question
    if _FOLLOWUP_CUES.search(q):
        return True
    # (e) leading "y + article/wh"
    if _FOLLOWUP_CUES_EXT_LEADING_Y.match(stripped):
        return True
    # (f) leading verb + definite article + noun
    if _FOLLOWUP_CUES_EXT_LEADING_VERB_DEFART.match(stripped):
        return True
    # (g) "tengo algún X" / "qué X tengo"
    if _FOLLOWUP_CUES_EXT_TENGO_ALGUN.match(stripped):
        return True
    return False


# Pronouns / demonstratives that, if still present after reformulation,
# signal the helper model failed to resolve the antecedent. Triggers the
# concat-with-last-user-turn fallback. `el`/`la`/`lo` excluded because
# they're also common articles and would fire on every sentence.
_UNRESOLVED_PRONOUN_RE = re.compile(
    r"\b(ella|él|eso|esa|ese|esto|esta|este|aquello|"
    r"ahí|ahi|allí|alli|allá|alla)\b",
    re.IGNORECASE,
)


_SCOPE_N_DAYS = re.compile(
    r"\bpr[oó]xim[oa]s?\s+(\d+)\s+d[ií]as?\b|\bnext\s+(\d+)\s+days?\b",
    re.IGNORECASE,
)
_SCOPE_WEEK = re.compile(
    r"\besta\s+semana\b|\bpr[oó]xima\s+semana\b|\bsemana\s+que\s+viene\b|"
    r"\bthis\s+week\b|\bnext\s+week\b",
    re.IGNORECASE,
)
_SCOPE_TOMORROW = re.compile(r"\bma(?:ñ|n)ana\b|\btomorrow\b", re.IGNORECASE)
_SCOPE_TODAY = re.compile(
    r"\bhoy\b|\bahora\b|\btoday\b|\beste\s+momento\b", re.IGNORECASE,
)


def _detect_time_scope(q: str) -> dict:
    """Return the time scope carried by the query. Priority: explicit N days
    beats week beats tomorrow beats today; unmarked queries get a 14-day
    reminders window + 7-day calendar window (user asking "qué tengo
    pendiente" wants a broad sweep).
    """
    m = _SCOPE_N_DAYS.search(q)
    if m:
        n = int(m.group(1) or m.group(2) or 0)
        n = max(1, min(30, n))
        return {
            "key": "n_days", "label": f"próximos {n} días",
            "reminders_horizon": n + 1, "calendar_ahead": n,
        }
    if _SCOPE_WEEK.search(q):
        return {
            "key": "week", "label": "esta semana",
            "reminders_horizon": 8, "calendar_ahead": 7,
        }
    if _SCOPE_TOMORROW.search(q):
        return {
            "key": "tomorrow", "label": "mañana",
            "reminders_horizon": 2, "calendar_ahead": 1,
        }
    if _SCOPE_TODAY.search(q):
        return {
            "key": "today", "label": "hoy",
            "reminders_horizon": 1, "calendar_ahead": 0,
        }
    return {
        "key": "general", "label": "próximas 2 semanas",
        "reminders_horizon": 14, "calendar_ahead": 7,
    }


def _fetch_calendar_ahead(days_ahead: int, max_events: int = 40) -> list[dict]:
    """Events from today through today+days_ahead via icalBuddy with relative
    date labels ("today" / "tomorrow" / "day after tomorrow" / YYYY-MM-DD).
    Returns [{date_label, title, time_range}]. Empty list on any failure —
    the brief layer tolerates silent missing calendar.
    """
    if not _apple_enabled() or days_ahead < 0:
        return []
    icb = _icalbuddy_path()
    if not icb:
        return []
    query = "eventsToday" if days_ahead == 0 else f"eventsToday+{days_ahead}"
    try:
        res = subprocess.run(
            [icb, "-nc", "-iep", "title,datetime", "-b", "", query],
            capture_output=True, text=True, errors="replace", timeout=10.0,
        )
    except (FileNotFoundError, subprocess.TimeoutExpired, OSError):
        return []
    if res.returncode != 0 or not (res.stdout or "").strip():
        return []

    events: list[dict] = []
    current: dict | None = None
    for raw in (res.stdout or "").splitlines():
        line = raw.rstrip()
        if not line:
            continue
        if not line.startswith(" ") and not line.startswith("\t"):
            if current and current.get("title"):
                events.append(current)
            current = {"title": line.strip(), "date_label": "", "time_range": ""}
            continue
        stripped = line.strip()
        if current is None:
            continue
        m = re.search(
            r"(\d{1,2}:\d{2}(?:\s*[AaPp][Mm])?)\s*-\s*(\d{1,2}:\d{2}(?:\s*[AaPp][Mm])?)",
            stripped,
        )
        if m:
            current["time_range"] = f"{m.group(1)}–{m.group(2)}"
            # date label is everything before the "at HH:MM" span
            label = stripped.split(" at ", 1)[0].strip()
            if label and label != stripped:
                current["date_label"] = label
        else:
            # date-only line (all-day event)
            current["date_label"] = stripped
    if current and current.get("title"):
        events.append(current)
    return events[:max_events]


def _collect_scoped_evidence(col, now: datetime, scope: dict) -> dict:
    """`_pendientes_collect` under the hood for services that are scope-free
    (gmail, mail, whatsapp, loops, contradictions), then override `reminders`
    and `calendar` with scope-aware fetches. WhatsApp always stays on — user
    wants the recent chat context attached even to non-WA questions.
    """
    ev = _pendientes_collect(col, now, days=14)
    from contextlib import suppress
    with suppress(Exception):
        ev["reminders"] = _fetch_reminders_due(
            now,
            horizon_days=int(scope["reminders_horizon"]),
            max_items=40,
        )
    with suppress(Exception):
        ev["calendar_range"] = _fetch_calendar_ahead(int(scope["calendar_ahead"]))
    # Ensure whatsapp is always populated even if `_pendientes_collect`
    # skipped it silently (bridge transiently down, etc.).
    if not ev.get("whatsapp"):
        with suppress(Exception):
            ev["whatsapp"] = _fetch_whatsapp_unread(hours=24, max_chats=8)
    return ev


def _format_pendientes_context(
    ev: dict, urgent: list[str], now: datetime, scope: dict,
) -> str:
    """Markdown evidence blob for the LLM. Scope label is front-and-centre so
    the model filters to the right window. WhatsApp is always included —
    even a non-WA question benefits from seeing recent chat context.
    """
    lines: list[str] = [
        f"# Evidencia de servicios · {now.strftime('%Y-%m-%d %H:%M')}",
        f"**Ventana temporal pedida:** {scope['label']}",
    ]

    if urgent:
        lines.append(f"\n## 🚨 Urgente / Overdue ({len(urgent)})")
        lines.extend(f"- {u}" for u in urgent)

    cal = ev.get("calendar_range") or ev.get("calendar") or []
    if cal:
        lines.append(f"\n## 📅 Calendar ({len(cal)} eventos)")
        for e in cal[:20]:
            when_bits = [b for b in (e.get("date_label"), e.get("time_range")) if b]
            if not when_bits and e.get("start"):
                when_bits = [e["start"]]
            when = " · ".join(when_bits) or "sin hora"
            lines.append(f"- {when} — {e.get('title','')}")

    rem = ev.get("reminders") or []
    if rem:
        lines.append(f"\n## 📌 Apple Reminders ({len(rem)})")
        for bucket in ("overdue", "today", "upcoming", "undated"):
            bucket_items = [x for x in rem if (x.get("bucket") or "") == bucket]
            for r in bucket_items[:10]:
                due = f" (due {r['due']})" if r.get("due") else ""
                lst = f" [{r['list']}]" if r.get("list") else ""
                lines.append(f"- **{bucket}**: {r.get('name','')}{due}{lst}")

    gm = ev.get("gmail") or {}
    awaiting = gm.get("awaiting_reply") or []
    if awaiting:
        lines.append(f"\n## 📧 Gmail awaiting reply ({len(awaiting)})")
        for m in awaiting[:8]:
            lines.append(
                f"- {m.get('days_old',0)}d · {m.get('subject','')} "
                f"— {m.get('from','')}"
            )

    mail = ev.get("mail_unread") or []
    if mail:
        lines.append(f"\n## 📬 Apple Mail unread 36h ({len(mail)})")
        for m in mail[:5]:
            vip = "VIP · " if m.get("is_vip") else ""
            lines.append(f"- {vip}{m.get('subject','')} — {m.get('sender','')}")

    wa = ev.get("whatsapp") or []
    if wa:
        total = sum(int(w.get("count", 0)) for w in wa)
        lines.append(
            f"\n## 💬 WhatsApp últimas 24h ({len(wa)} chats · {total} msgs)"
        )
        for w in wa[:10]:
            snip = (w.get("last_snippet") or "").strip()[:100]
            snip_part = f' — "{snip}"' if snip else ""
            lines.append(
                f"- {w.get('name','?')} ({w.get('count',0)} msgs){snip_part}"
            )

    stale = ev.get("loops_stale") or []
    activo = ev.get("loops_activo") or []
    if stale or activo:
        lines.append(
            f"\n## 📁 Vault loops ({len(stale)} stale · {len(activo)} activo)"
        )
        for it in stale[:5]:
            src = Path(it["source_note"]).stem
            lines.append(
                f"- stale {it.get('age_days',0)}d: "
                f"{it.get('loop_text','')[:100]} [[{src}]]"
            )
        for it in activo[:5]:
            src = Path(it["source_note"]).stem
            lines.append(
                f"- activo {it.get('age_days',0)}d: "
                f"{it.get('loop_text','')[:100]} [[{src}]]"
            )

    contrad = ev.get("contradictions") or []
    if contrad:
        lines.append(f"\n## ⚠ Contradicciones recientes ({len(contrad)})")
        for c in contrad[:3]:
            tgt = ", ".join(t.get("path", "") for t in c.get("targets", [])[:2])
            lines.append(f"- {c.get('subject_path','')} ↔ {tgt}")

    return "\n".join(lines)


def _pendientes_services_consulted(ev: dict) -> list[str]:
    """Labels of services that returned data — used for UI meta."""
    out: list[str] = []
    if (ev.get("gmail") or {}).get("awaiting_reply"):
        out.append("Gmail")
    if ev.get("mail_unread"):
        out.append("Apple Mail")
    if ev.get("whatsapp"):
        out.append("WhatsApp")
    if ev.get("reminders"):
        out.append("Reminders")
    if ev.get("calendar_range") or ev.get("calendar"):
        out.append("Calendar")
    if ev.get("loops_stale") or ev.get("loops_activo"):
        out.append("Vault loops")
    return out


def _build_tasks_system_rules(scope: dict) -> str:
    """SYSTEM prompt with the concrete time-window embedded. Keeping it per-
    call (instead of a global constant) avoids the 'Sin pendientes' parroting
    failure mode — the model now has to filter by scope, not repeat a fixed
    fallback string.
    """
    return (
        "Sos un asistente personal que organiza la agenda del usuario a "
        "partir de evidencia EN VIVO de sus servicios (Calendar, Reminders, "
        "Gmail, Apple Mail, WhatsApp, loops del vault).\n\n"
        f"VENTANA TEMPORAL: {scope['label']}.\n\n"
        "Reglas estrictas:\n"
        "1. Usá SOLO la evidencia del CONTEXTO. No inventes tareas, "
        "reuniones ni plazos que no estén literalmente listados.\n"
        "2. Filtrá los items a la ventana temporal indicada. "
        "Items de Reminders fuera de la ventana → omitilos.\n"
        "3. Respondé en español, estructurado con encabezados markdown y "
        "bullets. Omití secciones que queden vacías después del filtro.\n"
        "4. Orden: 🚨 Urgente/overdue → 📅 Agenda (Calendar) → "
        "📌 Reminders → 📧 Bandeja (mails) → 💬 WhatsApp → 📁 Vault loops.\n"
        "5. Sé conciso: una línea por item. Para eventos de Calendar incluí "
        "día + hora + título.\n"
        "6. Si tras filtrar por la ventana no queda NADA, respondé "
        "exactamente: 'No tengo nada trackeado para esa ventana.' y parate "
        "ahí — NO copies la frase 'Sin pendientes trackeados' del contexto."
    )


def _gen_tasks_response(sess: dict, question: str, history: list[dict]):
    """SSE generator for the tasks/agenda intent branch. Collects scope-aware
    evidence from ALL registered vaults (home + work) via rag's multi-vault
    collector, builds the structured context, streams the LLM answer.

    Delegates formatting + prompt rules to `rag.py` so web and `rag serve`
    (WhatsApp/bot surface) produce identical output — same sections, same
    `[home]`/`[work]` loop labels, same "SIEMPRE mostrar" prompt semantics.
    """
    now = datetime.now()
    scope = _detect_time_scope(question)
    try:
        all_vaults = resolve_vault_paths(["all"])
    except Exception:
        all_vaults = []
    try:
        if all_vaults:
            ev = _rag_collect_scoped_tasks_evidence_multi(all_vaults, now, scope)
        else:
            ev = _collect_scoped_evidence(get_db(), now, scope)
    except Exception as exc:
        yield _sse("error", {"message": f"servicios fallaron: {exc}"})
        return
    urgent = _pendientes_urgent(ev, now)
    services = (
        _rag_tasks_services_consulted(ev) if all_vaults
        else _pendientes_services_consulted(ev)
    )
    context = (
        _rag_format_scoped_tasks_context(ev, urgent, now, scope) if all_vaults
        else _format_pendientes_context(ev, urgent, now, scope)
    )
    system_rules = _rag_build_tasks_system_rules(scope)

    meta_bits: list[str] = [f"📋 ventana: {scope['label']}"]
    if services:
        meta_bits.append(", ".join(services))
    else:
        meta_bits.append("sin datos de servicios")
    if urgent:
        meta_bits.append(f"🚨 {len(urgent)} urgente(s)")
    yield _sse("meta", {"bits": meta_bits})

    if history:
        messages = (
            [{"role": "system", "content": f"{system_rules}\nCONTEXTO:\n{context}"}]
            + history
            + [{"role": "user", "content": question}]
        )
    else:
        messages = [{"role": "user", "content": (
            f"{system_rules}\nCONTEXTO:\n{context}\n\n"
            f"PREGUNTA: {question}\n\nRESPUESTA:"
        )}]

    parts: list[str] = []
    # Tasks layout with 6+ sections + per-item bullets overflows the default
    # 384 num_predict cap — LLM stops mid-"Vault loops". Bump to 800 so the
    # full brief lands even with both vaults' loops expanded.
    tasks_options = {**CHAT_OPTIONS, "num_predict": 800}
    try:
        for chunk in _OLLAMA_STREAM_CLIENT.chat(
            model=resolve_chat_model(),
            messages=messages,
            options=tasks_options,
            stream=True,
            keep_alive=chat_keep_alive(),
        ):
            delta = chunk.message.content or ""
            if delta:
                parts.append(delta)
                yield _sse("token", {"delta": delta})
    except Exception as exc:
        # Ver `[chat-stream-error] phase=synthesis` en /api/chat para
        # rationale — mismo bug de observabilidad acá. Este es el brief
        # de /api/tasks (streaming directo, sin tool loop).
        print(
            f"[chat-stream-error] phase=tasks_brief "
            f"exc_type={type(exc).__name__} exc={exc} "
            f"parts_so_far={len(parts)}",
            flush=True,
        )
        yield _sse("error", {"message": _sanitize_error_for_user(exc, phase="tasks_brief")})
        return

    full = "".join(parts)
    turn_id = new_turn_id()
    append_turn(sess, {
        "q": question,
        "a": full,
        "paths": [],
        "top_score": 0.0,
        "turn_id": turn_id,
        "mode": "tasks",
    })
    save_session(sess)
    log_query_event({
        "cmd": "web.tasks",
        "turn_id": turn_id,
        "session": sess["id"],
        "q": question,
        "scope": scope["key"],
        "scope_label": scope["label"],
        "services": services,
        "n_urgent": len(urgent),
        "n_reminders": len(ev.get("reminders") or []),
        "n_calendar": len(
            ev.get("calendar_range") or ev.get("calendar") or []
        ),
        "n_whatsapp": len(ev.get("whatsapp") or []),
        "n_loops_activo": len(ev.get("loops_activo") or []),
        "n_loops_stale": len(ev.get("loops_stale") or []),
    })

    yield _sse("done", {"turn_id": turn_id, "mode": "tasks"})


_TODAY_FM_SPLIT = re.compile(r"^---\s*\n.*?\n---\s*\n", re.DOTALL)


def _persist_today_brief(date_label: str, narrative: str) -> None:
    """Write the regenerated brief to `04-Archive/99-obsidian-system/99-AI/reviews/YYYY-MM-DD-evening.md` so
    subsequent `/api/home` calls hit the cached path. Mirrors the CLI
    `rag today` write — same frontmatter shape so contradiction detection
    and ambient skip rules treat both files the same. Silent-fail.
    """
    from contextlib import suppress

    now_iso = datetime.now().isoformat(timespec="seconds")
    fm_lines = [
        "---",
        f"created: '{now_iso}'",
        "type: evening-brief",
        "tags:",
        "- review",
        "- evening-brief",
        f"date: '{date_label}'",
        "source: web",
        "---",
    ]
    body = "\n".join(fm_lines) + f"\n\n# Evening brief — {date_label}\n\n{narrative}\n"
    path = VAULT_PATH / MORNING_FOLDER / f"{date_label}-evening.md"
    with suppress(Exception):
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(body, encoding="utf-8")


def _today_cached_narrative(date_label: str) -> str | None:
    """Return narrative from `04-Archive/99-obsidian-system/99-AI/reviews/YYYY-MM-DD-evening.md` if present.
    CLI `rag today` writes this file, so we avoid re-running the LLM when the
    brief was already generated today.
    """
    path = VAULT_PATH / MORNING_FOLDER / f"{date_label}-evening.md"
    if not path.is_file():
        return None
    try:
        raw = path.read_text(encoding="utf-8")
    except OSError:
        return None
    body = _TODAY_FM_SPLIT.sub("", raw, count=1).strip()
    # Strip the "# Evening brief — YYYY-MM-DD" header if present
    lines = body.splitlines()
    if lines and lines[0].startswith("# Evening brief"):
        body = "\n".join(lines[1:]).lstrip()
    return body or None


def _fetch_pagerank_top(col, n: int = 5) -> list[dict]:
    """Top-n notes by wikilink PageRank authority. Normalized pr (pr/max_pr).

    Reuses the module-level cache in rag.get_pagerank — cost is a dict sort
    after the first call. Silent-fail via caller suppressor.
    """
    pr_map = get_pagerank(col)
    if not pr_map:
        return []
    corpus = _load_corpus(col)
    ranked = sorted(pr_map.items(), key=lambda kv: -kv[1])[:n]
    # Defensive: `n <= 0` would make `ranked = []` despite a non-empty map,
    # and a hypothetical cache-invalidation race between `get_pagerank` and
    # `sorted` could also narrow it to zero rows. Return early rather than
    # IndexError-ing on `ranked[0][1]`.
    if not ranked:
        return []
    max_pr = ranked[0][1] or 1.0
    out: list[dict] = []
    for rank, (path, pr) in enumerate(ranked, start=1):
        title = _path_to_title(corpus, path) or Path(path).stem
        out.append({
            "title": title,
            "path": path,
            "pr": round(pr / max_pr, 4),
            "rank": rank,
        })
    return out


def _fetch_chrome_top_week(n: int = 5) -> list[dict]:
    """Top-n Chrome pages visited in the last 7 days.

    Chrome's `History` SQLite db is held open (locked) by the running
    browser — copy to a tmp file before reading. Chrome's `visit_time` is
    microseconds since 1601-01-01 UTC; convert to unix epoch for the iso
    stamp. Filters out chrome://, file://, and localhost.

    Known limitations (document, not fix — this endpoint is ambient signal,
    not a source of truth):
    - **Multi-profile**: only reads `Default/`. Chrome stores per-profile
      histories in sibling dirs (`Profile 1/History`, `Profile 2/History`, …).
      Users with multiple profiles get a partial top-n. `Local State` JSON
      would list them if we ever want to aggregate.
    - **WAL mode**: Chrome uses SQLite WAL; recent visits may live in
      `History-wal` until checkpointed. Copying only `History` misses the
      tail of "right now". Acceptable for a 7-day window.
    - **Guest mode / Incognito**: never written to disk by design — invisible
      to this fetcher, correctly so.
    """
    import shutil
    import sqlite3
    import tempfile

    src = Path.home() / "Library/Application Support/Google/Chrome/Default/History"
    if not src.is_file():
        return []

    CHROME_EPOCH_OFFSET = 11_644_473_600  # seconds between 1601-01-01 and 1970-01-01
    now_ts = time.time()
    week_ago_chrome = int((now_ts - 7 * 86_400 + CHROME_EPOCH_OFFSET) * 1_000_000)

    with tempfile.NamedTemporaryFile(suffix=".sqlite", delete=True) as tmp:
        shutil.copyfile(src, tmp.name)
        conn = sqlite3.connect(f"file:{tmp.name}?mode=ro", uri=True)
        try:
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                """
                SELECT u.url AS url, u.title AS title,
                       COUNT(v.id) AS visit_count,
                       MAX(v.visit_time) AS last_visit
                FROM urls u
                JOIN visits v ON v.url = u.id
                WHERE v.visit_time >= ?
                  AND u.url NOT LIKE 'chrome://%'
                  AND u.url NOT LIKE 'chrome-extension://%'
                  AND u.url NOT LIKE 'file://%'
                  AND u.url NOT LIKE '%://localhost%'
                  AND u.url NOT LIKE '%://127.0.0.1%'
                GROUP BY u.url
                ORDER BY visit_count DESC
                LIMIT ?
                """,
                (week_ago_chrome, n),
            ).fetchall()
        finally:
            conn.close()

    out: list[dict] = []
    for r in rows:
        last_unix = (r["last_visit"] / 1_000_000) - CHROME_EPOCH_OFFSET
        out.append({
            "title": r["title"] or "",
            "url": r["url"],
            "visit_count": int(r["visit_count"]),
            "last_visit_iso": datetime.fromtimestamp(last_unix).isoformat(timespec="seconds"),
        })
    return out


def _fetch_youtube_watched(n: int = 5, hours: int = 168) -> list[dict]:
    """Most-recent YouTube video pages opened in Chrome, last `hours` window.

    Same Chrome History DB + tmp-copy pattern as `_fetch_chrome_top_week`.
    Matches `youtube.com/watch`, `m.youtube.com/watch`, and `youtu.be/<id>`.
    Deduplicates by extracted video id (`v=...` for /watch, path segment
    for youtu.be) so re-opening the same video doesn't clutter the panel.
    Titles in Chrome history are the <title> tag at page load — the ad
    slate title occasionally leaks in instead of the real video title;
    acceptable noise for an ambient signal.
    """
    import shutil
    import sqlite3
    import tempfile
    from urllib.parse import urlparse, parse_qs

    src = Path.home() / "Library/Application Support/Google/Chrome/Default/History"
    if not src.is_file():
        return []

    CHROME_EPOCH_OFFSET = 11_644_473_600
    now_ts = time.time()
    window_start_chrome = int((now_ts - hours * 3_600 + CHROME_EPOCH_OFFSET) * 1_000_000)

    with tempfile.NamedTemporaryFile(suffix=".sqlite", delete=True) as tmp:
        shutil.copyfile(src, tmp.name)
        conn = sqlite3.connect(f"file:{tmp.name}?mode=ro", uri=True)
        try:
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                """
                SELECT u.url AS url, u.title AS title,
                       COUNT(v.id) AS visit_count,
                       MAX(v.visit_time) AS last_visit
                FROM urls u
                JOIN visits v ON v.url = u.id
                WHERE v.visit_time >= ?
                  AND (
                       u.url LIKE '%://www.youtube.com/watch%'
                    OR u.url LIKE '%://youtube.com/watch%'
                    OR u.url LIKE '%://m.youtube.com/watch%'
                    OR u.url LIKE '%://youtu.be/%'
                  )
                GROUP BY u.url
                ORDER BY last_visit DESC
                LIMIT 50
                """,
                (window_start_chrome,),
            ).fetchall()
        finally:
            conn.close()

    seen_ids: set[str] = set()
    out: list[dict] = []
    for r in rows:
        url = r["url"]
        try:
            parsed = urlparse(url)
            if parsed.netloc.endswith("youtu.be"):
                vid = parsed.path.strip("/").split("/")[0] or url
            else:
                vid = (parse_qs(parsed.query).get("v") or [url])[0]
        except Exception:
            vid = url
        if vid in seen_ids:
            continue
        seen_ids.add(vid)
        # "Video Title - YouTube" → "Video Title"; Chrome appends " - YouTube"
        # to the page <title> for every watch page.
        raw_title = (r["title"] or "").strip()
        title = raw_title[:-len(" - YouTube")].rstrip() if raw_title.endswith(" - YouTube") else raw_title
        last_unix = (r["last_visit"] / 1_000_000) - CHROME_EPOCH_OFFSET
        out.append({
            "title": title or url,
            "url": url,
            "video_id": vid,
            "visit_count": int(r["visit_count"]),
            "last_visit_iso": datetime.fromtimestamp(last_unix).isoformat(timespec="seconds"),
        })
        if len(out) >= n:
            break
    return out


# Baselines from CLAUDE.md (2026-04-16 post-quick-wins floor). Hardcoded
# because they're the reference against which the UI renders drift — changing
# the baseline needs a human decision, not an automatic update from the latest
# `rag eval` entry.
_EVAL_BASELINE = {
    "singles_hit5": 0.9048,
    "chains_hit5": 0.76,
    "chains_mrr": 0.58,
    "chain_success": 0.5556,
}


def _fetch_eval_trend(n: int = 10) -> dict | None:
    """Tail the last `n` eval runs from rag_eval_runs and pair with baseline.
    Returns None when the table is missing or empty (SQL-only since T10).
    """
    try:
        with _ragvec_state_conn() as conn:
            rows = conn.execute(
                "SELECT ts, extra_json FROM rag_eval_runs"
                " ORDER BY ts DESC LIMIT ?",
                (n,),
            ).fetchall()
    except Exception:
        return None
    if not rows:
        return None
    history: list[dict] = []
    for row in reversed(rows):
        entry: dict = {"ts": row[0]}
        if row[1]:
            try:
                entry.update(json.loads(row[1]))
            except Exception:
                pass
        history.append(entry)
    return {
        "latest": history[-1],
        "baseline": dict(_EVAL_BASELINE),
        "history": history,
    }


# Followup aging runs the full `find_followup_loops` pipeline, which costs one
# LLM judge call per open loop. Cheap per-loop but the list scales with vault
# size — 6h cache keeps /api/home under ~2s while still refreshing a few
# times per working day.
_FOLLOWUP_AGING_CACHE: dict = {"ts": 0.0, "payload": None}
_FOLLOWUP_AGING_TTL = 6 * 3600


def _fetch_followup_aging() -> dict | None:
    """Bucketize open-loops-by-age for the /api/home aging widget.

    Buckets: 0-7d, 8-30d, stale_30plus (age≥31 OR status=='stale'). Uses a
    90d window on `find_followup_loops` so the 30+ bucket actually sees aged
    items (the default CLI window is 30d). Resolved loops are dropped —
    they are, by definition, no longer aging.
    """
    now_ts = time.time()
    cached = _FOLLOWUP_AGING_CACHE
    if cached["payload"] is not None and now_ts - cached["ts"] < _FOLLOWUP_AGING_TTL:
        return cached["payload"]

    try:
        col = get_db()
        loops = find_followup_loops(col, VAULT_PATH, days=90)
    except Exception:
        return cached["payload"]  # serve last-known-good on failure

    open_loops = [it for it in loops if it.get("status") != "resolved"]
    buckets = {"0_7": 0, "8_30": 0, "stale_30plus": 0}
    for it in open_loops:
        age = int(it.get("age_days") or 0)
        status = it.get("status") or ""
        if age >= 31 or status == "stale":
            buckets["stale_30plus"] += 1
        elif age >= 8:
            buckets["8_30"] += 1
        else:
            buckets["0_7"] += 1

    sample = sorted(open_loops, key=lambda it: -int(it.get("age_days") or 0))[:5]
    sample_out = [
        {
            "note": it.get("source_note") or "",
            "loop": (it.get("loop_text") or "")[:140],
            "age_days": int(it.get("age_days") or 0),
            "status": it.get("status") or "",
        }
        for it in sample
    ]

    payload = {
        "buckets": buckets,
        "total": len(open_loops),
        "sample": sample_out,
    }
    _FOLLOWUP_AGING_CACHE["ts"] = now_ts
    _FOLLOWUP_AGING_CACHE["payload"] = payload
    return payload


def _fetch_drive_recent(now: datetime, max_items: int = 5) -> list[dict]:
    """Files modified in Google Drive in the last 48h. Thin wrapper over
    rag._fetch_drive_evidence — keeps the home panel's window smaller than
    morning's 5d so "recent" on the dashboard means actionable-today, not
    week-in-review. Silent-fail via the underlying fetcher → [].
    """
    ev = _fetch_drive_evidence(now, days=2, max_items=max_items) or {}
    return ev.get("files") or []


def _fetch_whatsapp_unreplied(hours: int = 48, max_chats: int = 5) -> list[dict]:
    """Chats whose **last** message is inbound and sits without a reply.

    Distinct from `_fetch_whatsapp_unread` (counts all inbound in window):
    this one only surfaces conversations where *you* owe the next move.
    A chat with 20 inbound-then-replied messages does not appear here; a
    chat with 1 inbound-and-ignored message does.

    Uses a window function to pick the latest row per chat, then filters
    to `is_from_me = 0`. Window: `hours` bounds how stale a last-inbound
    is allowed to be — anything older is almost certainly abandoned and
    pollutes the panel.

    Schema note: the bridge stores timestamps in RFC3339-ish text; SQLite's
    `datetime()` handles it but raw string comparison would not — keep the
    `datetime(timestamp)` wrappers.
    """
    if not WHATSAPP_DB_PATH.is_file():
        return []
    import sqlite3
    try:
        con = sqlite3.connect(f"file:{WHATSAPP_DB_PATH}?mode=ro", uri=True, timeout=5.0)
    except sqlite3.Error:
        return []
    try:
        con.row_factory = sqlite3.Row
        rows = con.execute(
            """
            WITH last_msg AS (
              SELECT chat_jid, content, is_from_me, timestamp,
                     ROW_NUMBER() OVER (
                       PARTITION BY chat_jid
                       ORDER BY datetime(timestamp) DESC
                     ) AS rn
              FROM messages
              WHERE chat_jid != ?
                AND chat_jid NOT LIKE '%status@broadcast'
                AND datetime(timestamp) > datetime('now', ?)
            )
            SELECT lm.chat_jid   AS jid,
                   c.name        AS name,
                   lm.content    AS last_content,
                   lm.timestamp  AS last_ts
            FROM last_msg lm
            LEFT JOIN chats c ON c.jid = lm.chat_jid
            WHERE lm.rn = 1 AND lm.is_from_me = 0
            ORDER BY datetime(lm.timestamp) DESC
            LIMIT ?
            """,
            (WHATSAPP_BOT_JID, f"-{int(hours)} hours", int(max_chats) * 3),
        ).fetchall()
    except sqlite3.Error:
        return []
    finally:
        con.close()

    now_ts = time.time()
    out: list[dict] = []
    for r in rows:
        raw_name = (r["name"] or "").strip()
        jid_prefix = (r["jid"] or "").split("@")[0]
        display_name = raw_name or jid_prefix
        # Same "real name" gate as _fetch_whatsapp_unread: digit-only names
        # are unresolved @lid participants, unhelpful to surface.
        if not any(ch.isalpha() for ch in display_name):
            continue
        snippet = (r["last_content"] or "").strip().replace("\n", " ")
        if len(snippet) > 120:
            snippet = snippet[:117] + "…"
        # Bridge stores naive local timestamps; `fromisoformat` parses either
        # naive or tz-aware and `.timestamp()` does the right thing for both.
        try:
            last_dt = datetime.fromisoformat((r["last_ts"] or "").replace("Z", "+00:00"))
            hours_waiting = max(0.0, (now_ts - last_dt.timestamp()) / 3600.0)
        except Exception:
            hours_waiting = 0.0
        out.append({
            "jid": r["jid"],
            "name": display_name,
            "last_snippet": snippet,
            "hours_waiting": round(hours_waiting, 1),
        })
        if len(out) >= max_chats:
            break
    return out


# MOZE (Money app) export — user drops `MOZE_YYYYMMDD_HHMMSS.csv` into the
# iCloud `/Finances` folder (que también aloja los `.xlsx` de resúmenes de
# tarjeta — ver `_fetch_credit_cards`). Tomamos el CSV más nuevo y lo
# parseamos local; sin red, sin API. Dates: MM/DD/YYYY (US); Price: ES
# decimals ("2026,74"). Gastos vienen como negativos → abs() para mostrar.
#
# Migración 2026-04-26: el dir antes era `/Backup`. Si en el futuro el user
# vuelve a moverlo, el override por env vive en `_FINANCE_DIR_ENV`.
_FINANCE_DIR_ENV = "OBSIDIAN_RAG_FINANCE_DIR"
_FINANCE_BACKUP_DIR = Path(
    os.environ.get(_FINANCE_DIR_ENV, "")
    or (Path.home() / "Library/Mobile Documents/com~apple~CloudDocs/Finances")
)
_FINANCE_CACHE: dict = {"key": None, "payload": None}


def _fetch_finance(now: datetime | None = None) -> dict | None:
    """Parse the latest MOZE_*.csv export into a home-panel summary.

    Returns None (silent-fail) if: no CSV found, iCloud folder missing,
    or no rows parseable. Cached by (path, mtime) — re-exports invalidate.
    """
    import calendar
    import csv as _csv
    from collections import Counter

    now = now or datetime.now()
    try:
        csvs = sorted(
            _FINANCE_BACKUP_DIR.glob("MOZE_*.csv"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
    except Exception:
        return None
    if not csvs:
        return None
    src = csvs[0]
    try:
        mtime = src.stat().st_mtime
    except Exception:
        return None
    key = (str(src), mtime)
    if _FINANCE_CACHE["key"] == key:
        return _FINANCE_CACHE["payload"]

    def pnum(s: str) -> float:
        s = (s or "").strip().replace(".", "").replace(",", ".")
        try:
            return float(s)
        except Exception:
            return 0.0

    rows: list[tuple[datetime, dict]] = []
    try:
        with src.open(newline="", encoding="utf-8") as fh:
            for r in _csv.DictReader(fh):
                raw = (r.get("Date") or "").strip()
                if not raw:
                    continue
                try:
                    d = datetime.strptime(raw, "%m/%d/%Y")
                except ValueError:
                    continue
                rows.append((d, r))
    except Exception:
        return None
    if not rows:
        return None

    def ym(d: datetime) -> tuple[int, int]:
        return (d.year, d.month)

    this = ym(now)
    anchor = datetime(now.year, now.month, 1)
    prev = (anchor.year - 1, 12) if anchor.month == 1 else (anchor.year, anchor.month - 1)

    expenses_this: dict[str, float] = {}
    expenses_prev: dict[str, float] = {}
    income_this: dict[str, float] = {}
    by_cat: dict[str, Counter] = {}

    for d, r in rows:
        cur = (r.get("Currency") or "").strip()
        typ = (r.get("Type") or "").strip()
        amt = abs(pnum(r.get("Price")))
        ymk = ym(d)
        if typ == "Expense":
            if ymk == this:
                expenses_this[cur] = expenses_this.get(cur, 0.0) + amt
                cat = (r.get("Main Category") or "—").strip() or "—"
                by_cat.setdefault(cur, Counter())[cat] += amt
            elif ymk == prev:
                expenses_prev[cur] = expenses_prev.get(cur, 0.0) + amt
        elif typ == "Income" and ymk == this:
            income_this[cur] = income_this.get(cur, 0.0) + amt

    rows.sort(key=lambda t: t[0], reverse=True)
    latest: list[dict] = []
    for d, r in rows:
        if len(latest) >= 5:
            break
        if (r.get("Type") or "").strip() not in ("Expense", "Income"):
            continue
        latest.append({
            "date": d.strftime("%Y-%m-%d"),
            "type": (r.get("Type") or "").strip(),
            "category": (r.get("Main Category") or "").strip(),
            "name": (r.get("Name") or "").strip(),
            "store": (r.get("Store") or "").strip(),
            "amount": pnum(r.get("Price")),
            "currency": (r.get("Currency") or "").strip(),
        })

    days_elapsed = now.day
    days_in_month = calendar.monthrange(now.year, now.month)[1]
    ars_this = expenses_this.get("ARS", 0.0)
    ars_prev = expenses_prev.get("ARS", 0.0)
    run_rate = ars_this / days_elapsed if days_elapsed else 0.0
    projected = run_rate * days_in_month
    delta_pct = ((ars_this - ars_prev) / ars_prev * 100.0) if ars_prev else None

    top_cats: list[dict] = []
    for name, amt in (by_cat.get("ARS") or Counter()).most_common(5):
        top_cats.append({
            "name": name,
            "amount": amt,
            "share": (amt / ars_this) if ars_this else 0.0,
        })

    usd_this = expenses_this.get("USD", 0.0) + expenses_this.get("USDB", 0.0)
    usd_prev = expenses_prev.get("USD", 0.0) + expenses_prev.get("USDB", 0.0)

    payload = {
        "source_file": src.name,
        "source_mtime": datetime.fromtimestamp(mtime).isoformat(timespec="seconds"),
        "month_label": now.strftime("%Y-%m"),
        "days_elapsed": days_elapsed,
        "days_in_month": days_in_month,
        "ars": {
            "this_month": ars_this,
            "prev_month": ars_prev,
            "delta_pct": delta_pct,
            "run_rate_daily": run_rate,
            "projected": projected,
            "income": income_this.get("ARS", 0.0),
            "top_categories": top_cats,
        },
        "usd": {
            "this_month": usd_this,
            "prev_month": usd_prev,
        },
        "latest": latest,
    }
    _FINANCE_CACHE["key"] = key
    _FINANCE_CACHE["payload"] = payload
    return payload


# Resúmenes de tarjeta de crédito — el banco emite un `.xlsx` por ciclo
# (Santander Río exporta `Último resumen - <Marca> <Últimos4>.xlsx`) y el
# user lo deja caer en el mismo dir iCloud `/Finances` que los CSV de MOZE.
# Naming esperado: empieza con "Último resumen" (con o sin acento) — usamos
# globs case-insensitive para tolerar variaciones del banco.
#
# Shape parseado (un dict por xlsx), todas las claves opcionales:
#   { "brand": "Visa", "last4": "1059", "holder": "Fernando ...",
#     "closing_date": "2026-03-26", "due_date": "2026-04-08",
#     "next_closing_date": "2026-04-30", "next_due_date": "2026-05-08",
#     "total_ars": 549438.75, "total_usd": 98.93,
#     "minimum_ars": ..., "minimum_usd": ...,
#     "top_purchases": [{date, description, amount, currency}, ...],
#     "source_file": "Último resumen - Visa 1059.xlsx",
#     "source_mtime": "2026-04-26T19:11:42" }
#
# Cache compartido: clave = tuple ordenado de (path, mtime) de TODOS los
# xlsx → re-export de cualquier tarjeta invalida; agregar/quitar tarjetas
# también. Silent-fail (devuelve `[]`) si no hay xlsx, falta openpyxl, o
# todos los archivos fallan al parsear.
_CARDS_CACHE: dict = {"key": None, "payload": None}
# Lock para read-then-write atómico (audit 2026-04-26): pre-fix dos threads
# concurrentes podían ver pareja inconsistente (key nuevo + payload viejo)
# si uno leía mientras otro escribía. Lock simple — el path no recursa.
_CARDS_CACHE_LOCK = threading.Lock()

# Regex para extraer marca + últimos 4 del nombre de archivo o sheet name.
# Tolera "Visa Crédito terminada en 1059", "Visa 1059", "Mastercard 5234",
# "Amex 3456". Case-insensitive.
_CARD_BRAND_RE = re.compile(
    r"\b(Visa|Master(?:card)?|Amex|American\s*Express|Cabal|Maestro|Naranja)\b",
    re.IGNORECASE,
)
_CARD_LAST4_RE = re.compile(r"(\d{4})(?!\d)")


def _parse_ars_or_usd(raw: object) -> tuple[float | None, str | None]:
    """Parsea celdas tipo `$549.438,75`, `U$S98,93`, `-$926,15`, o numérico
    crudo (openpyxl puede devolver float si la celda tiene formato número).
    Retorna `(amount, currency)` donde currency ∈ {"ARS", "USD", None}.

    Heurística decimal: si el string tiene `,` lo asumimos formato ES
    (decimal coma, miles punto) y normalizamos. Si solo tiene `.` puede ser
    formato US (`24.99`) — solo strippeamos puntos como miles si hay 3
    dígitos exactos después, sino el punto es decimal.
    """
    if raw is None:
        return (None, None)
    if isinstance(raw, (int, float)):
        return (float(raw), None)
    s = str(raw).strip()
    if not s:
        return (None, None)
    cur: str | None = None
    if s.upper().startswith("U$S") or "U$S" in s.upper() or s.upper().startswith("USD"):
        cur = "USD"
    elif s.startswith("$") or "ARS" in s.upper():
        cur = "ARS"
    # Strippeamos símbolos no-numéricos.
    cleaned = re.sub(r"[^\d,.\-]", "", s)
    if "," in cleaned:
        # Formato ES: punto = miles, coma = decimal.
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif cleaned.count(".") > 1:
        # Múltiples puntos → todos son miles ES sin decimal explícito.
        cleaned = cleaned.replace(".", "")
    # Else: un solo punto → decimal US (`24.99`), dejarlo como está.
    try:
        return (float(cleaned), cur)
    except ValueError:
        return (None, cur)


def _parse_card_date(raw: object) -> str | None:
    """DD/MM/YYYY → ISO `YYYY-MM-DD`. None si no parsea. Acepta
    `datetime.date|datetime` (openpyxl normaliza fechas si la celda tiene
    formato fecha) y strings con prefijo (`"Cierre: 26/03/2026"`).
    """
    if raw is None:
        return None
    if hasattr(raw, "strftime"):
        try:
            return raw.strftime("%Y-%m-%d")
        except Exception:
            return None
    s = str(raw).strip()
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if not m:
        return None
    try:
        d = datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        return d.strftime("%Y-%m-%d")
    except ValueError:
        return None


def _parse_credit_card_xlsx(path: Path) -> dict | None:
    """Parsea un `Último resumen - <Marca> <Últimos4>.xlsx` del banco a un
    dict normalizado. None si openpyxl no está disponible o el xlsx no
    tiene la estructura esperada (sin Total a pagar reconocible).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return None

    rows: list[tuple] = []
    sheet_name = ""
    try:
        ws = wb.active
        sheet_name = ws.title or ""
        # Lectura completa: el archivo es <100 filas — read_only modo
        # streaming, sin riesgo de memoria.
        for row in ws.iter_rows(values_only=True):
            rows.append(row)
    except Exception:
        return None
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if not rows:
        return None

    # Identificar marca + últimos 4 — primero del sheet name, sino del
    # nombre de archivo, sino del primer row con "terminada en".
    brand = None
    last4 = None
    for source in (sheet_name, path.stem):
        m_brand = _CARD_BRAND_RE.search(source)
        m_last4 = _CARD_LAST4_RE.search(source)
        if m_brand and not brand:
            brand = m_brand.group(1).title()
            if brand.lower() in ("master", "mastercard"):
                brand = "Mastercard"
            elif brand.lower() in ("amex", "american express"):
                brand = "Amex"
        if m_last4 and not last4:
            last4 = m_last4.group(1)
        if brand and last4:
            break

    # Pasada por filas: extraer secciones por anchor text.
    holder = None
    closing_date = None
    due_date = None
    next_closing_date = None
    next_due_date = None
    total_ars = None
    total_usd = None
    minimum_ars = None
    minimum_usd = None
    top_purchases: list[dict] = []
    # "Otros conceptos": impuestos / IVA / retenciones que el banco cobra
    # encima de los consumos. Típicamente: Impuesto de sellos, Iibb
    # percep-sant 3%, Iva rg 4240 21%, Db.rg 5617 30%. Suma ~10% del
    # total a pagar. Sin esto la suma de consumos no cuadra con el
    # "Total a pagar" — gap ~$55.683 reportado por el user 2026-04-26.
    other_charges: list[dict] = []

    def _row_text(r: tuple) -> str:
        return " ".join(str(c) for c in r if c is not None).strip()

    n = len(rows)
    i = 0
    in_purchases_block = False
    in_payments_block = False
    in_other_charges_block = False
    # Trackeamos la última fecha vista en filas de movimiento para
    # heredarla en filas que vienen con col[0] vacía (multi-consumo
    # mismo día). Reset al cerrar el bloque para no contaminar entre
    # tarjetas / secciones.
    last_seen_purchase_date: str | None = None
    while i < n:
        row = rows[i]
        text = _row_text(row).lower()

        # Header: "Tarjeta <Marca> ... terminada en NNNN"
        if not last4 and "terminada en" in text:
            m = _CARD_LAST4_RE.search(text)
            if m:
                last4 = m.group(1)
            m_brand = _CARD_BRAND_RE.search(text)
            if m_brand and not brand:
                brand = m_brand.group(1).title()

        # "Fecha de cierre" / "Fecha de vencimiento" — fila siguiente las trae
        if "fecha de cierre" in text and "fecha de vencimiento" in text and i + 1 < n:
            nxt = rows[i + 1]
            closing_date = _parse_card_date(nxt[0] if len(nxt) > 0 else None)
            due_date = _parse_card_date(nxt[1] if len(nxt) > 1 else None)
            i += 2
            continue

        # "Total a pagar" — fila siguiente: ARS | USD
        if "total a pagar" in text and i + 1 < n:
            nxt = rows[i + 1]
            v0, c0 = _parse_ars_or_usd(nxt[0] if len(nxt) > 0 else None)
            v1, c1 = _parse_ars_or_usd(nxt[1] if len(nxt) > 1 else None)
            for v, c in ((v0, c0), (v1, c1)):
                if v is None:
                    continue
                if c == "USD":
                    total_usd = v
                else:
                    total_ars = v if total_ars is None else total_ars
            i += 2
            continue

        # "Mínimo a pagar"
        if "mínimo a pagar" in text or "minimo a pagar" in text:
            if i + 1 < n:
                nxt = rows[i + 1]
                v0, c0 = _parse_ars_or_usd(nxt[0] if len(nxt) > 0 else None)
                v1, c1 = _parse_ars_or_usd(nxt[1] if len(nxt) > 1 else None)
                for v, c in ((v0, c0), (v1, c1)):
                    if v is None:
                        continue
                    if c == "USD":
                        minimum_usd = v
                    else:
                        minimum_ars = v if minimum_ars is None else minimum_ars
                i += 2
                continue

        # "Próximo resumen" — la fila tiene "Cierre: DD/MM/YYYY" en col 2
        if "próximo resumen" in text or "proximo resumen" in text:
            # Las dos siguientes filas tienen "Cierre: ..." y "Vencimiento: ..."
            for j in range(i + 1, min(i + 4, n)):
                trow = _row_text(rows[j]).lower()
                if "cierre" in trow:
                    # col 1 = actual, col 2 = próximo
                    cell = rows[j][1] if len(rows[j]) > 1 else None
                    next_closing_date = _parse_card_date(cell)
                if "vencimiento" in trow:
                    cell = rows[j][1] if len(rows[j]) > 1 else None
                    next_due_date = _parse_card_date(cell)
            i += 1
            continue

        # Holder: "<Marca> Crédito terminada en NNNN" + col 2 con "(Titular)"
        if holder is None and len(row) > 1 and row[1] and "titular" in str(row[1]).lower():
            holder = re.sub(r"\s*\(Titular\)\s*$", "", str(row[1])).strip() or None

        # Bloques de movimientos: "Pago de tarjeta y devoluciones" (skipear) /
        # "Tarjeta de <holder>" (capturar movimientos hasta "Total de ...")
        if "pago de tarjeta y devoluciones" in text:
            in_payments_block = True
            in_purchases_block = False
        elif text.startswith("tarjeta de ") and "terminada en" in text:
            in_purchases_block = True
            in_payments_block = False
            # Extraer holder si aún no lo tenemos
            if holder is None:
                m = re.search(r"tarjeta de (.+?)\s*-", text, re.IGNORECASE)
                if m:
                    holder = m.group(1).strip().title()
        elif text.startswith("total de ") and ("terminada en" in text or "tarjeta" in text):
            in_purchases_block = False
            in_payments_block = False
            in_other_charges_block = False
        elif text.startswith("otros conceptos"):
            in_purchases_block = False
            in_payments_block = False
            in_other_charges_block = True
        elif text.startswith("aviso importante") or text.startswith("total a pagar"):
            # Cierra el bloque de otros conceptos — el footer legal ya
            # no aporta data parseable.
            in_other_charges_block = False
        elif in_other_charges_block and len(row) >= 2:
            # Fila de "Otros conceptos": (descripción, monto ARS, ...).
            # Ejemplos del banco (Visa 1059 — 2026-04-26):
            #   "Impuesto de sellos" — $548,75
            #   "Iibb percep-sant 3%( 7)" — $902,95
            #   "Iva rg 4240 21%( 30098,37)" — $6.320,65
            #   "Db.rg 5617 30% ( 136276,07 )" — $40.882,82
            desc = (str(row[0]).strip() if row[0] else "").strip()
            if desc and desc.lower() not in ("descripción", "descripcion"):
                amt_ars, _ = _parse_ars_or_usd(row[1] if len(row) > 1 else None)
                if amt_ars is not None and abs(amt_ars) > 0:
                    other_charges.append({
                        "description": desc,
                        "amount": abs(amt_ars),
                        "currency": "ARS",
                    })
        elif in_purchases_block and len(row) >= 5:
            # Fila de movimiento: (fecha, descripción, cuotas, comprobante, ARS, USD)
            # Algunas filas tienen fecha vacía (continuación del día anterior).
            # El banco ahorra espacio mostrando la fecha solo en el primer
            # consumo del día — los siguientes vienen con col[0]=None y
            # heredan implícitamente la fecha de la fila anterior. Sin
            # esta inheritance, el render quedaba con `?` para esas filas
            # (≈9 de 14 movimientos en el test 2026-04-26).
            desc = (str(row[1]).strip() if row[1] else "").strip()
            if desc and desc.lower() not in ("descripción", "descripcion"):
                amt_ars, _ = _parse_ars_or_usd(row[4] if len(row) > 4 else None)
                amt_usd, _ = _parse_ars_or_usd(row[5] if len(row) > 5 else None)
                amount = None
                currency = None
                if amt_ars is not None:
                    amount = abs(amt_ars)
                    currency = "ARS"
                elif amt_usd is not None:
                    amount = abs(amt_usd)
                    currency = "USD"
                if amount is not None and amount > 0:
                    parsed_date = _parse_card_date(row[0] if len(row) > 0 else None)
                    if parsed_date:
                        # Cell tiene fecha → la usamos y la guardamos
                        # como "última vista" para que la próxima fila
                        # sin fecha la herede.
                        last_seen_purchase_date = parsed_date
                    elif last_seen_purchase_date:
                        # Fecha vacía + ya vimos una fecha previa →
                        # heredamos. Asume orden cronológico (que es lo
                        # que el banco emite).
                        parsed_date = last_seen_purchase_date
                    top_purchases.append({
                        "date": parsed_date,
                        "description": desc,
                        "amount": amount,
                        "currency": currency,
                    })

        i += 1

    # Sin total ni fechas reconocibles → xlsx no es un resumen válido.
    if total_ars is None and total_usd is None and not closing_date and not due_date:
        return None

    # Movimientos completos por moneda, ordenados por monto descendente.
    # `all_purchases_ars/usd` = TODOS los movimientos del ciclo (para
    # respuestas de "detalle"). `top_purchases_ars/usd` = top 5/3
    # (para summary cortos). Mantener ambos para retrocompat con
    # `_render_cards_answer` y otros consumers.
    all_ars_purchases = sorted(
        (p for p in top_purchases if p["currency"] == "ARS"),
        key=lambda p: p["amount"],
        reverse=True,
    )
    all_usd_purchases = sorted(
        (p for p in top_purchases if p["currency"] == "USD"),
        key=lambda p: p["amount"],
        reverse=True,
    )

    try:
        mtime = path.stat().st_mtime
    except Exception:
        mtime = 0.0

    return {
        "brand": brand,
        "last4": last4,
        "holder": holder,
        "closing_date": closing_date,
        "due_date": due_date,
        "next_closing_date": next_closing_date,
        "next_due_date": next_due_date,
        "total_ars": total_ars,
        "total_usd": total_usd,
        "minimum_ars": minimum_ars,
        "minimum_usd": minimum_usd,
        "top_purchases_ars": all_ars_purchases[:5],
        "top_purchases_usd": all_usd_purchases[:3],
        "all_purchases_ars": all_ars_purchases,
        "all_purchases_usd": all_usd_purchases,
        # Otros conceptos = impuestos/IVA/retenciones. ~10% típico del
        # total a pagar. Sin esto, `sum(purchases) != total_ars`.
        "other_charges": other_charges,
        "other_charges_total_ars": (
            sum(c["amount"] for c in other_charges) if other_charges else 0
        ),
        "source_file": path.name,
        "source_mtime": datetime.fromtimestamp(mtime).isoformat(timespec="seconds") if mtime else None,
    }


def _fetch_credit_cards(now: datetime | None = None) -> list[dict]:  # noqa: ARG001
    """Lista de resúmenes de tarjeta parseados de los xlsx en `/Finances`.

    Devuelve `[]` si no hay xlsx, openpyxl no está, o todos fallan al
    parsear. Cache por (paths, mtimes) compartidos — re-export de
    cualquier xlsx invalida el cache; agregar/quitar archivos también.

    Ordenado por `due_date` ascendente (vencimientos próximos primero) con
    ítems sin fecha al final.
    """
    try:
        # Glob case-insensitive: el banco a veces usa "Último" (con acento)
        # y a veces "Ultimo". Usamos 2 globs explícitos para no depender de
        # case-fold del FS (HFS+/APFS son case-insensitive por default pero
        # vault sync remoto puede no serlo).
        seen: set[Path] = set()
        for pattern in ("Último resumen*.xlsx", "Ultimo resumen*.xlsx"):
            for p in _FINANCE_BACKUP_DIR.glob(pattern):
                seen.add(p)
        files = sorted(seen, key=lambda p: p.name)
    except Exception:
        return []
    if not files:
        return []

    # Cache key: tupla con (str(path), mtime) de cada archivo, ordenada.
    try:
        cache_key = tuple(sorted((str(p), p.stat().st_mtime) for p in files))
    except Exception:
        cache_key = None

    # Snapshot bajo lock — evita pareja inconsistente.
    if cache_key is not None:
        with _CARDS_CACHE_LOCK:
            if _CARDS_CACHE.get("key") == cache_key:
                return _CARDS_CACHE.get("payload") or []

    cards: list[dict] = []
    for p in files:
        parsed = _parse_credit_card_xlsx(p)
        if parsed:
            cards.append(parsed)

    # Orden: due_date ASC; sin fecha al final.
    cards.sort(key=lambda c: (c.get("due_date") is None, c.get("due_date") or ""))

    if cache_key is not None:
        # Audit 2026-04-26 BUG #6 web — TOCTOU fix: si otro thread escribió
        # una key MÁS RECIENTE mientras computábamos, NO sobrescribimos
        # (sería un retroceso). Gen counter implícito vía max(mtime).
        with _CARDS_CACHE_LOCK:
            _existing = _CARDS_CACHE.get("key")
            _our_max_mtime = max((m for _, m in cache_key), default=0.0)
            _existing_max_mtime = (
                max((m for _, m in _existing), default=0.0)
                if _existing else 0.0
            )
            if _existing is None or _our_max_mtime >= _existing_max_mtime:
                _CARDS_CACHE["key"] = cache_key
                _CARDS_CACHE["payload"] = cards
            # else: skip write — el cache tiene data más fresca.
    return cards


# Cold `_home_compute` is 30–40s (12-fetcher fan-out, slowest = pendientes
# at ~25s). The user must NEVER block on that. Strategy:
#   1. Background pre-warmer thread recomputes every BG_INTERVAL, populating
#      _HOME_STATE["payload"] under lock.
#   2. /api/home reads from _HOME_STATE — instant 2ms regardless of TTL.
#   3. Stale-while-revalidate: if payload age > SOFT_TTL, serve stale and
#      kick a one-shot refresh thread (no double-compute thanks to the
#      `computing` flag).
#   4. First request before the pre-warmer finishes blocks once (cold path),
#      then every subsequent request is instant.
import threading
from fastapi.responses import Response

_HOME_LOCK = threading.Lock()
# Cache serialized bytes too — `/api/home` then becomes a raw bytes write
# (no per-request 35KB json.dumps under the GIL).
_HOME_STATE: dict = {
    "payload": None,
    "body": None,
    "ts": 0.0,
    "computing": False,
}
_HOME_SOFT_TTL = 120.0      # serve cached without bg-refresh under this age
# Pre-warmer cadence — bumped 25s → 300s on 2026-04-17 after chat latency
# blew up to 30-160s. Each cycle fans out 14 top-level + 9 sub-fetchers and can eat
# 6-27s of ollama + MPS + disk, which starves concurrent /api/chat requests
# waiting for the same ollama daemon. 5-minute refresh is plenty for a
# dashboard — the home page SWRs on visit anyway.
_HOME_BG_INTERVAL = 300.0
# Live chat-in-flight counter. While > 0, the prewarmer skips its cycle so
# chat gets exclusive use of ollama + the reranker. Simple counter (not a
# lock) because multiple concurrent chats are fine — they serialise on
# ollama anyway, and we only care about "nobody chatting" vs "someone is".
_CHAT_INFLIGHT = 0
_CHAT_INFLIGHT_LOCK = threading.Lock()
# Rolling window of recent home-compute totals (seconds). Powers the
# `degraded` SSE event: when an in-flight stream exceeds 2× the median
# of this window, we emit a one-shot `degraded` event with a probable
# cause (ollama state / memory pressure) so the UI can surface "esto
# está más lento de lo normal" instead of just spinning silently.
# Bounded to last 20 entries (≈10min at the 30s prewarmer cadence + 60s
# auto-refresh) — old enough to capture day-to-day variance, recent
# enough not to anchor on stale baselines after a restart.
_HOME_COMPUTE_HISTORY: list[float] = []
_HOME_COMPUTE_HISTORY_LOCK = threading.Lock()
_HOME_COMPUTE_HISTORY_MAX = 20
# Floor below which we don't bother computing degraded thresholds —
# avoids spamming "degraded" early on when we only have 1-2 samples
# from a fast warm cache (≈1.5s) and 2× would fire on any cold-ish run.
_HOME_COMPUTE_DEGRADED_FLOOR = 8.0


def _record_home_compute_total(
    elapsed_s: float,
    *,
    regenerate: bool = False,
    degraded: bool = False,
    degraded_cause: str | None = None,
) -> None:
    """Append a sample to the rolling window AND persist to SQL.

    Persisting to `rag_home_compute_metrics` lets the degraded
    threshold survive a web service restart — pre-persistence, the
    in-memory `_HOME_COMPUTE_HISTORY` reset on every kickstart, so
    the first 3 streams after restart fell back to the
    `_HOME_COMPUTE_DEGRADED_FLOOR` (8s) which is conservative and
    typically too high for warm-cache machines. SQL persistence makes
    the threshold reflect genuine recent baselines.

    Outliers (≤0 or >600s) are dropped — those would be measurement
    bugs, not real runs.
    """
    if elapsed_s <= 0 or elapsed_s > 600:
        return
    with _HOME_COMPUTE_HISTORY_LOCK:
        _HOME_COMPUTE_HISTORY.append(elapsed_s)
        if len(_HOME_COMPUTE_HISTORY) > _HOME_COMPUTE_HISTORY_MAX:
            _HOME_COMPUTE_HISTORY.pop(0)

    # Persist sync — single-row INSERT to telemetry.db is sub-ms warm
    # and benefits from `busy_timeout=30000` to absorb transient locks.
    # Going async via daemon thread caused races with conftest fixture
    # teardown in tests (`rag.DB_PATH` got restored before the daemon
    # could open its conn). Silent-fail per existing convention.
    try:
        from rag import _ragvec_state_conn
        with _ragvec_state_conn() as conn:
            conn.execute(
                "INSERT INTO rag_home_compute_metrics "
                "(ts, elapsed_s, regenerate, degraded, degraded_cause) "
                "VALUES (?, ?, ?, ?, ?)",
                (
                    datetime.now().isoformat(timespec="seconds"),
                    float(elapsed_s),
                    1 if regenerate else 0,
                    1 if degraded else 0,
                    degraded_cause,
                ),
            )
            conn.commit()
    except Exception:
        pass  # silent-fail per convention


def _hydrate_home_compute_history_from_sql() -> None:
    """Backfill `_HOME_COMPUTE_HISTORY` with the most recent samples
    from SQL on web service startup. Called from the lifespan hook.

    Without this, the threshold `max(floor, median × 2)` falls back
    to the floor every restart. Hydrating means the user sees
    consistent threshold behaviour across kickstarts.
    """
    try:
        from rag import _ragvec_state_conn
        with _ragvec_state_conn() as conn:
            rows = conn.execute(
                "SELECT elapsed_s FROM rag_home_compute_metrics "
                "ORDER BY ts DESC LIMIT ?",
                (_HOME_COMPUTE_HISTORY_MAX,),
            ).fetchall()
    except Exception:
        return
    if not rows:
        return
    samples = [float(r[0]) for r in reversed(rows)]
    with _HOME_COMPUTE_HISTORY_LOCK:
        _HOME_COMPUTE_HISTORY[:] = samples


def _home_compute_degraded_threshold() -> float:
    """Median × 2 over the rolling window, clamped to the floor.

    Median over mean → resistant to a single anomalous run. Less than
    3 samples → return floor so we don't fire prematurely after restart.
    """
    with _HOME_COMPUTE_HISTORY_LOCK:
        snap = list(_HOME_COMPUTE_HISTORY)
    if len(snap) < 3:
        return _HOME_COMPUTE_DEGRADED_FLOOR
    snap.sort()
    median = snap[len(snap) // 2]
    return max(_HOME_COMPUTE_DEGRADED_FLOOR, median * 2.0)


def _diagnose_home_slowdown() -> dict:
    """Best-effort sniff of *why* a home-compute is slow. Returns a
    dict with `cause: str` and `details: dict`.

    Probes are cheap — a 500ms ollama HTTP timeout, a vm_stat read, a
    reranker-load-time check. We never block the SSE stream more than
    ~1s on these.
    """
    import urllib.error as _ue
    import urllib.request as _ur

    details: dict = {}

    # 1. Ollama health — slow `/api/tags` typically means model swap or
    #    daemon unloading something big. 500ms budget covers warm and
    #    flags genuine wedges.
    ollama_state = "ok"
    ollama_t0 = time.time()
    try:
        req = _ur.Request("http://localhost:11434/api/tags", method="GET")
        with _ur.urlopen(req, timeout=0.5) as resp:
            resp.read()
    except _ue.URLError:
        ollama_state = "unreachable"
    except TimeoutError:
        ollama_state = "slow"
    except Exception:
        ollama_state = "error"
    details["ollama_ms"] = round((time.time() - ollama_t0) * 1000.0, 1)
    details["ollama_state"] = ollama_state

    # 2. Memory pressure — wired+active+compressed pct on macOS. >85%
    #    aligns with the watchdog threshold that evicts the chat model.
    try:
        import rag as _rag
        used_pct = _rag._system_memory_used_pct()
        if used_pct is not None:
            details["mem_used_pct"] = round(used_pct, 1)
    except Exception:
        pass

    # Pick the dominant cause for the UI label.
    if ollama_state in ("unreachable", "slow", "error"):
        cause = f"ollama_{ollama_state}"
    elif details.get("mem_used_pct", 0) >= 85:
        cause = "memory_pressure"
    else:
        cause = "unknown"
    return {"cause": cause, "details": details}
# Persist cache to disk so service restarts don't force a 30-45s cold
# compute on the next user visit. Disk payload is served immediately
# (stale OK — SWR refreshes in bg within seconds).
_HOME_CACHE_PATH = Path.home() / ".local/share/obsidian-rag/home_cache.json"
_HOME_DISK_TTL = 6 * 3600   # ignore disk cache older than this (stale-but-useful window)

_WARMING_BODY = json.dumps({
    "warming": True,
    "today": {"narrative": "", "narrative_source": "warming",
              "brief_path": None,
              "counts": {"recent_notes": 0, "inbox_today": 0, "todos": 0,
                         "new_contradictions": 0, "low_conf_queries": 0,
                         "total": 0},
              "evidence": {}},
    "urgent": [], "signals": {}, "tomorrow_calendar": [],
    "weather_forecast": None,
}).encode()


def _home_refresh(regenerate: bool = False) -> None:
    """Compute + publish under lock; never run two computes in parallel.
    Silent-fail keeps last good payload visible if a fetcher dependency
    (icalBuddy, whisper, ollama) is momentarily down.
    """
    with _HOME_LOCK:
        if _HOME_STATE["computing"]:
            return
        _HOME_STATE["computing"] = True
    t0 = time.time()
    try:
        payload = _home_compute(regenerate)
        body = json.dumps(payload, default=str).encode()
    except Exception:
        return
    finally:
        with _HOME_LOCK:
            _HOME_STATE["computing"] = False
    now = time.time()
    _HOME_STATE["payload"] = payload
    _HOME_STATE["body"] = body
    _HOME_STATE["ts"] = now
    elapsed = now - t0
    _record_home_compute_total(elapsed)
    print(f"[home-refresh] compute={elapsed:.2f}s regen={regenerate}", file=sys.stderr)
    _persist_home_cache(body, now)


def _persist_home_cache(body: bytes, ts: float) -> None:
    """Write the serialized payload + timestamp to disk so a service
    restart can serve it immediately (instead of forcing a 30-45s cold
    compute on the next visit). Atomic rename so a crash mid-write never
    leaves a half-file. Silent-fail — the cache is optional.
    """
    from contextlib import suppress

    with suppress(Exception):
        _HOME_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        tmp = _HOME_CACHE_PATH.with_suffix(".json.tmp")
        tmp.write_bytes(b'{"ts":' + str(ts).encode()
                        + b',"body":' + body + b'}')
        tmp.replace(_HOME_CACHE_PATH)


def _load_home_cache() -> None:
    """Hydrate `_HOME_STATE` from disk on server startup. If the file is
    present and younger than `_HOME_DISK_TTL`, the first user visit sees
    stale-but-useful data immediately; SWR refreshes in the background.
    """
    if not _HOME_CACHE_PATH.is_file():
        return
    try:
        raw = _HOME_CACHE_PATH.read_bytes()
        data = json.loads(raw)
        ts = float(data.get("ts", 0))
        body_obj = data.get("body")
        if not ts or body_obj is None:
            return
        age = time.time() - ts
        if age > _HOME_DISK_TTL:
            return
        body_bytes = json.dumps(body_obj, default=str).encode()
        _HOME_STATE["body"] = body_bytes
        _HOME_STATE["payload"] = body_obj
        _HOME_STATE["ts"] = ts
        print(f"[home-cache] hydrated from disk age={age:.0f}s size={len(body_bytes)}B",
              file=sys.stderr)
    except Exception as exc:
        print(f"[home-cache] load failed: {exc}", file=sys.stderr)


@app.get("/api/home")
def home_api(regenerate: bool = False) -> Response:
    _ensure_home_prewarmer()
    # Explicit user action — recompute LLM brief synchronously.
    if regenerate:
        _home_refresh(regenerate=True)
        body = _HOME_STATE["body"] or _WARMING_BODY
        return Response(content=body, media_type="application/json")

    body = _HOME_STATE["body"]
    age = time.time() - _HOME_STATE["ts"] if body else float("inf")

    if body and age < _HOME_SOFT_TTL:
        return Response(content=body, media_type="application/json")
    if body:
        # SWR: return stale immediately, refresh in bg
        threading.Thread(
            target=_home_refresh, name="home-swr", daemon=True,
        ).start()
        return Response(content=body, media_type="application/json")
    # First call ever (pre-warmer still racing). Return placeholder so
    # the frontend skeleton stays up; auto-refresh picks up real payload.
    threading.Thread(
        target=_home_refresh, name="home-cold", daemon=True,
    ).start()
    return Response(content=_WARMING_BODY, media_type="application/json")


@app.get("/api/home/stream")
async def home_stream(request: Request, regenerate: bool = False) -> StreamingResponse:
    """SSE companion to `/api/home`. Streams stage events as the 14-fetcher
    fan-out runs so the UI can show what's happening live (fetcher name +
    elapsed_ms on done, "timeout" / "error" on bail). Final `done` event
    carries the full payload (same shape as `/api/home`).

    Events:
    - `stage` ({stage, status, elapsed_ms, error?})
        status ∈ {"start","done","error","timeout"}
    - `done` ({payload})
    - `error` ({message}) — only on hard failure (today_evidence aborts).

    Disconnect handling: if the client closes the EventSource mid-stream
    (e.g. user navigates away), the generator detects it via
    `request.is_disconnected()` and (a) signals the worker thread to stop
    submitting new sub-fetchers via `cancel_event`, (b) bails out so the
    server stops yielding into a dead pipe. In-flight fetchers can't be
    interrupted (Python threads aren't preemptable), but the cancellation
    short-circuits any sub-stages that haven't started — which matters
    most for `signals` (9 internal fetchers).

    Bug fix 2026-04-27: apply the same per-IP slot cap used by
    dashboard_stream / system_memory_stream / system_cpu_stream. Without
    the cap a browser tab that keeps reconnecting (background EventSource
    auto-reconnect) can accumulate connections, each launching a full
    14-fetcher `_home_compute` fan-out with concurrent threads.
    """
    # Cap por IP — mismo patrón que dashboard_stream (audit 2026-04-25 R2).
    client_ip: str | None = None
    client_ip = request.client.host if request.client else "unknown"
    if not _sse_acquire_slot(client_ip):
        raise HTTPException(
            status_code=429,
            detail=f"too many concurrent streams (max {_SSE_MAX_PER_IP} per IP)",
        )

    import asyncio
    import queue as _queue
    from contextlib import suppress

    _ensure_home_prewarmer()
    q: _queue.Queue = _queue.Queue()
    SENTINEL = object()
    cancel_event = threading.Event()

    def _on_progress(stage: str, status: str, elapsed_ms: float, err: str | None) -> None:
        evt = {"stage": stage, "status": status, "elapsed_ms": round(elapsed_ms, 1)}
        if err:
            evt["error"] = err
        q.put(("stage", evt))

    def _runner() -> None:
        try:
            payload = _home_compute(
                regenerate=regenerate,
                progress=_on_progress,
                cancel_event=cancel_event,
            )
            q.put(("done", payload))
        except HTTPException as exc:
            q.put(("error", {"message": str(exc.detail)}))
        except Exception as exc:
            q.put(("error", {"message": str(exc)}))
        finally:
            q.put((SENTINEL, None))

    async def _gen():
        # Heartbeat so reverse-proxies / Safari don't kill the connection
        # while the slowest fetcher (signals, ~25s cold) is still running.
        # 14s is comfortably under the typical 30s idle-disconnect window.
        HEARTBEAT_S = 14.0
        # Disconnect probe — every iteration checks if the client is
        # still there. 1s is a balance between snappy cancel and
        # overhead (each probe is ~50µs).
        DISCONNECT_PROBE_S = 1.0
        # Hard cap on the SSE stream itself (not on the underlying
        # `_home_compute` worker — that one continues to use its own
        # per-fetcher budgets and writes to disk via the SWR path).
        #
        # 30s is empirically generous — `_home_compute` typical cold
        # is 6-10s, P95 is ~15s, the worst-recorded outlier in 7d was
        # 21.5s during an ollama eviction event. 30s gives ~40% margin
        # over that outlier.
        #
        # Pre-2026-04-24 cap was 90s, which was the SUM of nominal
        # per-fetcher budgets — but those run in parallel, not serial.
        # The 90s cap meant a wedged stream tied up an asyncio task +
        # one executor thread + a worker thread for 1.5min, hurting
        # the next visitor when several stale clients piled up
        # (browser background tab keeps the EventSource alive).
        # 30s frees those resources 3× faster on a wedge.
        HARD_CAP_S = 30.0

        worker = threading.Thread(target=_runner, name="home-stream", daemon=True)
        worker.start()

        # Initial event so the UI knows the stream is alive even before
        # the first fetcher finishes (today_evidence cold path = 4-30s).
        # Top-level stages + the 9 sub-fetchers of `signals` (which
        # itself is the dominant bottleneck on cold start). The UI
        # uses this list to seed placeholder chips so users see the
        # full surface from t=0.
        yield _sse("hello", {
            "stages": [
                "today", "signals", "tomorrow", "forecast",
                "pagerank", "chrome", "eval", "followup",
                "drive", "wa_unreplied", "bookmarks", "vaults",
                "finance", "cards", "youtube",
            ],
            "substages": {
                "signals": [
                    "signals.mail_unread", "signals.reminders",
                    "signals.calendar", "signals.whatsapp",
                    "signals.weather", "signals.gmail",
                    "signals.loops", "signals.contradictions",
                    "signals.low_conf",
                ],
            },
        })

        t0 = time.time()
        last_event = t0
        finished = False
        loop = asyncio.get_running_loop()
        # Degraded watchdog: emit a one-shot `degraded` event when the
        # in-flight stream exceeds 2× the rolling median (clamped to
        # _HOME_COMPUTE_DEGRADED_FLOOR). The probe runs `_diagnose_home_slowdown`
        # (cheap: 500ms ollama HEAD + vm_stat read) so the UI gets a
        # probable cause label instead of just "lento".
        degraded_threshold = _home_compute_degraded_threshold()
        degraded_emitted = False

        try:
            while not finished:
                elapsed = time.time() - t0
                if elapsed > HARD_CAP_S:
                    yield _sse("error", {
                        "message": f"home-compute excedió {HARD_CAP_S:.0f}s",
                    })
                    cancel_event.set()
                    break

                # One-shot degraded probe — fires once when we cross
                # threshold and haven't seen `done` yet. The probe runs
                # in the executor too (urllib + vm_stat are blocking).
                if not degraded_emitted and elapsed >= degraded_threshold:
                    diag = await loop.run_in_executor(None, _diagnose_home_slowdown)
                    yield _sse("degraded", {
                        "elapsed_ms": round(elapsed * 1000.0, 1),
                        "threshold_ms": round(degraded_threshold * 1000.0, 1),
                        **diag,
                    })
                    degraded_emitted = True
                    last_event = time.time()

                # Disconnect probe before blocking on the queue.
                with suppress(Exception):
                    if await request.is_disconnected():
                        cancel_event.set()
                        print("[home-stream] client disconnected, cancelling",
                              file=sys.stderr)
                        break

                # `queue.get` is sync + blocking — run in a thread so we
                # don't starve the event loop (which needs to serve
                # other requests + run is_disconnected probes).
                try:
                    kind, payload = await asyncio.wait_for(
                        loop.run_in_executor(
                            None, lambda: q.get(timeout=DISCONNECT_PROBE_S),
                        ),
                        timeout=DISCONNECT_PROBE_S + 0.5,
                    )
                except (_queue.Empty, asyncio.TimeoutError):
                    # No event in DISCONNECT_PROBE_S — emit heartbeat
                    # if we're approaching the proxy idle window.
                    if time.time() - last_event >= HEARTBEAT_S:
                        yield ": keepalive\n\n"
                        last_event = time.time()
                    continue

                last_event = time.time()
                if kind is SENTINEL:
                    finished = True
                    break
                yield _sse(kind, payload)
                if kind == "done":
                    # Persist + cache so the next /api/home is instant.
                    # Also feed the rolling history so the next stream's
                    # degraded threshold reflects this run.
                    with suppress(Exception):
                        body = json.dumps(payload, default=str).encode()
                        now = time.time()
                        _HOME_STATE["payload"] = payload
                        _HOME_STATE["body"] = body
                        _HOME_STATE["ts"] = now
                        _persist_home_cache(body, now)
                    _record_home_compute_total(time.time() - t0)
                    finished = True
                    break
                if kind == "error":
                    finished = True
                    break
        finally:
            # Always set the cancel event so a wedged worker doesn't
            # keep submitting new sub-fetchers behind our back. Already-
            # running fetchers will finish naturally (warming caches)
            # but won't queue more work.
            cancel_event.set()
            # Bug fix 2026-04-27: liberar el slot SSE al cerrar el stream
            # (limpio, cancel, o timeout). client_ip nunca es None aquí
            # porque el gate de arriba lo asigna siempre antes de entrar.
            if client_ip is not None:
                _sse_release_slot(client_ip)

    return StreamingResponse(
        _gen(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache, no-transform",
            "X-Accel-Buffering": "no",  # disable nginx-style proxy buffering
            "Connection": "keep-alive",
        },
    )


_HOME_PREWARMER_STARTED = False

# Prewarmer flipped a ON por default tras 2026-04-24 (cancel-on-disconnect +
# `_CHAT_INFLIGHT` skip-cycle ya en su lugar). Pre-flip era opt-in porque
# el fan-out (14 channel fetchers, varios pegando ollama) podía starvear
# `/api/chat` mid-cycle. La protección actual:
#
#   1. Skip-cycle: si `_CHAT_INFLIGHT > 0`, el loop duerme 10s y reintenta.
#      Un chat-en-curso garantiza que el prewarmer NO arranca un compute
#      nuevo encima.
#   2. Cancel-on-disconnect: si un cliente del SSE cierra la conexión,
#      `cancel_event` propaga `pool.shutdown(cancel_futures=True)` —
#      las pendientes se descartan, no se acumulan.
#   3. Si la prewarmer-cycle SÍ está corriendo cuando llega un chat,
#      ollama serializa los requests vía su propio scheduler (no hay
#      doble carga del chat-model — el prewarmer no lo toca, solo el
#      chat-prewarmer tiene contacto con el chat-model).
#
# Beneficio: visitas post-restart instantáneas en vez de pagar 6-10s
# cold compute la primera vez. Para deshabilitar (ej. perf debug):
# `OBSIDIAN_RAG_HOME_PREWARM=0` en el plist o en el shell del proceso.
_HOME_PREWARM_ENABLED = os.environ.get("OBSIDIAN_RAG_HOME_PREWARM", "1") not in ("0", "false", "no")

def _ensure_home_prewarmer() -> None:
    """Start the pre-warmer loop iff opt-in. Called from startup and from
    /api/home. Idempotent.
    """
    global _HOME_PREWARMER_STARTED
    if _HOME_PREWARMER_STARTED or not _HOME_PREWARM_ENABLED:
        _HOME_PREWARMER_STARTED = True  # poison so we don't retry the check
        return
    _HOME_PREWARMER_STARTED = True

    def loop() -> None:
        while True:
            try:
                # Skip the cycle while a chat request is in flight — each
                # home-compute fans out 14 top-level fetchers + 9 sub-
                # fetchers inside `signals` (some hitting ollama for
                # helper-model judgments) and would queue behind or evict
                # the chat model, turning a 3s chat into 60s+.
                if _CHAT_INFLIGHT > 0:
                    time.sleep(10)
                    continue
                _home_refresh(regenerate=False)
            except Exception:
                pass
            time.sleep(_HOME_BG_INTERVAL)

    threading.Thread(target=loop, name="home-prewarmer", daemon=True).start()


# Chat-model prewarmer. Root cause: ollama evicts the /chat model from VRAM
# between requests despite keep_alive=-1 when memory pressure builds up
# (observed: qwen2.5:7b dropped after ~5min idle while qwen2.5:3b + bge-m3
# + reranker stayed resident). Each cold reload = 4-7s prefill penalty on
# the next /api/chat, turning warm 2.9s into 10s+.
#
# Fix: every _CHAT_PREWARM_INTERVAL seconds, send a 1-token ping to the
# resolved chat model with keep_alive=-1. This re-pins the model in VRAM
# even if ollama decided to evict during idle. Skipped while a real chat
# is in flight (the chat itself keeps the model warm — no need to duplicate).
#
# Cost: ~100-200ms of ollama compute per cycle, ~5GB VRAM pinned continuously.
# Safe on 36GB unified memory with command-r NOT also resident.
_CHAT_PREWARM_INTERVAL = int(os.environ.get("OBSIDIAN_RAG_CHAT_PREWARM_INTERVAL", "240"))
_CHAT_PREWARMER_STARTED = False


def _ensure_chat_model_prewarmer() -> None:
    """Start the chat-model prewarm loop once. Idempotent."""
    global _CHAT_PREWARMER_STARTED
    if _CHAT_PREWARMER_STARTED:
        return
    _CHAT_PREWARMER_STARTED = True

    def loop() -> None:
        # First cycle runs after a short startup delay so the app finishes
        # booting (corpus cache, home hydration) before we pin the LLM.
        # Subsequent cycles run every _CHAT_PREWARM_INTERVAL.
        time.sleep(15)
        while True:
            try:
                if _CHAT_INFLIGHT > 0:
                    time.sleep(15)
                    continue
                model = _resolve_web_chat_model()
                # Tiny ping — 1 token is enough to re-pin KV cache + model weights.
                # keep_alive se resuelve via chat_keep_alive(model): -1 (forever)
                # para modelos chicos, "20m" para grandes (guard 2026-04-21 post
                # Mac-freeze regression). El prewarm interval (~240s) queda por
                # debajo del clamp de 20m, así que modelos grandes siguen warm
                # entre pings sin pinearse wired forever.
                # Uses the bounded streaming client so a stuck-load daemon can't
                # wedge this thread forever and mask the next real request.
                _OLLAMA_STREAM_CLIENT.chat(
                    model=model,
                    messages=[
                        {"role": "system", "content": _WEB_SYSTEM_PROMPT},
                        {"role": "user", "content": "."},
                    ],
                    options={"num_predict": 1, "num_ctx": _WEB_CHAT_NUM_CTX,
                             "temperature": 0, "seed": 42},
                    stream=False,
                    think=False,   # match the probe + main chat path
                    keep_alive=chat_keep_alive(model),
                )
                print(f"[chat-prewarm] {model} pinned", flush=True)
            except Exception as exc:
                # Silent fail: ollama down, model not loaded, network blip.
                # Next cycle retries. Never crash the daemon thread.
                print(f"[chat-prewarm] skipped: {exc}", flush=True)
            time.sleep(_CHAT_PREWARM_INTERVAL)

    threading.Thread(target=loop, name="chat-model-prewarmer", daemon=True).start()


_RERANKER_PREWARMER_STARTED = False


def _ensure_reranker_prewarmer() -> None:
    """Load the cross-encoder once in a background thread. Idempotent.

    Fires unconditionally — covers launches without RAG_RERANKER_NEVER_UNLOAD=1
    where _do_warmup skips the reranker. Eliminates the 9s cold-load hit on
    the first retrieve after startup.
    """
    global _RERANKER_PREWARMER_STARTED
    if _RERANKER_PREWARMER_STARTED:
        return
    _RERANKER_PREWARMER_STARTED = True

    def _load() -> None:
        try:
            from rag import get_reranker as _get_rr
            _get_rr()
            print("[prewarm] reranker loaded", flush=True)
        except Exception as exc:
            print(f"[prewarm] reranker skipped: {exc}", flush=True)

    threading.Thread(target=_load, name="reranker-prewarmer", daemon=True).start()


_CORPUS_PREWARMER_STARTED = False


def _ensure_corpus_prewarmer() -> None:
    """Warm the BM25 corpus cache in a background thread. Idempotent.

    _load_corpus is O(1) on warm re-runs (cached), so safe to call even if
    _do_warmup already triggered it for the same vault.
    """
    global _CORPUS_PREWARMER_STARTED
    if _CORPUS_PREWARMER_STARTED:
        return
    _CORPUS_PREWARMER_STARTED = True

    def _load() -> None:
        try:
            from rag import get_db_for
            for _name, path in resolve_vault_paths(None):
                try:
                    col = get_db_for(path)
                    if col.count():
                        _load_corpus(col)
                        print(f"[prewarm] corpus warm for {path}", flush=True)
                except Exception:
                    pass
                break
        except Exception as exc:
            print(f"[prewarm] corpus skipped: {exc}", flush=True)

    threading.Thread(target=_load, name="corpus-prewarmer", daemon=True).start()


# ── Response cache ──────────────────────────────────────────────────────
# LRU de respuestas completas de /api/chat para queries exactas repetidas.
# Cuando el mismo user query llega dentro del TTL y el vault no cambió,
# servimos la respuesta cacheada como SSE replay (<100ms wall) en lugar de
# re-correr retrieve + LLM (2-3s).
#
# Cache key: sha256(question|vault_scope|chat_model|vault_chunks_count)[:16].
# Incluir el chunks count efectivamente invalida el cache cuando el vault
# gana/pierde notas (count cambia → key cambia). TTL secundario para
# invalidar respuestas sin cambio de vault pero con información que puede
# haber envejecido (ej. `rag tune` refinó pesos → retrieval diferente).
#
# Pre-2026-04-25 era 300s (5min), bumped a 24h tras audit telemétrico que
# mostró 3% hit rate (28 entries / 919 queries 7d) vs el 15-25% esperado:
# el TTL corto tiraba la mayoría de respuestas antes de la siguiente
# query equivalente. El vault_chunks count en la key ya invalida cuando
# el contenido cambia, así que el TTL solo defiende contra "data
# externa" (calendar, gmail, WA) — pero esos paths ya tienen short-
# circuits separados (`_wa_in_query`, propose-intent, source-specific).
# Override via env si causa problemas: `RAG_WEB_CHAT_CACHE_TTL=300` para
# rollback al valor previo.
#
# No cacheamos cuando:
#   - hay history (follow-ups dependen del turno previo, el key no refleja eso)
#   - el response fue vacío/error
#   - _wa_in_query matcheó (WA data change rápidamente, datos frescos importan)
_CHAT_CACHE: "OrderedDict[str, dict]" = __import__("collections").OrderedDict()
_CHAT_CACHE_MAX = 100
_CHAT_CACHE_TTL = float(os.environ.get("RAG_WEB_CHAT_CACHE_TTL", "86400"))  # default 24h, override via env
_CHAT_CACHE_LOCK = threading.Lock()


def _chat_cache_key(question: str, vault_scope: str, model: str, vault_chunks: int) -> str:
    raw = f"{question.strip().lower()}|{vault_scope}|{model}|{vault_chunks}"
    return __import__("hashlib").sha256(raw.encode()).hexdigest()[:16]


def _chat_cache_get(key: str) -> dict | None:
    with _CHAT_CACHE_LOCK:
        entry = _CHAT_CACHE.get(key)
        if not entry:
            return None
        if time.time() - entry["ts"] > _CHAT_CACHE_TTL:
            del _CHAT_CACHE[key]
            return None
        _CHAT_CACHE.move_to_end(key)
        return dict(entry)


def _chat_cache_put(key: str, payload: dict) -> None:
    with _CHAT_CACHE_LOCK:
        _CHAT_CACHE[key] = {"ts": time.time(), **payload}
        _CHAT_CACHE.move_to_end(key)
        while len(_CHAT_CACHE) > _CHAT_CACHE_MAX:
            _CHAT_CACHE.popitem(last=False)


def _home_compute(
    regenerate: bool = False,
    progress: "Callable[[str, str, float, str | None], None] | None" = None,
    cancel_event: "threading.Event | None" = None,
) -> dict:
    """Centralizer — aggregates every information channel the user cares
    about into one JSON payload. Powers the home page.

    Channels:
    - today narrative (rag today) + vault evidence for the day
    - live signals (reminders, Apple Mail, Gmail, WhatsApp, calendar, weather)
    - derived signals (open loops, contradictions, low-confidence queries)
    - tomorrow agenda + multi-day weather forecast

    Per-channel silent-fail: if one source errors (e.g. Gmail OAuth expired,
    icalBuddy not installed), that key is missing/empty but the rest renders.
    `regenerate=true` forces a fresh LLM narrative (~10s); default reuses the
    cached brief from `04-Archive/99-obsidian-system/99-AI/reviews/<date>-evening.md` if present.

    `progress` (optional): callback invoked as
    `progress(stage, status, elapsed_ms, error_message)` where status is one
    of "start" | "done" | "error" | "timeout". Used by the SSE stream
    endpoint to surface live fetcher state to the UI.

    `cancel_event` (optional): if set mid-compute (e.g. SSE client
    disconnected), abandon waiting on remaining futures and shut the
    pool down with `cancel_futures=True`. Already-running fetchers
    finish naturally to warm caches; pending ones are dropped.
    """
    from concurrent.futures import ThreadPoolExecutor
    from contextlib import suppress

    now = datetime.now()
    date_label = now.strftime("%Y-%m-%d")
    col = get_db()

    # ── Fan out: vault evidence + live channels + forecast run in parallel.
    # `_pendientes_collect` internally fans out its 9 fetchers too, so the
    # critical path across the whole page is max(today_evidence, slowest
    # single fetcher, tomorrow_calendar, weather_forecast) instead of the sum.
    # Don't use `with ThreadPoolExecutor(...) as pool:` — its __exit__ calls
    # `shutdown(wait=True)` and blocks the whole endpoint until every future
    # finishes, defeating the per-future timeouts below. We submit, collect
    # with timeouts, and then `shutdown(wait=False)` to return fast while
    # any straggler (e.g. cold `_fetch_followup_aging` doing LLM-judge per
    # loop) keeps running in the background to warm its own cache.
    pool = ThreadPoolExecutor(max_workers=14, thread_name_prefix="home")
    timings: dict[str, float] = {}
    t_submit = time.time()

    def _timed(name: str, fn, *args):
        t0 = time.time()
        if progress is not None:
            with suppress(Exception):
                progress(name, "start", 0.0, None)
        try:
            result = fn(*args)
            elapsed_ms = (time.time() - t0) * 1000.0
            timings[name] = (time.time() - t0)
            if progress is not None:
                with suppress(Exception):
                    progress(name, "done", elapsed_ms, None)
            return result
        except Exception as exc:
            elapsed_ms = (time.time() - t0) * 1000.0
            timings[name] = (time.time() - t0)
            if progress is not None:
                with suppress(Exception):
                    progress(name, "error", elapsed_ms, str(exc))
            raise

    try:
        fut_today       = pool.submit(
            _timed, "today", _collect_today_evidence,
            now, VAULT_PATH, LOG_PATH, CONTRADICTION_LOG_PATH,
        )
        # Signals fans out 9 sub-fetchers internally — surface them as
        # `signals.<name>` sub-stages so the UI can pinpoint which inner
        # fetcher is the actual bottleneck (cold gmail / whatsapp /
        # mail_unread typically dominate the 5-7s of `signals` total).
        def _signals_progress(
            name: str, status: str, elapsed_ms: float, err: str | None,
        ) -> None:
            if progress is not None:
                with suppress(Exception):
                    progress(f"signals.{name}", status, elapsed_ms, err)

        fut_signals     = pool.submit(
            _timed, "signals",
            lambda: _pendientes_collect(col, now, 14, progress=_signals_progress),
        )
        fut_tomorrow    = pool.submit(_timed, "tomorrow", _fetch_calendar_ahead, 1, 10)
        fut_forecast    = pool.submit(_timed, "forecast", _fetch_weather_forecast)
        fut_pagerank    = pool.submit(_timed, "pagerank", _fetch_pagerank_top, col, 5)
        fut_chrome      = pool.submit(_timed, "chrome", _fetch_chrome_top_week, 5)
        fut_eval        = pool.submit(_timed, "eval", _fetch_eval_trend)
        fut_followup    = pool.submit(_timed, "followup", _fetch_followup_aging)
        fut_drive       = pool.submit(_timed, "drive", _fetch_drive_recent, now, 5)
        fut_wa_unreplied = pool.submit(_timed, "wa_unreplied", _fetch_whatsapp_unreplied, 48, 5)
        fut_bookmarks   = pool.submit(_timed, "bookmarks", _fetch_chrome_bookmarks_used, 48, 5)
        fut_vaults      = pool.submit(_timed, "vaults", _fetch_vault_activity, 48, 5)
        fut_finance     = pool.submit(_timed, "finance", _fetch_finance, now)
        fut_cards       = pool.submit(_timed, "cards", _fetch_credit_cards, now)
        fut_youtube     = pool.submit(_timed, "youtube", _fetch_youtube_watched, 5, 168)

        # Helper: wait on a future with budget, return default on timeout
        # or worker exception. Emits "timeout" progress event so the UI
        # can show "skipped (>Ns)" instead of "running" forever (the
        # worker keeps going in bg to warm caches but its data is
        # discarded by the main thread once the budget elapses).
        import concurrent.futures as _cf

        def _await(name: str, fut, timeout: float, default=None):
            # Honor cancel_event: if the SSE client bailed, stop waiting
            # for remaining futures (the worker thread keeps running to
            # warm caches but its data is no longer consumed).
            if cancel_event is not None and cancel_event.is_set():
                return default
            try:
                return fut.result(timeout=timeout)
            except _cf.TimeoutError:
                if progress is not None:
                    with suppress(Exception):
                        progress(name, "timeout", timeout * 1000.0, None)
                return default
            except Exception:
                # `_timed` already emitted "error" with the real elapsed.
                return default

        try:
            today_ev = fut_today.result(timeout=30)
        except _cf.TimeoutError as exc:
            if progress is not None:
                with suppress(Exception):
                    progress("today", "timeout", 30000.0, None)
            pool.shutdown(wait=False, cancel_futures=True)
            raise HTTPException(status_code=500, detail="today_evidence timeout") from exc
        except Exception as exc:
            pool.shutdown(wait=False, cancel_futures=True)
            raise HTTPException(status_code=500, detail=str(exc))

        # Pendientes itself fans out 9 fetchers; cold path measured ~25s.
        # Bump cap to 45s — the user never waits on this since pre-warmer
        # eats the cold compute. If it still times out, the next cycle
        # repopulates and SWR keeps stale signals visible meanwhile.
        signals = _await("signals", fut_signals, 45, default={}) or {}
        tomorrow_calendar = _await("tomorrow", fut_tomorrow, 10, default=[]) or []
        weather_forecast = _await("forecast", fut_forecast, 10, default=None)
        pagerank_top = _await("pagerank", fut_pagerank, 10, default=[]) or []
        chrome_top_week = _await("chrome", fut_chrome, 5, default=[]) or []
        eval_trend = _await("eval", fut_eval, 5, default=None)
        # followup_aging has its own 6h cache + LLM-judge per loop on cold.
        # If not ready within 2s, skip it this cycle — the bg thread keeps
        # computing and the next prewarmer pass (25s later) picks up the
        # warm cache. This shaves ~15s off the cold critical path.
        followup_aging = _await("followup", fut_followup, 2, default=None)
        drive_recent = _await("drive", fut_drive, 10, default=[]) or []
        whatsapp_unreplied = _await("wa_unreplied", fut_wa_unreplied, 10, default=[]) or []
        chrome_bookmarks = _await("bookmarks", fut_bookmarks, 5, default=[]) or []
        vault_activity = _await("vaults", fut_vaults, 10, default={}) or {}
        finance = _await("finance", fut_finance, 5, default=None)
        cards = _await("cards", fut_cards, 5, default=[]) or []
        youtube_watched = _await("youtube", fut_youtube, 5, default=[]) or []

        signals["pagerank_top"] = pagerank_top
        signals["chrome_top_week"] = chrome_top_week
        signals["eval_trend"] = eval_trend
        signals["followup_aging"] = followup_aging
        signals["drive_recent"] = drive_recent
        signals["whatsapp_unreplied"] = whatsapp_unreplied
        signals["chrome_bookmarks"] = chrome_bookmarks
        signals["vault_activity"] = vault_activity
        signals["finance"] = finance
        signals["cards"] = cards
        signals["youtube_watched"] = youtube_watched
    finally:
        # Detach stragglers; they continue warming caches for the next call.
        # On SSE-client disconnect (cancel_event set), drop pending
        # futures via cancel_futures=True so we don't keep submitting
        # work no one's waiting for. Already-running fetchers finish
        # naturally — Python threads can't be preempted.
        cancel = cancel_event is not None and cancel_event.is_set()
        pool.shutdown(wait=False, cancel_futures=cancel)
        if timings:
            ranked = sorted(timings.items(), key=lambda kv: -kv[1])
            summary = " ".join(f"{k}={v:.1f}s" for k, v in ranked if v >= 0.2)
            if summary:
                tag = "cancelled " if cancel else ""
                print(f"[home-compute] {tag}{summary} total={time.time() - t_submit:.1f}s",
                      file=sys.stderr)

    today_total = (
        len(today_ev.get("recent_notes") or [])
        + len(today_ev.get("inbox_today") or [])
        + len(today_ev.get("todos") or [])
        + len(today_ev.get("new_contradictions") or [])
        + len(today_ev.get("low_conf_queries") or [])
    )

    narrative = "" if today_total == 0 else (_today_cached_narrative(date_label) or "")
    narrative_source = "cached" if narrative else "none"
    # Default correlations: el path de cache (no-regenerate) NO corre el
    # correlator por costo. Si la UI quiere pintar el panel "🔗 Patrones
    # del día", solo aparece después de un regenerate. Es OK — los
    # patrones ya quedan visibles en el narrative del brief cacheado.
    today_correlations: dict | None = None
    # Only call the LLM when the caller explicitly asks. Default path stays
    # fast — if no cached brief exists yet, the UI shows "pendiente" and
    # offers a button that re-hits with regenerate=true.
    if today_total > 0 and regenerate:
        # Inject the monthly MOZE snapshot so the LLM narrative can cite
        # cuánto gastaste este mes — otherwise it refuses ("no encontré
        # finanzas") even though the home panel renders the numbers.
        finance_snapshot = signals.get("finance") if isinstance(signals, dict) else None
        if finance_snapshot:
            today_ev = {**today_ev, "finance": finance_snapshot}
        # Fetchers TODAY-only (gmail/wa/calendar/youtube received-or-seen
        # HOY 00:00 → now, NO ventana rolling). Solo se ejecutan al
        # regenerar el brief para no inflar la latencia del path
        # default `/api/home`. Paralelizados — el peor caso es ~max(5s
        # gmail API, 0.05s wa SQLite, 2s icalBuddy, 0.1s yt sqlite) ≈
        # 5s. Silent-fail por fetcher: si gmail falla el resto sigue.
        from concurrent.futures import ThreadPoolExecutor as _TPE
        with _TPE(max_workers=4, thread_name_prefix="today-extras") as _t_pool:
            fut_gmail_today = _t_pool.submit(_fetch_gmail_today, now, 8)
            fut_wa_today    = _t_pool.submit(_fetch_whatsapp_today, now, 8)
            fut_cal_today   = _t_pool.submit(_fetch_calendar_today, 15)
            fut_yt_today    = _t_pool.submit(_fetch_youtube_today, now, 5)
            try:
                gmail_today = fut_gmail_today.result(timeout=10) or []
            except Exception:
                gmail_today = []
            try:
                wa_today = fut_wa_today.result(timeout=5) or []
            except Exception:
                wa_today = []
            try:
                cal_today = fut_cal_today.result(timeout=10) or []
            except Exception:
                cal_today = []
            try:
                yt_today = fut_yt_today.result(timeout=5) or []
            except Exception:
                yt_today = []
        # Cross-source extras — NO los inventamos, ya los recolectamos
        # en paralelo arriba (`signals` + buckets sueltos + los 4 today
        # de recién). Pasarlos al LLM permite que el brief escriba
        # "esperás respuesta de 3 chats WhatsApp + 2 mails VIP + tenés
        # meet con Pablo mañana 10hs" en vez de un recap solo del vault.
        # El render del prompt ignora keys vacías así que es safe pasar
        # todo aunque algunos buckets hayan caído en timeout.
        signals_dict = signals if isinstance(signals, dict) else {}
        # Dedup: si un YouTube apareció HOY, sacarlo del bucket "últimos
        # 7d" para que el LLM no lo cite dos veces. El `youtube_recent`
        # queda como "lo de los últimos 7 días que NO viste hoy".
        yt_today_ids = {v.get("video_id") for v in yt_today if v.get("video_id")}
        yt_recent_raw = signals_dict.get("youtube_watched") or []
        yt_recent = [v for v in yt_recent_raw if v.get("video_id") not in yt_today_ids]
        extras = {
            # TODAY (corte exacto al 00:00 local, no ventana rolling)
            "gmail_today": gmail_today,
            "whatsapp_today": wa_today,
            "calendar_today": cal_today,
            "youtube_today": yt_today,
            # Buckets ya existentes (rolling windows / "stale" buckets)
            "gmail_unread": signals_dict.get("gmail") or {},
            "mail_unread": signals_dict.get("mail_unread") or [],
            "whatsapp_unreplied": signals_dict.get("whatsapp_unreplied")
                or signals_dict.get("whatsapp") or [],
            "tomorrow_calendar": tomorrow_calendar,
            "drive_recent": signals_dict.get("drive_recent") or [],
            "youtube_recent": yt_recent,
            "chrome_bookmarks": signals_dict.get("chrome_bookmarks") or [],
            "loops_stale": signals_dict.get("loops_stale") or [],
            "loops_activo": signals_dict.get("loops_activo") or [],
            "pagerank_top": signals_dict.get("pagerank_top") or [],
            "vault_activity": signals_dict.get("vault_activity") or {},
            "followup_aging": signals_dict.get("followup_aging") or {},
        }
        # Pre-correlate ANTES del LLM call. Un 7B aplanado descubre poco
        # cross-source por su cuenta — pasarle entidades + temas ya
        # matched evita los "X conecta con Y porque ambos son X" tautológicos
        # que generaba antes. Costo: <100ms (regex + dict ops, no IO).
        from rag.today_correlator import correlate_today_signals as _corr
        try:
            today_correlations = _corr(today_ev, extras)
        except Exception as exc:  # noqa: BLE001
            print(f"[today-correlator] failed: {exc}", file=sys.stderr)
            today_correlations = {"people": [], "topics": [], "time_overlaps": []}
        extras["correlations"] = today_correlations
        prompt = _render_today_prompt(date_label, today_ev, extras=extras)
        narrative = _generate_today_narrative(prompt)
        narrative_source = "generated" if narrative else "error"
        if narrative:
            _persist_today_brief(date_label, narrative)

    brief_rel: str | None = None
    path = VAULT_PATH / MORNING_FOLDER / f"{date_label}-evening.md"
    if path.is_file():
        with suppress(ValueError):
            brief_rel = str(path.relative_to(VAULT_PATH))

    urgent: list[str] = []
    with suppress(Exception):
        urgent = _pendientes_urgent(signals, now)

    return {
        "generated_at": now.isoformat(timespec="seconds"),
        "date": date_label,
        "today": {
            "narrative": narrative,
            "narrative_source": narrative_source,
            "brief_path": brief_rel,
            "counts": {
                "recent_notes": len(today_ev.get("recent_notes") or []),
                "inbox_today": len(today_ev.get("inbox_today") or []),
                "todos": len(today_ev.get("todos") or []),
                "new_contradictions": len(today_ev.get("new_contradictions") or []),
                "low_conf_queries": len(today_ev.get("low_conf_queries") or []),
                "total": today_total,
            },
            "evidence": today_ev,
            # Cross-source matches pre-armados (people + topics + time
            # overlaps). Solo populated cuando regenerate=true; en el
            # cache path queda None y el frontend oculta el panel.
            "correlations": today_correlations,
        },
        "urgent": urgent,
        "signals": signals,
        "tomorrow_calendar": tomorrow_calendar,
        "weather_forecast": weather_forecast,
    }


@app.get("/api/pendientes")
def pendientes_api(days: int = 14) -> dict:
    """Raw JSON view of the services layer. Useful for a dedicated pendientes
    card in the UI or for third-party clients; the chat endpoint uses the
    same `_pendientes_collect` internally when intent matches.
    """
    col = get_db()
    now = datetime.now()
    try:
        ev = _pendientes_collect(col, now, days=days)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    urgent = _pendientes_urgent(ev, now)
    return {
        "generated_at": now.isoformat(timespec="seconds"),
        "days": days,
        "urgent": urgent,
        "services_consulted": _pendientes_services_consulted(ev),
        "evidence": ev,
    }


_CONV_PENDING_PATH = Path.home() / ".local/share/obsidian-rag" / "conversation_turn_pending.jsonl"

# Track in-flight conversation-writer threads so the shutdown hook can
# drain them. We stay with daemon=True (a wedged SQL write must not block
# process exit) but give the event loop a 5s join window — enough for a
# normal SQL upsert + atomic file write. Turns still in flight after the
# timeout fall through to `_append_pending_conversation_turn`'s retry file
# via the exception path in `_persist_conversation_turn`, or were already
# persisted but logged after the deadline (acceptable).
_CONV_WRITERS_LOCK = threading.Lock()
_CONV_WRITERS: "set[threading.Thread]" = set()


def _spawn_conversation_writer(
    target_args: tuple,
    name: str,
) -> threading.Thread:
    """Launch `_persist_conversation_turn` on a daemon thread while
    registering it for the shutdown-drain hook. The wrapper removes the
    thread from the tracker on completion (success or failure) so the set
    doesn't grow unbounded across sessions.
    """
    def _wrapper() -> None:
        try:
            _persist_conversation_turn(*target_args)
        finally:
            with _CONV_WRITERS_LOCK:
                _CONV_WRITERS.discard(threading.current_thread())

    t = threading.Thread(target=_wrapper, name=name, daemon=True)
    with _CONV_WRITERS_LOCK:
        _CONV_WRITERS.add(t)
    t.start()
    return t


@_on_shutdown
def _drain_conversation_writers() -> None:
    """On server stop, give in-flight conversation writers up to 5s to
    finish. Any that don't make it land in the retry queue on disk via
    `_append_pending_conversation_turn` (the writer's own except path) and
    will be re-applied at next startup by `_retry_pending_conversation_turns`.

    We do NOT block indefinitely — a wedged SQL write should never prevent
    the process from exiting.
    """
    with _CONV_WRITERS_LOCK:
        pending = list(_CONV_WRITERS)
    if not pending:
        return
    import time as _time
    deadline = _time.monotonic() + 5.0
    for t in pending:
        remaining = deadline - _time.monotonic()
        if remaining <= 0:
            break
        t.join(timeout=remaining)
    with _CONV_WRITERS_LOCK:
        stragglers = [t for t in _CONV_WRITERS if t.is_alive()]
    if stragglers:
        try:
            _LOG_QUEUE.put((
                LOG_PATH,
                json.dumps({
                    "kind": "conversation_writer_shutdown_timeout",
                    "count": len(stragglers),
                }) + "\n",
            ))
        except Exception:
            pass


@_on_shutdown
def _shutdown_joblib_loky_pool() -> None:
    """Drain el joblib/loky reusable executor al stop del lifespan de FastAPI.

    El trabajo real vive en `rag/_shutdown.py` para compartir código con
    el CLI (`rag watch`, `rag serve`, etc.) — ver ese módulo para la
    doc completa del bug, el fix en tres pasos, y el trade-off de
    `kill_workers=True` durante SIGTERM.

    Acá solo envolvemos el helper en el decorator `@_on_shutdown` del
    lifespan, y mantenemos el nombre original para que los tests que
    verifican la registración (`test_shutdown_callback_registered`)
    sigan sin cambio.
    """
    from rag._shutdown import shutdown_joblib_loky_pool

    shutdown_joblib_loky_pool()


def _append_pending_conversation_turn(rec: dict) -> None:
    """Append a serialisable turn record to the pending-retry file.

    The file is a JSONL ring buffer consumed by _retry_pending_conversation_turns
    at server startup and (best-effort) periodically. Each record is a
    self-contained snapshot — vault_root str, session_id, question, answer,
    sources, confidence, iso timestamp — so re-hydration doesn't need any
    runtime context.
    """
    try:
        _CONV_PENDING_PATH.parent.mkdir(parents=True, exist_ok=True)
        with _CONV_PENDING_PATH.open("a", encoding="utf-8") as f:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
    except Exception:
        # Dead-letter failed too — nothing else we can do without blocking
        # the SSE response. The LOG_QUEUE error record above is the last
        # signal the operator gets.
        pass


def _sanitize_confidence(confidence) -> float:
    """Clamp NaN/±Inf/None/garbage confidences → 0.0 so they never leak
    into persisted state. `retrieve()` returns `float('-inf')` when the
    corpus was empty (meta-chat, zero vec rows); persisting that ended up
    serialised as `-Infinity` inside `conversation_turn_pending.jsonl`
    (valid json via allow_nan=True, but no portable YAML/strict-json
    consumer can round-trip it) and as the literal `-inf` inside the
    conversation frontmatter (`confidence_avg: -inf`), which then broke
    the turn-2 averaging with `float("-inf") * 1 + x` propagating as
    `-inf` forever. Callers that go through `_retry_pending_conversation_turns`
    may also pass `None`/str from malformed pending records — swallow
    both into the 0.0 fallback rather than raising.
    """
    import math
    try:
        c = float(confidence)
    except (TypeError, ValueError):
        return 0.0
    if math.isnan(c) or math.isinf(c):
        return 0.0
    return c


def _validate_retrieve_result(result: dict) -> dict:
    """Normaliza y valida el shape que devuelve `multi_retrieve()` in-place.

    El chat endpoint consume `result["docs"]`, `["metas"]`, `["scores"]`,
    `["confidence"]` en múltiples sitios (zip loops, frontmatter, LLM
    context). Si alguno está mal formado — longitudes distintas, metas
    que no son dicts, confidence que es None/NaN — el handler crashea
    con `KeyError`/`TypeError` a mitad del SSE stream (respuesta 500
    después de que el frontend ya mostró "generando…") o peor, silent
    mis-attribution: el `zip(metas, scores)` trunca al corto y el user
    ve fuentes con scores desalineados.

    Esta helper normaliza el dict UNA vez, inmediatamente después del
    retrieve, para que el resto del handler pueda asumir:
      - `docs`, `metas`, `scores` son listas del MISMO largo.
      - cada meta es un dict (nunca None, str, int).
      - `confidence` es un float válido (nunca NaN/Inf/None/str).

    Si las longitudes discrepan las trunca al minimo común y loguea
    `[retrieve-shape-mismatch]` para diagnóstico — mucho mejor que
    silent mis-attribution. La truncación es defensiva, no punitiva:
    preferimos que el user vea 2 fuentes correctas que 3 con 1
    desalineada. Muta `result` in-place porque los callers esperan el
    mismo objeto que devolvió `multi_retrieve` (el dict también carga
    `query_variants`, `filters_applied`, `fast_path`, `intent`, etc.
    que no queremos perder al copiar).

    Introducido 2026-04-24 como defense-in-depth tras audit que
    encontró 3 sitios con crashes latentes: `result["confidence"]`
    (line ~5517), `m["file"]` en set comprehension (line ~5519), y
    zips de metas/scores + docs/metas sin validación (lines 5556,
    5722). Todos pasaban tests con retrieves mockeados "perfectos"
    pero explotarían si `rag.py` devolviera un shape raro (bug
    downstream, data corruption, Ollama OOM mid-embedding, etc.).
    """
    docs = list(result.get("docs") or [])
    metas_raw = list(result.get("metas") or [])
    scores = list(result.get("scores") or [])

    # Coerce cada meta a dict — si viene None/str/int por bug downstream,
    # reemplazamos con {} vacío (el source_payload renderer sabe manejar
    # ese caso con .get()). Preservamos el orden para no desalinear
    # contra docs/scores.
    metas: list[dict] = []
    for m in metas_raw:
        metas.append(m if isinstance(m, dict) else {})

    # Truncar al minimo común si hay mismatch. Prefer warning silencioso
    # vs crash + mis-attribution.
    n = min(len(docs), len(metas), len(scores))
    if n < max(len(docs), len(metas), len(scores)):
        print(
            f"[retrieve-shape-mismatch] docs={len(docs)} metas={len(metas)} "
            f"scores={len(scores)} → truncating to {n} for safety",
            flush=True,
        )

    result["docs"] = docs[:n]
    result["metas"] = metas[:n]
    result["scores"] = scores[:n]
    # Confidence: siempre float válido. `_sanitize_confidence` ya cubre
    # NaN/Inf/None/str/garbage, pero lo llamamos explícito para que el
    # resto del handler pueda asumir que `result["confidence"]` es
    # siempre accesible sin KeyError y siempre numeric.
    result["confidence"] = _sanitize_confidence(result.get("confidence"))
    return result


def _persist_conversation_turn(
    vault_root: Path,
    session_id: str,
    question: str,
    answer: str,
    metas: list[dict],
    scores: list[float],
    confidence: float,
) -> None:
    sources_payload = [
        {"file": m.get("file", ""), "score": float(s)}
        for m, s in zip(metas, scores)
    ]
    clean_conf = _sanitize_confidence(confidence)
    ts = datetime.now(timezone.utc)
    try:
        turn = TurnData(
            question=question,
            answer=answer,
            sources=sources_payload,
            confidence=clean_conf,
            timestamp=ts,
        )
        path = write_turn(vault_root, session_id, turn)
        _LOG_QUEUE.put((
            LOG_PATH,
            json.dumps({
                "kind": "conversation_turn_written",
                "session_id": session_id,
                "path": str(path.relative_to(vault_root)),
            }) + "\n",
        ))
    except Exception as exc:
        # Persist the turn payload to a retry queue on disk so a transient
        # SQL / fs failure doesn't drop it silently. Consumed at startup by
        # _retry_pending_conversation_turns. Errors still surface via
        # LOG_QUEUE for operator visibility, but data is not lost.
        _append_pending_conversation_turn({
            "ts": ts.isoformat(timespec="seconds"),
            "vault_root": str(vault_root),
            "session_id": session_id,
            "question": question,
            "answer": answer,
            "sources": sources_payload,
            "confidence": clean_conf,
            "error": repr(exc),
        })
        _LOG_QUEUE.put((
            LOG_PATH,
            json.dumps({
                "kind": "conversation_turn_error",
                "session_id": session_id,
                "error": repr(exc),
                "queued_for_retry": True,
            }) + "\n",
        ))


def _retry_pending_conversation_turns() -> int:
    """Drain _CONV_PENDING_PATH on startup. Returns number of turns retried
    successfully. Best-effort: failures stay in the file for the next run.
    """
    if not _CONV_PENDING_PATH.is_file():
        return 0
    try:
        raw = _CONV_PENDING_PATH.read_text(encoding="utf-8")
    except Exception:
        return 0
    remaining: list[str] = []
    retried = 0
    for line in raw.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            rec = json.loads(line)
            ts_str = rec["ts"]
            ts = datetime.fromisoformat(ts_str)
            if ts.tzinfo is None:
                ts = ts.replace(tzinfo=timezone.utc)
            # Sanitize `confidence` on replay too — legacy pending files
            # written before _sanitize_confidence existed may carry
            # -Infinity / NaN, which round-trip through `float(...)` and
            # silently re-poison the restored frontmatter.
            turn = TurnData(
                question=rec["question"],
                answer=rec["answer"],
                sources=rec.get("sources", []),
                confidence=_sanitize_confidence(rec.get("confidence", 0.0)),
                timestamp=ts,
            )
            write_turn(Path(rec["vault_root"]), rec["session_id"], turn)
            retried += 1
        except Exception:
            # Keep the line for the next retry cycle.
            remaining.append(line)
    try:
        if remaining:
            _CONV_PENDING_PATH.write_text("\n".join(remaining) + "\n",
                                           encoding="utf-8")
        else:
            _CONV_PENDING_PATH.unlink(missing_ok=True)
    except Exception:
        pass
    return retried


def _build_retrieve_hint(intent: str | None) -> str | None:
    """Return a human-friendly hint for the SSE `status {stage:"retrieving"}`
    event based on the classified intent.

    Mapped to short, action-describing strings so the ticker label in the
    browser tells the user WHAT is being searched — not just that *something*
    is happening. Generic "semantic" (the default) returns None because it
    has no incremental info over the existing "buscando…" ticker.

    Intent vocabulary matches `rag.classify_intent` (see `rag.py` for the
    canonical set). Unknown intents return None — the client falls back to
    its legacy ticker copy.

    Added 2026-04-22 as a UX-latency fix: measured web retrieve p90 = 25s;
    during that wait the user had no signal about the retrieval path. The
    hint closes the gap without changing any actual timing.
    """
    if not intent:
        return None
    # Keep each hint ≤ 48 chars so it fits the thinking-line in one row
    # even on mobile. Avoid trailing period — the ticker already suffixes
    # the elapsed seconds after the label.
    return {
        "count":         "Contando notas…",
        "list":          "Listando notas…",
        "recent":        "Revisando notas recientes…",
        "agenda":        "Revisando tu agenda…",
        "entity_lookup": "Buscando por persona u organización…",
        "comparison":    "Comparando fuentes…",
        "synthesis":     "Sintetizando fuentes…",
        "create":        "Interpretando pedido de creación…",
    }.get(intent)


def _resolve_redo_question(turn_id: str) -> tuple[str | None, str | None]:
    """Resolve the original question from a previous turn_id.

    `turn_id` is not a first-class column of `rag_queries` — it lives in
    `extra_json` (see `_map_queries_row()`). We query by the JSON path
    and return `(q, session)` from the latest matching row, or (None, None)
    if not found.

    Scoped to rag_queries because every web chat turn is logged there via
    log_query_event(). The lookup is O(log n) via the index on `ts`, then
    a sequential scan over extra_json matches — fine for single-lookup
    semantics (redo is rare vs new chat).
    """
    try:
        from rag import _ragvec_state_conn
        with _ragvec_state_conn() as conn:
            row = conn.execute(
                "SELECT q, session FROM rag_queries"
                " WHERE json_extract(extra_json, '$.turn_id') = ?"
                " ORDER BY ts DESC LIMIT 1",
                (turn_id,),
            ).fetchone()
    except Exception as exc:  # noqa: BLE001 — log + fail gracefully
        print(f"[redo] sql lookup failed: {type(exc).__name__}: {exc}", flush=True)
        return None, None
    if row is None:
        return None, None
    q = (row[0] or "").strip() or None
    session = (row[1] or "").strip() or None
    return q, session


@app.get("/api/chat")
def chat_get_redirect():
    """Defensive redirect for users who bookmark or navigate to the API
    endpoint instead of the UI. `/api/chat` is the POST/SSE programmatic
    endpoint; the human-facing chat lives at `/chat`. Without this,
    hitting the URL in a browser returned `{"detail":"Method Not
    Allowed"}` (FastAPI's default 405) — ~256 such 405s observed in the
    web.log the first time the user tripped on it. 307 preserves the
    method if the caller is a programmatic client that for some reason
    issued GET; browsers just follow to the HTML UI.
    """
    from fastapi.responses import RedirectResponse  # noqa: PLC0415
    return RedirectResponse(url="/chat", status_code=307)


@app.post("/api/chat")
def chat(req: ChatRequest, request: Request) -> StreamingResponse:
    # Rate limit: 30 chat requests / 60s per IP. Each request pins the
    # chat model + reranker + embedder on MPS — a tight loop from a
    # runaway client can starve the daemon for legitimate work.
    client_ip = (request.client.host if request.client else "unknown")
    _check_rate_limit(_CHAT_BUCKETS, client_ip,
                      _CHAT_RATE_LIMIT, _CHAT_RATE_WINDOW)
    # Device classification for telemetry + downstream decisions. El
    # User-Agent header es el source más reliable — iPhone vs iPad vs Mac
    # desktop vs otra Mac. Loggeado en `rag_queries.extra_json.device`
    # en los 3 log_query_event sites abajo. Ver rag._classify_device
    # para la política + ejemplos.
    try:
        import rag as _rag_mod
        _client_device = _rag_mod._classify_device(
            request.headers.get("User-Agent", "")
        )
    except Exception:
        _client_device = "other"

    # ── /redo path ───────────────────────────────────────────────────────
    # If the client sent redo_turn_id, resolve the original question from
    # rag_queries (SQL), optionally prepending a hint to soft-steer the
    # regeneration. This keeps the endpoint uniform (same SSE stream, same
    # downstream retrieve + generate pipeline) — the only difference is
    # that `question` is loaded from persistence + possibly augmented.
    # Session preservation: if the client didn't pass session_id, we
    # inherit the one from the original turn, so the regenerated answer
    # lands in the same conversation.
    _redo_hint: str | None = None
    _redo_of: str | None = None
    if req.redo_turn_id:
        _orig_q, _orig_session = _resolve_redo_question(req.redo_turn_id)
        if not _orig_q:
            raise HTTPException(
                status_code=404,
                detail=f"turn_id '{req.redo_turn_id}' no encontrado en rag_queries",
            )
        _redo_of = req.redo_turn_id
        _redo_hint = (req.hint or "").strip() or None
        if _redo_hint:
            # Soft-steer: concatenate with a clear separator so the LLM
            # retriever sees both signals. We don't modify the original
            # q in SQL — the hint is ephemeral to this regeneration.
            question = f"{_orig_q} — enfocá en: {_redo_hint}"
        else:
            question = _orig_q
        # Inherit the session from the original turn unless the client
        # overrode it. Preserves conversation continuity.
        if not req.session_id and _orig_session:
            # Pydantic models are immutable by default but session_id
            # has no frozen=True guard — assign via object.__setattr__
            # to stay type-safe.
            object.__setattr__(req, "session_id", _orig_session)
    else:
        question = req.question.strip()
    if not question:
        raise HTTPException(status_code=400, detail="empty question")

    # Detect create-intent EARLY. When true we skip emitting `sources`
    # (vault citations are noise when the user is creating something new,
    # not asking about existing notes) and we bypass the read-intent
    # pre-router further down.
    is_propose_intent = _detect_propose_intent(question)

    # 2026-04-28 wave-8: track de tools y location del turno previo.
    # Se inicializan acá para que estén disponibles en todas las paths
    # (cache hit, metachat, retrieve, etc.) y se persistan al cierre.
    _last_weather_location: str | None = None
    _last_turn_tools: list[str] = []
    _last_turn_weather_location: str | None = None
    try:
        _prev_session_for_state = ensure_session(req.session_id, mode="chat") if req.session_id else None
        _prev_turns_state = (_prev_session_for_state or {}).get("turns") or []
        if _prev_turns_state:
            _last_t = _prev_turns_state[-1]
            _last_turn_tools = list(_last_t.get("tools_fired") or [])
            _last_turn_weather_location = _last_t.get("weather_location") or None
    except Exception:
        # Defense in depth — el state lookup no debe romper la chat call.
        _last_turn_tools = []
        _last_turn_weather_location = None

    # Meta-chat short-circuit — greetings / thanks / "what can you do".
    # 2026-04-21 Playwright probe (Fer F.): "hola" produced "Según tus
    # notas, tenés varias interacciones con diferentes contactos por
    # WhatsApp..." because `_WEB_SYSTEM_PROMPT` REGLA 1 forces the LLM
    # to engage with whatever retrieved context it got. Cheapest fix:
    # detect the bare social / meta turns up front and reply with a
    # canned line, skipping retrieval + tool-calling + LLM entirely.
    # Zero latency (<1ms) + no hallucination possible. See
    # `_detect_metachat_intent` for the matcher shape + rationale.
    is_metachat = (not is_propose_intent) and _detect_metachat_intent(question)

    # 2026-04-23: degenerate-query short-circuit. Inputs sin ≥2 chars
    # alfanuméricos (ej. "x", "?¡@#") caían al retrieve y devolvían
    # chunks random porque el matching semántico de un string casi
    # vacío no tiene signal. Devolvemos canned "no entendí, reformulá"
    # antes de tocar ninguna pieza pesada. No es metachat (no es saludo
    # ni thanks) — tiene su propio bucket para que el logging lo
    # distinga en analytics.
    is_degenerate = (
        not is_propose_intent
        and not is_metachat
        and _is_degenerate_query(question)
    )

    sid = req.session_id or f"web:{uuid.uuid4().hex[:12]}"
    sess = ensure_session(sid, mode="chat")
    vaults = _resolve_scope(req.vault_scope)
    if not vaults:
        raise HTTPException(status_code=400, detail=f"vault '{req.vault_scope}' no encontrado")

    def _emit_enrich(turn_id: str, q: str, answer: str, top_score: float):
        """Yield an `enrich` SSE event with cross-source signals (WA/Calendar/
        Reminders). 4s wall budget enforced via ThreadPoolExecutor.
        Soft-fail per source — never raises into the stream.

        Bug 2026-04-24 (Fer F. "quedó colgado"): previously used
        ``with ThreadPoolExecutor(max_workers=1) as _ex:``. When the inner
        ``.result(timeout=4.0)`` raised ``TimeoutError``, the exception
        propagated up and Python ran the ``__exit__`` of the context
        manager before reaching our except block. ``__exit__`` calls
        ``shutdown(wait=True)`` by default, which **waits for the still-
        running thread to finish**. If the thread was hung in osascript
        or ollama.chat (no timeout on those calls), the SSE generator
        blocked indefinitely — no ``[enrich] skipped`` log, no stream
        close, user sees a frozen chat.

        Fix: explicit submit + `shutdown(wait=False, cancel_futures=True)`
        in finally. The daemon worker thread may keep running briefly,
        but the SSE stream is released immediately. We don't need the
        result anyway once we've decided to skip.
        """
        print(f"[enrich] start turn={turn_id} top_score={top_score} answer_len={len(answer or '')}", flush=True)
        from concurrent.futures import ThreadPoolExecutor, TimeoutError as _FutTimeout
        _ex = ThreadPoolExecutor(max_workers=1, thread_name_prefix=f"enrich-{turn_id[:8]}")
        try:
            from rag import build_enrich_payload
            _fut = _ex.submit(build_enrich_payload, q, answer, top_score)
            _enrich = _fut.result(timeout=4.0)
            if _enrich:
                print(f"[enrich] hit lines={len(_enrich.get('lines', []))}", flush=True)
                return _sse("enrich", {"turn_id": turn_id, **_enrich})
            print("[enrich] no lines (empty payload)", flush=True)
        except _FutTimeout:
            print("[enrich] skipped: 4s budget exceeded", flush=True)
        except Exception as exc:
            print(f"[enrich] skipped: {type(exc).__name__}: {exc}", flush=True)
        finally:
            # Non-blocking shutdown: don't wait for a hung worker. Any
            # pending futures get cancelled; futures already running keep
            # going until they naturally finish (in a daemon thread — they
            # won't block process exit).
            _ex.shutdown(wait=False, cancel_futures=True)
        return None

    def _emit_grounding(turn_id: str, full: str, docs: list, metas: list, question: str) -> str | None:
        """Compute NLI grounding for the LLM response and return an SSE event
        string, or None if skipped or unavailable.
        Silent-fail — never raises into the stream.

        Bug 2026-04-27 (audit "RAG - Flujo api-chat - auditoría" #1): pre-fix
        `ground_claims_nli` ran synchronously inside the SSE generator. If
        the NLI helper hung (qwen2.5:3b stuck-load, MPS contention, ollama
        daemon not responsive), the generator blocked indefinitely — the
        client never received `done` and the PWA showed a frozen spinner.
        Fix: wrap the call in `ThreadPoolExecutor + result(timeout=4.0) +
        shutdown(wait=False, cancel_futures=True)`, the same pattern that
        `_emit_enrich` (just above) uses for the same kind of hang. The
        background worker thread may keep running briefly but the SSE
        stream is released within the timeout budget.
        """
        from concurrent.futures import ThreadPoolExecutor, TimeoutError as _FutTimeout
        try:
            import rag as _rag
            if not _rag._nli_grounding_enabled():
                return None
            if not full.strip() or not docs:
                return None
            _intent, _ = _rag.classify_intent(question, set(), set())
            if _intent in _rag._nli_skip_intents():
                return None
            claims = _rag.split_claims(full)
        except Exception as exc:
            print(f"[nli-grounding] skipped (pre-flight): {type(exc).__name__}: {exc}", flush=True)
            return None

        # Wall-time budget for the NLI helper. 4s matches `_emit_enrich`.
        # Override via env for hosts where bge-m3 + reranker + qwen3 fight
        # for MPS — bumping to 6-8s buys headroom without unbounded waits.
        try:
            _nli_budget_s = float(os.environ.get("RAG_NLI_GROUNDING_BUDGET_S", "4.0"))
        except ValueError:
            _nli_budget_s = 4.0

        _ex = ThreadPoolExecutor(max_workers=1, thread_name_prefix=f"grounding-{turn_id[:8]}")
        try:
            # `ground_claims_nli` has keyword-only args (after `*`), so
            # threshold_contradicts/max_claims MUST go as kwargs in submit().
            _fut = _ex.submit(
                _rag.ground_claims_nli,
                claims, docs, metas,
                threshold_contradicts=_rag._nli_contradicts_threshold(),
                max_claims=_rag._nli_max_claims(),
            )
            grounding = _fut.result(timeout=_nli_budget_s)
        except _FutTimeout:
            print(
                f"[nli-grounding] skipped: budget exceeded "
                f"({_nli_budget_s}s) turn={turn_id}",
                flush=True,
            )
            return None
        except Exception as exc:
            print(f"[nli-grounding] skipped: {type(exc).__name__}: {exc}", flush=True)
            return None
        finally:
            # Non-blocking shutdown: don't wait for a hung worker. Same
            # rationale as `_emit_enrich` — keeps the SSE stream moving
            # even if `ground_claims_nli` is stuck on a daemon thread.
            _ex.shutdown(wait=False, cancel_futures=True)

        try:
            if grounding is None or grounding.claims_total == 0:
                return None
            _metas_by_cid: dict = {}
            for _m in metas:
                _cid = _m.get("chunk_id") or _m.get("file", "")
                if _cid:
                    _metas_by_cid[_cid] = _m
            _claims_payload = []
            for _cg in grounding.claims:
                _note = ""
                if _cg.evidence_chunk_id:
                    _em = _metas_by_cid.get(_cg.evidence_chunk_id)
                    _note = str(_em.get("note", ""))[:60] if _em else str(_cg.evidence_chunk_id)[:60]
                _claims_payload.append({
                    "text": _cg.text,
                    "verdict": _cg.verdict,
                    "score": round(_cg.score, 3),
                    "evidence_span": (_cg.evidence_span or "")[:200],
                    "evidence_note": _note,
                })
            print(
                f"[nli-grounding] turn={turn_id} total={grounding.claims_total} "
                f"entails={grounding.claims_supported} contradicts={grounding.claims_contradicted} "
                f"nli_ms={grounding.nli_ms}",
                flush=True,
            )
            return _sse("grounding", {
                "turn_id": turn_id,
                "claims": _claims_payload,
                "total": grounding.claims_total,
                "supported": grounding.claims_supported,
                "contradicted": grounding.claims_contradicted,
                "neutral": grounding.claims_neutral,
                "nli_ms": grounding.nli_ms,
            })
        except Exception as exc:
            print(f"[nli-grounding] skipped: {type(exc).__name__}: {exc}", flush=True)
            return None

    def gen():
        # 2026-04-28 wave-8: nonlocal de _last_weather_location porque
        # asignamos a esa variable dentro de gen() (línea ~10395) y sin
        # nonlocal Python la considera local-only y los reads tempranos
        # (append_turn al final) tiran UnboundLocalError. _last_turn_*
        # son read-only para gen(), no requieren nonlocal.
        nonlocal _last_weather_location

        _t0 = time.perf_counter()
        yield _sse("session", {"id": sess["id"]})

        # Degenerate query short-circuit. Devuelve canned reply e
        # invita a reformular. No loggea como metachat (bucket propio
        # en analytics: `web.chat.degenerate`).
        if is_degenerate:
            reply = _pick_degenerate_reply(question)
            turn_id = new_turn_id()
            yield _sse("sources", {
                "items": [], "confidence": None, "metachat": True,
            })
            yield _sse("status", {"stage": "generating"})
            for i in range(0, len(reply), 40):
                yield _sse("token", {"delta": reply[i:i+40]})
            total_ms = int((time.perf_counter() - _t0) * 1000)
            yield _sse("done", {
                "turn_id": turn_id, "elapsed_ms": total_ms,
                "metachat": True,
            })
            try:
                append_turn(sess, {
                    "turn_id": turn_id, "q": question,
                    "a": reply[:500], "metachat": True,
                })
                save_session(sess)
            except Exception:
                pass
            try:
                log_query_event({
                    "cmd": "web.chat.degenerate", "q": question[:200],
                    "session": sess["id"], "answered": True,
                    "t_total": round(total_ms / 1000.0, 3),
                })
            except Exception:
                pass
            return

        # Finance/cards short-circuit (added 2026-04-26). El chat web
        # tarda 30-48s con el flow LLM completo (retrieve + tools +
        # qwen2.5:7b prompt+decode), pasando el timeout de 45s del
        # cliente Ollama → frontend muestra "LLM falló: timed out".
        # Para queries financieras con datos parsed (xlsx tarjeta,
        # CSV MOZE) NO necesitamos el LLM — el render determinístico
        # de `rag._finance_cards_comment` da la respuesta exacta en
        # <1s. Mismo patrón que `rag serve /query` y `/chat`.
        try:
            import rag as _rag_mod_for_fin
            _is_finance_q = _rag_mod_for_fin._is_finance_or_cards_query(question)
        except Exception:
            _is_finance_q = False
        if _is_finance_q:
            try:
                _t_fin = time.perf_counter()
                _finance_data = _fetch_finance()
                _cards_data = _fetch_credit_cards()
                _fin_answer = _rag_mod_for_fin._finance_cards_comment(
                    question, _finance_data, _cards_data,
                )
            except Exception as exc:
                _fin_answer = (
                    f"No pude leer los datos de tarjetas/MOZE: {str(exc)[:80]}"
                )

            turn_id = new_turn_id()
            yield _sse("sources", {
                "items": [],
                # confidence ALTA explícita evita que el frontend marque
                # `weakAnswer=true` y dispare el link "↗ buscar en internet"
                # — la respuesta es exacta, viene del banco, NO necesita
                # fallback a Google. Bug observado 2026-04-26.
                "confidence": 1.0,
                "intent": "finance",
            })
            yield _sse("status", {
                "stage": "generating",
                "intent": "finance",
                "hint": "Datos del banco / MOZE",
            })
            # Tokenizado en chunks chicos para mimic del UX normal —
            # el frontend tiene animación de typing.
            for i in range(0, len(_fin_answer), 40):
                yield _sse("token", {"delta": _fin_answer[i:i+40]})
            total_ms = int((time.perf_counter() - _t0) * 1000)
            yield _sse("done", {
                "turn_id": turn_id,
                "elapsed_ms": total_ms,
                "mode": "finance",
                # `source_specific: true` = bypass del fallback de
                # "buscar en internet" / cluster Google/YouTube/Wiki.
                # Mismo flag que usan los tools gmail_recent /
                # calendar_ahead / reminders_due — la data es local,
                # autoritativa, no necesita escape a la web.
                "source_specific": True,
            })
            try:
                append_turn(sess, {
                    "turn_id": turn_id, "q": question,
                    "a": _fin_answer[:500], "mode": "finance",
                })
                save_session(sess)
            except Exception:
                pass
            try:
                log_query_event({
                    "cmd": "web.chat.finance", "q": question[:200],
                    "session": sess["id"], "answered": True,
                    "t_total": round(total_ms / 1000.0, 3),
                })
            except Exception:
                pass
            return

        # Meta-chat short-circuit. Canned responses (varied across
        # WA-style variants so repeated "hola" doesn't always say the
        # same thing). Random seeded by minute so the same message in
        # a tight window picks the same variant; tests can monkey-patch
        # `_METACHAT_PICKER` to force a specific variant.
        if is_metachat:
            reply = _pick_metachat_reply(question)
            turn_id = new_turn_id()
            # Stream token-by-token to mimic the shape of a real answer
            # so the UI's token-append animation still plays — keeps UX
            # consistent whether or not retrieval ran.
            yield _sse("sources", {
                "items": [], "confidence": None, "metachat": True,
            })
            yield _sse("status", {"stage": "generating"})
            for i in range(0, len(reply), 40):
                yield _sse("token", {"delta": reply[i:i+40]})
            total_ms = int((time.perf_counter() - _t0) * 1000)
            yield _sse("done", {
                "turn_id": turn_id, "elapsed_ms": total_ms,
                "metachat": True,
            })
            # Log + persist turn so the session + analytics see it.
            try:
                append_turn(sess, {
                    "turn_id": turn_id, "q": question,
                    "a": reply[:500], "metachat": True,
                })
                save_session(sess)
            except Exception:
                pass
            try:
                # Metachat fires before retrieve(), so we classify intent
                # inline to keep telemetry consistent with the other paths.
                # Failure is non-fatal — the event still goes through.
                try:
                    import rag as _rag_m
                    _meta_intent, _ = _rag_m.classify_intent(question, set(), set())
                except Exception:
                    _meta_intent = None
                log_query_event({
                    "cmd": "web.chat.metachat", "q": question[:200],
                    "session": sess["id"], "answered": True,
                    "t_total": round(total_ms / 1000.0, 3),
                    "intent": _meta_intent,
                    # device: iphone/ipad/mac/linux/windows/android/other
                    # — habilita `SELECT device, AVG(...) GROUP BY 1` en analytics
                    "device": _client_device,
                })
            except Exception:
                pass
            return

        history = session_history(sess, window=SESSION_HISTORY_WINDOW)

        # Response cache — sirve respuesta cacheada si (question, scope, model,
        # vault_count) matchea + TTL viva + NO es follow-up. Skipped cuando hay
        # history porque la respuesta depende del contexto de la conversación,
        # que el key actual no refleja.
        # Nota: `_cache_key` queda pre-inicializado en None porque el topic-
        # shift gate (~línea 3914) puede reasignar `history = []` DESPUÉS de
        # este bloque; en ese caso, el PUT path abajo entraba con `_cache_key`
        # no-bound y deja un `UnboundLocalError` visible en web.log como
        # `[chat-cache] put failed: cannot access local variable '_cache_key'…`.
        _cache_key: str | None = None
        # ── Semantic cache (SQL, persistent, paraphrase-tolerant). 2026-04-23.
        # Second-layer fallback: exact-string LRU arriba captura repetidos
        # exactos dentro de TTL 5min. El semantic cache persiste 24h y
        # matchea paraphrases ("qué es ikigai" vs "qué es el ikigai") vía
        # cosine del embedding bge-m3. Llenamos _semantic_cache_emb/hash
        # acá para reusarlos en el PUT path al final (si hubo miss).
        _semantic_cache_emb = None
        _semantic_cache_hash = ""
        _semantic_cache_probe: dict | None = None
        if not history:
            try:
                from rag import get_db_for
                _vault_chunks = get_db_for(vaults[0][1]).count() if vaults else 0
            except Exception:
                _vault_chunks = 0
            _cache_key = _chat_cache_key(
                question, req.vault_scope or "", _resolve_web_chat_model(), _vault_chunks
            )
            _cached = _chat_cache_get(_cache_key)
            if _cached:
                # Replay completo como SSE. Status `cached` NO es redundante
                # aunque el `done` event también lleve `cached: True` — el
                # UI lo consume en app.js:2346 para mostrar el label "desde
                # caché" y detener el ticker inmediatamente (sub-100ms
                # replay no merece running counter). Sin este event, el
                # ticker quedaría en "retrieving" hasta el `done`. Audit
                # 2026-04-22 propuso removerlo como "1-2ms savings"; el
                # gap es UX-breaking, no perf win. NO remover.
                yield _sse("status", {"stage": "cached"})
                if not is_propose_intent:
                    yield _sse("sources", {
                        "items": _cached["sources_items"],
                        "confidence": _cached["top_score"],
                    })
                # Stream el texto en chunks chicos para mantener la ilusión
                # de streaming (el cliente ya espera SSE tokens). 40 chars ≈
                # 10 tokens — UI renderea suave sin el feel "pegote instantáneo".
                _text = _cached["text"]
                for i in range(0, len(_text), 40):
                    yield _sse("token", {"delta": _text[i:i+40]})
                _cached_total_ms = int((time.perf_counter() - _t0) * 1000)
                _cached_turn_id = new_turn_id()
                yield _sse("done", {
                    "turn_id": _cached_turn_id,
                    "top_score": _cached["top_score"],
                    "total_ms": _cached_total_ms,
                    "retrieve_ms": 0,
                    "ttft_ms": _cached_total_ms,
                    "llm_ms": 0,
                    "cached": True,
                })
                _enrich_evt = _emit_enrich(
                    _cached_turn_id, question, _text, float(_cached["top_score"]),
                )
                if _enrich_evt:
                    yield _enrich_evt
                print(f"[chat-cache] HIT {_cache_key} total={_cached_total_ms}ms", flush=True)
                # Emit a timing schema-stable line so downstream parsers
                # (tool_rounds/tool_ms/tool_names) don't see an NaN on
                # cache-hit turns. Values are 0/empty — tool loop skipped.
                print(
                    f"[chat-timing] model={_resolve_web_chat_model()} retrieve=0ms "
                    f"reform=0ms reform_outcome=skipped q_words={len(question.split())} "
                    f"wa=0ms llm_prefill=0ms llm_decode=0ms ttft={_cached_total_ms}ms "
                    f"total={_cached_total_ms}ms ctx_chars=0 tokens≈{len(_text.split())} "
                    # Defense-in-depth: el cached top_score pasa por
                    # `round(float(...), 3)` en el PUT (línea ~4476). Post
                    # fix del persist path queda finito, pero saneamos por
                    # las dudas para que el log de cache-hit no cargue un
                    # `-inf` si algún día alguien persiste crudo.
                    f"confidence={_sanitize_confidence(_cached['top_score']):.3f} variants=0 "
                    f"tool_rounds=0 tool_ms=0 tool_names= cached=1",
                    flush=True,
                )
                # Todavía appendeamos al session history — la conversación continúa normal
                append_turn(sess, {
                    "q": question,
                    "a": _text,
                    "paths": [s.get("file", "") for s in _cached["sources_items"]],
                    "top_score": _cached["top_score"],
                    "turn_id": new_turn_id(),
                })
                save_session(sess)
                return

        # ── Semantic cache (SQL) — segundo layer post LRU miss ────────────
        # Cubre: paraphrases + reinicios del server + TTL largo (24h) vs el
        # LRU (exact-match lowercased + 5min + in-memory).
        # Gates: no history, single-vault (corpus_hash es per-col),
        # no propose_intent (create actions NO son queries), cache enabled,
        # NO forced-tool-intent (weather, whatsapp_list_scheduled, reminders_due,
        # etc — el resultado es dinámico, cachearlo devuelve datos stale).
        try:
            _forced_intent_check_pairs = _detect_tool_intent(question)
        except Exception:
            _forced_intent_check_pairs = []
        _semantic_eligible = (
            not history
            and not is_propose_intent
            and len(vaults) == 1
            and not _forced_intent_check_pairs
        )
        # Audit 2026-04-24: distinguir "skipped por gate" de "lookup nunca
        # corrió" en el telemetry. Pre-fix `_semantic_cache_probe` quedaba
        # en None cuando `_semantic_eligible=False`, indistinguible del
        # caso de error pre-lookup. Ahora poblamos un probe explícito con
        # el primer gate que cortó — datos accionables para el próximo
        # tuning de cache: si `flags_skip` domina con reason=`history`,
        # tunear el cache key para incluir history hash sería más
        # impactful que aflojar el gate.
        # 2026-04-28 (eval cleanup wave 2): agregado forced_intent reason —
        # antes "clima en BA" se cacheaba y devolvía "última predicción
        # está para Santa Fe" porque match'eaba un embed viejo. Ahora si
        # la query rutea a una tool, salteamos el cache.
        if not _semantic_eligible:
            _skip_reason = (
                "history" if history else
                "propose_intent" if is_propose_intent else
                "multi_vault" if len(vaults) != 1 else
                "forced_intent" if _forced_intent_check_pairs else
                "unknown"
            )
            _semantic_cache_probe = {
                "result": "skipped",
                "reason": f"flags_skip:{_skip_reason}",
                "top_cosine": None,
                "candidates": 0,
            }
        if _semantic_eligible:
            try:
                from rag import (
                    embed as _rag_embed,
                    _corpus_hash_cached,
                    semantic_cache_lookup,
                    _semantic_cache_enabled,
                    query_embed_local as _rag_query_embed_local,
                    _local_embed_enabled as _rag_local_embed_enabled,
                )
                if _semantic_cache_enabled():
                    from rag import get_db_for as _rag_get_db_for
                    _sem_col = _rag_get_db_for(vaults[0][1])
                    # Doble-embed fix (audit perf 2026-04-26): pre-fix
                    # SIEMPRE iba por ollama HTTP (~140ms warm, 10s cold)
                    # y después `retrieve()` re-embeddeaba la misma query
                    # con `query_embed_local` (~30ms MPS) — 2× embed por
                    # query. Si el local está warm (Event set en steady-
                    # state web), usá ese; si no cae a ollama (mismo path
                    # que retrieve usaría tras `RAG_LOCAL_EMBED_WAIT_MS=0`).
                    _semantic_cache_emb = None
                    if _rag_local_embed_enabled():
                        _local_emb = _rag_query_embed_local([question])
                        if _local_emb:
                            _semantic_cache_emb = _local_emb[0]
                    if _semantic_cache_emb is None:
                        _semantic_cache_emb = _rag_embed([question])[0]
                    _semantic_cache_hash = _corpus_hash_cached(_sem_col)
                    _sem_hit, _semantic_cache_probe = semantic_cache_lookup(
                        _semantic_cache_emb, _semantic_cache_hash,
                        return_probe=True,
                    )
                    if _sem_hit is not None:
                        # Replay SSE same shape as LRU hit above. We
                        # synthesize `sources_items` from the cached paths +
                        # scores — minimal meta (file/note/folder/score/bar)
                        # since we don't have the full meta rows anymore.
                        from pathlib import Path as _SemP
                        _sem_paths = _sem_hit.get("paths") or []
                        _sem_scores = _sem_hit.get("scores") or []
                        _sem_top = _sem_hit.get("top_score")
                        if _sem_top is None:
                            _sem_top = 0.0
                        _sem_text = _sem_hit["response"]
                        # 2026-04-28 wave-7: aplicar PII redaction también
                        # al cache hit. Las entries pre-fix pueden tener
                        # PII en plain text. La invalidación por filter_version
                        # del corpus_hash debería evitarlo, pero defense-in-
                        # depth: redactar antes de yield.
                        _sem_text, _sem_redact_n = _redact_pii(_sem_text)
                        if _sem_redact_n > 0:
                            print(
                                f"[chat-pii-redact] phase=cache_hit "
                                f"redacted {_sem_redact_n} pattern(s) from "
                                f"cached response — old entry should be "
                                f"invalidated by corpus_hash bump",
                                flush=True,
                            )
                        # 2026-04-28 wave-8: idem para CJK leaks. Cache
                        # entries pre-fix pueden tener 汉字 incrustados
                        # (qwen2.5:7b leaks bajo presión de contexto en
                        # weather queries).
                        _sem_text_pre_strip_len = len(_sem_text)
                        _sem_text = _strip_foreign_scripts(_sem_text)
                        _sem_strip_n = _sem_text_pre_strip_len - len(_sem_text)
                        if _sem_strip_n > 0:
                            print(
                                f"[chat-cjk-strip] phase=cache_hit "
                                f"stripped {_sem_strip_n} foreign char(s) "
                                f"from cached response — filter_version "
                                f"bump should invalidate stale entries",
                                flush=True,
                            )
                        _sem_sources = [
                            {
                                "file": _p,
                                "note": _SemP(_p).stem,
                                "folder": str(_SemP(_p).parent),
                                "score": round(float(_s), 3),
                                "bar": _score_bar(float(_s)),
                            }
                            for _p, _s in zip(_sem_paths, _sem_scores)
                        ]
                        yield _sse("status", {"stage": "cached"})
                        if not is_propose_intent:
                            yield _sse("sources", {
                                "items": _sem_sources,
                                "confidence": _sem_top,
                            })
                        for _i in range(0, len(_sem_text), 40):
                            yield _sse("token", {"delta": _sem_text[_i:_i + 40]})
                        _sem_total_ms = int((time.perf_counter() - _t0) * 1000)
                        _sem_turn_id = new_turn_id()
                        yield _sse("done", {
                            "turn_id": _sem_turn_id,
                            "top_score": _sem_top,
                            "total_ms": _sem_total_ms,
                            "retrieve_ms": 0,
                            "ttft_ms": _sem_total_ms,
                            "llm_ms": 0,
                            "cached": True,
                            "cache_layer": "semantic",
                        })
                        _sem_enrich_evt = _emit_enrich(
                            _sem_turn_id, question, _sem_text, float(_sem_top),
                        )
                        if _sem_enrich_evt:
                            yield _sem_enrich_evt
                        print(
                            f"[chat-cache] SEMANTIC HIT "
                            f"cos={_sem_hit['cosine']:.3f} "
                            f"age={int(_sem_hit['age_seconds'] // 60)}m "
                            f"total={_sem_total_ms}ms",
                            flush=True,
                        )
                        # Emit the `[chat-timing]` schema-stable line so
                        # downstream parsers (tool_rounds/tool_ms) don't see
                        # NaN on cache-hit turns.
                        print(
                            f"[chat-timing] model={_resolve_web_chat_model()} "
                            f"retrieve=0ms reform=0ms reform_outcome=skipped "
                            f"q_words={len(question.split())} wa=0ms "
                            f"llm_prefill=0ms llm_decode=0ms "
                            f"ttft={_sem_total_ms}ms total={_sem_total_ms}ms "
                            f"ctx_chars=0 tokens≈{len(_sem_text.split())} "
                            f"confidence={_sanitize_confidence(_sem_top):.3f} "
                            f"variants=0 tool_rounds=0 tool_ms=0 tool_names= "
                            f"cached=1 cache_layer=semantic",
                            flush=True,
                        )
                        # Hydrate LRU so next-turn exact-match hits O(1).
                        if _cache_key is not None:
                            try:
                                _chat_cache_put(_cache_key, {
                                    "text": _sem_text,
                                    "sources_items": _sem_sources,
                                    "top_score": _sem_top,
                                })
                            except Exception:
                                pass
                        try:
                            append_turn(sess, {
                                "q": question,
                                "a": _sem_text,
                                "paths": _sem_paths,
                                "top_score": _sem_top,
                                "turn_id": _sem_turn_id,
                            })
                            save_session(sess)
                        except Exception:
                            pass
                        try:
                            log_query_event({
                                "cmd": "web.chat.cached_semantic",
                                "q": question[:200],
                                "session": sess["id"],
                                "turn_id": _sem_turn_id,
                                "answered": True,
                                "t_total": round(_sem_total_ms / 1000.0, 3),
                                "cache_hit": True,
                                "cache_probe": _semantic_cache_probe,
                                "cache_layer": "semantic",
                                "cache_cosine": round(_sem_hit["cosine"], 4),
                                "cache_age_seconds": int(_sem_hit["age_seconds"]),
                                "paths": _sem_paths,
                                "top_score": _sem_top,
                                "device": _client_device,
                            })
                        except Exception:
                            pass
                        return
            except Exception as _sem_exc:
                print(
                    f"[chat-cache] semantic lookup failed: "
                    f"{type(_sem_exc).__name__}: {_sem_exc}",
                    flush=True,
                )
                # Cache lookup is a perf optimization — fall back to the
                # normal pipeline. Reset emb/hash so PUT also skips.
                _semantic_cache_emb = None
                _semantic_cache_hash = ""

        # Fail fast if ollama is in the "stuck-load" state — accepts HTTP
        # but never responds. Two-layer probe: /api/tags is the cheap
        # "daemon listening?" check; the deep chat probe catches the mode
        # where tags responds but chat hangs (seen 2026-04-19 — tags OK,
        # /api/chat never flushed first chunk). On deep-probe failure we
        # auto-heal via `brew services restart ollama` and re-probe once
        # before giving up, so the user's next /api/chat request doesn't
        # need a manual panic-button press.
        if not _ollama_alive(timeout=2.0) or not _ollama_chat_probe(timeout_s=6.0):
            print("[ollama-preflight] stuck-load detected — auto-restarting", flush=True)
            if not _ollama_restart_if_stuck() or not _ollama_chat_probe(timeout_s=8.0):
                yield _sse("error", {
                    "message": "Ollama no responde (stuck-load). Auto-restart falló. "
                    "Probá: brew services restart ollama",
                })
                # Audit 2026-04-26 (BUG #31): emitir `done` siempre tras
                # `error` para que el cliente cierre el spinner y libere
                # el input. Pre-fix: error puro dejaba EventSource
                # esperando indefinidamente.
                #
                # Update 2026-04-28 (BUG #31 wave-2): incluir `top_score:
                # 0.0` además de `error: true`. El frontend usa
                # `top_score < 0.10` como gate para `appendFallbackCluster`
                # (los 3 botones Google/YouTube/Wikipedia). Pre-fix sin
                # ese campo el cliente NO mostraba el cluster cuando había
                # error → el user veía el banner rojo y nada más, sin
                # escape hatch. Con top_score=0.0 el frontend SÍ activa
                # el fallback.
                yield _sse("done", {"error": True, "top_score": 0.0})
                return
            print("[ollama-preflight] recovered via restart", flush=True)

        # Kick off WhatsApp fetch in parallel with retrieve so the SQLite
        # round-trip (25-180ms) overlaps with the heavier retrieval work
        # instead of stacking sequentially before the LLM call.
        #
        # Gated on query intent (2026-04-22): el fetch sólo se dispara
        # cuando la pregunta menciona WhatsApp — mismo regex que el
        # check más abajo para decidir la inyección al prompt. Pre-fix
        # el submit ocurría SIEMPRE aunque en ~70% de queries el
        # resultado se descartaba sin usarse. Ahorra 25-180ms de I/O
        # SQLite al messages.db del bridge en la gran mayoría del
        # tráfico web. El check es un regex plano sobre `question`
        # (disponible desde el inicio del endpoint), sub-microsegundo.
        from concurrent.futures import ThreadPoolExecutor
        _wa_in_query = bool(
            re.search(r"\b(whatsapp|\bwa\b|mensaje|chat de|último[s]? chat)",
                      question, re.IGNORECASE)
        )
        _wa_executor: ThreadPoolExecutor | None = None
        _wa_future = None
        _t_wa_start = time.perf_counter()
        if _wa_in_query:
            # Crear el executor + submit en un try/except que garantice
            # cleanup si submit falla (raro: RuntimeError si el executor
            # quedó broken, ResourceError si no hay threads libres).
            # 2026-04-24 audit: pre-fix, si `submit` lanzaba después de
            # que `ThreadPoolExecutor()` tuvo éxito, el executor quedaba
            # "vivo" pero sin future y el finally downstream (que chequea
            # `if _wa_future is None`) ni siquiera entraba al path con
            # el shutdown. Thread leak silencioso. Ahora si falla submit,
            # shutdown inmediato y dejamos ambos en None — el handler
            # downstream trata el caso como "WA fetch no fired".
            try:
                _wa_executor = ThreadPoolExecutor(max_workers=1)
                _wa_future = _wa_executor.submit(_fetch_whatsapp_unread, 24, 8)
            except Exception as _exc:
                print(
                    f"[chat-wa-executor-submit-failed] {type(_exc).__name__}: {_exc}",
                    flush=True,
                )
                if _wa_executor is not None:
                    _wa_executor.shutdown(wait=False)
                    _wa_executor = None
                _wa_future = None

        # Conversation-aware reformulation is a qwen2.5:3b call that costs
        # 0.6-2s. Only worth paying for short follow-ups that rely on prior
        # context ("y eso?", "más sobre ella"). Standalone questions ≥4 words
        # without anaphora cues skip it — the original query is already a
        # good search target.
        #
        # Two-layer strategy when a follow-up IS detected:
        #   1. reformulate_query rewrites the pronoun using history.
        #   2. If the output STILL contains an unresolved pronoun (helper
        #      got confused by noisy history — e.g. a prior turn cited a
        #      different entity by mistake), fall back to concatenating
        #      the last user turn with the current one. Empirically that
        #      concat anchors retrieval to the right entity without
        #      another LLM call. Example: "tenes algo mas sobre ella?"
        #      after "tenes notas sobre Grecia?" → helper may still emit
        #      "algo mas sobre ella" → concat produces "tenes notas sobre
        #      Grecia? tenes algo mas sobre ella?" which retrieves the
        #      Grecia notes at top (~+0.20 rerank score).
        # Early intent classification for the `retrieving` hint.
        # `classify_intent` is pure regex, sub-millisecond. Paying it once
        # extra here buys us a contextual hint for the user — "Contando
        # notas…" / "Buscando por persona u organización…" etc — instead
        # of the generic ticker label. Zero-latency UX win on the ~2-10s
        # retrieve window.
        #
        # Gap addressed: measured p50 web retrieve = 2.8s, p90 = 25s.
        # Before, the user stared at "buscando…" without knowing WHAT was
        # being searched. Now the status label reflects the actual intent
        # path the retriever will take, so long waits feel less opaque.
        _early_intent = None
        _early_hint = None
        try:
            import rag as _rag_mod_hint
            _early_intent, _ = _rag_mod_hint.classify_intent(
                question, set(), set(),
            )
            _early_hint = _build_retrieve_hint(_early_intent)
        except Exception:
            pass
        _retrieving_status = {"stage": "retrieving"}
        if _early_hint:
            _retrieving_status["hint"] = _early_hint
        if _early_intent:
            _retrieving_status["intent"] = _early_intent
        yield _sse("status", _retrieving_status)

        _t_reform_start = time.perf_counter()
        # Hybrid follow-up resolution (2026-04-26 rework). Earlier versions
        # of this code path skipped reformulate_query entirely in favour of
        # a free concat of the prior user turn — the comment claimed
        # "concat empíricamente lograba mismo rerank score con recall igual
        # o mejor". Re-running the queries.yaml chains golden eval against
        # 4 variants (base / always-concat / extended-regex / reform LLM)
        # showed the assumption was false: chain_success was 41.7% on the
        # web path vs 58.3% on the CLI eval path that uses reform-LLM. The
        # gap was concentrated on elliptical follow-ups whose regex didn't
        # fire (e.g. "listame los gastos en pesos" after "Cuanto devo a la
        # visa?" — reproduced from session web:cee69e81829c on 2026-04-26).
        #
        # Hybrid strategy:
        #   (a) `_looks_like_followup` matches → concat last user turn with
        #       current question. Cheap (0ms), captures the obvious cases
        #       (pronouns, ellipsis, "y + article"). Empirically gains
        #       ~+8pp chain_success vs raw.
        #   (b) Topic-shift gate dropped history (cosine < 0.4) → use raw.
        #       History was already cleared because the user moved on.
        #   (c) History exists, regex didn't fire, cosine in [0.4, 0.7) →
        #       borderline. Run `reformulate_query` (qwen2.5:3b helper,
        #       ~1-2s) for explicit pronoun resolution + entity anchoring.
        #       This is the band where concat is too noisy and raw misses
        #       the prior topic — e.g. "y cómo las indexás" needs "las" →
        #       "embeddings" rewrite that the regex can't do.
        #   (d) High cosine (≥ 0.7) without regex match → paraphrase or
        #       self-contained rephrase, raw query is fine.
        #
        # The `_topic_shifted` block below this one (build_person_context +
        # detect_topic_shift) used to live AFTER this section. It moved up
        # so we can read the cosine value here for the (c) gate.
        # `_helper_idle_unload_qwen3b` is a separate idle-sweeper; this
        # call respects whatever VRAM budget the caller already
        # configured.
        from rag import build_person_context as _build_person_ctx
        from rag import detect_topic_shift as _detect_topic_shift
        _person_ctx = _build_person_ctx(question)
        _person_block = f"{_person_ctx}\n\n---\n\n" if _person_ctx else ""
        _topic_shifted = False
        _topic_shift_reason = "no-history"
        _topic_shift_cosine: float | None = None
        if history:
            _topic_shifted, _topic_shift_reason, _topic_shift_cosine = (
                _detect_topic_shift(
                    question, history, person_fired=bool(_person_ctx),
                )
            )
            if _topic_shifted:
                history = []

        search_question = question
        _reform_fired = False
        _reform_used_concat = False
        _reform_used_llm = False
        last_user_q = next(
            (m["content"] for m in reversed(history) if m.get("role") == "user"),
            None,
        ) if history else None
        if history and last_user_q:
            if _looks_like_followup(question):
                # (a) Regex match → concat (cheap, 0ms).
                search_question = f"{last_user_q} {question}"
                _reform_fired = True
                _reform_used_concat = True
            elif (
                _topic_shift_cosine is not None
                and TOPIC_SHIFT_COSINE <= _topic_shift_cosine < REFORM_COSINE_HIGH
            ):
                # (c) Borderline cosine band → reform LLM.
                try:
                    search_question = reformulate_query(question, history)
                    _reform_fired = True
                    _reform_used_llm = True
                except Exception as _exc:
                    # Fail-safe: fall back to concat (better than raw, +8pp
                    # chain_success vs raw per the eval). Log for ops.
                    print(
                        f"[reform] LLM call failed, falling back to concat: "
                        f"{type(_exc).__name__}: {_exc}",
                        flush=True,
                    )
                    search_question = f"{last_user_q} {question}"
                    _reform_fired = True
                    _reform_used_concat = True
            # else (d): high cosine ≥ 0.7 or cosine=None (anaphoric/short/
            # person already handled above) → raw query is fine.
        _t_reform_end = time.perf_counter()

        try:
            _t_retrieve_start = time.perf_counter()
            # multi_query=False: the qwen2.5:3b paraphrase call costs 1-3s
            # and only marginally improves recall on chat-style questions.
            # k=4 + rerank_pool=5: k=1 + pool=2 starved the LLM context —
            # ambiguous queries (e.g. "dame info sobre Grecia" as proper
            # noun) leaked unrelated top-1 chunks (Contacts/_index), LLM
            # faithfully cited wrong source. Pool=5 matches bge-reranker
            # batch alignment (~210ms on MPS), k=4 gives the LLM real
            # context + room for graph expansion. Reverted if eval regresses
            # below singles hit@5 76.19 (CI lower bound) or chains
            # chain_success < 16.67.
            # Self-citation guard: drop this session's own episodic note +
            # the entire conversations/ folder from candidates. CLAUDE.md
            # marks these as system artifacts; surfacing them as retrieval
            # sources creates a feedback loop where the just-written turn
            # answers the next one (confirmed 2026-04-19).
            _own_conv = _own_conversation_path(sid)
            _exclude = {_own_conv} if _own_conv else None
            # Classify intent here (regex-only, no LLM call) and thread it
            # through `multi_retrieve`. Pre 2026-04-22 the web path never
            # passed intent → retrieve() saw intent=None → adaptive
            # routing could never kick in, and the 3 `log_query_event`
            # sites below had no intent to log (42% of traffic lost intent
            # in extra_json). See tests/test_intent_logging_web.py.
            #
            # Reuse del early-hint classification (2026-04-22): el path de
            # arriba ya clasificó `question` para armar el status hint.
            # Cuando el query es standalone (no follow-up) `question ==
            # search_question` → podemos reusar `_early_intent` sin la
            # segunda call a classify_intent. Follow-ups concatenan
            # `last_user_q + question` → el intent puede diferir, así que
            # recomputamos. Ahorro estimado: ~30-50μs × 70% de tráfico
            # (mayoría son queries standalone).
            if search_question == question and _early_intent is not None:
                _intent_for_log = _early_intent
            else:
                try:
                    import rag as _rag_mod
                    _intent_for_log, _ = _rag_mod.classify_intent(
                        search_question, set(), set())
                except Exception:
                    _intent_for_log = None
            # List-intent override (2026-04-28): si la query pide
            # explícitamente una lista amplia ("mostrame todas las
            # notas X", "cuántas notas tengo sobre Y"), subimos k y
            # rerank_pool. k=4 era OK para queries de pregunta corta
            # pero con queries de "lista" hacía que el LLM viera 4
            # chunks y o (a) los repitiera sin contexto adicional, o
            # (b) alucinara para llenar la respuesta. multi_query=True
            # también: para "lista todas X" valen las paraphrases
            # ("notas sobre X", "archivos con X", "menciones de X")
            # — el costo de la 2da call al qwen2.5:3b reformer (1-3s)
            # es aceptable para queries de exploración.
            #
            # Iteración 1 (mismo día): k=12 + multi_query=True saturaba
            # el num_ctx=4096 — Playwright autónomo midió ttft=107820ms
            # con timeout en synthesis. El context post-tools llegaba a
            # ~30k chars y qwen2.5:7b prefill no terminaba a tiempo.
            # Compromiso: k=8 (2x el default sin list intent) + pool=15
            # da más recall sin saturar el modelo. El user que quiere
            # 12+ resultados puede iterar pidiendo "más" en el chat.
            _is_list_intent = bool(_LIST_INTENT_RE.search(search_question))
            _retrieve_k = 8 if _is_list_intent else 4
            _retrieve_pool = 15 if _is_list_intent else 5
            _retrieve_multi_query = _is_list_intent
            # `caller="web"` (2026-04-28): impressions del chat web son
            # user-initiated (cada vez que el user manda un mensaje al
            # /api/chat). El default era "cli" — funcionalmente equivalente
            # como user signal pero ahora distinguimos para poder splitear
            # métricas por canal (CLI vs PWA vs WhatsApp listener vía
            # serve.chat). Ver doc en `rag.retrieve()` y commit `fd97829`.
            result = multi_retrieve(
                vaults, search_question, _retrieve_k, None, history, None, False,
                multi_query=_retrieve_multi_query, auto_filter=True, date_range=None,
                rerank_pool=_retrieve_pool, exclude_paths=_exclude,
                exclude_path_prefixes=(
                    # Episodic memory: el writer dejó de escribir acá tras
                    # 2026-04-25 pero archivos viejos pueden seguir en el vault
                    # hasta el próximo `rag consolidate`. Mantenemos el
                    # prefix para que no entren al index retroactivamente.
                    "00-Inbox/conversations/",
                    # Nueva ubicación canónica desde 2026-04-25 — sistema vive
                    # bajo 99-AI/, fuera del PARA del user.
                    "04-Archive/99-obsidian-system/99-AI/conversations/",
                ),
                intent=_intent_for_log,
                caller="web",
            )
            # Normalizar shape del result ANTES de cualquier acceso
            # downstream. Garantiza len(docs) == len(metas) == len(scores),
            # metas dicts (nunca None/str), confidence float válido.
            # Defense-in-depth contra bugs en rag.py o Ollama mid-embed.
            # Ver `_validate_retrieve_result` docstring para detalle.
            result = _validate_retrieve_result(result)
            _t_retrieve_end = time.perf_counter()
        except Exception as exc:
            # 2026-04-28 wave-7: sanitizar el error message antes de emitirlo
            # al user. Repro Conv 6: "cómo funciona el sistema RAG" → assertion
            # interna `bm25_search llamado en paralelo — es GIL-serialised por
            # diseño (CLAUDE.md línea 126, medido 3× slower paralelo en M3 Max)`
            # se filtraba directo al chat. Reemplazamos por mensaje user-friendly
            # y loggeamos el original en web.log para debug.
            _sanitized_msg = _sanitize_error_for_user(exc, phase="retrieve")
            print(
                f"[chat-error-sanitized] phase=retrieve "
                f"exc_type={type(exc).__name__} "
                f"original={str(exc)[:200]!r} "
                f"sent_to_user={_sanitized_msg!r}",
                flush=True,
            )
            yield _sse("error", {"message": _sanitized_msg})
            # BUG #31 wave-2 (2026-04-28): top_score=0.0 dispara fallback
            # cluster en frontend (Google/YouTube/Wikipedia) — sin esto el
            # user ve solo el banner de error rojo sin escape hatch.
            yield _sse("done", {"error": True, "top_score": 0.0})  # BUG #31 — emitir done
            return

        # Pre-compute forced tools ONCE here so both the empty-bail and
        # low-conf-bypass gates below can check it without re-running
        # the regex. Propose-intent turns skip the pre-router entirely
        # (line ~4894) so mirror that gate here to avoid enabling a
        # bypass path the pre-router wouldn't honour.
        # 2026-04-24: capturamos la LISTA además del bool para poder
        # reusarla más abajo al armar el `_system_msgs` — si alguno de
        # los tools disparados es source-specific (gmail_recent,
        # calendar_ahead, reminders_due) le agregamos al LLM un hint
        # que fuerza "si la sección pedida está vacía, decilo antes de
        # fallback-ear a otra fuente". El tool loop de más abajo también
        # llama _detect_tool_intent con el mismo `question` — se podría
        # DRY reutilizando esta variable, pero por ahora el regex scan
        # es sub-microsegundo y preservamos esa call local para no
        # acoplar los dos sitios.
        #
        # Bug 2026-04-27 (audit "RAG - Flujo api-chat - auditoría" #2):
        # esta call estaba SIN try/except — si `_detect_tool_intent`
        # reventaba (regex compile inesperada, tipo de retorno raro,
        # bug nuevo en un matcher agregado), la excepción propagaba
        # hacia `gen()`, Starlette cerraba la conexión, el frontend
        # veía "stream ended" sin mensaje de error y la PWA quedaba
        # con spinner congelado. La SEGUNDA call al final del flujo
        # (línea ~9701, dentro del pre-router) ya estaba protegida por
        # un try/except envolvente. Esta primera call no lo estaba.
        # Fix: try/except local que emite `error + done(error=True)`
        # para mantener el contrato del SSE stream (cliente recibe
        # cierre limpio con un mensaje de error que puede mostrar).
        try:
            _forced_tool_pairs: list[tuple[str, dict]] = (
                [] if is_propose_intent else _detect_tool_intent(question)
            )
        except Exception as exc:
            print(f"[chat] _detect_tool_intent failed: {type(exc).__name__}: {exc}", flush=True)
            yield _sse("error", {"message": f"detector de intents falló: {exc}"})
            # BUG #31 wave-2: top_score=0.0 dispara fallback cluster.
            yield _sse("done", {"error": True, "top_score": 0.0})
            return

        # 2026-04-28 wave-8: anaphoric carry-over de tools del turno previo.
        # Repro Conv 4 T2: "y mañana?" tras "qué hago hoy?" → el LLM
        # alucinaba propose_reminder. Fix: si el turno previo fired tools
        # de read-intent (calendar_ahead, reminders_due, weather, etc.) Y
        # la query actual matchea un patrón anaphoric — temporal ("y
        # mañana?") o referencial ("y de eso qué puedo posponer?") —
        # re-fire los mismos tools con args ajustados.
        #
        # UNION semantics — el `_detect_tool_intent` ya pudo haber emitido
        # `reminders_due/calendar_ahead/whatsapp_pending` por matchear
        # "mañana"/"hoy", pero NO `weather` (no hay keyword genérico).
        # Si el turno previo fired weather y el actual es anafórico,
        # ADD weather al set actual sin descartar lo que ya hay.
        _q_stripped = question.strip()
        _is_anaphoric_temporal = bool(_ANAPHORIC_TEMPORAL_RE.match(_q_stripped))
        _is_anaphoric_reference = bool(_ANAPHORIC_REFERENCE_RE.match(_q_stripped))
        if (
            not is_propose_intent
            and _last_turn_tools
            and (_is_anaphoric_temporal or _is_anaphoric_reference)
        ):
            _read_intent_tools = {
                "calendar_ahead", "reminders_due", "whatsapp_pending",
                "gmail_recent", "weather", "whatsapp_list_scheduled",
                "drive_search",
            }
            _existing_names = {n for n, _ in _forced_tool_pairs}
            _carryover_added: list[tuple[str, dict]] = []
            for name in _last_turn_tools:
                if name in _read_intent_tools and name not in _existing_names:
                    _carryover_added.append(
                        (name, _resolve_anaphoric_args(name, question, _last_turn_weather_location))
                    )
                    _existing_names.add(name)
            if _carryover_added:
                _forced_tool_pairs = list(_forced_tool_pairs) + _carryover_added
                _kind = "temporal" if _is_anaphoric_temporal else "reference"
                print(
                    f"[chat-anaphoric-carryover] kind={_kind} q={question!r} "
                    f"prev_tools={_last_turn_tools} "
                    f"existing={[n for n,_ in _forced_tool_pairs[:len(_forced_tool_pairs)-len(_carryover_added)]]} "
                    f"added={[n for n,_ in _carryover_added]}",
                    flush=True,
                )
        _has_forced_tools = bool(_forced_tool_pairs)
        # Capturamos location de weather si el pre-router la trajo —
        # para persistir en el next turn.
        for _n, _a in _forced_tool_pairs:
            if _n == "weather" and isinstance(_a, dict) and _a.get("location"):
                _last_weather_location = _a["location"]
                break

        if not result["docs"]:
            # Propose-intent turns don't need vault context — the tool
            # loop (propose_reminder / propose_calendar_event) creates
            # things out of thin air. Bailing here with "empty" made the
            # handler return "Sin resultados relevantes." for inputs like
            # "el 26 de Mayo es el cumple de Astor" (2026-04-21 Fer F.
            # Playwright report) which is absurd — the user is declaring
            # a date, not asking about one. Fall through to the tool
            # phase when the detector flagged intent.
            # 2026-04-22 (Fer F. user report, "qué tengo para hacer esta
            # semana?"): también fall-through cuando el pre-router
            # matchea tools deterministas (reminders_due, calendar_ahead,
            # finance_summary, gmail_recent, weather). El retrieve puede
            # venir vacío por razones transitorias (cold-start, corpus
            # sin notas lexicalmente cercanas al query social) y bailar
            # con "Sin resultados relevantes." + link a Google cuando
            # los tools hubieran listado los pendientes/eventos reales
            # es user-hostile. El tool loop abajo reemplaza CONTEXTO
            # entero con la salida de los tools, así que el LLM responde
            # sobre data autoritativa aun si el retrieve no aportó.
            if not is_propose_intent and not _has_forced_tools:
                yield _sse("empty", {"message": "Sin resultados relevantes."})
                return

        # Retrieval signals — previously surfaced as a UI meta bar
        # ("🟡 media · 0.8 · N variantes · M nota(s)"). Removed from the
        # UI per user request; instead we pass these signals as context
        # to the LLM so it can frame the answer (e.g. hedge when the
        # confidence is low, or acknowledge when only one note exists).
        _conf = float(result["confidence"])
        _, _conf_label = _confidence_badge(_conf)
        # `.get("file", "")` defensive: aunque `_validate_retrieve_result`
        # garantiza que cada meta es un dict, no podemos asumir que
        # siempre tiene la key `"file"` (rag.py metadata schema podría
        # migrar y dejar notas viejas sin el campo). Set con "" como
        # fallback dedupea correctamente — múltiples metas sin file
        # colapsan a una única entrada vacía en el count.
        _n_notes = len({m.get("file", "") for m in result["metas"]})
        _n_variants = len(result.get("query_variants", []))
        _filters = result.get("filters_applied") or {}
        retrieval_signals = {
            "confidence_label": _conf_label,
            "confidence_score": round(_conf, 2),
            "n_notes": _n_notes,
            "n_variants": _n_variants,
            "filters": _filters,
        }

        # Mention hits (entity dictionary) → synthetic source rows so the
        # user can see which definitions were injected. `score` set to a
        # sentinel high value (5.0) and bar to all-filled — these aren't
        # ranked by the retriever, they're authoritative user-provided.
        # Uses rag._match_mentions_in_query (same resolver that feeds
        # build_person_context) so the UI reflects exactly which mention
        # notes were folded into the LLM preamble.
        from rag import _match_mentions_in_query as _match_mentions_fn
        from pathlib import Path as _Path
        _mention_paths = _match_mentions_fn(question)
        _mention_sources = [
            {
                "file": rel,
                "note": _Path(rel).stem,
                "folder": str(_Path(rel).parent),
                "score": 5.0,
                "bar": "■■■■■",
            }
            for rel in _mention_paths
        ]

        yield _sse("sources", {
            "items": (
                []
                if is_propose_intent
                else _mention_sources + [
                    {**_source_payload(m, s), "bar": _score_bar(float(s))}
                    for m, s in zip(result["metas"], result["scores"])
                ]
            ),
            "confidence": round(_sanitize_confidence(result["confidence"]), 3),
            "propose_intent": is_propose_intent,
        })

        # Low-confidence bypass — cuando el vault no tiene info útil
        # (conf < CONFIDENCE_RERANK_MIN), saltamos el ollama.chat call
        # (5-8s de cold prefill + streamed decode) y devolvemos un
        # template fijo con la pregunta textual del usuario. Motivos
        # (2026-04-22):
        #   a) latencia cae a <100ms (solo retrieve) — el sistema
        #      percibido "sabe" que no tiene data en vez de esperar para
        #      recibir un "no tengo info" igual de corto del LLM.
        #   b) preserva nombres propios: el LLM sin CONTEXTO bajaba a
        #      apellidos parecidos ("Bizarrap" → "Bizarra") intentando
        #      matchear algo del pretraining. Con el template echoing
        #      exacto del input no hay superficie para alucinar.
        # Gates (cualquiera en false → path normal):
        #   - conf < CONFIDENCE_RERANK_MIN (0.015, global gate en
        #     rag.py) → retrieve básicamente devolvió vacío.
        #   - NO mention hits: si el query pega con 99 Mentions @/ el
        #     usuario preguntó por una entidad conocida y la respuesta
        #     debe pasar por el LLM con el mention preamble.
        #   - NO propose_intent: los turns de "recordame X" / "agendá
        #     Y" usan el tool loop para crear cosas, no consultan el
        #     vault.
        # El `low_conf_bypass=True` en el `done` payload le indica al
        # frontend que renderee el cluster de fallback ("¿querés buscar
        # en Google/YouTube/Wikipedia?") en lugar del inline web-search
        # link del path weakAnswer.
        # 2026-04-22: gate extra `not _has_forced_tools`. Queries que
        # matchean el pre-router deterministic (e.g. "qué tengo para hacer
        # esta semana?" → reminders_due + calendar_ahead) NO deben caer al
        # template "No tengo info sobre '...' en tus notas." aunque el
        # retrieve haya bajado debajo de CONFIDENCE_RERANK_MIN — los tools
        # van a inyectar la data real. Sin este check, el user reporta
        # "Sin resultados" cuando el sistema sí tiene forma de responder.
        # Audit 2026-04-26 (BUG #6): per-source threshold en lugar
        # del global CONFIDENCE_RERANK_MIN=0.015. Si Phase 1.h baja el
        # gate de WA a 0.008, el web seguía cortando a 0.015 → falsos
        # refuses cross-source.
        _top_meta = (result.get("metas") or [{}])[0] if result.get("metas") else {}
        _top_src = _top_meta.get("source") if isinstance(_top_meta, dict) else None
        try:
            from rag import confidence_threshold_for_source as _conf_thresh_fn
            _conf_threshold = _conf_thresh_fn(_top_src) if _top_src else CONFIDENCE_RERANK_MIN
        except Exception:
            _conf_threshold = CONFIDENCE_RERANK_MIN
        _low_conf_bypass = (
            not is_propose_intent
            and not _mention_paths
            and not _has_forced_tools
            and float(result["confidence"]) < _conf_threshold
        )
        if _low_conf_bypass:
            _t_retrieve_ms = int((_t_retrieve_end - _t_retrieve_start) * 1000)
            # Escape mínimo de `"` para no romper el template cuando el
            # usuario mandó comillas literales. UTF-8 + emojis intactos.
            _q_safe = question.replace('"', '\\"')
            _bypass_text = f'No tengo info sobre "{_q_safe}" en tus notas.'
            print(
                f"[chat-bypass] conf={float(result['confidence']):.4f} "
                f"reason=low_conf q_words={len(question.split())} "
                f"retrieve_ms={_t_retrieve_ms}",
                flush=True,
            )
            yield _sse("status", {"stage": "generating"})
            # Stream en chunks de 40 chars — mismo shape que el
            # cache-hit path (~línea 3800). El UI ya dispara el
            # token-ticker en el primer delta, así que la percepción
            # es idéntica a una respuesta "en vivo" aunque sea canned.
            for _i in range(0, len(_bypass_text), 40):
                yield _sse("token", {"delta": _bypass_text[_i:_i + 40]})
            _bypass_turn_id = new_turn_id()
            _bypass_total_ms = int((time.perf_counter() - _t0) * 1000)
            _bypass_top_score = round(
                _sanitize_confidence(result["confidence"]), 3,
            )
            yield _sse("done", {
                "turn_id": _bypass_turn_id,
                "top_score": _bypass_top_score,
                "total_ms": _bypass_total_ms,
                "retrieve_ms": _t_retrieve_ms,
                "ttft_ms": _bypass_total_ms,
                "llm_ms": 0,
                "low_conf_bypass": True,
                "bypassed": True,
            })
            # Emitimos un `[chat-timing]` shape-stable para los parsers
            # downstream (dashboards + grep-tooling asumen una línea por
            # turn). Mismo approach que el cache-hit path (~3820).
            # `bypassed=1` como tag terminal distingue del cached=1.
            print(
                f"[chat-timing] model={_resolve_web_chat_model()} "
                f"retrieve={_t_retrieve_ms}ms "
                f"reform=0ms reform_outcome=skipped "
                f"q_words={len(question.split())} "
                f"wa=0ms llm_prefill=0ms llm_decode=0ms "
                f"ttft={_bypass_total_ms}ms total={_bypass_total_ms}ms "
                f"ctx_chars=0 tokens≈{len(_bypass_text.split())} "
                f"confidence={_bypass_top_score:.3f} variants=0 "
                f"tool_rounds=0 tool_ms=0 tool_names= bypassed=1",
                flush=True,
            )
            # Persistimos la conversación + logueamos igual que el path
            # normal para que analytics y el episodic writer vean el
            # turno. Tests garantizan que bypass=True queda en el log.
            try:
                append_turn(sess, {
                    "q": question,
                    "a": _bypass_text,
                    "paths": [],
                    "top_score": _bypass_top_score,
                    "turn_id": _bypass_turn_id,
                    "low_conf_bypass": True,
                })
                save_session(sess)
            except Exception:
                pass
            try:
                log_query_event({
                    "cmd": "web.chat.low_conf_bypass",
                    "turn_id": _bypass_turn_id,
                    "session": sess["id"],
                    "q": question,
                    "paths": [],
                    "scores": [],
                    "top_score": _bypass_top_score,
                    "t_retrieve": round(_t_retrieve_ms / 1000.0, 3),
                    "t_gen": 0.0,
                    "low_conf_bypass": True,
                    # Intent echoed from the retrieve result so telemetry
                    # is symmetric with the `cmd=web` path above.
                    "intent": result.get("intent") if isinstance(result, dict) else None,
                    "device": _client_device,
                    # Stage timing parity with the main web path
                    # (2026-04-22). In the bypass path there's no LLM
                    # call — ttft_ms == total_ms and llm_* are zero.
                    "ttft_ms": int(_bypass_total_ms),
                    "llm_prefill_ms": 0,
                    "llm_decode_ms": 0,
                    "total_ms": int(_bypass_total_ms),
                })
            except Exception:
                pass
            _spawn_conversation_writer(
                target_args=(
                    VAULT_PATH,
                    sess["id"],
                    req.question,
                    _bypass_text,
                    result["metas"],
                    result["scores"],
                    result["confidence"],
                ),
                name=f"conv-writer-{_bypass_turn_id[:8]}",
            )
            _enrich_evt = _emit_enrich(
                _bypass_turn_id, question, _bypass_text,
                float(result["confidence"]),
            )
            if _enrich_evt:
                yield _enrich_evt
            return

        is_multi = len(vaults) > 1
        # Cap each chunk at 500 chars on the fast path. With 4 chunks that's
        # ~2000 chars ≈ 500 tokens de contexto — chico para que prefill
        # uncached < 6s, grande para que REGLA 4 obligatoria aterrice.
        _WEB_CHUNK_CAP = 500
        # Audit security 2026-04-26 (HIGH): pre-fix el fast-path emitía
        # chunks al LLM con f-string crudo, BYPASSEANDO `_format_chunk_for_llm`
        # → no `_redact_sensitive` (OTPs/CBU/tokens leakean al LLM) ni
        # fences `<<<CHUNK>>>...<<<END_CHUNK>>>` (la REGLA 0 de "chunks son
        # DATA no instrucciones" no aplica). Cross-source corpus (Gmail/
        # WhatsApp) puede meter prompt-injection en este path. Fix:
        # delegar al helper como los otros 4 callers (CLI query, chat,
        # serve, build_progressive_context).
        from rag import _format_chunk_for_llm as _rag_format_chunk
        context = "\n\n---\n\n".join(
            (f"[vault: {m.get('_vault', '?')}]\n" if is_multi else "")
            + _rag_format_chunk(d[:_WEB_CHUNK_CAP], m, role="nota")
            for d, m in zip(result["docs"], result["metas"])
        )

        # Collect the WA fetch we kicked off before retrieve. By now it's
        # almost always done (retrieve dominates), so .result() is a no-op.
        # Telemetry: log the post-retrieve wait, not wall time — so a 0ms
        # `wa_wait` means the SQLite read overlapped fully with retrieve.
        _t_wa_wait_start = time.perf_counter()
        if _wa_future is None:
            # WA fetch wasn't dispatched — query doesn't mention WhatsApp.
            # Empty list signals the downstream injection block to skip.
            wa_recent = []
        else:
            try:
                wa_recent = _wa_future.result(timeout=2.0)
            except Exception:
                wa_recent = []
            finally:
                if _wa_executor is not None:
                    _wa_executor.shutdown(wait=False)
        _t_wa_end = time.perf_counter()
        _t_wa_wait_ms = int((_t_wa_end - _t_wa_wait_start) * 1000)
        # `_wa_in_query` ya se computó arriba (antes del submit) — mismo
        # regex. Pre-2026-04-22 se recalculaba acá con mismo valor; el fix
        # del fetch condicional lo movió al inicio del endpoint para
        # gate-ear también la submit del WA fetch (no sólo la injection).
        if wa_recent and _wa_in_query:
            total_msgs = sum(int(w.get("count", 0)) for w in wa_recent)
            wa_block_lines = [
                f"[contexto auxiliar: WhatsApp últimas 24h · "
                f"{len(wa_recent)} chats · {total_msgs} msgs]"
            ]
            for w in wa_recent[:8]:
                snip = (w.get("last_snippet") or "").strip()[:100]
                snip_part = f' — "{snip}"' if snip else ""
                wa_block_lines.append(
                    f"- {w.get('name','?')} ({w.get('count',0)} "
                    f"msgs){snip_part}"
                )
            context = context + "\n\n---\n\n" + "\n".join(wa_block_lines)

        # Retrieval signals block removido 2026-04-18: costaba ~300 chars
        # = ~75 tokens = ~150ms de prefill uncached en CADA request. El
        # hedging que calibraba (no-refusal, humildad en confianza baja,
        # engancharse con contexto débil) ya está expresado en REGLA 1 +
        # REGLA 4 del _WEB_SYSTEM_PROMPT (que está byte-identical + cached).
        # Empíricamente el modelo respeta esas reglas sin repetir las
        # señales cada turno — redundante con el system prompt.
        # retrieval_signals dict sigue disponible para telemetría
        # (_n_notes, confidence) pero no se inyecta al prompt.

        # Person-mention enrichment: when the query names someone listed in
        # 99 Mentions @/, prepend that note's body + Apple Contacts data so
        # the LLM knows who the user is talking about before reading the
        # retrieved chunks. Pre-LLM only — never leaked to episodic notes.
        # NOTE: this supersedes the earlier web-local `_mentions_block` —
        # `build_person_context` does the same job with richer output
        # (Apple Contacts metadata + structured format).
        # Person-mention enrichment + topic-shift gate moved up to the
        # search-question selection block (~line 8410) so the cosine value
        # from `detect_topic_shift` can drive the hybrid reform-vs-concat
        # routing. `_person_ctx`, `_person_block`, `_topic_shifted`, and
        # `_topic_shift_reason` are already populated at this point. See
        # the "Hybrid follow-up resolution (2026-04-26 rework)" comment for
        # rationale.

        # Build messages so the system prompt is BYTE-IDENTICAL across all
        # requests — that's the whole game for ollama prefix caching. The
        # context (which changes per query) goes in the user message AFTER
        # any history. With history, this puts the cacheable prefix as:
        #   [system (cached)] + [history turns (partially cached)] + user
        #
        # EXCEPTION: create-intent ("recordame X", "mañana viene Y a casa,
        # calendarizalo"). Vault retrieval is NOISE for those — the LLM
        # otherwise engancha con notas tangenciales and hallucinates
        # "open file 04-Archive/Calendar.md" instead of calling the
        # propose_* tool. Skip the CONTEXTO block entirely in that case.
        if is_propose_intent:
            user_content = f"{_person_block}{question}"
        else:
            user_content = (
                f"{_person_block}CONTEXTO:\n{context}\n\n"
                f"PREGUNTA: {question}\n\nRESPUESTA:"
            )
        # Tool-deciding messages: include _WEB_TOOL_ADDENDUM as a second
        # system message so it's byte-identical across tool-deciding calls
        # (ollama prefix cache stays warm across rounds). The final
        # streaming call below STRIPS the addendum + `tools=` param so its
        # cache state matches the pre-tool-calling era byte-for-byte.
        #
        # On create-intent turns we REPLACE the full stack with just the
        # focused override — the normal _WEB_SYSTEM_PROMPT + addendum pin
        # the model to "summarise the vault" behaviour that competes with
        # the tool call. Prefix cache is cold for that first create-intent
        # turn, acceptable cost given how rare they are compared to reads.
        #
        # History is also SKIPPED on create-intent. If previous turns in
        # the same session show the assistant responding with text (no
        # tool_calls) — which happened constantly during debug before the
        # fix — the model mirrors that pattern and refuses to emit
        # tool_calls this turn either (tool_rounds=0 logged). A create
        # intent is a self-contained action anyway: no narrative continuity
        # with prior turns, no pronouns to resolve against earlier context.
        if is_propose_intent:
            _system_msgs: list[dict] = [
                {"role": "system",
                 "content": _build_propose_create_override(datetime.now())},
            ]
            _turn_history: list[dict] = []
        else:
            _system_msgs = [
                {"role": "system", "content": _WEB_SYSTEM_PROMPT},
                {"role": "system", "content": _WEB_TOOL_ADDENDUM},
            ]
            # Source-specific intent hint (2026-04-24, Fer F. user report):
            # si el pre-router fired un tool source-specific (gmail_recent,
            # calendar_ahead, reminders_due), agregamos un 3er system msg
            # turn-scoped que le dice al LLM "el user preguntó por X, anclá
            # ahí, si está vacío decilo antes de fallback-ear". Antes el LLM
            # veía "### Mails" vacía en CONTEXTO y seguía respondiendo sobre
            # WhatsApp/notas como si fueran la respuesta principal, sin
            # reconocer que el user pidió mails específicamente.
            #
            # Prefix cache impact: el hint es turn-scoped y se agrega en
            # ~10-20% del tráfico (estimado por cobertura de las 3 regex
            # source-specific). Para esos turns el prefix cache de ollama
            # queda cold en el 3er system msg (los 2 primeros siguen
            # byte-identical). Tradeoff aceptable: el hint evita respuestas
            # user-hostile — prefijos 1&2 siguen dominando la latencia.
            _src_hint = _build_source_intent_hint(
                [name for name, _args in _forced_tool_pairs]
            )
            if _src_hint:
                _system_msgs.append({"role": "system", "content": _src_hint})
            _turn_history = history or []
        tool_messages: list[dict] = (
            _system_msgs
            + _turn_history
            + [{"role": "user", "content": user_content}]
        )

        # Log context size for prefill analysis (pre-tool expansion).
        _ctx_chars = sum(len(m.get("content","")) for m in tool_messages)

        # Fast-path LLM options.
        # num_ctx: measured from web.log (75 timing rows, 2026-04-17):
        #   P50 ctx ≈ 4889 chars / 1.35 char-per-token ≈ 3622 tok
        #   P95 ctx ≈ 6065 chars ≈ 4492 tok
        #   P99 ctx ≈ 7751 chars ≈ 5742 tok
        # Budget breakdown per request:
        #   system prompt ~1100 tok + retrieved context ~2500–4500 tok
        #   + history (trimmed) + current query
        # 2560 was silently truncating every request above ~1900 tok of
        # retrieved context. 5120 covers P99 (5742 tok) with ~10% headroom.
        # Don't raise further without evidence — larger ctx has real prefill
        # cost on command-r:35b.
        # num_predict 256: REGLA 5 (web prompt) pide 2-4 oraciones, lo que
        # aterriza en ~80-180 tok; con la nueva REGLA 4 (preservar URLs) las
        # respuestas pueden llegar a ~200 tok. Cap previo de 60 cortaba a
        # media oración. Decode budget: qwen2.5:7b ≈ 30 tok/s → 256 tok ≈ 8s.
        # Prefix caching is automatic in ollama when system prompts match
        # byte-for-byte across requests — see _WEB_SYSTEM_PROMPT for the
        # stable definition that keeps the cache warm. cache_prompt was
        # tested as an explicit option but is not in ollama's accepted set
        # and made some calls hang silently — leave it out.
        #
        # Fast-path dispatch (2026-04-22, Improvement #3 Fase D web-wire):
        # Cuando retrieve() devuelve `fast_path=True` (adaptive routing
        # judged it semantic + high confidence OR WA majority + threshold
        # per `RAG_WA_FAST_PATH`), switch to _LOOKUP_MODEL (qwen2.5:3b) +
        # `_LOOKUP_NUM_CTX` (4096). Pre-fix el endpoint ignoraba el flag
        # y corría qwen2.5:7b siempre — audit medio 8.5s gen p50 en web
        # vs 3.1s potencial en fast-path (CLI data, 7d). El switch es
        # local al streaming block (tool loop de arriba sigue usando el
        # modelo completo porque queremos tool-calling robustness).
        # ── Mode gate (2026-04-24) ──────────────────────────────────
        # `mode` is a user-facing knob sent from the web UI that overrides
        # the adaptive `fast_path` marker produced by `retrieve()`. Three
        # branches:
        #
        #   - mode="auto"  → honour whatever retrieve() decided
        #                    (`result["fast_path"]`). This is the default
        #                    and mirrors pre-mode behaviour.
        #   - mode="fast"  → force `_fast_path=True` (qwen2.5:3b +
        #                    _LOOKUP_NUM_CTX). User explicitly asks for
        #                    the quick/lookup model even if retrieve()
        #                    judged the query complex.
        #   - mode="deep"  → force `_fast_path=False` (full chat model +
        #                    _WEB_CHAT_NUM_CTX). User explicitly asks for
        #                    the full model even when retrieve() thought
        #                    fast-path was fine.
        #
        # Invalid / unknown values → silent fallback to "auto". We never
        # 400 here: old curl scripts, MCP clients, and PWA builds that
        # pre-date this knob keep working transparently.
        #
        # IMPORTANT: the `_fast_path_downgraded` logic below (pre-router
        # detects tools + context would bloat) stays ORTHOGONAL to `mode`.
        # Even with `mode="fast"`, if the pre-router fires tools the
        # downgrade kicks in to protect prefill latency. That's a safety
        # rail against context bloat that users shouldn't be able to
        # bypass by picking the wrong mode — the downgrade isn't about
        # what the user wanted, it's about what prefill can physically
        # handle on the small model with 3K+ tokens of tool output.
        # `mode_origin` captures whether the client sent ANY value for the
        # field (including garbage) vs. not sending it at all. This lets
        # analytics measure how often users override the adaptive routing
        # separately from how often they hit us with broken payloads.
        mode_origin = "request" if req.mode is not None else "default"
        _raw_mode = (req.mode or "").strip().lower()
        if _raw_mode in {"auto", "fast", "deep"}:
            mode = _raw_mode
        else:
            # Unknown / empty-after-strip / accented / typo → silent "auto".
            # Covers "rápido", "thinking", "", whitespace-only, etc. We
            # never 400: old curl scripts, MCP, and pre-feature PWA builds
            # that don't send a mode (or send a different vocabulary) must
            # keep working unchanged.
            mode = "auto"

        if mode == "fast":
            _fast_path = True
        elif mode == "deep":
            _fast_path = False
        else:  # "auto"
            _fast_path = bool(result.get("fast_path", False))

        # Propose-intent hard-override: cualquier query que requiera un
        # propose_* tool (reminder / calendar_event / whatsapp_send)
        # necesita tool-calling confiable, y qwen2.5:3b (el modelo del
        # fast_path) NO lo hace bien — entra en un loop de sampling que
        # puede tardar >120s y timeoutea. Observed 2026-04-24 (Fer F.):
        # "Enviale un mensaje a Grecia [...texto largo...]" con
        # is_propose_intent=True y retrieve().fast_path=True corría con
        # qwen2.5:3b, el tool loop loopeaba y el turno fallaba con
        # "LLM falló: timed out" tras el budget de _OLLAMA_TOOL_TIMEOUT.
        # El tradeoff: +3-5s de prefill del 7b vs. turno que completa vs
        # turno que crashea → claramente vale el upgrade.
        if is_propose_intent and _fast_path:
            _fast_path = False
        # Tracks si hubo downgrade runtime por pre-router tools (ver bloque
        # de pre-router más abajo). False por default; el log del endpoint
        # lo expone como `fast_path_downgraded` en extra_json para que el
        # analytics pueda medir cuán seguido el adaptive routing pierde
        # efectividad por entradas híbridas tool+semantic. El downgrade
        # corre aunque el user haya pedido `mode="fast"` — es protección
        # contra prefill bloat, no una decisión de routing.
        _fast_path_downgraded = False
        _web_model_full = _resolve_web_chat_model()
        _web_model = _LOOKUP_MODEL if _fast_path else _web_model_full
        _web_num_ctx = _LOOKUP_NUM_CTX if _fast_path else _WEB_CHAT_NUM_CTX
        _WEB_CHAT_OPTIONS = {
            **CHAT_OPTIONS,
            "num_ctx": _web_num_ctx,
            "num_predict": 256,
        }

        yield _sse("status", {"stage": "generating"})

        # MPS flush sleep removed 2026-04-18: the 20-40s GPU contention
        # it defended against was actually caused by num_ctx mismatch
        # (5120 vs 4096 loaded) forcing ollama KV reinit — fixed at commit
        # 79f6b8e. Without the sleep, warm /api/chat drops by 200ms cleanly.
        # If prefill variance returns to 20s+ range, first suspect num_ctx
        # drift, not MPS contention.

        print(
            f"[chat-model-keepalive] model={_web_model} keep_alive={OLLAMA_KEEP_ALIVE}"
            f" num_ctx={_WEB_CHAT_OPTIONS['num_ctx']} fast_path={_fast_path}"
            f" mode={mode} mode_origin={mode_origin}",
            flush=True,
        )

        # ── ollama-native tool loop ───────────────────────────────────
        # Up to 3 tool-deciding rounds against CHAT_TOOLS. Each round is a
        # non-streaming ollama.chat with tools= set. If the LLM emits
        # tool_calls we execute them (serial bucket first, then parallel
        # bucket via ThreadPoolExecutor) and append `role:"tool"` messages
        # before the next round. When tool_calls comes back empty we break
        # out and fall through to the plain streaming call below.
        #
        # tool_rounds = rounds where ≥1 tool actually executed. When the
        # LLM returns no tool_calls in round 1, tool_rounds stays 0 —
        # that's "no tools needed".
        _TOOL_ROUND_CAP = 3
        tool_rounds = 0
        tool_ms_total = 0
        tool_names_called: list[str] = []

        def _unwrap_args(args):
            """command-r wraps args as {"tool_name":..., "parameters":{...}}.
            Unwrap so we can call the Python function directly (copied from
            rag.py `rag do`)."""
            if isinstance(args, dict) and "parameters" in args and isinstance(args["parameters"], (dict, str)):
                params = args["parameters"]
                if isinstance(params, str):
                    try:
                        params = json.loads(params)
                    except Exception:
                        params = {}
                args = params
            if not isinstance(args, dict):
                return {}
            return {k: v for k, v in args.items() if v not in ("", None)}

        try:
            from concurrent.futures import ThreadPoolExecutor as _ToolExecutor
            from concurrent.futures import as_completed as _as_completed

            # ── Pre-router: keyword-forced tools ─────────────────────────
            # Runs BEFORE the LLM tool-deciding call. Guarantees that
            # unambiguous queries (gastos → finance_summary, pendientes →
            # reminders_due, etc.) always fetch fresh data — the LLM's
            # default bias (REGLA 1 "engancháte con CONTEXTO") previously
            # swallowed these queries into stale vault summaries.
            # Results are appended as a single `role:"system"` block with
            # an explicit "DATOS FRESCOS" prefix so the LLM treats them as
            # authoritative over the retrieved vault context.
            #
            # Exception: create-intent queries ("mañana tengo daily a las
            # 10am") match `reminders_due` / `calendar_ahead` via the
            # shared `_PLANNING_PAT` (any "mañana" triggers it). Feeding
            # those read-intent results into context when the user is
            # CREATING something makes the LLM hallucinate that the event
            # already exists. Skip the pre-router entirely when propose
            # intent is detected — let the LLM decide cleanly.
            _propose_intent = is_propose_intent
            # 2026-04-28 wave-8: usamos `_forced_tool_pairs` ya computados al
            # entrar a gen() (línea ~10362) en vez de re-correr
            # `_detect_tool_intent` acá, porque el original perdía el carry-
            # over anafórico (Conv B T2 "y en Barcelona?": el carryover
            # populaba _forced_tool_pairs con [('weather', {'location': 'Barcelona'})]
            # pero esta línea lo ignoraba y re-detectaba [], saltando la
            # ejecución del tool entera).
            _forced_tools = [] if _propose_intent else list(_forced_tool_pairs)
            if _forced_tools:
                _f_serial = [(n, a) for n, a in _forced_tools if n not in PARALLEL_SAFE]
                _f_parallel = [(n, a) for n, a in _forced_tools if n in PARALLEL_SAFE]
                # Fan-out: emit all `tool` events upfront so the frontend
                # renders chips before any blocks on execution.
                for _n, _a in _forced_tools:
                    yield _sse("status", {"stage": "tool", "name": _n, "args": _a})
                _forced_results: list[tuple[str, str, int]] = []
                _pre_serial_sum_ms = 0
                _pre_parallel_max_ms = 0
                # Serial bucket (vault-touching tools under GIL/MPS).
                for _n, _a in _f_serial:
                    _t0 = time.perf_counter()
                    # Eval 2026-04-28: inyectar el query original al post-LLM
                    # para que los validators (_validate_calendar_recurrence,
                    # _validate_reminder_*) puedan recuperar info que el LLM
                    # dropeó.
                    if _n in ("propose_calendar_event", "propose_reminder") and "_original_query" not in _a:
                        _a["_original_query"] = question
                    try:
                        _res = TOOL_FNS[_n](**_a)
                    except Exception as _exc:
                        _res = f"Error: {_exc}"
                    _ms = int((time.perf_counter() - _t0) * 1000)
                    _pre_serial_sum_ms += _ms
                    _forced_results.append((_n, str(_res), _ms))
                    yield _sse("status", {"stage": "tool_done", "name": _n, "ms": _ms})
                    if _prop := _maybe_emit_proposal(_n, str(_res)):
                        yield _prop
                    tool_names_called.append(_n)
                # Parallel bucket (pure IO fetchers, no shared state).
                if _f_parallel:
                    def _run_tool(name: str, args: dict):
                        _t0 = time.perf_counter()
                        # Eval 2026-04-28: inyectar el query original (ver
                        # comentario en el bucket serial arriba).
                        if name in ("propose_calendar_event", "propose_reminder") and "_original_query" not in args:
                            args["_original_query"] = question
                        try:
                            _r = TOOL_FNS[name](**args)
                        except Exception as _exc:
                            _r = f"Error: {_exc}"
                        return name, str(_r), int((time.perf_counter() - _t0) * 1000)
                    with _ToolExecutor(max_workers=5) as _pool:
                        _futs = [_pool.submit(_run_tool, n, a) for n, a in _f_parallel]
                        for _fut in _as_completed(_futs):
                            _n, _res, _ms = _fut.result()
                            _pre_parallel_max_ms = max(_pre_parallel_max_ms, _ms)
                            _forced_results.append((_n, _res, _ms))
                            yield _sse("status", {"stage": "tool_done", "name": _n, "ms": _ms})
                            if _prop := _maybe_emit_proposal(_n, _res):
                                yield _prop
                            tool_names_called.append(_n)
                # Replace CONTEXTO entirely with tool output. Prior attempts
                # (append as system msg / prepend to user) failed: REGLA 1
                # in _WEB_SYSTEM_PROMPT ("engancháte SIEMPRE con el
                # CONTEXTO") pinned the LLM on vault retrieval. When the
                # pre-router fires we have fresh authoritative data — the
                # vault context is noise. Replacing the CONTEXTO block (the
                # canonical anchor for REGLA 1) makes the LLM use the tool
                # output as its sole evidence.
                #
                # Each tool output goes through `_format_forced_tool_output`
                # to render it as tidy markdown (Spanish bucket labels,
                # dedup, explicit empty states) instead of raw JSON under a
                # `## {tool_name}` header. Pre-fix qwen2.5:7b dropped
                # undated items and occasionally seeded `[[calendar_ahead]]`
                # citation artifacts because the tool name leaked as a
                # wikilink-ish token. See helper docstring for details.
                #
                # 2026-04-24 (iteración 2, Fer F. report): excepción al
                # replacement cuando TODOS los tools disparados vinieron
                # vacíos — en ese caso preservamos el CONTEXTO original del
                # vault y agregamos el tool output como sección explícita
                # "CONSULTAS EN VIVO (todas vacías)". Racional: si el user
                # preguntó por "últimos mails" y `gmail_recent` vino vacío,
                # pero el retrieve pulló `03-Resources/Gmail/2026-04-22.md`
                # (digest indexado de mails), queremos que el LLM pueda
                # usar ese digest como fallback concreto en vez de
                # contestar "te dejo otras fuentes" en abstracto. Con el
                # tool output reemplazando, el vault material se descartaba
                # y el LLM no tenía de qué resumir.
                # Tool output truncation (eval 2026-04-28 fix): outputs grandes de
                # gmail/whatsapp_search/drive disparaban timeout de 45-90s en la 2da
                # ronda del LLM. Recortamos a top-N items + summary para acelerar la
                # síntesis. Lazy import para deploy parcial.
                try:
                    from rag._tool_output_helpers import truncate_tool_output_for_synthesis
                except Exception:
                    truncate_tool_output_for_synthesis = lambda name, raw, **kw: raw  # type: ignore

                _datos_block = ""
                for _n, _res, _ in _forced_results:
                    _res_truncated = truncate_tool_output_for_synthesis(_n, _res)
                    _datos_block += "\n" + _format_forced_tool_output(_n, _res_truncated) + "\n"
                _all_tools_empty = all(
                    _is_empty_tool_output(_n, _res)
                    for _n, _res, _ in _forced_results
                )
                for _msg in reversed(tool_messages):
                    if _msg.get("role") == "user":
                        if _all_tools_empty:
                            _msg["content"] = (
                                f"{_person_block}CONTEXTO:\n{context}\n\n"
                                f"CONSULTAS EN VIVO (todas vacías — el "
                                f"pre-router consultó estas fuentes y no "
                                f"encontró nada reciente):\n{_datos_block}\n\n"
                                f"PREGUNTA: {question}\n\nRESPUESTA:"
                            )
                        else:
                            _msg["content"] = (
                                f"CONTEXTO (datos en vivo, no del vault):\n{_datos_block}\n\n"
                                f"PREGUNTA: {question}\n\nRESPUESTA:"
                            )
                        break
                tool_rounds += 1
                tool_ms_total += _pre_serial_sum_ms + _pre_parallel_max_ms

                # External sources re-emit (2026-04-24, iter 3 + 4):
                # Cuando drive_search / whatsapp_pending disparan, agregamos
                # sus resultados como sources para que la UI muestre un link
                # directo al doc/chat *antes* de las fuentes del vault. El
                # frontend reemplaza la lista en cada evento — esta segunda
                # emisión pisa la inicial. `_drive_source_items` primero
                # (Drive suele tener el match más específico), luego
                # `_wa_source_items` (chats pendientes). Dedupe por `file`
                # lo maneja el frontend (`seen.add(s.file)`).
                _drive_source_items: list[dict] = []
                _wa_source_items: list[dict] = []
                for _n, _res, _ in _forced_results:
                    if _n == "drive_search":
                        try:
                            _drive_payload = json.loads(_res)
                        except Exception:
                            continue
                        if not isinstance(_drive_payload, dict):
                            continue
                        for _f in (_drive_payload.get("files") or [])[:5]:
                            if not isinstance(_f, dict):
                                continue
                            _link = str(_f.get("link") or "").strip()
                            _name = str(_f.get("name") or "").strip()
                            if not _link or not _link.startswith(("http://", "https://")):
                                continue
                            _drive_source_items.append({
                                "file": _link,
                                "note": _name or "(sin nombre)",
                                "folder": "Google Drive",
                                "score": 5.0,
                                "bar": "■■■■■",
                            })
                    elif _n == "whatsapp_pending":
                        try:
                            _wa_payload = json.loads(_res)
                        except Exception:
                            continue
                        if not isinstance(_wa_payload, list):
                            continue
                        for _chat in _wa_payload[:5]:
                            if not isinstance(_chat, dict):
                                continue
                            _name = str(_chat.get("name") or "").strip()
                            _jid = str(_chat.get("jid") or "").strip()
                            if not _name or not _jid:
                                continue
                            # JID → deeplink al chat. Tres formas:
                            #   - `@s.whatsapp.net` = DM, phone en prefix
                            #     → `https://wa.me/<phone>` (web/app).
                            #   - `@g.us` = grupo → WhatsApp Web no
                            #     tiene deeplink a grupos públicos,
                            #     pero usamos un hash fragment con el
                            #     jid para que cada grupo tenga un URL
                            #     único (evita colapso en el dedup del
                            #     frontend que usa `s.file` como key).
                            #     El browser ignora el fragment al
                            #     resolver, WA Web abre al inbox.
                            #   - `@lid` = participante sin resolver en
                            #     grupo, mismo tratamiento que grupo.
                            _phone = _jid.split("@")[0] if "@" in _jid else ""
                            if _jid.endswith("@s.whatsapp.net") and _phone.isdigit():
                                _link = f"https://wa.me/{_phone}"
                            else:
                                _link = f"https://web.whatsapp.com/#{_jid}"
                            _wa_source_items.append({
                                "file": _link,
                                "note": _name,
                                "folder": "WhatsApp",
                                "score": 5.0,
                                "bar": "■■■■■",
                            })
                if _drive_source_items or _wa_source_items:
                    yield _sse("sources", {
                        "items": (
                            _drive_source_items
                            + _wa_source_items
                            + _mention_sources
                            + [
                                {**_source_payload(_m, _s), "bar": _score_bar(float(_s))}
                                for _m, _s in zip(result["metas"], result["scores"])
                            ]
                        ),
                        "confidence": round(_sanitize_confidence(result["confidence"]), 3),
                        "propose_intent": False,
                    })

                # Fast-path downgrade cuando el pre-router fired tools.
                # Racional (2026-04-24, medido en prod el 2026-04-23):
                # `_fast_path` fue calibrado por `retrieve()` para queries
                # semánticas simples donde el CONTEXTO es ~500 tok de vault
                # chunks — qwen2.5:3b es eficiente ahí. Pero cuando el
                # pre-router matchea (reminders_due, calendar_ahead,
                # finance_summary, weather, gmail_recent) el CONTEXTO se
                # REEMPLAZA con la salida formateada de esos tools — fácil
                # 2-4K tokens de listas. qwen2.5:3b en M3 Max prefillea a
                # ~2.5ms/tok → 3K tokens = 7.5s sólo en prefill, vs qwen2.5:7b
                # que prefillea a ~0.5ms/tok (mejor flash-attention
                # throughput en modelos más grandes) → 3K tokens = 1.5s.
                # Ejemplo medido: "qué pendientes tengo" fast_path=1 +
                # tool_rounds=1 → prefill=11595ms, total=16.3s. Post-
                # downgrade (qwen2.5:7b) mismo tool output → prefill estimado
                # ~2s, total ~5s. Gate: si pre-router fired AND fast_path
                # estaba en True, disable fast-path PARA ESTA QUERY — usar
                # _web_model_full + _WEB_CHAT_NUM_CTX para el streaming
                # final. Todo lo calculado arriba (fast_path marker en el
                # log, status SSE) se mantiene consistente: el marker en
                # telemetría es bool(result["fast_path"]) y refleja lo que
                # retrieve() decidió ANTES de ver los tools — así el
                # analytics sigue midiendo la cobertura del adaptive
                # routing. El downgrade es local al código de streaming.
                # Rollback: exportar el env var con cualquier valor
                # truthy para restaurar el comportamiento pre-fix.
                #
                # Mode interaction (2026-04-24): el gate usa
                # `_could_have_fast` — True cuando (a) retrieve() decidió
                # fast_path O (b) el user mandó `mode="fast"`. Loggeamos
                # `fast_path_downgraded=True` aunque el user haya pedido
                # `mode="deep"` siempre que retrieve() hubiera ido por
                # fast: es un evento semántico que dice "había intención
                # de fast-path y el pre-router la bloqueó". El cambio
                # efectivo de modelo sólo aplica si ya estábamos en fast
                # (si `_fast_path` era False por `mode="deep"`, ya
                # estábamos en full, no hay nada que downgradear).
                _could_have_fast = bool(result.get("fast_path", False)) or mode == "fast"
                if _could_have_fast and os.environ.get(
                    "RAG_FAST_PATH_KEEP_WITH_TOOLS", ""
                ).strip().lower() in ("", "0", "false", "no"):
                    _fast_path_downgraded = True
                    if _fast_path:
                        _web_model = _web_model_full
                        _WEB_CHAT_OPTIONS = {
                            **CHAT_OPTIONS,
                            "num_ctx": _WEB_CHAT_NUM_CTX,
                            "num_predict": 256,
                        }
                        print(
                            f"[chat-fast-path-downgrade] pre-router fired "
                            f"({len(_forced_tools)} tools); switching "
                            f"{_LOOKUP_MODEL}→{_web_model} num_ctx="
                            f"{_WEB_CHAT_OPTIONS['num_ctx']}",
                            flush=True,
                        )

            # Gate the LLM tool-deciding loop. Empirically the LLM rarely
            # picks tools beyond what the pre-router already fired (REGLA 1
            # pinning + pre-router-injected data satisfies the query), and
            # the tool-deciding call is a non-streaming cold-prefill
            # (~10-30s on 9k-char ctx) that burns latency for near-zero
            # benefit. Skip it by default — go straight to final streaming.
            # Opt back in with RAG_WEB_TOOL_LLM_DECIDE=1 for queries where
            # chaining multiple tools based on first-round results matters.
            #
            # Exception: propose_reminder / propose_calendar_event can ONLY
            # be invoked via the LLM tool-decide loop (arg extraction is
            # LLM-hard). `_detect_propose_intent` flips the gate on for
            # create-intent queries so the user doesn't have to flip
            # RAG_WEB_TOOL_LLM_DECIDE just to record something.
            _llm_tool_decide = os.environ.get(
                "RAG_WEB_TOOL_LLM_DECIDE", ""
            ).strip() not in ("", "0", "false", "no")
            # `_propose_intent` set earlier (pre-router gate); reuse to avoid
            # re-running the regex.
            #
            # Safety-net fallback (2026-04-24, user report iter 7): si el
            # pre-router regex NO engancho tool alguno Y la retrieve del
            # vault devolvió confianza débil (< CONFIDENCE_DEEP_THRESHOLD
            # = 0.10 — "vault really failed" según la calibración global),
            # prendemos el LLM tool-decide round como safety-net. Cubre
            # el gap de sinónimos/phrasings que el regex no anticipó —
            # p.ej. "qué correspondencia tengo?", "mostrame mi agenda de
            # mayo", "tenés algo en el inbox?". Sin este fallback la
            # query caía al vault retrieve y si el vault no tenía nada
            # semánticamente cercano, el user quedaba sin respuesta útil.
            #
            # Tradeoff: +10-30s en las queries que caen en este branch
            # (cold prefill de qwen2.5:7b sobre ~9k chars). Estimado
            # <10% del tráfico — el 80%+ tiene match de regex o vault
            # strong, el 10%+ restante cae a bypass canned antes de
            # llegar acá. La racional empírica original ("el LLM rara
            # vez agrega valor si el pre-router no matcheó") era válida
            # cuando el regex cubría la mayoría de los casos; cuando el
            # regex falla, el LLM ES el plan B.
            from rag import CONFIDENCE_DEEP_THRESHOLD as _CONF_DEEP_THRESHOLD
            _pre_router_missed = not _forced_tool_pairs
            _vault_retrieve_weak = (
                float(result.get("confidence") or 0.0) < _CONF_DEEP_THRESHOLD
            )
            # Mention-hit guard: cuando el query pega con el resolver de
            # Mentions @/ (p.ej. "qué sabés de Bizarrap" → match contra
            # `99 Mentions @/bizarrap.md`), `build_person_context` ya
            # inyectó un preamble authoritative en el system prompt del
            # turn. No hay gap que tapar con el LLM tool-decide — el LLM
            # tiene la data que necesita. Skippeamos el fallback para no
            # pagar 10-30s extras en queries que ya están respondidas.
            _has_mention_hit = bool(_mention_paths)
            _llm_fallback_needed = (
                _pre_router_missed
                and _vault_retrieve_weak
                and not _has_mention_hit
            )
            _skip_llm_tool_round = (
                not _llm_tool_decide
                and not _propose_intent
                and not _llm_fallback_needed
            )
            # Log explícito cuando el fallback se activa — útil para tunear
            # el threshold si vemos que dispara demasiado seguido (cost)
            # o muy poco (cobertura insuficiente). `[chat-llm-fallback]`
            # tag es greppable desde el web.log.
            if _llm_fallback_needed and not _llm_tool_decide and not _propose_intent:
                print(
                    f"[chat-llm-fallback] pre-router miss + conf="
                    f"{float(result.get('confidence') or 0.0):.3f} "
                    f"< {_CONF_DEEP_THRESHOLD} — running LLM tool-decide "
                    f"as safety net",
                    flush=True,
                )

            for _round_idx in range(_TOOL_ROUND_CAP):
                if _skip_llm_tool_round:
                    # Pre-router matched nothing and opt-in not set → skip
                    # the LLM decide round entirely. `break` avoids the
                    # `else:` clause firing (which would nudge the model
                    # with a "cap reached" system message).
                    break
                # On create-intent turns, narrow the tool surface to JUST
                # the 2 propose_* tools. With all 9 tools visible qwen2.5:7b
                # gets decision paralysis and often returns a clarifying
                # question instead of calling the tool. With only 2 tools
                # present and the override telling it to call one of them,
                # the choice is forced.
                _round_tools = CHAT_TOOLS
                if is_propose_intent:
                    _round_tools = [
                        fn for fn in CHAT_TOOLS if fn.__name__ in PROPOSAL_TOOL_NAMES
                    ]
                # Use the tool-decision client with a wider timeout than
                # the streaming one: non-streaming + `tools=` schema of
                # all 12 chat tools can push qwen2.5:7b > 45s on long
                # inputs. See `_OLLAMA_TOOL_CLIENT` comment above.
                _tr = _OLLAMA_TOOL_CLIENT.chat(
                    model=_web_model,
                    messages=tool_messages,
                    tools=_round_tools,
                    options=CHAT_TOOL_OPTIONS,
                    stream=False,
                    think=False,   # see _ollama_chat_probe for rationale
                    keep_alive=chat_keep_alive(_web_model),
                )
                _tmsg = _tr.message
                _tcalls = list(_tmsg.tool_calls or [])
                if not _tcalls:
                    # Terminal assistant turn — no tools. Don't append it to
                    # tool_messages; if we did, the final streaming call
                    # would see a completed answer and emit zero tokens.
                    # The final call regenerates from the clean prompt.
                    break
                # Persist the assistant turn so the LLM sees its own tool_calls
                # when we come back around (and the final streaming call too).
                tool_messages.append({
                    "role": "assistant",
                    "content": _tmsg.content or "",
                    "tool_calls": [tc.model_dump() for tc in _tcalls],
                })

                # Split into serial (not parallel-safe) and parallel buckets.
                _serial: list[tuple[str, dict]] = []
                _parallel: list[tuple[str, dict]] = []
                for _tc in _tcalls:
                    _name = _tc.function.name
                    _args = _unwrap_args(_tc.function.arguments or {})
                    if _name in PARALLEL_SAFE:
                        _parallel.append((_name, _args))
                    else:
                        _serial.append((_name, _args))

                _round_tool_names: list[str] = []
                _round_parallel_max_ms = 0
                _round_serial_sum_ms = 0

                # Tool output truncation (eval 2026-04-28 fix): mismo motivo
                # que en el pre-router (ver bloque `_datos_block` arriba) —
                # los outputs grandes hacían que la 2da ronda del LLM
                # timeoutiara. Lazy import para deploy parcial.
                try:
                    from rag._tool_output_helpers import truncate_tool_output_for_synthesis as _trunc_tool_out
                except Exception:
                    _trunc_tool_out = lambda name, raw, **kw: raw  # type: ignore

                # Serial bucket first — vault search is the gating signal.
                for _name, _args in _serial:
                    yield _sse("status", {"stage": "tool", "name": _name, "args": _args})
                    _t_tool_start = time.perf_counter()
                    _fn = TOOL_FNS.get(_name)
                    # Eval 2026-04-28: inyectar el query original al post-LLM
                    # para que los validators (_validate_calendar_recurrence,
                    # _validate_reminder_*) puedan recuperar info que el LLM
                    # dropeó.
                    if _name in ("propose_calendar_event", "propose_reminder") and "_original_query" not in _args:
                        _args["_original_query"] = question
                    try:
                        if _fn is None:
                            _out = f"Error: tool '{_name}' no existe"
                        else:
                            _ret = _fn(**_args)
                            _out = _ret if isinstance(_ret, str) else json.dumps(_ret, ensure_ascii=False)
                    except Exception as _exc:
                        _out = f"Error: {_exc}"
                    _elapsed_ms = int((time.perf_counter() - _t_tool_start) * 1000)
                    _round_serial_sum_ms += _elapsed_ms
                    _round_tool_names.append(_name)
                    yield _sse("status", {"stage": "tool_done", "name": _name, "ms": _elapsed_ms})
                    if _prop := _maybe_emit_proposal(_name, _out):
                        yield _prop
                    tool_messages.append({
                        "role": "tool",
                        "name": _name,
                        "content": _trunc_tool_out(_name, _out),
                    })

                # Parallel bucket: emit all `tool` fan-out events first, then
                # submit + wait. Per spec, tool_done events arrive as futures
                # resolve (interleaved).
                if _parallel:
                    for _name, _args in _parallel:
                        yield _sse("status", {"stage": "tool", "name": _name, "args": _args})

                    def _exec_one(name_args):
                        _name, _args = name_args
                        _t0 = time.perf_counter()
                        _fn = TOOL_FNS.get(_name)
                        # Eval 2026-04-28: inyectar el query original (ver
                        # comentario en el bucket serial arriba).
                        if _name in ("propose_calendar_event", "propose_reminder") and "_original_query" not in _args:
                            _args["_original_query"] = question
                        try:
                            if _fn is None:
                                _out = f"Error: tool '{_name}' no existe"
                            else:
                                _ret = _fn(**_args)
                                _out = _ret if isinstance(_ret, str) else json.dumps(_ret, ensure_ascii=False)
                        except Exception as _exc:
                            _out = f"Error: {_exc}"
                        return _name, _out, int((time.perf_counter() - _t0) * 1000)

                    _ex = _ToolExecutor(max_workers=5)
                    try:
                        _futures = [_ex.submit(_exec_one, na) for na in _parallel]
                        for _fut in _as_completed(_futures):
                            _name, _out, _elapsed_ms = _fut.result()
                            if _elapsed_ms > _round_parallel_max_ms:
                                _round_parallel_max_ms = _elapsed_ms
                            _round_tool_names.append(_name)
                            yield _sse("status", {"stage": "tool_done", "name": _name, "ms": _elapsed_ms})
                            if _prop := _maybe_emit_proposal(_name, _out):
                                yield _prop
                            tool_messages.append({
                                "role": "tool",
                                "name": _name,
                                "content": _trunc_tool_out(_name, _out),
                            })
                    finally:
                        _ex.shutdown(wait=False)

                if _round_tool_names:
                    tool_rounds += 1
                    tool_names_called.extend(_round_tool_names)
                    # Critical-path ms this round = serial sum + parallel max.
                    tool_ms_total += _round_serial_sum_ms + _round_parallel_max_ms
                    # 2026-04-28 wave-3: si esta ronda ejecutó tools, salimos
                    # del loop directo — la siguiente ronda llamaría al LLM
                    # con `tools=` schema (~23 tools) + tool results en
                    # contexto. Empíricamente: en 678 queries históricos NUNCA
                    # se necesitó round 2 productivo (todos `tool_rounds=1`),
                    # pero a veces tardaba >90s causando "LLM falló: timed
                    # out". Skipear esa ronda nos lleva directo a la synthesis
                    # call (sin `tools=`), más rápida y sin colgarse. Si en el
                    # futuro queremos chained tool execution real, aflojar
                    # esta guard y agregar telemetry para detectar cuántas
                    # queries lo necesitan.
                    break
            else:
                # Round cap reached without the model emitting an empty
                # tool_calls turn — nudge it to close out with what it has.
                tool_messages.append({
                    "role": "system",
                    "content": "Alcanzado cap de herramientas; respondé con lo que tenés.",
                })
        except Exception as exc:
            # Ver `[chat-stream-error] phase=synthesis` (más abajo) para
            # rationale del logging — mismo bug de observabilidad acá.
            # Este es el except del tool-decision loop (cliente non-stream
            # con tools= schema, budget de _OLLAMA_TOOL_TIMEOUT=90s).
            print(
                f"[chat-stream-error] phase=tool_decision "
                f"exc_type={type(exc).__name__} exc={exc} "
                f"tool_rounds={tool_rounds} "
                f"tools_so_far={','.join(tool_names_called) or 'none'} "
                f"messages_so_far={len(tool_messages)}",
                flush=True,
            )
            yield _sse("error", {"message": _sanitize_error_for_user(exc, phase="tool_decision")})
            # BUG #31 wave-2: top_score=0.0 dispara fallback cluster.
            yield _sse("done", {"error": True, "top_score": 0.0})  # BUG #31
            return

        # Final streaming answer call: strip _WEB_TOOL_ADDENDUM and `tools=`
        # so the prefix cache for plain-text generation matches the pre
        # tool-calling era byte-for-byte. We keep the full tool loop
        # (assistant tool_calls + tool messages) so the LLM sees what it
        # learned; without `tools=` ollama won't re-invoke anything.
        final_messages = [m for m in tool_messages if not (
            m.get("role") == "system" and m.get("content") == _WEB_TOOL_ADDENDUM
        )]

        # num_ctx PINNED to `_WEB_CHAT_NUM_CTX` (= _WEB_CHAT_OPTIONS["num_ctx"])
        # — sin variación entre turnos. Ver más abajo el rationale extendido.
        #
        # Historia:
        #
        # 2026-04-25 (developer-1) introdujo aquí un `_adaptive_num_ctx =
        # min(4096, max(1024, (_final_ctx_chars // 4) + 512))` con la idea
        # de ahorrar prefill en queries cortas (saludos, propose intent
        # simple). Telemetría esperada: -1s prefill en queries de <1.5K
        # tokens.
        #
        # 2026-04-28 repro Playwright detectó el costo oculto: cada call
        # con `num_ctx ≠ <currently-loaded num_ctx>` fuerza a ollama a
        # reinicializar la KV cache del modelo. En MPS bajo memory
        # pressure ese reinit puede tomar 60-120s. Concretamente vimos
        #
        #   [chat-stream-error] phase=synthesis exc=timed out ttft_ms=90004
        #     ctx_chars=8277 num_ctx=2581 tools=whatsapp_pending,calendar_
        #     ahead,reminders_due got_first_token=False
        #
        # con el modelo loaded a `CONTEXT 4096` (visible en `ollama ps`).
        # El "ahorro" de 1s en queries cortas no compensa los 90s de
        # timeout en queries que mezclan adaptive=2581 con loaded=4096.
        #
        # El propio comentario de `_WEB_CHAT_NUM_CTX` (línea ~2700) dice:
        #
        #   "Shared num_ctx for every call to the chat model from this
        #    server — the real /api/chat, the preflight probe, and the
        #    background prewarmer must all pass the same value. ollama
        #    re-initialises the KV cache when a model call arrives with a
        #    different num_ctx than the currently-loaded one [...]"
        #
        # El adaptive contradecía ese invariante. Lo borramos; si en el
        # futuro queremos volver a optimizar prefill para queries cortas,
        # la ruta correcta es:
        #   (a) usar el fast_path (qwen2.5:3b con num_ctx=4096) que ya
        #       acelera por modelo más chico sin tocar KV cache, o
        #   (b) cambiar el num_ctx loaded del modelo (vía prewarm) para
        #       que coincida con el adaptive — pero entonces hay que
        #       garantizar que TODOS los callers (prewarm, probe, /api/
        #       tasks) usen el mismo valor.
        #
        # Por ahora, num_ctx=4096 fijo en todos lados. Mantenemos el
        # _final_ctx_chars solo para el log diagnóstico de [chat-stream-
        # error] y por si en el futuro queremos truncar tool output
        # cuando excede el budget (línea ~2685: "mejor truncar el tool
        # output ANTES de mandarlo al LLM").
        _final_ctx_chars = sum(len(m.get("content", "")) for m in final_messages)
        _adaptive_num_ctx = _WEB_CHAT_OPTIONS["num_ctx"]
        _stream_options = dict(_WEB_CHAT_OPTIONS)
        # No override de num_ctx — pin al valor pre-loaded.

        parts: list[str] = []
        # Pipeline orden: raw_tool_call → citation → iberian.
        # raw_tool_call PRIMERO porque puede swallow los primeros 120 chars
        # antes de que pasen al resto del pipeline (early-exit si el LLM
        # leaked tool-call syntax).
        raw_tool_filter = _RawToolCallStripper()
        stripper = _InlineCitationStripper()
        # 2026-04-23: chain de filtro de leaks portugueses/gallegos
        # (qwen2.5:7b se contagia del CONTEXTO de WhatsApp con
        # contactos brasileros). Idempotente sobre texto en español —
        # el costo es un regex scan por emit. Buffer adds ~1 palabra
        # de lag al streaming, irrelevante para UX.
        iberian = _IberianLeakFilter()
        # 2026-04-28 wave-7: PII redaction filter al final del pipeline.
        # Cubre el caso de chunks split (label + valor llegan en chunks
        # distintos) que el iberian filter ya redacta inline pero solo
        # mismo-chunk. Greppable: incrementa pii.
        pii_filter = _PiiRedactFilter()

        # 2026-04-28 wave-8 fix: el filter `_strip_foreign_scripts` estaba
        # definido (líneas ~1504-1531) pero NUNCA se llamaba, así que los
        # leaks de CJK que sí ocurren bajo presión de contexto (qwen2.5:7b
        # mete 汉字 en respuestas de weather, observado eval Conv B T1
        # "Según el clima预报，今天在马德里...") salían tal cual al user.
        # Greppable: incrementa cjk_strip_count.
        _foreign_strip_count = 0

        def _emit(token_text: str) -> str:
            """Helper: pasa token_text por strip_foreign_scripts → pii_filter
            y devuelve lo que se debe yieldear (puede ser '' si está holdeado).
            """
            nonlocal _foreign_strip_count
            cleaned = _strip_foreign_scripts(token_text)
            if cleaned != token_text:
                _foreign_strip_count += len(token_text) - len(cleaned)
            return pii_filter.feed(cleaned)

        _t_llm_start = time.perf_counter()
        _first_token_logged = False
        try:
            for chunk in _OLLAMA_STREAM_CLIENT.chat(
                model=_web_model,
                messages=final_messages,
                options=_stream_options,
                stream=True,
                think=False,   # the user-facing stream never includes a
                               # <think> preamble. Thinking-capable models
                               # (qwen3+, deepseek-r1, qwq) otherwise emit
                               # tokens with empty content.delta and the
                               # UI sees 0-token responses (measured on
                               # qwen3.6 2026-04-20).
                keep_alive=chat_keep_alive(_web_model),
            ):
                delta = chunk.message.content or ""
                if not delta:
                    continue
                if not _first_token_logged:
                    _t_first_token = time.perf_counter()
                    _first_token_logged = True
                pre = raw_tool_filter.feed(delta)
                if not pre:
                    continue
                filtered = stripper.feed(pre)
                if filtered:
                    cleaned = iberian.feed(filtered)
                    if cleaned:
                        emit_text = _emit(cleaned)
                        if emit_text:
                            parts.append(emit_text)
                            yield _sse("token", {"delta": emit_text})
            # Flush raw_tool_filter primero (puede emitir clarificación).
            raw_tail = raw_tool_filter.flush()
            if raw_tail:
                filtered_raw_tail = stripper.feed(raw_tail)
                if filtered_raw_tail:
                    cleaned_raw_tail = iberian.feed(filtered_raw_tail)
                    if cleaned_raw_tail:
                        emit_text = _emit(cleaned_raw_tail)
                        if emit_text:
                            parts.append(emit_text)
                            yield _sse("token", {"delta": emit_text})
            # Flush any tail that was held back waiting for a close-paren.
            tail = stripper.flush()
            if tail:
                cleaned_tail = iberian.feed(tail)
                if cleaned_tail:
                    emit_text = _emit(cleaned_tail)
                    if emit_text:
                        parts.append(emit_text)
                        yield _sse("token", {"delta": emit_text})
            # Final drain del buffer iberian (residual que no vio boundary).
            final_tail = iberian.flush()
            if final_tail:
                emit_text = _emit(final_tail)
                if emit_text:
                    parts.append(emit_text)
                    yield _sse("token", {"delta": emit_text})
            # Final drain del PII filter (cualquier label sin valor que
            # quedó en buffer al cierre del stream).
            pii_tail = pii_filter.flush()
            if pii_tail:
                parts.append(pii_tail)
                yield _sse("token", {"delta": pii_tail})
            # 2026-04-28 wave-5: log si emitimos clarificación de raw tool call.
            if raw_tool_filter._was_raw_tool_call:
                print(
                    f"[chat-raw-tool-call] LLM leaked tool-call syntax — "
                    f"replaced with clarification. tools_fired={tool_names_called or 'none'} "
                    f"is_propose={is_propose_intent}",
                    flush=True,
                )
            # 2026-04-28 wave-7: log si redactamos PII.
            if pii_filter._redact_count > 0:
                print(
                    f"[chat-pii-redact] redacted {pii_filter._redact_count} "
                    f"PII pattern(s) from response. tools_fired="
                    f"{tool_names_called or 'none'}",
                    flush=True,
                )
            # 2026-04-28 wave-8: log si stripped CJK/foreign tokens.
            if _foreign_strip_count > 0:
                print(
                    f"[chat-cjk-strip] stripped {_foreign_strip_count} "
                    f"foreign char(s) from streaming output. tools_fired="
                    f"{tool_names_called or 'none'}",
                    flush=True,
                )
        except Exception as exc:
            # Diagnóstico: sin este print el server NO loggea NADA cuando el
            # frontend muestra "LLM falló: timed out". Debugear regresiones
            # post-eval era imposible (los chat-timing de éxito sí están en
            # web.log, los failures eran black-box). Capturamos el tipo de
            # exception, ttft hasta donde llegamos, ctx_chars, num_ctx
            # efectivo y qué tools corrieron en este turno — todo lo
            # necesario para distinguir "modelo cold-load" vs "context
            # explotó" vs "ollama daemon wedged" sin re-instrumentar.
            _err_ttft_ms = (
                int((time.perf_counter() - _t_llm_start) * 1000)
                if not _first_token_logged
                else int((_t_first_token - _t_llm_start) * 1000)
            )
            print(
                f"[chat-stream-error] phase=synthesis exc_type={type(exc).__name__} "
                f"exc={exc} ttft_ms={_err_ttft_ms} ctx_chars={_final_ctx_chars} "
                f"num_ctx={_adaptive_num_ctx} tools={','.join(tool_names_called) or 'none'} "
                f"got_first_token={_first_token_logged} parts_so_far={len(parts)}",
                flush=True,
            )
            # 2026-04-28 wave-5: auto-recovery on synthesis ReadTimeout.
            # Si llegamos acá con ttft=90s+ y got_first_token=False, el daemon
            # está wedged — el preflight `_ollama_chat_probe` pasó pero el
            # /api/chat real cuelga. Spawn restart en background thread (no
            # bloqueamos el response al user, que ya se va a ir con error)
            # para que la PRÓXIMA request no caiga en el mismo wedge.
            # Threshold 75s+ porque eso es más que cualquier cold-load
            # legítimo (15s peor caso documentado en chat-timing logs).
            _exc_type_str = type(exc).__name__
            if (
                _exc_type_str in ("ReadTimeout", "RemoteProtocolError")
                and _err_ttft_ms >= 75000
                and not _first_token_logged
            ):
                def _bg_restart():
                    try:
                        print(f"[ollama-auto-recovery] triggering after {_exc_type_str} ttft={_err_ttft_ms}ms", flush=True)
                        ok = _ollama_restart_if_stuck()
                        print(f"[ollama-auto-recovery] result ok={ok}", flush=True)
                    except Exception as _bg_exc:
                        print(f"[ollama-auto-recovery] error: {_bg_exc!r}", flush=True)
                threading.Thread(target=_bg_restart, daemon=True, name="ollama-auto-recovery").start()
            yield _sse("error", {"message": _sanitize_error_for_user(exc, phase="synthesis")})
            # BUG #31 wave-2: top_score=0.0 dispara fallback cluster.
            yield _sse("done", {"error": True, "top_score": 0.0})  # BUG #31
            return

        full = "".join(parts)
        _t_done = time.perf_counter()

        # Timing breakdown for diagnostics
        _t_reform_ms = int((_t_reform_end - _t_reform_start) * 1000)
        _t_retrieve_ms = int((_t_retrieve_end - _t_retrieve_start) * 1000)
        _t_wa_ms = _t_wa_wait_ms
        _t_llm_prefill_ms = int((_t_first_token - _t_llm_start) * 1000) if _first_token_logged else -1
        _t_llm_decode_ms = int((_t_done - (_t_first_token if _first_token_logged else _t_llm_start)) * 1000)
        _t_total_ms = int((_t_done - _t0) * 1000)
        _t_ttft_ms = int(((_t_first_token if _first_token_logged else _t_done) - _t0) * 1000)
        _tok_count = len(full.split())
        _q_words = len(question.split())
        _reform_outcome = (
            "concat" if _reform_used_concat
            else "rewritten" if _reform_fired
            else "skipped"
        )
        _tool_names_str = ",".join(tool_names_called)
        print(
            f"[chat-timing] model={_resolve_web_chat_model()} retrieve={_t_retrieve_ms}ms "
            f"reform={_t_reform_ms}ms reform_outcome={_reform_outcome} q_words={_q_words} "
            f"wa={_t_wa_ms}ms llm_prefill={_t_llm_prefill_ms}ms "
            f"llm_decode={_t_llm_decode_ms}ms ttft={_t_ttft_ms}ms "
            f"total={_t_total_ms}ms ctx_chars={_ctx_chars} tokens≈{_tok_count} "
            # `retrieve()` devuelve float('-inf') con corpus vacío;
            # imprimirlo crudo deja `confidence=-inf` en logs que grep
            # + dashboards interpretan como "error sin sanitizar". Sanear
            # al print para que el log sea shape-stable (float finito).
            f"confidence={_sanitize_confidence(result['confidence']):.3f} "
            f"variants={len(result.get('query_variants',[]))} "
            f"tool_rounds={tool_rounds} tool_ms={tool_ms_total} tool_names={_tool_names_str} "
            f"topic_shift={_topic_shift_reason}{'!' if _topic_shifted else ''}",
            flush=True,
        )

        turn_id = new_turn_id()
        append_turn(sess, {
            "q": question,
            "a": full,
            "paths": [m.get("file", "") for m in result["metas"]],
            "top_score": round(_sanitize_confidence(result["confidence"]), 3),
            "turn_id": turn_id,
            # 2026-04-28 wave-8: track tools fired this turn para que el
            # próximo turn pueda detectar follow-ups anafóricos como
            # "y mañana?" tras "qué hago hoy?" y re-fire los mismos read-
            # intent tools con date shift, en lugar de pasar a propose_*.
            "tools_fired": list(tool_names_called),
            # Track la última weather location explícita para carry-over
            # en follow-ups tipo "y mañana?" o "cuánta lluvia?".
            "weather_location": _last_weather_location,
        })
        save_session(sess)
        log_query_event({
            "cmd": "web",
            "turn_id": turn_id,
            "session": sess["id"],
            "q": question,
            "paths": [m.get("file", "") for m in result["metas"]],
            "scores": [round(float(s), 2) for s in result["scores"]],
            "top_score": round(_sanitize_confidence(result["confidence"]), 2),
            "t_retrieve": round(_t_retrieve_ms / 1000.0, 3),
            "t_gen": round(max(0, _t_total_ms - _t_retrieve_ms) / 1000.0, 3),
            "topic_shifted": _topic_shifted,
            "topic_shift_reason": _topic_shift_reason,
            # Intent classified in the pipeline, echoed back via
            # `result["intent"]`. Closes the 42% gap in intent telemetry
            # measured 2026-04-22.
            "intent": result.get("intent"),
            # Device classificado del User-Agent — iphone/ipad/mac/linux/
            # windows/android/other. Habilita analytics por device y es
            # prerequisito para layout adaptativo (mac-term mobile).
            "device": _client_device,
            # Redo metadata — non-null only when the client called /redo
            # via redo_turn_id. Let analytics count how often users
            # regenerate + whether hint-assisted redos land better
            # answers than bare redos (top_score delta).
            "redo_of_turn_id": _redo_of,
            "redo_hint": _redo_hint,
            # Stage-level timing (ms, integer) — 2026-04-22 gap fix.
            # Pre-fix these values were emitted only to the SSE `done`
            # event for the browser, never persisted. That meant the
            # dashboard couldn't distinguish between "retrieve is slow"
            # vs "LLM is slow" vs "TTFT is slow" — only the coarse
            # t_retrieve + t_gen columns. Persisting here unlocks:
            #   - ttft_ms: percepción del user. Sub-segundo = sweet
            #     spot. El delta vs total_ms mide cuánto "espera"
            #     explícitamente el user.
            #   - llm_prefill_ms: cost del context assembly (antes de
            #     emitir el primer token). Útil para medir impacto de
            #     prefix-cache hits + context size changes.
            #   - llm_decode_ms: tokens/s puro. Foco para speculative
            #     decoding (GC#2.D).
            #   - total_ms: crosscheck t_retrieve + t_gen (column
            #     values) vs lo que realmente midió el endpoint.
            "ttft_ms": int(_t_ttft_ms),
            "llm_prefill_ms": int(_t_llm_prefill_ms),
            "llm_decode_ms": int(_t_llm_decode_ms),
            "total_ms": int(_t_total_ms),
            # Adaptive num_ctx (2026-04-25, developer-1) — el ceiling KV
            # efectivo que pasamos al streaming call. Pre-fix era fijo en
            # 4096; ahora es min(4096, max(1024, char_count//4 + 512)),
            # con cap igual al pre-fix value para que no rompa prefix
            # caching. Cruzar contra `llm_prefill_ms` para medir si bajar
            # el ctx para queries chicas reduce prefill (esperado: queries
            # de saludo / propose intent caen ~1.0s, vault queries
            # largas no cambian). El campo `final_ctx_chars` es el input
            # crudo a la heurística (todos los message.content del
            # streaming call concatenados), útil para reverse-engineering
            # del cálculo si la heurística necesita re-tuning.
            "adaptive_num_ctx": int(_adaptive_num_ctx),
            "final_ctx_chars": int(_final_ctx_chars),
            # fast_path marker (Improvement #3): True when adaptive routing
            # judged the query eligible for the small-model fast path
            # (semantic intent + top-1 rerank > 0.6, OR WA majority per
            # `RAG_WA_FAST_PATH`). Desde 2026-04-22 el endpoint SÍ honra
            # el flag: `_web_model` switcha a `_LOOKUP_MODEL` (qwen2.5:3b)
            # + `num_ctx=_LOOKUP_NUM_CTX` (4096). Pre-fix loggeaba el
            # marker pero seguía generando con qwen2.5:7b → pérdida de
            # 2.75× speedup medido en prod (CLI 7d). Esta columna ahora
            # surfacea la ganancia real cuando `fast_path=True`.
            #
            # NOTE (2026-04-24): loggeamos el `_fast_path` EFECTIVO — el
            # valor realmente usado en la generación tras aplicar el
            # mode gate. Si el user mandó `mode="fast"` o `mode="deep"`,
            # puede diferir de `result["fast_path"]` (la decisión del
            # adaptive routing). Para distinguir override del user vs
            # decisión automática, cruzar contra `mode` + `mode_origin`
            # más abajo.
            "fast_path": bool(_fast_path),
            # Downgrade marker — True cuando el pre-router fired tools y
            # el endpoint cambió runtime de _LOOKUP_MODEL (qwen2.5:3b) a
            # qwen2.5:7b porque el contexto inflado por el tool output
            # era letal para prefill en el modelo chico (ver bloque de
            # downgrade en pre-router). Desacoplado del marker `fast_path`
            # — `fast_path=True + fast_path_downgraded=True` significa
            # "retrieve() pensó fast-path, pero la realidad del pre-router
            # mandó downgrade". Util para medir cobertura real del
            # adaptive routing vs lo que realmente corrió. También fires
            # cuando el user pidió `mode="fast"` y el pre-router matched
            # tools — el downgrade es safety rail de prefill, no routing.
            "fast_path_downgraded": bool(_fast_path_downgraded),
            # Chat mode override (2026-04-24, feature "modo rápido/profundo"):
            # captura el knob user-facing que el web UI manda. `chat_mode`
            # es uno de {"auto","fast","deep"} post-validación (invalid →
            # silent "auto"). `chat_mode_origin` es "request" cuando el
            # cliente incluyó el campo (valor válido O inválido) y
            # "default" cuando el campo estaba ausente/null. Juntos
            # contestan: (a) cuán seguido los users overridean el
            # adaptive routing? (b) en qué dirección? (c) hay clientes
            # rotos mandando garbage? Esencial para decidir si el
            # adaptive routing necesita retuning.
            #
            # **Nota sobre el nombre**: loggeamos `chat_mode`, no `mode`,
            # porque la tabla `rag_queries` YA tiene una columna `mode`
            # (string libre para command type: "query"/"chat"/"do"/etc).
            # Si usáramos `"mode"` como key en el extra_json dict, el
            # merge de log_query_event la mapearía a la columna
            # preexistente y perderíamos el valor del UI. El prefix
            # `chat_` separa el namespace y mantiene la analítica limpia.
            "chat_mode": mode,
            "chat_mode_origin": mode_origin,
            # Retrieval stage timing — desde 2026-04-23, también para
            # el web endpoint. Incluye embed_ms / sem_ms / bm25_ms /
            # rrf_ms / reranker_ms / total_ms. Pre-fix solo el CLI
            # `query` + `chat` persistían este dict; el web emitía el
            # t_retrieve agregado pero no el breakdown interno. Con
            # `timing` en extra_json, el SQL puede distinguir "retrieve
            # lento por embed (warmup race)" vs "retrieve lento por
            # reranker (cold MPS)" sin tener que reproducir en CLI.
            # Requisito para diagnosticar los outliers de web que
            # mostraban `retrieve_ms` 4-6s warm sin desglose accesible.
            "timing": _round_timing_ms(result.get("timing")),
            # Audit 2026-04-24: pre-fix `cache_probe` se loggeaba SOLO en
            # el HIT path (line 5717+, cuando el lookup matcheaba). Para
            # los 998 web queries del último período (mayoría misses) este
            # campo nunca se persistía → `rag cache stats --days 7`
            # mostraba 5 eligible (sólo de CLI) cuando en realidad había
            # ~600 web queries pasando por el lookup. El bug ocultaba el
            # telemetry crítico para diagnosticar por qué el cache tiene
            # 0 hits efectivos. Post-fix loggeamos el probe SIEMPRE — el
            # caller (line 5520) inicializa la var a None y la setea
            # cuando el lookup corre, así que es safe en todos los paths.
            "cache_probe": _semantic_cache_probe,
            "cache_hit": False,
            "cache_layer": None,
            # Audit 2026-04-25: trackeamos cuántas iteraciones del
            # sufficiency loop corrieron y por qué exit-eamos. `None`
            # cuando `deep_retrieve()` no se invocó (top_score >=
            # CONFIDENCE_DEEP_THRESHOLD ya en la primera pasada, o
            # el caller pidió `no_deep`). `1 + "high_confidence_bypass"`
            # cuando el early-exit por alta confianza activó (audit
            # 2026-04-25 — commit e81251f). Permite SQL tipo
            # "qué fracción de queries usa más de 1 iter?" + medir
            # hit rate del bypass sin instrumentación adicional.
            # Ver `rag/__init__.py:19515-19521` (donde el dict se
            # populates) y los campos `_DEEP_HIGH_CONF_BYPASS` /
            # `_DEEP_LOW_CONF_BYPASS` para los thresholds.
            "deep_retrieve_iterations": result.get("deep_retrieve_iterations"),
            "deep_retrieve_exit_reason": result.get("deep_retrieve_exit_reason"),
        })

        yield _sse("done", {
            "turn_id": turn_id,
            "top_score": round(_sanitize_confidence(result["confidence"]), 3),
            "total_ms": _t_total_ms,
            "retrieve_ms": _t_retrieve_ms,
            "ttft_ms": _t_ttft_ms,
            "llm_ms": _t_llm_decode_ms,
            # Source-specific flag (2026-04-24, Fer F. user report iter 4):
            # true cuando el pre-router disparó un tool mapeado a una
            # fuente concreta del user (gmail_recent / calendar_ahead /
            # reminders_due — las que viven en `_SOURCE_INTENT_META`).
            # El frontend lo usa para apagar CTAs de "buscar en internet"
            # y el cluster YouTube, que no tienen sentido cuando la
            # pregunta es inherentemente sobre data local — Google no
            # sabe qué mails pendientes tengo yo, ni qué recordatorios
            # cargué en Apple Reminders. Tools como `weather` /
            # `finance_summary` NO flaguean (no son "fuentes" buscables,
            # son passthrough / resúmenes).
            "source_specific": any(
                name in _SOURCE_INTENT_META
                for name, _args in _forced_tool_pairs
            ),
        })

        _spawn_conversation_writer(
            target_args=(
                VAULT_PATH,
                sess["id"],
                req.question,
                full,
                result["metas"],
                result["scores"],
                result["confidence"],
            ),
            name=f"conv-writer-{turn_id[:8]}",
        )

        _enrich_evt = _emit_enrich(turn_id, question, full, float(result["confidence"]))
        if _enrich_evt:
            yield _enrich_evt

        _grounding_evt = _emit_grounding(turn_id, full, result["docs"], result["metas"], question)
        if _grounding_evt:
            yield _grounding_evt

        # Cache store — sólo queries sin history (no follow-ups) y con
        # respuesta no-vacía. Keya con el vault_chunks count computado en el
        # cache-get path arriba (mismo scope/model/vault state → misma key).
        # `_cache_key is not None` cubre el caso donde el turno arrancó con
        # history (no key computado) y el topic-shift gate reasignó
        # `history = []` (línea ~3914), dejando este branch alcanzable sin
        # key válida. Sin el guard, se tiraba `UnboundLocalError` en cada
        # topic-shift turn.
        if not history and full.strip() and _cache_key is not None:
            try:
                _sources_items = [
                    {**_source_payload(m, s), "bar": _score_bar(float(s))}
                    for m, s in zip(result["metas"], result["scores"])
                ]
                _chat_cache_put(_cache_key, {
                    "text": full,
                    "sources_items": _sources_items,
                    "top_score": round(_sanitize_confidence(result["confidence"]), 3),
                })
                print(f"[chat-cache] PUT {_cache_key} len={len(full)}", flush=True)
            except Exception as _cache_err:
                print(f"[chat-cache] put failed: {_cache_err}", flush=True)
            # Semantic cache store (SQL, persistent). Piggyback on the
            # embed + corpus_hash already computed in the lookup path
            # above (~line 4814). background=True is fine here — web
            # server is a long-running daemon so the atexit drop that
            # bit `rag query` CLI doesn't apply. Same store gates as
            # any other caller (refusal detector, low top_score,
            # empty response) fire inside `semantic_cache_store`.
            if _semantic_cache_emb is not None and _semantic_cache_hash:
                try:
                    from rag import semantic_cache_store as _rag_sem_store
                    _sem_cache_paths = list(dict.fromkeys(
                        m.get("file", "")
                        for m in (result.get("metas") or [])
                        if m.get("file")
                    ))
                    _rag_sem_store(
                        _semantic_cache_emb,
                        question=question,
                        response=full,
                        paths=_sem_cache_paths,
                        scores=[
                            round(float(s), 2)
                            for s in (result.get("scores") or [])[:len(_sem_cache_paths)]
                        ],
                        top_score=round(
                            _sanitize_confidence(result["confidence"]), 2,
                        ),
                        intent=None,
                        corpus_hash=_semantic_cache_hash,
                        background=True,
                    )
                    print(
                        f"[chat-cache] SEMANTIC PUT hash={_semantic_cache_hash} "
                        f"len={len(full)}",
                        flush=True,
                    )
                except Exception as _sem_put_err:
                    print(
                        f"[chat-cache] semantic put failed: "
                        f"{type(_sem_put_err).__name__}: {_sem_put_err}",
                        flush=True,
                    )

    def guarded():
        """Increment/decrement the global chat-in-flight counter around
        the real generator. The home-prewarmer loop checks this counter
        and skips cycles while > 0 so chat gets exclusive ollama time.
        Decrement runs in `finally`, which executes on normal completion
        AND when the client disconnects (generator .close()).
        """
        global _CHAT_INFLIGHT
        with _CHAT_INFLIGHT_LOCK:
            _CHAT_INFLIGHT += 1
        try:
            yield from gen()
        finally:
            with _CHAT_INFLIGHT_LOCK:
                _CHAT_INFLIGHT = max(0, _CHAT_INFLIGHT - 1)

    # Bug fix 2026-04-27: add anti-buffering headers so reverse proxies
    # (Caddy, nginx) don't buffer the SSE stream, causing the user to
    # see nothing until the response completes. Same set as home_stream /
    # dashboard_stream.
    return StreamingResponse(
        guarded(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache, no-transform",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


@app.get("/dashboard")
def dashboard_page() -> FileResponse:
    return FileResponse(STATIC_DIR / "dashboard.html")


# ── /transcripts — dashboard de telemetría del whisper learning loop ─────────
# Phase 2 Step 3.b. Página HTML server-rendered (no SPA) con stats agregadas
# de las transcripciones de audio, correcciones acumuladas, y top vocab terms
# que el sistema está aprendiendo. Sirve como observability + para calibrar
# el threshold del LLM auto-correct (si muchos audios caen en el bucket
# `avg_logprob < -0.8`, conviene bajar el threshold; si pocos, está bien
# como está).
#
# Doc del plan en el vault:
# `04-Archive/99-obsidian-system/99-AI/system/whatsapp-whisper-learning/plan.md`

def _esc(s: str) -> str:
    """HTML escape para evitar XSS desde texto del usuario (chats, transcripts)."""
    return (str(s) if s is not None else "")\
        .replace("&", "&amp;")\
        .replace("<", "&lt;")\
        .replace(">", "&gt;")\
        .replace('"', "&quot;")


@app.get("/transcripts", response_class=HTMLResponse)
def transcripts_dashboard(nofresh: int = 0) -> HTMLResponse:
    """Dashboard de telemetría del whisper learning loop.

    Args:
        nofresh: si !=0, suprime el meta-refresh auto cada 60s. Útil cuando
            estás leyendo la página y no querés que se recargue scrollando.
            Ejemplo: `/transcripts?nofresh=1`.
    """
    try:
        with _ragvec_state_conn() as conn:
            # Stats globales (últimos 30 días)
            cutoff_30d = time.time() - 30 * 86400
            row_global = conn.execute(
                "SELECT COUNT(*), AVG(avg_logprob), AVG(duration_s) "
                "FROM rag_audio_transcripts "
                "WHERE transcribed_at > ?",
                (cutoff_30d,),
            ).fetchone()
            n_30d = row_global[0] or 0
            avg_lp_30d = row_global[1]
            avg_dur_30d = row_global[2]
            n_total = conn.execute(
                "SELECT COUNT(*) FROM rag_audio_transcripts"
            ).fetchone()[0]
            # Correcciones — totales y por source
            n_corrections = conn.execute(
                "SELECT COUNT(*) FROM rag_audio_corrections"
            ).fetchone()[0]
            corr_by_source = dict(conn.execute(
                "SELECT source, COUNT(*) FROM rag_audio_corrections GROUP BY source"
            ).fetchall())
            # Logprob histogram — buckets de 0.2
            buckets = [
                ("> -0.2 (excelente)", "avg_logprob > -0.2"),
                ("-0.4 a -0.2 (alta)", "avg_logprob > -0.4 AND avg_logprob <= -0.2"),
                ("-0.6 a -0.4 (media)", "avg_logprob > -0.6 AND avg_logprob <= -0.4"),
                ("-0.8 a -0.6 (baja)", "avg_logprob > -0.8 AND avg_logprob <= -0.6"),
                ("< -0.8 (LLM correct)", "avg_logprob <= -0.8"),
            ]
            hist = []
            for label, where in buckets:
                n = conn.execute(
                    f"SELECT COUNT(*) FROM rag_audio_transcripts WHERE {where}"
                ).fetchone()[0]
                hist.append((label, n))
            # Vocab — totales y por source
            n_vocab = conn.execute(
                "SELECT COUNT(*) FROM rag_whisper_vocab"
            ).fetchone()[0]
            vocab_by_source = dict(conn.execute(
                "SELECT source, COUNT(*) FROM rag_whisper_vocab GROUP BY source"
            ).fetchall())
            last_vocab_refresh = conn.execute(
                "SELECT MAX(refreshed_at) FROM rag_whisper_vocab"
            ).fetchone()[0]
            # Top vocab por weight (50)
            top_vocab = conn.execute(
                "SELECT term, weight, source FROM rag_whisper_vocab "
                "ORDER BY weight DESC LIMIT 50"
            ).fetchall()
            # Últimas 30 transcripciones — incluye `duration_s` para correlacionar
            # con logprob (audios cortos típicamente tienen logprob más bajo
            # por menos contexto; audios largos pueden tener errores en
            # fragmentos específicos).
            recent_transcripts = conn.execute(
                "SELECT audio_path, transcribed_at, avg_logprob, model, "
                "       text, corrected_text, correction_source, chat_id, "
                "       duration_s "
                "FROM rag_audio_transcripts "
                "ORDER BY transcribed_at DESC LIMIT 30"
            ).fetchall()
            # Top 20 correcciones recientes
            recent_corrections = conn.execute(
                "SELECT ts, source, original, corrected "
                "FROM rag_audio_corrections "
                "ORDER BY ts DESC LIMIT 20"
            ).fetchall()
            # Patrones repetidos en correcciones — mismo algoritmo que
            # `rag whisper patterns` CLI. Imported lazy para no creates
            # circular import si rag_whisper_learning todavía no está.
            try:
                from rag_whisper_learning.patterns import find_correction_patterns
                patterns_data = find_correction_patterns(min_count=2)
            except Exception:
                patterns_data = []
            # Heatmap por hora del día (últimos 30 días, hora local).
            # Útil para visualizar cuándo el user manda audios típicamente.
            hour_counts: dict[int, int] = {h: 0 for h in range(24)}
            try:
                for row in conn.execute(
                    "SELECT CAST(strftime('%H', transcribed_at, 'unixepoch', 'localtime') AS INTEGER) AS hour, "
                    "       COUNT(*) "
                    "FROM rag_audio_transcripts "
                    "WHERE transcribed_at > ? "
                    "GROUP BY hour",
                    (cutoff_30d,),
                ):
                    h = int(row[0]) if row[0] is not None else 0
                    if 0 <= h < 24:
                        hour_counts[h] = int(row[1] or 0)
            except Exception:
                pass
            # Heatmap semanal día×hora (últimos 60 días para tener suficiente
            # signal por celda). Sqlite `strftime('%w', ...)` devuelve 0-6
            # (0=domingo, 1=lunes, ..., 6=sábado). Reordenamos visualmente
            # a lunes-domingo para coincidir con el calendario europeo /
            # locale español.
            cutoff_60d = time.time() - 60 * 86400
            week_counts: dict[tuple[int, int], int] = {
                (d, h): 0 for d in range(7) for h in range(24)
            }
            try:
                for row in conn.execute(
                    "SELECT CAST(strftime('%w', transcribed_at, 'unixepoch', 'localtime') AS INTEGER) AS dow, "
                    "       CAST(strftime('%H', transcribed_at, 'unixepoch', 'localtime') AS INTEGER) AS hour, "
                    "       COUNT(*) "
                    "FROM rag_audio_transcripts "
                    "WHERE transcribed_at > ? "
                    "GROUP BY dow, hour",
                    (cutoff_60d,),
                ):
                    d = int(row[0]) if row[0] is not None else 0
                    h = int(row[1]) if row[1] is not None else 0
                    if 0 <= d < 7 and 0 <= h < 24:
                        week_counts[(d, h)] = int(row[2] or 0)
            except Exception:
                pass
    except Exception as exc:
        return HTMLResponse(
            f"<!doctype html><html><body>"
            f"<h1>Whisper transcripts</h1>"
            f"<p>Error reading state: {_esc(str(exc))}</p>"
            f"<p>¿`telemetry.db` accesible? ¿`RAG_STATE_SQL=1` en el listener?</p>"
            f"</body></html>",
            status_code=500,
        )

    # Renderizar HTML
    last_refresh_str = "—"
    if last_vocab_refresh:
        ago = time.time() - last_vocab_refresh
        ago_h = int(ago / 3600)
        ago_m = int((ago % 3600) / 60)
        last_refresh_str = (
            f"{datetime.fromtimestamp(last_vocab_refresh).strftime('%Y-%m-%d %H:%M')} "
            f"({ago_h}h {ago_m}m ago)"
        )
    avg_lp_str = f"{avg_lp_30d:.3f}" if avg_lp_30d is not None else "—"
    avg_dur_str = f"{avg_dur_30d:.1f}s" if avg_dur_30d is not None else "—"

    # Tabla histogram — barras usan width inline (data-driven) + class .bar (color del theme).
    max_hist = max((n for _, n in hist), default=1)
    hist_rows = "\n".join(
        f'<tr><td>{_esc(label)}</td>'
        f'<td style="text-align:right">{n}</td>'
        f'<td><div class="bar" style="width:{int(200 * n / max(1, max_hist))}px"></div></td></tr>'
        for label, n in hist
    )

    # Tabla top vocab
    vocab_rows = "\n".join(
        f'<tr><td class="weight-cell">{w:.2f}</td>'
        f'<td><span class="text-mono">{_esc(s)}</span></td>'
        f'<td><strong>{_esc(t)}</strong></td></tr>'
        for t, w, s in top_vocab
    )

    # Tabla transcripts
    def fmt_ts(ts: float | None) -> str:
        if ts is None:
            return "—"
        return datetime.fromtimestamp(ts).strftime("%m-%d %H:%M")

    def fmt_lp(lp: float | None) -> str:
        if lp is None:
            return '<span class="text-mono">—</span>'
        cls = "lp-good" if lp > -0.4 else ("lp-mid" if lp > -0.8 else "lp-bad")
        return f'<span class="{cls}">{lp:.2f}</span>'

    def fmt_duration(d: float | None) -> str:
        """Formatea duración en `Xs` o `Xm Ys` según length."""
        if d is None or d <= 0:
            return "—"
        if d < 60:
            return f"{d:.0f}s"
        m = int(d // 60)
        s = int(d % 60)
        return f"{m}m {s:02d}s"

    transcript_rows = ""
    for r in recent_transcripts:
        path, ts, lp, model, text, corrected, corr_src, chat_id, duration = r
        chat_label = (chat_id or "")[-20:] if chat_id else "—"
        text_disp = (corrected or text or "")[:100]
        marker = ""
        if corr_src == "llm":
            marker = ' <span class="pill pill-llm">llm</span>'
        elif corr_src == "explicit":
            marker = ' <span class="pill pill-fix">/fix</span>'
        elif corr_src == "vault_diff":
            marker = ' <span class="pill pill-vault">vault</span>'
        transcript_rows += (
            f'<tr>'
            f'<td><span class="text-mono">{_esc(fmt_ts(ts))}</span></td>'
            f'<td><span class="text-mono">{_esc(fmt_duration(duration))}</span></td>'
            f'<td>{fmt_lp(lp)}</td>'
            f'<td><span class="text-mono">{_esc((model or "")[:25])}</span></td>'
            f'<td><span class="text-mono">{_esc(chat_label)}</span></td>'
            f'<td>{_esc(text_disp)}{marker}</td>'
            f'</tr>\n'
        )
    if not transcript_rows:
        transcript_rows = '<tr><td colspan="6" class="meta" style="text-align:center;padding:20px">sin transcripciones logueadas — mandá un audio por WhatsApp para que aparezca acá</td></tr>'

    # Tabla corrections
    correction_rows = ""
    for r in recent_corrections:
        ts, src, orig, corr = r
        src_cls = {"explicit": "src-explicit", "llm": "src-llm", "vault_diff": "src-vault"}.get(src, "")
        correction_rows += (
            f'<tr>'
            f'<td><span class="text-mono">{_esc(fmt_ts(ts))}</span></td>'
            f'<td><span class="{src_cls}">{_esc(src)}</span></td>'
            f'<td class="text-orig">{_esc((orig or "")[:80])}</td>'
            f'<td class="text-fixed">{_esc((corr or "")[:80])}</td>'
            f'</tr>\n'
        )
    if not correction_rows:
        correction_rows = '<tr><td colspan="4" class="meta" style="text-align:center;padding:20px">sin correcciones todavía — usá <code>/fix &lt;texto&gt;</code> en WhatsApp para corregir un audio</td></tr>'

    # Stats correcciones por source
    corr_summary_parts = []
    for src in ("explicit", "llm", "vault_diff"):
        n = corr_by_source.get(src, 0)
        if n > 0:
            cls = {"explicit": "src-explicit", "llm": "src-llm", "vault_diff": "src-vault"}[src]
            corr_summary_parts.append(f'<span class="{cls}">{n} {src}</span>')
    corr_summary = " · ".join(corr_summary_parts) if corr_summary_parts else "0"

    # Patterns repetidos — render compact list. Si no hay (count<2 en todos),
    # placeholder con expectativa.
    if patterns_data:
        pattern_rows_html = ""
        for p in patterns_data[:10]:  # cap top 10
            src_parts = []
            if p.sources.get("explicit", 0) > 0:
                src_parts.append(f'<span class="src-explicit">{p.sources["explicit"]} /fix</span>')
            if p.sources.get("llm", 0) > 0:
                src_parts.append(f'<span class="src-llm">{p.sources["llm"]} llm</span>')
            if p.sources.get("vault_diff", 0) > 0:
                src_parts.append(f'<span class="src-vault">{p.sources["vault_diff"]} vault</span>')
            src_label = " · ".join(src_parts) if src_parts else "—"
            pattern_rows_html += (
                f'<tr>'
                f'<td><span class="text-mono">×{p.count}</span></td>'
                f'<td class="text-orig">{_esc(p.original)}</td>'
                f'<td><span class="text-mono">→</span></td>'
                f'<td class="text-fixed">{_esc(p.corrected)}</td>'
                f'<td><span class="text-mono">{src_label}</span></td>'
                f'</tr>\n'
            )
        patterns_html = (
            f'<table>'
            f'<thead><tr><th>count</th><th>whisper dijo</th><th></th>'
            f'<th>corrección</th><th>fuente</th></tr></thead>'
            f'<tbody>{pattern_rows_html}</tbody>'
            f'</table>'
        )
    else:
        patterns_html = (
            '<p class="meta" style="padding:8px 0">'
            'sin patrones repetidos todavía (necesita ≥2 correcciones del mismo '
            'swap). cuando uses <code>/fix</code> seguido para corregir el mismo '
            'error, va a aparecer acá. también detectado vía <code>rag whisper patterns</code>.'
            '</p>'
        )

    # Heatmap por hora del día — render como una row de 24 cells coloreadas
    # según count (intensidad creciente). Si todos están en 0, muestra una
    # row plana con texto info.
    max_hour = max(hour_counts.values()) if hour_counts else 0
    if max_hour > 0:
        heatmap_cells = []
        for h in range(24):
            n = hour_counts.get(h, 0)
            # Intensidad de 0 a 1 — usa accent color con alpha variable.
            alpha = 0.0 if max_hour == 0 else (n / max_hour)
            # Background: blend del accent + bg-elev. Mostrar count en hover.
            bg = f"background:rgba(88,166,255,{alpha:.2f})" if alpha > 0 else "background:var(--border-soft)"
            label = f"{h:02d}"
            heatmap_cells.append(
                f'<td class="heatmap-cell" title="{h:02d}:00 — {n} audio(s)" style="{bg}">'
                f'<div class="hour-label">{label}</div>'
                f'<div class="hour-count">{n if n > 0 else ""}</div>'
                f'</td>'
            )
        heatmap_html = (
            f'<div class="heatmap-wrap">'
            f'<table class="heatmap"><tr>{"".join(heatmap_cells)}</tr></table>'
            f'<p class="meta">cada celda = 1 hora del día (30d) · intensidad ∝ count · max={max_hour}</p>'
            f'</div>'
        )
    else:
        heatmap_html = (
            '<p class="meta" style="padding:14px 0">'
            'sin audios en últimos 30d para construir heatmap. '
            'cuando llegue el primer audio, esta sección se va a poblar con '
            'la distribución horaria.'
            '</p>'
        )

    # Heatmap semanal día×hora — matriz 7×24. Reordenamos las rows visualmente
    # de lunes a domingo (sqlite devuelve 0=Sun..6=Sat, transformamos al
    # mapping europeo lun=0, mar=1, ..., dom=6).
    # `dow_sql_to_visual`: 0(Sun)→6, 1(Mon)→0, 2(Tue)→1, ..., 6(Sat)→5.
    dow_to_visual = {0: 6, 1: 0, 2: 1, 3: 2, 4: 3, 5: 4, 6: 5}
    visual_labels = ["lun", "mar", "mié", "jue", "vie", "sáb", "dom"]
    max_week = max(week_counts.values()) if week_counts else 0
    if max_week > 0:
        # Header con horas (00-23). Primera col es label del día.
        header_cells = '<th></th>' + ''.join(
            f'<th class="week-hour-hdr">{h:02d}</th>' for h in range(24)
        )
        rows_html = []
        for visual_idx, day_label in enumerate(visual_labels):
            sql_dow = next(d for d, v in dow_to_visual.items() if v == visual_idx)
            cells: list[str] = []
            cells.append(f'<th class="week-day-hdr">{day_label}</th>')
            for h in range(24):
                n = week_counts.get((sql_dow, h), 0)
                alpha = (n / max_week) if max_week > 0 else 0.0
                bg = (
                    f"background:rgba(88,166,255,{alpha:.2f})"
                    if alpha > 0
                    else "background:var(--border-soft)"
                )
                cells.append(
                    f'<td class="week-cell" title="{day_label} {h:02d}:00 — {n} audio(s)" style="{bg}">'
                    f'{n if n > 0 else ""}'
                    f'</td>'
                )
            rows_html.append(f'<tr>{"".join(cells)}</tr>')
        week_html = (
            f'<div class="heatmap-wrap">'
            f'<table class="week-heatmap"><thead><tr>{header_cells}</tr></thead>'
            f'<tbody>{"".join(rows_html)}</tbody></table>'
            f'<p class="meta">7×24 = 168 celdas (60d) · intensidad ∝ count · max={max_week}</p>'
            f'</div>'
        )
    else:
        week_html = (
            '<p class="meta" style="padding:14px 0">'
            'sin audios en últimos 60d para construir heatmap semanal. '
            'cuando hayan suficientes audios, vas a ver patrones tipo '
            '"lunes mañana" o "viernes a la tarde".'
            '</p>'
        )

    # Stats vocab por source
    vocab_summary_parts = []
    for src in ("corrections", "contacts", "notes", "chats"):
        n = vocab_by_source.get(src, 0)
        if n > 0:
            vocab_summary_parts.append(f'{n} {src}')
    vocab_summary = " · ".join(vocab_summary_parts) if vocab_summary_parts else "0"

    html = f"""<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="color-scheme" content="dark">
{"<!-- Auto-refresh suppressed via ?nofresh=1 -->" if nofresh else "<!-- Auto-refresh cada 60s. Override con ?nofresh=1 -->"}
{"" if nofresh else '<meta http-equiv="refresh" content="60">'}
<title>whisper transcripts — rag</title>
<style>
  /* Dark mode FIJO. Paleta cercana al GitHub dark + Obsidian default
     theme para coherencia con el resto del stack del usuario.
     Decisión 2026-04-25: removido el override `prefers-color-scheme:
     light` — el resto del stack (chat/dashboard/status/home) tiene tema
     oscuro como default y el user lo prefiere así, no querer que la
     página se flipee a light cuando el OS está en modo claro. Si en el
     futuro queremos agregar toggle manual (como dashboard.html), se
     hace sobre `data-theme="light"` en el <html>, NO con media query. */
  :root {{
    --bg: #0d1117;
    --bg-elev: #161b22;
    --bg-elev-2: #1f242c;
    --border: #30363d;
    --border-soft: #21262d;
    --text: #e6edf3;
    --text-muted: #8b949e;
    --text-dim: #6e7681;
    --accent: #58a6ff;
    --green: #3fb950;
    --orange: #d29922;
    --red: #f85149;
  }}
  html, body {{ background: var(--bg); }}
  body {{
    font: 14px/1.5 -apple-system, "SF Pro Text", system-ui, sans-serif;
    max-width: 1100px;
    margin: 20px auto;
    padding: 0 16px;
    color: var(--text);
  }}
  h1 {{ font-size: 22px; margin: 0 0 6px 0; color: var(--text); font-weight: 600; }}
  h2 {{ font-size: 16px; margin: 28px 0 8px 0; color: var(--text); border-bottom: 1px solid var(--border); padding-bottom: 6px; font-weight: 600; }}
  .summary {{ display: flex; flex-wrap: wrap; gap: 12px; margin: 16px 0 24px 0; }}
  .stat {{
    background: var(--bg-elev);
    padding: 12px 16px;
    border-radius: 8px;
    min-width: 150px;
    border: 1px solid var(--border-soft);
  }}
  .stat .num {{ font-size: 22px; font-weight: 600; color: var(--text); font-variant-numeric: tabular-nums; }}
  .stat .lbl {{ font-size: 11px; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.05em; margin-top: 2px; }}
  table {{ border-collapse: collapse; width: 100%; background: var(--bg); }}
  th, td {{ padding: 6px 10px; text-align: left; border-bottom: 1px solid var(--border-soft); vertical-align: top; }}
  th {{ background: var(--bg-elev); font-weight: 600; font-size: 12px; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.04em; border-bottom: 1px solid var(--border); }}
  tr:hover td {{ background: var(--bg-elev); }}
  .grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 24px; }}
  @media (max-width: 800px) {{ .grid-2 {{ grid-template-columns: 1fr; }} }}
  .meta {{ color: var(--text-muted); font-size: 12px; margin-top: 4px; }}
  code {{ background: var(--bg-elev-2); padding: 2px 6px; border-radius: 4px; font-size: 12px; font-family: ui-monospace, "SF Mono", Menlo, monospace; color: var(--text); }}
  /* Pills de marker: rebajan el fondo en dark mode para no quemar la vista */
  .pill {{ padding: 1px 6px; border-radius: 4px; font-size: 10px; font-weight: 500; }}
  .pill-llm {{ background: rgba(88, 166, 255, 0.18); color: var(--accent); }}
  .pill-fix {{ background: rgba(63, 185, 80, 0.18); color: var(--green); }}
  .pill-vault {{ background: rgba(210, 153, 34, 0.18); color: var(--orange); }}
  /* Histogram bars: usa accent del theme */
  .bar {{ background: var(--accent); height: 14px; border-radius: 2px; }}
  /* Logprob colors */
  .lp-good {{ color: var(--green); font-variant-numeric: tabular-nums; }}
  .lp-mid {{ color: var(--orange); font-variant-numeric: tabular-nums; }}
  .lp-bad {{ color: var(--red); font-variant-numeric: tabular-nums; }}
  /* Source colors en correcciones */
  .src-explicit {{ color: var(--green); }}
  .src-llm {{ color: var(--accent); }}
  .src-vault {{ color: var(--orange); }}
  .text-orig {{ color: var(--red); }}
  .text-fixed {{ color: var(--green); }}
  .text-mono {{ color: var(--text-muted); font-size: 11px; font-family: ui-monospace, "SF Mono", Menlo, monospace; }}
  .weight-cell {{ text-align: right; color: var(--text-muted); font-variant-numeric: tabular-nums; }}
  /* Topnav consistente con /dashboard y /home — links a las otras páginas */
  .topnav {{ display: flex; gap: 14px; padding: 8px 0; margin: 0 0 8px 0; border-bottom: 1px solid var(--border-soft); font-size: 13px; }}
  .topnav a {{ color: var(--text-muted); text-decoration: none; padding: 4px 0; border-bottom: 2px solid transparent; transition: border-color .12s ease, color .12s ease; }}
  .topnav a:hover {{ color: var(--text); }}
  .topnav a.active {{ color: var(--text); border-bottom-color: var(--accent); }}
  /* Heatmap por hora — 24 cells de ancho equal, height fijo, color del accent
     con alpha = count / max. Hover deja un title con el count exacto. */
  .heatmap-wrap {{ margin: 8px 0 4px 0; }}
  .heatmap {{ width: 100%; border-collapse: separate; border-spacing: 2px; table-layout: fixed; }}
  .heatmap-cell {{
    height: 38px;
    border-radius: 3px;
    text-align: center;
    border-bottom: none !important;
    padding: 4px 2px !important;
    cursor: default;
  }}
  .heatmap-cell .hour-label {{ font-size: 10px; color: var(--text-dim); font-variant-numeric: tabular-nums; line-height: 1; }}
  .heatmap-cell .hour-count {{ font-size: 11px; color: var(--text); font-weight: 600; line-height: 1.4; font-variant-numeric: tabular-nums; }}
  /* Heatmap semanal 7×24 — más denso que el horario, cells más chicas. */
  .week-heatmap {{ width: 100%; border-collapse: separate; border-spacing: 2px; table-layout: fixed; }}
  .week-heatmap th, .week-heatmap td {{ padding: 0 !important; border-bottom: none !important; }}
  .week-heatmap .week-day-hdr {{ width: 40px; text-align: right; padding-right: 8px !important; color: var(--text-muted); font-size: 11px; font-weight: 500; text-transform: lowercase; letter-spacing: 0; background: transparent; }}
  .week-heatmap .week-hour-hdr {{ font-size: 10px; color: var(--text-dim); font-weight: 400; font-variant-numeric: tabular-nums; text-transform: none; letter-spacing: 0; background: transparent; }}
  .week-cell {{
    height: 22px;
    border-radius: 2px;
    text-align: center;
    font-size: 10px;
    color: var(--text);
    font-variant-numeric: tabular-nums;
    cursor: default;
  }}
  /* Refresh button al lado del título h1: link simple que recarga la página
     sin esperar el auto-refresh de 60s. Visible pero no intrusivo. */
  .refresh-btn {{
    color: var(--text-muted);
    text-decoration: none;
    font-size: 16px;
    margin-left: 8px;
    padding: 0 6px;
    border-radius: 4px;
    transition: color .12s ease, background .12s ease;
    display: inline-block;
  }}
  .refresh-btn:hover {{ color: var(--accent); background: var(--bg-elev); }}
</style>
</head>
<body>
  <nav class="topnav">
    <a href="/">home</a>
    <a href="/chat">chat</a>
    <a href="/dashboard">dashboard</a>
    <a href="/status">status</a>
    <a href="/transcripts" class="active">transcripts</a>
    <a href="https://ra.ai/agents/">agents</a>
  </nav>
  <h1>whisper transcripts <a href="/transcripts" class="refresh-btn" title="recargar ahora">↻</a></h1>
  <p class="meta">phase 2 learning loop · vocab refresh: {_esc(last_refresh_str)}{" · auto-refresh OFF" if nofresh else ""}</p>

  <div class="summary">
    <div class="stat"><div class="num">{n_30d}</div><div class="lbl">audios 30d ({n_total} total)</div></div>
    <div class="stat"><div class="num">{avg_lp_str}</div><div class="lbl">avg logprob (30d)</div></div>
    <div class="stat"><div class="num">{avg_dur_str}</div><div class="lbl">avg duración (30d)</div></div>
    <div class="stat"><div class="num">{n_corrections}</div><div class="lbl">correcciones totales</div></div>
    <div class="stat"><div class="num">{n_vocab}</div><div class="lbl">vocab terms</div></div>
  </div>

  <div class="grid-2">
    <div>
      <h2>logprob histogram</h2>
      <p class="meta">distribución de confianza de las transcripciones. logprob &lt; -0.8 dispara LLM auto-correct.</p>
      <table>{hist_rows}</table>
    </div>
    <div>
      <h2>correcciones por source</h2>
      <p class="meta">{corr_summary}</p>
      <h2 style="margin-top: 24px">vocab por source</h2>
      <p class="meta">{vocab_summary}</p>
    </div>
  </div>

  <h2>distribución horaria (30d)</h2>
  {heatmap_html}

  <h2>distribución semanal (60d)</h2>
  {week_html}

  <h2>últimas 30 transcripciones</h2>
  <table>
    <thead><tr><th>fecha</th><th>duración</th><th>logprob</th><th>modelo</th><th>chat</th><th>texto</th></tr></thead>
    <tbody>{transcript_rows}</tbody>
  </table>

  <h2>últimas 20 correcciones</h2>
  <p class="meta">source <span class="src-explicit">explicit</span> = comando /fix · <span class="src-llm">llm</span> = qwen2.5:7b auto-correct · <span class="src-vault">vault_diff</span> = nota editada</p>
  <table>
    <thead><tr><th>fecha</th><th>source</th><th>original</th><th>corregido</th></tr></thead>
    <tbody>{correction_rows}</tbody>
  </table>

  <h2>patrones repetidos (≥2 swaps)</h2>
  <p class="meta">errores sistemáticos del modelo whisper que se corrigieron varias veces. señal fuerte para el vocab. mismo algoritmo que <code>rag whisper patterns</code>.</p>
  {patterns_html}

  <h2>top 50 vocab terms (peso DESC)</h2>
  <p class="meta">se inyecta al <code>--prompt</code> de whisper en cada transcripción. corregido vía <code>rag whisper vocab refresh</code>.</p>
  <table>
    <thead><tr><th style="text-align:right">peso</th><th>source</th><th>term</th></tr></thead>
    <tbody>{vocab_rows}</tbody>
  </table>
</body>
</html>"""
    return HTMLResponse(content=html)


# ── /status — service health page ────────────────────────────────────
# Una página sola que lista TODOS los componentes que el sistema necesita
# para funcionar — db, ollama, tunnel, whatsapp, los 20+ daemons +
# scheduled jobs — con un semáforo (verde/amarillo/rojo) por servicio.
# Pensado para "prendí la Mac y chequeo qué está vivo de un vistazo" sin
# tener que acordarse de 20 labels de launchctl.
#
# Semántica de estados:
#   ok       (verde)  → el servicio está funcionando como corresponde.
#   warn     (amarillo) → loaded pero aún no corrió, o info faltante
#                         (ej. scheduled sin primera corrida todavía).
#   down     (rojo)   → debería estar corriendo y no está, o última
#                        corrida terminó con exit != 0.
#
# Tres "kinds" de check:
#   daemon    → KeepAlive=true; debería estar running 24/7.
#   scheduled → StartCalendarInterval/StartInterval/RunAtLoad-oneshot;
#               se gradúa por last-exit-code (pasa la mayor parte del
#               tiempo not-running por diseño).
#   probe     → HTTP/filesystem probe directo (Ollama, DB file, web
#               self, vault, tunnel URL freshness).
#
# Las checks corren en ThreadPoolExecutor para que el endpoint responda
# en <500ms aun con ~25 targets (cada launchctl print son ~20-40ms).


_STATUS_LAUNCHD_UID: int | None = None


def _launchd_uid_cached() -> int:
    global _STATUS_LAUNCHD_UID
    if _STATUS_LAUNCHD_UID is None:
        _STATUS_LAUNCHD_UID = os.getuid()
    return _STATUS_LAUNCHD_UID


def _launchctl_print_fields(label: str, timeout: float = 3.0) -> dict[str, str] | None:
    """Run `launchctl print gui/<uid>/<label>` and parse top-level fields.

    Returns None si el servicio no está cargado (launchctl exit != 0).
    Sólo extrae las keys top-level del bloque principal — las nested
    (endpoints, sockets, domain) las descartamos porque inflan el output
    y no las usamos. La heurística es "línea que empieza con exactamente
    UN tab"; nested bloques empiezan con 2+ tabs.
    """
    try:
        out = subprocess.run(
            ["/bin/launchctl", "print", f"gui/{_launchd_uid_cached()}/{label}"],
            capture_output=True,
            text=True, errors="replace",
            timeout=timeout,
            check=False,
        )
    except (subprocess.TimeoutExpired, FileNotFoundError, OSError):
        return None
    if out.returncode != 0:
        return None
    fields: dict[str, str] = {}
    for line in out.stdout.splitlines():
        # Top-level key: exactamente un leading tab, no dos.
        if not (line.startswith("\t") and not line.startswith("\t\t")):
            continue
        stripped = line.strip()
        if " = " not in stripped:
            continue
        k, _, v = stripped.partition(" = ")
        k = k.strip()
        v = v.strip()
        if k and k not in fields:
            fields[k] = v
    return fields


def _status_fmt_size(n: int) -> str:
    """Byte count → human string (5 MB, 123 KB, 1.2 GB)."""
    if n < 1024:
        return f"{n} B"
    size = float(n)
    for unit in ("KB", "MB", "GB", "TB"):
        size /= 1024.0
        if size < 1024.0:
            # 1 decimal si es <10, 0 si es >=10 (más compacto)
            return f"{size:.1f} {unit}" if size < 10 else f"{size:.0f} {unit}"
    return f"{size:.0f} PB"


def _status_fmt_age(seconds: float) -> str:
    """Segundos → "hace 3m" / "hace 2h" / "hace 5d"."""
    seconds = max(0.0, seconds)
    if seconds < 60:
        return f"hace {int(seconds)}s"
    mins = seconds / 60.0
    if mins < 60:
        return f"hace {int(mins)}m"
    hours = mins / 60.0
    if hours < 24:
        return f"hace {int(hours)}h"
    days = hours / 24.0
    return f"hace {int(days)}d"


def _status_grade_daemon(label: str, name: str) -> dict:
    """Grade KeepAlive daemon: running = ok, anything else = down.

    El payload incluye `label` + `running` + `loaded` para que la UI
    decida qué botón mostrar (start / stop) y si la acción puede
    fallar de antemano.
    """
    info = _launchctl_print_fields(label)
    if info is None:
        return {"id": label, "name": name, "kind": "daemon", "status": "down",
                "detail": "no cargado en launchd",
                "label": label, "loaded": False, "running": False}
    state = info.get("state", "")
    pid = info.get("pid")
    last_exit = info.get("last exit code", "")
    if state == "running":
        detail = f"pid {pid}" if pid else "running"
        resp = {"id": label, "name": name, "kind": "daemon", "status": "ok",
                "detail": detail,
                "label": label, "loaded": True, "running": True}
        if pid:
            resp["pid"] = pid
        return resp
    det = "not running"
    if last_exit and last_exit not in ("(never exited)", "0"):
        det = f"crashed · exit {last_exit}"
    return {"id": label, "name": name, "kind": "daemon", "status": "down",
            "detail": det,
            "label": label, "loaded": True, "running": False}


def _status_grade_scheduled(label: str, name: str) -> dict:
    """Grade scheduled/oneshot job: last-exit-code drives the light.

    - state=running → ok (corriendo justo ahora)
    - exit 0 → ok
    - exit N (nonzero) → down
    - never exited + runs=0 → warn (loaded, no corrió todavía)
    - never exited + runs>0 → ok (inusual pero válido)
    - no cargado → warn (no down porque algunos están desactivados a mano)
    """
    info = _launchctl_print_fields(label)
    if info is None:
        return {"id": label, "name": name, "kind": "scheduled", "status": "warn",
                "detail": "no cargado",
                "label": label, "loaded": False, "running": False}
    state = info.get("state", "")
    runs = info.get("runs", "0")
    last_exit = info.get("last exit code", "(never exited)")
    if state == "running":
        pid = info.get("pid", "?")
        return {"id": label, "name": name, "kind": "scheduled", "status": "ok",
                "detail": f"corriendo · pid {pid} · runs {runs}",
                "label": label, "loaded": True, "running": True,
                "pid": str(pid)}
    if last_exit == "0":
        return {"id": label, "name": name, "kind": "scheduled", "status": "ok",
                "detail": f"última OK · runs {runs}",
                "label": label, "loaded": True, "running": False}
    if last_exit == "(never exited)":
        if runs == "0":
            return {"id": label, "name": name, "kind": "scheduled", "status": "warn",
                    "detail": "aún no corrió",
                    "label": label, "loaded": True, "running": False}
        return {"id": label, "name": name, "kind": "scheduled", "status": "ok",
                "detail": f"runs {runs}",
                "label": label, "loaded": True, "running": False}
    return {"id": label, "name": name, "kind": "scheduled", "status": "down",
            "detail": f"última exit {last_exit} · runs {runs}",
            "label": label, "loaded": True, "running": False}


def _status_probe_self() -> dict:
    return {"id": "web-self", "name": "Web server (FastAPI)", "kind": "probe",
            "status": "ok", "detail": f"pid {os.getpid()}"}


def _status_probe_ollama() -> dict:
    t0 = time.monotonic()
    ok = _ollama_alive(timeout=2.0)
    ms = int((time.monotonic() - t0) * 1000)
    if ok:
        return {"id": "ollama", "name": "Ollama (LLM runtime)", "kind": "probe",
                "status": "ok", "detail": f"/api/tags {ms}ms"}
    return {"id": "ollama", "name": "Ollama (LLM runtime)", "kind": "probe",
            "status": "down", "detail": f"no responde /api/tags ({ms}ms)"}


def _status_probe_rag_db() -> dict:
    """Check ragvec.db exists + is readable. Size + mtime for visibility."""
    try:
        from rag import DB_PATH as _DB_DIR  # noqa: PLC0415
        dbf = _DB_DIR / "ragvec.db"
        if not dbf.is_file():
            return {"id": "rag-db", "name": "RAG DB (ragvec)", "kind": "probe",
                    "status": "down", "detail": f"falta {dbf}"}
        st = dbf.stat()
        age = time.time() - st.st_mtime
        return {"id": "rag-db", "name": "RAG DB (ragvec)", "kind": "probe",
                "status": "ok",
                "detail": f"{_status_fmt_size(st.st_size)} · modif. {_status_fmt_age(age)}"}
    except Exception as e:
        return {"id": "rag-db", "name": "RAG DB (ragvec)", "kind": "probe",
                "status": "down", "detail": f"{type(e).__name__}: {e}"}


def _status_probe_telemetry_db() -> dict:
    try:
        from rag import DB_PATH as _DB_DIR  # noqa: PLC0415
        dbf = _DB_DIR / "telemetry.db"
        if not dbf.is_file():
            return {"id": "telemetry-db", "name": "Telemetry DB", "kind": "probe",
                    "status": "warn", "detail": "aún no creada"}
        st = dbf.stat()
        age = time.time() - st.st_mtime
        return {"id": "telemetry-db", "name": "Telemetry DB", "kind": "probe",
                "status": "ok",
                "detail": f"{_status_fmt_size(st.st_size)} · modif. {_status_fmt_age(age)}"}
    except Exception as e:
        return {"id": "telemetry-db", "name": "Telemetry DB", "kind": "probe",
                "status": "warn", "detail": f"{type(e).__name__}: {e}"}


def _status_probe_vault() -> dict:
    try:
        if not VAULT_PATH.is_dir():
            return {"id": "vault", "name": "Obsidian vault", "kind": "probe",
                    "status": "down", "detail": f"no existe {VAULT_PATH}"}
        return {"id": "vault", "name": "Obsidian vault", "kind": "probe",
                "status": "ok", "detail": str(VAULT_PATH)}
    except Exception as e:
        return {"id": "vault", "name": "Obsidian vault", "kind": "probe",
                "status": "down", "detail": f"{type(e).__name__}: {e}"}


def _status_probe_wa_db() -> dict:
    try:
        if not WHATSAPP_DB_PATH.is_file():
            return {"id": "wa-db", "name": "WhatsApp bridge SQLite", "kind": "probe",
                    "status": "warn", "detail": "no existe (bridge todavía no inició)"}
        st = WHATSAPP_DB_PATH.stat()
        age = time.time() - st.st_mtime
        return {"id": "wa-db", "name": "WhatsApp bridge SQLite", "kind": "probe",
                "status": "ok",
                "detail": f"{_status_fmt_size(st.st_size)} · modif. {_status_fmt_age(age)}"}
    except Exception as e:
        return {"id": "wa-db", "name": "WhatsApp bridge SQLite", "kind": "probe",
                "status": "warn", "detail": f"{type(e).__name__}: {e}"}


def _status_probe_tunnel_url() -> dict:
    """Read ~/.local/share/obsidian-rag/cloudflared-url.txt to surface the
    current public URL and its freshness. Tunnel watcher re-writes this
    file whenever it re-establishes the quick-tunnel.
    """
    try:
        url_file = Path.home() / ".local/share/obsidian-rag/cloudflared-url.txt"
        if not url_file.is_file():
            return {"id": "tunnel-url", "name": "Tunnel URL", "kind": "probe",
                    "status": "warn", "detail": "sin archivo cloudflared-url.txt"}
        url = url_file.read_text(encoding="utf-8", errors="replace").strip()
        st = url_file.stat()
        age = time.time() - st.st_mtime
        if not url:
            return {"id": "tunnel-url", "name": "Tunnel URL", "kind": "probe",
                    "status": "warn", "detail": "archivo vacío"}
        return {"id": "tunnel-url", "name": "Tunnel URL", "kind": "probe",
                "status": "ok", "detail": f"{url} · {_status_fmt_age(age)}",
                "meta": {"url": url}}
    except Exception as e:
        return {"id": "tunnel-url", "name": "Tunnel URL", "kind": "probe",
                "status": "warn", "detail": f"{type(e).__name__}: {e}"}


# Catálogo de servicios. Orden = orden de aparición en la UI.
# "category" agrupa en secciones; "kind" dispatchea al helper correcto.
_STATUS_CATALOG: list[dict] = [
    # Core: sin esto no hay sistema.
    {"category": "core", "category_label": "Core", "kind": "self", "id": "web-self"},
    {"category": "core", "category_label": "Core", "kind": "ollama", "id": "ollama"},
    {"category": "core", "category_label": "Core", "kind": "rag_db", "id": "rag-db"},
    {"category": "core", "category_label": "Core", "kind": "telemetry_db", "id": "telemetry-db"},
    {"category": "core", "category_label": "Core", "kind": "vault", "id": "vault"},
    {"category": "core", "category_label": "Core", "kind": "daemon",
     "target": "com.fer.obsidian-rag-web", "name": "Web daemon (launchd)"},
    {"category": "core", "category_label": "Core", "kind": "scheduled",
     "target": "com.fer.ollama-env", "name": "Ollama env shim"},

    # Tunnel: HTTPS público vía Cloudflare.
    {"category": "tunnel", "category_label": "Cloudflare tunnel", "kind": "daemon",
     "target": "com.fer.obsidian-rag-cloudflare-tunnel", "name": "Tunnel (quick)"},
    {"category": "tunnel", "category_label": "Cloudflare tunnel", "kind": "daemon",
     "target": "com.fer.obsidian-rag-cloudflare-tunnel-watcher", "name": "Tunnel watcher"},
    {"category": "tunnel", "category_label": "Cloudflare tunnel", "kind": "daemon",
     "target": "com.fer.cloudflared", "name": "cloudflared (base)"},
    {"category": "tunnel", "category_label": "Cloudflare tunnel", "kind": "tunnel_url",
     "id": "tunnel-url"},

    # WhatsApp: bridge + listener + vault sync.
    {"category": "whatsapp", "category_label": "WhatsApp", "kind": "daemon",
     "target": "com.fer.whatsapp-bridge", "name": "Bridge (Go)"},
    {"category": "whatsapp", "category_label": "WhatsApp", "kind": "daemon",
     "target": "com.fer.whatsapp-listener", "name": "Listener (ambient agent)"},
    {"category": "whatsapp", "category_label": "WhatsApp", "kind": "scheduled",
     "target": "com.fer.whatsapp-vault-sync", "name": "Vault sync"},
    # Worker que entrega mensajes WhatsApp programados (cola de
    # rag_whatsapp_scheduled). Si este daemon está caído, los mensajes
    # programados via /chat o dashboard no se mandan — el user lo ve
    # en /status y puede reactivarlo con `rag setup`.
    {"category": "whatsapp", "category_label": "WhatsApp", "kind": "scheduled",
     "target": "com.fer.obsidian-rag-wa-scheduled-send",
     "name": "Scheduled send worker"},
    {"category": "whatsapp", "category_label": "WhatsApp", "kind": "wa_db", "id": "wa-db"},

    # Cross-source ingesters.
    {"category": "ingest", "category_label": "Cross-source ingesters", "kind": "scheduled",
     "target": "com.fer.obsidian-rag-ingest-gmail", "name": "Gmail"},
    {"category": "ingest", "category_label": "Cross-source ingesters", "kind": "scheduled",
     "target": "com.fer.obsidian-rag-ingest-calendar", "name": "Calendar"},
    {"category": "ingest", "category_label": "Cross-source ingesters", "kind": "scheduled",
     "target": "com.fer.obsidian-rag-ingest-reminders", "name": "Reminders"},
    {"category": "ingest", "category_label": "Cross-source ingesters", "kind": "scheduled",
     "target": "com.fer.obsidian-rag-ingest-whatsapp", "name": "WhatsApp → vault"},
    {"category": "ingest", "category_label": "Cross-source ingesters", "kind": "scheduled",
     "target": "com.fer.obsidian-rag-wa-tasks", "name": "WA tasks extractor"},

    # Briefs + automation.
    {"category": "briefs", "category_label": "Briefs + automation", "kind": "scheduled",
     "target": "com.fer.obsidian-rag-wake-up", "name": "Wake-up pack (04:00)"},
    {"category": "briefs", "category_label": "Briefs + automation", "kind": "scheduled",
     "target": "com.fer.obsidian-rag-morning", "name": "Morning brief"},
    {"category": "briefs", "category_label": "Briefs + automation", "kind": "scheduled",
     "target": "com.fer.obsidian-rag-today", "name": "Today brief"},
    {"category": "briefs", "category_label": "Briefs + automation", "kind": "scheduled",
     "target": "com.fer.obsidian-rag-digest", "name": "Digest"},
    {"category": "briefs", "category_label": "Briefs + automation", "kind": "daemon",
     "target": "com.fer.obsidian-rag-watch", "name": "rag watch (ambient)"},
    {"category": "briefs", "category_label": "Briefs + automation", "kind": "scheduled",
     "target": "com.fer.morning-briefing", "name": "Morning briefing (legacy)"},

    # Maintenance + vault health.
    {"category": "maintenance", "category_label": "Maintenance + vault health", "kind": "scheduled",
     "target": "com.fer.obsidian-rag-maintenance", "name": "Maintenance"},
    {"category": "maintenance", "category_label": "Maintenance + vault health", "kind": "scheduled",
     "target": "com.fer.obsidian-rag-online-tune", "name": "Online tune (ranker-vivo)"},
    {"category": "maintenance", "category_label": "Maintenance + vault health", "kind": "scheduled",
     "target": "com.fer.obsidian-rag-patterns", "name": "Patterns (radar)"},
    {"category": "maintenance", "category_label": "Maintenance + vault health", "kind": "scheduled",
     "target": "com.fer.obsidian-rag-emergent", "name": "Emergent radar"},
    {"category": "maintenance", "category_label": "Maintenance + vault health", "kind": "scheduled",
     "target": "com.fer.obsidian-rag-archive", "name": "Archive"},
    {"category": "maintenance", "category_label": "Maintenance + vault health", "kind": "scheduled",
     "target": "com.fer.obsidian-rag-consolidate", "name": "Consolidate"},
    {"category": "maintenance", "category_label": "Maintenance + vault health", "kind": "scheduled",
     "target": "com.fer.mcp-orphan-reaper", "name": "MCP orphan reaper"},

    # Optional: servicios que corren pero no son críticos para el chat.
    {"category": "optional", "category_label": "Optional", "kind": "daemon",
     "target": "com.fer.obsidian-rag-serve", "name": "rag serve (HTTP API)"},
]


def _status_dispatch_one(entry: dict) -> dict:
    """Run the correct grader for a single catalog entry."""
    kind = entry["kind"]
    if kind == "self":
        r = _status_probe_self()
    elif kind == "ollama":
        r = _status_probe_ollama()
    elif kind == "rag_db":
        r = _status_probe_rag_db()
    elif kind == "telemetry_db":
        r = _status_probe_telemetry_db()
    elif kind == "vault":
        r = _status_probe_vault()
    elif kind == "wa_db":
        r = _status_probe_wa_db()
    elif kind == "tunnel_url":
        r = _status_probe_tunnel_url()
    elif kind == "daemon":
        r = _status_grade_daemon(entry["target"], entry.get("name", entry["target"]))
    elif kind == "scheduled":
        r = _status_grade_scheduled(entry["target"], entry.get("name", entry["target"]))
    else:
        r = {"id": entry.get("id", entry.get("target", kind)),
             "name": entry.get("name", kind), "kind": kind,
             "status": "warn", "detail": f"kind desconocido: {kind}"}
    r["category"] = entry["category"]
    r["category_label"] = entry["category_label"]
    return r


# Cache corto del payload entero — si la UI auto-refreshea cada 5s, no
# queremos hacer 25 subprocess.run por cada request. 3s es suficiente
# para que un F5 manual vea cambios recientes sin hammerear launchctl.
_STATUS_CACHE: dict = {"ts": 0.0, "payload": None}
_STATUS_CACHE_TTL = 3.0
_STATUS_CACHE_LOCK = threading.Lock()


def _status_build_payload() -> dict:
    """Run todos los checks en paralelo y agruparlos por categoría.

    El payload final tiene shape:
      {
        "generated_at": "ISO-8601 UTC",
        "overall": "ok" | "degraded" | "down",
        "counts": {"ok": N, "warn": N, "down": N},
        "categories": [
           {"id": "core", "label": "Core", "services": [ {...}, ... ]},
           ...
        ]
      }

    `overall` surfaces un estado agregado del sistema:
      - down  si algún servicio de `core` está down (el chat no funciona)
      - down  si Ollama o rag-db están down (lo mismo)
      - degraded si hay algún warn o down en no-core
      - ok    si todo verde
    """
    import concurrent.futures  # noqa: PLC0415

    results: list[dict] = [{} for _ in _STATUS_CATALOG]
    # Thread pool con >1 worker aun para checks rápidos — el cuello es
    # los ~25 `launchctl print` (20-40ms c/u). Con 8 workers, ~100ms
    # total vs. 800ms seriales.
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as pool:
        futs = {pool.submit(_status_dispatch_one, entry): i
                for i, entry in enumerate(_STATUS_CATALOG)}
        for fut in concurrent.futures.as_completed(futs):
            i = futs[fut]
            try:
                results[i] = fut.result()
            except Exception as e:
                entry = _STATUS_CATALOG[i]
                results[i] = {
                    "id": entry.get("id", entry.get("target", "?")),
                    "name": entry.get("name", "?"),
                    "kind": entry.get("kind", "?"),
                    "status": "down",
                    "detail": f"error del check: {type(e).__name__}: {e}",
                    "category": entry["category"],
                    "category_label": entry["category_label"],
                }

    # Agrupar manteniendo el orden del catálogo.
    categories: list[dict] = []
    cat_index: dict[str, int] = {}
    counts = {"ok": 0, "warn": 0, "down": 0}
    core_down = False
    any_non_ok = False
    for svc in results:
        cat_id = svc["category"]
        if cat_id not in cat_index:
            cat_index[cat_id] = len(categories)
            categories.append({"id": cat_id, "label": svc["category_label"],
                               "services": []})
        categories[cat_index[cat_id]]["services"].append(svc)
        st = svc.get("status", "down")
        counts[st] = counts.get(st, 0) + 1
        if st != "ok":
            any_non_ok = True
        if cat_id == "core" and st == "down":
            core_down = True

    overall = "ok"
    if core_down:
        overall = "down"
    elif any_non_ok:
        overall = "degraded"

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "overall": overall,
        "counts": counts,
        "categories": categories,
    }


@app.get("/api/status")
def api_status(nocache: int = 0) -> dict:
    """JSON health dump para la /status page. `?nocache=1` fuerza refresh.

    Side effect: alimenta `rag_status_samples` con un sample del estado
    actual de los servicios core (rate-limited a 60s globalmente). La
    persistencia es necesaria para el heatmap del card #5 — el resto de
    los cards se computan on-demand de fuentes pre-existentes.
    """
    now = time.monotonic()
    with _STATUS_CACHE_LOCK:
        cached = _STATUS_CACHE["payload"]
        fresh = (now - _STATUS_CACHE["ts"]) < _STATUS_CACHE_TTL
    if nocache or not fresh or cached is None:
        payload = _status_build_payload()
        with _STATUS_CACHE_LOCK:
            _STATUS_CACHE["payload"] = payload
            _STATUS_CACHE["ts"] = now
        # Hook de persistencia para el heatmap del uptime card (#5).
        # Rate-limited internamente a 60s; cualquier excepción se traga
        # para no bajar el endpoint principal.
        _persist_status_samples(payload)
        return payload
    return cached


# ── /api/status/action — start / stop launchd-controlled services ────
# Permite "trigger digest now" o "stop ambient agent" desde la UI sin
# tener que abrir terminal y acordarse del label exacto. Whitelist
# estricta: solo labels que ya estan en _STATUS_CATALOG (los `target`
# de daemon + scheduled). Cualquier otro label devuelve 400.
#
# Acciones soportadas:
#   start -> `launchctl kickstart gui/<uid>/<label>` (launchd corre el
#            job ahora, sin esperar al siguiente trigger). Si esta
#            running, no-op (kickstart no relanza por default).
#   stop  -> `launchctl kill SIGTERM gui/<uid>/<label>` (manda senal al
#            proceso; para daemons KeepAlive launchd lo restartea solo,
#            para scheduled/oneshot se va y queda not-running).
#
# El server NO tiene auth (ver CLAUDE.md sobre el modo LAN-exposed):
# el whitelist es la unica defensa. Es OK porque los labels son todos
# `com.fer.*` propios del usuario y el subset es chico (~25). Si
# alguien con acceso de red dispara un kickstart del digest, el peor
# caso es que el digest corra unos minutos antes - no perdida de datos.


def _status_actionable_labels() -> set[str]:
    """Set of labels que la UI puede start/stop. Derivado del catalogo
    para no hardcodear dos veces - si un servicio esta en el catalogo
    como daemon/scheduled, es actionable."""
    return {
        e["target"] for e in _STATUS_CATALOG
        if e.get("kind") in ("daemon", "scheduled") and e.get("target")
    }


class StatusActionRequest(BaseModel):
    label: str
    action: str  # "start" | "stop"


@app.post("/api/status/action")
def status_action(req: StatusActionRequest) -> dict:
    """Disparar o parar un servicio launchd controlado. Whitelist-only."""
    label = (req.label or "").strip()
    action = (req.action or "").strip().lower()
    if action not in ("start", "stop"):
        raise HTTPException(status_code=400,
                            detail=f"action invalida: {action!r} (esperaba start|stop)")
    if label not in _status_actionable_labels():
        raise HTTPException(status_code=400,
                            detail=f"label no whitelisted: {label!r}")

    target = f"gui/{_launchd_uid_cached()}/{label}"
    if action == "start":
        cmd = ["/bin/launchctl", "kickstart", target]
    else:  # stop
        cmd = ["/bin/launchctl", "kill", "SIGTERM", target]

    try:
        out = subprocess.run(
            cmd, capture_output=True, text=True, errors="replace", timeout=10.0, check=False,
        )
    except (subprocess.TimeoutExpired, FileNotFoundError, OSError) as e:
        raise HTTPException(status_code=500,
                            detail=f"launchctl fallo: {type(e).__name__}: {e}")

    # Invalidar cache para que el proximo /api/status refleje el cambio
    # (kickstart cambia state a running en <1s; sin esto la UI mostraria
    # info vieja por hasta 3s - el TTL del cache).
    with _STATUS_CACHE_LOCK:
        _STATUS_CACHE["ts"] = 0.0
        _STATUS_CACHE["payload"] = None

    return {
        "ok": out.returncode == 0,
        "label": label,
        "action": action,
        "returncode": out.returncode,
        "stdout": (out.stdout or "").strip(),
        "stderr": (out.stderr or "").strip(),
    }


# ── /api/status/latency — serie p50/p95 para el sparkline del /status ─
# Lee rag_queries (cmd LIKE 'web%', extra_json.total_ms) y bucketiza las
# últimas 24h en percentiles horarios con window functions (nearest-rank,
# mismo método que test_analytics_p50_over_ttft_ms). Suma un baseline 7d
# para que la UI muestre "last-hour p95 vs typical p95" — la regresión
# de latencia es el síntoma más común que el semáforo binario no captura.
#
# Presupuesto: típicamente ~24h × ~50 queries/h ~= 1200 rows (ix_rag_
# queries_cmd_ts covering index la vuelve sub-50ms). Cacheado 60s porque
# los buckets horarios no cambian mid-hour y el /status polea cada 10s.
#
# Schema del payload (contract con status.js):
#   {
#     "window_hours": 24, "bucket": "hour",
#     "series": [                  # 25 elementos (hora actual incluida)
#       {"ts": "2026-04-24T12:00:00Z", "count": 42,
#        "p50_ms": 2100, "p95_ms": 8700, "p99_ms": 14200}, ...
#     ],
#     "summary": {
#       "p50_1h_ms": ..., "p95_1h_ms": ...,
#       "p50_baseline_ms": ..., "p95_baseline_ms": ...,   # 7d percentiles
#       "delta_p95_pct": -15.3,                           # vs baseline
#       "count_24h": 342,
#     },
#   }
# Valores null para buckets vacíos (el frontend les pone gap en el SVG).

_LATENCY_CACHE_TTL = 60.0
_LATENCY_CACHE: dict = {"ts": 0.0, "payload": None}
_LATENCY_CACHE_LOCK = threading.Lock()


def _status_latency_build_payload() -> dict:
    """Build the latency series + summary payload for /api/status/latency.

    Percentiles use the nearest-rank method per SQL window function — the
    same pattern as test_web_stage_timing_persisted::test_analytics_p50_
    over_ttft_ms so the contract stays consistent across analytics surfaces.
    Empty buckets are filled with nulls so the frontend gets a predictable
    25-element series (24 hours back + current) regardless of traffic.
    """
    now = datetime.now(timezone.utc)
    series: list[dict] = []
    summary: dict = {
        "p50_1h_ms": None,
        "p95_1h_ms": None,
        "p50_baseline_ms": None,
        "p95_baseline_ms": None,
        "delta_p95_pct": None,
        "count_24h": 0,
    }

    try:
        with _ragvec_state_conn() as conn:
            # Per-hour percentiles for the last 24h. `cmd LIKE 'web%'`
            # cubre 'web' + 'web.chat.low_conf_bypass' (y cualquier futuro
            # 'web.chat.*' que mantenga el prefix convention).
            rows = conn.execute(
                """
                WITH base AS (
                    SELECT
                        strftime('%Y-%m-%dT%H:00:00Z', ts) AS bucket,
                        CAST(json_extract(extra_json, '$.total_ms') AS INTEGER) AS v
                    FROM rag_queries
                    WHERE cmd LIKE 'web%'
                      AND ts >= datetime('now', '-24 hours')
                      AND json_extract(extra_json, '$.total_ms') IS NOT NULL
                ),
                ranked AS (
                    SELECT
                        bucket, v,
                        ROW_NUMBER() OVER (PARTITION BY bucket ORDER BY v) AS rn,
                        COUNT(*) OVER (PARTITION BY bucket) AS n
                    FROM base
                )
                SELECT
                    bucket,
                    MAX(n) AS count,
                    MIN(CASE WHEN rn * 1.0 / n >= 0.5  THEN v END) AS p50,
                    MIN(CASE WHEN rn * 1.0 / n >= 0.95 THEN v END) AS p95,
                    MIN(CASE WHEN rn * 1.0 / n >= 0.99 THEN v END) AS p99
                FROM ranked
                GROUP BY bucket
                ORDER BY bucket
                """,
            ).fetchall()
            by_bucket = {r[0]: r for r in rows}

            # Emitir 25 buckets (24h atrás + hora actual) en orden cronológico.
            # Los huecos quedan con count=0/None para que el sparkline dibuje
            # un gap visible en vez de interpolar una línea falsa.
            current = now.replace(minute=0, second=0, microsecond=0)
            for i in range(24, -1, -1):
                t = current - timedelta(hours=i)
                key = t.strftime("%Y-%m-%dT%H:00:00Z")
                r = by_bucket.get(key)
                if r is not None:
                    series.append({
                        "ts": key,
                        "count": int(r[1] or 0),
                        "p50_ms": int(r[2]) if r[2] is not None else None,
                        "p95_ms": int(r[3]) if r[3] is not None else None,
                        "p99_ms": int(r[4]) if r[4] is not None else None,
                    })
                else:
                    series.append({
                        "ts": key,
                        "count": 0,
                        "p50_ms": None,
                        "p95_ms": None,
                        "p99_ms": None,
                    })

            # Last-hour = último bucket con data (puede no ser la hora
            # actual si aún no hubo queries esta hora).
            last_with_data = next(
                (s for s in reversed(series) if s["count"] > 0), None,
            )
            if last_with_data is not None:
                summary["p50_1h_ms"] = last_with_data["p50_ms"]
                summary["p95_1h_ms"] = last_with_data["p95_ms"]
            summary["count_24h"] = sum(s["count"] for s in series)

            # Baseline 7d (un solo percentile global, no por hora) — sirve
            # como "typical p95" contra el que comparamos la última hora.
            baseline = conn.execute(
                """
                WITH base AS (
                    SELECT CAST(json_extract(extra_json, '$.total_ms') AS INTEGER) AS v
                    FROM rag_queries
                    WHERE cmd LIKE 'web%'
                      AND ts >= datetime('now', '-7 days')
                      AND json_extract(extra_json, '$.total_ms') IS NOT NULL
                ),
                ranked AS (
                    SELECT v,
                        ROW_NUMBER() OVER (ORDER BY v) AS rn,
                        COUNT(*) OVER () AS n
                    FROM base
                )
                SELECT
                    MIN(CASE WHEN rn * 1.0 / n >= 0.5  THEN v END) AS p50,
                    MIN(CASE WHEN rn * 1.0 / n >= 0.95 THEN v END) AS p95
                FROM ranked
                """,
            ).fetchone()
            if baseline is not None:
                summary["p50_baseline_ms"] = int(baseline[0]) if baseline[0] is not None else None
                summary["p95_baseline_ms"] = int(baseline[1]) if baseline[1] is not None else None
                if summary["p95_1h_ms"] and summary["p95_baseline_ms"]:
                    delta = (
                        (summary["p95_1h_ms"] - summary["p95_baseline_ms"])
                        / summary["p95_baseline_ms"]
                        * 100.0
                    )
                    summary["delta_p95_pct"] = round(delta, 1)
    except Exception as e:
        # Telemetry DB locked / schema missing: devolver payload neutral
        # en vez de 500ear el /status. El frontend muestra "—" y el
        # semáforo global sigue funcionando con el resto de sus probes.
        print(f"[status_latency] warn: build failed: {type(e).__name__}: {e}", flush=True)

    return {
        "window_hours": 24,
        "bucket": "hour",
        "series": series,
        "summary": summary,
    }


@app.get("/api/status/latency")
def status_latency(nocache: int = 0) -> dict:
    """Hourly p50/p95/p99 series of /api/chat total_ms over the last 24h
    plus a 7d baseline for delta computation. Drives the latency sparkline
    on /status. Cached 60s; `?nocache=1` forces refresh."""
    now = time.monotonic()
    with _LATENCY_CACHE_LOCK:
        cached = _LATENCY_CACHE["payload"]
        fresh = (now - _LATENCY_CACHE["ts"]) < _LATENCY_CACHE_TTL
    if nocache or not fresh or cached is None:
        payload = _status_latency_build_payload()
        with _LATENCY_CACHE_LOCK:
            _LATENCY_CACHE["payload"] = payload
            _LATENCY_CACHE["ts"] = now
        return payload
    return cached


# ── /api/status/errors — error-budget para el donut del /status ──────
# Feed el card #2 del insights grid. Lee los 2 jsonl de errores que el
# rag.py ya mantiene:
#
#   ~/.local/share/obsidian-rag/silent_errors.jsonl
#     — una línea por `except Exception: pass` migrado a `_silent_log`
#       (contradict parse, reranker unload, OAuth refresh, wiki ingest,
#       semantic cache, etc.). Schema: {ts, where, exc_type, exc}.
#
#   ~/.local/share/obsidian-rag/sql_state_errors.jsonl
#     — una línea por fail de SQL write/read post-retry-budget (1756 en
#       6 días en el audit de 2026-04-24, mayoritariamente `database is
#       locked` por contention). Schema: {ts, event, err}.
#
# El rollup lógico vive también en rag.py/rag_stats (`_rollup`), pero
# ahí está enterrado dentro de una función grande. Reimplementamos
# acá en ~15 líneas para no forzar un refactor en rag.py.
#
# Presupuesto: scan lineal de ambos files (~50-200 MB combinados en
# prod típico). Cacheado 30s — los errores son "eventos", no continuos,
# y 30s de latencia en la dashboard está bien. La lectura usa un cutoff
# por `ts` ISO parsing; líneas malformed se skipean tolerantes.
#
# Schema del payload:
#   {
#     "window_hours": 24,
#     "total_errors": 458,
#     "by_source": {"silent": 120, "sql": 338},
#     "breakdown": [                    # top-N con count desc
#       {"key": "contradictions_sql_read_failed", "count": 289, "source": "sql"},
#       ...
#     ],
#     "total_errors_prev_24h": 312,      # 24-48h atrás — para delta
#     "delta_pct": 46.8,                  # vs prev_24h
#   }

_ERRORS_CACHE_TTL = 30.0
_ERRORS_CACHE: dict = {"ts": 0.0, "payload": None}
_ERRORS_CACHE_LOCK = threading.Lock()

# Cuántos `where/event` distintos mostramos en la breakdown antes de
# colapsar el resto en `other`. 6 es el sweet spot — entra en la card
# sin scroll y cubre el 90% de los errores típicos según el audit.
_ERRORS_BREAKDOWN_TOP_N = 6


def _rollup_error_log(path: Path, key_field: str, source: str,
                      cutoff: datetime, upper: datetime | None = None,
                      ) -> tuple[int, dict[str, int]]:
    """Read a JSONL error log (silent_errors / sql_state_errors) and
    roll up by `key_field` for entries whose `ts` is in [cutoff, upper).

    Returns (total_count, {key: count, ...}).
    Tolerant: missing ts / malformed JSON / missing key_field are skipped
    silently — these are best-effort diagnostic logs, not auditable.
    """
    total = 0
    counts: dict[str, int] = {}
    if not path.is_file():
        return total, counts
    try:
        with path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = json.loads(line)
                except Exception:
                    continue
                ts = rec.get("ts")
                if not ts:
                    continue
                try:
                    rec_dt = datetime.fromisoformat(ts.rstrip("Z"))
                except Exception:
                    continue
                if rec_dt < cutoff:
                    continue
                if upper is not None and rec_dt >= upper:
                    continue
                key = rec.get(key_field) or "(unknown)"
                counts[key] = counts.get(key, 0) + 1
                total += 1
    except Exception as e:
        print(f"[status_errors] warn: rollup of {path.name} failed: {type(e).__name__}: {e}", flush=True)
    return total, counts


def _status_errors_build_payload() -> dict:
    """Build the error-budget payload for /api/status/errors.

    Two passes per log file: current 24h window + prev 24h window (for
    the delta vs yesterday). The prev window is best-effort — if the
    files were rotated recently, we return 0 and the frontend shows
    `sin baseline` in the delta slot.
    """
    now = datetime.now()
    cutoff_24h = now - timedelta(hours=24)
    cutoff_48h = now - timedelta(hours=48)

    payload: dict = {
        "window_hours": 24,
        "total_errors": 0,
        "by_source": {"silent": 0, "sql": 0},
        "breakdown": [],
        "total_errors_prev_24h": 0,
        "delta_pct": None,
    }

    try:
        # Current window: last 24h.
        silent_total, silent_counts = _rollup_error_log(
            SILENT_ERRORS_LOG_PATH, "where", "silent", cutoff_24h,
        )
        sql_total, sql_counts = _rollup_error_log(
            _SQL_STATE_ERROR_LOG, "event", "sql", cutoff_24h,
        )

        # Previous window: 24-48h ago. Mismo rollup con ventana desplazada.
        silent_prev, _ = _rollup_error_log(
            SILENT_ERRORS_LOG_PATH, "where", "silent", cutoff_48h, cutoff_24h,
        )
        sql_prev, _ = _rollup_error_log(
            _SQL_STATE_ERROR_LOG, "event", "sql", cutoff_48h, cutoff_24h,
        )

        payload["total_errors"] = silent_total + sql_total
        payload["by_source"] = {"silent": silent_total, "sql": sql_total}
        payload["total_errors_prev_24h"] = silent_prev + sql_prev

        # Merge both dicts into a single (key, count, source) list, sort
        # by count desc, slice top-N, collapse the rest into `other`.
        merged: list[tuple[str, int, str]] = []
        for key, count in silent_counts.items():
            merged.append((key, count, "silent"))
        for key, count in sql_counts.items():
            merged.append((key, count, "sql"))
        merged.sort(key=lambda t: t[1], reverse=True)

        top = merged[:_ERRORS_BREAKDOWN_TOP_N]
        rest = merged[_ERRORS_BREAKDOWN_TOP_N:]
        breakdown: list[dict] = [
            {"key": k, "count": c, "source": s} for (k, c, s) in top
        ]
        if rest:
            breakdown.append({
                "key": "(other)",
                "count": sum(c for _k, c, _s in rest),
                "source": "mixed",
            })
        payload["breakdown"] = breakdown

        # Delta vs prev 24h. Guard contra división por cero cuando prev=0
        # (nuevo sistema o log rotado): reportamos None en vez de inf.
        if payload["total_errors_prev_24h"] > 0:
            delta = (
                (payload["total_errors"] - payload["total_errors_prev_24h"])
                / payload["total_errors_prev_24h"]
                * 100.0
            )
            payload["delta_pct"] = round(delta, 1)
    except Exception as e:
        print(f"[status_errors] warn: build failed: {type(e).__name__}: {e}", flush=True)

    return payload


@app.get("/api/status/errors")
def status_errors(nocache: int = 0) -> dict:
    """Error-budget de las últimas 24h: total + breakdown top-N por
    `where`/`event` + delta vs las 24h previas. Drives el donut card
    del /status. Cacheado 30s; ?nocache=1 fuerza refresh.

    Fuentes: silent_errors.jsonl (swallowed exceptions) + sql_state_
    errors.jsonl (SQL write/read fails post-retry-budget). No cubre
    /api/chat error rate directo — esas fallas típicamente no logean
    una row en rag_queries, así que no tenemos denominador limpio de
    "requests". Mostramos counts absolutos + delta vs ayer, que es
    honesto y alineado con cómo vive la telemetría hoy.
    """
    now = time.monotonic()
    with _ERRORS_CACHE_LOCK:
        cached = _ERRORS_CACHE["payload"]
        fresh = (now - _ERRORS_CACHE["ts"]) < _ERRORS_CACHE_TTL
    if nocache or not fresh or cached is None:
        payload = _status_errors_build_payload()
        with _ERRORS_CACHE_LOCK:
            _ERRORS_CACHE["payload"] = payload
            _ERRORS_CACHE["ts"] = now
        return payload
    return cached


# ── /api/status/freshness — cuándo corrió por última vez cada fuente ──
# Feed el card #3 del insights grid. Los ingestores dispersos del sistema
# tienen "last-run" en distintos lugares (plist con StartInterval, log
# mtime, sqlite, ...). Consolidamos acá en un shape uniforme para que
# la UI muestre la tabla con "hace Xm" + chip de drift vs SLA.
#
# Estrategia por fuente: el mtime del stdout log de su launchd job es
# el indicador más simple + universal. Cada ingestor escribe al menos
# un log line cuando corre (header de timestamps, payload count, etc.),
# así que mtime del log === last-run. Evita tocar los ingestores para
# agregar un "marker file" y funciona hoy sin migración.
#
# El SLA (cadence expected) se lee del `StartInterval` en el plist.
# Para vault no hay StartInterval (es un watch daemon continuous), así
# que hardcodeamos 900s como target de freshness — si en 15min no se
# movió un archivo del vault, el watch está zombie o el vault quieto
# (ambos ameritan el "warn" amarillo).
#
# Schema del payload:
#   {
#     "window_hours": 24,
#     "sources": [
#       {
#         "id": "vault", "label": "vault",
#         "last_run_ts": "2026-04-24T20:19:04",  # ISO local (no UTC)
#         "age_seconds": 240,
#         "sla_seconds": 900,
#         "drift_ratio": 0.27,                   # age/sla
#         "status": "ok",                         # ok | warn | stale | unknown
#         "detail": "watch daemon · hace 4m"
#       }, ...
#     ],
#     "sources_healthy": 5,
#     "sources_total": 6,
#   }
#
# Thresholds:
#   drift < 1.0         → ok (dentro del SLA)
#   drift 1.0 .. 3.0    → warn (atrasado pero tolerable)
#   drift ≥ 3.0         → stale (el ingestor probablemente se wedgeó)
#   file missing        → unknown (no corrió nunca o el log fue borrado)

_FRESHNESS_CACHE_TTL = 30.0
_FRESHNESS_CACHE: dict = {"ts": 0.0, "payload": None}
_FRESHNESS_CACHE_LOCK = threading.Lock()

# Catálogo de fuentes. El orden acá define el orden en la UI (stable,
# lo que permite al user aprender la tabla "de memoria"). `sla_seconds`
# para vault es el único hardcoded (no tiene StartInterval porque es
# watcher continuous); el resto se lee del plist.
_FRESHNESS_SOURCES: tuple[dict, ...] = (
    {
        "id": "vault",
        "label": "vault",
        "log_name": "watch.log",
        "launchd_label": "com.fer.obsidian-rag-watch",
        "sla_fallback_s": 900,  # 15min — si el watcher no indexa nada en
                                # 15m, o está zombie o el vault está
                                # idle (ambos vale marcar warn).
        "kind": "continuous",
    },
    {
        "id": "whatsapp",
        "label": "whatsapp",
        "log_name": "ingest-whatsapp.log",
        "launchd_label": "com.fer.obsidian-rag-ingest-whatsapp",
        "sla_fallback_s": 900,
        "kind": "scheduled",
    },
    {
        "id": "gmail",
        "label": "gmail",
        "log_name": "ingest-gmail.log",
        "launchd_label": "com.fer.obsidian-rag-ingest-gmail",
        "sla_fallback_s": 3600,
        "kind": "scheduled",
    },
    {
        "id": "calendar",
        "label": "calendar",
        "log_name": "ingest-calendar.log",
        "launchd_label": "com.fer.obsidian-rag-ingest-calendar",
        "sla_fallback_s": 3600,
        "kind": "scheduled",
    },
    {
        "id": "reminders",
        "label": "reminders",
        "log_name": "ingest-reminders.log",
        "launchd_label": "com.fer.obsidian-rag-ingest-reminders",
        "sla_fallback_s": 3600,
        "kind": "scheduled",
    },
    {
        "id": "drive",
        "label": "drive",
        "log_name": "ingest-drive.log",
        "launchd_label": "com.fer.obsidian-rag-ingest-drive",
        "sla_fallback_s": 3600,
        "kind": "scheduled",
    },
)


def _read_start_interval_s(plist_label: str) -> int | None:
    """Extract `StartInterval` (seconds) from a user launchd plist. Returns
    None if the plist doesn't exist or doesn't declare StartInterval (some
    jobs use StartCalendarInterval or are continuous watchers — those get
    handled via `sla_fallback_s` in the catalog)."""
    plist_path = Path.home() / "Library" / "LaunchAgents" / f"{plist_label}.plist"
    if not plist_path.is_file():
        return None
    try:
        content = plist_path.read_text(encoding="utf-8")
    except Exception:
        return None
    # Simple regex — no need for plistlib here; `<key>StartInterval</key>
    # <integer>NNN</integer>` es stable-enough en plists generados por
    # launchd / plutil. Si alguien edita a mano con formato raro, cae
    # al fallback.
    import re
    m = re.search(
        r"<key>\s*StartInterval\s*</key>\s*<integer>\s*(\d+)\s*</integer>",
        content,
    )
    if m:
        try:
            return int(m.group(1))
        except ValueError:
            return None
    return None


def _fmt_age_spanish(age_s: float) -> str:
    """'hace 4m', 'hace 2h', 'hace 3d' — copy rioplatense consistente."""
    if age_s < 0:
        return "justo ahora"
    if age_s < 60:
        return f"hace {int(age_s)}s"
    if age_s < 3600:
        return f"hace {int(age_s / 60)}m"
    if age_s < 86400:
        return f"hace {int(age_s / 3600)}h"
    return f"hace {int(age_s / 86400)}d"


def _status_freshness_build_payload() -> dict:
    """Build the freshness matrix payload. One row per _FRESHNESS_SOURCES
    entry. Errors per-row are logged but don't fail the whole endpoint —
    una fuente caída no debería impedir que las otras 5 se muestren."""
    log_dir = Path.home() / ".local" / "share" / "obsidian-rag"
    now = time.time()
    rows: list[dict] = []
    healthy = 0

    for src in _FRESHNESS_SOURCES:
        row: dict = {
            "id": src["id"],
            "label": src["label"],
            "kind": src["kind"],
            "last_run_ts": None,
            "age_seconds": None,
            "sla_seconds": None,
            "drift_ratio": None,
            "status": "unknown",
            "detail": "sin log",
        }
        try:
            # SLA: plist StartInterval > catalog fallback.
            sla = _read_start_interval_s(src["launchd_label"]) or src["sla_fallback_s"]
            row["sla_seconds"] = int(sla)

            log_path = log_dir / src["log_name"]
            if not log_path.is_file():
                # Log no existe todavía (ingestor nunca corrió). Dejamos
                # status=unknown; la UI lo muestra en gris.
                row["detail"] = "nunca corrió"
                rows.append(row)
                continue

            stat = log_path.stat()
            age_s = now - stat.st_mtime
            row["last_run_ts"] = datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds")
            row["age_seconds"] = int(age_s)
            row["drift_ratio"] = round(age_s / sla, 2) if sla > 0 else None

            # Thresholds: <1× ok, 1-3× warn, ≥3× stale.
            if sla > 0:
                r = age_s / sla
                if r < 1.0:
                    row["status"] = "ok"
                    healthy += 1
                elif r < 3.0:
                    row["status"] = "warn"
                else:
                    row["status"] = "stale"
            else:
                row["status"] = "unknown"

            # Detalle legible: "hace Xm / SLA Ym"
            sla_str = _fmt_age_spanish(sla).replace("hace ", "")
            row["detail"] = f"{_fmt_age_spanish(age_s)} / SLA {sla_str}"
        except Exception as e:
            row["status"] = "unknown"
            row["detail"] = f"error: {type(e).__name__}"
            print(f"[status_freshness] warn: source {src['id']!r} failed: {e}", flush=True)
        rows.append(row)

    return {
        "window_hours": 24,
        "sources": rows,
        "sources_healthy": healthy,
        "sources_total": len(_FRESHNESS_SOURCES),
    }


@app.get("/api/status/freshness")
def status_freshness(nocache: int = 0) -> dict:
    """Freshness matrix: por fuente (vault, whatsapp, gmail, calendar,
    reminders, drive), cuándo fue la última vez que corrió y si está
    dentro del SLA. Drives el card #3 del /status. Cacheado 30s;
    ?nocache=1 fuerza refresh.

    Mide "last run" via mtime del stdout log del launchd job. Es un
    proxy — asume que cada ingestor escribe al log cuando corre (todos
    los actuales lo hacen, al menos un timestamp header). Si alguno
    deja de hacerlo, hay que migrarlo a un marker file explícito.
    """
    now_mono = time.monotonic()
    with _FRESHNESS_CACHE_LOCK:
        cached = _FRESHNESS_CACHE["payload"]
        fresh = (now_mono - _FRESHNESS_CACHE["ts"]) < _FRESHNESS_CACHE_TTL
    if nocache or not fresh or cached is None:
        payload = _status_freshness_build_payload()
        with _FRESHNESS_CACHE_LOCK:
            _FRESHNESS_CACHE["payload"] = payload
            _FRESHNESS_CACHE["ts"] = now_mono
        return payload
    return cached


# ── /api/status/logs — eventos recientes WARN/ERROR ──────────────────
# Feed el card #4 del insights grid. Lee individual events (no rollup)
# de los mismos jsonl que el endpoint /errors:
#
#   silent_errors.jsonl  → level=WARN  (swallowed exceptions, no user-
#                          facing impact pero útiles para diagnóstico)
#   sql_state_errors.jsonl → level=ERROR (post-retry-budget, persistent)
#
# Los `.log` plain de los daemons (ingest-*.log, watch.log, etc.) NO
# se consumen acá — son texto semi-estructurado, dump de stdout, sin
# timestamp por línea. Parsearlos requeriría regex frágiles + heurísticas
# de severity. Mejor approach futuro: que los daemons escriban a un
# `runtime_events.jsonl` estructurado. Hasta entonces, el card cubre
# el ~95% de la señal real (los 2 jsonl actuales tienen 1000+ entries
# por día en prod).
#
# Schema del payload:
#   {
#     "window_seconds": 3600,
#     "limit": 50,
#     "level_filter": "all",   # "all" | "warn" | "error"
#     "events": [              # ordenados desc por ts
#       {
#         "ts": "2026-04-24T22:50:00",
#         "ts_age_s": 234,
#         "level": "ERROR",     # WARN | ERROR
#         "source": "sql",      # silent | sql
#         "where": "queries_sql_write_failed",
#         "exc_type": "OperationalError",
#         "message": "database is locked"
#       }, ...
#     ],
#     "total_in_window": 458,   # antes del cap por limit
#     "truncated": true,         # si total > limit
#   }

_LOGS_CACHE_TTL = 15.0  # más corto que errors (30s) porque queremos
                          # ver eventos nuevos rápido — log-tail debería
                          # sentirse "live"
_LOGS_CACHE: dict = {}    # cached por (window, limit, level) tuple
_LOGS_CACHE_LOCK = threading.Lock()

# Allowed values para los query params; sirven también para el contract
# tests + el frontend pickea de acá.
_LOGS_VALID_LEVELS = {"all", "warn", "error"}
_LOGS_DEFAULT_WINDOW_S = 3600
_LOGS_DEFAULT_LIMIT = 50
_LOGS_MAX_LIMIT = 200    # hard cap para no devolver MB de jsonl
_LOGS_MAX_WINDOW_S = 86400  # 24h cap


def _read_jsonl_events(path: Path, key_field: str, level: str,
                       cutoff: datetime) -> list[dict]:
    """Read a JSONL log file and return raw events with `ts >= cutoff`,
    each as a normalized dict ready for merging.

    Records that fail to parse / lack ts / are out-of-window are skipped.
    Field name `key_field` differs per log (silent → 'where', sql →
    'event'); both get normalized to `where` in the output for the
    frontend.
    """
    events: list[dict] = []
    if not path.is_file():
        return events
    try:
        with path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = json.loads(line)
                except Exception:
                    continue
                ts = rec.get("ts")
                if not ts:
                    continue
                try:
                    rec_dt = datetime.fromisoformat(ts.rstrip("Z"))
                except Exception:
                    continue
                if rec_dt < cutoff:
                    continue
                where = rec.get(key_field) or "(unknown)"
                # Mensaje + exc_type vienen distintos en cada log:
                #   silent: {exc_type, exc}
                #   sql:    {err: "RuntimeError('...')"} → solo string
                if "exc_type" in rec:
                    exc_type = rec.get("exc_type") or "(unknown)"
                    message = rec.get("exc") or ""
                else:
                    # SQL: el err es un repr del exception. Intentamos
                    # extraer el type prefix ("OperationalError(...)")
                    # con una regex simple. Si falla, exc_type queda
                    # vacío y el frontend muestra solo el mensaje.
                    err_str = str(rec.get("err") or "")
                    import re
                    m = re.match(r"^([A-Z][A-Za-z0-9_]*)\((.*)\)$", err_str, re.DOTALL)
                    if m:
                        exc_type = m.group(1)
                        message = m.group(2).strip("'\"")
                    else:
                        exc_type = ""
                        message = err_str
                events.append({
                    "ts": ts,
                    "ts_dt": rec_dt,
                    "level": level,
                    "source": "silent" if key_field == "where" else "sql",
                    "where": where,
                    "exc_type": exc_type,
                    "message": message[:500],  # truncate to keep payload
                })
    except Exception as e:
        print(f"[status_logs] warn: read of {path.name} failed: {type(e).__name__}: {e}", flush=True)
    return events


def _status_logs_build_payload(
    window_s: int, limit: int, level_filter: str,
) -> dict:
    """Build the recent-events payload for /api/status/logs.

    Reads silent + sql jsonl, filters by ts within `window_s`, optionally
    by level (all|warn|error), sorts desc by ts, caps at `limit`.
    """
    now = datetime.now()
    cutoff = now - timedelta(seconds=window_s)
    events: list[dict] = []

    if level_filter in ("all", "warn"):
        events.extend(_read_jsonl_events(
            SILENT_ERRORS_LOG_PATH, "where", "WARN", cutoff,
        ))
    if level_filter in ("all", "error"):
        events.extend(_read_jsonl_events(
            _SQL_STATE_ERROR_LOG, "event", "ERROR", cutoff,
        ))

    # Sort desc por ts (más recientes primero).
    events.sort(key=lambda e: e["ts_dt"], reverse=True)
    total = len(events)
    truncated = total > limit
    if truncated:
        events = events[:limit]

    # Convertir ts_dt a ts_age_s (segundos relativos a ahora) para que
    # el frontend pueda formatear "hace Xm" sin parsear fechas. Sacamos
    # el ts_dt del payload (no es serializable y el ts string ya está).
    out_events = []
    now_ts = now.timestamp()
    for e in events:
        age_s = max(0, int(now_ts - e["ts_dt"].timestamp()))
        out_events.append({
            "ts": e["ts"],
            "ts_age_s": age_s,
            "level": e["level"],
            "source": e["source"],
            "where": e["where"],
            "exc_type": e["exc_type"],
            "message": e["message"],
        })

    return {
        "window_seconds": window_s,
        "limit": limit,
        "level_filter": level_filter,
        "events": out_events,
        "total_in_window": total,
        "truncated": truncated,
    }


@app.get("/api/status/logs")
def status_logs(
    since_seconds: int = _LOGS_DEFAULT_WINDOW_S,
    limit: int = _LOGS_DEFAULT_LIMIT,
    level: str = "all",
    nocache: int = 0,
) -> dict:
    """Feed del card #4 (log-tail). Eventos individuales WARN/ERROR de
    silent_errors.jsonl + sql_state_errors.jsonl, ordenados desc por
    timestamp, cap por `limit`.

    Query params:
      since_seconds: ventana de tiempo (default 1h, max 24h)
      limit: cuántos eventos devolver (default 50, max 200)
      level: "all" | "warn" | "error" (filtra por sink)
      nocache: 1 fuerza refresh (default usa el cache de 15s)

    Cache key incluye los 3 query params — cada combinación tiene su
    propio TTL bucket. Sin esto, alternar filtros en el UI siempre
    devolvería el primer query cacheado.
    """
    # Validate + clamp.
    if level not in _LOGS_VALID_LEVELS:
        raise HTTPException(
            status_code=400,
            detail=f"level inválido: {level!r} (esperaba {sorted(_LOGS_VALID_LEVELS)})",
        )
    window_s = max(1, min(int(since_seconds), _LOGS_MAX_WINDOW_S))
    lim = max(1, min(int(limit), _LOGS_MAX_LIMIT))

    cache_key = (window_s, lim, level)
    now_mono = time.monotonic()
    with _LOGS_CACHE_LOCK:
        entry = _LOGS_CACHE.get(cache_key)
        fresh = entry is not None and (now_mono - entry["ts"]) < _LOGS_CACHE_TTL
    if nocache or not fresh:
        payload = _status_logs_build_payload(window_s, lim, level)
        with _LOGS_CACHE_LOCK:
            _LOGS_CACHE[cache_key] = {"payload": payload, "ts": now_mono}
        return payload
    return entry["payload"]


# ── /api/status/uptime — heatmap 7d × 24h por servicio core ──────────
# Feed el card #5 del insights grid. Este es el único de los 5 cards
# que requiere persistencia histórica — los anteriores se computan
# on-demand de fuentes existentes (rag_queries para latency, jsonl
# logs para errors+log-tail, file mtimes para freshness).
#
# Strategy: hookear `_persist_status_samples` al final de cada
# `_status_build_payload`. Como /api/status se cachea 3s + se llama
# cada 10s desde el frontend cuando hay un user mirando, samplear
# cada call sería ~6 inserts/min × 5 services = 30/min = 43k/día.
# Rate-limit a 60s mantiene ~5 inserts/min × 5 services = 7.2k/día,
# que × 7d = 50k rows total. Manejable para un sqlite con índice
# (service_id, ts).
#
# Cuando NO hay user mirando /status, no samplea — es la limitación
# aceptada de v1. Pro version: cron-driven probe cada 60s sin importar
# si hay user. TODO si el heatmap muestra muchos huecos en horas que
# el user no estaba activo.
#
# Initial state: tabla vacía → heatmap todo gris ("sin datos"). Después
# de 1h de uso, top-right corner se llena. Después de 7d, full.
#
# Schema del payload /api/status/uptime:
#   {
#     "window_days": 7,
#     "services": [
#       {
#         "id": "web-self",
#         "label": "Web daemon",
#         "uptime_pct_7d": 99.4,
#         "buckets": [
#           # 168 buckets ordenados de viejo a nuevo, fila por día
#           # × hora. {"ts": "2026-04-18T00", "uptime_pct": 100.0,
#           # "samples": 5} o null si no hay samples.
#           ...
#         ]
#       }, ...
#     ]
#   }

# Servicios core que rastreamos en el heatmap. Hardcoded porque queremos
# UI estable (no surgir/desaparecer filas según qué servicios estén en el
# catálogo en un momento dado).
_UPTIME_TRACKED_SERVICES: tuple[tuple[str, str], ...] = (
    ("web-self", "web"),
    ("ollama", "ollama"),
    ("rag-db", "rag-db"),
    ("telemetry-db", "telemetry"),
    ("vault", "vault"),
)
_UPTIME_TRACKED_IDS = frozenset(s[0] for s in _UPTIME_TRACKED_SERVICES)

_STATUS_SAMPLE_RATE_LIMIT_S = 60.0
_STATUS_SAMPLE_LAST_TS = {"ts": 0.0}  # monotonic time of last sample write
_STATUS_SAMPLE_LOCK = threading.Lock()

# Pruning del rag_status_samples — corremos un DELETE de filas viejas
# de vez en cuando para mantener la tabla acotada. No agregamos un cron
# especial; cada N samples chequeamos. 7d cap para el query + 1d de
# safety buffer = 8d.
_UPTIME_PRUNE_AFTER_S = 8 * 86400  # 8 días
_UPTIME_PRUNE_EVERY_N_SAMPLES = 200  # ~prune cada 4-6h en uso normal


def _persist_status_samples(payload: dict) -> None:
    """Insert one row per `_UPTIME_TRACKED_SERVICES` con su current status.

    Rate-limited a `_STATUS_SAMPLE_RATE_LIMIT_S` (1 minute) globalmente —
    aunque /api/status se llame cada 10s, sólo escribimos cada 60s. El
    rate-limit usa monotonic time, no ts ISO, para evitar problemas de
    clock skew.

    Best-effort: cualquier excepción se traga con un print al stderr.
    Una falla de SQL acá no debería bajar el endpoint /api/status que
    es lo que está sirviendo al user en este momento.
    """
    now = time.monotonic()
    with _STATUS_SAMPLE_LOCK:
        if (now - _STATUS_SAMPLE_LAST_TS["ts"]) < _STATUS_SAMPLE_RATE_LIMIT_S:
            return
        _STATUS_SAMPLE_LAST_TS["ts"] = now

    ts_iso = datetime.now().isoformat(timespec="seconds")
    samples: list[tuple[str, str, str]] = []
    for cat in payload.get("categories", []):
        for svc in cat.get("services", []):
            sid = svc.get("id")
            if sid in _UPTIME_TRACKED_IDS:
                samples.append((ts_iso, sid, svc.get("status") or "unknown"))

    if not samples:
        return

    try:
        with _ragvec_state_conn() as conn:
            conn.executemany(
                "INSERT OR IGNORE INTO rag_status_samples "
                "(ts, service_id, status) VALUES (?, ?, ?)",
                samples,
            )
            # Pruning ocasional. Hacemos un COUNT cheap y, si crece más
            # allá del cap esperado, borramos lo viejo. 1-in-N
            # comprobación para que el cost amortice.
            import random
            if random.random() < (1.0 / _UPTIME_PRUNE_EVERY_N_SAMPLES):
                conn.execute(
                    "DELETE FROM rag_status_samples WHERE ts < datetime('now', '-8 days')"
                )
    except Exception as e:
        print(f"[status_samples] warn: persist failed: {type(e).__name__}: {e}", flush=True)


# ── Periodic background probe ────────────────────────────────────────
# Cron-driven sampling para que el heatmap se llene aunque nadie esté
# mirando /status. Single daemon thread inside FastAPI lifecycle: arranca
# en `_on_startup`, se apaga vía Event en `_on_shutdown`. Sleep de 60s
# coincide con `_STATUS_SAMPLE_RATE_LIMIT_S` así cada tick efectivamente
# inserta — sin doble-pisarse con el rate-limit cuando además un user
# está activo (el rate-limit gana, el thread queda no-op para esa tick).
#
# Por qué thread interno y no launchd cron separado:
#   - Cero archivos nuevos que mantener (plist, log paths, kickstart
#     handlers en /api/status/action).
#   - Sampling matchea "is the web server alive?" — si el server está
#     down, no hay heatmap que llenar tampoco. Es la semántica correcta
#     para un dashboard local-first single-user.
#   - El thread comparte el cache de `_status_build_payload` así si un
#     user hizo un request hace <3s, el thread reusa esa data en vez de
#     re-probar 25 servicios en paralelo.
#
# Trade-off aceptado: cuando el server está dormido (mac suspended), no
# samplea — el heatmap muestra gris para esas horas. Si en algún momento
# el user quiere "completitud aunque la mac esté apagada", la solución
# es un launchd job externo (cron) que persista igual desde otra máquina,
# o usar una probe externa (ej. UptimeRobot). Out of scope acá.

_STATUS_PROBE_PERIOD_S = 60.0
_STATUS_PROBE_THREAD: "threading.Thread | None" = None
_STATUS_PROBE_STOP = threading.Event()


def _status_periodic_probe_loop() -> None:
    """Daemon thread loop: cada `_STATUS_PROBE_PERIOD_S` segundos invoca
    `_status_build_payload + _persist_status_samples` para alimentar el
    heatmap del card #5.

    Usa el mismo `_status_build_payload` que `/api/status` así el cache
    queda warm para el próximo request del user. La persistencia es
    rate-limited globalmente (`_STATUS_SAMPLE_LOCK`), así que aunque
    coincida con un request del user en la misma ventana, sólo se
    inserta una vez.

    Best-effort: cualquier excepción se traga + sigue. El loop sale solo
    cuando `_STATUS_PROBE_STOP.is_set()` (vía `_on_shutdown`). Logging
    a stderr para que las fallas queden visibles en `web.error.log`
    sin spammar el log feliz path (sólo se imprime en error).
    """
    while not _STATUS_PROBE_STOP.is_set():
        try:
            payload = _status_build_payload()
            # Update el cache módulo-level así un user que llega justo
            # ahora consume la data fresca sin re-build (3s TTL del
            # `/api/status` cache).
            with _STATUS_CACHE_LOCK:
                _STATUS_CACHE["payload"] = payload
                _STATUS_CACHE["ts"] = time.monotonic()
            _persist_status_samples(payload)
        except Exception as exc:
            print(f"[status-probe] tick failed: {type(exc).__name__}: {exc}",
                  file=sys.stderr, flush=True)
        # Wait con cancelación temprana — `Event.wait(timeout)` retorna
        # True ni bien `set()` se llama, así que el shutdown no espera
        # los 60s completos.
        _STATUS_PROBE_STOP.wait(timeout=_STATUS_PROBE_PERIOD_S)


@_on_startup
def _start_status_probe_thread() -> None:
    """Arrancar el daemon thread del probe periódico al startup.

    `daemon=True` para que no bloquee la salida del proceso si por algún
    motivo el shutdown callback no llega (ej. SIGKILL). El thread normal
    sale via `_STATUS_PROBE_STOP.set()` desde `_stop_status_probe_thread`.
    """
    global _STATUS_PROBE_THREAD
    _STATUS_PROBE_STOP.clear()
    _STATUS_PROBE_THREAD = threading.Thread(
        target=_status_periodic_probe_loop,
        name="status-probe",
        daemon=True,
    )
    _STATUS_PROBE_THREAD.start()


@_on_shutdown
def _stop_status_probe_thread() -> None:
    """Señalar al probe thread que pare + esperarlo brevemente.

    Tope de 2s al join para no bloquear la shutdown sequence — si el
    thread está mid-build_payload (que puede tardar varios segundos por
    los probes paralelos), aceptamos cortarlo. Como es daemon, el
    proceso lo cierra de cualquier forma; el join es sólo cleanup
    cosmético para que no salga un warning de Python al exit.
    """
    _STATUS_PROBE_STOP.set()
    if _STATUS_PROBE_THREAD is not None:
        _STATUS_PROBE_THREAD.join(timeout=2.0)


# Cache para el endpoint /uptime — el query es relativamente caro
# (window function sobre 50k rows agrupadas por hora), TTL más largo
# que los otros cards porque los buckets de uptime cambian de a poco.
_UPTIME_CACHE_TTL = 90.0
_UPTIME_CACHE: dict = {"ts": 0.0, "payload": None}
_UPTIME_CACHE_LOCK = threading.Lock()


def _status_uptime_build_payload() -> dict:
    """Compute 7d × 24h heatmap por servicio. Para cada bucket horario,
    `uptime_pct = ok_count / total_count * 100`. Buckets sin samples
    quedan con `uptime_pct=null` y la UI los pinta gris transparente.

    Genera 168 buckets por servicio en orden cronológico (viejo → nuevo)
    para que el frontend pueda renderizar la grid sin re-ordenar.
    """
    now = datetime.now()
    bucket_count = 168
    services_out: list[dict] = []

    try:
        with _ragvec_state_conn() as conn:
            for service_id, label in _UPTIME_TRACKED_SERVICES:
                # Bucket por hora: count total + count(status='ok'). El
                # uptime_pct sale luego en Python para tener control sobre
                # los nulls. Filter por service_id para hit del covering
                # index.
                rows = conn.execute(
                    """
                    SELECT
                        strftime('%Y-%m-%dT%H', ts) AS bucket,
                        COUNT(*) AS total,
                        SUM(CASE WHEN status = 'ok' THEN 1 ELSE 0 END) AS ok_count
                    FROM rag_status_samples
                    WHERE service_id = ?
                      AND ts >= datetime('now', '-7 days')
                    GROUP BY bucket
                    """,
                    (service_id,),
                ).fetchall()
                by_bucket = {r[0]: (int(r[1] or 0), int(r[2] or 0)) for r in rows}

                # Generar 168 buckets en orden cronológico. Cada bucket es
                # el inicio de la hora (HH:00) en local time.
                buckets: list[dict] = []
                current = now.replace(minute=0, second=0, microsecond=0)
                ok_total = 0
                samples_total = 0
                for i in range(bucket_count - 1, -1, -1):
                    t = current - timedelta(hours=i)
                    key = t.strftime("%Y-%m-%dT%H")
                    bucket_data = by_bucket.get(key)
                    if bucket_data is not None and bucket_data[0] > 0:
                        total, ok_count = bucket_data
                        uptime = (ok_count / total) * 100.0
                        ok_total += ok_count
                        samples_total += total
                        buckets.append({
                            "ts": key + ":00:00",
                            "uptime_pct": round(uptime, 1),
                            "samples": total,
                        })
                    else:
                        buckets.append({
                            "ts": key + ":00:00",
                            "uptime_pct": None,
                            "samples": 0,
                        })

                uptime_7d = (ok_total / samples_total * 100.0) if samples_total > 0 else None
                services_out.append({
                    "id": service_id,
                    "label": label,
                    "uptime_pct_7d": round(uptime_7d, 2) if uptime_7d is not None else None,
                    "samples_7d": samples_total,
                    "buckets": buckets,
                })
    except Exception as e:
        print(f"[status_uptime] warn: build failed: {type(e).__name__}: {e}", flush=True)
        # Returns lista vacía — frontend muestra "sin datos" en heatmap.

    return {
        "window_days": 7,
        "bucket_hours": 1,
        "services": services_out,
    }


@app.get("/api/status/uptime")
def status_uptime(nocache: int = 0) -> dict:
    """7-día heatmap horario por servicio core. Lee rag_status_samples
    (poblado por _persist_status_samples cada 60s al hit /api/status).
    Cacheado 90s — los buckets horarios cambian lento. ?nocache=1 fuerza
    refresh.

    Cuando recién se instala (table vacía), todos los buckets vienen
    null y la UI muestra el heatmap gris con un tooltip "esperando
    datos". Después de unas horas de uso normal, los recientes se
    pueblan; después de 7d el cuadro está full.
    """
    now_mono = time.monotonic()
    with _UPTIME_CACHE_LOCK:
        cached = _UPTIME_CACHE["payload"]
        fresh = (now_mono - _UPTIME_CACHE["ts"]) < _UPTIME_CACHE_TTL
    if nocache or not fresh or cached is None:
        payload = _status_uptime_build_payload()
        with _UPTIME_CACHE_LOCK:
            _UPTIME_CACHE["payload"] = payload
            _UPTIME_CACHE["ts"] = now_mono
        return payload
    return cached


@app.get("/status")
def status_page() -> FileResponse:
    return FileResponse(STATIC_DIR / "status.html")


# ── /logs — dashboard de logs del sistema ─────────────────────────────
# Página + endpoints para inspeccionar todos los archivos de log que
# generan los daemons del stack (ingest-*, watch, web, wa-scheduled-send,
# anticipate, reminder-wa-push, etc.). A diferencia del card #4 del
# /status que sólo lee 2 jsonl estructurados, acá listamos los ~80
# archivos `.log` + `.error.log` que escriben los daemons launchd.
#
# Decisiones:
#   - Sources: ~/.local/share/obsidian-rag/ + ~/.local/share/whatsapp-listener/
#     (los 2 stacks compartidos del sistema). Todo lo que termine en .log
#     o .error.log entra; los .jsonl quedan fuera (formato distinto, ya
#     cubiertos por el card #4 de /status).
#   - Agrupamos por "service": el nombre sin .log / .error.log. Cada
#     service tiene un par (stdout, stderr) + status agregado: OK si no
#     hay errores recientes, WARN si hay warnings, ERROR si .error.log
#     tiene contenido reciente o líneas con patterns de error.
#   - Tail eficiente: para archivos grandes (web.log ~1.6MB) seekeamos
#     desde el final en chunks de 64KB hasta juntar N líneas. No leemos
#     el archivo entero a memoria.
#   - Heurística de level por línea: regex case-insensitive sobre
#     "error|exception|traceback|fatal|panic|critical" → error;
#     "warn|warning|deprec" → warn. Evitamos falsos positivos de
#     "failed=0" / "errors=0" (líneas de stats que NO son errores).
#
# Seguridad: el `name` del file viene de la lista enumerada del backend.
# El handler valida que el path resuelto está dentro de los allowed
# LOG_DIRS (defensa en profundidad contra path traversal).

# Directorios donde viven los logs de los daemons del stack. Si en algún
# futuro se agrega otro stack, sumarlo acá.
_LOG_DIRS: tuple[Path, ...] = (
    Path.home() / ".local/share/obsidian-rag",
    Path.home() / ".local/share/whatsapp-listener",
)

# Heurística para clasificar líneas. Sin word boundary INICIAL así
# matchea CamelCase tipo `OperationalError` o `UserWarning` (que NO
# tienen boundary entre `l-E` o `r-W`), pero con boundary final + un
# lookahead negativo `(?!s?[\s:=]+0\b)` que evita clasificar como error
# las stats "failed=0" / "errors: 0".
_LOG_RE_ERROR = re.compile(
    r"(error|exception|traceback|failed|fatal|panic|critical)"
    r"(?!s?[\s:=]+0\b)\b",
    re.IGNORECASE,
)
_LOG_RE_WARN = re.compile(r"(warning|warn|deprec|retry)\b", re.IGNORECASE)
# Heartbeat / status normales — overrideamos cualquier match accidental.
_LOG_RE_OK = re.compile(r"^\[heartbeat\]|alive=true|✓\s|status=ok\b", re.IGNORECASE)
# Líneas claramente informacionales (uvicorn / stdlib logging). Si una
# línea empieza con `INFO:` / `DEBUG:` etc., NO la clasificamos como
# error aunque mencione "error" en una URL/path. Bug histórico:
# `INFO:    "GET /api/logs/file?name=...error.log... 200 OK"` aparecía
# como ERROR sólo porque la URL contenía "error.log".
_LOG_RE_INFO_PREFIX = re.compile(r"^\s*(INFO|DEBUG|TRACE|NOTICE)[\s:]", re.IGNORECASE)

# Patrones de timestamp que aparecen en los logs del stack. En orden de
# prioridad — el primer match gana. Soportamos:
#   - JSONL `{"ts": "2026-04-26T19:51:39", ...}` (silent_errors,
#     sql_state_errors, collection_ops, behavior, etc.)
#   - ISO con T: `2026-04-26T19:47:50` (heartbeat, eval, varios daemons)
#   - ISO con espacio: `2026-04-26 17:05:22` (cloudflared-watcher,
#     algunos scripts shell)
# Capturamos el timestamp en formato ISO normalizado (sin Z ni offset).
_LOG_RE_TS_JSONL = re.compile(r'"ts"\s*:\s*"(\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2})')
_LOG_RE_TS_ISO_T = re.compile(r"\b(\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2})")
_LOG_RE_TS_ISO_SPACE = re.compile(r"^\s*(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})")


def _extract_log_ts(line: str) -> str | None:
    """Extraer timestamp ISO de una línea de log si está presente.

    Retorna el timestamp en formato `YYYY-MM-DDTHH:MM:SS` (T como
    separador, sin zona horaria), o None si la línea no tiene timestamp
    detectable.

    Cubre los formatos vistos en el stack al 2026-04-26: JSONL con
    `"ts": "..."`, ISO con T (heartbeat / silent_errors), e ISO con
    espacio (cloudflared-watcher).
    """
    if not line:
        return None
    m = _LOG_RE_TS_JSONL.search(line)
    if m:
        return m.group(1).replace(" ", "T")
    m = _LOG_RE_TS_ISO_T.search(line)
    if m:
        return m.group(1)
    m = _LOG_RE_TS_ISO_SPACE.match(line)
    if m:
        return m.group(1).replace(" ", "T")
    return None

# Cap conservador. La página puede pedir tail=N hasta este max para no
# devolver MB de log al cliente.
_LOG_FILE_TAIL_DEFAULT = 500
_LOG_FILE_TAIL_MAX = 5000

# Ventana para "errores recientes" en la lista de archivos. Si .error.log
# tiene mtime más viejo que esto, igual lo marcamos como con errores
# pero con menor severidad visual ("histórico"). Default: 24h.
_LOG_RECENT_ERROR_WINDOW_S = 86400

# Cache para /api/logs (la lista de archivos). El stat() de 80 files es
# rápido (~5ms), pero la heurística de "error_count_recent" lee la cola
# de varios files para detectar errores → más caro. Cache 10s mantiene
# la página snappy sin perder responsiveness real.
_LOGS_INDEX_CACHE: dict = {"ts": 0.0, "payload": None}
_LOGS_INDEX_CACHE_TTL = 10.0
_LOGS_INDEX_CACHE_LOCK = threading.Lock()


def _classify_log_line(line: str) -> str:
    """Devolver 'error' | 'warn' | 'ok' | 'info' para una línea de log."""
    if not line.strip():
        return "info"
    if _LOG_RE_OK.search(line):
        return "ok"
    # `INFO:`/`DEBUG:` prefix gana sobre los matches de "error" en URLs.
    if _LOG_RE_INFO_PREFIX.match(line):
        return "info"
    if _LOG_RE_ERROR.search(line):
        return "error"
    if _LOG_RE_WARN.search(line):
        return "warn"
    return "info"


def _read_tail_lines(path: Path, max_lines: int, max_bytes: int = 4 * 1024 * 1024) -> list[str]:
    """Leer las últimas `max_lines` líneas de `path` sin cargar el archivo
    entero a memoria. Seek desde el final en chunks; cap absoluto por
    `max_bytes` así no nos vamos al ratón con archivos gigantes.

    Retorna lista en orden cronológico (vieja → nueva).
    """
    if not path.is_file():
        return []
    try:
        with path.open("rb") as f:
            f.seek(0, 2)
            total = f.tell()
            if total == 0:
                return []
            chunk_size = 64 * 1024
            # Acumulamos chunks desde el final hasta tener max_lines + 1
            # newlines (uno extra para tolerar el caso de archivos que
            # terminan en `\n`, que crean una "línea vacía" final que
            # luego descartamos). El slice por max_lines lo hacemos al
            # final, una sola vez, para que la lógica sea simple +
            # correcta sin cortar de más.
            buf = b""
            pos = total
            read_total = 0
            target_newlines = max_lines + 1
            while pos > 0 and buf.count(b"\n") < target_newlines and read_total < max_bytes:
                read_size = min(chunk_size, pos)
                pos -= read_size
                f.seek(pos)
                chunk = f.read(read_size)
                read_total += read_size
                buf = chunk + buf
            # Splitear y decodear con replace (no morir con bytes
            # inválidos en tracebacks Python).
            lines = [b.decode("utf-8", errors="replace") for b in buf.split(b"\n")]
            # Archivos terminados en `\n` generan una línea vacía final
            # que es ruido — descartar.
            while lines and lines[-1] == "":
                lines.pop()
            # Slice a max_lines del final. Si el primer fragmento fue
            # leído parcialmente (pos > 0) descartamos esa primer línea
            # porque puede estar cortada por la mitad.
            if pos > 0 and len(lines) > 0:
                lines = lines[1:]
            if len(lines) > max_lines:
                lines = lines[-max_lines:]
            return lines
    except Exception:
        return []


def _classify_file_status(error_log_path: Path, recent_lines: list[str]) -> tuple[str, int]:
    """Decidir el status agregado de un service: 'ok' | 'warn' | 'error'.

    - error si .error.log existe + size > 0 + mtime reciente, o si las
      últimas N líneas del .log tienen pattern de error.
    - warn si las últimas N líneas tienen pattern de warn.
    - ok en otro caso.

    Devuelve (status, error_count_recent).
    """
    error_count = 0
    has_warn = False
    # Inspeccionar .error.log: si existe, tiene size > 0 y mtime
    # reciente, contar líneas con pattern.
    if error_log_path.is_file():
        try:
            stat = error_log_path.stat()
            if stat.st_size > 0:
                age_s = time.time() - stat.st_mtime
                if age_s < _LOG_RECENT_ERROR_WINDOW_S:
                    # Leer las últimas 200 líneas del error log y
                    # contar las que matchean error.
                    err_lines = _read_tail_lines(error_log_path, 200)
                    for ln in err_lines:
                        lvl = _classify_log_line(ln)
                        if lvl == "error":
                            error_count += 1
                        elif lvl == "warn":
                            has_warn = True
        except Exception:
            pass
    # También revisar las recent lines del stdout — algunos daemons
    # imprimen tracebacks ahí en vez de stderr.
    for ln in recent_lines:
        lvl = _classify_log_line(ln)
        if lvl == "error":
            error_count += 1
        elif lvl == "warn":
            has_warn = True
    if error_count > 0:
        return ("error", error_count)
    if has_warn:
        return ("warn", 0)
    return ("ok", 0)


def _human_size(n: int) -> str:
    """Format bytes to human-readable (B / KB / MB)."""
    if n < 1024:
        return f"{n} B"
    if n < 1024 * 1024:
        return f"{n / 1024:.1f} KB"
    return f"{n / (1024 * 1024):.1f} MB"


def _resolve_log_path(name: str) -> Path | None:
    """Resolve un log name (ej. 'watch.log' o 'whatsapp-listener/listener.log')
    al Path absoluto, validando que esté dentro de uno de los _LOG_DIRS.
    Retorna None si no existe o el path está fuera del allowlist.
    """
    if not name or ".." in name or name.startswith("/"):
        return None
    # El name viene como "<dir-slug>/<filename>" o sólo "<filename>"
    # (en cuyo caso asumimos el primer dir, obsidian-rag).
    if "/" in name:
        dir_slug, _, filename = name.partition("/")
        if "/" in filename or ".." in filename:
            return None
        # dir_slug → mapear a uno de _LOG_DIRS por el último componente
        # del path.
        for d in _LOG_DIRS:
            if d.name == dir_slug:
                candidate = (d / filename).resolve()
                try:
                    candidate.relative_to(d.resolve())
                except ValueError:
                    return None
                if candidate.is_file():
                    return candidate
                return None
        return None
    # Sin slash → buscar en _LOG_DIRS[0] (obsidian-rag).
    candidate = (_LOG_DIRS[0] / name).resolve()
    try:
        candidate.relative_to(_LOG_DIRS[0].resolve())
    except ValueError:
        return None
    if candidate.is_file():
        return candidate
    return None


def _build_logs_index_payload() -> dict:
    """Enumerar todos los .log + .error.log de los _LOG_DIRS, agrupar por
    service, y devolver con metadata + status.
    """
    services_by_key: dict[str, dict] = {}

    for log_dir in _LOG_DIRS:
        if not log_dir.is_dir():
            continue
        dir_slug = log_dir.name
        # Listar .log + .error.log. Ignoramos otros ext.
        for path in sorted(log_dir.iterdir()):
            if not path.is_file():
                continue
            name = path.name
            if not (name.endswith(".log") or name.endswith(".stdout.log") or name.endswith(".stderr.log")):
                continue
            # Derivar service key. "wa-scheduled-send.error.log" → "wa-scheduled-send"
            # "watch.log" → "watch". "cloudflared-watcher.stdout.log" → "cloudflared-watcher".
            base = name
            kind = "stdout"
            if base.endswith(".error.log"):
                base = base[: -len(".error.log")]
                kind = "stderr"
            elif base.endswith(".stderr.log"):
                base = base[: -len(".stderr.log")]
                kind = "stderr"
            elif base.endswith(".stdout.log"):
                base = base[: -len(".stdout.log")]
                kind = "stdout"
            elif base.endswith(".log"):
                base = base[: -len(".log")]
                kind = "stdout"
            key = f"{dir_slug}::{base}"
            try:
                stat = path.stat()
            except Exception:
                continue
            file_meta = {
                "name": name,
                "ref": f"{dir_slug}/{name}",
                "kind": kind,
                "size_bytes": stat.st_size,
                "size_human": _human_size(stat.st_size),
                "mtime": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
                "mtime_age_s": max(0, int(time.time() - stat.st_mtime)),
                "abs_path": str(path),
            }
            svc = services_by_key.setdefault(
                key,
                {
                    "service": base,
                    "dir": dir_slug,
                    "files": [],
                },
            )
            svc["files"].append(file_meta)

    # Para cada service: status agregado + preview.
    services_out: list[dict] = []
    for key in sorted(services_by_key.keys()):
        svc = services_by_key[key]
        # Buscar el stdout y el stderr.
        stdout_file = next((f for f in svc["files"] if f["kind"] == "stdout"), None)
        stderr_file = next((f for f in svc["files"] if f["kind"] == "stderr"), None)

        # Preview: tail de N líneas del stdout (o del stderr si no hay
        # stdout). Sólo última línea, para mostrar en el sidebar.
        primary = stdout_file or stderr_file
        preview = ""
        recent_lines: list[str] = []
        if primary:
            recent_lines = _read_tail_lines(Path(primary["abs_path"]), 100)
            for ln in reversed(recent_lines):
                if ln.strip():
                    preview = ln.strip()[:200]
                    break

        # Status agregado: error si el stderr tiene contenido reciente,
        # warn si las recent_lines del stdout matchean warn, ok si nada.
        stderr_path = Path(stderr_file["abs_path"]) if stderr_file else (
            (Path(primary["abs_path"]).parent / (svc["service"] + ".error.log"))
            if primary else None
        )
        if stderr_path is None:
            status, err_count = ("ok", 0)
        else:
            status, err_count = _classify_file_status(stderr_path, recent_lines)

        # mtime más reciente entre los files.
        mtime_max = max((f["mtime_age_s"] for f in svc["files"]), default=10**9)
        # El archivo más reciente determina el "primary mtime" para
        # ordenar la lista por actividad.
        mtime_min_age = min((f["mtime_age_s"] for f in svc["files"]), default=10**9)
        # ¿Está totalmente vacío? (todos los files tienen size 0)
        all_empty = all(f["size_bytes"] == 0 for f in svc["files"])

        services_out.append({
            "service": svc["service"],
            "dir": svc["dir"],
            "status": status,
            "error_count_recent": err_count,
            "files": svc["files"],
            "preview": preview,
            "mtime_age_s": mtime_min_age,
            "all_empty": all_empty,
        })

    # Sort: services con error primero, después warn, después ok; dentro
    # de cada bucket, los más activos (mtime_min_age menor) primero.
    status_order = {"error": 0, "warn": 1, "ok": 2}
    services_out.sort(
        key=lambda s: (status_order.get(s["status"], 3), s["mtime_age_s"], s["service"])
    )

    n_error = sum(1 for s in services_out if s["status"] == "error")
    n_warn = sum(1 for s in services_out if s["status"] == "warn")
    n_ok = sum(1 for s in services_out if s["status"] == "ok")

    return {
        "scanned_at": datetime.now().isoformat(timespec="seconds"),
        "services": services_out,
        "totals": {
            "services": len(services_out),
            "error": n_error,
            "warn": n_warn,
            "ok": n_ok,
        },
    }


@app.get("/api/logs")
def logs_index(nocache: int = 0) -> dict:
    """Lista de todos los archivos de log del stack, agrupados por service,
    con status agregado (ok/warn/error) y preview de la última línea.

    Cacheado `_LOGS_INDEX_CACHE_TTL` segundos. ?nocache=1 fuerza refresh.
    """
    now_mono = time.monotonic()
    with _LOGS_INDEX_CACHE_LOCK:
        cached = _LOGS_INDEX_CACHE["payload"]
        fresh = (now_mono - _LOGS_INDEX_CACHE["ts"]) < _LOGS_INDEX_CACHE_TTL
    if nocache or not fresh or cached is None:
        payload = _build_logs_index_payload()
        with _LOGS_INDEX_CACHE_LOCK:
            _LOGS_INDEX_CACHE["payload"] = payload
            _LOGS_INDEX_CACHE["ts"] = now_mono
        return payload
    return cached


@app.get("/api/logs/file")
def logs_file(
    name: str,
    tail: int = _LOG_FILE_TAIL_DEFAULT,
    q: str = "",
    only_errors: int = 0,
) -> dict:
    """Devolver el tail de un archivo de log específico, con cada línea
    clasificada (ok/warn/error/info) y opcionalmente filtrada por query.

    Args:
        name: ref del archivo, formato "<dir-slug>/<filename>" (ej.
            "obsidian-rag/watch.log"). El backend valida contra el
            allowlist de _LOG_DIRS.
        tail: cantidad de líneas a devolver (default 500, max 5000).
        q: substring filter case-insensitive. Si vacío, devuelve todo.
        only_errors: 1 → sólo líneas clasificadas como warn o error.
    """
    path = _resolve_log_path(name)
    if path is None:
        raise HTTPException(status_code=404, detail=f"log no encontrado: {name!r}")

    tail_n = max(1, min(int(tail), _LOG_FILE_TAIL_MAX))
    raw_lines = _read_tail_lines(path, tail_n)

    q_lower = (q or "").strip().lower()
    out_lines: list[dict] = []
    counts = {"ok": 0, "info": 0, "warn": 0, "error": 0}
    # Numeración relativa al final: la última línea es la más alta. No
    # tenemos line numbers absolutos sin scanear el archivo entero, así
    # que devolvemos el offset reverse desde el final.
    total = len(raw_lines)
    # Forward-fill de timestamp: si una línea no tiene ts pero la anterior
    # sí, la marcamos como `ts_inferred=true` heredando ese timestamp. Útil
    # para tracebacks multi-línea (la línea con `Traceback` tiene ts, las
    # frames siguientes no, y queremos agruparlas visualmente).
    last_ts: str | None = None
    for idx, ln in enumerate(raw_lines):
        level = _classify_log_line(ln)
        counts[level] = counts.get(level, 0) + 1
        ts = _extract_log_ts(ln)
        ts_inferred = False
        if ts:
            last_ts = ts
        elif last_ts is not None:
            ts = last_ts
            ts_inferred = True
        if q_lower and q_lower not in ln.lower():
            continue
        if only_errors and level not in ("warn", "error"):
            continue
        out_lines.append({
            "n": total - idx,  # 1 = última línea (más reciente)
            "text": ln,
            "level": level,
            "ts": ts,
            "ts_inferred": ts_inferred,
        })

    try:
        stat = path.stat()
        size_bytes = stat.st_size
        mtime = datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds")
    except Exception:
        size_bytes = 0
        mtime = None

    return {
        "name": name,
        "abs_path": str(path),
        "size_bytes": size_bytes,
        "size_human": _human_size(size_bytes),
        "mtime": mtime,
        "tail_requested": tail_n,
        "lines_total": total,
        "lines_returned": len(out_lines),
        "counts": counts,
        "lines": out_lines,
        "filtered_by_query": bool(q_lower),
        "filtered_by_level": bool(only_errors),
    }


# ── /api/logs/errors — feed global de errores ─────────────────────────
# Agrega TODAS las líneas con level=error de TODOS los services en una
# sola lista ordenada por timestamp desc. Útil para responder "¿qué
# falló en el sistema en la última hora?" sin tener que clickear service
# por service.
#
# Cap implícito: por archivo leemos hasta `_LOG_GLOBAL_TAIL_PER_FILE`
# líneas (default 300). En archivos chiquitos no importa; en `web.log`
# que tiene 1.6M, esto es ~10% del tail ⇒ dominado por la cantidad de
# .log + .error.log que iteramos (~80) × 300 = 24k líneas escaneadas
# total, ~30ms en SSD.
#
# Para los logs sin timestamps nativos, NO podemos ordenar globalmente
# (no sabemos cuándo ocurrió cada línea). En esos casos usamos el mtime
# del archivo como fallback — todas las líneas sin ts del file heredan
# `mtime - i*1s` donde i es la posición desde el final. Inferred timestamp
# se devuelve con el flag `ts_synthetic=true` así el frontend lo muestra
# en cursiva opacity baja.

_LOG_GLOBAL_TAIL_PER_FILE = 300
_LOG_GLOBAL_DEFAULT_WINDOW_S = 3600  # 1h
_LOG_GLOBAL_MAX_WINDOW_S = 7 * 86400  # 7 días
_LOG_GLOBAL_MAX_LINES = 1000

_LOG_GLOBAL_CACHE: dict = {}
_LOG_GLOBAL_CACHE_TTL = 8.0
_LOG_GLOBAL_CACHE_LOCK = threading.Lock()


def _build_global_errors_payload(window_s: int, level_filter: str) -> dict:
    """Iterar todos los .log + .error.log, extraer las líneas matching
    el level_filter ('error', 'warn_error', 'all'), mergear por ts desc.

    Para líneas sin ts: fallback a (mtime - offset_from_tail) — no es
    exacto, pero ordena bien dentro del mismo archivo y mantiene los
    archivos modificados recientemente arriba del feed. Marcamos esas
    líneas con `ts_synthetic=true` para que el frontend las distinga.
    """
    cutoff_ts = time.time() - window_s
    out_lines: list[dict] = []
    files_scanned = 0
    files_skipped_old = 0

    for log_dir in _LOG_DIRS:
        if not log_dir.is_dir():
            continue
        dir_slug = log_dir.name
        for path in sorted(log_dir.iterdir()):
            if not path.is_file():
                continue
            name = path.name
            if not (name.endswith(".log") or name.endswith(".stdout.log") or
                    name.endswith(".stderr.log")):
                continue
            try:
                stat = path.stat()
            except Exception:
                continue
            if stat.st_size == 0:
                continue
            if stat.st_mtime < cutoff_ts:
                # Archivo no modificado dentro de la ventana → skip.
                files_skipped_old += 1
                continue
            files_scanned += 1
            # Derivar service + kind del nombre (mismo mapping que el index).
            base = name
            kind = "stdout"
            if base.endswith(".error.log"):
                base = base[: -len(".error.log")]; kind = "stderr"
            elif base.endswith(".stderr.log"):
                base = base[: -len(".stderr.log")]; kind = "stderr"
            elif base.endswith(".stdout.log"):
                base = base[: -len(".stdout.log")]; kind = "stdout"
            elif base.endswith(".log"):
                base = base[: -len(".log")]; kind = "stdout"

            raw = _read_tail_lines(path, _LOG_GLOBAL_TAIL_PER_FILE)
            mtime_dt = datetime.fromtimestamp(stat.st_mtime)
            n_total = len(raw)
            last_ts: str | None = None
            for idx, ln in enumerate(raw):
                level = _classify_log_line(ln)
                # Filter por level.
                if level_filter == "error" and level != "error":
                    continue
                if level_filter == "warn_error" and level not in ("warn", "error"):
                    continue
                ts = _extract_log_ts(ln)
                ts_inferred = False
                ts_synthetic = False
                if ts:
                    last_ts = ts
                elif last_ts is not None:
                    ts = last_ts
                    ts_inferred = True
                else:
                    # Fallback: mtime - (offset from tail) seg. Para que las
                    # líneas más cerca del final del archivo queden con ts
                    # más reciente.
                    secs_back = (n_total - 1 - idx)
                    synthetic_dt = mtime_dt - timedelta(seconds=secs_back)
                    ts = synthetic_dt.isoformat(timespec="seconds")
                    ts_synthetic = True
                # Filter por window — si el ts (real o sintético) cae
                # antes de cutoff, skip.
                try:
                    ts_dt = datetime.fromisoformat(ts)
                    if ts_dt.timestamp() < cutoff_ts:
                        continue
                except Exception:
                    pass
                out_lines.append({
                    "ts": ts,
                    "ts_inferred": ts_inferred,
                    "ts_synthetic": ts_synthetic,
                    "level": level,
                    "service": base,
                    "dir": dir_slug,
                    "kind": kind,
                    "ref": f"{dir_slug}/{name}",
                    "text": ln,
                })

    # Sort por ts desc (más reciente primero). Cap a _LOG_GLOBAL_MAX_LINES.
    out_lines.sort(key=lambda x: x["ts"], reverse=True)
    truncated = len(out_lines) > _LOG_GLOBAL_MAX_LINES
    if truncated:
        out_lines = out_lines[:_LOG_GLOBAL_MAX_LINES]

    # Counts por nivel + por service.
    counts_by_level = {"error": 0, "warn": 0, "ok": 0, "info": 0}
    counts_by_service: dict[str, int] = {}
    for ln in out_lines:
        counts_by_level[ln["level"]] = counts_by_level.get(ln["level"], 0) + 1
        counts_by_service[ln["service"]] = counts_by_service.get(ln["service"], 0) + 1
    top_services = sorted(counts_by_service.items(), key=lambda x: -x[1])[:10]

    return {
        "window_seconds": window_s,
        "level_filter": level_filter,
        "lines": out_lines,
        "lines_total": len(out_lines),
        "truncated": truncated,
        "files_scanned": files_scanned,
        "files_skipped_old": files_skipped_old,
        "counts_by_level": counts_by_level,
        "top_services": [{"service": s, "count": c} for s, c in top_services],
        "scanned_at": datetime.now().isoformat(timespec="seconds"),
    }


@app.get("/api/logs/errors")
def logs_global_errors(
    since_seconds: int = _LOG_GLOBAL_DEFAULT_WINDOW_S,
    level: str = "error",
    nocache: int = 0,
) -> dict:
    """Feed global de errores agregado de TODOS los logs.

    Args:
        since_seconds: ventana de tiempo (default 1h, max 7d).
        level: "error" | "warn_error" (incluye warn) | "all".
        nocache: 1 fuerza refresh del cache.

    Response shape: ver `_build_global_errors_payload`.
    """
    if level not in ("error", "warn_error", "all"):
        raise HTTPException(
            status_code=400,
            detail=f"level inválido: {level!r} (esperaba error|warn_error|all)",
        )
    window_s = max(60, min(int(since_seconds), _LOG_GLOBAL_MAX_WINDOW_S))
    cache_key = (window_s, level)
    now_mono = time.monotonic()
    with _LOG_GLOBAL_CACHE_LOCK:
        entry = _LOG_GLOBAL_CACHE.get(cache_key)
        fresh = entry is not None and (now_mono - entry["ts"]) < _LOG_GLOBAL_CACHE_TTL
    if nocache or not fresh:
        payload = _build_global_errors_payload(window_s, level)
        with _LOG_GLOBAL_CACHE_LOCK:
            _LOG_GLOBAL_CACHE[cache_key] = {"payload": payload, "ts": now_mono}
        return payload
    return entry["payload"]


# ── /api/diagnose-error — LLM-powered error diagnosis ─────────────────
# El user clickea el botón "🩺 fix con IA" al lado de una línea con
# level=error en el viewer, y recibe streaming del LLM con un análisis
# del error: causa probable, severidad, propuesta de fix.
#
# El frontend (web/static/diagnose-modal.js) consume el SSE con eventos:
#   - {type: "model", name: "..."}
#   - {type: "token", content: "..."}
#   - {type: "done"}
#   - {type: "error", message: "..."}
#
# Scope (Nivel 1):
#   - SOLO diagnóstico textual. NO ejecuta comandos, NO edita archivos.
#   - El LLM recibe el error + ~20 líneas de contexto + service.
#   - El LLM puede SUGERIR comandos en bloques ```bash``` que el user
#     decide ejecutar a mano (copy-paste a terminal).
#
# El frontend tiene un endpoint `/api/diagnose-error/execute` para
# auto-ejecutar comandos del LLM, pero por seguridad ese endpoint está
# deshabilitado en este server (returns 503). Habilitarlo requiere una
# whitelist conservadora de comandos seguros + audit log + diseño
# pensado en otra entrega.

class _DiagnoseErrorRequest(BaseModel):
    error_text: str = Field(..., description="La línea de log con el error")
    service: str = Field("", description="Service de origen (ej. 'watch')")
    file: str = Field("", description="Label del archivo de origen")
    line_n: int = Field(0, description="Número de línea reverso (1=más reciente)")
    timestamp: str | None = Field(None, description="Timestamp ISO de la línea")
    context_lines: list[str] = Field(default_factory=list, description="Líneas previas (cronológicamente)")

    @field_validator("error_text")
    @classmethod
    def _error_not_empty(cls, v: str) -> str:
        v = (v or "").strip()
        if not v:
            raise ValueError("error_text vacío")
        if len(v) > 4000:
            v = v[:4000] + "…(truncado)"
        return v


# Model preference para el feature diagnose-error. command-r:35b es ~2-3×
# mejor que qwen2.5:7b en razonamiento técnico sobre stack traces + bash
# (bench informal 2026-04-26). Si command-r no está instalado, caemos a
# resolve_chat_model() del rag.py que devuelve el chat default del user.
#
# Cacheado durante la vida del proceso — la lista de modelos no cambia
# salvo que el user haga `ollama pull` mientras el web está corriendo,
# en cuyo caso un restart del daemon lo recoge.
_DIAGNOSE_MODEL_PREFERENCE: tuple[str, ...] = (
    "command-r:latest",
    "command-r:35b",
    "qwen3:30b-a3b",
)
_DIAGNOSE_MODEL_RESOLVED: str | None = None


def _resolve_diagnose_model() -> str:
    """Devuelve el primer modelo de _DIAGNOSE_MODEL_PREFERENCE instalado;
    fallback a resolve_chat_model() del rag.py (suele ser qwen2.5:7b)."""
    global _DIAGNOSE_MODEL_RESOLVED
    if _DIAGNOSE_MODEL_RESOLVED is not None:
        return _DIAGNOSE_MODEL_RESOLVED
    try:
        available = {m.model for m in ollama.list().models}
    except Exception:
        available = set()
    for candidate in _DIAGNOSE_MODEL_PREFERENCE:
        if candidate in available:
            _DIAGNOSE_MODEL_RESOLVED = candidate
            return candidate
    _DIAGNOSE_MODEL_RESOLVED = resolve_chat_model()
    return _DIAGNOSE_MODEL_RESOLVED


_DIAGNOSE_ERROR_SYSTEM_PROMPT = """\
Sos un asistente experto en el stack `obsidian-rag` de Fer (Fernando Ferrari).
El user te muestra una línea de log con un error y pide diagnóstico.

Stack relevante:
- Local-first RAG sobre vault Obsidian, single-file `rag.py` (~50k líneas)
  + `web/server.py` (FastAPI) + daemons launchd (watch, ingest-*, anticipate,
  reminder-wa-push, wa-scheduled-send, etc.).
- SQLite-vec (`ragvec.db`) con escrituras concurrentes — `database is
  locked` es el patrón típico de contención, recoverable.
- Ollama local + sentence-transformers + reranker (BGE).

Errores frecuentes:
- `OperationalError: no such column: ...` → falta migration de schema.
- `database is locked` → contención SQLite, recoverable. Serio sólo si
  se acumulan decenas seguidos en pocos minutos.
- `UserWarning: leaked semaphore` → tqdm/loky multi-process. Patch ya
  documentado en `web/server.py` líneas iniciales.
- `another row available` → bug real: `LIMIT 1` faltante en SQL o join
  que duplica.

Formato de respuesta (markdown, español rioplatense):

## Qué está pasando
1-2 oraciones: causa probable + severidad (ok/warning/serio).

## Cómo arreglarlo
Pasos concretos. Si está documentado en CLAUDE.md o docs/, apuntá ahí.
Si no estás seguro, decilo y pedí más contexto en vez de inventar.

## Comandos sugeridos
Si hay comandos shell que ayudan, ponelos en bloques ```bash```. Cada
comando en su propia línea, sin pipes (`|`), redirects (`>`), command
substitution (`$()`), ni encadenamiento (`;`, `&&`). El user va a poder
clickear "▶ ejecutar" y el server los corre directo — pero hay una
WHITELIST estricta del lado del server, así que sólo estas formas pasan:

- `launchctl kickstart -k <label>` — reiniciar un daemon. Label debe
  matchear `com.fer.obsidian-rag-*` o `com.fer.whatsapp-*`. Aceptamos
  el prefix `gui/501/` opcional. Ejemplos OK:
  - `launchctl kickstart -k com.fer.obsidian-rag-watch`
  - `launchctl kickstart -k gui/501/com.fer.obsidian-rag-wa-scheduled-send`
- `launchctl list com.fer.obsidian-rag-<service>` — ver estado.
- `launchctl print gui/501/com.fer.obsidian-rag-<service>` — info detallada.
- `tail [-n N] /Users/fer/.local/share/obsidian-rag/<archivo>.log` — leer log.
  Soportamos también `tail -50 <path>`. NO uses `-f` (se cuelga).
- `head [-n N] <log_path>` — primeras N líneas.
- `wc -l <log_path>` — contar líneas.
- `cat <log_path>` — todo el archivo.
- `ls -la /Users/fer/.local/share/obsidian-rag/` — listar logs.
- `rag stats` / `rag status` / `rag vault list` — CLI read-only.

Cualquier otro comando va a ser RECHAZADO por la whitelist con un 403.
NUNCA sugieras `rm`, `mv`, `cp`, `sudo`, `bash -c`, `python -c`, `git push`,
`kill`, ni nada con shell metachars. Si la solución requiere algo así,
NO lo pongas en un bloque ```bash``` — describilo en prosa para que el
user lo haga a mano.

NO inventes paths, archivos, o líneas que no estén en el contexto.
Si el "error" parece un falso positivo del clasificador, decilo.
"""


def _build_diagnose_error_prompt(req: _DiagnoseErrorRequest) -> str:
    parts = []
    if req.timestamp:
        parts.append(f"**Timestamp**: {req.timestamp}")
    if req.service:
        parts.append(f"**Service**: `{req.service}`")
    if req.file:
        parts.append(f"**Archivo**: `{req.file}`")
    if req.line_n:
        parts.append(f"**Línea N°** (reverse desde el final): {req.line_n}")
    parts.append("")
    if req.context_lines:
        parts.append("**Contexto previo** (las líneas anteriores en el log):")
        parts.append("```")
        for ln in req.context_lines[-20:]:
            parts.append(ln)
        parts.append("```")
        parts.append("")
    parts.append("**Línea con el error**:")
    parts.append("```")
    parts.append(req.error_text)
    parts.append("```")
    parts.append("")
    parts.append("Diagnosticá según el formato del system prompt.")
    return "\n".join(parts)


@app.post("/api/diagnose-error")
def diagnose_error(req: _DiagnoseErrorRequest, request: Request) -> StreamingResponse:
    """SSE streaming del LLM con el diagnóstico del error.

    Rate-limited (mismo bucket que /api/chat). El SSE emite eventos:
    `model` (al inicio con el nombre del modelo), `token` (cada chunk
    del response), `done` (al final), `error` (si algo falla).
    """
    client_ip = (request.client.host if request.client else "unknown")
    _check_rate_limit(_CHAT_BUCKETS, client_ip,
                      _CHAT_RATE_LIMIT, _CHAT_RATE_WINDOW)

    user_prompt = _build_diagnose_error_prompt(req)
    model = _resolve_diagnose_model()

    def _stream():
        try:
            yield f"data: {json.dumps({'type': 'model', 'name': model})}\n\n"
            stream = ollama.chat(
                model=model,
                messages=[
                    {"role": "system", "content": _DIAGNOSE_ERROR_SYSTEM_PROMPT},
                    {"role": "user", "content": user_prompt},
                ],
                stream=True,
                options={
                    "temperature": 0.3,
                    "seed": 42,
                    "num_predict": 800,
                    "num_ctx": 4096,
                },
                keep_alive=chat_keep_alive(),
            )
            for chunk in stream:
                msg = chunk.get("message") if isinstance(chunk, dict) else getattr(chunk, "message", None)
                content = ""
                if msg is not None:
                    content = (msg.get("content") if isinstance(msg, dict)
                               else getattr(msg, "content", "")) or ""
                if content:
                    yield f"data: {json.dumps({'type': 'token', 'content': content})}\n\n"
                done = chunk.get("done") if isinstance(chunk, dict) else getattr(chunk, "done", False)
                if done:
                    yield f"data: {json.dumps({'type': 'done'})}\n\n"
                    break
        except Exception as e:
            err_msg = f"{type(e).__name__}: {e}"
            yield f"data: {json.dumps({'type': 'error', 'message': err_msg})}\n\n"

    return StreamingResponse(
        _stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache, no-transform",
            "X-Accel-Buffering": "no",
        },
    )


# ── /api/diagnose-error/execute — auto-ejecución segura del LLM ───────
# El user pidió "que resuelva solo el problema". Esto requiere ejecutar
# los comandos shell que el LLM sugiere. Como ejecutar shell arbitrario
# producido por un LLM es peligroso, aplicamos varias capas de defensa:
#
#   1. WHITELIST estricta de binarios + validators de argumentos. El
#      comando se rechaza si su primer token no está en la whitelist.
#   2. NO usamos `shell=True` — pasamos argv directo a subprocess. Eso
#      elimina shell metachars (`;`, `|`, `&`, `$()`, backticks, etc.)
#      como vector de ataque incluso si el shlex parser fallara.
#   3. Detección defensiva de metachars en el string CRUDO antes de
#      shlex.split — si el LLM emite `tail watch.log; rm -rf /` lo
#      cortamos antes de tocarlo.
#   4. Argumentos validados por whitelist entry — `tail` sólo acepta
#      paths bajo `~/.local/share/obsidian-rag/`, `launchctl kickstart`
#      sólo acepta labels que matchean `com.fer.obsidian-rag-*`, etc.
#   5. Timeout de 15s en subprocess.run para que un comando colgado
#      no quede vivo.
#   6. Audit log JSONL en `~/.local/share/obsidian-rag/diagnose_executions.jsonl`
#      con timestamp + comando original + comando ejecutado + exit_code
#      + stdout/stderr (truncados). Auditable post-mortem.
#   7. Rate limit (mismo bucket que /api/chat).
#
# Lo que NO está permitido:
#   - rm, mv, cp, dd (filesystem mutator)
#   - sudo, su (privilege escalation)
#   - bash, sh, zsh, python -c (arbitrary code)
#   - git push, git reset --hard, git checkout (repo mutator)
#   - kill -9 (signal a procesos arbitrarios — `launchctl kickstart -k`
#     mata el daemon target específico de forma controlada)
#   - cualquier cosa con stdin redirigido o pipes

_LOG_DIR_ABS = (Path.home() / ".local/share/obsidian-rag").resolve()
_DIAGNOSE_AUDIT_LOG = _LOG_DIR_ABS / "diagnose_executions.jsonl"
_DIAGNOSE_TIMEOUT_S = 15.0
_DIAGNOSE_OUTPUT_TRUNCATE = 8000  # bytes por stdout/stderr


def _is_safe_log_path(arg: str) -> bool:
    """Path debe estar dentro de _LOG_DIR_ABS."""
    if not arg:
        return False
    try:
        p = Path(arg).expanduser().resolve()
    except Exception:
        return False
    try:
        p.relative_to(_LOG_DIR_ABS)
    except ValueError:
        return False
    return True


_LAUNCHCTL_LABEL_RE = re.compile(r"^com\.fer\.(obsidian-rag-|whatsapp-)[a-z0-9_-]+$")
_LAUNCHCTL_GUI_RE = re.compile(r"^gui/\d+/com\.fer\.(obsidian-rag-|whatsapp-)[a-z0-9_-]+$")


def _is_safe_launchctl_target(arg: str) -> bool:
    return bool(_LAUNCHCTL_LABEL_RE.match(arg) or _LAUNCHCTL_GUI_RE.match(arg))


def _is_self_daemon_target(arg: str) -> bool:
    """¿El target apunta al propio web daemon? Si sí, rechazamos kickstart
    porque mataríamos al proceso mid-request. List/print SÍ se permiten
    (son read-only, no afectan al daemon)."""
    return "obsidian-rag-web" in arg


def _is_int_string(arg: str) -> bool:
    return arg.lstrip("-").isdigit() and len(arg) <= 6


def _validate_launchctl_args(args: list[str]) -> bool:
    """`launchctl kickstart [-k|-p|-s] <target>` | `list [<label>]` |
    `print <target>`. Hard-rejects kickstart del propio daemon web."""
    if not args:
        return False
    sub = args[0]
    if sub == "kickstart":
        flags_ok = {"-k", "-p", "-s"}
        rest = args[1:]
        i = 0
        while i < len(rest) and rest[i].startswith("-"):
            if rest[i] not in flags_ok:
                return False
            i += 1
        if len(rest) - i != 1:
            return False
        target = rest[i]
        if not _is_safe_launchctl_target(target):
            return False
        # Hard-defense: kickstart al propio daemon mata el endpoint
        # mid-request. Lo rechazamos sin importar lo que diga el prompt.
        if _is_self_daemon_target(target):
            return False
        return True
    if sub == "list":
        if len(args) == 1:
            return True
        if len(args) == 2:
            return _is_safe_launchctl_target(args[1])
        return False
    if sub == "print":
        if len(args) != 2:
            return False
        return _is_safe_launchctl_target(args[1])
    return False


def _validate_tail_head_args(args: list[str]) -> bool:
    """`tail [-n N] <log_path>` o `tail -<N> <log_path>` (BSD)."""
    if not args:
        return False
    path = None
    i = 0
    while i < len(args):
        a = args[i]
        if a == "-n" or a == "-c":
            if i + 1 >= len(args) or not _is_int_string(args[i + 1]):
                return False
            i += 2
            continue
        if a == "-f":
            return False  # follow se cuelga
        if a.startswith("-") and a != "-" and _is_int_string(a):
            i += 1
            continue
        if a.startswith("-"):
            return False
        if path is not None:
            return False
        path = a
        i += 1
    return path is not None and _is_safe_log_path(path)


def _validate_wc_args(args: list[str]) -> bool:
    if not args:
        return False
    flags_ok = {"-l", "-c", "-w", "-m"}
    path = None
    for a in args:
        if a in flags_ok:
            continue
        if a.startswith("-"):
            return False
        if path is not None:
            return False
        path = a
    return path is not None and _is_safe_log_path(path)


def _validate_cat_args(args: list[str]) -> bool:
    if len(args) != 1:
        return False
    if args[0].startswith("-"):
        return False
    return _is_safe_log_path(args[0])


def _validate_rag_args(args: list[str]) -> bool:
    """Sólo subcomandos read-only de la CLI."""
    if not args:
        return False
    READ_ONLY_SUBCMDS = {"stats", "status", "vault", "version", "config"}
    if args[0] not in READ_ONLY_SUBCMDS:
        return False
    if args[0] == "vault":
        if len(args) == 1:
            return True
        if args[1] in {"list", "info", "ls"}:
            return len(args) <= 3 and all(not a.startswith("/") for a in args[2:])
        return False
    if len(args) > 2:
        return False
    if len(args) == 2 and ("/" in args[1] or args[1].startswith("-")):
        return False
    return True


def _validate_ls_args(args: list[str]) -> bool:
    flags_ok = {"-l", "-a", "-la", "-al", "-h", "-lh", "-lah", "-laH"}
    path = None
    for a in args:
        if a in flags_ok:
            continue
        if a.startswith("-"):
            return False
        if path is not None:
            return False
        path = a
    return path is not None and _is_safe_log_path(path)


def _which_safe(name: str, candidates: tuple[str, ...]) -> str | None:
    for c in candidates:
        if Path(c).is_file():
            return c
    return None


_SAFE_LAUNCHCTL = _which_safe("launchctl", ("/bin/launchctl",))
_SAFE_TAIL = _which_safe("tail", ("/usr/bin/tail",))
_SAFE_HEAD = _which_safe("head", ("/usr/bin/head",))
_SAFE_WC = _which_safe("wc", ("/usr/bin/wc",))
_SAFE_CAT = _which_safe("cat", ("/bin/cat",))
_SAFE_LS = _which_safe("ls", ("/bin/ls",))
_SAFE_RAG = _which_safe("rag", (
    str(Path.home() / ".local/bin/rag"),
    "/usr/local/bin/rag",
))

_DIAGNOSE_COMMAND_REGISTRY: dict[str, tuple[str | None, "Callable[[list[str]], bool]"]] = {
    "launchctl": (_SAFE_LAUNCHCTL, _validate_launchctl_args),
    "tail": (_SAFE_TAIL, _validate_tail_head_args),
    "head": (_SAFE_HEAD, _validate_tail_head_args),
    "wc": (_SAFE_WC, _validate_wc_args),
    "cat": (_SAFE_CAT, _validate_cat_args),
    "ls": (_SAFE_LS, _validate_ls_args),
    "rag": (_SAFE_RAG, _validate_rag_args),
}

# Defense-in-depth: detección de metachars en raw string antes de shlex.
# Audit 2026-04-26 BUG #9: agregar `\x00` (null byte) y otros whitespace
# control. shlex/execvp manejan null como string-end pero el audit log
# guarda `command_original` truncado → forensics se rompen post-mortem.
_DANGEROUS_METACHARS_RE = re.compile(r"[;&|<>`$()\n\r\t\v\f\x00]")


def _validate_safe_command(cmd_str: str) -> tuple[list[str] | None, str]:
    """Validar un comando shell crudo contra la whitelist.

    Retorna `(argv, "")` si es seguro y se puede ejecutar, o `(None,
    reason)` con un mensaje de error legible.
    """
    cmd = (cmd_str or "").strip()
    if not cmd:
        return None, "comando vacío"
    if len(cmd) > 500:
        return None, "comando demasiado largo (>500 chars)"
    m = _DANGEROUS_METACHARS_RE.search(cmd)
    if m:
        return None, f"metacharacter shell prohibido: {m.group(0)!r}"
    import shlex
    try:
        argv = shlex.split(cmd, posix=True)
    except ValueError as e:
        return None, f"comando mal formado: {e}"
    if not argv:
        return None, "comando vacío post-parse"
    binary_name = argv[0]
    if "/" in binary_name:
        return None, "el comando debe ser un nombre simple, no un path"
    entry = _DIAGNOSE_COMMAND_REGISTRY.get(binary_name)
    if entry is None:
        return None, f"comando '{binary_name}' no está en la whitelist"
    abs_path, validator = entry
    if abs_path is None:
        return None, f"binario '{binary_name}' no encontrado en el sistema"
    if not validator(argv[1:]):
        return None, f"argumentos inválidos para '{binary_name}'"
    return [abs_path] + argv[1:], ""


def _audit_diagnose_execution(record: dict) -> None:
    """Append-only audit log de ejecuciones. Best-effort."""
    try:
        _DIAGNOSE_AUDIT_LOG.parent.mkdir(parents=True, exist_ok=True)
        with _DIAGNOSE_AUDIT_LOG.open("a", encoding="utf-8") as f:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
    except Exception as e:
        print(f"[diagnose-execute] audit log write failed: "
              f"{type(e).__name__}: {e}", file=sys.stderr, flush=True)


class _DiagnoseExecuteRequest(BaseModel):
    command: str = Field(..., description="Comando shell que el LLM sugirió")


@app.post("/api/diagnose-error/execute")
def diagnose_error_execute(req: _DiagnoseExecuteRequest, request: Request) -> dict:
    """Auto-ejecutar un comando del LLM, validado contra una whitelist.

    Retorna `{exit_code, stdout, stderr, command_executed, command_original,
    duration_s, timed_out}`. Si el comando no pasa la validación,
    retorna 403 con `{detail}` describiendo por qué.

    Hard timeout 15s. Audit log a `~/.local/share/obsidian-rag/diagnose_executions.jsonl`.
    """
    client_ip = (request.client.host if request.client else "unknown")
    _check_rate_limit(_CHAT_BUCKETS, client_ip,
                      _CHAT_RATE_LIMIT, _CHAT_RATE_WINDOW)

    cmd_original = req.command
    argv, reason = _validate_safe_command(cmd_original)
    if argv is None:
        record = {
            "ts": datetime.now().isoformat(timespec="seconds"),
            "command_original": cmd_original,
            "rejected": True,
            "reason": reason,
            "client_ip": client_ip,
        }
        _audit_diagnose_execution(record)
        raise HTTPException(
            status_code=403,
            detail=f"comando rechazado por la whitelist: {reason}",
        )

    t0 = time.monotonic()
    try:
        result = subprocess.run(
            argv,
            capture_output=True,
            text=True, errors="replace",
            timeout=_DIAGNOSE_TIMEOUT_S,
            shell=False,
            check=False,
            cwd=str(Path.home()),
            env={
                "PATH": "/usr/bin:/bin:/usr/sbin:/sbin",
                "HOME": str(Path.home()),
                "LANG": "en_US.UTF-8",
            },
        )
        stdout = (result.stdout or "")[:_DIAGNOSE_OUTPUT_TRUNCATE]
        stderr = (result.stderr or "")[:_DIAGNOSE_OUTPUT_TRUNCATE]
        exit_code = result.returncode
        timed_out = False
    except subprocess.TimeoutExpired as e:
        stdout = (e.stdout.decode("utf-8", "replace") if e.stdout else "")[:_DIAGNOSE_OUTPUT_TRUNCATE]
        stderr = (
            (e.stderr.decode("utf-8", "replace") if e.stderr else "")
            + f"\n[timeout: {_DIAGNOSE_TIMEOUT_S}s]"
        )[:_DIAGNOSE_OUTPUT_TRUNCATE]
        exit_code = 124
        timed_out = True
    except Exception as e:
        stdout = ""
        stderr = f"{type(e).__name__}: {e}"
        exit_code = -1
        timed_out = False
    duration_s = round(time.monotonic() - t0, 3)

    record = {
        "ts": datetime.now().isoformat(timespec="seconds"),
        "command_original": cmd_original,
        "command_executed": argv,
        "exit_code": exit_code,
        "duration_s": duration_s,
        "stdout_len": len(stdout),
        "stderr_len": len(stderr),
        "timed_out": timed_out,
        "client_ip": client_ip,
        "stdout_preview": stdout[:300],
        "stderr_preview": stderr[:300],
    }
    _audit_diagnose_execution(record)

    return {
        "exit_code": exit_code,
        "stdout": stdout,
        "stderr": stderr,
        "command_executed": argv,
        "command_original": cmd_original,
        "duration_s": duration_s,
        "timed_out": timed_out,
    }


# ── /api/auto-fix — agent loop que resuelve solo el problema ──────────
# El user pidió "no tiene que decirte qué hacer, tiene que hacerlo solo".
# Esto es un agent loop: el LLM diagnostica + ejecuta comandos + verifica
# en un ciclo, hasta resolver el error o agotar el cap de turnos.
#
# Flow por turno:
#   1. LLM recibe el contexto acumulado (error original + resultados de
#      acciones previas) y devuelve un JSON con shape:
#        { "thought": "..", "action": "<cmd>" | null, "done": bool, "summary": "" }
#   2. Si action no-null: el server lo valida con _validate_safe_command
#      (misma whitelist que /execute), lo ejecuta con timeout, y el
#      stdout/stderr se inyecta al contexto del próximo turno.
#   3. Si done=true: el loop termina y emite el summary al frontend.
#   4. Cap a _AUTO_FIX_MAX_TURNS para evitar bucles infinitos.
#
# El SSE emite eventos:
#   - {type: "turn", n: 1}
#   - {type: "thought", text: "..."}
#   - {type: "action", command: "..."}
#   - {type: "action_result", exit_code: 0, stdout: "...", stderr: "..."}
#   - {type: "action_rejected", command: "...", reason: "..."}
#   - {type: "done", summary: "..."}
#   - {type: "max_turns_reached"}
#   - {type: "error", message: "..."}

_AUTO_FIX_MAX_TURNS = 8
_AUTO_FIX_OUTPUT_TRUNCATE = 2000  # truncado del stdout/stderr inyectado al LLM

_AUTO_FIX_SYSTEM_PROMPT = """\
Sos un agente que resuelve errores del stack obsidian-rag de Fer.

Recibís un error en un log y tenés que diagnosticar Y resolver el
problema en un ciclo de hasta 6 turnos. NO le das al user instrucciones
para que él haga algo — vos hacés el trabajo ejecutando comandos.

Stack relevante:
- Daemons launchd: com.fer.obsidian-rag-{watch,web,wa-scheduled-send,
  ingest-{calendar,gmail,drive,whatsapp,reminders},anticipate,
  reminder-wa-push,maintenance,morning,today, ...}.
- Logs: /Users/fer/.local/share/obsidian-rag/<servicename>.log y
  <servicename>.error.log.
- SQLite-vec con escrituras concurrentes (database is locked es típico,
  recoverable).

Tools disponibles (whitelist estricta — cualquier otra cosa es rechazada):
- `launchctl kickstart -k gui/501/<label>` — reiniciar daemon. IMPORTANTE:
  el label DEBE venir prefixed con `gui/501/`, sino macOS lo rechaza.
  Ejemplo correcto: `launchctl kickstart -k gui/501/com.fer.obsidian-rag-watch`
  Ejemplo INCORRECTO: `launchctl kickstart -k com.fer.obsidian-rag-watch`
  (tira `Unrecognized target specifier`).
- `launchctl list com.fer.obsidian-rag-<service>` — ver estado del daemon
  (este SÍ usa label desnudo, sin gui/501/).
- `launchctl print gui/501/com.fer.obsidian-rag-<service>` — info detallada.
- `tail [-n N] <log_path>` — leer últimas líneas (NO uses -f, se cuelga).
- `head [-n N] <log_path>` — primeras líneas.
- `wc -l <log_path>` — contar líneas.
- `cat <log_path>` — todo el archivo.
- `ls -la <dir>` — listar files (sólo bajo el log dir).
- `rag stats` / `rag status` / `rag vault list` — CLI read-only.

Workflow esperado (sé EFICIENTE — máximo 1-2 turnos de investigación):
1. Investigá UNA vez: ej. `launchctl list <label>` o `tail -50 <log>`.
   NO hagas tail múltiples veces — la primera lectura ya te debería
   dar suficiente contexto. Si necesitás MÁS líneas, usá `tail -n 200`
   en el siguiente turno, NO repitas `tail -50`.
2. Decidí el fix: kickstart del daemon (caso típico) o no-acción
   (si es un error transient/aislado).
3. Aplicá el fix con la sintaxis correcta (`gui/501/<label>` para kickstart).
4. Verificá: `tail -10` post-restart o `launchctl list` para confirmar PID nuevo.
5. Devolvé done=true con summary.

Errores comunes y fix asociado:
- "database is locked" + REPETIDO (≥3 ocurrencias en últimos 5 min):
  kickstart del daemon → `launchctl kickstart -k gui/501/<label>`.
  Si es 1-2 ocurrencias aisladas: NO requiere acción (el daemon retrió
  bien). Devolvé done=true marcándolo como aislado.
- "OperationalError: no such column" → schema desincronizado. NO se
  resuelve sin tocar código. Devolvé done=true con summary explicando
  que requiere intervención humana (schema migration).
- "UserWarning: leaked semaphore" → ruido de tqdm/loky. Falso positivo,
  no es serio. Devolvé done=true marcándolo como ignorable.
- "another row available" → bug SQL real (LIMIT 1 faltante). No se
  resuelve con kickstart. Devolvé done=true explicando que requiere
  fix de código.

FORMATO DE RESPUESTA (responder SIEMPRE con JSON válido):
{
  "thought": "explicación corta de qué vas a hacer ahora (≤2 frases)",
  "action": "<comando exacto sin pipes ni metachars>" o null,
  "done": false,
  "summary": ""
}

Cuando termines (resuelto o no-resoluble):
{
  "thought": "última observación",
  "action": null,
  "done": true,
  "summary": "qué hiciste / qué pasó / qué requiere atención manual"
}

Reglas:
- NUNCA emitas comandos con `;`, `&&`, `|`, `>`, `$()`, backticks. La
  whitelist los rechaza y perdés un turno.
- NUNCA inventes paths que no estén en el contexto.
- NUNCA reinicies el daemon `obsidian-rag-web` (com.fer.obsidian-rag-web).
  Vos vivís adentro de ese daemon — kickstartearlo te mata mid-request
  y el user pierde la conexión sin ver el resultado. Si el error es
  del daemon web, devolvé done=true explicando qué viste pero pediendo
  que el user reinicie a mano.
- Si después de 2-3 acciones no encontrás progreso, devolvé done=true
  con summary explicando qué intentaste y qué requiere review humano.
- Sé conservador: si dudás entre kickstart y no-acción, prefiero no-acción.
"""


def _build_initial_auto_fix_user_prompt(req: "_AutoFixRequest") -> str:
    parts = ["Tenés que resolver este error:"]
    parts.append("")
    if req.timestamp:
        parts.append(f"**Timestamp**: {req.timestamp}")
    if req.service:
        parts.append(f"**Service**: `{req.service}`")
    if req.file:
        parts.append(f"**Archivo**: `{req.file}`")
    parts.append("")
    if req.context_lines:
        parts.append("**Contexto previo del log** (las líneas anteriores):")
        parts.append("```")
        for ln in req.context_lines[-15:]:
            parts.append(ln)
        parts.append("```")
        parts.append("")
    parts.append("**Línea con el error**:")
    parts.append("```")
    parts.append(req.error_text)
    parts.append("```")
    parts.append("")
    parts.append(
        f"Tenés hasta {_AUTO_FIX_MAX_TURNS} turnos. Empezá investigando "
        "(ver estado del daemon o leer más del log) antes de actuar. "
        "Respondé con JSON según el schema del system prompt."
    )
    return "\n".join(parts)


class _AutoFixRequest(BaseModel):
    error_text: str = Field(..., description="La línea de log con el error")
    service: str = Field("", description="Service de origen (ej. 'watch')")
    file: str = Field("", description="Label del archivo")
    line_n: int = Field(0, description="Número de línea reverso")
    timestamp: str | None = Field(None, description="Timestamp ISO")
    context_lines: list[str] = Field(default_factory=list)

    @field_validator("error_text")
    @classmethod
    def _error_not_empty(cls, v: str) -> str:
        v = (v or "").strip()
        if not v:
            raise ValueError("error_text vacío")
        if len(v) > 4000:
            v = v[:4000] + "…(truncado)"
        return v


def _execute_whitelisted_command(cmd_str: str) -> dict:
    """Ejecutar un comando del whitelist y devolver dict con resultado.

    Reusa _validate_safe_command + _audit_diagnose_execution. Si el
    comando es rechazado, devuelve `{rejected: True, reason}`. Si pasa,
    `{exit_code, stdout, stderr, command_executed, duration_s, timed_out}`.
    """
    argv, reason = _validate_safe_command(cmd_str)
    if argv is None:
        record = {
            "ts": datetime.now().isoformat(timespec="seconds"),
            "command_original": cmd_str,
            "rejected": True,
            "reason": reason,
            "via": "auto-fix",
        }
        _audit_diagnose_execution(record)
        return {"rejected": True, "reason": reason}

    t0 = time.monotonic()
    try:
        result = subprocess.run(
            argv,
            capture_output=True,
            text=True, errors="replace",
            timeout=_DIAGNOSE_TIMEOUT_S,
            shell=False,
            check=False,
            cwd=str(Path.home()),
            env={
                "PATH": "/usr/bin:/bin:/usr/sbin:/sbin",
                "HOME": str(Path.home()),
                "LANG": "en_US.UTF-8",
            },
        )
        stdout = (result.stdout or "")[:_DIAGNOSE_OUTPUT_TRUNCATE]
        stderr = (result.stderr or "")[:_DIAGNOSE_OUTPUT_TRUNCATE]
        exit_code = result.returncode
        timed_out = False
    except subprocess.TimeoutExpired as e:
        stdout = (e.stdout.decode("utf-8", "replace") if e.stdout else "")[:_DIAGNOSE_OUTPUT_TRUNCATE]
        stderr = (
            (e.stderr.decode("utf-8", "replace") if e.stderr else "")
            + f"\n[timeout: {_DIAGNOSE_TIMEOUT_S}s]"
        )[:_DIAGNOSE_OUTPUT_TRUNCATE]
        exit_code = 124
        timed_out = True
    except Exception as e:
        stdout = ""
        stderr = f"{type(e).__name__}: {e}"
        exit_code = -1
        timed_out = False
    duration_s = round(time.monotonic() - t0, 3)

    record = {
        "ts": datetime.now().isoformat(timespec="seconds"),
        "command_original": cmd_str,
        "command_executed": argv,
        "exit_code": exit_code,
        "duration_s": duration_s,
        "stdout_len": len(stdout),
        "stderr_len": len(stderr),
        "timed_out": timed_out,
        "via": "auto-fix",
        "stdout_preview": stdout[:300],
        "stderr_preview": stderr[:300],
    }
    _audit_diagnose_execution(record)

    return {
        "rejected": False,
        "exit_code": exit_code,
        "stdout": stdout,
        "stderr": stderr,
        "command_executed": argv,
        "duration_s": duration_s,
        "timed_out": timed_out,
    }


@app.post("/api/auto-fix")
def auto_fix(req: _AutoFixRequest, request: Request) -> StreamingResponse:
    """Agent loop que diagnostica + resuelve el error solo.

    En cada turno: el LLM emite un JSON con `{thought, action, done,
    summary}`. Si action no-null, el server lo ejecuta y manda el
    resultado al LLM como contexto del próximo turno. Loop hasta
    `done=true` o `_AUTO_FIX_MAX_TURNS` (6).

    Streams SSE con eventos: turn, thought, action, action_result,
    action_rejected, done, max_turns_reached, error.
    """
    client_ip = (request.client.host if request.client else "unknown")
    _check_rate_limit(_CHAT_BUCKETS, client_ip,
                      _CHAT_RATE_LIMIT, _CHAT_RATE_WINDOW)

    initial_user_prompt = _build_initial_auto_fix_user_prompt(req)
    model = resolve_chat_model()

    def _stream():
        messages = [
            {"role": "system", "content": _AUTO_FIX_SYSTEM_PROMPT},
            {"role": "user", "content": initial_user_prompt},
        ]
        yield f"data: {json.dumps({'type': 'model', 'name': model})}\n\n"

        for turn_n in range(1, _AUTO_FIX_MAX_TURNS + 1):
            yield f"data: {json.dumps({'type': 'turn', 'n': turn_n})}\n\n"
            try:
                resp = ollama.chat(
                    model=model,
                    messages=messages,
                    format="json",
                    options={
                        "temperature": 0.2,
                        "seed": 42,
                        "num_predict": 400,
                        "num_ctx": 6144,
                    },
                    keep_alive=chat_keep_alive(),
                )
                raw_content = (resp.message.content if hasattr(resp, "message") else
                               resp["message"]["content"])
            except Exception as e:
                yield f"data: {json.dumps({'type': 'error', 'message': f'LLM call failed: {e}'})}\n\n"
                return

            try:
                data = json.loads(raw_content or "{}")
            except json.JSONDecodeError as e:
                yield f"data: {json.dumps({'type': 'error', 'message': f'LLM did not return valid JSON: {e}'})}\n\n"
                return

            thought = (data.get("thought") or "").strip()
            action = data.get("action")
            done = bool(data.get("done"))
            summary = (data.get("summary") or "").strip()

            if thought:
                yield f"data: {json.dumps({'type': 'thought', 'text': thought})}\n\n"

            if done:
                yield f"data: {json.dumps({'type': 'done', 'summary': summary})}\n\n"
                return

            if not action:
                # No action y no done — el LLM se confundió. Cortamos.
                yield f"data: {json.dumps({'type': 'done', 'summary': summary or 'el agente no decidió siguiente acción'})}\n\n"
                return

            yield f"data: {json.dumps({'type': 'action', 'command': action})}\n\n"
            result = _execute_whitelisted_command(action)
            if result.get("rejected"):
                yield f"data: {json.dumps({'type': 'action_rejected', 'command': action, 'reason': result['reason']})}\n\n"
                # Inyectar al LLM la rejection para que intente otra cosa.
                messages.append({"role": "assistant", "content": raw_content})
                messages.append({
                    "role": "user",
                    "content": f"La acción `{action}` fue rechazada por la whitelist: {result['reason']}. Intentá con otra acción permitida o devolvé done=true si no podés resolver.",
                })
                continue

            yield f"data: {json.dumps({'type': 'action_result', 'exit_code': result['exit_code'], 'stdout': result['stdout'][:_AUTO_FIX_OUTPUT_TRUNCATE], 'stderr': result['stderr'][:_AUTO_FIX_OUTPUT_TRUNCATE], 'duration_s': result['duration_s']})}\n\n"

            # Inyectar el resultado al contexto para el próximo turno.
            messages.append({"role": "assistant", "content": raw_content})
            obs = (
                f"Resultado de `{action}`:\n"
                f"exit_code: {result['exit_code']}\n"
                f"stdout: {result['stdout'][:_AUTO_FIX_OUTPUT_TRUNCATE]}\n"
                f"stderr: {result['stderr'][:500]}"
            )
            messages.append({"role": "user", "content": obs})

        # Si salió del loop sin done, max_turns_reached.
        yield f"data: {json.dumps({'type': 'max_turns_reached', 'limit': _AUTO_FIX_MAX_TURNS})}\n\n"

    return StreamingResponse(
        _stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache, no-transform",
            "X-Accel-Buffering": "no",
        },
    )


# ── /api/auto-fix-devin — delegar al agente Devin ─────────────────────
# El user pidió "para arreglar los errores tenes que usar devin no otro
# modelo". El endpoint /api/auto-fix con qwen2.5:7b + whitelist propia
# era una solución restringida; esto lo reemplaza delegando a la CLI de
# Devin (https://cli.devin.ai) que corre en el mismo repo y tiene acceso
# real al código, permissions del user, tools completos (read/edit/exec),
# y sus propias rules + skills configurados en `.devin/`.
#
# Cómo funciona:
#   1. Construimos un prompt con el error + contexto (igual que /auto-fix).
#   2. Spawneamos `devin -p "<prompt>"` como subprocess (modo non-
#      interactive — procesa el prompt y sale).
#   3. Capturamos stdout+stderr en vivo, stream al cliente via SSE.
#   4. Al terminar: evento `done` con el output completo + exit_code.
#
# Seguridad: NO necesitamos whitelist porque Devin respeta las
# permission rules del user configuradas en `.devin/config.json`
# (deny-list: rm -rf, sudo, force-push, reset --hard, branch -D; ask-list
# para ops sensibles; allow-list para el workflow normal del RAG).
# Devin sabe las convenciones del proyecto via CLAUDE.md cargado como rule.
#
# Costo: cada invocación de Devin consume ACUs (unidades pagas de
# Cognition). El user está consciente (pidió explícitamente usar Devin).
# Rate-limited con el mismo bucket que /api/chat para evitar spam.
#
# Latencia: Devin tarda 30-120s (depende del error). El modal muestra
# un spinner + stream del progreso en vivo.
#
# Limitaciones:
#   - `devin -p` devuelve el output completo al final, no necesariamente
#     streaming progresivo durante su trabajo (depende de cómo la CLI
#     flushee stdout). Igual lo tratamos como stream — si todo sale al
#     final, el modal muestra el resultado al final.
#   - Si la CLI requiere confirmación interactiva (workspace trust, etc.),
#     se cuelga. Pasamos --respect-workspace-trust false para evitarlo.

_DEVIN_BIN = _which_safe("devin", (
    str(Path.home() / ".local/bin/devin"),
    "/usr/local/bin/devin",
))
_AUTO_FIX_DEVIN_TIMEOUT_S = 300.0  # 5min cap


def _build_devin_prompt(req: "_AutoFixRequest") -> str:
    """Arma el prompt para pasar a `devin -p`. En una sola línea porque
    el CLI acepta el prompt como string positional arg."""
    parts = [
        "Error detectado en el stack obsidian-rag. Diagnosticá Y resolvé:",
    ]
    if req.service:
        parts.append(f"service={req.service}")
    if req.file:
        parts.append(f"archivo={req.file}")
    if req.timestamp:
        parts.append(f"ts={req.timestamp}")
    parts.append(f"error='{req.error_text[:500]}'")
    if req.context_lines:
        ctx = " | ".join(ln[:120] for ln in req.context_lines[-6:])
        parts.append(f"contexto_previo='{ctx[:600]}'")
    parts.append(
        "Investigá el problema (launchctl list, tail del log), identificá "
        "la causa, aplicá el fix (kickstart del daemon si aplica, o fix "
        "de código si requiere), verificá que quedó resuelto, y devolvé "
        "un resumen corto de qué hiciste. NO reinicies obsidian-rag-web "
        "(matarías el proceso que te invocó). Si no podés resolver solo "
        "y requiere decisión humana, explicalo en el resumen."
    )
    return " ".join(parts)


@app.post("/api/auto-fix-devin")
def auto_fix_devin(req: _AutoFixRequest, request: Request) -> StreamingResponse:
    """Delegar la resolución del error al agente Devin (CLI `devin -p`).

    Stream SSE con eventos:
      - {type: "start", cmd: ["devin", "-p", ...]}
      - {type: "output", chunk: "..."}      (stdout/stderr en vivo)
      - {type: "done", exit_code: N, output: "<full output>"}
      - {type: "error", message: "..."}
    """
    client_ip = (request.client.host if request.client else "unknown")
    _check_rate_limit(_CHAT_BUCKETS, client_ip,
                      _CHAT_RATE_LIMIT, _CHAT_RATE_WINDOW)

    if _DEVIN_BIN is None:
        raise HTTPException(
            status_code=503,
            detail=(
                "CLI `devin` no encontrado en el sistema. Instalá desde "
                "https://cli.devin.ai o configurá el path."
            ),
        )

    prompt = _build_devin_prompt(req)
    # Limitamos el prompt para no pasarle miles de chars a un arg de CLI.
    if len(prompt) > 3000:
        prompt = prompt[:3000] + "…(truncado)"

    cmd = [
        _DEVIN_BIN,
        "-p", prompt,
        "--respect-workspace-trust", "false",
        # --permission-mode dangerous: auto-aprobar todas las tool calls.
        # El user pidió explícitamente auto-fix 100% autónomo. Sin esto,
        # Devin al toparse con un edit/exec que requiera aprobación queda
        # esperando stdin (que tiene DEVNULL) → cuelgue indefinido.
        # Valores válidos en la CLI: "auto" (default, auto-aprueba read-only)
        # o "dangerous" (auto-aprueba todo). No existe "bypass" — verificado
        # con `devin --help` 2026-04-26.
        "--permission-mode", "dangerous",
    ]

    def _stream():
        import select  # local import — select solo se usa acá

        yield f"data: {json.dumps({'type': 'start', 'cmd': cmd})}\n\n"
        # Log de la invocación al audit.
        _audit_diagnose_execution({
            "ts": datetime.now().isoformat(timespec="seconds"),
            "command_original": f"devin -p '{prompt[:200]}…'",
            "command_executed": cmd[:2],  # no loguear todo el prompt (PII)
            "via": "auto-fix-devin",
            "prompt_len": len(prompt),
            "service": req.service,
            "error_preview": req.error_text[:200],
        })

        full_output_chunks: list[str] = []
        exit_code = -1
        try:
            # cwd = repo root para que Devin tenga contexto del código.
            # stdin=DEVNULL explícito — si devin pide confirmación interactiva
            # por algo no cubierto por --permission-mode dangerous, preferimos
            # que falle rápido antes que bloquearse esperando input.
            # TERM=dumb + NO_COLOR=1 para minimizar ANSI en el stream.
            proc = subprocess.Popen(
                cmd,
                cwd=str(ROOT),
                stdin=subprocess.DEVNULL,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,  # merge stderr al stdout para un solo stream
                bufsize=0,  # unbuffered — leemos raw bytes
                env={
                    **os.environ,  # Devin CLI necesita su env completo
                    "TERM": "dumb",
                    "NO_COLOR": "1",
                    "FORCE_COLOR": "0",
                },
            )

            # Lectura NON-BLOCKING con select+os.read.
            #
            # Antes usábamos readline() con bufsize=1, pero eso bloquea
            # esperando '\n'. Devin durante la fase "pensando" emite
            # spinners/dots/ANSI sin newline, entonces readline() se colgaba
            # y el cliente veía "Devin investigando…" eternamente hasta el
            # final. Ahora leemos bytes crudos a medida que llegan y los
            # pusheamos inmediatamente al SSE.
            #
            # Además emitimos 'heartbeat' cada 3s cuando no hay output, para
            # que el modal sepa que Devin sigue vivo y pueda mostrar un
            # contador (ej. "Devin investigando… (27s)").
            fd = proc.stdout.fileno()
            t0 = time.monotonic()
            last_heartbeat = 0
            while True:
                if proc.poll() is not None:
                    # Proceso terminó — drenar lo que quede del pipe.
                    try:
                        remaining = proc.stdout.read()
                        if isinstance(remaining, bytes):
                            remaining = remaining.decode("utf-8", errors="replace")
                    except Exception:
                        remaining = ""
                    if remaining:
                        full_output_chunks.append(remaining)
                        yield f"data: {json.dumps({'type': 'output', 'chunk': remaining})}\n\n"
                    break

                elapsed = time.monotonic() - t0
                if elapsed > _AUTO_FIX_DEVIN_TIMEOUT_S:
                    proc.terminate()
                    yield f"data: {json.dumps({'type': 'error', 'message': f'timeout después de {_AUTO_FIX_DEVIN_TIMEOUT_S}s'})}\n\n"
                    try:
                        proc.wait(timeout=5)
                    except subprocess.TimeoutExpired:
                        proc.kill()
                    return

                # Poll stdout por hasta 1s. Si hay datos, leer un chunk
                # (hasta 4KB) y emitirlo. Si no, chequear si toca heartbeat.
                ready, _, _ = select.select([fd], [], [], 1.0)
                if ready:
                    try:
                        raw = os.read(fd, 4096)
                    except (OSError, ValueError):
                        raw = b""
                    if raw:
                        chunk = raw.decode("utf-8", errors="replace")
                        full_output_chunks.append(chunk)
                        yield f"data: {json.dumps({'type': 'output', 'chunk': chunk})}\n\n"
                else:
                    # Sin datos nuevos: emitir heartbeat cada ~3s.
                    elapsed_int = int(elapsed)
                    if elapsed_int - last_heartbeat >= 3:
                        last_heartbeat = elapsed_int
                        yield f"data: {json.dumps({'type': 'heartbeat', 'elapsed_s': elapsed_int})}\n\n"

            exit_code = proc.returncode if proc.returncode is not None else -1
        except Exception as e:
            yield f"data: {json.dumps({'type': 'error', 'message': f'{type(e).__name__}: {e}'})}\n\n"
            return

        full_output = "".join(full_output_chunks)
        yield f"data: {json.dumps({'type': 'done', 'exit_code': exit_code, 'output': full_output[:16000]})}\n\n"

    return StreamingResponse(
        _stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache, no-transform",
            "X-Accel-Buffering": "no",
        },
    )


# ══════════════════════════════════════════════════════════════════════
# ── Auto-fix queue: DB + scanner + worker ──────────────────────────────
# ══════════════════════════════════════════════════════════════════════
# El user pidió: "meté los errores en una DB y dedespues que vaya
# agarrando uno por uno y los resuelva 100% automático".
#
# Diseño:
#   1. Schema: tabla rag_error_queue (sqlite-vec, state DB). Cada error
#      único (por signature hash) es una fila. Si se repite, incrementa
#      occurrence_count y last_seen_at.
#   2. Scanner: thread interno del web daemon que cada N minutos lee
#      todos los logs (reusa `_build_global_errors_payload`) y hace UPSERT.
#   3. Worker: OTRO thread interno que cada M minutos toma el siguiente
#      error con status=pending (ordenado por occurrence_count desc,
#      luego last_seen_at desc), marca processing, spawn `devin -p`,
#      parsea el STATUS de la respuesta, marca resolved/failed/etc.
#   4. Hard cap: max N invocaciones de Devin por hora para controlar
#      ACUs. Default conservador: 5/hora.
#
# Estados:
#   pending      — nuevo, esperando worker
#   processing   — Devin está trabajando
#   resolved     — Devin reportó STATUS: resolved
#   needs-human  — Devin reportó STATUS: needs-human (fix requiere decisión)
#   no-action    — Devin reportó STATUS: no-action (falso positivo / transient)
#   failed       — Devin crashó, exit != 0, o output sin STATUS
#   skipped      — el worker decidió no procesar (ya tuvo attempts >= 3)
#
# Idempotencia: si un error resuelto vuelve a aparecer, NO volvemos a
# procesar — el dedupe por signature lo mantiene en status=resolved
# (solo incrementa occurrence_count). Para "reopen" manualmente, el
# user hace UPDATE status='pending' o borra la row via CLI/DB tool.

_ERROR_QUEUE_DDL = (
    "CREATE TABLE IF NOT EXISTS rag_error_queue ("
    " id INTEGER PRIMARY KEY AUTOINCREMENT,"
    " detected_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,"
    " service TEXT NOT NULL,"
    " file_ref TEXT NOT NULL,"
    " error_signature TEXT NOT NULL UNIQUE,"
    " error_text TEXT NOT NULL,"
    " context_lines TEXT,"
    " error_ts TEXT,"
    " first_seen_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,"
    " last_seen_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,"
    " occurrence_count INTEGER NOT NULL DEFAULT 1,"
    " status TEXT NOT NULL DEFAULT 'pending',"
    " attempts INTEGER NOT NULL DEFAULT 0,"
    " started_at TEXT,"
    " completed_at TEXT,"
    " duration_s REAL,"
    " devin_exit_code INTEGER,"
    " devin_output TEXT,"
    " resolution_status TEXT,"
    " resolution_reason TEXT"
    ")"
)
_ERROR_QUEUE_INDEXES = (
    "CREATE INDEX IF NOT EXISTS ix_error_queue_status ON rag_error_queue(status)",
    "CREATE INDEX IF NOT EXISTS ix_error_queue_signature ON rag_error_queue(error_signature)",
    "CREATE INDEX IF NOT EXISTS ix_error_queue_service_ts "
    "ON rag_error_queue(service, last_seen_at DESC)",
)

_ERROR_QUEUE_DDL_DONE = False
_ERROR_QUEUE_DDL_LOCK = threading.Lock()


def _ensure_error_queue_table() -> None:
    """DDL idempotente del rag_error_queue. Corre al primer uso."""
    global _ERROR_QUEUE_DDL_DONE
    if _ERROR_QUEUE_DDL_DONE:
        return
    with _ERROR_QUEUE_DDL_LOCK:
        if _ERROR_QUEUE_DDL_DONE:
            return
        try:
            with _ragvec_state_conn() as conn:
                conn.execute(_ERROR_QUEUE_DDL)
                for idx in _ERROR_QUEUE_INDEXES:
                    conn.execute(idx)
                conn.commit()
            _ERROR_QUEUE_DDL_DONE = True
        except Exception as e:
            print(f"[error-queue] DDL failed: {type(e).__name__}: {e}",
                  file=sys.stderr, flush=True)


def _compute_error_signature(service: str, error_text: str) -> str:
    """Hash normalizado para dedupe. Errores "similares" (mismo exception
    type + mismo mensaje estructural) → mismo signature → se dedupean.

    Normalización:
      - Lowercase
      - Absolute paths (`/Users/...`, `/tmp/...`) → `PATH`
      - Filenames con extensión (`_index.md`, `2026-04.md`) → `FILE`
      - Números puros → `N` (para `line 123` = `line 456`)
      - Whitespace colapsado

    El orden importa: primero paths completos, después filenames,
    después números. Si hacés números primero, `2026-04.md` queda
    `N-N.md` que ya no matchea el regex de filename.
    """
    import hashlib
    t = error_text.lower()
    # 1. Paths absolutos primero.
    t = re.sub(r"/[\w/.-]+", "PATH", t)
    # 2. Filenames con extensión (incluso dentro de texto).
    t = re.sub(r"\b[\w.-]+\.(md|py|txt|log|json|jsonl|db|html|js|css|yml|yaml)\b",
               "FILE", t)
    # 3. Números sueltos.
    t = re.sub(r"\b\d+\b", "N", t)
    t = re.sub(r"\s+", " ", t).strip()
    t = t[:500]
    return hashlib.sha256(f"{service}:{t}".encode("utf-8")).hexdigest()[:16]


def _enqueue_error(
    service: str, file_ref: str, error_text: str,
    context_lines: list[str] | None = None, error_ts: str | None = None,
) -> tuple[int, bool]:
    """UPSERT del error. Retorna `(row_id, was_new)`.

    Si el signature ya existía:
      - Incrementa occurrence_count
      - Actualiza last_seen_at
      - NO toca status (errores ya resueltos que vuelven quedan resolved)
    Si es nuevo: inserta con status='pending'.
    """
    _ensure_error_queue_table()
    sig = _compute_error_signature(service, error_text)
    ctx_json = json.dumps(context_lines or [], ensure_ascii=False)[:2000]
    try:
        with _ragvec_state_conn() as conn:
            existing = conn.execute(
                "SELECT id FROM rag_error_queue WHERE error_signature = ?",
                (sig,),
            ).fetchone()
            if existing:
                conn.execute(
                    "UPDATE rag_error_queue SET "
                    "occurrence_count = occurrence_count + 1, "
                    "last_seen_at = CURRENT_TIMESTAMP "
                    "WHERE id = ?",
                    (existing[0],),
                )
                conn.commit()
                return existing[0], False
            cursor = conn.execute(
                "INSERT INTO rag_error_queue "
                "(service, file_ref, error_signature, error_text, "
                " context_lines, error_ts) VALUES (?, ?, ?, ?, ?, ?)",
                (service, file_ref, sig, error_text[:4000], ctx_json, error_ts),
            )
            conn.commit()
            return cursor.lastrowid, True
    except Exception as e:
        print(f"[error-queue] enqueue failed: {type(e).__name__}: {e}",
              file=sys.stderr, flush=True)
        return -1, False


# Rate limit del worker — max N invocaciones de Devin por hora.
# Cada invocación tarda ~60-120s y consume ACUs pagas, así que el cap
# es conservador. Ajustable con env var.
_AUTO_FIX_WORKER_HOURLY_CAP = int(os.environ.get("RAG_AUTO_FIX_HOURLY_CAP", "5"))
_AUTO_FIX_WORKER_INVOCATIONS: list[float] = []  # monotonic ts de invocaciones


def _worker_can_invoke_devin() -> tuple[bool, str]:
    """¿El worker puede disparar otra invocación de Devin ahora?"""
    now = time.monotonic()
    # Filtrar invocaciones de la última hora.
    one_hour_ago = now - 3600
    _AUTO_FIX_WORKER_INVOCATIONS[:] = [t for t in _AUTO_FIX_WORKER_INVOCATIONS if t > one_hour_ago]
    if len(_AUTO_FIX_WORKER_INVOCATIONS) >= _AUTO_FIX_WORKER_HOURLY_CAP:
        return False, (
            f"rate limit: {_AUTO_FIX_WORKER_HOURLY_CAP} invocaciones/hora alcanzado "
            f"(próxima slot en {int((_AUTO_FIX_WORKER_INVOCATIONS[0] + 3600 - now) / 60)}min)"
        )
    return True, ""


def _parse_devin_resolution_status(output: str) -> tuple[str, str]:
    """Extraer STATUS: y REASON: del output de Devin. Retorna
    `(status, reason)` donde status ∈ {resolved, no-action, needs-human, failed}.

    Si no encuentra un STATUS: marker, devuelve ('failed', 'no status marker').
    """
    m = re.search(r"STATUS:\s*(resolved|no-action|needs-human|failed)\b",
                  output, re.IGNORECASE)
    if not m:
        return "failed", "output sin marker STATUS:"
    status = m.group(1).lower()
    reason_match = re.search(r"REASON:\s*(.+?)(?:\n|$)", output)
    reason = reason_match.group(1).strip() if reason_match else ""
    return status, reason[:500]


def _get_next_pending_error() -> dict | None:
    """Siguiente error para procesar. Prioridad:
    1. occurrence_count DESC (errores más frecuentes primero)
    2. last_seen_at DESC (más recientes primero)
    3. attempts < 3 (si ya falló 3 veces, skip)
    """
    _ensure_error_queue_table()
    try:
        with _ragvec_state_conn() as conn:
            row = conn.execute(
                "SELECT id, service, file_ref, error_text, context_lines, "
                "error_ts, occurrence_count, attempts "
                "FROM rag_error_queue "
                "WHERE status = 'pending' AND attempts < 3 "
                "ORDER BY occurrence_count DESC, last_seen_at DESC "
                "LIMIT 1"
            ).fetchone()
            if not row:
                return None
            return {
                "id": row[0], "service": row[1], "file_ref": row[2],
                "error_text": row[3],
                "context_lines": json.loads(row[4]) if row[4] else [],
                "error_ts": row[5],
                "occurrence_count": row[6], "attempts": row[7],
            }
    except Exception as e:
        print(f"[error-queue] get_next failed: {type(e).__name__}: {e}",
              file=sys.stderr, flush=True)
        return None


def _process_error_with_devin(error_id: int) -> dict:
    """Spawn `devin -p` con el contexto del error en la DB y update la row.

    Retorna el resultado `{exit_code, output, resolution_status, ...}`.
    """
    _ensure_error_queue_table()
    # Fetch error row.
    try:
        with _ragvec_state_conn() as conn:
            row = conn.execute(
                "SELECT service, file_ref, error_text, context_lines, error_ts, attempts "
                "FROM rag_error_queue WHERE id = ?",
                (error_id,),
            ).fetchone()
    except Exception as e:
        return {"error": f"fetch row failed: {e}"}
    if not row:
        return {"error": "row no existe"}

    service, file_ref, error_text, context_json, error_ts, attempts = row
    context_lines = json.loads(context_json) if context_json else []

    # Marcar como processing + increment attempts.
    with _ragvec_state_conn() as conn:
        conn.execute(
            "UPDATE rag_error_queue SET status = 'processing', "
            "started_at = CURRENT_TIMESTAMP, attempts = attempts + 1 "
            "WHERE id = ?",
            (error_id,),
        )
        conn.commit()

    # Build prompt + ejecutar devin.
    req = _AutoFixRequest(
        error_text=error_text,
        service=service,
        file=file_ref,
        line_n=0,
        timestamp=error_ts,
        context_lines=context_lines,
    )
    prompt = _build_devin_prompt(req) + (
        "\n\nCuando termines, terminá la respuesta con una línea:\n"
        "STATUS: resolved | no-action | needs-human | failed\n"
        "REASON: <explicación breve>\n"
        "Esto es CRÍTICO para que el worker autónomo pueda clasificar "
        "tu respuesta."
    )
    if len(prompt) > 3000:
        prompt = prompt[:3000] + "…(truncado)"

    if _DEVIN_BIN is None:
        _finalize_error(error_id, -1, "", "failed", "devin CLI no encontrado", 0.0)
        _AUTO_FIX_WORKER_INVOCATIONS.append(time.monotonic())
        return {"error_id": error_id, "exit_code": -1, "resolution_status": "failed"}

    # Worker autónomo: mismas flags que el endpoint `/api/auto-fix-devin`.
    # --permission-mode dangerous es obligatorio acá — sin ello el subprocess
    # se cuelga indefinidamente cuando Devin intenta un edit/exec y pide
    # aprobación (no hay humano leyendo stdin).
    cmd = [
        _DEVIN_BIN, "-p", prompt,
        "--respect-workspace-trust", "false",
        "--permission-mode", "dangerous",
    ]
    _AUTO_FIX_WORKER_INVOCATIONS.append(time.monotonic())
    t0 = time.monotonic()
    try:
        proc = subprocess.run(
            cmd,
            cwd=str(ROOT),
            stdin=subprocess.DEVNULL,  # evitar bloqueo esperando input
            capture_output=True,
            text=True,
            timeout=_AUTO_FIX_DEVIN_TIMEOUT_S,
            shell=False,
            env={**os.environ, "TERM": "dumb", "NO_COLOR": "1"},
        )
        output = (proc.stdout or "") + ("\n" + proc.stderr if proc.stderr else "")
        exit_code = proc.returncode
        duration_s = round(time.monotonic() - t0, 3)
        if exit_code == 0:
            resolution_status, reason = _parse_devin_resolution_status(output)
        else:
            resolution_status, reason = "failed", f"exit {exit_code}"
        _finalize_error(error_id, exit_code, output, resolution_status, reason, duration_s)
        return {
            "error_id": error_id, "exit_code": exit_code,
            "resolution_status": resolution_status, "reason": reason,
            "duration_s": duration_s,
        }
    except subprocess.TimeoutExpired:
        duration_s = round(time.monotonic() - t0, 3)
        _finalize_error(error_id, 124, "", "failed",
                        f"timeout {_AUTO_FIX_DEVIN_TIMEOUT_S}s", duration_s)
        return {"error_id": error_id, "exit_code": 124, "resolution_status": "failed"}
    except Exception as e:
        duration_s = round(time.monotonic() - t0, 3)
        _finalize_error(error_id, -1, "", "failed",
                        f"{type(e).__name__}: {e}", duration_s)
        return {"error_id": error_id, "exit_code": -1, "resolution_status": "failed"}


def _finalize_error(
    error_id: int, exit_code: int, output: str,
    resolution_status: str, reason: str, duration_s: float,
) -> None:
    """Marcar la row como completada con resolution_status."""
    # resolution_status → row.status mapping.
    status_map = {
        "resolved": "resolved",
        "no-action": "no-action",
        "needs-human": "needs-human",
        "failed": "failed",
    }
    new_status = status_map.get(resolution_status, "failed")
    try:
        with _ragvec_state_conn() as conn:
            conn.execute(
                "UPDATE rag_error_queue SET "
                "status = ?, completed_at = CURRENT_TIMESTAMP, "
                "duration_s = ?, devin_exit_code = ?, "
                "devin_output = ?, resolution_status = ?, resolution_reason = ? "
                "WHERE id = ?",
                (new_status, duration_s, exit_code, output[:16000],
                 resolution_status, reason, error_id),
            )
            conn.commit()
    except Exception as e:
        print(f"[error-queue] finalize failed: {type(e).__name__}: {e}",
              file=sys.stderr, flush=True)


# ── Scanner + Worker threads ──────────────────────────────────────────
# Dos daemon threads — uno popla la queue, el otro la procesa. Ambos
# despiertan periódicamente y corren una iteración.

_SCANNER_PERIOD_S = 300.0  # 5min
_WORKER_PERIOD_S = 120.0   # 2min
_SCANNER_THREAD: "threading.Thread | None" = None
_WORKER_THREAD: "threading.Thread | None" = None
_ERROR_QUEUE_STOP = threading.Event()
# Flag para habilitar el worker automático. Default: off para que el user
# explícitamente lo active (evita sorpresas con ACUs). Se controla via
# env var o via /api/logs/queue/config.
_WORKER_AUTO_ENABLED = os.environ.get("RAG_AUTO_FIX_WORKER", "0") == "1"


def _scanner_loop() -> None:
    """Loop: cada _SCANNER_PERIOD_S, escanea logs + enqueue errores nuevos."""
    while not _ERROR_QUEUE_STOP.is_set():
        try:
            payload = _build_global_errors_payload(3600, "error")
            enqueued = 0
            seen = 0
            for ln in payload.get("lines", []):
                seen += 1
                _, was_new = _enqueue_error(
                    service=ln.get("service") or "unknown",
                    file_ref=ln.get("ref") or "",
                    error_text=ln.get("text") or "",
                    context_lines=[],  # Scanner no tiene contexto líneas; el feed global no lo provee
                    error_ts=ln.get("ts"),
                )
                if was_new:
                    enqueued += 1
            if enqueued > 0:
                print(f"[error-queue scanner] enqueued {enqueued} new errors "
                      f"(seen {seen})", flush=True)
        except Exception as e:
            print(f"[error-queue scanner] tick failed: {type(e).__name__}: {e}",
                  file=sys.stderr, flush=True)
        _ERROR_QUEUE_STOP.wait(timeout=_SCANNER_PERIOD_S)


def _worker_loop() -> None:
    """Loop: cada _WORKER_PERIOD_S, si hay pending + bajo el rate limit,
    procesa 1 error con Devin."""
    while not _ERROR_QUEUE_STOP.is_set():
        try:
            if not _WORKER_AUTO_ENABLED:
                _ERROR_QUEUE_STOP.wait(timeout=_WORKER_PERIOD_S)
                continue
            can, reason = _worker_can_invoke_devin()
            if not can:
                _ERROR_QUEUE_STOP.wait(timeout=_WORKER_PERIOD_S)
                continue
            next_error = _get_next_pending_error()
            if not next_error:
                _ERROR_QUEUE_STOP.wait(timeout=_WORKER_PERIOD_S)
                continue
            print(f"[error-queue worker] processing error id={next_error['id']} "
                  f"service={next_error['service']} "
                  f"count={next_error['occurrence_count']}", flush=True)
            result = _process_error_with_devin(next_error["id"])
            print(f"[error-queue worker] done id={next_error['id']} "
                  f"status={result.get('resolution_status')} "
                  f"duration={result.get('duration_s')}s", flush=True)
        except Exception as e:
            print(f"[error-queue worker] tick failed: {type(e).__name__}: {e}",
                  file=sys.stderr, flush=True)
        _ERROR_QUEUE_STOP.wait(timeout=_WORKER_PERIOD_S)


@_on_startup
def _start_error_queue_threads() -> None:
    """Arrancar scanner + worker threads al boot del web daemon."""
    global _SCANNER_THREAD, _WORKER_THREAD
    _ERROR_QUEUE_STOP.clear()
    _SCANNER_THREAD = threading.Thread(
        target=_scanner_loop, name="error-queue-scanner", daemon=True)
    _SCANNER_THREAD.start()
    _WORKER_THREAD = threading.Thread(
        target=_worker_loop, name="error-queue-worker", daemon=True)
    _WORKER_THREAD.start()


@_on_shutdown
def _stop_error_queue_threads() -> None:
    """Señalar a los threads que paren."""
    _ERROR_QUEUE_STOP.set()
    if _SCANNER_THREAD is not None:
        _SCANNER_THREAD.join(timeout=2.0)
    if _WORKER_THREAD is not None:
        _WORKER_THREAD.join(timeout=2.0)


# ── API endpoints ─────────────────────────────────────────────────────

@app.get("/api/logs/queue")
def logs_queue_list(status: str = "all", limit: int = 100) -> dict:
    """Listar entries del queue. Filtro por status opcional."""
    _ensure_error_queue_table()
    valid_statuses = {"all", "pending", "processing", "resolved",
                      "needs-human", "no-action", "failed", "skipped"}
    if status not in valid_statuses:
        raise HTTPException(status_code=400,
                            detail=f"status inválido: {status!r}")
    lim = max(1, min(int(limit), 500))
    try:
        with _ragvec_state_conn() as conn:
            if status == "all":
                rows = conn.execute(
                    "SELECT id, service, file_ref, error_text, status, "
                    "occurrence_count, attempts, first_seen_at, last_seen_at, "
                    "completed_at, duration_s, resolution_status, resolution_reason "
                    "FROM rag_error_queue "
                    "ORDER BY CASE status "
                    "  WHEN 'processing' THEN 0 "
                    "  WHEN 'pending' THEN 1 "
                    "  WHEN 'needs-human' THEN 2 "
                    "  WHEN 'failed' THEN 3 "
                    "  WHEN 'resolved' THEN 4 "
                    "  ELSE 5 END, "
                    "last_seen_at DESC "
                    "LIMIT ?",
                    (lim,),
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT id, service, file_ref, error_text, status, "
                    "occurrence_count, attempts, first_seen_at, last_seen_at, "
                    "completed_at, duration_s, resolution_status, resolution_reason "
                    "FROM rag_error_queue WHERE status = ? "
                    "ORDER BY last_seen_at DESC LIMIT ?",
                    (status, lim),
                ).fetchall()
            # Counts por status.
            counts = dict(conn.execute(
                "SELECT status, COUNT(*) FROM rag_error_queue GROUP BY status"
            ).fetchall())
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {e}") from e

    entries = [{
        "id": r[0], "service": r[1], "file_ref": r[2],
        "error_text": r[3], "status": r[4],
        "occurrence_count": r[5], "attempts": r[6],
        "first_seen_at": r[7], "last_seen_at": r[8],
        "completed_at": r[9], "duration_s": r[10],
        "resolution_status": r[11], "resolution_reason": r[12],
    } for r in rows]

    can_invoke, rate_reason = _worker_can_invoke_devin()
    return {
        "entries": entries,
        "counts_by_status": counts,
        "total": sum(counts.values()),
        "worker_enabled": _WORKER_AUTO_ENABLED,
        "worker_rate_limit": {
            "hourly_cap": _AUTO_FIX_WORKER_HOURLY_CAP,
            "current_hour_count": len(_AUTO_FIX_WORKER_INVOCATIONS),
            "can_invoke_now": can_invoke,
            "reason": rate_reason if not can_invoke else "",
        },
    }


@app.get("/api/logs/queue/{error_id}")
def logs_queue_get(error_id: int) -> dict:
    """Detalle de un error (incluye context_lines + devin_output completo)."""
    _ensure_error_queue_table()
    try:
        with _ragvec_state_conn() as conn:
            row = conn.execute(
                "SELECT id, service, file_ref, error_signature, error_text, "
                "context_lines, error_ts, first_seen_at, last_seen_at, "
                "occurrence_count, status, attempts, started_at, completed_at, "
                "duration_s, devin_exit_code, devin_output, "
                "resolution_status, resolution_reason "
                "FROM rag_error_queue WHERE id = ?",
                (error_id,),
            ).fetchone()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {e}") from e
    if not row:
        raise HTTPException(status_code=404, detail="error no encontrado")
    return {
        "id": row[0], "service": row[1], "file_ref": row[2],
        "error_signature": row[3], "error_text": row[4],
        "context_lines": json.loads(row[5]) if row[5] else [],
        "error_ts": row[6], "first_seen_at": row[7], "last_seen_at": row[8],
        "occurrence_count": row[9], "status": row[10], "attempts": row[11],
        "started_at": row[12], "completed_at": row[13],
        "duration_s": row[14], "devin_exit_code": row[15],
        "devin_output": row[16], "resolution_status": row[17],
        "resolution_reason": row[18],
    }


@app.post("/api/logs/queue/process-next")
def logs_queue_process_next() -> dict:
    """Procesar manualmente el siguiente pending. Útil para testing y
    para cuando el worker está desactivado."""
    can, reason = _worker_can_invoke_devin()
    if not can:
        raise HTTPException(status_code=429, detail=reason)
    next_error = _get_next_pending_error()
    if not next_error:
        return {"status": "no-pending", "message": "no hay errores pending"}
    result = _process_error_with_devin(next_error["id"])
    return {"status": "processed", **result}


@app.post("/api/logs/queue/scan-now")
def logs_queue_scan_now() -> dict:
    """Forzar un scan manual de los logs + enqueue. Útil después de que
    un daemon empezó a fallar y querés ver los errores en la queue ya."""
    try:
        payload = _build_global_errors_payload(3600, "error")
        enqueued = 0
        for ln in payload.get("lines", []):
            _, was_new = _enqueue_error(
                service=ln.get("service") or "unknown",
                file_ref=ln.get("ref") or "",
                error_text=ln.get("text") or "",
                context_lines=[], error_ts=ln.get("ts"),
            )
            if was_new:
                enqueued += 1
        return {
            "status": "ok", "scanned": len(payload.get("lines", [])),
            "new_entries": enqueued,
        }
    except Exception as e:
        raise HTTPException(status_code=500,
                            detail=f"{type(e).__name__}: {e}") from e


class _WorkerConfigRequest(BaseModel):
    enabled: bool = Field(..., description="Habilitar o no el worker auto")
    hourly_cap: int | None = Field(None, description="Cap de invocaciones/hora")


@app.post("/api/logs/queue/config")
def logs_queue_config(req: _WorkerConfigRequest) -> dict:
    """Toggle del worker automático + ajuste del rate limit."""
    global _WORKER_AUTO_ENABLED, _AUTO_FIX_WORKER_HOURLY_CAP
    _WORKER_AUTO_ENABLED = bool(req.enabled)
    if req.hourly_cap is not None:
        _AUTO_FIX_WORKER_HOURLY_CAP = max(0, min(int(req.hourly_cap), 50))
    return {
        "worker_enabled": _WORKER_AUTO_ENABLED,
        "hourly_cap": _AUTO_FIX_WORKER_HOURLY_CAP,
    }


@app.delete("/api/logs/queue/{error_id}")
def logs_queue_delete(error_id: int) -> dict:
    """Borrar un entry del queue. Útil para limpiar falsos positivos."""
    _ensure_error_queue_table()
    try:
        with _ragvec_state_conn() as conn:
            cursor = conn.execute(
                "DELETE FROM rag_error_queue WHERE id = ?", (error_id,)
            )
            conn.commit()
            if cursor.rowcount == 0:
                raise HTTPException(status_code=404, detail="error no encontrado")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500,
                            detail=f"{type(e).__name__}: {e}") from e
    return {"status": "deleted", "id": error_id}


@app.get("/logs")
def logs_page() -> FileResponse:
    return FileResponse(STATIC_DIR / "logs.html")


def _collect_screentime_daily(
    st_start: datetime, st_end: datetime, st_days: int,
) -> list[dict]:
    """Daily screentime buckets via a single SQL scan of knowledgeC.db.

    Replaces N separate `_collect_screentime` calls (one per day). Groups
    /app/usage sessions by local date in SQL so we touch the db once.
    """
    import sqlite3
    from rag import SCREENTIME_DB, _SCREENTIME_COCOA_OFFSET  # noqa: PLC0415

    path = SCREENTIME_DB
    daily: list[dict] = []
    if not path.is_file():
        return daily
    start_ts = st_start.timestamp() - _SCREENTIME_COCOA_OFFSET
    end_ts = st_end.timestamp() - _SCREENTIME_COCOA_OFFSET
    try:
        uri = f"file:{path}?mode=ro&immutable=1"
        conn = sqlite3.connect(uri, uri=True, timeout=2.0)
        try:
            rows = conn.execute(
                """
                SELECT date(ZSTARTDATE + ?, 'unixepoch', 'localtime') AS day,
                       SUM(ZENDDATE - ZSTARTDATE) AS secs
                FROM ZOBJECT
                WHERE ZSTREAMNAME = '/app/usage'
                  AND ZSTARTDATE >= ?
                  AND ZSTARTDATE < ?
                  AND (ZENDDATE - ZSTARTDATE) >= 5
                GROUP BY day
                ORDER BY day
                """,
                (_SCREENTIME_COCOA_OFFSET, start_ts, end_ts),
            ).fetchall()
        finally:
            conn.close()
    except Exception:
        return daily
    by_day = {day: int(secs or 0) for day, secs in rows if day}
    for d in range(st_days):
        day_start = (st_end - timedelta(days=st_days - 1 - d)).replace(
            hour=0, minute=0, second=0, microsecond=0,
        )
        key = day_start.strftime("%Y-%m-%d")
        daily.append({"day": key, "secs": by_day.get(key, 0)})
    return daily


# TTL cache for /api/dashboard. Polling clients hit this every minute;
# aggregating 7 JSONL files + 8 sqlite queries per hit is wasteful when
# the underlying logs barely move between polls. SSE stream pushes live
# deltas, so a 30s cache is invisible to the user.
_DASHBOARD_CACHE: dict[int, tuple[float, dict]] = {}
_DASHBOARD_TTL = 30.0


@app.get("/api/dashboard")
def dashboard_api(days: int = 30) -> dict:
    now_ts = time.time()
    hit = _DASHBOARD_CACHE.get(days)
    if hit and now_ts - hit[0] < _DASHBOARD_TTL:
        return hit[1]
    payload = _dashboard_compute(days)
    _DASHBOARD_CACHE[days] = (now_ts, payload)
    return payload


@app.get("/api/status/home")
def status_home() -> dict:
    """Health del home page compute — surface el threshold actual,
    los últimos N totales (para sparkline en el status page) y un
    conteo de degraded events en las últimas 24h.

    Diseño: SOLO lee `rag_home_compute_metrics` (sin cómputos pesados,
    sub-50ms warm). El threshold es el mismo que usa el SSE generator,
    derivado de la rolling window in-memory. Si la window está fría
    (post-restart sin hidratar todavía), devuelve `floor` con un flag
    `cold=true` para que la UI pinte un disclaimer apropiado.
    """
    from rag import _ragvec_state_conn, _sql_read_with_retry

    def _do() -> dict:
        with _ragvec_state_conn() as conn:
            recent = conn.execute(
                "SELECT ts, elapsed_s, regenerate, degraded, degraded_cause "
                "FROM rag_home_compute_metrics "
                "ORDER BY ts DESC LIMIT 50"
            ).fetchall()
            cutoff = (datetime.now() - timedelta(hours=24)).isoformat(timespec="seconds")
            counts = conn.execute(
                "SELECT "
                " COUNT(*) AS total, "
                " COALESCE(SUM(degraded), 0) AS degraded_n, "
                " COALESCE(SUM(regenerate), 0) AS regen_n "
                "FROM rag_home_compute_metrics WHERE ts >= ?",
                (cutoff,),
            ).fetchone()
            cause_breakdown = conn.execute(
                "SELECT degraded_cause, COUNT(*) AS n "
                "FROM rag_home_compute_metrics "
                "WHERE ts >= ? AND degraded = 1 AND degraded_cause IS NOT NULL "
                "GROUP BY degraded_cause ORDER BY n DESC",
                (cutoff,),
            ).fetchall()
        return {
            "samples": [
                {
                    "ts": r[0],
                    "elapsed_s": float(r[1]),
                    "regenerate": bool(r[2]),
                    "degraded": bool(r[3]),
                    "degraded_cause": r[4],
                }
                for r in recent
            ],
            "window_24h": {
                "total": int(counts[0] or 0),
                "degraded": int(counts[1] or 0),
                "regenerate": int(counts[2] or 0),
                "by_cause": [
                    {"cause": r[0], "count": int(r[1])}
                    for r in cause_breakdown
                ],
            },
        }

    payload = _sql_read_with_retry(_do, "status_home_sql_read_failed", default=None)
    if payload is None:
        payload = {
            "samples": [],
            "window_24h": {"total": 0, "degraded": 0, "regenerate": 0,
                           "by_cause": []},
        }

    with _HOME_COMPUTE_HISTORY_LOCK:
        history_size = len(_HOME_COMPUTE_HISTORY)
    payload["threshold_s"] = round(_home_compute_degraded_threshold(), 2)
    payload["floor_s"] = _HOME_COMPUTE_DEGRADED_FLOOR
    payload["cold"] = history_size < 3
    payload["history_size"] = history_size
    return payload


def _dashboard_compute(days: int = 30) -> dict:
    """Aggregate telemetry into dashboard metrics.

    SQL-only since T10. On any exception, logs to sql_state_errors.jsonl and
    returns an empty-shape payload so the dashboard JS can still render
    (the aggregator handles empty lists uniformly).
    """
    try:
        return _dashboard_compute_sql(days)
    except Exception as exc:
        _log_sql_state_error("dashboard_sql_compute_failed", err=repr(exc))
        payload = _dashboard_aggregate(
            queries=[], all_queries=[], fb_entries=[], ambient_entries=[],
            contra_entries=[], filing_recent=[], tune_entries=[],
            surface_entries=[], days=days,
        )
        # Shape parity: always emit `signals` key so the dashboard JS
        # doesn't branch between "normal" and "fallback" payloads.
        payload.setdefault("signals", {"counts": {}, "by_source": {}, "window_days": days})
        return payload


# ── Dashboard event-row helpers ─────────────────────────────────────────────
# The SQL and JSONL readers converge on a common "event dict" shape (matches
# the JSONL line schema) before the aggregator runs, so only the data source
# differs. `_dashboard_aggregate()` is the single source of truth for every
# field the dashboard JS consumes.


def _dashboard_sql_row_to_event(row, json_cols: tuple[str, ...] = ()) -> dict:
    """Rehydrate a sqlite3.Row into the JSONL-style event dict.

    - `*_json` columns whose base name (`_json` suffix stripped) is listed in
      `json_cols` are decoded and re-keyed to their JSONL equivalent (e.g.
      `paths_json` → `paths`).
    - `extra_json` is decoded and its keys merged into the top level (never
      overwriting an explicit column).
    - Other `*_json` columns are decoded in place if `json_cols` doesn't map
      them — safe default so aggregate callers never see raw JSON strings.
    """
    ev: dict = {}
    try:
        keys = row.keys()
    except Exception:
        return ev
    for k in keys:
        v = row[k]
        if v is None:
            continue
        if k == "extra_json":
            try:
                extra = json.loads(v)
                if isinstance(extra, dict):
                    for ek, ev_val in extra.items():
                        ev.setdefault(ek, ev_val)
            except Exception:
                pass
            continue
        if k.endswith("_json"):
            base = k[:-5]
            try:
                decoded = json.loads(v)
            except Exception:
                decoded = None
            target = base if base in json_cols else k
            if decoded is not None:
                ev[target] = decoded
            continue
        ev[k] = v
    return ev


def _dashboard_compute_sql(days: int = 30) -> dict:
    """SQL-path reader: fetch rag_* rows since cutoff, normalise to JSONL-shape
    event dicts, then call the shared aggregator.

    Tables read: rag_queries, rag_feedback, rag_ambient, rag_contradictions,
    rag_filing_log, rag_tune, rag_surface_log. `all_queries` is the unwindowed
    snapshot of rag_queries (used only for the total_queries_all_time KPI).
    """
    from rag import _ragvec_state_conn, _sql_query_window  # local import keeps module-load light
    cutoff = datetime.now() - timedelta(days=days)
    cutoff_iso = cutoff.isoformat(timespec="seconds")

    # Well-earlier floor — we still need "all time" count + full feedback for
    # corrective-miss/neg-reason surfacing (JSONL path uses the full file).
    ancient_iso = "0000-01-01T00:00:00"

    with _ragvec_state_conn() as conn:
        q_rows_window = _sql_query_window(conn, "rag_queries", cutoff_iso)
        # all_queries: full history — same semantics as JSONL (unbounded scan).
        q_rows_all = _sql_query_window(conn, "rag_queries", ancient_iso)
        fb_rows_all = _sql_query_window(conn, "rag_feedback", ancient_iso)
        amb_rows = _sql_query_window(conn, "rag_ambient", cutoff_iso)
        contra_rows = _sql_query_window(conn, "rag_contradictions", cutoff_iso)
        filing_rows = _sql_query_window(conn, "rag_filing_log", cutoff_iso)
        tune_rows = _sql_query_window(conn, "rag_tune", ancient_iso)
        surface_rows = _sql_query_window(conn, "rag_surface_log", ancient_iso)
        # Implicit signals KPI (2026-04-22): count rag_behavior events
        # grouped by type in the dashboard window. Feeds the "signals"
        # panel so the user can see at a glance whether the ranker-vivo
        # is being fed (opens, copies, saves, corrective-path
        # selections, etc.). Keyed on `event`; source breakdown
        # (web/cli) is a secondary dimension.
        behavior_signal_counts: dict[str, int] = {}
        behavior_signal_by_source: dict[str, dict[str, int]] = {}
        try:
            # Two grouped SELECTs — cheap even over 30 days of events;
            # rag_behavior has an ix_rag_behavior_event_ts index for this
            # exact access pattern.
            for ev_type, n in conn.execute(
                "SELECT event, COUNT(*) FROM rag_behavior "
                "WHERE ts >= ? GROUP BY event",
                (cutoff_iso,),
            ).fetchall():
                if ev_type:
                    behavior_signal_counts[str(ev_type)] = int(n)
            for ev_type, src, n in conn.execute(
                "SELECT event, source, COUNT(*) FROM rag_behavior "
                "WHERE ts >= ? GROUP BY event, source",
                (cutoff_iso,),
            ).fetchall():
                if ev_type and src:
                    behavior_signal_by_source.setdefault(
                        str(ev_type), {}
                    )[str(src)] = int(n)
        except Exception as exc:
            # Silently degrade — the dashboard shouldn't 500 because of
            # a missing column or a schema mismatch from an older DB.
            _log_sql_state_error(
                "dashboard_signals_sql_read_failed", err=repr(exc),
            )

    _Q_JSON = ("variants", "paths", "scores", "filters", "bad_citations")
    _F_JSON = ("paths",)
    _C_JSON = ("contradicts",)
    _FIL_JSON = ("neighbors",)
    _T_JSON = ("baseline", "best")

    all_queries = [_dashboard_sql_row_to_event(r, _Q_JSON) for r in q_rows_all]
    queries = [_dashboard_sql_row_to_event(r, _Q_JSON) for r in q_rows_window]
    fb_entries = [_dashboard_sql_row_to_event(r, _F_JSON) for r in fb_rows_all]
    # rag_ambient stores extra fields in `payload_json`, not `extra_json`.
    ambient_entries: list[dict] = []
    for r in amb_rows:
        ev = {}
        try:
            keys = r.keys()
        except Exception:
            keys = []
        for k in keys:
            v = r[k]
            if v is None:
                continue
            if k == "payload_json":
                try:
                    payload = json.loads(v)
                    if isinstance(payload, dict):
                        for pk, pv in payload.items():
                            ev.setdefault(pk, pv)
                except Exception:
                    pass
                continue
            ev[k] = v
        ambient_entries.append(ev)
    contra_entries = [_dashboard_sql_row_to_event(r, _C_JSON) for r in contra_rows]
    # Flatten `singles`/`chains` back onto each tune entry so the JSONL aggregator
    # sees `baseline`/`best` dicts as-is.
    tune_entries = [_dashboard_sql_row_to_event(r, _T_JSON) for r in tune_rows]
    # Filing rows — add an implicit `cmd` marker so the aggregator counts them
    # as filing events (it treats every entry in filing_recent as a filing).
    filing_recent = [_dashboard_sql_row_to_event(r, _FIL_JSON) for r in filing_rows]
    # Surface rows — aggregator counts events by `cmd in {"surface_pair","surface_run"}`,
    # and rag_surface_log preserves `cmd` as a column.
    surface_entries = [_dashboard_sql_row_to_event(r) for r in surface_rows]

    payload = _dashboard_aggregate(
        queries=queries,
        all_queries=all_queries,
        fb_entries=fb_entries,
        ambient_entries=ambient_entries,
        contra_entries=contra_entries,
        filing_recent=filing_recent,
        tune_entries=tune_entries,
        surface_entries=surface_entries,
        days=days,
    )
    # Signals panel — additive, non-breaking. Dashboard.js reads
    # payload.signals optionally; old clients just ignore it.
    payload["signals"] = {
        "counts": behavior_signal_counts,
        "by_source": behavior_signal_by_source,
        "window_days": days,
    }
    return payload


def _dashboard_aggregate(
    *,
    queries: list[dict],
    all_queries: list[dict],
    fb_entries: list[dict],
    ambient_entries: list[dict],
    contra_entries: list[dict],
    filing_recent: list[dict],
    tune_entries: list[dict],
    surface_entries: list[dict],
    days: int,
) -> dict:
    """Build the full dashboard payload from normalised event dicts.

    This is the shared body both the JSONL and SQL paths call with identical
    semantics. The dashboard JS key-contract (`queries_per_day`, `hours`,
    `cmds`, `sources`, `score_distribution`, `hot_topics`, `chat_keywords`,
    `feedback`, etc.) is fixed here — don't drop keys without updating
    web/static/dashboard.js.
    """
    import statistics
    from collections import defaultdict

    data_dir = Path.home() / ".local/share/obsidian-rag"
    cutoff = datetime.now() - timedelta(days=days)

    def _ts(e: dict) -> datetime | None:
        raw = e.get("ts") or e.get("timestamp")
        if not raw:
            return None
        try:
            return datetime.fromisoformat(raw)
        except Exception:
            return None

    def _pct(values: list[float], p: float) -> float:
        if not values:
            return 0.0
        s = sorted(values)
        idx = int(len(s) * p / 100)
        return round(s[min(idx, len(s) - 1)], 3)

    # ── Queries ──────────────────────────────────────────────────────
    scores = [float(e["top_score"]) for e in queries if isinstance(e.get("top_score"), (int, float))]
    t_retrieves = [float(e["t_retrieve"]) for e in queries if isinstance(e.get("t_retrieve"), (int, float))]
    t_gens = [float(e["t_gen"]) for e in queries if isinstance(e.get("t_gen"), (int, float)) and e["t_gen"] > 0]
    t_totals = [
        float(e["t_retrieve"]) + float(e["t_gen"])
        for e in queries
        if isinstance(e.get("t_retrieve"), (int, float)) and isinstance(e.get("t_gen"), (int, float)) and e["t_gen"] > 0
    ]

    # Per-day counts
    queries_per_day: dict[str, int] = defaultdict(int)
    latency_per_day: dict[str, list[float]] = defaultdict(list)
    for e in queries:
        t = _ts(e)
        if t:
            day = t.strftime("%Y-%m-%d")
            queries_per_day[day] += 1
            tr = e.get("t_retrieve")
            tg = e.get("t_gen")
            if isinstance(tr, (int, float)) and isinstance(tg, (int, float)) and tg > 0:
                latency_per_day[day].append(float(tr) + float(tg))

    # Hourly heatmap
    hours = defaultdict(int)
    for e in queries:
        t = _ts(e)
        if t:
            hours[t.hour] += 1

    # Command breakdown
    cmds = defaultdict(int)
    for e in queries:
        cmds[e.get("cmd", "?")] += 1

    # Source breakdown. `cmd: serve*` is exclusively the rag-serve fast path
    # (only the WhatsApp listener uses it), so any serve event is whatsapp
    # regardless of session prefix.
    sources = defaultdict(int)
    for e in queries:
        sid = e.get("session") or ""
        cmd = e.get("cmd") or ""
        if sid.startswith("wa:") or cmd.startswith("serve"):
            sources["whatsapp"] += 1
        elif sid.startswith("web:"):
            sources["web"] += 1
        elif sid.startswith("tg:"):
            sources["legacy"] += 1
        else:
            sources["cli"] += 1

    # Score distribution (10 buckets)
    score_buckets = [0] * 10
    for s in scores:
        idx = min(int(max(0.0, min(s, 1.0)) * 10), 9)
        score_buckets[idx] += 1

    # Hot topics
    topics: dict[str, int] = defaultdict(int)
    for e in queries:
        q = e.get("q", "")
        if q:
            words = q.lower().split()[:3]
            topics[" ".join(words)] += 1
    hot_topics = sorted(topics.items(), key=lambda x: -x[1])[:15]

    # Chat keyword cloud — per-word frequency over user-initiated queries
    # (chat/web/whatsapp), not ambient/eval synthetic traffic. Tokenises
    # on word boundaries, strips accents for collation (so "qué" and "que"
    # collapse), drops stopwords + words shorter than 3 chars. Returns the
    # top 60 for the UI cloud; the renderer sizes by log(count).
    _ES_EN_STOPWORDS = {
        "que", "de", "la", "el", "en", "y", "a", "los", "las", "un", "una",
        "del", "se", "por", "con", "para", "mi", "mis", "me", "tu", "tus",
        "su", "sus", "al", "lo", "le", "les", "es", "esta", "este", "esto",
        "esa", "ese", "eso", "como", "cuando", "donde", "quien", "cual",
        "cuales", "sobre", "entre", "desde", "hasta", "sin", "pero", "o",
        "u", "ni", "mas", "muy", "ya", "si", "no", "tambien", "hay", "fue",
        "son", "soy", "ser", "estoy", "estan", "estar", "he", "has", "ha",
        "hemos", "han", "puedo", "puede", "pueden", "podria", "podrian",
        "the", "a", "an", "is", "are", "was", "were", "be", "been", "being",
        "do", "does", "did", "have", "has", "had", "and", "or", "but",
        "if", "then", "else", "when", "where", "what", "which", "who",
        "how", "why", "to", "of", "in", "on", "for", "with", "from",
        "by", "as", "at", "into", "out", "up", "down", "off", "over",
        "this", "that", "these", "those", "it", "its", "i", "you", "he",
        "she", "we", "they", "me", "my", "your", "our", "their",
        "algo", "nada", "todo", "todos", "otra", "otro", "otros", "otras",
        "solo", "sólo", "tan", "tanto", "tener", "tiene", "tienes",
        "hacer", "hace", "haces", "hice", "hicimos", "voy", "vas", "va",
        "vamos", "van", "dame", "darme", "tengo", "tenes", "tenés",
        "estas", "estos", "esas", "esos", "cuantos", "cuantas",
        "hola", "gracias", "porfa", "plis", "oka", "dale", "listo",
        "quiero", "quieres", "quisiera", "necesito", "necesitas",
        "decime", "dime", "decir", "dijo", "dije", "saber", "sabes",
    }
    _KEYWORD_RE = re.compile(r"[A-Za-zÁÉÍÓÚÑáéíóúñü0-9][A-Za-zÁÉÍÓÚÑáéíóúñü0-9_\-']{2,}")
    def _fold(s: str) -> str:
        import unicodedata
        nfd = unicodedata.normalize("NFD", s.lower())
        return "".join(ch for ch in nfd if not unicodedata.combining(ch))
    keyword_counts: dict[str, int] = defaultdict(int)
    for e in queries:
        sid = e.get("session") or ""
        cmd = e.get("cmd") or ""
        # User-initiated surfaces only — exclude eval/ambient/morning/etc.
        if not (
            sid.startswith(("wa:", "web:", "tg:"))
            or cmd in ("query", "chat", "web", "serve", "serve-web")
        ):
            continue
        q = (e.get("q") or "").strip()
        if not q:
            continue
        for match in _KEYWORD_RE.findall(q):
            folded = _fold(match)
            if len(folded) < 3 or folded in _ES_EN_STOPWORDS:
                continue
            keyword_counts[folded] += 1
    chat_keywords = [
        {"word": w, "count": c}
        for w, c in sorted(keyword_counts.items(), key=lambda x: -x[1])[:5]
        if c >= 2  # skip hapax — noise dominates
    ]

    # Latency time series (daily p50)
    latency_ts = {}
    for day, vals in sorted(latency_per_day.items()):
        latency_ts[day] = {
            "p50": round(statistics.median(vals), 2) if vals else 0,
            "p95": _pct(vals, 95),
            "count": len(vals),
        }

    # ── Health metrics ───────────────────────────────────────────────
    # Only count queries that actually go through retrieval (have scores)
    retrieval_queries = [e for e in queries if isinstance(e.get("top_score"), (int, float)) and e.get("cmd") in ("query", "chat", "web", None)]
    n_retrieval = len(retrieval_queries)

    gated = sum(1 for e in queries if e.get("gated_low_confidence"))
    gate_rate = round(gated / n_retrieval * 100, 1) if n_retrieval else 0

    bad_citation_queries = sum(1 for e in queries if e.get("bad_citations"))
    bad_citation_rate = round(bad_citation_queries / n_retrieval * 100, 1) if n_retrieval else 0
    bad_citation_total = sum(len(e.get("bad_citations", [])) for e in queries)

    # Score quality bands
    retrieval_scores = [float(e["top_score"]) for e in retrieval_queries if isinstance(e.get("top_score"), (int, float))]
    n_scored = len(retrieval_scores)
    score_high = sum(1 for s in retrieval_scores if s >= 0.3)     # strong match
    score_mid = sum(1 for s in retrieval_scores if 0.05 <= s < 0.3)  # acceptable
    score_low = sum(1 for s in retrieval_scores if s < 0.05)      # weak/miss

    # Score trend per day (daily mean)
    score_per_day: dict[str, list[float]] = defaultdict(list)
    for e in retrieval_queries:
        t = _ts(e)
        s = e.get("top_score")
        if t and isinstance(s, (int, float)):
            score_per_day[t.strftime("%Y-%m-%d")].append(float(s))
    score_trend = {
        day: round(statistics.mean(vals), 3)
        for day, vals in sorted(score_per_day.items())
    }

    # Latency ratio (retrieve vs generate dominance)
    avg_ret = statistics.mean(t_retrieves) if t_retrieves else 0
    avg_gen = statistics.mean(t_gens) if t_gens else 0
    retrieve_pct = round(avg_ret / (avg_ret + avg_gen) * 100, 1) if (avg_ret + avg_gen) > 0 else 0

    # Answer length trend (is the LLM being verbose or terse?)
    answer_lens = [e["answer_len"] for e in queries if isinstance(e.get("answer_len"), (int, float)) and e["answer_len"] > 0]
    avg_answer_len = round(statistics.mean(answer_lens)) if answer_lens else 0

    health = {
        "retrieval_queries": n_retrieval,
        "gate_rate": gate_rate,
        "gated_count": gated,
        "bad_citation_rate": bad_citation_rate,
        "bad_citation_queries": bad_citation_queries,
        "bad_citation_total": bad_citation_total,
        "score_high": score_high,
        "score_mid": score_mid,
        "score_low": score_low,
        "score_high_pct": round(score_high / n_scored * 100, 1) if n_scored else 0,
        "score_mid_pct": round(score_mid / n_scored * 100, 1) if n_scored else 0,
        "score_low_pct": round(score_low / n_scored * 100, 1) if n_scored else 0,
        "score_trend": score_trend,
        "retrieve_pct": retrieve_pct,
        "generate_pct": round(100 - retrieve_pct, 1),
        "avg_answer_len": avg_answer_len,
    }

    # ── Feedback ─────────────────────────────────────────────────────
    # fb_entries is the full-history feedback log (SQL or JSONL). Peer
    # work on master added a windowed view for the KPI widget + all-time
    # companions — keep both so the dashboard can show "last N days" and
    # "all time" side-by-side.
    fb_recent = [e for e in fb_entries if (t := _ts(e)) and t >= cutoff]
    fb_pos = sum(1 for e in fb_recent if e.get("rating") == 1)
    fb_neg = sum(1 for e in fb_recent if e.get("rating") == -1)
    fb_pos_all = sum(1 for e in fb_entries if e.get("rating") == 1)
    fb_neg_all = sum(1 for e in fb_entries if e.get("rating") == -1)

    neg_path_counts: dict[str, int] = defaultdict(int)
    pos_path_counts: dict[str, int] = defaultdict(int)
    for e in fb_entries:
        rating = e.get("rating")
        for p in e.get("paths") or []:
            if rating == -1:
                neg_path_counts[p] += 1
            elif rating == 1:
                pos_path_counts[p] += 1

    # Corrective paths: user-supplied "the right note was here" — pure retrieval misses,
    # i.e. golden-eval candidates. Most actionable signal in the whole pipeline.
    corrective_misses: list[dict] = []
    for e in fb_entries:
        cp = e.get("corrective_path")
        if not cp:
            continue
        if cp in (e.get("paths") or []):
            continue
        corrective_misses.append({
            "ts": e.get("ts", ""),
            "q": (e.get("q") or "")[:140],
            "missing_path": cp,
            "retrieved": list((e.get("paths") or [])[:3]),
        })
    corrective_misses.sort(key=lambda x: x["ts"], reverse=True)

    # Recent negative reasons — qualitative why-it-failed signal.
    neg_reasons: list[dict] = []
    for e in fb_entries:
        if e.get("rating") != -1 or not e.get("reason"):
            continue
        neg_reasons.append({
            "ts": e.get("ts", ""),
            "q": (e.get("q") or "")[:140],
            "reason": e["reason"][:200],
            "paths": list((e.get("paths") or [])[:2]),
        })
    neg_reasons.sort(key=lambda x: x["ts"], reverse=True)

    # Calibration mismatch: cross feedback w/ queries by turn_id.
    # - false_confident = retrieved with high score but user said -1
    # - false_gated     = gate refused (low score) but user later said +1 / supplied corrective_path
    queries_by_turn = {e.get("turn_id"): e for e in queries if e.get("turn_id")}
    false_confident = 0
    false_gated = 0
    for fb in fb_entries:
        q_ev = queries_by_turn.get(fb.get("turn_id"))
        if not q_ev:
            continue
        ts = q_ev.get("top_score")
        if not isinstance(ts, (int, float)):
            continue
        if fb.get("rating") == -1 and ts >= 0.20:
            false_confident += 1
        if fb.get("rating") == 1 and ts < 0.05:
            false_gated += 1
        if fb.get("corrective_path") and ts < 0.05:
            false_gated += 1

    # Feedback per day (positive vs negative trend)
    fb_pos_per_day: dict[str, int] = defaultdict(int)
    fb_neg_per_day: dict[str, int] = defaultdict(int)
    for e in fb_recent:
        t = _ts(e)
        if not t:
            continue
        day = t.strftime("%Y-%m-%d")
        if e.get("rating") == 1:
            fb_pos_per_day[day] += 1
        elif e.get("rating") == -1:
            fb_neg_per_day[day] += 1

    feedback_actionable = {
        "total": len(fb_entries),
        "recent_pos": sum(1 for e in fb_recent if e.get("rating") == 1),
        "recent_neg": sum(1 for e in fb_recent if e.get("rating") == -1),
        "net_satisfaction": (
            round((fb_pos - fb_neg) / (fb_pos + fb_neg) * 100, 1)
            if (fb_pos + fb_neg) > 0 else None
        ),
        "top_negative_paths": [
            {"path": p, "count": c, "pos_count": pos_path_counts.get(p, 0)}
            for p, c in sorted(neg_path_counts.items(), key=lambda x: -x[1])[:5]
        ],
        "top_positive_paths": [
            {"path": p, "count": c}
            for p, c in sorted(pos_path_counts.items(), key=lambda x: -x[1])[:5]
        ],
        "corrective_misses": corrective_misses[:5],
        "n_corrective_misses": len(corrective_misses),
        "negative_reasons": neg_reasons[:5],
        "false_confident": false_confident,
        "false_gated": false_gated,
        "per_day_pos": dict(sorted(fb_pos_per_day.items())),
        "per_day_neg": dict(sorted(fb_neg_per_day.items())),
    }

    # ── Ambient ──────────────────────────────────────────────────────
    # ambient_entries is already provided by the caller; filter to window here.
    ambient_recent = [e for e in ambient_entries if (t := _ts(e)) and t >= cutoff]
    ambient_wikilinks = sum(e.get("wikilinks_applied", 0) for e in ambient_recent)
    ambient_per_day: dict[str, int] = defaultdict(int)
    for e in ambient_recent:
        t = _ts(e)
        if t:
            ambient_per_day[t.strftime("%Y-%m-%d")] += 1

    # ── Contradictions ───────────────────────────────────────────────
    contra_recent = [e for e in contra_entries if (t := _ts(e)) and t >= cutoff]
    contra_found = [e for e in contra_recent if e.get("contradicts") and not e.get("skipped")]
    contra_per_day: dict[str, int] = defaultdict(int)
    for e in contra_found:
        t = _ts(e)
        if t:
            contra_per_day[t.strftime("%Y-%m-%d")] += 1

    # ── Surface ──────────────────────────────────────────────────────
    surface_pairs = [e for e in surface_entries if e.get("cmd") == "surface_pair"]
    surface_runs = [e for e in surface_entries if e.get("cmd") == "surface_run"]

    # ── Filing ───────────────────────────────────────────────────────
    # filing_recent is already windowed by the caller.

    # ── Tune ─────────────────────────────────────────────────────────
    tune_history = []
    for e in tune_entries:
        t = _ts(e)
        tune_history.append({
            "ts": e.get("ts"),
            "samples": e.get("samples"),
            "baseline_hit": e.get("baseline", {}).get("hit"),
            "baseline_mrr": e.get("baseline", {}).get("mrr"),
            "best_hit": e.get("best", {}).get("hit"),
            "best_mrr": e.get("best", {}).get("mrr"),
            "delta": e.get("delta"),
        })

    # ── Sessions ─────────────────────────────────────────────────────
    sessions_dir = data_dir / "sessions"
    session_count = 0
    session_turns_total = 0
    wa_sessions = 0
    if sessions_dir.is_dir():
        for sf in sessions_dir.glob("*.json"):
            try:
                sess = json.loads(sf.read_text(encoding="utf-8"))
                session_count += 1
                turns = sess.get("turns", [])
                session_turns_total += len(turns)
                if sess.get("id", "").startswith("wa:"):
                    wa_sessions += 1
            except Exception:
                continue

    # ── Screen Time (macOS knowledgeC.db) ────────────────────────────
    # Cap at 7 days — dashboard day-pickers of 30/90 don't help: knowledge
    # store only keeps ~7-10d of /app/usage before CoreDuet aggregates it
    # away. Querying a wider window returns gaps, not more data.
    st_days = min(max(int(days), 1), 7)
    st_end = datetime.now()
    st_start = st_end - timedelta(days=st_days)
    screentime_agg = _collect_screentime(st_start, st_end)
    screentime_daily = _collect_screentime_daily(st_start, st_end, st_days)
    screentime_total_str = (
        _fmt_hm(int(screentime_agg.get("total_secs") or 0))
        if screentime_agg.get("available") else ""
    )

    # ── Index stats ──────────────────────────────────────────────────
    index_stats = {}
    try:
        col = get_db()
        n_chunks = col.count()
        corpus = _load_corpus(col)
        pr = get_pagerank(col)
        top_pr = sorted(pr.items(), key=lambda x: -x[1])[:5] if pr else []
        index_stats = {
            "chunks": n_chunks,
            "notes_files": len(corpus.get("outlinks", {})),
            "notes_titles": len(corpus.get("title_to_paths", {})),
            "notes": len(corpus.get("outlinks", {})),
            "tags": len(corpus.get("tags", set())),
            "folders": len(corpus.get("folders", set())),
            "top_pagerank": [{"path": p, "score": round(s, 4)} for p, s in top_pr],
        }
    except Exception:
        pass

    return {
        "period_days": days,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "kpis": {
            "total_queries": len(queries),
            "total_queries_all_time": len(all_queries),
            "avg_latency": round(statistics.mean(t_totals), 2) if t_totals else None,
            "avg_retrieve": round(statistics.mean(t_retrieves), 2) if t_retrieves else None,
            "avg_generate": round(statistics.mean(t_gens), 2) if t_gens else None,
            "feedback_positive": fb_pos,
            "feedback_negative": fb_neg,
            "feedback_positive_all_time": fb_pos_all,
            "feedback_negative_all_time": fb_neg_all,
            "sessions": session_count,
            "wa_sessions": wa_sessions,
            "ambient_hooks": len(ambient_recent),
            "ambient_wikilinks": ambient_wikilinks,
            "contradictions_found": len(contra_found),
            "surface_pairs": len(surface_pairs),
            "filings": len(filing_recent),
        },
        "health": health,
        "index": index_stats,
        "queries_per_day": dict(sorted(queries_per_day.items())),
        "latency_per_day": latency_ts,
        "hours": {str(h): hours.get(h, 0) for h in range(24)},
        "cmds": dict(sorted(cmds.items(), key=lambda x: -x[1])),
        "sources": dict(sources),
        "score_distribution": score_buckets,
        "score_stats": {
            "p50": _pct(scores, 50),
            "p95": _pct(scores, 95),
            "min": round(min(scores), 3) if scores else 0,
            "max": round(max(scores), 3) if scores else 0,
        },
        "latency_stats": {
            "retrieve_p50": _pct(t_retrieves, 50),
            "retrieve_p95": _pct(t_retrieves, 95),
            "generate_p50": _pct(t_gens, 50),
            "generate_p95": _pct(t_gens, 95),
            "total_p50": _pct(t_totals, 50),
            "total_p95": _pct(t_totals, 95),
        },
        "hot_topics": [{"topic": t, "count": c} for t, c in hot_topics],
        "chat_keywords": chat_keywords,
        "ambient_per_day": dict(sorted(ambient_per_day.items())),
        "contradictions_per_day": dict(sorted(contra_per_day.items())),
        "tune_history": tune_history,
        "surface_runs": len(surface_runs),
        "filing_confidence": [round(e.get("confidence", 0), 2) for e in filing_recent],
        "feedback": feedback_actionable,
        "screentime": {
            "available": screentime_agg.get("available", False),
            "window_days": st_days,
            "total_secs": int(screentime_agg.get("total_secs") or 0),
            "total_label": screentime_total_str,
            "top_apps": screentime_agg.get("top_apps") or [],
            "categories": screentime_agg.get("categories") or {},
            "daily": screentime_daily,
        },
    }


# ── Real-time stream ─────────────────────────────────────────────────────────
# Poll rag_* SQL tables for new rows and push them as SSE events. The browser
# polls /api/dashboard for the heavy aggregations; this endpoint streams
# deltas so KPIs and the live feed update without waiting for the next poll.
#
# Post-T10 note: writers no longer touch the `queries.jsonl` / `behavior.jsonl`
# / `ambient.jsonl` / `contradictions.jsonl` files, so the previous tail-based
# implementation emitted zero deltas in production. SQL ID-poll is the only
# viable source now. Each kind maps to a table + a row→event mapper that
# rebuilds the JSONL-equivalent shape `_stream_payload` expects.

# Rows returned per poll per table. Caps runaway growth if a long burst
# happens between polls; anything over this is silently skipped — the
# next poll picks up the newest rows (the UI feed is a live tail, not an
# audit log, so losing intermediate events during a burst is acceptable).
_STREAM_ROW_CAP = 100


def _row_to_query_ev(row: dict) -> dict:
    """Rebuild the JSONL-shaped event `_stream_payload('query', ...)` expects
    from a rag_queries row. `extra_json` carries legacy keys like `gated_
    low_confidence` and `error` — unpack when present, ignore otherwise.
    """
    extra: dict = {}
    raw_extra = row.get("extra_json")
    if raw_extra:
        try:
            extra = json.loads(raw_extra) or {}
        except Exception:
            extra = {}
    paths: list = []
    raw_paths = row.get("paths_json")
    if raw_paths:
        try:
            paths = json.loads(raw_paths) or []
        except Exception:
            paths = []
    bad: list = []
    raw_bad = row.get("bad_citations_json")
    if raw_bad:
        try:
            bad = json.loads(raw_bad) or []
        except Exception:
            bad = []
    return {
        "ts": row.get("ts"),
        "q": row.get("q"),
        "cmd": row.get("cmd"),
        "session": row.get("session"),
        "top_score": row.get("top_score"),
        "t_retrieve": row.get("t_retrieve"),
        "t_gen": row.get("t_gen"),
        "paths": paths,
        "bad_citations": bad,
        "gated_low_confidence": extra.get("gated_low_confidence"),
        "error": extra.get("error"),
    }


def _row_to_feedback_ev(row: dict) -> dict:
    """rag_feedback row → event shape `_stream_payload('feedback', ...)`."""
    extra: dict = {}
    raw_extra = row.get("extra_json")
    if raw_extra:
        try:
            extra = json.loads(raw_extra) or {}
        except Exception:
            extra = {}
    paths: list = []
    raw_paths = row.get("paths_json")
    if raw_paths:
        try:
            paths = json.loads(raw_paths) or []
        except Exception:
            paths = []
    return {
        "ts": row.get("ts"),
        "rating": row.get("rating"),
        "q": row.get("q"),
        "paths": paths,
        "reason": extra.get("reason"),
        "corrective_path": extra.get("corrective_path"),
    }


def _row_to_ambient_ev(row: dict) -> dict:
    """rag_ambient row → event shape `_stream_payload('ambient', ...)`.
    `wikilinks_applied` lives in `payload_json`; absent → 0."""
    payload: dict = {}
    raw_payload = row.get("payload_json")
    if raw_payload:
        try:
            payload = json.loads(raw_payload) or {}
        except Exception:
            payload = {}
    return {
        "ts": row.get("ts"),
        "path": row.get("path"),
        "wikilinks_applied": payload.get("wikilinks_applied", 0),
    }


def _row_to_contradiction_ev(row: dict) -> dict:
    """rag_contradictions row → event shape `_stream_payload('contradiction',
    ...)`. The dashboard expects `path` (subject) and `contradicts` (list)."""
    contradicts: list = []
    raw = row.get("contradicts_json")
    if raw:
        try:
            contradicts = json.loads(raw) or []
        except Exception:
            contradicts = []
    skipped_val = row.get("skipped")
    return {
        "ts": row.get("ts"),
        "path": row.get("subject_path"),
        "contradicts": contradicts,
        "skipped": bool(skipped_val) and skipped_val not in ("", "0", "false", "no"),
    }


# (table, mapper) per stream kind. Mappers rebuild the JSONL-shaped dict that
# `_stream_payload(kind, ev)` consumes — keeps the downstream shaping logic
# (which the dashboard JS already depends on) byte-identical.
_STREAM_SOURCES: dict[str, tuple[str, Callable[[dict], dict]]] = {
    "query":         ("rag_queries",        _row_to_query_ev),
    "feedback":      ("rag_feedback",       _row_to_feedback_ev),
    "ambient":       ("rag_ambient",        _row_to_ambient_ev),
    "contradiction": ("rag_contradictions", _row_to_contradiction_ev),
}


# Audit 2026-04-25 finding R2-Performance #3: límite de SSE streams
# concurrentes por IP. Cada conexión al `/api/dashboard/stream` mantiene
# un poll cada 1.5s contra SQLite indefinidamente; un solo browser con
# N tabs del dashboard abiertas multiplica el costo. 10 tabs × 24h ≈
# 57k ciclos de polling acumulados — fácil de gatillar accidentalmente.
#
# Tracking: contador in-memory por IP protegido por lock. Se incrementa
# al entrar al handler y se decrementa en el `finally` del generator
# (cuando el cliente cierra el stream o cancela la request).
#
# Default 3 streams por IP cubre el caso real (PWA + 1 desktop tab + 1
# mobile) sin permitir runaway. Configurable por env var.
_SSE_MAX_PER_IP_DEFAULT = 3
try:
    _SSE_MAX_PER_IP = int(os.environ.get("RAG_SSE_MAX_PER_IP", _SSE_MAX_PER_IP_DEFAULT))
    if _SSE_MAX_PER_IP < 1:
        _SSE_MAX_PER_IP = _SSE_MAX_PER_IP_DEFAULT
except (TypeError, ValueError):
    _SSE_MAX_PER_IP = _SSE_MAX_PER_IP_DEFAULT
_SSE_CONNECTIONS_PER_IP: dict[str, int] = {}
_SSE_CONNECTIONS_LOCK = _threading.Lock()


def _sse_acquire_slot(ip: str) -> bool:
    """Reserva un slot SSE para `ip`. Devuelve False si ya hay
    `_SSE_MAX_PER_IP` conexiones activas. Audit 2026-04-25 R2-Performance #3."""
    with _SSE_CONNECTIONS_LOCK:
        current = _SSE_CONNECTIONS_PER_IP.get(ip, 0)
        if current >= _SSE_MAX_PER_IP:
            return False
        _SSE_CONNECTIONS_PER_IP[ip] = current + 1
        return True


def _sse_release_slot(ip: str) -> None:
    """Libera un slot SSE para `ip` (idempotente con el contador a 0)."""
    with _SSE_CONNECTIONS_LOCK:
        current = _SSE_CONNECTIONS_PER_IP.get(ip, 0)
        if current <= 1:
            _SSE_CONNECTIONS_PER_IP.pop(ip, None)
        else:
            _SSE_CONNECTIONS_PER_IP[ip] = current - 1


def _stream_max_id(conn, table: str) -> int:
    """Starting cursor for a stream — connection-time high-water mark so the
    client only sees rows written AFTER it connects. Missing table or no rows
    → 0."""
    try:
        row = conn.execute(f"SELECT COALESCE(MAX(id), 0) FROM {table}").fetchone()
        return int(row[0]) if row else 0
    except Exception:
        return 0


def _stream_fetch_since(conn, table: str, last_id: int) -> list[dict]:
    """Return up to `_STREAM_ROW_CAP` new rows with `id > last_id` ordered by
    id ascending. Empty list on any error (connection churn, transient lock)
    — the next poll retries."""
    try:
        cur = conn.execute(
            f"SELECT * FROM {table} WHERE id > ? ORDER BY id LIMIT ?",
            (last_id, _STREAM_ROW_CAP),
        )
        cols = [c[0] for c in cur.description]
        return [dict(zip(cols, r)) for r in cur.fetchall()]
    except Exception:
        return []


def _do_poll_sql(cursors: dict) -> list[tuple[str, dict]]:
    """Sync helper: open one SQL connection, scan all stream sources, return
    list of (kind, payload) pairs for rows found since the last cursor.

    Extracted so `dashboard_stream` can offload the blocking I/O to a thread
    pool via `asyncio.to_thread`, keeping the async event loop unblocked.
    Cursor state is mutated in-place so callers see the updated positions.
    """
    import rag as _rag

    events: list[tuple[str, dict]] = []
    try:
        with _rag._ragvec_state_conn() as conn:
            for kind, (table, mapper) in _STREAM_SOURCES.items():
                if cursors[kind] is None:
                    cursors[kind] = _stream_max_id(conn, table)
                    continue
                rows = _stream_fetch_since(conn, table, cursors[kind])
                if not rows:
                    continue
                cursors[kind] = max(int(r.get("id") or 0) for r in rows)
                for r in rows:
                    ev = mapper(r)
                    events.append((kind, _stream_payload(kind, ev)))
    except Exception:
        pass
    return events


@app.get("/api/dashboard/stream")
async def dashboard_stream(request: Request = None) -> StreamingResponse:  # type: ignore[assignment]
    """SSE: poll rag_* SQL tables for new rows and push them as events.

    Each kind tracks its own `last_id` cursor, initialised at connection
    time from `MAX(id)` so the client only receives rows inserted during
    the connection. Per poll, one SQL connection is opened, all 4 tables
    are scanned, then the connection is closed — keeps total connection
    time short (WAL readers don't block writers but we prefer no-op
    polls to not hold the conn).

    Audit 2026-04-25 finding R2-Performance #3: cap por IP de streams
    concurrentes. Si la IP ya tiene `_SSE_MAX_PER_IP` conexiones
    abiertas, devolvemos 429 antes de armar el generator. El slot se
    libera en el `finally` del generator (sea por cierre limpio o
    cancelación del cliente).

    `request` es opcional para tests que invocan el handler directo
    (sin pasar por el routing de FastAPI) — en producción FastAPI lo
    inyecta siempre. Si viene None, salteamos el rate-limit por IP.
    """
    # Cap por IP — el primer chequeo es gate-keeping, el slot real se
    # reserva en `_sse_acquire_slot` con un lock para que dos requests
    # simultáneas no se pasen el límite por race. Si `request is None`
    # (caso test), el slot no se gestiona y `client_ip` queda en None.
    client_ip: str | None = None
    if request is not None:
        client_ip = (request.client.host if request.client else "unknown")
        if not _sse_acquire_slot(client_ip):
            raise HTTPException(
                status_code=429,
                detail=f"too many concurrent streams (max {_SSE_MAX_PER_IP} per IP)",
            )

    # Cursors per kind — populated lazily on first poll so connection-setup
    # errors don't kill the SSE stream (the dashboard uses this endpoint as
    # a heartbeat in addition to a data feed).
    cursors: dict[str, int | None] = {kind: None for kind in _STREAM_SOURCES}

    async def gen():
        last_heartbeat = time.time()
        # Initial hello so the client can flip the indicator immediately.
        yield _sse("hello", {"t": time.time(), "tracking": list(_STREAM_SOURCES)})
        try:
            while True:
                # Offload blocking SQL I/O to a thread so the async event
                # loop stays free for concurrent /api/chat and other SSE
                # streams. `_do_poll_sql` mutates `cursors` in-place.
                events = await asyncio.to_thread(_do_poll_sql, cursors)
                for kind, payload in events:
                    yield _sse(kind, payload)
                now = time.time()
                if now - last_heartbeat >= 15:
                    yield _sse("heartbeat", {"t": now})
                    last_heartbeat = now
                await asyncio.sleep(1.5)
        except asyncio.CancelledError:
            return
        finally:
            # Audit 2026-04-25 R2-Performance #3: liberamos el slot tanto si
            # el stream terminó limpio como si el cliente canceló (el cleanup
            # de FastAPI dispara CancelledError al cerrar la conexión).
            # `client_ip is None` cuando el handler lo invocó un test directo
            # sin Request — no hay slot que liberar.
            if client_ip is not None:
                _sse_release_slot(client_ip)

    return StreamingResponse(
        gen(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


def _stream_payload(kind: str, ev: dict) -> dict:
    """Trim raw log entries to the fields the dashboard needs."""
    if kind == "query":
        sid = ev.get("session") or ""
        # Whatsapp listener uses session ids like `wa:<jid>` OR custom strings
        # like `<runtime>:<vault>` (rag serve from whatsapp-listener) — treat
        # `serve` cmd as whatsapp by default since that's the only consumer.
        cmd = ev.get("cmd") or ""
        if sid.startswith("wa:") or cmd.startswith("serve"):
            source = "whatsapp"
        elif sid.startswith("web:"):
            source = "web"
        elif sid.startswith("tg:"):
            source = "legacy"
        else:
            source = "cli"
        tr = ev.get("t_retrieve")
        tg = ev.get("t_gen")
        latency = (
            round(float(tr) + float(tg), 2)
            if isinstance(tr, (int, float)) and isinstance(tg, (int, float))
            else None
        )
        # phase: received → in-flight, error → failed, anything else → completed
        if cmd == "serve.received":
            phase = "in_flight"
        elif cmd == "serve.error":
            phase = "error"
        else:
            phase = "done"
        return {
            "ts": ev.get("ts"),
            "q": (ev.get("q") or "")[:160],
            "score": ev.get("top_score"),
            "latency": latency,
            "source": source,
            "cmd": cmd,
            "phase": phase,
            "error": (ev.get("error") or "")[:200] if ev.get("error") else None,
            "gated": bool(ev.get("gated_low_confidence")),
            "bad_citations": len(ev.get("bad_citations") or []),
            "n_paths": len(ev.get("paths") or []),
        }
    if kind == "feedback":
        return {
            "ts": ev.get("ts"),
            "rating": ev.get("rating"),
            "q": (ev.get("q") or "")[:160],
            "reason": (ev.get("reason") or "")[:200],
            "corrective_path": ev.get("corrective_path"),
            "n_paths": len(ev.get("paths") or []),
        }
    if kind == "ambient":
        return {
            "ts": ev.get("ts"),
            "path": ev.get("path"),
            "wikilinks_applied": ev.get("wikilinks_applied", 0),
        }
    if kind == "contradiction":
        return {
            "ts": ev.get("ts"),
            "path": ev.get("path"),
            "contradicts": ev.get("contradicts"),
            "skipped": bool(ev.get("skipped")),
        }
    return ev


# ── /learning ─────────────────────────────────────────────────
# Dashboard secundario, focalizado en el LOOP de aprendizaje del RAG: cómo
# evoluciona el ranker, cuánto feedback (explícito + implícito) entra,
# qué tan bien funciona la cache, qué entities/contradictions emergen, etc.
# El endpoint principal `/api/dashboard` cubre métricas operativas (latencia,
# top topics, cmd breakdown). Este otro responde la pregunta "el sistema
# está aprendiendo o estancado?".
#
# Cache TTL: 60s (vs 30s del operativo) — los datos del learning loop
# cambian más lento (eval runs son nightly, tune runs son nightly, weights
# se reescriben 1/día). Sirve directo desde RAM si hay hit warm.

_LEARNING_CACHE: dict[int, tuple[float, dict]] = {}
_LEARNING_CACHE_LOCK = threading.Lock()  # audit 2026-04-26 — race en SSE concurrente
_LEARNING_TTL = 60.0


@app.get("/learning")
def learning_page() -> FileResponse:
    """HTML estático del learning dashboard. El contenido se hidrata vía
    `/api/learning` (initial load) + `/api/learning/stream`
    (SSE, snapshot completo cada 30s)."""
    return FileResponse(STATIC_DIR / "learning.html")


@app.get("/api/learning")
def learning_api(days: int = 30) -> dict:
    """Snapshot completo de las 11 secciones + KPIs hero. Cada sección es
    independiente — un SQL error en una NO contamina las otras (las
    funciones de `web.learning_queries` envuelven sus reads con
    `_sql_read_with_retry` y devuelven shape vacío en error)."""
    now_ts = time.time()
    with _LEARNING_CACHE_LOCK:
        hit = _LEARNING_CACHE.get(days)
        if hit and now_ts - hit[0] < _LEARNING_TTL:
            return hit[1]
    # Lazy import: módulo nuevo, no queremos cargarlo al import-time del
    # servidor si nadie pide /learning. Después del primer hit
    # queda cacheado por el import system; las llamadas subsiguientes son
    # gratis.
    from web.learning_queries import (
        anticipatory,
        behavior,
        feedback_explicit,
        feedback_implicit,
        kpis,
        query_learning,
        ranker_weights,
        retrieval_quality,
        routing_learning,
        score_calibration,
        vault_intelligence,
        verdict,
        whisper_learning,
    )
    payload = {
        "meta": {
            "window_days": days,
            "generated_at": datetime.now(timezone.utc).isoformat(),
        },
        # Veredicto: ¿aprende cada uno de los 12 sistemas? Origen: el
        # diagnóstico manual del 2026-04-26 que detectó loop roto en
        # anticipatory + 3 sistemas dormidos. Auto-generado para que la
        # próxima vez algo se rompa, esté visible en /learning.
        "verdict": verdict(),
        "kpis": kpis(days),
        "sections": {
            "retrieval_quality": retrieval_quality(days),
            "ranker_weights": ranker_weights(days),
            "score_calibration": score_calibration(),
            "feedback_explicit": feedback_explicit(days),
            "feedback_implicit": feedback_implicit(days),
            "behavior": behavior(days),
            "query_learning": query_learning(days),
            "anticipatory": anticipatory(days),
            "routing_learning": routing_learning(days),
            "whisper_learning": whisper_learning(days),
            "vault_intelligence": vault_intelligence(days),
        },
    }
    with _LEARNING_CACHE_LOCK:
        _LEARNING_CACHE[days] = (now_ts, payload)
    return payload


# ── /api/learning/health ─────────────────────────────────────────
# Semáforo del sistema: bloque "verde / amarillo / rojo" arriba del dashboard
# que un usuario sin conocimiento técnico interpreta en 2s. Computa 6 señales
# (acierto en preguntas simples + complejas, servicios, vault al día, errores
# 24h, velocidad de respuesta) y aplica worst-case wins. Ver `system_health`
# en `web/learning_queries.py` para la lógica + thresholds.
#
# TTL=15s — más corto que `/api/learning` (60s) porque queremos
# que el banner refresque rápido. La señal `services` invoca `launchctl list`
# (subprocess, ~50ms) — el resto son SQL reads cacheados por _ragvec_state_conn.

_LEARNING_HEALTH_CACHE: tuple[float, dict] | None = None
_LEARNING_HEALTH_TTL = 15.0


@app.get("/api/learning/health")
def learning_health_api() -> dict:
    """Devuelve el semáforo del sistema. Shape estable:

        {
          "level": "green" | "yellow" | "red",
          "headline": "Todo funcionando bien" | ...,
          "summary": "El sistema responde bien y está al día.",
          "signals": [
            {"key": ..., "label": ..., "level": ...,
             "value_text": ..., "value_raw": ..., "tooltip": ...,
             "explanation": ...},
            ...
          ],
          "checked_at": "2026-04-26T..."
        }

    Cache TTL=15s. Cada señal corre en try/except independiente — la falla
    de una NO afecta a las otras."""
    global _LEARNING_HEALTH_CACHE
    now_ts = time.time()
    hit = _LEARNING_HEALTH_CACHE
    if hit and now_ts - hit[0] < _LEARNING_HEALTH_TTL:
        return hit[1]
    from web.learning_queries import system_health
    payload = system_health()
    _LEARNING_HEALTH_CACHE = (now_ts, payload)
    return payload


@app.get("/api/learning/stream")
async def learning_stream(request: Request = None) -> StreamingResponse:  # type: ignore[assignment]
    """SSE: cada 30s emite un snapshot completo del payload de
    /api/learning. Heartbeat con comentario `: keep-alive` cada
    15s entre snapshots para que proxies/CDNs no maten la conexión por
    inactividad. El cliente reconecta automáticamente con EventSource si
    se cae."""
    # Audit 2026-04-26 BUG #5: agregar slot cap.
    client_ip: str | None = None
    if request is not None:
        client_ip = (request.client.host if request.client else "unknown")
        if not _sse_acquire_slot(client_ip):
            raise HTTPException(status_code=429,
                detail=f"too many concurrent streams (max {_SSE_MAX_PER_IP} per IP)")

    async def gen():
        last_snapshot_ts = 0.0
        try:
            # learning_api() es sync con I/O — sacarlo del event loop.
            # Audit 2026-04-26 BUG #1 web: pre-fix bloqueaba el loop por
            # toda la duración (10s+ con 11 secciones SQL).
            payload = await asyncio.to_thread(learning_api, days=30)
            yield f"event: snapshot\ndata: {json.dumps(payload, ensure_ascii=False)}\n\n"
            last_snapshot_ts = time.time()
            while True:
                if request is not None and await request.is_disconnected():
                    break
                now = time.time()
                if now - last_snapshot_ts >= 30:
                    payload = await asyncio.to_thread(learning_api, days=30)
                    yield f"event: snapshot\ndata: {json.dumps(payload, ensure_ascii=False)}\n\n"
                    last_snapshot_ts = now
                else:
                    yield ": keep-alive\n\n"
                await asyncio.sleep(15)
        except asyncio.CancelledError:
            return
        finally:
            if client_ip is not None:
                _sse_release_slot(client_ip)

    return StreamingResponse(
        gen(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


# ── RAG-system memory sampler ──────────────────────────────────────────
# Tracks RSS of processes that are part of the rag stack we built:
#   rag        — obsidian-rag python (watch, morning, today, digest, web,
#                mcp, chat, query, launchd-spawned jobs, their children)
#   ollama     — ollama serve + per-model runners (LLM + embeddings)
#   sqlite-vec — sqlite-vec-gui streamlit inspector pointed at ragvec.db
#   whatsapp   — whatsapp-bridge, whatsapp-listener, whatsapp-mcp, vault-sync
#
# System-wide processes (browser, Claude Code, unrelated python) are
# intentionally excluded — this chart answers "how much RAM does OUR
# stack use?", not "what's eating the machine?".
#
# Zero-dep: parses `ps -axo rss=,command=` on the full command line so
# we can match interpreter path + script path (short `comm=` collapses
# everything to `python3`). RSS on macOS double-counts shared memory
# across helper processes, so absolute numbers overstate real usage;
# trends are what the chart is for.
from collections import deque

_MEMORY_STATE_PATH = Path.home() / ".local/share/obsidian-rag/rag_memory.jsonl"
_MEMORY_BUFFER_MAX = 1440  # 24h @ 60s
_MEMORY_SAMPLE_INTERVAL = 60.0
_MEMORY_BUFFER: deque = deque(maxlen=_MEMORY_BUFFER_MAX)
_MEMORY_LOCK = threading.Lock()

_MEMORY_CATEGORIES = ("rag", "ollama", "sqlite-vec", "whatsapp")

# Order matters: first regex to match claims the process. `whatsapp`
# goes before `rag` because whatsapp-mcp lives under a path that also
# matches obsidian-rag venvs in edge cases; `ollama` before `rag`
# likewise for safety.
_RAG_PROC_MATCHERS: tuple[tuple[str, "re.Pattern[str]"], ...] = (
    ("whatsapp",   re.compile(r"whatsapp-(bridge|listener|vault-sync|mcp)")),
    ("ollama",     re.compile(r"(?:^|/)ollama(?:\s|$)")),
    ("sqlite-vec", re.compile(r"sqlite-vec-gui\b")),
    ("rag",        re.compile(r"obsidian-rag")),
)


def _classify_rag_proc(cmd: str) -> str | None:
    for cat, rx in _RAG_PROC_MATCHERS:
        if rx.search(cmd):
            return cat
    return None


def _rag_proc_label(cmd: str, cat: str) -> str:
    """Friendly, aggregatable label for the `top` list."""
    if cat == "rag":
        if "obsidian-rag-mcp" in cmd:
            return "obsidian-rag-mcp"
        if "web/server.py" in cmd:
            return "rag web"
        m = re.search(r"/rag\s+(\w[\w-]*)", cmd)
        if m:
            return f"rag {m.group(1)}"
        if "resource_tracker" in cmd:
            return "rag (resource_tracker)"
        return "rag (python)"
    if cat == "ollama":
        if "ollama runner" in cmd:
            m = re.search(r"sha256-([0-9a-f]{8,})", cmd)
            return f"ollama runner ({m.group(1)[:8]})" if m else "ollama runner"
        if "ollama serve" in cmd:
            return "ollama serve"
        return "ollama"
    if cat == "sqlite-vec":
        return "sqlite-vec-gui"
    if cat == "whatsapp":
        for tag in ("whatsapp-bridge", "whatsapp-listener", "whatsapp-vault-sync", "whatsapp-mcp"):
            if tag in cmd:
                return tag
        return "whatsapp"
    return cmd.split()[0].rsplit("/", 1)[-1]


def _read_vm_stat() -> dict:
    """Parse `vm_stat` to surface pressure + free pages."""
    try:
        out = subprocess.run(
            ["vm_stat"], capture_output=True, text=True, errors="replace", timeout=2, check=False,
        ).stdout
    except Exception:
        return {}
    page_size = 4096
    stats: dict[str, float] = {}
    for line in out.splitlines():
        if "page size of" in line:
            m = re.search(r"page size of (\d+)", line)
            if m:
                page_size = int(m.group(1))
            continue
        if ":" not in line:
            continue
        key, val = line.split(":", 1)
        val = val.strip().rstrip(".")
        try:
            stats[key.strip()] = float(val)
        except ValueError:
            continue
    mb = lambda pages: round(pages * page_size / (1024 * 1024), 1)
    return {
        "free_mb": mb(stats.get("Pages free", 0)),
        "active_mb": mb(stats.get("Pages active", 0)),
        "inactive_mb": mb(stats.get("Pages inactive", 0)),
        "wired_mb": mb(stats.get("Pages wired down", 0)),
        "compressed_mb": mb(stats.get("Pages occupied by compressor", 0)),
    }


def _sample_memory() -> dict | None:
    """One sample: RSS totals per rag-stack category + top processes + vm_stat."""
    try:
        out = subprocess.run(
            ["ps", "-axo", "rss=,command="],
            capture_output=True, text=True, errors="replace", timeout=5, check=False,
        ).stdout
    except Exception:
        return None

    by_cat: dict[str, float] = {k: 0.0 for k in _MEMORY_CATEGORIES}
    procs: dict[str, dict] = {}
    for line in out.splitlines():
        line = line.strip()
        if not line:
            continue
        parts = line.split(None, 1)
        if len(parts) != 2:
            continue
        try:
            rss_kb = int(parts[0])
        except ValueError:
            continue
        cmd = parts[1]
        cat = _classify_rag_proc(cmd)
        if cat is None:
            continue
        rss_mb = rss_kb / 1024.0
        by_cat[cat] += rss_mb
        label = _rag_proc_label(cmd, cat)
        slot = procs.get(label)
        if slot is None:
            procs[label] = {"name": label, "mb": rss_mb, "cat": cat}
        else:
            slot["mb"] += rss_mb

    total = round(sum(by_cat.values()), 1)
    by_cat_rounded = {k: round(v, 1) for k, v in by_cat.items()}
    top = sorted(procs.values(), key=lambda d: d["mb"], reverse=True)[:10]
    top_list = [{"name": d["name"], "mb": round(d["mb"], 1), "cat": d["cat"]} for d in top]

    return {
        "ts": datetime.now().isoformat(timespec="seconds"),
        "total_mb": total,
        "by_category": by_cat_rounded,
        "top": top_list,
        "vm": _read_vm_stat(),
    }


def _memory_load_history() -> None:
    """Warm the ring buffer from the JSONL sidecar (last N lines)."""
    if not _MEMORY_STATE_PATH.exists():
        return
    try:
        with _MEMORY_STATE_PATH.open("r", encoding="utf-8") as fh:
            lines = fh.readlines()[-_MEMORY_BUFFER_MAX:]
    except Exception:
        return
    for raw in lines:
        raw = raw.strip()
        if not raw:
            continue
        try:
            _MEMORY_BUFFER.append(json.loads(raw))
        except Exception:
            continue


def _metrics_background_default() -> bool:
    """True cuando las writes de los samplers (CPU/memory metrics) deben ir
    al background queue en vez de sync.

    Default ON tras audit 2026-04-24: los samplers corren en daemon threads
    cada 60s y contenían por el WAL lock de telemetry.db con el resto de
    los writers (queries, behavior, impressions, semantic_cache,
    feedback_golden). Pre-fix había 66 `memory_sql_write_failed` + 34
    `cpu_sql_write_failed` en `sql_state_errors.jsonl` (subset de los 1756
    errores totales del periodo 19-24 Abr). Los reads del dashboard +
    `rag insights` toleran 1-2s de delay sin regresión observable.

    Override: `RAG_METRICS_ASYNC=0` fuerza sync (útil si se sospecha que
    el daemon worker está saturado).
    """
    val = os.environ.get("RAG_METRICS_ASYNC", "").strip().lower()
    return val not in ("0", "false", "no")


def _persist_with_sqlite_retry(
    write_fn, error_tag: str, *, attempts: int = 8,
) -> None:
    """Run a one-shot SQL-write closure with transient-lock retry.

    `_memory_persist` + `_cpu_persist` (and historically anything else
    firing from the per-minute samplers) were dropping ~30 samples/day
    to `sql_state_errors.jsonl` with `database is locked`, because the
    bare `_ragvec_state_conn() + _sql_append_event` path has no retry —
    if any other writer holds the WAL write-lock longer than the
    `busy_timeout=30s` window (e.g. during `_write_feedback_golden_sql`
    under embed latency, or a concurrent contradiction worker), the
    sample is lost.

    **2026-04-23 tuning**: bumped attempts 3→8 + backoff max 0.25→0.6
    (total budget ~4s vs 0.75s pre). Audit del sql_state_errors.jsonl
    mostró 258 `queries_sql_write_failed` + 64 samples perdidos en las
    últimas semanas — el old budget de 3 intentos × ~0.35s se quedaba
    corto bajo bursts coordinados (memory+cpu samplers alineados cada
    60s + el queries writer concurrente). 4s es suficiente para esperar
    un WAL checkpoint stall sin bloquear indefinidamente. Callers que
    están en el hot path del usuario (donde 4s de delay es visible)
    deben pasar `attempts=3` explícito para preservar el comportamiento
    tight; los samplers + backfill writers usan el default bumped.

    Expanded error transience: además de "locked", también reintenta
    "disk I/O error" — audit mostró 92 `disk I/O error` transitorios
    que el primer intento fallaba y el segundo pasaba limpio.
    """
    import random as _r
    import sqlite3 as _sqlite3
    import time as _t
    for attempt in range(attempts):
        try:
            write_fn()
            return
        except _sqlite3.OperationalError as exc:
            msg = str(exc).lower()
            _is_transient = (
                "locked" in msg
                or "disk i/o" in msg
                or "disk io" in msg
            )
            if not _is_transient or attempt == attempts - 1:
                _log_sql_state_error(error_tag, err=repr(exc))
                return
            _t.sleep(0.15 + _r.random() * 0.45)
        except Exception as exc:
            _log_sql_state_error(error_tag, err=repr(exc))
            return


def _memory_persist(sample: dict) -> None:
    """Memory metrics sampler write — daemon thread every 60s.

    Fire-and-forget via background queue (2026-04-24): pre-fix había 66
    `memory_sql_write_failed` en `sql_state_errors.jsonl` por contención
    WAL contra los otros writers de telemetry.db. El sampler no tiene
    consumidor sync — los reads (dashboard, `rag insights`) toleran
    delays 1-2s sin problema.

    Rollback: `RAG_METRICS_ASYNC=0` fuerza sync (útil si se sospecha que
    el daemon worker está saturado y samples se están perdiendo en la
    cola — aunque ese caso es más una señal de un bug en otro lado).
    """
    def _do() -> None:
        with _ragvec_state_conn() as conn:
            _sql_append_event(conn, "rag_memory_metrics",
                               _map_memory_row(sample))

    if _metrics_background_default():
        _enqueue_background_sql(_do, "memory_sql_write_failed")
    else:
        _persist_with_sqlite_retry(_do, "memory_sql_write_failed")


def _memory_trim_file() -> None:
    """Keep the JSONL file bounded at ~2x buffer so it doesn't grow forever."""
    try:
        if not _MEMORY_STATE_PATH.exists():
            return
        with _MEMORY_STATE_PATH.open("r", encoding="utf-8") as fh:
            lines = fh.readlines()
        cap = _MEMORY_BUFFER_MAX * 2
        if len(lines) <= cap:
            return
        tail = lines[-cap:]
        with _MEMORY_STATE_PATH.open("w", encoding="utf-8") as fh:
            fh.writelines(tail)
    except Exception:
        pass


@_on_startup
def _start_memory_sampler() -> None:
    _memory_load_history()

    def loop() -> None:
        import random as _r
        # Initial jitter: duerma 0-30s random antes del primer sample para
        # no alinearse con otros samplers (cpu, memory_pressure_watchdog,
        # queries writer) que arrancan al mismo startup. Sin jitter cada 60s
        # los 3 writers colisionaban simultáneo, saturaban el WAL lock, y
        # uno perdía el sample (~64/semana en el audit 2026-04-23).
        time.sleep(_r.uniform(0, 30))
        trim_counter = 0
        while True:
            sample = _sample_memory()
            if sample:
                with _MEMORY_LOCK:
                    _MEMORY_BUFFER.append(sample)
                _memory_persist(sample)
                trim_counter += 1
                if trim_counter >= 60:  # once an hour
                    _memory_trim_file()
                    trim_counter = 0
            # Per-cycle jitter ±5s para evitar drift hacia realineamiento
            # después del primer offset. Rango [55, 65] sobre interval 60s.
            time.sleep(_MEMORY_SAMPLE_INTERVAL + _r.uniform(-5, 5))

    threading.Thread(target=loop, name="memory-sampler", daemon=True).start()


@app.get("/api/system-memory")
def system_memory_api(minutes: int = 360) -> dict:
    """Return the last `minutes` of samples (default 6h) + a fresh current.

    Categories: python, browser, ollama, node, claude, other (all MB).
    Used for initial backfill; the live stream is `/api/system-memory/stream`.
    """
    minutes = max(5, min(minutes, 1440))
    cutoff = datetime.now() - timedelta(minutes=minutes)
    with _MEMORY_LOCK:
        all_samples = list(_MEMORY_BUFFER)

    samples = []
    for s in all_samples:
        ts = s.get("ts") or ""
        try:
            if datetime.fromisoformat(ts) >= cutoff:
                samples.append(s)
        except Exception:
            continue

    current = _sample_memory()
    return {
        "categories": list(_MEMORY_CATEGORIES),
        "samples": samples,
        "current": current,
        "interval_s": _MEMORY_SAMPLE_INTERVAL,
        "live_interval_s": _MEMORY_LIVE_INTERVAL,
        "window_minutes": minutes,
    }


# Live SSE stream. Emits a fresh sample every `_MEMORY_LIVE_INTERVAL` seconds
# so the chart updates in real time instead of waiting for the 60s persistent
# tick. The persistent 60s sampler (for the JSONL sidecar + historical
# buffer) runs independently — live ticks are NOT persisted to avoid
# bloating the JSONL at 2s cadence.
_MEMORY_LIVE_INTERVAL = 2.0


@app.get("/api/system-memory/stream")
async def system_memory_stream(request: Request = None) -> StreamingResponse:  # type: ignore[assignment]
    # Audit 2026-04-26: agregar slot cap como dashboard/stream para evitar
    # N tabs × infinite-loop por IP saturando el server.
    client_ip: str | None = None
    if request is not None:
        client_ip = (request.client.host if request.client else "unknown")
        if not _sse_acquire_slot(client_ip):
            raise HTTPException(status_code=429,
                detail=f"too many concurrent streams (max {_SSE_MAX_PER_IP} per IP)")

    async def gen():
        try:
            first = await asyncio.to_thread(_sample_memory)
            if first:
                yield _sse("sample", first)
            while True:
                await asyncio.sleep(_MEMORY_LIVE_INTERVAL)
                sample = await asyncio.to_thread(_sample_memory)
                if not sample:
                    continue
                yield _sse("sample", sample)
        finally:
            if client_ip is not None:
                _sse_release_slot(client_ip)

    return StreamingResponse(
        gen(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache, no-transform",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


# ── RAG-stack CPU sampler ──────────────────────────────────────────────
# Same scope as the memory sampler (rag, ollama, sqlite-vec, whatsapp).
# macOS `ps -o %cpu` is a decaying average, which smears transient load
# — instead we compute instantaneous CPU% from deltas of cumulative CPU
# time across two snapshots (`ps -axo pid,cputime,command`). Each caller
# (live SSE stream, persistent sampler, etc.) owns its own prev-state
# dict so concurrent streams don't fight over the baseline.
#
# Returned percentages are in "% of one core" (can exceed 100 on
# multi-threaded processes — ollama runners routinely peg several
# cores during generation).
_CPU_STATE_PATH = Path.home() / ".local/share/obsidian-rag/rag_cpu.jsonl"
_CPU_BUFFER_MAX = 1440  # 24h @ 60s
_CPU_SAMPLE_INTERVAL = 60.0
_CPU_LIVE_INTERVAL = 2.0
_CPU_BUFFER: deque = deque(maxlen=_CPU_BUFFER_MAX)
_CPU_LOCK = threading.Lock()
_CPU_NCORES = os.cpu_count() or 1


def _parse_cputime(s: str) -> float | None:
    """Parse macOS ps cputime — `M:SS.FF` or `H:MM:SS.FF` — into seconds."""
    try:
        parts = s.strip().split(":")
        if len(parts) == 2:
            return int(parts[0]) * 60 + float(parts[1])
        if len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + float(parts[2])
    except (ValueError, TypeError):
        pass
    return None


def _sample_cpu(prev_state: dict) -> dict | None:
    """Instantaneous CPU% per rag-stack category via cputime delta.

    `prev_state` holds the previous snapshot across calls:
        {"ts": monotonic_seconds, "per_pid": {pid: cputime_s}}
    First invocation only primes the state and returns None.
    """
    now_mono = time.monotonic()
    try:
        out = subprocess.run(
            ["ps", "-axo", "pid=,cputime=,command="],
            capture_output=True, text=True, errors="replace", timeout=5, check=False,
        ).stdout
    except Exception:
        return None

    per_pid_cpu: dict[int, float] = {}
    per_pid_cmd: dict[int, str] = {}
    for line in out.splitlines():
        line = line.strip()
        if not line:
            continue
        parts = line.split(None, 2)
        if len(parts) != 3:
            continue
        try:
            pid = int(parts[0])
        except ValueError:
            continue
        cpu_s = _parse_cputime(parts[1])
        if cpu_s is None:
            continue
        per_pid_cpu[pid] = cpu_s
        per_pid_cmd[pid] = parts[2]

    prev_per_pid = prev_state.get("per_pid") or {}
    prev_ts = prev_state.get("ts")
    prev_state["ts"] = now_mono
    prev_state["per_pid"] = per_pid_cpu

    if prev_ts is None:
        return None
    interval = now_mono - prev_ts
    if interval <= 0:
        return None

    by_cat: dict[str, float] = {k: 0.0 for k in _MEMORY_CATEGORIES}
    procs: dict[str, dict] = {}
    for pid, cpu_s in per_pid_cpu.items():
        prev_cpu = prev_per_pid.get(pid)
        if prev_cpu is None:
            continue
        delta = cpu_s - prev_cpu
        if delta < 0:
            continue  # clock anomaly or pid reuse — skip
        pct = (delta / interval) * 100.0
        if pct < 0.05:
            continue
        cmd = per_pid_cmd[pid]
        cat = _classify_rag_proc(cmd)
        if cat is None:
            continue
        by_cat[cat] += pct
        label = _rag_proc_label(cmd, cat)
        slot = procs.get(label)
        if slot is None:
            procs[label] = {"name": label, "pct": pct, "cat": cat}
        else:
            slot["pct"] += pct

    total = round(sum(by_cat.values()), 1)
    by_cat_rounded = {k: round(v, 1) for k, v in by_cat.items()}
    top = sorted(procs.values(), key=lambda d: d["pct"], reverse=True)[:10]
    top_list = [{"name": d["name"], "pct": round(d["pct"], 1), "cat": d["cat"]} for d in top]

    return {
        "ts": datetime.now().isoformat(timespec="seconds"),
        "total_pct": total,
        "by_category": by_cat_rounded,
        "top": top_list,
        "ncores": _CPU_NCORES,
        "interval_s": round(interval, 2),
    }


def _cpu_load_history() -> None:
    if not _CPU_STATE_PATH.exists():
        return
    try:
        with _CPU_STATE_PATH.open("r", encoding="utf-8") as fh:
            lines = fh.readlines()[-_CPU_BUFFER_MAX:]
    except Exception:
        return
    for raw in lines:
        raw = raw.strip()
        if not raw:
            continue
        try:
            _CPU_BUFFER.append(json.loads(raw))
        except Exception:
            continue


def _cpu_persist(sample: dict) -> None:
    """CPU metrics sampler — daemon every 60s. Same rationale que
    `_memory_persist`: fire-and-forget via `_enqueue_background_sql`,
    override con `RAG_METRICS_ASYNC=0`."""
    def _do() -> None:
        with _ragvec_state_conn() as conn:
            _sql_append_event(conn, "rag_cpu_metrics",
                               _map_cpu_row(sample))

    if _metrics_background_default():
        _enqueue_background_sql(_do, "cpu_sql_write_failed")
    else:
        _persist_with_sqlite_retry(_do, "cpu_sql_write_failed")


def _cpu_trim_file() -> None:
    try:
        if not _CPU_STATE_PATH.exists():
            return
        with _CPU_STATE_PATH.open("r", encoding="utf-8") as fh:
            lines = fh.readlines()
        cap = _CPU_BUFFER_MAX * 2
        if len(lines) <= cap:
            return
        with _CPU_STATE_PATH.open("w", encoding="utf-8") as fh:
            fh.writelines(lines[-cap:])
    except Exception:
        pass


@_on_startup
def _start_cpu_sampler() -> None:
    _cpu_load_history()

    def loop() -> None:
        import random as _r
        state: dict = {}
        # Prime the baseline, then wait a full interval before the first
        # real sample so the delta is meaningful.
        _sample_cpu(state)
        # Initial jitter 0-30s para desynchroñar el primer write con el
        # memory-sampler. Ver el comentario equivalente en
        # `_start_memory_sampler` para el razonamiento.
        time.sleep(_r.uniform(0, 30))
        trim_counter = 0
        while True:
            # Per-cycle jitter ±5s; evita la re-sincronización progresiva
            # hacia colisiones con otros samplers.
            time.sleep(_CPU_SAMPLE_INTERVAL + _r.uniform(-5, 5))
            sample = _sample_cpu(state)
            if sample:
                with _CPU_LOCK:
                    _CPU_BUFFER.append(sample)
                _cpu_persist(sample)
                trim_counter += 1
                if trim_counter >= 60:
                    _cpu_trim_file()
                    trim_counter = 0

    threading.Thread(target=loop, name="cpu-sampler", daemon=True).start()


@app.get("/api/system-cpu")
def system_cpu_api(minutes: int = 360) -> dict:
    """Last `minutes` of rag-stack CPU samples + a freshly computed current.

    Values are `% of one core` per category. Can exceed 100 per category
    (sum across multithreaded procs). `ncores` is included so the client
    can render a secondary scale as % of total CPU if desired.
    """
    minutes = max(5, min(minutes, 1440))
    cutoff = datetime.now() - timedelta(minutes=minutes)
    with _CPU_LOCK:
        all_samples = list(_CPU_BUFFER)

    samples = []
    for s in all_samples:
        ts = s.get("ts") or ""
        try:
            if datetime.fromisoformat(ts) >= cutoff:
                samples.append(s)
        except Exception:
            continue

    # Snapshot current via a short-gap sample (sleep in a worker thread
    # would block this sync handler; just use a small synchronous window).
    probe: dict = {}
    _sample_cpu(probe)
    time.sleep(_CPU_LIVE_INTERVAL)
    current = _sample_cpu(probe)

    return {
        "categories": list(_MEMORY_CATEGORIES),
        "samples": samples,
        "current": current,
        "interval_s": _CPU_SAMPLE_INTERVAL,
        "live_interval_s": _CPU_LIVE_INTERVAL,
        "window_minutes": minutes,
        "ncores": _CPU_NCORES,
    }


@app.get("/api/system-cpu/stream")
async def system_cpu_stream(request: Request = None) -> StreamingResponse:  # type: ignore[assignment]
    # Audit 2026-04-26: agregar slot cap (gemelo de system-memory/stream).
    client_ip: str | None = None
    if request is not None:
        client_ip = (request.client.host if request.client else "unknown")
        if not _sse_acquire_slot(client_ip):
            raise HTTPException(status_code=429,
                detail=f"too many concurrent streams (max {_SSE_MAX_PER_IP} per IP)")

    async def gen():
        try:
            state: dict = {}
            await asyncio.to_thread(_sample_cpu, state)
            while True:
                await asyncio.sleep(_CPU_LIVE_INTERVAL)
                sample = await asyncio.to_thread(_sample_cpu, state)
                if not sample:
                    continue
                yield _sse("sample", sample)
        finally:
            if client_ip is not None:
                _sse_release_slot(client_ip)

    return StreamingResponse(
        gen(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache, no-transform",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


@app.get("/api/system-metrics")
def system_metrics_api(hours: int = 24) -> dict:
    """Last `hours` of rows from rag_cpu_metrics + rag_memory_metrics (default 24h).

    Each row has by_category already parsed from JSON.
    Returns empty lists when the tables are missing or the window is empty.
    """
    hours = max(1, min(hours, 168))
    cutoff = (datetime.now() - timedelta(hours=hours)).isoformat(timespec="seconds")
    cpu_rows: list[dict] = []
    mem_rows: list[dict] = []
    try:
        with _ragvec_state_conn() as conn:
            try:
                for r in conn.execute(
                    "SELECT ts, total_pct, ncores, interval_s, by_category_json, top_json"
                    " FROM rag_cpu_metrics WHERE ts >= ? ORDER BY ts ASC",
                    (cutoff,),
                ).fetchall():
                    row: dict = {
                        "ts": r[0], "total_pct": r[1],
                        "ncores": r[2], "interval_s": r[3],
                    }
                    if r[4]:
                        try:
                            row["by_category"] = json.loads(r[4])
                        except Exception:
                            pass
                    if r[5]:
                        try:
                            row["top"] = json.loads(r[5])
                        except Exception:
                            pass
                    cpu_rows.append(row)
            except Exception:
                pass
            try:
                for r in conn.execute(
                    "SELECT ts, total_mb, by_category_json, top_json, vm_json"
                    " FROM rag_memory_metrics WHERE ts >= ? ORDER BY ts ASC",
                    (cutoff,),
                ).fetchall():
                    row = {"ts": r[0], "total_mb": r[1]}
                    if r[2]:
                        try:
                            row["by_category"] = json.loads(r[2])
                        except Exception:
                            pass
                    if r[3]:
                        try:
                            row["top"] = json.loads(r[3])
                        except Exception:
                            pass
                    if r[4]:
                        try:
                            row["vm"] = json.loads(r[4])
                        except Exception:
                            pass
                    mem_rows.append(row)
            except Exception:
                pass
    except Exception:
        pass
    return {"hours": hours, "cpu": cpu_rows, "memory": mem_rows}


if __name__ == "__main__":
    import uvicorn
    # Bind host: default 127.0.0.1 (localhost-only, el estándar). Para
    # exponer el server al LAN (ej. PWA en iPhone, accede por
    # `http://192.168.x.x:8765`), setear OBSIDIAN_RAG_BIND_HOST=0.0.0.0.
    # Ojo: sin auth, cualquiera en el mismo WiFi puede leer el vault —
    # úsalo sólo en red doméstica confiable.
    bind_host = os.environ.get("OBSIDIAN_RAG_BIND_HOST", "127.0.0.1").strip() or "127.0.0.1"
    uvicorn.run(app, host=bind_host, port=8765, log_level="info")
