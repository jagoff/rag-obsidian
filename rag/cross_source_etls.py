"""Cross-source ETLs — extracted from rag/__init__.py 2026-05-04.

Each function writes `.md` notes to the vault so the regular `_run_index`
rglob absorbs them. All helpers follow the same contract:
- Silent-fail: return ``{ok: False, reason: "..."}`` instead of raising.
- Hash-skip: only write if file content changed (``_atomic_write_if_changed``).
- Stats dict return for logging.

Imports from ``rag/__init__.py`` are lazy (inside each function body) to
avoid circular-import issues.
"""
from __future__ import annotations

import contextlib
import json
import os
import re
import subprocess
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path

from rag._constants import _GOOGLE_TOKEN_PATH


def _etl_log_swallow(scope: str, exc: BaseException) -> None:
    """Wrapper local sobre `rag._silent_log` con lazy-import para evitar
    circular import. Cualquier fallo del logger se traga.

    Audit 2026-05-04: pre-fix había 35/59 except silent en este archivo
    con SOLO 1 logueado. Cuando un ETL fallaba (Gmail OAuth caducó, Drive
    API rate-limit, MOZE CSV mal formado, Spotify token expired), no
    quedaba traza en silent_errors.jsonl → debugging a ciegas.
    """
    try:
        from rag import _silent_log  # noqa: PLC0415 — lazy
        _silent_log(scope, exc)
    except Exception:  # pragma: no cover — never re-raise
        pass

__all__ = [
    # MOZE helpers
    "MOZE_BACKUP_DIR",
    "TARJETAS_BACKUP_DIR",
    "MOZE_VAULT_SUBPATH",
    "MOZE_MONTH_ES",
    "_moze_pnum",
    "_moze_fmt_ars",
    "_moze_parse_latest",
    "_moze_render_month",
    "_sync_moze_notes",
    # Credit card helpers
    "TARJETAS_VAULT_SUBPATH",
    "_CARD_BRAND_RE",
    "_CARD_LAST4_RE",
    "_parse_ars_or_usd",
    "_parse_card_date",
    "_parse_credit_card_xlsx",
    "_card_note_filename",
    "_card_render_note",
    "_sync_credit_cards_notes",
    # WhatsApp ETL
    "_WHATSAPP_ETL_SCRIPT",
    "_WHATSAPP_ETL_RE",
    "_sync_whatsapp_notes",
    # External-source ETL constants
    "_REMINDERS_VAULT_SUBPATH",
    "_CALENDAR_VAULT_SUBPATH",
    "_CHROME_VAULT_SUBPATH",
    "_YOUTUBE_VAULT_SUBPATH",
    "_GMAIL_VAULT_SUBPATH",
    "_GDRIVE_VAULT_SUBPATH",
    "_GITHUB_VAULT_SUBPATH",
    "_CLAUDE_VAULT_SUBPATH",
    "_YOUTUBE_TRANSCRIPTS_SUBPATH",
    "_SPOTIFY_VAULT_SUBPATH",
    "_SPOTIFY_CREDS_PATH",
    "_SPOTIFY_TOKEN_PATH",
    "_SPOTIFY_SCOPES",
    "_SPOTIFY_TOP_TTL_DAYS",
    "_GOOGLE_KEYS_CANDIDATES",
    "_GOOGLE_TOKEN_PATH",
    "_GOOGLE_SCOPES",
    # OAuth / Chrome helpers
    "_harden_oauth_cache_perms",
    "_CHROME_HISTORY_PATH",
    "_CHROME_EPOCH_OFFSET_S",
    "_CHROME_SKIP_PREFIXES",
    "_CHROME_SKIP_PATTERNS",
    "_YOUTUBE_WATCH_RE",
    "_atomic_write_if_changed",
    # Reminders + Calendar ETLs
    "_sync_reminders_notes",
    "_sync_apple_calendar_notes",
    # Chrome helpers + ETL
    "_unix_to_chrome_ts",
    "_read_chrome_visits",
    "_sync_chrome_history",
    # Google helpers + ETLs
    "_google_keys_path",
    "_load_google_credentials",
    "_decode_gmail_body",
    "_sync_gmail_notes",
    "_sync_gdrive_notes",
    # GitHub
    "_GH_EVENT_LABELS",
    "_gh_run",
    "_sync_github_activity",
    # Claude Code transcripts
    "_CLAUDE_PROJECTS_ROOT",
    "_CLAUDE_INDEX_WINDOW_DAYS",
    "_CLAUDE_TURN_BODY_CAP",
    "_SECRET_PATTERNS",
    "_redact_secrets",
    "_claude_extract_turn",
    "_sync_claude_code_transcripts",
    # YouTube transcripts
    "_YT_TRANSCRIPT_LANG_PRIORITY",
    "_YT_TRANSCRIPT_BATCH",
    "_YT_VIDEO_ID_RE",
    "_collect_youtube_video_ids",
    "_fetch_yt_transcript_for_index",
    "_sync_youtube_transcripts",
    # Spotify
    "_spotify_client",
    "_sync_spotify_notes",
    # Screen Time
    "SCREENTIME_VAULT_SUBPATH",
    "_SCREENTIME_BACKFILL_DAYS",
    "_SCREENTIME_DAILY_RE",
    "_SCREENTIME_MONTHLY_RE",
    "_sync_screentime_notes",
    "_render_screentime_daily_md",
    "_render_screentime_monthly_md",
    "_render_screentime_index_md",
]

# ── MOZE finanzas ─────────────────────────────────────────────────────────────

# MOZE (Tally4 app) export → vive en su propio container iCloud desde el
# 2026-05-04. Antes compartía dir con los xlsx de tarjetas (CloudDocs/Finances)
# pero el user separó las fuentes. Override: `OBSIDIAN_RAG_MOZE_DIR`.
MOZE_BACKUP_DIR = Path(
    os.environ.get("OBSIDIAN_RAG_MOZE_DIR", "")
    or (Path.home() / "Library/Mobile Documents/iCloud~amoos~Tally4/Documents")
)
# Tarjetas de crédito (xlsx del banco) + PDFs de transferencias siguen en
# `CloudDocs/Finances`. Override: `OBSIDIAN_RAG_FINANCE_DIR` (nombre legacy
# preservado para no romper setups existentes que ya lo tenían apuntando ahí).
TARJETAS_BACKUP_DIR = Path(
    os.environ.get("OBSIDIAN_RAG_FINANCE_DIR", "")
    or (Path.home() / "Library/Mobile Documents/com~apple~CloudDocs/Finances")
)
MOZE_VAULT_SUBPATH = os.environ.get(
    "OBSIDIAN_RAG_MOZE_FOLDER", "02-Areas/Personal/Finanzas/MOZE"
)
MOZE_MONTH_ES = [
    "", "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
]


def _moze_cache_dir() -> Path | None:
    """Cache dir donde `tally4_realm.ensure_moze_csv` deja los CSV
    generados a partir del backup `.zip` de Tally4. Lazy import para no
    cargar el módulo en boot si nunca se necesita."""
    try:
        from rag.integrations.tally4_realm import CACHE_DIR
        return CACHE_DIR
    except Exception:
        return None


def _moze_pnum(s: str) -> float:
    """Parse MOZE numeric column: ES decimals, optional thousand dots."""
    s = (s or "").strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def _moze_fmt_ars(n: float) -> str:
    """$3.470.209 — ES-AR thousands, no decimals."""
    v = int(round(abs(n)))
    s = f"{v:,}".replace(",", ".")
    return f"${s}"


def _moze_parse_latest() -> tuple[Path, list[tuple[datetime, dict]]] | None:
    """Find newest MOZE_*.csv and parse into (date, row) tuples. Dates are
    MM/DD/YYYY; rows with unparseable dates are skipped.

    Post 2026-05-04: si Tally4 dejó un `MOZE_*.zip` más nuevo que el
    último CSV, lo extraemos primero (silent-fail) — ver
    `rag.integrations.tally4_realm`.
    """
    try:
        from rag.integrations import tally4_realm
        tally4_realm.ensure_moze_csv(MOZE_BACKUP_DIR)
    except Exception as exc:
        _etl_log_swallow("moze_tally4_ensure_csv", exc)

    try:
        csvs: list[Path] = []
        for d in (MOZE_BACKUP_DIR, _moze_cache_dir()):
            if d and d.exists():
                csvs.extend(d.glob("MOZE_*.csv"))
        csvs.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    except Exception as exc:
        _etl_log_swallow("moze_csv_listing", exc)
        return None
    if not csvs:
        return None
    src = csvs[0]
    rows: list[tuple[datetime, dict]] = []
    try:
        import csv
        with src.open(newline="", encoding="utf-8") as fh:
            for r in csv.DictReader(fh):
                raw = (r.get("Date") or "").strip()
                if not raw:
                    continue
                try:
                    d = datetime.strptime(raw, "%m/%d/%Y")
                except ValueError:
                    continue
                rows.append((d, r))
    except Exception as exc:
        _etl_log_swallow("moze_csv_parse", exc)
        return None
    return (src, rows) if rows else None


def _moze_render_month(year: int, month: int, rows: list[tuple[datetime, dict]],
                       prev_rows: list[tuple[datetime, dict]],
                       source_name: str, source_mtime: float) -> str:
    """Render one monthly note. `rows`/`prev_rows` are pre-filtered to month.
    Deterministic — same inputs → same output — so the file hash is stable
    and the indexer skips unchanged months.
    """
    from collections import Counter

    def _sum_by(rs, cur_only="ARS"):
        tot = 0.0
        for _, r in rs:
            if (r.get("Currency") or "").strip() != cur_only:
                continue
            if (r.get("Type") or "").strip() != "Expense":
                continue
            tot += abs(_moze_pnum(r.get("Price")))
        return tot

    ars_this = _sum_by(rows, "ARS")
    ars_prev = _sum_by(prev_rows, "ARS")
    usd_this = sum(
        abs(_moze_pnum(r.get("Price")))
        for _, r in rows
        if (r.get("Type") or "").strip() == "Expense"
        and (r.get("Currency") or "").strip() in ("USD", "USDB")
    )
    income_ars = sum(
        abs(_moze_pnum(r.get("Price")))
        for _, r in rows
        if (r.get("Type") or "").strip() == "Income"
        and (r.get("Currency") or "").strip() == "ARS"
    )

    cat_tot: Counter = Counter()
    for _, r in rows:
        if (r.get("Currency") or "").strip() != "ARS":
            continue
        if (r.get("Type") or "").strip() != "Expense":
            continue
        cat = (r.get("Main Category") or "—").strip() or "—"
        cat_tot[cat] += abs(_moze_pnum(r.get("Price")))

    delta_pct = ((ars_this - ars_prev) / ars_prev * 100.0) if ars_prev else None
    month_label_es = f"{MOZE_MONTH_ES[month]} {year}"

    lines: list[str] = []
    lines.append("---")
    lines.append("type: finanzas")
    lines.append("source: MOZE")
    lines.append(f"month: {year:04d}-{month:02d}")
    lines.append(f"gasto_ars: {int(round(ars_this))}")
    if usd_this:
        lines.append(f"gasto_usd: {int(round(usd_this))}")
    if income_ars:
        lines.append(f"ingreso_ars: {int(round(income_ars))}")
    lines.append("tags: [finanzas, moze, gastos]")
    lines.append("ambient: skip")
    lines.append(f"source_file: {source_name}")
    lines.append("---")
    lines.append("")
    lines.append(f"# Finanzas — {month_label_es}")
    lines.append("")

    lines.append("## Resumen")
    lines.append("")
    lines.append(f"- Gasto ARS: **{_moze_fmt_ars(ars_this)}**")
    if ars_prev:
        delta_txt = f"{delta_pct:+.1f}%" if delta_pct is not None else "—"
        lines.append(f"- Mes anterior: {_moze_fmt_ars(ars_prev)} ({delta_txt})")
    if usd_this:
        lines.append(f"- Gasto USD: ${usd_this:,.2f}")
    if income_ars:
        lines.append(f"- Ingreso ARS: {_moze_fmt_ars(income_ars)}")
    lines.append(f"- Transacciones: {sum(1 for _, r in rows if (r.get('Type') or '') in ('Expense','Income'))}")
    lines.append("")

    if cat_tot:
        lines.append("## Top categorías (ARS)")
        lines.append("")
        for name, amt in cat_tot.most_common(10):
            share = (amt / ars_this * 100.0) if ars_this else 0.0
            lines.append(f"- {name}: {_moze_fmt_ars(amt)} ({share:.0f}%)")
        lines.append("")

    lines.append("## Transacciones")
    lines.append("")
    ordered = sorted(rows, key=lambda t: (t[0], t[1].get("Time") or ""))
    for d, r in ordered:
        typ = (r.get("Type") or "").strip()
        if typ not in ("Expense", "Income", "Receivable"):
            continue
        cur = (r.get("Currency") or "").strip()
        cat = (r.get("Main Category") or "").strip()
        sub = (r.get("Subcategory") or "").strip()
        name = (r.get("Name") or "").strip()
        store = (r.get("Store") or "").strip()
        amt = _moze_pnum(r.get("Price"))
        sign = "+" if typ == "Income" else "-"
        amt_str = _moze_fmt_ars(amt) if cur == "ARS" else f"${abs(amt):,.2f}"
        parts = [
            d.strftime("%Y-%m-%d"),
            typ,
            cat,
        ]
        if sub and sub != cat:
            parts.append(sub)
        if name:
            parts.append(name)
        if store:
            parts.append(store)
        parts.append(f"{sign}{amt_str} {cur}")
        lines.append(f"- {' · '.join(parts)}")
    lines.append("")

    return "\n".join(lines)


def _sync_moze_notes(vault_root: Path) -> dict:
    """Regenerate monthly MOZE notes under `{vault_root}/{MOZE_VAULT_SUBPATH}/`.

    Per-month note: `YYYY-MM.md`. Rewrite only if content changed (compare by
    file text) so the indexer's hash-based skip still works. Also writes an
    `_index.md` at the folder root with a summary table of recent months.

    Returns stats for logging. Silent-fail if no CSV is found.
    """
    from collections import defaultdict

    parsed = _moze_parse_latest()
    if not parsed:
        return {"ok": False, "reason": "no_csv"}
    src, rows = parsed

    by_month: dict[tuple[int, int], list[tuple[datetime, dict]]] = defaultdict(list)
    for d, r in rows:
        by_month[(d.year, d.month)].append((d, r))

    target_dir = vault_root / MOZE_VAULT_SUBPATH
    try:
        target_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        return {"ok": False, "reason": f"mkdir: {exc}"}

    months_sorted = sorted(by_month.keys())
    src_mtime = src.stat().st_mtime
    written = 0
    skipped = 0
    current_set = set()

    for i, (year, month) in enumerate(months_sorted):
        fname = f"{year:04d}-{month:02d}.md"
        current_set.add(fname)
        month_rows = by_month[(year, month)]
        prev_rows: list[tuple[datetime, dict]] = []
        if i > 0:
            prev_rows = by_month[months_sorted[i - 1]]
        body = _moze_render_month(year, month, month_rows, prev_rows, src.name, src_mtime)
        path = target_dir / fname
        existing = path.read_text(encoding="utf-8") if path.is_file() else ""
        # Strip `generated_at`-style volatile fields — we don't emit any, so
        # a direct compare is enough; hash-based index skip follows naturally.
        if existing == body:
            skipped += 1
            continue
        path.write_text(body, encoding="utf-8")
        written += 1

    # Prune stale month files (e.g., a month that got fully deleted in MOZE).
    for p in target_dir.glob("*.md"):
        if p.name == "_index.md":
            continue
        if p.name not in current_set:
            try:
                p.unlink()
            except Exception as exc:
                _etl_log_swallow("moze_prune_stale_month", exc)

    # Roll-up index note — gives chat queries like "cuánto gasté este año"
    # a single surface to land on instead of 12 per-month notes.
    idx_lines = [
        "---",
        "type: finanzas",
        "source: MOZE",
        "tags: [finanzas, moze, indice]",
        "ambient: skip",
        f"source_file: {src.name}",
        "---",
        "",
        "# Finanzas — índice mensual (MOZE)",
        "",
        f"Fuente: `{src.name}` (export Money app).",
        "",
        "| Mes | Gasto ARS | Δ vs mes ant |",
        "|---|---:|---:|",
    ]
    prev_tot = 0.0
    for (year, month) in months_sorted:
        mrows = by_month[(year, month)]
        tot = sum(
            abs(_moze_pnum(r.get("Price")))
            for _, r in mrows
            if (r.get("Type") or "").strip() == "Expense"
            and (r.get("Currency") or "").strip() == "ARS"
        )
        delta = ""
        if prev_tot:
            delta_pct = (tot - prev_tot) / prev_tot * 100.0
            delta = f"{delta_pct:+.1f}%"
        idx_lines.append(
            f"| [[{year:04d}-{month:02d}]] | {_moze_fmt_ars(tot)} | {delta} |"
        )
        prev_tot = tot
    idx_body = "\n".join(idx_lines) + "\n"
    idx_path = target_dir / "_index.md"
    if not idx_path.is_file() or idx_path.read_text(encoding="utf-8") != idx_body:
        idx_path.write_text(idx_body, encoding="utf-8")
        written += 1
    else:
        skipped += 1

    return {
        "ok": True,
        "source": src.name,
        "months_total": len(months_sorted),
        "months_written": written,
        "months_skipped": skipped,
        "target": str(target_dir.relative_to(vault_root)),
    }


# ── Resúmenes de tarjeta de crédito → notas mensuales ────────────────────────

TARJETAS_VAULT_SUBPATH = os.environ.get(
    "OBSIDIAN_RAG_TARJETAS_FOLDER", "02-Areas/Personal/Finanzas/Tarjetas"
)

# Regex para extraer marca + últimos 4 del nombre de archivo o sheet name.
_CARD_BRAND_RE = re.compile(
    r"\b(Visa|Master(?:card)?|Amex|American\s*Express|Cabal|Maestro|Naranja)\b",
    re.IGNORECASE,
)
_CARD_LAST4_RE = re.compile(r"(\d{4})(?!\d)")


def _parse_ars_or_usd(raw: object) -> tuple[float | None, str | None]:
    """Parsea celdas tipo `$549.438,75`, `U$S98,93`, `-$926,15`, o numérico
    crudo (openpyxl puede devolver float si la celda tiene formato número).
    Retorna `(amount, currency)` donde currency ∈ {"ARS", "USD", None}.

    Heurística decimal: si el string tiene `,` lo asumimos formato ES
    (decimal coma, miles punto) y normalizamos. Si solo tiene `.` puede ser
    formato US (`24.99`) — solo strippeamos puntos como miles si hay 3
    dígitos exactos después, sino el punto es decimal.
    """
    if raw is None:
        return (None, None)
    if isinstance(raw, (int, float)):
        return (float(raw), None)
    s = str(raw).strip()
    if not s:
        return (None, None)
    cur: str | None = None
    if s.upper().startswith("U$S") or "U$S" in s.upper() or s.upper().startswith("USD"):
        cur = "USD"
    elif s.startswith("$") or "ARS" in s.upper():
        cur = "ARS"
    # Strippeamos símbolos no-numéricos.
    cleaned = re.sub(r"[^\d,.\-]", "", s)
    if "," in cleaned:
        # Formato ES: punto = miles, coma = decimal.
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif cleaned.count(".") > 1:
        # Múltiples puntos → todos son miles ES sin decimal explícito.
        cleaned = cleaned.replace(".", "")
    # Else: un solo punto → decimal US (`24.99`), dejarlo como está.
    try:
        return (float(cleaned), cur)
    except ValueError:
        return (None, cur)


def _parse_card_date(raw: object) -> str | None:
    """DD/MM/YYYY → ISO `YYYY-MM-DD`. None si no parsea. Acepta
    `datetime.date|datetime` (openpyxl normaliza fechas si la celda tiene
    formato fecha) y strings con prefijo (`"Cierre: 26/03/2026"`).
    """
    if raw is None:
        return None
    if hasattr(raw, "strftime"):
        try:
            return raw.strftime("%Y-%m-%d")
        except Exception:
            return None
    s = str(raw).strip()
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if not m:
        return None
    try:
        d = datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        return d.strftime("%Y-%m-%d")
    except ValueError:
        return None


def _parse_credit_card_xlsx(path: Path) -> dict | None:
    """Parsea un `Último resumen - <Marca> <Últimos4>.xlsx` del banco a un
    dict normalizado. None si openpyxl no está disponible o el xlsx no
    tiene la estructura esperada (sin Total a pagar reconocible).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        _etl_log_swallow("credit_card_xlsx_load", exc)
        return None

    rows: list[tuple] = []
    sheet_name = ""
    try:
        ws = wb.active
        sheet_name = ws.title or ""
        # Lectura completa: el archivo es <100 filas — read_only modo
        # streaming, sin riesgo de memoria.
        for row in ws.iter_rows(values_only=True):
            rows.append(row)
    except Exception as exc:
        _etl_log_swallow("credit_card_xlsx_iter_rows", exc)
        return None
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if not rows:
        return None

    # Identificar marca + últimos 4 — primero del sheet name, sino del
    # nombre de archivo, sino del primer row con "terminada en".
    brand = None
    last4 = None
    for source in (sheet_name, path.stem):
        m_brand = _CARD_BRAND_RE.search(source)
        m_last4 = _CARD_LAST4_RE.search(source)
        if m_brand and not brand:
            brand = m_brand.group(1).title()
            if brand.lower() in ("master", "mastercard"):
                brand = "Mastercard"
            elif brand.lower() in ("amex", "american express"):
                brand = "Amex"
        if m_last4 and not last4:
            last4 = m_last4.group(1)
        if brand and last4:
            break

    # Pasada por filas: extraer secciones por anchor text.
    holder = None
    closing_date = None
    due_date = None
    next_closing_date = None
    next_due_date = None
    total_ars = None
    total_usd = None
    minimum_ars = None
    minimum_usd = None
    top_purchases: list[dict] = []

    def _row_text(r: tuple) -> str:
        return " ".join(str(c) for c in r if c is not None).strip()

    n = len(rows)
    i = 0
    in_purchases_block = False
    while i < n:
        row = rows[i]
        text = _row_text(row).lower()

        # Header: "Tarjeta <Marca> ... terminada en NNNN"
        if not last4 and "terminada en" in text:
            m = _CARD_LAST4_RE.search(text)
            if m:
                last4 = m.group(1)
            m_brand = _CARD_BRAND_RE.search(text)
            if m_brand and not brand:
                brand = m_brand.group(1).title()

        # "Fecha de cierre" / "Fecha de vencimiento" — fila siguiente las trae
        if "fecha de cierre" in text and "fecha de vencimiento" in text and i + 1 < n:
            nxt = rows[i + 1]
            closing_date = _parse_card_date(nxt[0] if len(nxt) > 0 else None)
            due_date = _parse_card_date(nxt[1] if len(nxt) > 1 else None)
            i += 2
            continue

        # "Total a pagar" — fila siguiente: ARS | USD
        if "total a pagar" in text and i + 1 < n:
            nxt = rows[i + 1]
            v0, c0 = _parse_ars_or_usd(nxt[0] if len(nxt) > 0 else None)
            v1, c1 = _parse_ars_or_usd(nxt[1] if len(nxt) > 1 else None)
            for v, c in ((v0, c0), (v1, c1)):
                if v is None:
                    continue
                if c == "USD":
                    total_usd = v
                else:
                    total_ars = v if total_ars is None else total_ars
            i += 2
            continue

        # "Mínimo a pagar"
        if "mínimo a pagar" in text or "minimo a pagar" in text:
            if i + 1 < n:
                nxt = rows[i + 1]
                v0, c0 = _parse_ars_or_usd(nxt[0] if len(nxt) > 0 else None)
                v1, c1 = _parse_ars_or_usd(nxt[1] if len(nxt) > 1 else None)
                for v, c in ((v0, c0), (v1, c1)):
                    if v is None:
                        continue
                    if c == "USD":
                        minimum_usd = v
                    else:
                        minimum_ars = v if minimum_ars is None else minimum_ars
                i += 2
                continue

        # "Próximo resumen" — la fila tiene "Cierre: DD/MM/YYYY" en col 2
        if "próximo resumen" in text or "proximo resumen" in text:
            # Las dos siguientes filas tienen "Cierre: ..." y "Vencimiento: ..."
            for j in range(i + 1, min(i + 4, n)):
                trow = _row_text(rows[j]).lower()
                if "cierre" in trow:
                    cell = rows[j][1] if len(rows[j]) > 1 else None
                    next_closing_date = _parse_card_date(cell)
                if "vencimiento" in trow:
                    cell = rows[j][1] if len(rows[j]) > 1 else None
                    next_due_date = _parse_card_date(cell)
            i += 1
            continue

        # Holder: "<Marca> Crédito terminada en NNNN" + col 2 con "(Titular)"
        if holder is None and len(row) > 1 and row[1] and "titular" in str(row[1]).lower():
            holder = re.sub(r"\s*\(Titular\)\s*$", "", str(row[1])).strip() or None

        # Bloques de movimientos: "Pago de tarjeta y devoluciones" (skipear) /
        # "Tarjeta de <holder>" (capturar movimientos hasta "Total de ...")
        if "pago de tarjeta y devoluciones" in text:
            in_purchases_block = False
        elif text.startswith("tarjeta de ") and "terminada en" in text:
            in_purchases_block = True
            if holder is None:
                m = re.search(r"tarjeta de (.+?)\s*-", text, re.IGNORECASE)
                if m:
                    holder = m.group(1).strip().title()
        elif text.startswith("total de ") and ("terminada en" in text or "tarjeta" in text):
            in_purchases_block = False
        elif text.startswith("otros conceptos"):
            in_purchases_block = False
        elif in_purchases_block and len(row) >= 5:
            # Fila de movimiento: (fecha, descripción, cuotas, comprobante, ARS, USD)
            desc = (str(row[1]).strip() if row[1] else "").strip()
            if desc and desc.lower() not in ("descripción", "descripcion"):
                amt_ars, _ = _parse_ars_or_usd(row[4] if len(row) > 4 else None)
                amt_usd, _ = _parse_ars_or_usd(row[5] if len(row) > 5 else None)
                amount = None
                currency = None
                if amt_ars is not None:
                    amount = abs(amt_ars)
                    currency = "ARS"
                elif amt_usd is not None:
                    amount = abs(amt_usd)
                    currency = "USD"
                if amount is not None and amount > 0:
                    top_purchases.append({
                        "date": _parse_card_date(row[0] if len(row) > 0 else None),
                        "description": desc,
                        "amount": amount,
                        "currency": currency,
                    })

        i += 1

    # Sin total ni fechas reconocibles → xlsx no es un resumen válido.
    if total_ars is None and total_usd is None and not closing_date and not due_date:
        return None

    # Top 5 movimientos por monto absoluto, separando ARS/USD.
    ars_purchases = sorted(
        (p for p in top_purchases if p["currency"] == "ARS"),
        key=lambda p: p["amount"],
        reverse=True,
    )[:5]
    usd_purchases = sorted(
        (p for p in top_purchases if p["currency"] == "USD"),
        key=lambda p: p["amount"],
        reverse=True,
    )[:3]

    try:
        mtime = path.stat().st_mtime
    except Exception:
        mtime = 0.0

    return {
        "brand": brand,
        "last4": last4,
        "holder": holder,
        "closing_date": closing_date,
        "due_date": due_date,
        "next_closing_date": next_closing_date,
        "next_due_date": next_due_date,
        "total_ars": total_ars,
        "total_usd": total_usd,
        "minimum_ars": minimum_ars,
        "minimum_usd": minimum_usd,
        "top_purchases_ars": ars_purchases,
        "top_purchases_usd": usd_purchases,
        "source_file": path.name,
        "source_mtime": datetime.fromtimestamp(mtime).isoformat(timespec="seconds") if mtime else None,
        "all_purchases": top_purchases,  # extra para el render de nota
    }


def _card_note_filename(card: dict) -> str | None:
    """Construye el filename de la nota .md para una tarjeta parseada."""
    brand = (card.get("brand") or "").strip()
    last4 = (card.get("last4") or "").strip()
    if not brand or not last4:
        return None
    cycle = card.get("closing_date") or card.get("due_date")
    if not cycle:
        return None
    yyyy_mm = cycle[:7]  # "2026-03-26" → "2026-03"
    return f"Tarjeta-{brand}-{last4}-{yyyy_mm}.md"


def _card_render_note(card: dict) -> str:
    """Renderiza una nota markdown para un resumen de tarjeta parseado."""
    brand = card.get("brand") or "?"
    last4 = card.get("last4") or "????"
    holder = card.get("holder") or ""
    closing = card.get("closing_date") or ""
    due = card.get("due_date") or ""
    next_closing = card.get("next_closing_date") or ""
    next_due = card.get("next_due_date") or ""
    total_ars = card.get("total_ars")
    total_usd = card.get("total_usd")
    minimum_ars = card.get("minimum_ars")
    minimum_usd = card.get("minimum_usd")
    src_file = card.get("source_file") or ""
    cycle_yyyy_mm = (closing or due or "")[:7]

    def _fmt_ars(v: float | None) -> str:
        if v is None:
            return "—"
        s = f"{v:,.2f}"  # `1,234,567.89` → en-US
        return s.replace(",", "_TMP_").replace(".", ",").replace("_TMP_", ".")

    def _fmt_usd(v: float | None) -> str:
        if v is None:
            return "—"
        return f"{v:,.2f}".replace(",", "_TMP_").replace(".", ",").replace("_TMP_", ".")

    fm: list[str] = ["---"]
    fm.append("type: finanzas")
    fm.append("source: tarjeta")
    fm.append(f"brand: {brand}")
    fm.append(f"last4: \"{last4}\"")
    if holder:
        fm.append(f"holder: {holder}")
    if cycle_yyyy_mm:
        fm.append(f"cycle: {cycle_yyyy_mm}")
    if closing:
        fm.append(f"closing_date: {closing}")
    if due:
        fm.append(f"due_date: {due}")
    if next_closing:
        fm.append(f"next_closing_date: {next_closing}")
    if next_due:
        fm.append(f"next_due_date: {next_due}")
    if total_ars is not None:
        fm.append(f"total_ars: {total_ars:.2f}")
    if total_usd is not None:
        fm.append(f"total_usd: {total_usd:.2f}")
    if minimum_ars is not None:
        fm.append(f"minimum_ars: {minimum_ars:.2f}")
    if minimum_usd is not None:
        fm.append(f"minimum_usd: {minimum_usd:.2f}")
    tags = ["finanzas", "tarjeta", brand.lower()]
    fm.append(f"tags: [{', '.join(tags)}]")
    fm.append("ambient: skip")
    if src_file:
        fm.append(f"source_file: {src_file}")
    fm.append("---")
    fm.append("")

    title = f"# Tarjeta {brand} ·{last4} — ciclo {cycle_yyyy_mm or 'sin fecha'}"
    body: list[str] = [title, ""]
    if holder:
        body.append(f"Titular: **{holder}**.")
        body.append("")

    parts: list[str] = []
    if total_ars is not None:
        parts.append(f"total ARS **${_fmt_ars(total_ars)}**")
    if total_usd is not None:
        parts.append(f"total USD **U$S{_fmt_usd(total_usd)}**")
    if parts:
        body.append("Total a pagar este ciclo: " + " · ".join(parts) + ".")
    if closing or due:
        body.append(f"Fecha de cierre: **{closing or '—'}** · Fecha de vencimiento: **{due or '—'}**.")
    if minimum_ars is not None or minimum_usd is not None:
        mins: list[str] = []
        if minimum_ars is not None:
            mins.append(f"ARS **${_fmt_ars(minimum_ars)}**")
        if minimum_usd is not None:
            mins.append(f"USD **U$S{_fmt_usd(minimum_usd)}**")
        body.append("Mínimo a pagar: " + " · ".join(mins) + ".")
    if next_closing or next_due:
        body.append(f"Próximo ciclo — cierre: {next_closing or '—'} · vencimiento: {next_due or '—'}.")
    body.append("")

    ars = card.get("top_purchases_ars") or []
    if ars:
        body.append("## Top movimientos ARS")
        body.append("")
        body.append("| Fecha | Descripción | Monto |")
        body.append("|---|---|---:|")
        for p in ars:
            d = p.get("date") or ""
            desc = (p.get("description") or "").replace("|", "\\|")
            amt = p.get("amount")
            body.append(f"| {d} | {desc} | ${_fmt_ars(amt)} |")
        body.append("")

    usd = card.get("top_purchases_usd") or []
    if usd:
        body.append("## Top movimientos USD")
        body.append("")
        body.append("| Fecha | Descripción | Monto |")
        body.append("|---|---|---:|")
        for p in usd:
            d = p.get("date") or ""
            desc = (p.get("description") or "").replace("|", "\\|")
            amt = p.get("amount")
            body.append(f"| {d} | {desc} | U$S{_fmt_usd(amt)} |")
        body.append("")

    all_p = card.get("all_purchases") or []
    if all_p:
        body.append("## Todos los movimientos")
        body.append("")
        for p in all_p:
            d = p.get("date") or ""
            desc = (p.get("description") or "").replace("|", "\\|")
            cur = p.get("currency") or ""
            amt = p.get("amount")
            if cur == "USD":
                body.append(f"- {d} — {desc} — U$S{_fmt_usd(amt)}")
            else:
                body.append(f"- {d} — {desc} — ${_fmt_ars(amt)}")
        body.append("")

    body.append(f"_Fuente: `{src_file}` (export del banco)._")
    body.append("")

    return "\n".join(fm + body)


def _sync_credit_cards_notes(vault_root: Path) -> dict:
    """Regenerate per-cycle credit-card notes from the xlsx exports in
    `TARJETAS_BACKUP_DIR` (`CloudDocs/Finances` por default). Mirrors the
    MOZE pattern: one note per (card, cycle), hash-skip if content is
    unchanged, prune notes whose source xlsx no longer exists.
    """
    try:
        seen: set[Path] = set()
        for pattern in ("Último resumen*.xlsx", "Ultimo resumen*.xlsx"):
            for p in TARJETAS_BACKUP_DIR.glob(pattern):
                seen.add(p)
        files = sorted(seen, key=lambda p: p.name)
    except Exception as exc:
        return {"ok": False, "reason": f"glob: {exc}"}
    if not files:
        return {"ok": False, "reason": "no_xlsx"}

    target_dir = vault_root / TARJETAS_VAULT_SUBPATH
    try:
        _vr = vault_root.resolve()
        _td = target_dir.resolve()
        _td.relative_to(_vr)
    except Exception:
        return {"ok": False, "reason": f"target escapa vault: {target_dir}"}
    try:
        target_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        return {"ok": False, "reason": f"mkdir: {exc}"}

    written = 0
    skipped = 0
    parse_failed = 0
    current_set: set[str] = set()

    for xlsx in files:
        card = _parse_credit_card_xlsx(xlsx)
        if not card:
            parse_failed += 1
            continue
        fname = _card_note_filename(card)
        if not fname:
            parse_failed += 1
            continue
        current_set.add(fname)
        body = _card_render_note(card)
        path = target_dir / fname
        existing = path.read_text(encoding="utf-8") if path.is_file() else ""
        if existing == body:
            skipped += 1
            continue
        path.write_text(body, encoding="utf-8")
        written += 1

    for p in target_dir.glob("Tarjeta-*.md"):
        if p.name not in current_set:
            try:
                p.unlink()
            except Exception as exc:
                _etl_log_swallow("credit_card_prune_stale", exc)

    if not current_set and parse_failed:
        return {"ok": False, "reason": "no_parsed"}

    return {
        "ok": True,
        "files_total": len(files),
        "files_written": written,
        "files_skipped": skipped,
        "files_parse_failed": parse_failed,
        "target": str(target_dir.relative_to(vault_root)),
    }


# ── WhatsApp ETL ──────────────────────────────────────────────────────────────

_WHATSAPP_ETL_SCRIPT = Path.home() / ".local/bin/whatsapp-to-vault"
_WHATSAPP_ETL_RE = re.compile(
    r"wrote\s+(\d+)\s+files,\s+(\d+)\s+unchanged,\s+(\d+)\s+\(chat, month\)\s+buckets,\s+(\d+)\s+chats"
)


def _sync_whatsapp_notes(vault_root: Path) -> dict:
    """Trigger the WhatsApp → vault ETL script and parse its summary line.

    Mirrors the MOZE pre-index pattern: produces `.md` files in
    `<vault>/03-Resources/WhatsApp/<chat>/YYYY-MM.md` so the regular rglob
    picks them up. Subprocess to keep it as a single source of truth — the
    same script that the `com.fer.whatsapp-vault-sync` launchd plist runs
    every 15 min. Silent-fail when the script is missing (other machines).
    """
    if not _WHATSAPP_ETL_SCRIPT.is_file():
        return {"ok": False, "reason": "script_missing"}
    try:
        proc = subprocess.run(
            [str(_WHATSAPP_ETL_SCRIPT)],
            capture_output=True, timeout=60, text=True,
        )
    except (subprocess.TimeoutExpired, FileNotFoundError, OSError) as exc:
        return {"ok": False, "reason": str(exc)[:120]}
    out = (proc.stdout or "").strip()
    err = (proc.stderr or "").strip()
    if proc.returncode != 0:
        return {"ok": False, "reason": err[:160] or f"rc={proc.returncode}"}
    m = _WHATSAPP_ETL_RE.search(out)
    if not m:
        return {"ok": True, "raw": out[:160]}
    return {
        "ok": True,
        "files_written": int(m.group(1)),
        "files_unchanged": int(m.group(2)),
        "buckets": int(m.group(3)),
        "chats": int(m.group(4)),
        "target": "03-Resources/WhatsApp",
    }


# ── External-source ETLs ──────────────────────────────────────────────────────
# Same pattern as MOZE / WhatsApp: produce `.md` files inside the vault so the
# regular `_run_index` rglob absorbs them. Each helper is silent-fail and
# returns a stats dict for logging. Triggered from `_run_index` after the
# WhatsApp sync, before the vault scan.

_REMINDERS_VAULT_SUBPATH = "03-Resources/Reminders"
_CALENDAR_VAULT_SUBPATH = "03-Resources/Calendar"
_CHROME_VAULT_SUBPATH = "03-Resources/Chrome"
_YOUTUBE_VAULT_SUBPATH = "03-Resources/YouTube"
_GMAIL_VAULT_SUBPATH = "03-Resources/Gmail"
_GDRIVE_VAULT_SUBPATH = "03-Resources/GoogleDrive"
_GITHUB_VAULT_SUBPATH = "03-Resources/GitHub"
_CLAUDE_VAULT_SUBPATH = "03-Resources/Claude"
_YOUTUBE_TRANSCRIPTS_SUBPATH = "03-Resources/YouTube/transcripts"
_SPOTIFY_VAULT_SUBPATH = "03-Resources/Spotify"
_SPOTIFY_CREDS_PATH = Path.home() / ".config/obsidian-rag/spotify_client.json"
_SPOTIFY_TOKEN_PATH = Path.home() / ".config/obsidian-rag/spotify_token.json"
_SPOTIFY_SCOPES = "user-read-recently-played user-top-read"
_SPOTIFY_TOP_TTL_DAYS = 7  # weekly refresh of _top.md

# OAuth keys: reuse the gmail-mcp client config so the user doesn't manage two
# Google Cloud OAuth apps. Token is stored in our own config dir so the
# scopes (gmail + drive readonly) are independent of gmail-mcp's own token.
_GOOGLE_KEYS_CANDIDATES = (
    Path.home() / ".config/obsidian-rag/google_credentials.json",
    Path.home() / ".gmail-mcp/gcp-oauth.keys.json",
)
_GOOGLE_SCOPES = (
    "https://www.googleapis.com/auth/gmail.readonly",
    "https://www.googleapis.com/auth/drive.readonly",
)


def _harden_oauth_cache_perms() -> None:
    """One-shot chmod of OAuth cache files + their containing dir."""
    cfg_dir = Path.home() / ".config/obsidian-rag"
    if cfg_dir.is_dir():
        try:
            cfg_dir.chmod(0o700)
        except OSError as exc:
            _etl_log_swallow("oauth_cache_chmod_dir", exc)
    for tok in (_GOOGLE_TOKEN_PATH, _SPOTIFY_TOKEN_PATH):
        if tok.is_file():
            try:
                os.chmod(tok, 0o600)
            except OSError as exc:
                _etl_log_swallow("oauth_cache_chmod_token", exc)


_harden_oauth_cache_perms()

_CHROME_HISTORY_PATH = Path.home() / "Library/Application Support/Google/Chrome/Default/History"
# Chrome epoch is 1601-01-01 UTC microseconds (Windows FILETIME).
_CHROME_EPOCH_OFFSET_S = 11644473600
# URL prefixes / patterns we never want indexed — they're navigation noise.
_CHROME_SKIP_PREFIXES = (
    "chrome://", "chrome-extension://", "about:", "edge://", "view-source:",
    "data:", "javascript:", "file:///",
)
_CHROME_SKIP_PATTERNS = (
    re.compile(r"^https?://(www\.)?google\.[^/]+/search\?"),
    re.compile(r"^https?://(www\.)?google\.[^/]+/url\?"),
    re.compile(r"^https?://(www\.)?bing\.com/search\?"),
    re.compile(r"^https?://(duckduckgo\.com|search\.brave\.com)/\?"),
)
_YOUTUBE_WATCH_RE = re.compile(r"^https?://(www\.|m\.)?youtube\.com/watch\?(?:.*&)?v=([\w\-]+)")


def _atomic_write_if_changed(target: Path, body: str) -> bool:
    """Write `body` to `target` only if its contents changed. Returns True on
    write, False on skip. Indexing relies on hash-skip — rewriting bytes that
    haven't changed forces re-embed for nothing.
    """
    target.parent.mkdir(parents=True, exist_ok=True)
    if target.exists():
        try:
            if target.read_text(encoding="utf-8") == body:
                return False
        except OSError:
            pass
    tmp = target.with_suffix(target.suffix + ".tmp")
    tmp.write_text(body, encoding="utf-8")
    os.replace(tmp, target)
    return True


def _sync_reminders_notes(vault_root: Path) -> dict:
    """Snapshot Apple Reminders to a daily note. Pending only, horizon 180 days
    + undated. Completed-reminders fetch is intentionally NOT included.
    """
    from rag import _apple_enabled, _fetch_reminders_due  # lazy
    if not _apple_enabled():
        return {"ok": False, "reason": "apple_disabled"}
    now = datetime.now()
    pending = _fetch_reminders_due(now, horizon_days=180, max_items=500)
    if not pending:
        return {"ok": True, "files_written": 0, "reason": "no_data"}

    by_bucket: dict[str, list[dict]] = {}
    for item in pending:
        by_bucket.setdefault(item["bucket"], []).append(item)

    today = now.strftime("%Y-%m-%d")
    fm_lines = [
        "---",
        "source: apple-reminders",
        f"snapshot_at: {now.isoformat(timespec='seconds')}",
        f"pending_count: {len(pending)}",
        "tags:",
        "- apple-reminders",
        "- system-snapshot",
        "---",
        "",
        f"# Apple Reminders — {today}",
        "",
    ]
    body_lines: list[str] = list(fm_lines)
    for bucket_key, label in (
        ("overdue", "Overdue"),
        ("today", "Hoy"),
        ("upcoming", "Próximos"),
        ("undated", "Sin fecha"),
    ):
        items = by_bucket.get(bucket_key) or []
        if not items:
            continue
        body_lines.append(f"## {label} ({len(items)})")
        body_lines.append("")
        for it in items:
            due = it["due"] or "—"
            list_tag = f" `[{it['list']}]`" if it.get("list") else ""
            body_lines.append(f"- **{it['name']}** · {due}{list_tag}")
        body_lines.append("")
    body = "\n".join(body_lines)

    target = vault_root / _REMINDERS_VAULT_SUBPATH / f"{today}.md"
    written = _atomic_write_if_changed(target, body)
    return {
        "ok": True,
        "files_written": 1 if written else 0,
        "pending": len(pending),
        "completed": 0,
        "target": _REMINDERS_VAULT_SUBPATH,
    }


def _sync_apple_calendar_notes(vault_root: Path, days_ahead: int = 90) -> dict:
    """Snapshot upcoming Apple Calendar events to per-week notes. Requires
    icalBuddy (`brew install ical-buddy`); returns silently when missing.
    """
    from rag import _apple_enabled, _icalbuddy_path, _fetch_calendar_ahead  # lazy
    if not _apple_enabled():
        return {"ok": False, "reason": "apple_disabled"}
    if not _icalbuddy_path():
        return {"ok": False, "reason": "icalbuddy_missing"}
    events = _fetch_calendar_ahead(days_ahead=days_ahead, max_events=200)
    if not events:
        return {"ok": True, "files_written": 0, "reason": "no_events"}
    now = datetime.now()
    iso_year, iso_week, _ = now.isocalendar()
    week_label = f"{iso_year}-W{iso_week:02d}"

    fm_lines = [
        "---",
        "source: apple-calendar",
        f"snapshot_at: {now.isoformat(timespec='seconds')}",
        f"window_days: {days_ahead}",
        f"event_count: {len(events)}",
        "tags:",
        "- apple-calendar",
        "- system-snapshot",
        "---",
        "",
        f"# Calendar — semana {week_label} (próximos {days_ahead}d)",
        "",
    ]
    body_lines: list[str] = list(fm_lines)
    current_label = None
    for ev in events:
        label = ev.get("date_label") or "(sin fecha)"
        if label != current_label:
            body_lines.append(f"## {label}")
            body_lines.append("")
            current_label = label
        time_range = ev.get("time_range") or ""
        time_part = f"`{time_range}` · " if time_range else ""
        body_lines.append(f"- {time_part}{ev.get('title', '(sin título)')}")
    body_lines.append("")
    body = "\n".join(body_lines)

    target = vault_root / _CALENDAR_VAULT_SUBPATH / f"{week_label}.md"
    written = _atomic_write_if_changed(target, body)
    return {
        "ok": True,
        "files_written": 1 if written else 0,
        "events": len(events),
        "target": _CALENDAR_VAULT_SUBPATH,
    }


def _unix_to_chrome_ts(unix_s: float) -> int:
    return int((unix_s + _CHROME_EPOCH_OFFSET_S) * 1_000_000)


def _read_chrome_visits(history_db: Path, hours: int = 48) -> list[dict]:
    """Read distinct URLs visited in the last `hours` from Chrome History.
    Chrome locks the SQLite while the browser runs — we copy to /tmp and read
    the snapshot. Empty list on any error.
    """
    from rag import _chrome_to_unix_ts  # lazy — defined in integrations.chrome_bookmarks
    if not history_db.is_file():
        return []
    import shutil
    import sqlite3 as _sqlite3
    import tempfile
    tmp = Path(tempfile.gettempdir()) / "obsidian-rag-chrome-history.db"
    try:
        shutil.copy2(history_db, tmp)
    except OSError:
        return []
    try:
        con = _sqlite3.connect(f"file:{tmp}?mode=ro", uri=True)
        con.row_factory = _sqlite3.Row
        cutoff = _unix_to_chrome_ts(time.time() - hours * 3600)
        rows = con.execute(
            "SELECT url, title, visit_count, last_visit_time "
            "FROM urls WHERE last_visit_time > ? "
            "ORDER BY last_visit_time DESC",
            (cutoff,),
        ).fetchall()
        con.close()
    except _sqlite3.Error:
        return []
    finally:
        try:
            tmp.unlink()
        except OSError as exc:
            _etl_log_swallow("chrome_history_tmp_unlink", exc)

    out: list[dict] = []
    seen: set[str] = set()
    for r in rows:
        url = (r["url"] or "").strip()
        if not url or url in seen:
            continue
        if any(url.startswith(p) for p in _CHROME_SKIP_PREFIXES):
            continue
        if any(p.match(url) for p in _CHROME_SKIP_PATTERNS):
            continue
        seen.add(url)
        out.append({
            "url": url,
            "title": (r["title"] or "").strip() or url,
            "visit_count": int(r["visit_count"] or 0),
            "ts": _chrome_to_unix_ts(int(r["last_visit_time"] or 0)),
        })
    return out


def _sync_chrome_history(vault_root: Path, hours: int = 48) -> dict:
    """Daily snapshot of Chrome history (last `hours`, dedup by exact URL).
    Also derives a YouTube-only note from URLs matching watch?v=… so YouTube
    activity surfaces independently in retrieval. Hash-skipped when content
    matches the existing day file.
    """
    import sys as _sys
    _chrome_hist_path = getattr(_sys.modules.get("rag"), "_CHROME_HISTORY_PATH", _CHROME_HISTORY_PATH)
    visits = _read_chrome_visits(_chrome_hist_path, hours=hours)
    if not visits:
        return {"ok": False, "reason": "no_visits_or_chrome_locked"}
    now = datetime.now()
    today = now.strftime("%Y-%m-%d")

    chrome_fm = [
        "---",
        "source: chrome-history",
        f"snapshot_at: {now.isoformat(timespec='seconds')}",
        f"window_hours: {hours}",
        f"url_count: {len(visits)}",
        "tags:",
        "- chrome-history",
        "- system-snapshot",
        "---",
        "",
        f"# Chrome history — {today} (últimas {hours}h)",
        "",
    ]
    chrome_lines: list[str] = list(chrome_fm)
    for v in visits:
        ts = datetime.fromtimestamp(v["ts"]).strftime("%H:%M")
        title = v["title"].replace("|", "·")
        chrome_lines.append(f"- `{ts}` [{title}]({v['url']})")
    chrome_body = "\n".join(chrome_lines) + "\n"

    chrome_target = vault_root / _CHROME_VAULT_SUBPATH / f"{today}.md"
    chrome_written = _atomic_write_if_changed(chrome_target, chrome_body)

    yt_videos: list[dict] = []
    seen_vid: set[str] = set()
    for v in visits:
        m = _YOUTUBE_WATCH_RE.match(v["url"])
        if not m:
            continue
        vid = m.group(2)
        if vid in seen_vid:
            continue
        seen_vid.add(vid)
        yt_videos.append({
            "video_id": vid,
            "title": v["title"],
            "url": f"https://www.youtube.com/watch?v={vid}",
            "ts": v["ts"],
        })

    yt_written = 0
    if yt_videos:
        yt_fm = [
            "---",
            "source: youtube-via-chrome",
            f"snapshot_at: {now.isoformat(timespec='seconds')}",
            f"window_hours: {hours}",
            f"video_count: {len(yt_videos)}",
            "tags:",
            "- youtube",
            "- system-snapshot",
            "---",
            "",
            f"# YouTube watched — {today} (últimas {hours}h, vía Chrome)",
            "",
        ]
        yt_lines: list[str] = list(yt_fm)
        for v in yt_videos:
            ts = datetime.fromtimestamp(v["ts"]).strftime("%H:%M")
            title = v["title"].replace("|", "·")
            yt_lines.append(f"- `{ts}` [{title}]({v['url']})")
        yt_body = "\n".join(yt_lines) + "\n"
        yt_target = vault_root / _YOUTUBE_VAULT_SUBPATH / f"{today}.md"
        yt_written = 1 if _atomic_write_if_changed(yt_target, yt_body) else 0

    return {
        "ok": True,
        "files_written": (1 if chrome_written else 0) + yt_written,
        "urls": len(visits),
        "youtube_videos": len(yt_videos),
        "target": _CHROME_VAULT_SUBPATH,
    }


def _google_keys_path() -> Path | None:
    for p in _GOOGLE_KEYS_CANDIDATES:
        if p.is_file():
            return p
    return None


def _load_google_credentials(allow_interactive: bool = True) -> "google.oauth2.credentials.Credentials | None":
    """Return Google OAuth `Credentials` for Gmail + Drive (readonly), or None.

    Lookup order: cached token → refresh if expired → first-time interactive
    browser flow (only when `allow_interactive` and stdin is a TTY). Token is
    persisted to `_GOOGLE_TOKEN_PATH` so subsequent runs are silent.
    """
    from rag import _silent_log, _write_secret_file  # lazy
    try:
        from google.oauth2.credentials import Credentials
        from google.auth.transport.requests import Request
        from google_auth_oauthlib.flow import InstalledAppFlow
    except ImportError:
        return None

    creds = None
    if _GOOGLE_TOKEN_PATH.is_file():
        try:
            creds = Credentials.from_authorized_user_file(
                str(_GOOGLE_TOKEN_PATH), list(_GOOGLE_SCOPES)
            )
        except Exception:
            creds = None
    if creds and creds.valid:
        return creds
    if creds and creds.expired and creds.refresh_token:
        try:
            creds.refresh(Request())
            _write_secret_file(_GOOGLE_TOKEN_PATH, creds.to_json())
            return creds
        except Exception as exc:
            _silent_log('google_token_refresh', exc)
    if not allow_interactive or not sys.stdin.isatty():
        return None
    keys = _google_keys_path()
    if not keys:
        return None
    try:
        flow = InstalledAppFlow.from_client_secrets_file(str(keys), list(_GOOGLE_SCOPES))
        creds = flow.run_local_server(port=0, open_browser=True)
    except Exception as exc:
        _etl_log_swallow("google_oauth_flow_failed", exc)
        return None
    _write_secret_file(_GOOGLE_TOKEN_PATH, creds.to_json())
    return creds


def _decode_gmail_body(payload: dict) -> str:
    """Walk a Gmail API `payload` tree, prefer text/plain, fall back to HTML
    stripped of tags. Returns empty string when the message has no body parts.
    """
    import base64
    def _decode(data: str) -> str:
        try:
            return base64.urlsafe_b64decode(data.encode("ascii")).decode("utf-8", errors="ignore")
        except Exception:
            return ""

    def _walk(node: dict, want_mime: str) -> str:
        if node.get("mimeType") == want_mime and (node.get("body") or {}).get("data"):
            return _decode(node["body"]["data"])
        for child in node.get("parts") or []:
            found = _walk(child, want_mime)
            if found:
                return found
        return ""

    plain = _walk(payload, "text/plain")
    if plain:
        return plain
    html = _walk(payload, "text/html")
    if not html:
        return ""
    # Drop <style> + <script> block contents before stripping tags.
    html = re.sub(
        r"<(style|script)\b[^>]*>.*?</\1\s*>", " ", html,
        flags=re.DOTALL | re.IGNORECASE,
    )
    return re.sub(r"<[^>]+>", " ", html)


def _sync_gmail_notes(vault_root: Path, hours: int = 48, max_messages: int = 30, body_cap: int = 5000) -> dict:
    """Snapshot recent Gmail to a daily note. Subject + headers + body (capped)
    per message. Hash-skipped when content unchanged.
    """
    import sys as _sys
    _cred_fn = getattr(_sys.modules.get("rag"), "_load_google_credentials", _load_google_credentials)
    creds = _cred_fn()
    if creds is None:
        return {"ok": False, "reason": "no_google_credentials"}
    try:
        from googleapiclient.discovery import build
    except ImportError:
        return {"ok": False, "reason": "google_api_missing"}
    try:
        gm = build("gmail", "v1", credentials=creds, cache_discovery=False)
        days = max(1, int((hours + 23) // 24))
        resp = gm.users().messages().list(
            userId="me", q=f"newer_than:{days}d", maxResults=max_messages,
        ).execute()
        ids = [m["id"] for m in (resp.get("messages") or [])]
    except Exception as exc:
        return {"ok": False, "reason": f"list_failed: {str(exc)[:120]}"}
    if not ids:
        return {"ok": True, "files_written": 0, "reason": "no_messages"}

    messages: list[dict] = []
    for mid in ids:
        try:
            msg = gm.users().messages().get(
                userId="me", id=mid, format="full",
            ).execute()
        except Exception as exc:
            _etl_log_swallow("gmail_message_fetch", exc)
            continue
        headers = {h["name"].lower(): h["value"] for h in (msg.get("payload", {}).get("headers") or [])}
        body = _decode_gmail_body(msg.get("payload") or {})
        body = re.sub(r"\s+", " ", body).strip()[:body_cap]
        messages.append({
            "id": mid,
            "subject": headers.get("subject", "(sin subject)"),
            "from": headers.get("from", "?"),
            "date": headers.get("date", ""),
            "snippet": (msg.get("snippet") or "").strip(),
            "body": body,
        })

    now = datetime.now()
    today = now.strftime("%Y-%m-%d")
    fm = [
        "---",
        "source: gmail",
        f"snapshot_at: {now.isoformat(timespec='seconds')}",
        f"window_hours: {hours}",
        f"message_count: {len(messages)}",
        "tags:",
        "- gmail",
        "- system-snapshot",
        "---",
        "",
        f"# Gmail — {today} (últimas {hours}h)",
        "",
    ]
    for m in messages:
        fm.append(f"## {m['subject']}")
        fm.append("")
        fm.append(f"**From:** {m['from']}  ")
        fm.append(f"**Date:** {m['date']}  ")
        if m["snippet"]:
            fm.append(f"**Snippet:** {m['snippet']}")
        fm.append("")
        if m["body"]:
            fm.append(m["body"])
            fm.append("")
    body_text = "\n".join(fm) + "\n"
    target = vault_root / _GMAIL_VAULT_SUBPATH / f"{today}.md"
    written = _atomic_write_if_changed(target, body_text)
    return {
        "ok": True,
        "files_written": 1 if written else 0,
        "messages": len(messages),
        "target": _GMAIL_VAULT_SUBPATH,
    }


def _sync_gdrive_notes(vault_root: Path, hours: int = 48, max_docs: int = 4, body_cap: int = 8000) -> dict:
    """Snapshot the last `max_docs` Google Docs/Sheets/Slides modified in the
    window. Title + exported text body per doc. Hash-skipped.
    """
    import sys as _sys
    _cred_fn = getattr(_sys.modules.get("rag"), "_load_google_credentials", _load_google_credentials)
    creds = _cred_fn()
    if creds is None:
        return {"ok": False, "reason": "no_google_credentials"}
    try:
        from googleapiclient.discovery import build
    except ImportError:
        return {"ok": False, "reason": "google_api_missing"}
    try:
        dv = build("drive", "v3", credentials=creds, cache_discovery=False)
        cutoff = (datetime.now(timezone.utc) - timedelta(hours=hours)).strftime("%Y-%m-%dT%H:%M:%SZ")
        mime_filter = " or ".join(
            f"mimeType = '{m}'" for m in (
                "application/vnd.google-apps.document",
                "application/vnd.google-apps.spreadsheet",
                "application/vnd.google-apps.presentation",
            )
        )
        q = f"(modifiedTime > '{cutoff}') and ({mime_filter}) and trashed = false"
        resp = dv.files().list(
            q=q, orderBy="modifiedTime desc", pageSize=max_docs,
            fields="files(id, name, mimeType, modifiedTime, owners(displayName), webViewLink)",
        ).execute()
        files = resp.get("files") or []
    except Exception as exc:
        return {"ok": False, "reason": f"list_failed: {str(exc)[:120]}"}
    if not files:
        return {"ok": True, "files_written": 0, "reason": "no_docs"}

    EXPORT_MIME = {
        "application/vnd.google-apps.document": "text/plain",
        "application/vnd.google-apps.spreadsheet": "text/csv",
        "application/vnd.google-apps.presentation": "text/plain",
    }
    docs: list[dict] = []
    for f in files:
        export_mime = EXPORT_MIME.get(f["mimeType"], "text/plain")
        try:
            body = dv.files().export(fileId=f["id"], mimeType=export_mime).execute()
            if isinstance(body, bytes):
                body = body.decode("utf-8", errors="ignore")
            body = body.strip()[:body_cap]
        except Exception:
            body = ""
        docs.append({
            "id": f["id"],
            "name": f.get("name", "(sin nombre)"),
            "mime": f["mimeType"].split(".")[-1],
            "modified": f.get("modifiedTime", ""),
            "owner": (f.get("owners") or [{}])[0].get("displayName", "?"),
            "link": f.get("webViewLink", ""),
            "body": body,
        })

    now = datetime.now()
    today = now.strftime("%Y-%m-%d")
    fm = [
        "---",
        "source: google-drive",
        f"snapshot_at: {now.isoformat(timespec='seconds')}",
        f"window_hours: {hours}",
        f"doc_count: {len(docs)}",
        "tags:",
        "- google-drive",
        "- system-snapshot",
        "---",
        "",
        f"# Google Drive — {today} (últimos {len(docs)} docs últimas {hours}h)",
        "",
    ]
    for d in docs:
        fm.append(f"## {d['name']}")
        fm.append("")
        fm.append(f"**Tipo:** {d['mime']} · **Modificado:** {d['modified']} · **Owner:** {d['owner']}")
        if d["link"]:
            fm.append(f"**Link:** {d['link']}")
        fm.append("")
        if d["body"]:
            fm.append(d["body"])
            fm.append("")
    body_text = "\n".join(fm) + "\n"
    target = vault_root / _GDRIVE_VAULT_SUBPATH / f"{today}.md"
    written = _atomic_write_if_changed(target, body_text)
    return {
        "ok": True,
        "files_written": 1 if written else 0,
        "docs": len(docs),
        "target": _GDRIVE_VAULT_SUBPATH,
    }


# ── GitHub activity ───────────────────────────────────────────────────────────

_GH_EVENT_LABELS = {
    "PushEvent": "push",
    "PullRequestEvent": "pull-request",
    "IssueCommentEvent": "issue-comment",
    "IssuesEvent": "issue",
    "PullRequestReviewEvent": "pr-review",
    "PullRequestReviewCommentEvent": "pr-review-comment",
    "CreateEvent": "create",
    "DeleteEvent": "delete",
    "ForkEvent": "fork",
    "WatchEvent": "star",
    "ReleaseEvent": "release",
}


def _gh_run(args: list[str], timeout: float = 10.0) -> tuple[int, str, str]:
    """Run a `gh` command. Returns (rc, stdout, stderr)."""
    try:
        proc = subprocess.run(
            ["gh", *args], capture_output=True, timeout=timeout, text=True,
        )
    except (FileNotFoundError, subprocess.TimeoutExpired, OSError) as exc:
        return 127, "", str(exc)[:160]
    return proc.returncode, proc.stdout or "", proc.stderr or ""


def _sync_github_activity(vault_root: Path, hours: int = 48) -> dict:
    """Snapshot recent GitHub activity (push/PR/issues/stars) plus open PRs.
    Uses the already-authenticated `gh` CLI; silent-fail on missing/unauth.
    """
    rc, login_out, _ = _gh_run(["api", "user", "--jq", ".login"])
    if rc != 0:
        return {"ok": False, "reason": "gh_unavailable_or_unauth"}
    user = login_out.strip()
    if not user:
        return {"ok": False, "reason": "gh_no_login"}

    rc, events_raw, err = _gh_run(["api", f"users/{user}/events?per_page=100"])
    if rc != 0:
        return {"ok": False, "reason": f"events_failed: {err[:120]}"}
    try:
        events = json.loads(events_raw)
    except json.JSONDecodeError:
        return {"ok": False, "reason": "events_parse_failed"}

    cutoff = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(hours=hours)
    fresh: list[dict] = []
    for ev in events:
        ts_raw = ev.get("created_at", "")
        try:
            ts = datetime.strptime(ts_raw, "%Y-%m-%dT%H:%M:%SZ")
        except ValueError:
            continue
        if ts < cutoff:
            continue
        fresh.append({
            "ts": ts,
            "type": ev.get("type", "?"),
            "repo": (ev.get("repo") or {}).get("name", "?"),
            "payload": ev.get("payload") or {},
        })

    rc, pr_raw, _ = _gh_run([
        "api", "search/issues",
        "-X", "GET",
        "-f", f"q=is:pr is:open author:{user}",
    ])
    open_prs: list[dict] = []
    if rc == 0:
        try:
            for it in (json.loads(pr_raw).get("items") or [])[:20]:
                open_prs.append({
                    "title": it.get("title", "?"),
                    "url": it.get("html_url", ""),
                    "repo": (it.get("repository_url", "") or "").split("/repos/")[-1],
                    "number": it.get("number"),
                })
        except json.JSONDecodeError:
            pass

    if not fresh and not open_prs:
        return {"ok": True, "files_written": 0, "reason": "no_activity"}

    by_type: dict[str, list[dict]] = {}
    for ev in fresh:
        by_type.setdefault(ev["type"], []).append(ev)

    today = datetime.now().strftime("%Y-%m-%d")
    fm = [
        "---",
        "source: github",
        f"snapshot_date: {today}",
        f"window_hours: {hours}",
        f"event_count: {len(fresh)}",
        f"open_pr_count: {len(open_prs)}",
        "tags:",
        "- github",
        "- system-snapshot",
        "---",
        "",
        f"# GitHub activity — {today} (últimas {hours}h, usuario {user})",
        "",
    ]
    for ev_type, items in sorted(by_type.items()):
        label = _GH_EVENT_LABELS.get(ev_type, ev_type)
        fm.append(f"## {label} ({len(items)})")
        fm.append("")
        for ev in items:
            ts = ev["ts"].strftime("%Y-%m-%d %H:%M")
            repo = ev["repo"]
            p = ev["payload"]
            detail = ""
            if ev_type == "PushEvent":
                commits = p.get("commits") or []
                msgs = " · ".join((c.get("message", "").split("\n", 1)[0])[:80] for c in commits[:3])
                detail = f"{len(commits)} commit(s) {msgs}"
            elif ev_type in ("PullRequestEvent", "PullRequestReviewEvent", "PullRequestReviewCommentEvent"):
                pr = p.get("pull_request") or {}
                detail = f"{p.get('action','?')} #{pr.get('number','?')} {pr.get('title','')[:80]}"
            elif ev_type == "IssuesEvent":
                iss = p.get("issue") or {}
                detail = f"{p.get('action','?')} #{iss.get('number','?')} {iss.get('title','')[:80]}"
            elif ev_type == "IssueCommentEvent":
                iss = p.get("issue") or {}
                detail = f"comentó #{iss.get('number','?')} {iss.get('title','')[:80]}"
            elif ev_type == "WatchEvent":
                detail = "starred"
            elif ev_type == "CreateEvent":
                detail = f"creó {p.get('ref_type','?')} {p.get('ref','') or ''}"
            elif ev_type == "ReleaseEvent":
                rel = p.get("release") or {}
                detail = f"release {rel.get('tag_name','')}"
            else:
                detail = ""
            fm.append(f"- `{ts}` {repo} — {detail}")
        fm.append("")

    if open_prs:
        fm.append(f"## Open PRs ({len(open_prs)})")
        fm.append("")
        for pr in open_prs:
            fm.append(f"- {pr['repo']}#{pr['number']} [{pr['title']}]({pr['url']})")
        fm.append("")

    body = "\n".join(fm) + "\n"
    target = vault_root / _GITHUB_VAULT_SUBPATH / f"{today}.md"
    written = _atomic_write_if_changed(target, body)
    return {
        "ok": True,
        "files_written": 1 if written else 0,
        "events": len(fresh),
        "open_prs": len(open_prs),
        "target": _GITHUB_VAULT_SUBPATH,
    }


# ── Claude Code transcripts ───────────────────────────────────────────────────

_CLAUDE_PROJECTS_ROOT = Path.home() / ".claude/projects"
_CLAUDE_INDEX_WINDOW_DAYS = 30
_CLAUDE_TURN_BODY_CAP = 8000

_SECRET_PATTERNS = [
    (re.compile(r"sk-(?:proj-|ant-)?[A-Za-z0-9_\-]{20,}"), "[REDACTED-OPENAI/ANTHROPIC]"),
    (re.compile(r"ghp_[A-Za-z0-9]{30,}"),                  "[REDACTED-GH-PAT]"),
    (re.compile(r"github_pat_[A-Za-z0-9_]{20,}"),          "[REDACTED-GH-PAT-NEW]"),
    (re.compile(r"AKIA[A-Z0-9]{16}"),                      "[REDACTED-AWS-KEY]"),
    (re.compile(r"AIza[A-Za-z0-9_\-]{35}"),                "[REDACTED-GOOGLE-API]"),
    (re.compile(r"xox[baprs]-[A-Za-z0-9\-]{10,}"),         "[REDACTED-SLACK]"),
    (re.compile(r"(?i)(?<![A-Za-z0-9_])(?:api[_-]?key|secret|password|token)\s*[:=]\s*[\"']?[A-Za-z0-9_\-./]{16,}"),
                                                            "[REDACTED-KV]"),
]


def _redact_secrets(text: str) -> str:
    for pat, rep in _SECRET_PATTERNS:
        text = pat.sub(rep, text)
    return text


def _claude_extract_turn(record: dict) -> tuple[str, str, str] | None:
    """Pull (role, ts, body) from one Claude Code transcript line. Returns
    None when the record is internal (tool result, summary, etc.) and
    shouldn't be rendered as a chat turn.
    """
    rec_type = record.get("type") or ""
    msg = record.get("message") or {}
    if rec_type not in ("user", "assistant"):
        return None
    role = msg.get("role") or rec_type
    ts = (record.get("timestamp") or "").replace("T", " ").split(".")[0]
    content = msg.get("content")
    if isinstance(content, str):
        body = content
    elif isinstance(content, list):
        parts = []
        for block in content:
            if not isinstance(block, dict):
                continue
            if block.get("type") == "text" and block.get("text"):
                parts.append(block["text"])
            elif block.get("type") == "tool_use":
                tool = block.get("name", "?")
                parts.append(f"[tool_use:{tool}]")
            elif block.get("type") == "tool_result":
                parts.append("[tool_result]")
        body = "\n".join(parts)
    else:
        body = ""
    body = _redact_secrets(body.strip())
    import sys as _sys
    _body_cap = getattr(_sys.modules.get("rag"), "_CLAUDE_TURN_BODY_CAP", _CLAUDE_TURN_BODY_CAP)
    if len(body) > _body_cap:
        body = body[:_body_cap] + "\n\n[…body truncado]"
    if not body:
        return None
    return role, ts, body


def _sync_claude_code_transcripts(vault_root: Path) -> dict:
    """Convert Claude Code session JSONL → per-session markdown. Walks
    `~/.claude/projects/<slug>/*.jsonl` modified within the last 30 days,
    redacts common secret shapes, hash-skips via `_atomic_write_if_changed`.
    """
    import sys as _sys
    _rag = _sys.modules.get("rag")
    _projects_root = getattr(_rag, "_CLAUDE_PROJECTS_ROOT", _CLAUDE_PROJECTS_ROOT)
    if not _projects_root.is_dir():
        return {"ok": False, "reason": "no_claude_projects_dir"}
    cutoff_mtime = time.time() - (_CLAUDE_INDEX_WINDOW_DAYS * 86400)
    written = 0
    total = 0
    skipped = 0
    for project_dir in sorted(_projects_root.iterdir()):
        if not project_dir.is_dir():
            continue
        for jsonl in sorted(project_dir.glob("*.jsonl")):
            try:
                stat = jsonl.stat()
            except OSError:
                continue
            if stat.st_mtime < cutoff_mtime:
                continue
            total += 1
            try:
                lines = jsonl.read_text(encoding="utf-8", errors="ignore").splitlines()
            except OSError:
                continue
            turns: list[tuple[str, str, str]] = []
            for line in lines:
                if not line.strip():
                    continue
                try:
                    rec = json.loads(line)
                except json.JSONDecodeError:
                    continue
                t = _claude_extract_turn(rec)
                if t:
                    turns.append(t)
            if not turns:
                skipped += 1
                continue
            session_id = jsonl.stem
            started = turns[0][1] or "?"
            ended = turns[-1][1] or "?"
            fm = [
                "---",
                "source: claude-code",
                f"project: {project_dir.name}",
                f"session_id: {session_id}",
                f"started_at: {started}",
                f"ended_at: {ended}",
                f"turn_count: {len(turns)}",
                "tags:",
                "- claude-code",
                "- system-snapshot",
                "---",
                "",
                f"# Claude Code session — {project_dir.name} / {session_id}",
                "",
            ]
            for role, ts, body in turns:
                fm.append(f"## {role} · {ts}")
                fm.append("")
                fm.append(body)
                fm.append("")
            body_text = "\n".join(fm) + "\n"
            target = vault_root / _CLAUDE_VAULT_SUBPATH / project_dir.name / f"{session_id}.md"
            if _atomic_write_if_changed(target, body_text):
                written += 1
            else:
                skipped += 1

    # Prune vault transcripts whose source JSONL is gone or older than the window.
    pruned = 0
    vault_claude_dir = vault_root / _CLAUDE_VAULT_SUBPATH
    if vault_claude_dir.is_dir():
        for project_vault_dir in vault_claude_dir.iterdir():
            if not project_vault_dir.is_dir():
                continue
            source_project = _projects_root / project_vault_dir.name
            for md_file in project_vault_dir.glob("*.md"):
                session_id = md_file.stem
                source_jsonl = source_project / f"{session_id}.jsonl"
                stale = False
                if not source_jsonl.is_file():
                    stale = True
                else:
                    try:
                        if source_jsonl.stat().st_mtime < cutoff_mtime:
                            stale = True
                    except OSError:
                        stale = True
                if stale:
                    with contextlib.suppress(OSError):
                        md_file.unlink()
                        pruned += 1
            with contextlib.suppress(OSError):
                if not any(project_vault_dir.iterdir()):
                    project_vault_dir.rmdir()

    if not total and not pruned:
        return {"ok": True, "files_written": 0, "reason": "no_recent_sessions"}
    return {
        "ok": True,
        "files_written": written,
        "sessions_seen": total,
        "skipped": skipped,
        "pruned": pruned,
        "target": _CLAUDE_VAULT_SUBPATH,
    }


# ── YouTube transcripts ───────────────────────────────────────────────────────

_YT_TRANSCRIPT_LANG_PRIORITY = ("es", "es-419", "en", "en-US")
_YT_TRANSCRIPT_BATCH = 10
_YT_VIDEO_ID_RE = re.compile(r"youtube\.com/watch\?v=([\w\-]{6,})")


def _collect_youtube_video_ids(vault_root: Path) -> list[tuple[str, str]]:
    """Read recent YouTube daily notes, pull (video_id, title) pairs."""
    yt_dir = vault_root / "03-Resources/YouTube"
    if not yt_dir.is_dir():
        return []
    seen: set[str] = set()
    out: list[tuple[str, str]] = []
    for md in sorted(yt_dir.glob("*.md")):
        try:
            text = md.read_text(encoding="utf-8")
        except OSError:
            continue
        for line in text.splitlines():
            m = _YT_VIDEO_ID_RE.search(line)
            if not m:
                continue
            vid = m.group(1)
            if vid in seen:
                continue
            seen.add(vid)
            title_match = re.search(r"\[([^\]]+)\]\(", line)
            title = title_match.group(1).strip() if title_match else vid
            out.append((vid, title))
    return out


def _fetch_yt_transcript_for_index(video_id: str) -> tuple[str, str] | None:
    """Returns (lang, transcript_text) or None on miss."""
    try:
        from youtube_transcript_api import YouTubeTranscriptApi
    except ImportError:
        return None
    try:
        api = YouTubeTranscriptApi()
        listing = api.list(video_id)
    except Exception as exc:
        _etl_log_swallow("yt_transcript_list", exc)
        return None
    transcript = None
    chosen_lang = None
    for lang in _YT_TRANSCRIPT_LANG_PRIORITY:
        try:
            transcript = listing.find_transcript([lang])
            chosen_lang = lang
            break
        except Exception:
            # Per-lang miss esperado — el video no tiene ese idioma. NO
            # loggear (sino se llena el log con noise por cada lang miss).
            continue
    if transcript is None:
        try:
            transcript = next(iter(listing))
            chosen_lang = transcript.language_code
        except Exception as exc:
            _etl_log_swallow("yt_transcript_iter_fallback", exc)
            return None
    try:
        fetched = transcript.fetch()
    except Exception as exc:
        _etl_log_swallow("yt_transcript_fetch", exc)
        return None
    snippets = getattr(fetched, "snippets", None) or fetched
    parts = [getattr(s, "text", None) or s.get("text", "") for s in snippets]
    text = " ".join(p for p in parts if p).strip()
    if not text:
        return None
    return chosen_lang or "?", text


def _sync_youtube_transcripts(vault_root: Path, batch: int = _YT_TRANSCRIPT_BATCH) -> dict:
    """For each video referenced in recent YouTube daily notes, fetch its
    transcript once. Caps at `batch` per run.
    """
    import sys as _sys
    _yt_fetch = getattr(_sys.modules.get("rag"), "_fetch_yt_transcript_for_index", _fetch_yt_transcript_for_index)
    videos = _collect_youtube_video_ids(vault_root)
    if not videos:
        return {"ok": True, "files_written": 0, "reason": "no_videos"}
    target_dir = vault_root / _YOUTUBE_TRANSCRIPTS_SUBPATH
    target_dir.mkdir(parents=True, exist_ok=True)
    written = 0
    fetched = 0
    failed = 0
    for vid, title in videos:
        target = target_dir / f"{vid}.md"
        if target.is_file():
            continue
        if fetched >= batch:
            break
        result = _yt_fetch(vid)
        fetched += 1
        if not result:
            failed += 1
            continue
        lang, text = result
        url = f"https://www.youtube.com/watch?v={vid}"
        body = (
            "---\n"
            "source: youtube-transcript\n"
            f"video_id: {vid}\n"
            f"language: {lang}\n"
            f"url: {url}\n"
            "tags:\n"
            "- youtube-transcript\n"
            "- system-snapshot\n"
            "---\n\n"
            f"# {title}\n\n"
            f"{url}\n\n"
            f"{text}\n"
        )
        if _atomic_write_if_changed(target, body):
            written += 1
    return {
        "ok": True,
        "files_written": written,
        "fetched_this_run": fetched,
        "failed_this_run": failed,
        "videos_known": len(videos),
        "target": _YOUTUBE_TRANSCRIPTS_SUBPATH,
    }


# ── Spotify ───────────────────────────────────────────────────────────────────

def _spotify_client(allow_interactive: bool = True) -> "spotipy.Spotify | None":
    """Return an authenticated `spotipy.Spotify` instance, or None."""
    import sys as _sys
    _rag = _sys.modules.get("rag")
    _creds_path = getattr(_rag, "_SPOTIFY_CREDS_PATH", _SPOTIFY_CREDS_PATH)
    _token_path = getattr(_rag, "_SPOTIFY_TOKEN_PATH", _SPOTIFY_TOKEN_PATH)
    if not _creds_path.is_file():
        return None
    try:
        import spotipy
        from spotipy.oauth2 import SpotifyOAuth
    except ImportError:
        return None
    try:
        creds = json.loads(_creds_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    cid = creds.get("client_id")
    secret = creds.get("client_secret")
    redirect = creds.get("redirect_uri", "http://localhost:8888/callback")
    if not (cid and secret):
        return None
    open_browser = bool(allow_interactive and sys.stdin.isatty())
    auth = SpotifyOAuth(
        client_id=cid, client_secret=secret, redirect_uri=redirect,
        scope=_SPOTIFY_SCOPES, cache_path=str(_token_path),
        open_browser=open_browser,
    )
    try:
        token = auth.get_cached_token()
        if not token or auth.is_token_expired(token):
            if not open_browser:
                return None
            token = auth.get_access_token(as_dict=True)
        if not token:
            return None
        try:
            if _token_path.is_file():
                os.chmod(_token_path, 0o600)
        except OSError as exc:
            _etl_log_swallow("spotify_token_chmod", exc)
        return spotipy.Spotify(auth=token["access_token"])
    except Exception as exc:
        _etl_log_swallow("spotify_oauth_token", exc)
        return None


def _sync_spotify_notes(vault_root: Path, max_recent: int = 50) -> dict:
    """Snapshot Spotify recently-played + (weekly) top tracks."""
    import sys as _sys
    _rag = _sys.modules.get("rag")
    _token_path = getattr(_rag, "_SPOTIFY_TOKEN_PATH", _SPOTIFY_TOKEN_PATH)
    _sp_client = getattr(_rag, "_spotify_client", _spotify_client)
    if not _token_path.is_file():
        return {"ok": False, "reason": "no_spotify_token"}
    sp = _sp_client(allow_interactive=False)
    if sp is None:
        return {"ok": False, "reason": "no_spotify_credentials"}

    try:
        recent = sp.current_user_recently_played(limit=max_recent)
    except Exception as exc:
        return {"ok": False, "reason": f"recent_failed: {str(exc)[:120]}"}

    items = recent.get("items") or []
    now = datetime.now()
    today = now.strftime("%Y-%m-%d")

    fm = [
        "---",
        "source: spotify",
        f"snapshot_date: {today}",
        f"track_count: {len(items)}",
        "tags:",
        "- spotify",
        "- system-snapshot",
        "---",
        "",
        f"# Spotify recently played — {today}",
        "",
    ]
    for it in items:
        track = it.get("track") or {}
        name = track.get("name") or "(sin título)"
        artists = ", ".join(a.get("name", "?") for a in (track.get("artists") or []))
        album = (track.get("album") or {}).get("name", "")
        played = (it.get("played_at") or "").replace("T", " ").split(".")[0]
        url = (track.get("external_urls") or {}).get("spotify", "")
        link = f"[{name}]({url})" if url else name
        fm.append(f"- `{played}` {link} — {artists}{f' · _{album}_' if album else ''}")
    body = "\n".join(fm) + "\n"
    target = vault_root / _SPOTIFY_VAULT_SUBPATH / f"{today}.md"
    written_recent = _atomic_write_if_changed(target, body)

    top_target = vault_root / _SPOTIFY_VAULT_SUBPATH / "_top.md"
    written_top = 0
    needs_top = (
        not top_target.is_file()
        or (time.time() - top_target.stat().st_mtime) > _SPOTIFY_TOP_TTL_DAYS * 86400
    )
    if needs_top:
        try:
            top_tracks = sp.current_user_top_tracks(limit=20, time_range="short_term")
            top_artists = sp.current_user_top_artists(limit=20, time_range="short_term")
        except Exception:
            top_tracks = top_artists = None
        if top_tracks and top_artists:
            t_items = top_tracks.get("items") or []
            a_items = top_artists.get("items") or []
            tfm = [
                "---",
                "source: spotify-top",
                f"refreshed_date: {today}",
                "window: short_term (4 weeks)",
                "tags:",
                "- spotify",
                "- system-snapshot",
                "---",
                "",
                "# Spotify Top — últimas 4 semanas",
                "",
                f"## Top tracks ({len(t_items)})",
                "",
            ]
            for t in t_items:
                artists = ", ".join(a.get("name", "?") for a in (t.get("artists") or []))
                url = (t.get("external_urls") or {}).get("spotify", "")
                name = t.get("name", "?")
                link = f"[{name}]({url})" if url else name
                tfm.append(f"- {link} — {artists}")
            tfm += ["", f"## Top artists ({len(a_items)})", ""]
            for a in a_items:
                url = (a.get("external_urls") or {}).get("spotify", "")
                name = a.get("name", "?")
                genres = ", ".join((a.get("genres") or [])[:3])
                link = f"[{name}]({url})" if url else name
                tfm.append(f"- {link}{f' · {genres}' if genres else ''}")
            top_body = "\n".join(tfm) + "\n"
            if _atomic_write_if_changed(top_target, top_body):
                written_top = 1

    return {
        "ok": True,
        "files_written": (1 if written_recent else 0) + written_top,
        "recently_played": len(items),
        "refreshed_top": bool(written_top),
        "target": _SPOTIFY_VAULT_SUBPATH,
    }


# ── Screen Time persistence (daily + monthly) ─────────────────────────────────

SCREENTIME_VAULT_SUBPATH = "03-Resources/Screentime"
_SCREENTIME_BACKFILL_DAYS = 30
_SCREENTIME_DAILY_RE = re.compile(r"^\d{4}-\d{2}-\d{2}\.md$")
_SCREENTIME_MONTHLY_RE = re.compile(r"^\d{4}-\d{2}\.md$")


def _sync_screentime_notes(
    vault_root: Path,
    days: int = _SCREENTIME_BACKFILL_DAYS,
    db_path: Path | None = None,
) -> dict:
    """Persist Screen Time per-app foreground usage as vault notes."""
    from collections import defaultdict
    from datetime import datetime, timedelta
    from rag import _collect_screentime  # lazy

    target_dir = vault_root / SCREENTIME_VAULT_SUBPATH
    try:
        target_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        return {"ok": False, "reason": f"mkdir: {exc}"}

    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    day_data: dict[str, dict] = {}  # "YYYY-MM-DD" → screentime dict
    months: dict[str, list[str]] = defaultdict(list)  # "YYYY-MM" → ["YYYY-MM-DD", ...]

    db_was_available = False
    for d in range(days, -1, -1):
        day_start = today - timedelta(days=d)
        day_end = day_start + timedelta(days=1)
        st = _collect_screentime(day_start, day_end, db_path=db_path)
        if not st.get("available"):
            if d == days and not (target_dir / "_index.md").is_file():
                return {"ok": False, "reason": "no_data"}
            continue
        db_was_available = True
        if int(st.get("total_secs") or 0) < 60:
            continue
        day_str = day_start.strftime("%Y-%m-%d")
        month_str = day_start.strftime("%Y-%m")
        day_data[day_str] = st
        months[month_str].append(day_str)

    if not db_was_available:
        return {"ok": False, "reason": "no_data"}
    if not day_data:
        return {"ok": True, "files_written": 0, "days_total": 0,
                "target": str(target_dir.relative_to(vault_root))}

    written = 0
    skipped = 0
    current_set: set[str] = set()

    # Daily notes
    for day_str, st in day_data.items():
        body = _render_screentime_daily_md(day_str, st)
        path = target_dir / f"{day_str}.md"
        current_set.add(path.name)
        existing = path.read_text(encoding="utf-8") if path.is_file() else ""
        if existing == body:
            skipped += 1
            continue
        path.write_text(body, encoding="utf-8")
        written += 1

    # Monthly aggregates
    for month_str, day_list in sorted(months.items()):
        body = _render_screentime_monthly_md(
            month_str,
            [(d, day_data[d]) for d in sorted(day_list)],
        )
        path = target_dir / f"{month_str}.md"
        current_set.add(path.name)
        existing = path.read_text(encoding="utf-8") if path.is_file() else ""
        if existing == body:
            skipped += 1
            continue
        path.write_text(body, encoding="utf-8")
        written += 1

    # Index — tabla mensual rolling
    idx_body = _render_screentime_index_md(months, day_data)
    idx_path = target_dir / "_index.md"
    current_set.add(idx_path.name)
    existing = idx_path.read_text(encoding="utf-8") if idx_path.is_file() else ""
    if existing != idx_body:
        idx_path.write_text(idx_body, encoding="utf-8")
        written += 1
    else:
        skipped += 1

    # Prune días que ya cayeron fuera de la ventana de backfill (>30d).
    for p in target_dir.glob("*.md"):
        if p.name in current_set:
            continue
        if _SCREENTIME_DAILY_RE.match(p.name):
            continue
        if _SCREENTIME_MONTHLY_RE.match(p.name):
            continue

    total_secs = sum(int(st.get("total_secs") or 0) for st in day_data.values())
    return {
        "ok": True,
        "files_written": written,
        "files_skipped": skipped,
        "days_total": len(day_data),
        "months_total": len(months),
        "total_secs": total_secs,
        "target": str(target_dir.relative_to(vault_root)),
    }


def _render_screentime_daily_md(day_str: str, st: dict) -> str:
    """Daily note: top apps + categorías. Determinístico para que el
    hash-skip funcione."""
    from rag import _fmt_hm  # lazy
    total = int(st.get("total_secs") or 0)
    top_apps = (st.get("top_apps") or [])[:10]
    cats = st.get("categories") or {}

    lines = [
        "---",
        "type: screentime",
        f"date: {day_str}",
        f"total_active_secs: {total}",
        "ambient: skip",
        "tags: [screentime, productividad]",
        "---",
        "",
        f"# Pantalla · {day_str} · {_fmt_hm(total)} activo",
        "",
        "## Top apps",
    ]
    if top_apps:
        for a in top_apps:
            lines.append(f"- {a['label']} · {_fmt_hm(int(a['secs']))}")
    else:
        lines.append("- _sin actividad registrada_")

    if cats:
        lines.append("")
        lines.append("## Por categoría")
        order = ["code", "notas", "comms", "browser", "media", "otros"]
        for k in order:
            v = int(cats.get(k, 0) or 0)
            if v >= 60:
                lines.append(f"- {k} · {_fmt_hm(v)}")

    return "\n".join(lines) + "\n"


def _render_screentime_monthly_md(month_str: str, days: list[tuple[str, dict]]) -> str:
    """Monthly aggregate: top apps del mes + por categoría + tabla diaria."""
    from collections import defaultdict
    from rag import _fmt_hm  # lazy

    total_secs = sum(int(st.get("total_secs") or 0) for _, st in days)
    apps_total: dict[str, dict] = defaultdict(lambda: {"label": "", "secs": 0})
    cats_total: dict[str, int] = defaultdict(int)
    for _day, st in days:
        for a in (st.get("top_apps") or []):
            bundle = a.get("bundle", "")
            apps_total[bundle]["label"] = a.get("label", bundle)
            apps_total[bundle]["secs"] += int(a.get("secs") or 0)
        for k, v in (st.get("categories") or {}).items():
            cats_total[k] += int(v or 0)

    top_apps_sorted = sorted(apps_total.items(), key=lambda kv: -kv[1]["secs"])[:15]

    lines = [
        "---",
        "type: screentime-monthly",
        f"month: {month_str}",
        f"total_active_secs: {total_secs}",
        f"days_active: {len(days)}",
        "ambient: skip",
        "tags: [screentime, productividad]",
        "---",
        "",
        f"# Pantalla · {month_str} · {_fmt_hm(total_secs)} activo ({len(days)} días)",
        "",
        "## Top apps del mes",
    ]
    for _bundle, info in top_apps_sorted:
        lines.append(f"- {info['label']} · {_fmt_hm(info['secs'])}")

    lines.append("")
    lines.append("## Por categoría")
    order = ["code", "notas", "comms", "browser", "media", "otros"]
    for k in order:
        v = int(cats_total.get(k, 0))
        if v >= 60:
            lines.append(f"- {k} · {_fmt_hm(v)}")

    lines.append("")
    lines.append("## Por día")
    lines.append("| Día | Total | Top app | Top categoría |")
    lines.append("|---|---|---|---|")
    for day_str, st in days:
        total = int(st.get("total_secs") or 0)
        top = (st.get("top_apps") or [{}])[0]
        top_label = top.get("label", "—")
        top_secs = int(top.get("secs") or 0)
        cats = st.get("categories") or {}
        top_cat = max(cats.items(), key=lambda kv: kv[1])[0] if cats else "—"
        lines.append(
            f"| [[{day_str}]] | {_fmt_hm(total)} | "
            f"{top_label} ({_fmt_hm(top_secs)}) | {top_cat} |"
        )

    return "\n".join(lines) + "\n"


def _render_screentime_index_md(
    months: dict[str, list[str]], day_data: dict[str, dict]
) -> str:
    """Index note — tabla mensual con totales + top categoría."""
    from collections import defaultdict
    from rag import _fmt_hm  # lazy

    lines = [
        "---",
        "type: screentime-index",
        "ambient: skip",
        "tags: [screentime, indice, productividad]",
        "---",
        "",
        "# Pantalla — índice mensual",
        "",
        "Fuente: `~/Library/Application Support/Knowledge/knowledgeC.db` (CoreDuet).",
        "Ventana: macOS retiene ~30 días — estas notas persisten lo histórico.",
        "",
        "| Mes | Total activo | Días | Top categoría |",
        "|---|---|---:|---|",
    ]
    for month_str in sorted(months.keys()):
        day_list = months[month_str]
        total = sum(int(day_data[d].get("total_secs") or 0) for d in day_list)
        cats_total: dict[str, int] = defaultdict(int)
        for d in day_list:
            for k, v in (day_data[d].get("categories") or {}).items():
                cats_total[k] += int(v or 0)
        top_cat = (
            max(cats_total.items(), key=lambda kv: kv[1])[0] if cats_total else "—"
        )
        lines.append(
            f"| [[{month_str}]] | {_fmt_hm(total)} | {len(day_list)} | {top_cat} |"
        )
    return "\n".join(lines) + "\n"
