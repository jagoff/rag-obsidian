"""Chat upload file handling helpers.

This module owns upload-oriented parsing/sanitizing logic. ``web.server`` keeps
the FastAPI routes for compatibility, but delegates the file work here so the
server module stays closer to an HTTP orchestration layer.
"""
from __future__ import annotations

import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as _ET

from fastapi import HTTPException, UploadFile

# Register HEIC/HEIF reader in PIL once. Without this iPhone photos (HEIC
# default) are passthrough in ``sanitize_image_exif`` and keep GPS coords when
# copied to the iCloud vault.
try:
    import pillow_heif as _pillow_heif  # noqa: PLC0415

    _pillow_heif.register_heif_opener()
    HEIC_AVAILABLE = True
except ImportError:
    HEIC_AVAILABLE = False


CHAT_UPLOAD_TEXT_MAX_CHARS = 12_000
CHAT_UPLOAD_TEXT_SUFFIXES = {
    ".md", ".markdown", ".txt", ".log", ".csv", ".tsv", ".json", ".jsonl",
    ".yaml", ".yml", ".xml", ".html", ".htm", ".css", ".js", ".jsx",
    ".ts", ".tsx", ".py", ".sh", ".bash", ".zsh", ".sql", ".tf",
    ".tfvars", ".hcl", ".ini", ".cfg", ".conf", ".env",
}
CHAT_UPLOAD_DOC_SUFFIXES = {".pdf", ".docx", ".xlsx"}
CHAT_UPLOAD_IMAGE_SUFFIXES = {
    ".jpg", ".jpeg", ".png", ".heic", ".heif", ".webp", ".gif",
}

# suffix -> PIL format. HEIC/HEIF are added only when pillow-heif is available.
SANITIZABLE_FORMATS: dict[str, str] = {
    ".jpg": "JPEG",
    ".jpeg": "JPEG",
    ".png": "PNG",
    ".webp": "WEBP",
    ".gif": "GIF",
}
if HEIC_AVAILABLE:
    SANITIZABLE_FORMATS[".heic"] = "JPEG"
    SANITIZABLE_FORMATS[".heif"] = "JPEG"


def safe_chat_upload_filename(name: str, fallback: str = "archivo") -> str:
    base = Path(name or fallback).name.strip() or fallback
    base = re.sub(r"[\x00-\x1f/\\:]+", "-", base)
    base = re.sub(r"\s+", " ", base).strip(" .")
    return base[:160] or fallback


def sanitize_image_exif(
    raw_bytes: bytes,
    suffix: str,
    *,
    sanitizable_formats: dict[str, str] | None = None,
) -> bytes:
    """Re-encode an image without EXIF/GPS/metadata when supported."""
    formats = sanitizable_formats if sanitizable_formats is not None else SANITIZABLE_FORMATS
    fmt = formats.get(suffix.lower())
    if fmt is None:
        return raw_bytes
    try:
        import io  # noqa: PLC0415
        from PIL import Image  # noqa: PLC0415

        with Image.open(io.BytesIO(raw_bytes)) as img:
            img.load()
            stripped = Image.frombytes(img.mode, img.size, img.tobytes())
            buf = io.BytesIO()
            stripped.save(buf, format=fmt)
            return buf.getvalue()
    except Exception:
        return raw_bytes


def truncate_chat_upload_text(
    text: str,
    *,
    max_chars: int = CHAT_UPLOAD_TEXT_MAX_CHARS,
) -> tuple[str, bool]:
    text = (text or "").replace("\x00", "").strip()
    if len(text) <= max_chars:
        return text, False
    marker = (
        "\n\n[archivo truncado para el chat: se incluyeron los primeros "
        f"{max_chars} caracteres]"
    )
    keep = max(0, max_chars - len(marker))
    return text[:keep].rstrip() + marker, True


def decode_chat_upload_text(raw: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "utf-16", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def looks_like_binary(raw: bytes) -> bool:
    if not raw:
        return False
    sample = raw[:4096]
    if b"\x00" in sample:
        return True
    control = sum(1 for b in sample if b < 32 and b not in (9, 10, 13))
    return control / max(1, len(sample)) > 0.08


def extract_pdf_text_for_chat(path: Path) -> str:
    try:
        import pypdf  # noqa: PLC0415
    except ImportError:
        return ""
    try:
        reader = pypdf.PdfReader(path)
        parts: list[str] = []
        for page in reader.pages[:25]:
            txt = page.extract_text() or ""
            if txt.strip():
                parts.append(txt.strip())
            if sum(len(p) for p in parts) >= CHAT_UPLOAD_TEXT_MAX_CHARS:
                break
        return "\n\n".join(parts)
    except Exception:
        return ""


def extract_docx_text_for_chat(path: Path) -> str:
    try:
        with zipfile.ZipFile(path) as zf:
            raw_xml = zf.read("word/document.xml")
    except (KeyError, zipfile.BadZipFile, OSError):
        return ""
    try:
        root = _ET.fromstring(raw_xml)
    except _ET.ParseError:
        return ""
    ns = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    paras: list[str] = []
    for para in root.iter(f"{ns}p"):
        bits: list[str] = []
        for node in para.iter():
            if node.tag == f"{ns}t" and node.text:
                bits.append(node.text)
            elif node.tag == f"{ns}tab":
                bits.append("\t")
            elif node.tag in {f"{ns}br", f"{ns}cr"}:
                bits.append("\n")
        line = "".join(bits).strip()
        if line:
            paras.append(line)
        if sum(len(p) for p in paras) >= CHAT_UPLOAD_TEXT_MAX_CHARS:
            break
    return "\n".join(paras)


def extract_xlsx_text_for_chat(path: Path) -> str:
    try:
        import openpyxl  # noqa: PLC0415
    except ImportError:
        return ""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return ""
    parts: list[str] = []
    try:
        for sheet_name in wb.sheetnames[:8]:
            ws = wb[sheet_name]
            parts.append(f"## Hoja: {sheet_name}")
            for row in ws.iter_rows(values_only=True):
                vals = ["" if cell is None else str(cell) for cell in row]
                line = "\t".join(vals).strip()
                if line:
                    parts.append(line)
                if sum(len(p) for p in parts) >= CHAT_UPLOAD_TEXT_MAX_CHARS:
                    return "\n".join(parts)
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return "\n".join(parts)


def copy_chat_upload_to_vault(raw: bytes, filename: str, file_hash: str) -> str | None:
    try:
        import rag as _rag  # noqa: PLC0415
        from datetime import datetime as _dt  # noqa: PLC0415

        vault_root = _rag.VAULT_PATH
        if not vault_root.is_dir():
            return None
        target_dir = vault_root / "00-Inbox" / "chat-uploads"
        target_dir.mkdir(parents=True, exist_ok=True)
        safe_name = safe_chat_upload_filename(filename)
        ts = _dt.now().strftime("%Y%m%d-%H%M%S")
        target = target_dir / f"{ts}-{file_hash[:8]}-{safe_name}"
        existing = list(target_dir.glob(f"*-{file_hash[:8]}-{safe_name}"))
        if existing:
            return str(existing[0].relative_to(vault_root))
        target.write_bytes(raw)
        return str(target.relative_to(vault_root))
    except OSError:
        return None


async def read_upload_capped(file: UploadFile, max_bytes: int) -> bytes:
    raw = await file.read(max_bytes + 1)
    if not raw:
        raise HTTPException(status_code=400, detail="archivo vacío")
    if len(raw) > max_bytes:
        raise HTTPException(
            status_code=413,
            detail=f"archivo muy grande (>{max_bytes} bytes)",
        )
    return raw


def extract_chat_upload_text(path: Path, raw: bytes, content_type: str) -> str:
    suffix = path.suffix.lower()
    ctype = (content_type or "").lower()
    if suffix in CHAT_UPLOAD_TEXT_SUFFIXES or ctype.startswith("text/") or ctype in {
        "application/json",
        "application/xml",
        "application/x-yaml",
        "application/yaml",
    }:
        if looks_like_binary(raw):
            return ""
        return decode_chat_upload_text(raw)
    if suffix == ".pdf":
        return extract_pdf_text_for_chat(path)
    if suffix == ".docx":
        return extract_docx_text_for_chat(path)
    if suffix == ".xlsx":
        return extract_xlsx_text_for_chat(path)
    if not looks_like_binary(raw):
        return decode_chat_upload_text(raw)
    return ""


__all__ = [
    "CHAT_UPLOAD_DOC_SUFFIXES",
    "CHAT_UPLOAD_IMAGE_SUFFIXES",
    "CHAT_UPLOAD_TEXT_MAX_CHARS",
    "CHAT_UPLOAD_TEXT_SUFFIXES",
    "HEIC_AVAILABLE",
    "SANITIZABLE_FORMATS",
    "copy_chat_upload_to_vault",
    "decode_chat_upload_text",
    "extract_chat_upload_text",
    "extract_docx_text_for_chat",
    "extract_pdf_text_for_chat",
    "extract_xlsx_text_for_chat",
    "looks_like_binary",
    "read_upload_capped",
    "safe_chat_upload_filename",
    "sanitize_image_exif",
    "truncate_chat_upload_text",
]
