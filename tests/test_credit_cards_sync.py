"""Unit tests for `rag._sync_credit_cards_notes`.

Mirrors the MOZE pattern: produce 1 markdown note per (card, cycle) under
`{vault}/04-Archive/99-obsidian-system/99-AI/external-ingest/Finanzas/Tarjetas/`, hash-skip if content
unchanged, prune notes whose source xlsx no longer exists.

Tests use real openpyxl-built xlsx fixtures (same builder used by
`test_credit_cards_parser.py`) and a temp vault under `tmp_path` so the
sync writes real .md files we can read back and assert on.

`TARJETAS_BACKUP_DIR` (the source dir of xlsx del banco) is monkeypatched
a un tmp_path subdir per test — we never touch the real iCloud `/Finances`
folder. Post 2026-05-04 split, MOZE vive en otro container y NO interfiere
con el sync de tarjetas.
"""
from __future__ import annotations

from pathlib import Path

import pytest

from tests.test_credit_cards_parser import _make_card_xlsx


# ── Helpers ────────────────────────────────────────────────────────────────


@pytest.fixture
def isolated_finances(tmp_path, monkeypatch):
    """Fresh temp dir for xlsx + temp vault + monkeypatched TARJETAS_BACKUP_DIR.

    Returns `(finance_dir, vault_root)`. Tests drop xlsx files in
    `finance_dir` and call `_sync_credit_cards_notes(vault_root)`.

    `_sync_credit_cards_notes` resuelve el binding en `rag.cross_source_etls`,
    así que monkeypatcheamos ambos paths para que el override tome efecto:

    - `rag.cross_source_etls.TARJETAS_BACKUP_DIR` — el lookup real
    - `rag.TARJETAS_BACKUP_DIR` — re-export vía `from rag.cross_source_etls import *`,
      lo dejamos coherente para tests que lean de `rag` directo.
    """
    import rag
    import rag.cross_source_etls as _cse

    finance_dir = tmp_path / "Finances"
    finance_dir.mkdir()
    vault_root = tmp_path / "vault"
    vault_root.mkdir()

    # Point the sync at our test dir, not iCloud.
    monkeypatch.setattr(_cse, "TARJETAS_BACKUP_DIR", finance_dir)
    monkeypatch.setattr(rag, "TARJETAS_BACKUP_DIR", finance_dir, raising=False)

    return finance_dir, vault_root


# ── Tests ──────────────────────────────────────────────────────────────────


def test_sync_writes_one_note_per_card_cycle(isolated_finances):
    """Single xlsx → 1 .md under `04-Archive/99-obsidian-system/99-AI/external-ingest/Finanzas/Tarjetas/`."""
    import rag

    finance_dir, vault_root = isolated_finances
    _make_card_xlsx(finance_dir / "Último resumen - Visa 1059.xlsx")

    stats = rag._sync_credit_cards_notes(vault_root)

    assert stats["ok"] is True
    assert stats["files_total"] == 1
    assert stats["files_written"] == 1
    assert stats["files_skipped"] == 0
    assert stats["files_parse_failed"] == 0
    assert stats["target"] == "04-Archive/99-obsidian-system/99-AI/external-ingest/Finanzas/Tarjetas"

    target_dir = vault_root / "04-Archive/99-obsidian-system/99-AI/external-ingest/Finanzas/Tarjetas"
    notes = list(target_dir.glob("*.md"))
    assert len(notes) == 1
    assert notes[0].name == "Tarjeta-Visa-1059-2026-03.md"


def test_sync_note_has_structured_frontmatter(isolated_finances):
    """The .md must have a YAML frontmatter the indexer can read for
    metadata-aware retrieval. type=finanzas + source=tarjeta + brand +
    last4 + cycle + totals + tags must be present and well-formed.
    """
    import rag

    finance_dir, vault_root = isolated_finances
    _make_card_xlsx(finance_dir / "Último resumen - Visa 1059.xlsx")
    rag._sync_credit_cards_notes(vault_root)

    note = (vault_root / "04-Archive/99-obsidian-system/99-AI/external-ingest/Finanzas/Tarjetas/"
            "Tarjeta-Visa-1059-2026-03.md")
    body = note.read_text(encoding="utf-8")

    # Frontmatter is a YAML block delimited by `---`.
    assert body.startswith("---\n")
    fm_end = body.index("---\n", 4)  # second `---` line
    fm = body[4:fm_end]

    # Required keys.
    assert "type: finanzas" in fm
    assert "source: tarjeta" in fm
    assert "brand: Visa" in fm
    assert 'last4: "1059"' in fm  # quoted to avoid YAML int coercion
    assert "cycle: 2026-03" in fm
    assert "closing_date: 2026-03-26" in fm
    assert "due_date: 2026-04-08" in fm
    assert "total_ars: 549438.75" in fm
    assert "total_usd: 98.93" in fm
    # Tags must include the brand-lowercased.
    assert "tags: [finanzas, tarjeta, visa]" in fm
    # ambient: skip prevents the daily ambient signals loop from sucking
    # these in as conversational context — same convention as MOZE notes.
    assert "ambient: skip" in fm


def test_sync_note_body_contains_purchases(isolated_finances):
    """Body must include enough text for retrieval: TL;DR resumen + tablas
    + lista plana de TODOS los movimientos (con descripciones que el
    retriever puede matchear, ej. 'Merpago*idilicadeco').
    """
    import rag

    finance_dir, vault_root = isolated_finances
    _make_card_xlsx(finance_dir / "Último resumen - Visa 1059.xlsx")
    rag._sync_credit_cards_notes(vault_root)

    note = (vault_root / "04-Archive/99-obsidian-system/99-AI/external-ingest/Finanzas/Tarjetas/"
            "Tarjeta-Visa-1059-2026-03.md")
    body = note.read_text(encoding="utf-8")

    # Title + cycle anchored.
    assert "# Tarjeta Visa ·1059 — ciclo 2026-03" in body
    # TL;DR cita literal del total — REGLA 1.b del web system prompt
    # depende de que el monto exacto aparezca en el chunk recuperado.
    assert "549.438,75" in body
    assert "98,93" in body
    assert "**2026-03-26**" in body  # closing_date como bold
    # Tablas top movimientos — separadas por moneda.
    assert "## Top movimientos ARS" in body
    assert "## Top movimientos USD" in body
    # Sección "Todos los movimientos" — habilita retrieval por descripción.
    assert "## Todos los movimientos" in body
    assert "Merpago*idilicadeco" in body
    assert "Apple.com/bill" in body  # USD purchase, ambas monedas listadas


def test_sync_is_idempotent(isolated_finances):
    """Run 2x → second run is a no-op (hash-skip)."""
    import rag

    finance_dir, vault_root = isolated_finances
    _make_card_xlsx(finance_dir / "Último resumen - Visa 1059.xlsx")

    s1 = rag._sync_credit_cards_notes(vault_root)
    s2 = rag._sync_credit_cards_notes(vault_root)

    assert s1["files_written"] == 1
    assert s2["files_written"] == 0
    assert s2["files_skipped"] == 1


def test_sync_handles_multiple_cards_separate_notes(isolated_finances):
    """Two cards (Visa 1059 + Mastercard 5234) → two separate .md notes,
    one per (brand, last4, cycle).
    """
    import rag

    finance_dir, vault_root = isolated_finances
    _make_card_xlsx(
        finance_dir / "Último resumen - Visa 1059.xlsx",
        brand="Visa", last4="1059",
    )
    _make_card_xlsx(
        finance_dir / "Último resumen - Mastercard 5234.xlsx",
        brand="Mastercard", last4="5234",
    )
    stats = rag._sync_credit_cards_notes(vault_root)

    assert stats["files_total"] == 2
    assert stats["files_written"] == 2

    target_dir = vault_root / "04-Archive/99-obsidian-system/99-AI/external-ingest/Finanzas/Tarjetas"
    names = sorted(p.name for p in target_dir.glob("*.md"))
    assert names == [
        "Tarjeta-Mastercard-5234-2026-03.md",
        "Tarjeta-Visa-1059-2026-03.md",
    ]


def test_sync_different_cycles_create_different_notes(isolated_finances):
    """Same card, two cycles (e.g. user has both March and February xlsx
    in /Finances) → two notes, both preserved as historical record.
    """
    import rag

    finance_dir, vault_root = isolated_finances
    _make_card_xlsx(
        finance_dir / "Último resumen - Visa 1059 - feb.xlsx",
        closing_date="26/02/2026", due_date="08/03/2026",
    )
    _make_card_xlsx(
        finance_dir / "Último resumen - Visa 1059.xlsx",
        closing_date="26/03/2026", due_date="08/04/2026",
    )

    stats = rag._sync_credit_cards_notes(vault_root)
    target_dir = vault_root / "04-Archive/99-obsidian-system/99-AI/external-ingest/Finanzas/Tarjetas"
    names = sorted(p.name for p in target_dir.glob("*.md"))

    assert stats["files_written"] == 2
    assert names == [
        "Tarjeta-Visa-1059-2026-02.md",
        "Tarjeta-Visa-1059-2026-03.md",
    ]


def test_sync_prunes_orphan_notes(isolated_finances):
    """If a previously-synced xlsx is deleted, the corresponding .md is
    pruned on the next run (mirrors MOZE prune behavior).
    """
    import rag

    finance_dir, vault_root = isolated_finances
    xlsx = _make_card_xlsx(finance_dir / "Último resumen - Visa 1059.xlsx")
    rag._sync_credit_cards_notes(vault_root)

    target_dir = vault_root / "04-Archive/99-obsidian-system/99-AI/external-ingest/Finanzas/Tarjetas"
    note = target_dir / "Tarjeta-Visa-1059-2026-03.md"
    assert note.is_file()

    # User deletes xlsx (e.g. moved to archive).
    xlsx.unlink()
    stats = rag._sync_credit_cards_notes(vault_root)

    # No xlsx left → silent-fail with no_xlsx.
    assert stats["ok"] is False
    assert stats["reason"] == "no_xlsx"
    # But the orphan note is NOT pruned in this path because we exit
    # before the prune loop — the prune only fires when we have at least
    # one valid xlsx, otherwise we'd nuke everything if the user
    # temporarily moves /Finances elsewhere. This is a deliberate safety
    # tradeoff documented in `_sync_credit_cards_notes`.

    # If we add a different xlsx, the orphan should get pruned.
    _make_card_xlsx(
        finance_dir / "Último resumen - Mastercard 5234.xlsx",
        brand="Mastercard", last4="5234",
    )
    stats = rag._sync_credit_cards_notes(vault_root)
    assert stats["ok"] is True

    remaining = sorted(p.name for p in target_dir.glob("*.md"))
    assert remaining == ["Tarjeta-Mastercard-5234-2026-03.md"]
    assert not note.is_file()  # pruned


def test_sync_silent_fails_when_no_xlsx(isolated_finances):
    """Empty `/Finances` → `{ok: False, reason: 'no_xlsx'}`. Doesn't
    raise, doesn't create the target dir.
    """
    import rag

    _, vault_root = isolated_finances
    stats = rag._sync_credit_cards_notes(vault_root)
    assert stats == {"ok": False, "reason": "no_xlsx"}


def test_sync_skips_xlsx_with_missing_brand_or_last4(isolated_finances):
    """An xlsx whose brand+last4 the parser can't recover → skipped (not
    written, not crashing). Counts in `files_parse_failed`.
    """
    import rag

    finance_dir, vault_root = isolated_finances
    # Filename has no brand/last4 hint. The parser tries the sheet name
    # too — set it to something that ALSO doesn't match.
    bad = _make_card_xlsx(
        finance_dir / "Último resumen - sin nada.xlsx",
        brand="Visa", last4="1059",  # builder still injects in body
    )
    # Manually reopen and clobber the sheet name + the "terminada en"
    # row so the parser can't recover the metadata.
    from openpyxl import load_workbook
    wb = load_workbook(bad)
    ws = wb.active
    ws.title = "Resumen"  # generic sheet name, no brand
    # Wipe the row that says "Tarjeta Visa Crédito terminada en 1059".
    for r_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
        for c in row:
            if c.value and "terminada en" in str(c.value).lower():
                c.value = "Resumen del ciclo"
                break
    wb.save(bad)

    stats = rag._sync_credit_cards_notes(vault_root)
    # The xlsx was found and parsed, but `_card_note_filename` returned
    # None (no brand/last4) → counted as parse_failed.
    assert stats["ok"] is True or stats["reason"] == "no_parsed"
    target_dir = vault_root / "04-Archive/99-obsidian-system/99-AI/external-ingest/Finanzas/Tarjetas"
    if target_dir.is_dir():
        assert list(target_dir.glob("*.md")) == []
