"""Finances integration — ingester para PDFs de tarjeta de crédito y Excels de movimientos.

Escanea la carpeta iCloud `~/Library/Mobile Documents/com~apple~CloudDocs/Finances`,
extrae texto de PDFs (ocrmac/pypdf) y Excels (openpyxl), y escribe chunks al corpus
RAG con source="finances". Searchable vía `rag query --source finances`.

## Surfaces

- `ingest()` — corre el ingester completo: escanea archivos, extrae texto,
  chunkea, y upserta a la collection sqlite-vec. Idempotente.
- `_extract_text_from_pdf(path)` — extrae texto de PDF usando ocrmac o pypdf.
- `_extract_text_from_excel(path)` — extrae texto de Excel usando openpyxl.

## Invariantes

- Silent-fail: si la carpeta no existe o un archivo falla, se loguea y se continua.
- Idempotente: re-running re-indexa solo archivos modificados (mtime check).
- PDFs: prioriza ocrmac (macOS Vision), fallback a pypdf si hay texto extraíble.
- Excels: lee todas las hojas, extrae texto de celdas no-vacías.
"""

from __future__ import annotations

import logging
import os
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

logger = logging.getLogger("rag.integrations.finances")

# Default location of the iCloud-synced Finances folder. Overridable via env
# `RAG_FINANCES_PATH` for testing or non-default iCloud setups.
_DEFAULT_FINANCES_PATH = (
    Path.home()
    / "Library/Mobile Documents/com~apple~CloudDocs/Finances"
)

# Chunk parameters
_CHUNK_MAX_CHARS = 800
_CHUNK_MIN_CHARS = 150


def _finances_path() -> Path:
    """Resolves the Finances path, honoring RAG_FINANCES_PATH env override."""
    env = os.environ.get("RAG_FINANCES_PATH")
    return Path(env) if env else _DEFAULT_FINANCES_PATH


def _extract_text_from_pdf(path: Path) -> str:
    """Extract text from PDF using pdftotext (poppler), fallback to pypdf + ocrmac."""
    # Try pdftotext first (best for banking PDFs)
    try:
        import subprocess
        result = subprocess.run(
            ["pdftotext", str(path), "-"],
            capture_output=True,
            text=True,
            timeout=30
        )
        if result.returncode == 0 and result.stdout.strip():
            text = result.stdout.strip()
            logger.info("pdftotext extracted %d chars", len(text))
            return text
    except FileNotFoundError:
        logger.info("pdftotext not found, falling back to pypdf")
    except subprocess.TimeoutExpired:
        logger.warning("pdftotext timeout, falling back to pypdf")
    except Exception as exc:
        logger.warning("pdftotext failed: %s, falling back to pypdf", exc)

    # Fallback to pypdf
    try:
        import pypdf
    except ImportError:
        logger.warning("pypdf not available for PDF extraction")
        return ""

    try:
        reader = pypdf.PdfReader(path)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        text = text.strip()
        # If pypdf extracted very little text, try OCR
        if len(text) < 100:
            logger.info("pypdf extracted minimal text (%d chars), trying OCR", len(text))
            return _extract_text_from_pdf_ocr(path)
        return text
    except Exception as exc:
        logger.warning("Failed to extract PDF text with pypdf: %s", exc)
        return _extract_text_from_pdf_ocr(path)


def _extract_text_from_pdf_ocr(path: Path) -> str:
    """Extract text from PDF using pdf2image + ocrmac (macOS Vision)."""
    try:
        from pdf2image import convert_from_path
        from ocrmac import ocrmac as ocrmac_mod
    except ImportError as e:
        logger.warning("pdf2image or ocrmac not available for PDF OCR: %s", e)
        return ""

    try:
        # Convert PDF to images
        images = convert_from_path(str(path), dpi=200)
        text = ""
        for img in images:
            # OCR each image with ocrmac
            result = ocrmac_mod.OCR(img, language_preference=["es-ES", "en-US"])
            text += result.recognize() + "\n"
        return text.strip()
    except Exception as exc:
        logger.warning("Failed to extract PDF text with pdf2image+ocrmac: %s", exc)
        return ""


def _extract_text_from_excel(path: Path) -> str:
    """Extract text from Excel using xlrd (.xls) or openpyxl (.xlsx).

    Some .xls files are actually HTML exports from banking systems.
    Falls back to text extraction if xlrd fails.
    """
    suffix = path.suffix.lower()

    if suffix == ".xls":
        try:
            import xlrd
        except ImportError:
            logger.warning("xlrd not available for .xls extraction")
            return ""

        try:
            wb = xlrd.open_workbook(path)
            text_parts = []
            for sheet_idx in range(wb.nsheets):
                sheet = wb.sheet_by_index(sheet_idx)
                text_parts.append(f"## Hoja: {sheet.name}\n")
                for row_idx in range(sheet.nrows):
                    row = sheet.row(row_idx)
                    row_text = " | ".join(str(cell.value) if cell.value else "" for cell in row)
                    if row_text.strip():
                        text_parts.append(row_text)
            return "\n".join(text_parts).strip()
        except Exception as exc:
            logger.warning("Failed to extract .xls text with xlrd (trying fallback): %s", exc)
            # Fallback: try to read as text (some .xls are HTML exports)
            try:
                return path.read_text(encoding="utf-8", errors="replace")
            except Exception as exc2:
                logger.warning("Text fallback also failed: %s", exc2)
                return ""

    if suffix == ".xlsx":
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl not available for .xlsx extraction")
            return ""

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            text_parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                text_parts.append(f"## Hoja: {sheet_name}\n")
                for row in ws.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip():
                        text_parts.append(row_text)
            return "\n".join(text_parts).strip()
        except Exception as exc:
            logger.warning("Failed to extract .xlsx text with openpyxl: %s", exc)
            return ""

    logger.warning("Unsupported Excel format: %s", suffix)
    return ""


def _simple_chunks(text: str, max_chars: int = _CHUNK_MAX_CHARS,
                  min_chars: int = _CHUNK_MIN_CHARS) -> list[str]:
    """Simple chunking by size (no semantic splitting like vault)."""
    if not text or len(text) < min_chars:
        return [text] if text else []

    chunks = []
    current = []
    current_len = 0

    for paragraph in text.split("\n\n"):
        para_len = len(paragraph)
        if current_len + para_len <= max_chars:
            current.append(paragraph)
            current_len += para_len
        else:
            if current:
                chunks.append("\n\n".join(current))
            # Start new chunk
            if para_len > max_chars:
                # Split oversized paragraph by sentences
                sentences = paragraph.split(". ")
                current = []
                current_len = 0
                for sent in sentences:
                    sent = sent.strip()
                    if not sent:
                        continue
                    sent_len = len(sent) + 1  # +1 for period/space
                    if current_len + sent_len <= max_chars:
                        current.append(sent + ".")
                        current_len += sent_len
                    else:
                        if current:
                            chunks.append(" ".join(current))
                        current = [sent + "."]
                        current_len = sent_len
            else:
                current = [paragraph]
                current_len = para_len

    if current:
        chunks.append("\n\n".join(current))

    # Merge undersized chunks
    merged = []
    for chunk in chunks:
        if merged and len(merged[-1]) < min_chars:
            merged[-1] += "\n\n" + chunk
        else:
            merged.append(chunk)

    return merged


@dataclass(frozen=True)
class FinanceFile:
    path: Path
    name: str
    mtime: float
    file_type: str  # "pdf" or "excel"
    text: str


def _scan_finances_folder(base_path: Path) -> list[FinanceFile]:
    """Scan Finances folder recursively and extract text from supported files."""
    if not base_path.is_dir():
        logger.warning("Finances folder not found: %s", base_path)
        return []

    files = []
    for item in base_path.rglob("*"):
        if item.is_dir():
            continue
        if item.name.startswith("."):
            continue

        suffix = item.suffix.lower()
        if suffix not in {".pdf", ".xlsx", ".xls"}:
            continue

        mtime = item.stat().st_mtime
        file_type = "pdf" if suffix == ".pdf" else "excel"

        if file_type == "pdf":
            text = _extract_text_from_pdf(item)
        else:
            text = _extract_text_from_excel(item)

        if text:
            files.append(FinanceFile(
                path=item,
                name=str(item.relative_to(base_path)),
                mtime=mtime,
                file_type=file_type,
                text=text,
            ))

    return files


def _get_ingest_state(conn) -> dict[str, float]:
    """Load last mtime per file from state table."""
    try:
        rows = conn.execute(
            "SELECT file_path, last_mtime FROM rag_finances_state"
        ).fetchall()
        return {row["file_path"]: row["last_mtime"] for row in rows}
    except Exception:
        # Table doesn't exist yet
        return {}


def _save_ingest_state(conn, file_path: str, mtime: float) -> None:
    """Upsert mtime for a file in state table."""
    conn.execute(
        "INSERT OR REPLACE INTO rag_finances_state (file_path, last_mtime) VALUES (?, ?)",
        (file_path, mtime)
    )


def ingest(path: Path | str | None = None) -> dict[str, Any]:
    """Run the full Finances → sqlite-vec ingester. Idempotent.

    Returns a dict with counts:
      `{folder, files_scanned, files_updated, chunks_written, elapsed_ms}`
    """
    t0 = time.time()
    base = Path(path) if path else _finances_path()

    files = _scan_finances_folder(base)
    if not files:
        return {
            "folder": str(base),
            "files_scanned": 0,
            "files_updated": 0,
            "chunks_written": 0,
            "elapsed_ms": round((time.time() - t0) * 1000, 1),
            "skipped": True,
            "reason": "no_files_found",
        }

    # Import rag for collection access
    try:
        import rag
    except ImportError:
        logger.error("Failed to import rag module")
        return {
            "folder": str(base),
            "files_scanned": len(files),
            "files_updated": 0,
            "chunks_written": 0,
            "elapsed_ms": round((time.time() - t0) * 1000, 1),
            "error": "rag_import_failed",
        }

    # Get collection
    try:
        col = rag.get_db()
    except Exception as exc:
        logger.error("Failed to get collection: %s", exc)
        return {
            "folder": str(base),
            "files_scanned": len(files),
            "files_updated": 0,
            "chunks_written": 0,
            "elapsed_ms": round((time.time() - t0) * 1000, 1),
            "error": f"collection_failed: {exc}",
        }

    # State table for incremental updates
    with rag._ragvec_state_conn() as state_conn:
        state_conn.execute(
            "CREATE TABLE IF NOT EXISTS rag_finances_state ("
            " file_path TEXT PRIMARY KEY,"
            " last_mtime REAL NOT NULL"
            ")"
        )

        # Load last mtimes
        last_mtimes = _get_ingest_state(state_conn)

        # Filter files to process (modified or new)
        to_process = []
        for f in files:
            last_mtime = last_mtimes.get(str(f.path), 0.0)
            if f.mtime > last_mtime:
                to_process.append(f)

        if not to_process:
            return {
                "folder": str(base),
                "files_scanned": len(files),
                "files_updated": 0,
                "chunks_written": 0,
                "elapsed_ms": round((time.time() - t0) * 1000, 1),
            }

        # Process files
        total_chunks = 0
        for f in to_process:
            chunks = _simple_chunks(f.text)
            if not chunks:
                continue

            # Generate embeddings
            embed_prefixes = [f"[source=finances | file={f.name} | type={f.file_type}] {c}"
                              for c in chunks]
            embeddings = rag.embed(embed_prefixes)

            # Build metadata
            ids = []
            docs = []
            metas = []
            for i, chunk in enumerate(chunks):
                doc_id = f"finances://{f.name}::{i}"
                ids.append(doc_id)
                docs.append(chunk)
                metas.append({
                    "file": f"finances://{f.name}",
                    "note": f"Finances: {f.name}",
                    "folder": "Finances",
                    "tags": "",
                    "hash": "",
                    "outlinks": "",
                    "source": "finances",
                    "created_ts": f.mtime,
                    "file_name": f.name,
                    "file_type": f.file_type,
                    "file_path": str(f.path),
                })

            # Delete existing chunks for this file
            try:
                file_key = f"finances://{f.name}"
                existing = col.get(where={"file": file_key}, include=[])
                if existing.get("ids"):
                    col.delete(ids=existing["ids"])
            except Exception as exc:
                logger.warning("finances: failed to delete existing chunks for %s: %s", f.name, exc)

            # Add new chunks
            if ids and docs and metas:
                col.add(
                    ids=ids,
                    documents=docs,
                    metadatas=metas,
                    embeddings=embeddings,
                )
                total_chunks += len(ids)

            # Update state
            _save_ingest_state(state_conn, str(f.path), f.mtime)

    return {
        "folder": str(base),
        "files_scanned": len(files),
        "files_updated": len(to_process),
        "chunks_written": total_chunks,
        "elapsed_ms": round((time.time() - t0) * 1000, 1),
    }


__all__ = ["ingest"]
