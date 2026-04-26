"""Tests para el parser de resúmenes de tarjeta de crédito (xlsx que el banco
emite y el user deja en iCloud `/Finances`). Cubre:

1. ``_parse_credit_card_xlsx`` — extracción de cada campo desde un xlsx
   generado en tmp_path con la misma forma que produce Santander Río.
2. ``_fetch_credit_cards`` — glob del dir, agregación, ordenamiento por
   fecha de vencimiento, cache hit/miss por (path, mtime).
3. Resiliencia: dir vacío, xlsx malformado, openpyxl ausente (smoke).
4. Helpers ``_parse_ars_or_usd`` y ``_parse_card_date``.

El xlsx de fixture se genera con openpyxl directamente para no shippear un
binario al repo (que también obligaría a re-exportarlo cada vez que el banco
cambie el formato).

Origen del feature: 2026-04-26 — el user mudó la fuente de datos financiera
de iCloud `/Backup` a `/Finances` y agregó los `.xlsx` de resúmenes de
tarjeta junto a los CSV de MOZE. Decisión del user: panel separado
"tarjetas" en home + tool nuevo `credit_cards_summary` (sin notas en vault).
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest

from web.server import (
    _fetch_credit_cards,
    _format_cards_block,
    _is_empty_tool_output,
    _parse_ars_or_usd,
    _parse_card_date,
    _parse_credit_card_xlsx,
)


# ── Fixture builder ─────────────────────────────────────────────────────────


def _make_card_xlsx(
    path: Path,
    *,
    brand: str = "Visa",
    last4: str = "1059",
    holder: str = "Fernando Raul Ferrari",
    closing_date: str = "26/03/2026",
    due_date: str = "08/04/2026",
    next_closing_date: str = "30/04/2026",
    next_due_date: str = "08/05/2026",
    total_ars: str = "$549.438,75",
    total_usd: str = "U$S98,93",
    minimum_ars: str | None = "$549.438,75",
    minimum_usd: str | None = "U$S98,93",
    purchases: list[tuple[str, str, str | None, str | None]] | None = None,
) -> Path:
    """Genera un .xlsx con la forma de un resumen de Santander Río — la misma
    estructura que vimos en el archivo real durante el bring-up. Cada
    `purchases` es (fecha, descripción, monto_ars, monto_usd) — solo uno de
    los dos montos suele estar populado.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = f"{brand} {last4}"

    # Layout copiado del xlsx real, fila por fila (rows en blanco
    # son intencionales — el parser tolera celdas vacías).
    rows: list[list] = [
        [],
        ["Movimientos del resumen"],
        [f"Tarjeta {brand} Crédito terminada en {last4}"],
        [],
        ["Fecha de cierre", "Fecha de vencimiento"],
        [closing_date, due_date],
        ["Total a pagar"],
        [total_ars, total_usd],
        ["Mínimo a pagar"],
        [minimum_ars or "", minimum_usd or ""],
        [],
        [],
        ["Tarjetas incluidas en el resumen", "Tarjeta de", "Total en pesos", "Total en dólares"],
        [f"{brand} Crédito terminada en {last4}", f"{holder} (Titular)", "$493.756,15", "U$S98,93"],
        [],
        [],
        ["Cierres y vencimientos"],
        ["Resumen actual", "Próximo resumen"],
        [f"Cierre: {closing_date}", f"Cierre: {next_closing_date}"],
        [f"Vencimiento: {due_date}", f"Vencimiento: {next_due_date}"],
        [],
        [],
        ["Pago de tarjeta y devoluciones"],
        ["Fecha", "Descripción", "Cuotas", "Comprobante", "Monto en pesos", "Monto en dólares"],
        ["06/03/2026", "Su pago en pesos", "", "-", "$-926.148,78", None],
        [],
        [],
        [f"Tarjeta de {holder} - {brand} Crédito terminada en {last4}"],
        ["Fecha", "Descripción", "Cuotas", "Comprobante", "Monto en pesos", "Monto en dólares"],
    ]
    if purchases is None:
        purchases = [
            ("29/01/2026", "Merpago*idilicadeco", "$153.333,33", None),
            ("25/02/2026", "Dlo*starlink", "$34.000,00", None),
            ("26/02/2026", "Apple.com/bill", None, "U$S24,99"),
            ("14/03/2026", "Claude.ai subscr in1t", None, "U$S20,00"),
            ("25/03/2026", "Starlink", "$33.160,00", None),
        ]
    for fecha, desc, ars_amt, usd_amt in purchases:
        rows.append([fecha, desc, "", "000000K", ars_amt, usd_amt])
    rows.append([f"Total de {brand} Crédito terminada en {last4}", None, None, None, "$493.756,15", "U$S98,93"])
    rows.append([])
    rows.append(["Otros conceptos"])
    rows.append(["Descripción", "Monto en pesos"])
    rows.append(["Impuesto de sellos $", "$548,75"])

    for r in rows:
        ws.append(r)
    wb.save(path)
    return path


# ── Helpers: _parse_ars_or_usd / _parse_card_date ───────────────────────────


@pytest.mark.parametrize("raw,expected", [
    ("$549.438,75", (549438.75, "ARS")),
    ("U$S98,93", (98.93, "USD")),
    ("USD 24.99", (24.99, "USD")),
    ("$-926.148,78", (-926148.78, "ARS")),
    ("", (None, None)),
    (None, (None, None)),
    ("garbage", (None, None)),
    (12345.67, (12345.67, None)),  # numeric passthrough
    ("123,45", (123.45, None)),    # ES decimals sin símbolo
])
def test_parse_ars_or_usd(raw, expected):
    assert _parse_ars_or_usd(raw) == expected


@pytest.mark.parametrize("raw,expected", [
    ("26/03/2026", "2026-03-26"),
    ("8/4/2026", "2026-04-08"),
    ("Cierre: 26/03/2026", "2026-03-26"),
    ("Vencimiento: 08/05/2026", "2026-05-08"),
    ("", None),
    (None, None),
    ("not a date", None),
    ("31/02/2026", None),  # día inválido
])
def test_parse_card_date(raw, expected):
    assert _parse_card_date(raw) == expected


# ── _parse_credit_card_xlsx ─────────────────────────────────────────────────


def test_parse_xlsx_extracts_all_fields(tmp_path):
    xlsx = _make_card_xlsx(tmp_path / "Último resumen - Visa 1059.xlsx")
    result = _parse_credit_card_xlsx(xlsx)

    assert result is not None
    assert result["brand"] == "Visa"
    assert result["last4"] == "1059"
    assert result["holder"] == "Fernando Raul Ferrari"
    assert result["closing_date"] == "2026-03-26"
    assert result["due_date"] == "2026-04-08"
    assert result["next_closing_date"] == "2026-04-30"
    assert result["next_due_date"] == "2026-05-08"
    assert result["total_ars"] == 549438.75
    assert result["total_usd"] == 98.93
    assert result["minimum_ars"] == 549438.75
    assert result["minimum_usd"] == 98.93
    assert result["source_file"] == "Último resumen - Visa 1059.xlsx"
    assert result["source_mtime"] is not None


def test_parse_xlsx_top_purchases_split_by_currency(tmp_path):
    xlsx = _make_card_xlsx(tmp_path / "Último resumen - Visa 1059.xlsx")
    result = _parse_credit_card_xlsx(xlsx)

    # Top ARS sorted desc.
    ars = result["top_purchases_ars"]
    assert len(ars) >= 3
    amounts = [p["amount"] for p in ars]
    assert amounts == sorted(amounts, reverse=True)
    assert ars[0]["description"] == "Merpago*idilicadeco"
    assert ars[0]["amount"] == 153333.33
    assert ars[0]["currency"] == "ARS"

    # USD también ordenado desc.
    usd = result["top_purchases_usd"]
    assert len(usd) >= 2
    assert usd[0]["amount"] == 24.99
    assert usd[0]["currency"] == "USD"


def test_parse_xlsx_excludes_payments_block(tmp_path):
    """El bloque "Pago de tarjeta y devoluciones" tiene el pago al banco
    como movimiento negativo grande ($-926.148,78). NO debe aparecer en
    `top_purchases_ars` — solo movimientos del bloque del titular cuentan.
    """
    xlsx = _make_card_xlsx(tmp_path / "Último resumen - Visa 1059.xlsx")
    result = _parse_credit_card_xlsx(xlsx)

    descs = [p["description"] for p in result["top_purchases_ars"]]
    assert not any("Su pago en pesos" in d for d in descs), (
        f"Pago al banco no debería estar en top_purchases_ars; got {descs}"
    )


def test_parse_xlsx_brand_normalization(tmp_path):
    """Mastercard / Master se normalizan a "Mastercard"; Amex / American
    Express → "Amex".
    """
    xlsx = _make_card_xlsx(
        tmp_path / "Último resumen - Mastercard 5234.xlsx",
        brand="Mastercard", last4="5234",
    )
    result = _parse_credit_card_xlsx(xlsx)
    assert result is not None
    assert result["brand"] == "Mastercard"
    assert result["last4"] == "5234"


def test_parse_xlsx_invalid_returns_none(tmp_path):
    """Un xlsx que no tiene la estructura esperada (sin Total a pagar ni
    fechas) devuelve None — no es un resumen válido.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Esto no es un resumen"])
    ws.append(["Random", "data"])
    bogus = tmp_path / "Último resumen - Garbage.xlsx"
    wb.save(bogus)

    assert _parse_credit_card_xlsx(bogus) is None


def test_parse_xlsx_missing_file_returns_none(tmp_path):
    assert _parse_credit_card_xlsx(tmp_path / "nonexistent.xlsx") is None


# ── _fetch_credit_cards ─────────────────────────────────────────────────────


@pytest.fixture
def cards_dir(tmp_path, monkeypatch):
    """Redirige `_FINANCE_BACKUP_DIR` al tmp_path y limpia el cache.
    Devuelve el dir para que el test caller pueble los xlsx.
    """
    import web.server as srv

    monkeypatch.setattr(srv, "_FINANCE_BACKUP_DIR", tmp_path)
    # Limpiar cache para no contaminar entre tests.
    srv._CARDS_CACHE["key"] = None
    srv._CARDS_CACHE["payload"] = None
    return tmp_path


def test_fetch_returns_empty_when_no_xlsx(cards_dir):
    assert _fetch_credit_cards() == []


def test_fetch_parses_single_xlsx(cards_dir):
    _make_card_xlsx(cards_dir / "Último resumen - Visa 1059.xlsx")

    result = _fetch_credit_cards()
    assert len(result) == 1
    assert result[0]["brand"] == "Visa"
    assert result[0]["last4"] == "1059"


def test_fetch_sorts_by_due_date_ascending(cards_dir):
    """Las tarjetas más próximas a vencer aparecen primero — el panel del
    home muestra primero la accionable.
    """
    _make_card_xlsx(
        cards_dir / "Último resumen - Visa 1059.xlsx",
        last4="1059", due_date="08/05/2026",
    )
    _make_card_xlsx(
        cards_dir / "Último resumen - Mastercard 5234.xlsx",
        brand="Mastercard", last4="5234", due_date="03/05/2026",
    )

    result = _fetch_credit_cards()
    assert len(result) == 2
    assert result[0]["last4"] == "5234"  # vence 03/05 (más cerca)
    assert result[1]["last4"] == "1059"  # vence 08/05


def test_fetch_caches_by_path_and_mtime(cards_dir):
    """Mismo (paths, mtimes) → mismo payload (cache hit). Tocar el archivo
    invalida el cache.
    """
    import os
    import time

    import web.server as srv

    xlsx = _make_card_xlsx(cards_dir / "Último resumen - Visa 1059.xlsx")
    first = _fetch_credit_cards()
    assert len(first) == 1
    cached_payload = srv._CARDS_CACHE["payload"]
    assert cached_payload is not None

    # Sin tocar el archivo: devuelve el mismo objeto exacto desde cache.
    second = _fetch_credit_cards()
    assert second is cached_payload

    # Touch (mtime cambia) → cache miss, re-parse.
    time.sleep(0.01)
    new_mtime = xlsx.stat().st_mtime + 10
    os.utime(xlsx, (new_mtime, new_mtime))
    third = _fetch_credit_cards()
    assert third is not cached_payload
    assert len(third) == 1  # mismo contenido, payload object distinto


def test_fetch_skips_unparseable_xlsx_silently(cards_dir):
    """Un xlsx malformado no rompe el batch — los demás siguen funcionando.
    """
    from openpyxl import Workbook

    # Uno bueno, uno garbage.
    _make_card_xlsx(cards_dir / "Último resumen - Visa 1059.xlsx")
    wb = Workbook()
    wb.active.append(["garbage"])
    wb.save(cards_dir / "Último resumen - Garbage.xlsx")

    result = _fetch_credit_cards()
    # Solo el válido sobrevivió.
    assert len(result) == 1
    assert result[0]["brand"] == "Visa"


def test_fetch_ignores_non_xlsx_files(cards_dir):
    """Archivos que no matchean el glob (CSV de MOZE, .DS_Store, otros
    .xlsx que no empiezan con "Último resumen") se ignoran.
    """
    (cards_dir / "MOZE_20260426.csv").write_text("dummy")
    (cards_dir / ".DS_Store").write_text("garbage")
    (cards_dir / "Otra cosa.xlsx").write_text("also dummy")
    _make_card_xlsx(cards_dir / "Último resumen - Visa 1059.xlsx")

    result = _fetch_credit_cards()
    assert len(result) == 1


# ── _format_cards_block ─────────────────────────────────────────────────────


def test_format_cards_empty_list_returns_placeholder():
    out = _format_cards_block(json.dumps([]))
    assert "### Tarjetas" in out
    assert "Sin resúmenes" in out


def test_format_cards_renders_brand_total_and_due():
    payload = [{
        "brand": "Visa",
        "last4": "1059",
        "total_ars": 549438.75,
        "total_usd": 98.93,
        "due_date": "2026-04-08",
        "closing_date": "2026-03-26",
        "top_purchases_ars": [
            {"description": "Merpago*idilicadeco", "amount": 153333.33, "currency": "ARS"},
        ],
        "top_purchases_usd": [
            {"description": "Apple.com/bill", "amount": 24.99, "currency": "USD"},
        ],
    }]
    out = _format_cards_block(json.dumps(payload))

    assert "### Tarjetas" in out
    assert "Visa" in out and "1059" in out
    assert "Total a pagar" in out
    # ES thousands separator (`.`) con cualquier rounding de cents
    assert "$549.43" in out or "$549.44" in out
    assert "U$S98.93" in out or "U$S98,93" in out
    assert "Vence:" in out and "2026-04-08" in out
    assert "Cierre:" in out and "2026-03-26" in out
    assert "Merpago*idilicadeco" in out
    assert "Apple.com/bill" in out


def test_format_cards_minimum_omitted_when_equals_total():
    """Si mínimo == total (banks suelen igualarlos cuando no hay
    refinanciación), no se renderiza la línea de "Mínimo:" — sería ruido
    visual sin información nueva.
    """
    payload = [{
        "brand": "Visa", "last4": "1059",
        "total_ars": 549438.75, "minimum_ars": 549438.75,
        "due_date": "2026-04-08",
    }]
    out = _format_cards_block(json.dumps(payload))
    assert "Mínimo:" not in out


def test_format_cards_malformed_passthrough():
    out = _format_cards_block("not json")
    assert "### Tarjetas" in out
    assert "not json" in out


# ── _is_empty_tool_output integration ───────────────────────────────────────


def test_is_empty_credit_cards_summary():
    assert _is_empty_tool_output("credit_cards_summary", "[]") is True
    assert _is_empty_tool_output("credit_cards_summary", json.dumps([
        {"brand": "Visa", "last4": "1059"},
    ])) is False
    # Malformed JSON → conservative False (no asumimos que es empty).
    assert _is_empty_tool_output("credit_cards_summary", "garbage") is False
